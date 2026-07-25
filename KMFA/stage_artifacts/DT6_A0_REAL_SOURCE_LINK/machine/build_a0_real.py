"""从真实税务源建 A0 权威基准(源可溯,整数分)。
铁律:真实金额只写私有产物;KMOS 只收 public-safe(源文件/表/行/笔数/哈希,零金额)。
Owner 优先级 税务>银行>金蝶>红圈:本 A0 以税务源 开票纳税汇总表 为权威。"""
import openpyxl, collections, hashlib, json, sys

SRC = "/d/KMFA_MetaData/财务/开票纳税/2024-2025.12开票、纳税 资金汇总表.xlsx"
SRC_NAME = "财务/开票纳税/2024-2025.12开票、纳税 资金汇总表.xlsx"
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

def cents(x):
    return int(round(float(x) * 100)) if isinstance(x, (int, float)) else 0

# —— 成本:按内容类别聚合 付款金额(收支类别=项目成本),源可溯到行 ——
ws = wb["2025项目成本明细"]
cost_by_cat = collections.OrderedDict()
cost_rows = 0
cost_total_cents = 0
for ridx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cat, pay, content = r[3], r[5], r[7]
    if cat and "项目成本" in str(cat) and isinstance(pay, (int, float)):
        c = str(content) if content else "未分类"
        cost_by_cat.setdefault(c, {"cents": 0, "rows": []})
        cost_by_cat[c]["cents"] += cents(pay)
        cost_by_cat[c]["rows"].append(ridx)
        cost_rows += 1
        cost_total_cents += cents(pay)

# —— 收入:开票外部不含税收入 按主体聚合(1-12月),源可溯 ——
ws2 = wb["2025年开票及纳税汇总"]
rev_by_entity = collections.OrderedDict()
rev_total_cents = 0
for ridx, r in enumerate(ws2.iter_rows(min_row=3, values_only=True), start=3):
    name, typ = r[1], r[2]
    if name and typ and "外部" in str(typ):
        months = [c for c in r[3:15] if isinstance(c, (int, float))]
        s = sum(cents(m) for m in months)
        rev_by_entity[str(name)] = {"cents": s, "row": ridx}
        rev_total_cents += s
wb.close()

# —— 私有 A0(含真实金额)——只落私有 scratch,永不进 KMOS ——
private = {
    "a0_id": "KMFA-A0-REAL-20260725-FROM-TAX",
    "authority_source": SRC_NAME,
    "authority_priority": "税务>银行>金蝶>红圈(Owner 2026-07-25)",
    "unit": "integer_cents",
    "cost_by_category": {k: v["cents"] for k, v in cost_by_cat.items()},
    "cost_total_cents": cost_total_cents,
    "cost_rows": cost_rows,
    "revenue_by_entity": {k: v["cents"] for k, v in rev_by_entity.items()},
    "revenue_total_cents": rev_total_cents,
}
priv_bytes = json.dumps(private, ensure_ascii=False, sort_keys=True).encode("utf-8")
priv_hash = "sha256:" + hashlib.sha256(priv_bytes).hexdigest()
with open("/out/a0_real_baseline_private.json", "w", encoding="utf-8") as f:
    json.dump(private, f, ensure_ascii=False, indent=2)

# —— 公开安全证据:只留结构/源可溯/笔数/哈希。零金额、且零业务标识名
#    (类别名可能含人名、主体名是公司名,均属私有,不进公开 KMOS)——
public_safe = {
    "schema_version": "kmfa.a0_real_source_registration.v1",
    "a0_id": private["a0_id"],
    "authority_source_file": SRC_NAME,
    "authority_priority": private["authority_priority"],
    "unit": "integer_cents",
    "cost": {
        "source_sheet": "2025项目成本明细",
        "filter": "收支类别==项目成本 的 付款金额",
        "category_count": len(cost_by_cat),
        "record_rows": cost_rows,
        "amounts_public": False,
        "names_public": False,
    },
    "revenue": {
        "source_sheet": "2025年开票及纳税汇总",
        "filter": "外部不含税收入 1-12月 按主体",
        "entity_count": len(rev_by_entity),
        "amounts_public": False,
        "names_public": False,
    },
    "private_baseline_sha256": priv_hash,
    "private_only_note": "真实金额与业务标识名(类别名/主体名)仅存私有产物;本 manifest 只有笔数/源trace/哈希,可入公开 KMOS。",
}
with open("/out/a0_real_source_manifest.json", "w", encoding="utf-8") as f:
    json.dump(public_safe, f, ensure_ascii=False, indent=2)

print("A0-BUILT")
print("成本笔数:", cost_rows, "| 类别数:", len(cost_by_cat), "| 收入主体数:", len(rev_by_entity))
print("私有基准哈希:", priv_hash)
# 公开 manifest 自检:不得含任何真实金额或业务标识名
pub_str = json.dumps(public_safe, ensure_ascii=False)
leak_names = [n for n in list(cost_by_cat.keys()) + list(rev_by_entity.keys()) if n.strip() and n.strip() in pub_str]
print("公开manifest泄漏业务名:", leak_names if leak_names else "无")
print("公开manifest内容:", pub_str)
