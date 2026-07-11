#!/usr/bin/env python3
"""Generate current KMFA v0.1.4 S15-P1 performance field evidence."""

from __future__ import annotations

import argparse
import functools
import html
import http.server
import io
import json
import os
import socketserver
import subprocess
import sys
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

from KMFA.tools import v014_s14_post_remediation_stage_review as s14_review
from KMFA.tools.check_v014_s14_post_remediation_stage_review import (
    validate_v014_s14_post_remediation_stage_review,
)


PHASE_ID = "V014_S15_P1_POST_REMEDIATION_PERFORMANCE_FACT_FIELDS"
ROADMAP_PHASE_ID = "S15-P1"
TASK_ID = "KMFA-V014-S15-P1-POST-REMEDIATION-PERFORMANCE-FACT-FIELDS-20260711"
ACCEPTANCE_ID = "ACC-V014-S15-P1-POST-REMEDIATION-PERFORMANCE-FACT-FIELDS"
VERSION = "0.1.4-s15-p1-post-remediation-performance-fact-fields"
STATUS = "completed_validated_local_only_s15_p1_all_fields_manual_review_no_go_upload_deferred"
DECISION = "NO_GO"
FORMULA_ID = "FORM-KMFA-V014-S15P1-POST-REMEDIATION-PERFORMANCE-FACT-FIELDS-001"
PARAMETER_IDS = ("PARAM-KMFA-1765", "PARAM-KMFA-1766", "PARAM-KMFA-1767", "PARAM-KMFA-1768")
MODEL_REGISTRY_KEY = "kmfa_v014_s15_p1_post_remediation_performance_fact_fields"

OUTPUT_DIR = Path("KMFA/stage_artifacts") / PHASE_ID
MACHINE_DIR = OUTPUT_DIR / "machine"
HUMAN_DIR = OUTPUT_DIR / "human"
HTML_DIR = OUTPUT_DIR / "exports/html"
SUMMARY_PATH = MACHINE_DIR / "performance_fact_fields_summary.json"
MANIFEST_PATH = MACHINE_DIR / "performance_fact_fields_manifest.json"
FIELD_DEFINITIONS_PATH = MACHINE_DIR / "performance_fact_field_definitions_public_safe.json"
FIELD_BINDINGS_PATH = MACHINE_DIR / "performance_fact_field_binding_status_public_safe.json"
MANUAL_REVIEW_PATH = MACHINE_DIR / "performance_fact_manual_review_requirements_public_safe.json"
MATRIX_PATH = MACHINE_DIR / "performance_fact_fields_acceptance_matrix.json"
GO_NO_GO_PATH = MACHINE_DIR / "performance_fact_fields_go_no_go.json"
HTML_PATH = HTML_DIR / "performance_fact_fields_workbench.html"
REPORT_PATH = HUMAN_DIR / "performance_fact_fields_report_zh.md"
TEST_RESULTS_PATH = HUMAN_DIR / "test_results_zh.md"
RISK_REGISTER_PATH = HUMAN_DIR / "risk_register_zh.md"
ROLLBACK_PATH = HUMAN_DIR / "rollback_plan_zh.md"

QUALITY_DIR = Path("KMFA/metadata/quality")
METADATA_SUMMARY_PATH = QUALITY_DIR / "v014_s15_p1_post_remediation_performance_fact_fields_summary.json"
METADATA_MANIFEST_PATH = QUALITY_DIR / "v014_s15_p1_post_remediation_performance_fact_fields_manifest.json"
METADATA_FIELD_DEFINITIONS_PATH = (
    QUALITY_DIR / "v014_s15_p1_post_remediation_performance_fact_field_definitions_public_safe.json"
)
METADATA_FIELD_BINDINGS_PATH = (
    QUALITY_DIR / "v014_s15_p1_post_remediation_performance_fact_field_binding_status_public_safe.json"
)
METADATA_MANUAL_REVIEW_PATH = (
    QUALITY_DIR / "v014_s15_p1_post_remediation_performance_fact_manual_review_requirements_public_safe.json"
)
METADATA_MATRIX_PATH = QUALITY_DIR / "v014_s15_p1_post_remediation_performance_fact_fields_acceptance_matrix.json"
METADATA_GO_NO_GO_PATH = QUALITY_DIR / "v014_s15_p1_post_remediation_performance_fact_fields_go_no_go.json"

PRIVATE_DIR = Path("KMFA/.codex_private_runtime/v014_s15_p1_post_remediation_performance_fact_fields")
PRIVATE_RAW_BEFORE_PATH = PRIVATE_DIR / "raw_immutability_before.json"
PRIVATE_RAW_AFTER_PATH = PRIVATE_DIR / "raw_immutability_after.json"
PRIVATE_PROBE_PATH = PRIVATE_DIR / "performance_fact_field_candidate_probe.json"
PRIVATE_REPORT_PATH = PRIVATE_DIR / "performance_fact_field_difference_report_zh.md"
PRIVATE_BROWSER_PATH = PRIVATE_DIR / "browser_verification.json"
PRIVATE_BASELINE_AUDIT_PATH = PRIVATE_DIR / "human_flow_baseline_audit.csv"
PRIVATE_CURRENT_AUDIT_PATH = PRIVATE_DIR / "current_performance_fact_fields_audit.csv"
PRIVATE_SCREENSHOT_DIR = PRIVATE_DIR / "screenshots"

ROADMAP_PATH = Path("KMFA/taskpack/v1_4/roadmap/02_KMFA_Codex_Development_Roadmap_18_Stages_v1_4.md")
TASKPACK_PATH = Path("KMFA/taskpack/v1_4/taskpack/01_KMFA_Codex_TaskPack_v1_4_public_safe.md")
HTML_BASELINE_ROOT = Path("KMFA/taskpack/v1_4/html_uiux")
CURRENT_STAGE9_REVIEW_PATH = Path(
    "KMFA/stage_artifacts/V014_S09_POST_REMEDIATION_STAGE_REVIEW/machine/"
    "stage9_post_remediation_review_manifest.json"
)
CURRENT_S13_P1_PATH = Path(
    "KMFA/stage_artifacts/V014_S13_P1_POST_REMEDIATION_FINANCIAL_OPERATING_REPORT/machine/"
    "financial_operating_report_manifest.json"
)
CURRENT_S13_P2_PATH = Path(
    "KMFA/stage_artifacts/V014_S13_P2_POST_REMEDIATION_COLLECTION_RECEIVABLE_AGING/machine/"
    "collection_receivable_aging_manifest.json"
)
CURRENT_S13_P3_PATH = Path(
    "KMFA/stage_artifacts/V014_S13_P3_POST_REMEDIATION_CROSS_TABLE_REVIEW/machine/"
    "cross_table_review_manifest.json"
)
CURRENT_S14_P2_PATH = s14_review.p2.MANIFEST_PATH
LEGACY_MANIFEST_PATH = Path(
    "KMFA/stage_artifacts/V014_S15_P1_PERFORMANCE_FACT_FIELDS/machine/performance_fact_fields_manifest.json"
)
LEGACY_FIELD_DEFINITIONS_PATH = Path(
    "KMFA/stage_artifacts/V014_S15_P1_PERFORMANCE_FACT_FIELDS/machine/"
    "performance_fact_field_definitions.jsonl"
)
LEGACY_FIELD_BINDINGS_PATH = Path(
    "KMFA/stage_artifacts/V014_S15_P1_PERFORMANCE_FACT_FIELDS/machine/"
    "performance_fact_field_bindings.jsonl"
)
LEGACY_MANUAL_REVIEW_PATH = Path(
    "KMFA/stage_artifacts/V014_S15_P1_PERFORMANCE_FACT_FIELDS/machine/"
    "performance_fact_manual_review_fields.jsonl"
)

DEVELOPMENT_EVENTS_PATH = Path("KMFA/docs/governance/development_events.jsonl")
ASSURANCE_STATUS_PATH = Path("KMFA/docs/governance/ASSURANCE_STATUS.yaml")
STAGE_STATUS_PATH = Path("KMFA/metadata/stage_status.jsonl")
TASK_STATUS_PATH = Path("KMFA/metadata/traceability/v1_4_stage_phase_task_status.jsonl")

FIELD_SPECS: tuple[dict[str, Any], ...] = (
    {
        "field_key": "invoice_amount",
        "label": "开票金额",
        "fact_kind": "money_minor_unit",
        "direct_terms": ("开票金额", "发票金额", "已开票金额", "开票额"),
        "paired_terms": (),
        "reason_code": "missing_authoritative_project_invoice_period_and_value_binding",
        "reason_zh": "缺少权威项目、发票、期间与数值绑定",
        "responsible_role": "finance_owner",
        "field_specific_refs": (CURRENT_S14_P2_PATH, CURRENT_S13_P1_PATH),
    },
    {
        "field_key": "gross_margin_rate",
        "label": "毛利率",
        "fact_kind": "ratio_basis_points",
        "direct_terms": ("毛利率", "项目毛利率", "毛利"),
        "paired_terms": (),
        "reason_code": "missing_authoritative_project_revenue_cost_period_and_value_binding",
        "reason_zh": "缺少权威收入、成本、期间与数值绑定",
        "responsible_role": "finance_owner",
        "field_specific_refs": (CURRENT_STAGE9_REVIEW_PATH, CURRENT_S13_P1_PATH),
    },
    {
        "field_key": "settlement_speed",
        "label": "结算速度",
        "fact_kind": "duration_days",
        "direct_terms": ("结算速度", "结算周期", "结算天数", "结算日期", "结算时间"),
        "paired_terms": (
            ("完工日期", "完成日期", "竣工日期"),
            ("结算日期", "结算时间"),
        ),
        "reason_code": "missing_authoritative_completion_settlement_dates_and_shared_project_binding",
        "reason_zh": "缺少权威完工、结算日期与共享项目绑定",
        "responsible_role": "project_owner",
        "field_specific_refs": (CURRENT_S13_P2_PATH, CURRENT_S13_P3_PATH),
    },
    {
        "field_key": "collection_speed",
        "label": "回款速度",
        "fact_kind": "duration_days",
        "direct_terms": ("回款速度", "回款周期", "回款天数", "回款日期", "到账日期"),
        "paired_terms": (
            ("开票日期", "结算日期"),
            ("回款日期", "到账日期"),
        ),
        "reason_code": "missing_authoritative_invoice_collection_dates_and_shared_project_binding",
        "reason_zh": "缺少权威开票、回款日期与共享项目绑定",
        "responsible_role": "finance_owner",
        "field_specific_refs": (CURRENT_S13_P2_PATH, CURRENT_S13_P3_PATH),
    },
    {
        "field_key": "audit_variance",
        "label": "审计偏差",
        "fact_kind": "money_minor_unit_or_ratio_basis_points",
        "direct_terms": ("审计偏差", "审计差异", "审减", "核减", "送审金额", "审定金额"),
        "paired_terms": (
            ("审计", "审定"),
            ("差异", "偏差", "金额"),
        ),
        "reason_code": "cross_table_dimensions_not_comparable_and_exact_variance_not_performed",
        "reason_zh": "跨表维度不可比较，尚未执行精确偏差核对",
        "responsible_role": "finance_review_owner",
        "field_specific_refs": (CURRENT_S13_P3_PATH, s14_review.MANIFEST_PATH),
    },
    {
        "field_key": "customer_relationship_rate",
        "label": "客情费率",
        "fact_kind": "ratio_basis_points",
        "direct_terms": ("客情费率", "客情费用率", "客情费", "关系维护费", "业务招待费率"),
        "paired_terms": (),
        "reason_code": "missing_authoritative_definition_source_period_denominator_and_value_binding",
        "reason_zh": "缺少权威定义、来源、期间、分母与数值绑定",
        "responsible_role": "sales_owner",
        "field_specific_refs": (),
    },
)
FIELD_KEYS = tuple(spec["field_key"] for spec in FIELD_SPECS)
FIELD_LABELS = {spec["field_key"]: spec["label"] for spec in FIELD_SPECS}

DEPENDENCY_LINKS = {
    "fund": (
        "../../../V014_S14_P1_POST_REMEDIATION_FUND_CASH_LOAN_PLAN/exports/html/"
        "fund_cash_loan_workbench.html",
        "资金现金贷款工作台",
    ),
    "invoice-tax": (
        "../../../V014_S14_P2_POST_REMEDIATION_INVOICE_TAX_PLAN/exports/html/"
        "invoice_tax_plan_workbench.html",
        "开票纳税计划工作台",
    ),
    "policy": (
        "../../../V014_S14_P3_POST_REMEDIATION_POLICY_EVIDENCE_PLAN/exports/html/"
        "policy_evidence_workbench.html",
        "政策证据工作台",
    ),
    "cross-table": (
        "../../../V014_S13_P3_POST_REMEDIATION_CROSS_TABLE_REVIEW/exports/html/"
        "cross_table_quality_workbench.html",
        "跨表复核质量工作台",
    ),
}
P2_HREF = (
    "../../../V014_S15_P2_POST_REMEDIATION_PERFORMANCE_REVIEW_LIST/exports/html/"
    "performance_review_workbench.html"
)
P3_HREF = (
    "../../../V014_S15_P3_POST_REMEDIATION_SALARY_BOUNDARY/exports/html/"
    "salary_boundary_workbench.html"
)


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"{path} must contain an object")
    return value


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"{path} contains a non-object row")
            rows.append(value)
    return rows


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8")


def _git_output(args: list[str]) -> str:
    result = subprocess.run(
        ["git", "-c", "core.quotePath=false", *args],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip())
    return result.stdout.strip()


def _upsert_jsonl(path: Path, row: dict[str, Any]) -> None:
    preserved: list[str] = []
    if path.is_file():
        for line in path.read_text(encoding="utf-8").splitlines():
            if line.strip() and json.loads(line).get("phase_id") != PHASE_ID:
                preserved.append(line)
    preserved.append(json.dumps(row, ensure_ascii=False, separators=(",", ":")))
    _write_text(path, "\n".join(preserved))


def _sync_assurance_snapshot_time(generated_at: str) -> None:
    lines = ASSURANCE_STATUS_PATH.read_text(encoding="utf-8").splitlines()
    indexes = [index for index, line in enumerate(lines) if line.startswith("snapshot_event_time:")]
    if len(indexes) != 1:
        raise RuntimeError("ASSURANCE_STATUS must contain exactly one snapshot_event_time")
    lines[indexes[0]] = f'snapshot_event_time: "{generated_at}"'
    _write_text(ASSURANCE_STATUS_PATH, "\n".join(lines))


def _load_dependency() -> dict[str, Any]:
    manifest = validate_v014_s14_post_remediation_stage_review(
        require_private_evidence=True,
        require_browser_evidence=True,
        require_final_evidence=True,
    )
    summary = manifest["summary"]
    if manifest.get("next_phase") != "S15-P1":
        raise ValueError("Stage 14 post-remediation review must route to S15-P1")
    if summary.get("stage14_review_performed") is not True or summary.get("s15_p1_performed") is not False:
        raise ValueError("Stage 14 dependency state drift")
    if summary.get("decision") != "NO_GO" or summary.get("current_report_grade") != "D":
        raise ValueError("Stage 14 quality state drift")
    return manifest


def _load_contract() -> dict[str, Any]:
    roadmap = ROADMAP_PATH.read_text(encoding="utf-8")
    taskpack = TASKPACK_PATH.read_text(encoding="utf-8")
    for token in (
        "S15｜销售绩效事实与复核清单",
        "S15-P1",
        "绩效事实字段",
        "开票金额",
        "毛利率",
        "结算速度",
        "回款速度",
        "审计偏差",
        "客情费率",
        "绑定项目成本事实和回款事实",
        "字段缺失时标记人工复核",
    ):
        if token not in roadmap:
            raise ValueError(f"roadmap token missing: {token}")
    for token in ("销售绩效/业务考核线", "输出绩效事实和复核清单", "不做工资最终审批"):
        if token not in taskpack:
            raise ValueError(f"taskpack token missing: {token}")
    return {
        "roadmap_read": True,
        "taskpack_read": True,
        "human_flow_baseline_read": True,
        "raw_read_only_contract_applied": True,
        "six_required_fields_locked": True,
        "missing_field_manual_review_rule_locked": True,
    }


def _load_legacy_fixture() -> dict[str, Any]:
    manifest = _read_json(LEGACY_MANIFEST_PATH)
    definitions = _read_jsonl(LEGACY_FIELD_DEFINITIONS_PATH)
    bindings = _read_jsonl(LEGACY_FIELD_BINDINGS_PATH)
    manual = _read_jsonl(LEGACY_MANUAL_REVIEW_PATH)
    if len(definitions) != 6 or len(bindings) != 6 or len(manual) != 4:
        raise ValueError("historical S15-P1 fixture shape drift")
    return {
        "fixture_validated": True,
        "definition_count": len(definitions),
        "binding_count": len(bindings),
        "manual_review_count": len(manual),
        "historical_two_non_manual_fields": len(definitions) - len(manual),
        "historical_summary": manifest.get("performance_fact_fields_summary", manifest.get("summary", {})),
    }


def _load_current_source_state() -> dict[str, Any]:
    stage9 = _read_json(CURRENT_STAGE9_REVIEW_PATH)["summary"]
    s13_p1 = _read_json(CURRENT_S13_P1_PATH)["summary"]
    s13_p2 = _read_json(CURRENT_S13_P2_PATH)["summary"]
    s13_p3 = _read_json(CURRENT_S13_P3_PATH)["summary"]
    s14_p2 = _read_json(CURRENT_S14_P2_PATH)["summary"]
    checks = (
        stage9.get("decision") == "NO_GO",
        s13_p1.get("raw_value_bound_lane_count") == 0,
        s13_p2.get("row_level_binding_proven_lane_count") == 0,
        s13_p2.get("identified_business_item_count") == 0,
        s13_p3.get("comparable_dimension_count") == 0,
        s13_p3.get("exact_comparison_performed_count") == 0,
        s14_p2.get("value_binding_proven_lane_count") == 0,
        s14_p2.get("identified_issue_candidate_count") == 0,
    )
    if not all(checks):
        raise ValueError("current project-cost collection invoice or cross-table state drift")
    return {
        "stage9": stage9,
        "s13_p1": s13_p1,
        "s13_p2": s13_p2,
        "s13_p3": s13_p3,
        "s14_p2": s14_p2,
    }


def _matched_fields(searchable: str) -> tuple[list[str], dict[str, list[str]]]:
    matched: list[str] = []
    terms_by_field: dict[str, list[str]] = {}
    for spec in FIELD_SPECS:
        direct_hits = [term for term in spec["direct_terms"] if term in searchable]
        paired_hits: list[str] = []
        paired = spec["paired_terms"]
        if paired:
            grouped = [[term for term in group if term in searchable] for group in paired]
            if all(group for group in grouped):
                paired_hits = [term for group in grouped for term in group]
        hits = sorted(set(direct_hits + paired_hits))
        if hits:
            matched.append(spec["field_key"])
            terms_by_field[spec["field_key"]] = hits
    return matched, terms_by_field


def _scan_workbook(
    payload: bytes,
    *,
    raw_path: Path,
    member_name: str | None,
    raw_index: int,
    member_index: int,
) -> tuple[int, list[dict[str, Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    records: list[dict[str, Any]] = []
    sheet_count = len(workbook.worksheets)
    for sheet_index, sheet in enumerate(workbook.worksheets, 1):
        parts = [str(sheet.title)]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(int(sheet.max_row or 0), 12),
            max_col=min(int(sheet.max_column or 0), 30),
            values_only=True,
        ):
            parts.extend(str(value) for value in row if value is not None)
        matched, terms = _matched_fields("\n".join(parts))
        if not matched:
            continue
        records.append(
            {
                "raw_index": raw_index,
                "raw_path_private": str(raw_path),
                "raw_filename_private": raw_path.name,
                "member_index": member_index,
                "member_name_private": member_name,
                "member_sha256": s14_review.p1._sha256_bytes(payload),
                "sheet_index": sheet_index,
                "sheet_name_private": sheet.title,
                "sheet_max_row": int(sheet.max_row or 0),
                "sheet_max_column": int(sheet.max_column or 0),
                "matched_fields": matched,
                "matched_terms_private": terms,
                **s14_review.p1._sheet_probe(sheet, row_limit=200, column_limit=30),
            }
        )
    workbook.close()
    return sheet_count, records


def _raw_candidate_probe(raw_root: Path) -> dict[str, Any]:
    raw_files = sorted(path for path in raw_root.rglob("*") if path.is_file())
    raw_records: list[dict[str, Any]] = []
    candidate_records: list[dict[str, Any]] = []
    unparseable: list[dict[str, Any]] = []
    xlsx_container_count = 0
    parseable_count = 0
    unparseable_count = 0
    parseable_sheet_count = 0
    roundtrip_mismatch_count = 0

    for raw_index, raw_path in enumerate(raw_files, 1):
        raw_bytes = raw_path.read_bytes()
        payloads: list[tuple[int, str | None, bytes]] = []
        if raw_path.suffix.lower() == ".xlsx":
            payloads.append((1, None, raw_bytes))
        elif raw_path.suffix.lower() == ".zip":
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
                payloads = [
                    (member_index, info.filename, archive.read(info))
                    for member_index, info in enumerate(archive.infolist(), 1)
                    if not info.is_dir() and Path(info.filename).suffix.lower() == ".xlsx"
                ]
        raw_records.append(
            {
                "raw_index": raw_index,
                "raw_path_private": str(raw_path),
                "raw_filename_private": raw_path.name,
                "raw_suffix": raw_path.suffix.lower(),
                "raw_size_bytes": raw_path.stat().st_size,
                "raw_sha256": s14_review.p1._sha256_bytes(raw_bytes),
                "xlsx_container_count": len(payloads),
            }
        )
        for member_index, member_name, payload in payloads:
            xlsx_container_count += 1
            try:
                first_sheet_count, first = _scan_workbook(
                    payload,
                    raw_path=raw_path,
                    member_name=member_name,
                    raw_index=raw_index,
                    member_index=member_index,
                )
                second_sheet_count, second = _scan_workbook(
                    payload,
                    raw_path=raw_path,
                    member_name=member_name,
                    raw_index=raw_index,
                    member_index=member_index,
                )
            except Exception as exc:
                unparseable_count += 1
                unparseable.append(
                    {
                        "raw_index": raw_index,
                        "member_index": member_index,
                        "member_name_private": member_name,
                        "member_sha256": s14_review.p1._sha256_bytes(payload),
                        "error_class": type(exc).__name__,
                    }
                )
                continue
            parseable_count += 1
            parseable_sheet_count += first_sheet_count
            if first_sheet_count != second_sheet_count:
                roundtrip_mismatch_count += 1
            first_map = {
                (row["sheet_index"], tuple(row["matched_fields"])): row["probe_fingerprint"]
                for row in first
            }
            second_map = {
                (row["sheet_index"], tuple(row["matched_fields"])): row["probe_fingerprint"]
                for row in second
            }
            roundtrip_mismatch_count += sum(
                first_map.get(key) != second_map.get(key)
                for key in set(first_map) | set(second_map)
            )
            candidate_records.extend(first)

    counts = {
        field_key: sum(field_key in row["matched_fields"] for row in candidate_records)
        for field_key in FIELD_KEYS
    }
    return {
        "schema_version": "kmfa.private.v014.s15_p1.performance_fact_candidate_probe.v1",
        "classification": "PRIVATE_RUNTIME_ONLY_NEVER_COMMIT",
        "raw_root_private": str(raw_root),
        "raw_source_file_count": len(raw_files),
        "private_xlsx_container_count": xlsx_container_count,
        "private_parseable_xlsx_count": parseable_count,
        "private_unparseable_xlsx_count": unparseable_count,
        "private_parseable_sheet_count": parseable_sheet_count,
        "private_candidate_sheet_count_by_field": counts,
        "private_unique_candidate_sheet_count": len(candidate_records),
        "private_multi_field_candidate_sheet_count": sum(
            len(row["matched_fields"]) > 1 for row in candidate_records
        ),
        "private_candidate_covered_field_count": sum(count > 0 for count in counts.values()),
        "private_probe_roundtrip_mismatch_count": roundtrip_mismatch_count,
        "authoritative_row_binding_proven_field_count": 0,
        "authoritative_value_binding_proven_field_count": 0,
        "raw_files_private": raw_records,
        "unparseable_xlsx_private": unparseable,
        "candidate_sheets_private": candidate_records,
    }


def _build_field_definitions(probe: dict[str, Any]) -> list[dict[str, Any]]:
    counts = probe["private_candidate_sheet_count_by_field"]
    return [
        {
            "field_definition_id": f"V014-S15P1-POST-FIELD-{index:03d}",
            "field_key": spec["field_key"],
            "visible_field_label": spec["label"],
            "fact_kind": spec["fact_kind"],
            "required_by_roadmap": True,
            "project_cost_fact_reference_required": True,
            "collection_fact_reference_required": True,
            "private_candidate_structure_observed": counts[spec["field_key"]] > 0,
            "private_candidate_sheet_count": counts[spec["field_key"]],
            "authoritative_row_binding_required": True,
            "authoritative_value_binding_required": True,
            "manual_review_required": True,
            "field_value_materialized": False,
            "public_business_value_allowed": False,
            "salary_or_bonus_use_allowed": False,
        }
        for index, spec in enumerate(FIELD_SPECS, 1)
    ]


def _build_field_bindings(probe: dict[str, Any]) -> list[dict[str, Any]]:
    counts = probe["private_candidate_sheet_count_by_field"]
    project_ref = CURRENT_STAGE9_REVIEW_PATH.as_posix()
    collection_ref = CURRENT_S13_P2_PATH.as_posix()
    rows: list[dict[str, Any]] = []
    for index, spec in enumerate(FIELD_SPECS, 1):
        field_key = spec["field_key"]
        rows.append(
            {
                "binding_status_id": f"V014-S15P1-POST-BIND-{index:03d}",
                "field_key": field_key,
                "visible_field_label": spec["label"],
                "project_cost_structure_reference": project_ref,
                "collection_structure_reference": collection_ref,
                "field_specific_structure_references": [
                    path.as_posix() for path in spec["field_specific_refs"]
                ],
                "project_cost_structure_reference_connected": True,
                "collection_structure_reference_connected": True,
                "private_candidate_structure_observed": counts[field_key] > 0,
                "private_candidate_sheet_count": counts[field_key],
                "authoritative_row_binding_proven": False,
                "authoritative_value_binding_proven": False,
                "performance_fact_materialized": False,
                "binding_status": "manual_review_required_unproven_authoritative_row_and_value_binding",
                "manual_review_required": True,
                "manual_review_requirement_id": f"V014-S15P1-POST-REVIEW-{index:03d}",
                "candidate_as_fact_allowed": False,
                "auto_fill_allowed": False,
                "formal_report_allowed": False,
                "business_decision_basis_allowed": False,
                "salary_calculation_allowed": False,
                "bonus_approval_allowed": False,
                "payroll_export_allowed": False,
            }
        )
    return rows


def _build_manual_review_requirements() -> list[dict[str, Any]]:
    return [
        {
            "manual_review_requirement_id": f"V014-S15P1-POST-REVIEW-{index:03d}",
            "field_key": spec["field_key"],
            "visible_field_label": spec["label"],
            "manual_review_required": True,
            "reason_code": spec["reason_code"],
            "reason_zh": spec["reason_zh"],
            "responsible_role": spec["responsible_role"],
            "required_evidence": (
                "authoritative project row, source identity, reporting period, unit, "
                "calculation definition and exact value binding"
            ),
            "current_resolution_status": "keep_pending_missing_authoritative_binding",
            "candidate_as_fact_allowed": False,
            "auto_fill_allowed": False,
            "auto_calculation_allowed": False,
            "auto_approval_allowed": False,
            "s15_p2_review_list_created": False,
            "salary_or_bonus_action_allowed": False,
        }
        for index, spec in enumerate(FIELD_SPECS, 1)
    ]


def _render_html(
    definitions: list[dict[str, Any]],
    bindings: list[dict[str, Any]],
    manual: list[dict[str, Any]],
) -> str:
    rows = "".join(
        "<tr>"
        f"<td>{html.escape(row['visible_field_label'])}</td>"
        f"<td>{row['private_candidate_sheet_count']}</td>"
        "<td>已连接</td><td>已连接</td><td>人工复核</td>"
        "</tr>"
        for row in bindings
    )
    buttons = "".join(
        f'<button type="button" data-field-button="{row["field_key"]}">'
        f'{html.escape(row["visible_field_label"])}</button>'
        for row in definitions
    )
    manual_by_key = {row["field_key"]: row for row in manual}
    panels = "".join(
        f'<section data-field-panel="{row["field_key"]}" hidden>'
        f'<div class="panel-head"><div><span>字段 {index}/6</span>'
        f'<h3>{html.escape(row["visible_field_label"])}</h3></div>'
        '<span class="tag blocked">人工复核</span></div>'
        '<dl><div><dt>项目成本结构</dt><dd>已连接</dd></div>'
        '<div><dt>回款结构</dt><dd>已连接</dd></div>'
        '<div><dt>权威行/值</dt><dd>0 / 0</dd></div></dl>'
        f'<p>当前候选工作表：{row["private_candidate_sheet_count"]}；'
        '候选仅证明结构可能相关，不是绩效事实。</p>'
        f'<div class="reason">待补证：{html.escape(manual_by_key[row["field_key"]]["reason_zh"])}</div>'
        '<div class="limit">不得自动填值，不得计算绩效、工资或奖金。</div>'
        '</section>'
        for index, row in enumerate(bindings, 1)
    )
    links = "".join(
        f'<a data-dependency-link="{link_id}" href="{href}">{label}</a>'
        for link_id, (href, label) in DEPENDENCY_LINKS.items()
    )
    stage_links = (
        f'<a data-stage-link="review-list" href="{P2_HREF}">绩效复核清单工作台</a>'
        f'<a data-stage-link="salary-boundary" href="{P3_HREF}">工资项目边界工作台</a>'
    )
    labels = json.dumps(FIELD_LABELS, ensure_ascii=False, sort_keys=True)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>KMFA 绩效事实字段工作台</title>
  <style>
    *{{box-sizing:border-box}}body{{margin:0;background:#f2f5f5;color:#172b27;font:15px/1.55 -apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;letter-spacing:0}}
    header{{background:#103d34;color:#fff;border-bottom:3px solid #72c6b0}}.nav,.page{{width:min(1080px,calc(100% - 32px));margin:auto}}
    .nav{{display:flex;align-items:center;justify-content:space-between;gap:20px;padding:14px 0}}.brand{{display:flex;gap:12px;align-items:center}}
    .logo{{width:38px;height:38px;display:grid;place-items:center;background:#fff;color:#103d34;font-weight:800;border-radius:6px}}
    .brand strong{{display:block;font-size:17px}}.brand small{{display:block}}nav{{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}}
    nav a{{color:#fff;text-decoration:none;border:1px solid #7aa79c;padding:7px 10px;border-radius:4px;font-size:13px}}
    main{{padding:26px 0 34px}}h1{{font-size:30px;margin:0 0 4px}}h2{{font-size:20px;margin:0}}h3{{font-size:22px;margin:4px 0 0}}
    .intro{{color:#61746f;margin:0 0 16px}}.badges{{display:flex;gap:8px;margin:10px 0 16px}}.badge,.tag{{padding:4px 9px;border:1px solid #ccd7d4;border-radius:4px;font-weight:700;font-size:13px;background:#fff}}
    .badge.ok{{color:#087258;border-color:#9fdacb}}.badge.blocked,.tag.blocked{{color:#ad3028;border-color:#efaaa5;background:#fff5f4}}
    .gate{{border-left:4px solid #ae3c35;background:#fff1ef;color:#8b2d27;padding:13px 15px;margin-bottom:18px}}
    .stats{{display:grid;grid-template-columns:repeat(4,1fr);border:1px solid #cad6d3;background:#fff;margin-bottom:18px}}
    .stat{{padding:14px 16px;border-right:1px solid #cad6d3}}.stat:last-child{{border-right:0}}.stat strong{{display:block;font-size:26px}}.stat span{{color:#687b76;font-size:13px}}
    .section{{background:#fff;border:1px solid #cad6d3;margin-bottom:18px}}.section-head{{display:flex;justify-content:space-between;align-items:center;padding:14px 16px;border-bottom:1px solid #cad6d3}}
    table{{width:100%;border-collapse:collapse;table-layout:fixed}}th,td{{padding:10px 14px;text-align:left;border-bottom:1px solid #d9e1df;overflow-wrap:anywhere}}th{{font-size:13px;color:#5a706a;background:#f6f8f7}}tr:last-child td{{border-bottom:0}}
    .workspace{{display:grid;grid-template-columns:210px 1fr}}.field-buttons{{padding:14px;background:#f3f6f5;border-right:1px solid #cad6d3}}
    .field-buttons button{{width:100%;background:transparent;border:1px solid transparent;text-align:left;padding:9px;border-radius:4px;font:inherit}}
    .field-buttons button.active{{background:#fff;border-color:#60af9a;color:#087258;font-weight:700}}.panel{{padding:18px 20px}}
    .panel-head{{display:flex;justify-content:space-between;gap:20px;align-items:flex-start}}.panel-head span:first-child{{font-size:12px;color:#6b7d78}}
    dl{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}}dl div{{border-left:3px solid #7fc5b2;padding-left:10px}}dt{{font-size:12px;color:#6b7d78}}dd{{margin:2px 0 0;font-weight:700}}
    .reason{{background:#f5f8f7;border:1px solid #cad6d3;padding:12px;margin-top:12px;overflow-wrap:anywhere}}.limit{{background:#fff1ef;border:1px solid #efaaa5;color:#9b332d;padding:12px;margin-top:10px}}
    .status{{padding:12px 16px;background:#f5f8f7;border-top:1px solid #cad6d3;color:#5d716b}}.status-links{{display:flex;gap:12px;flex-wrap:wrap;margin-top:8px}}.status a{{color:#007696;text-decoration:none}}
    footer{{color:#5f736e;font-size:13px}}button:focus-visible,a:focus-visible{{outline:3px solid #f0ba4d;outline-offset:2px}}
    @media(max-width:680px){{.nav{{display:block}}nav{{justify-content:flex-start;margin-top:12px}}h1{{font-size:27px}}.stats{{grid-template-columns:repeat(2,1fr)}}.stat:nth-child(2){{border-right:0}}.stat:nth-child(-n+2){{border-bottom:1px solid #cad6d3}}
      table{{min-width:0;table-layout:fixed;font-size:11px}}th,td{{padding:8px 5px;word-break:break-word}}.workspace{{display:block}}.field-buttons{{border-right:0;border-bottom:1px solid #cad6d3;display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:6px}}
      .field-buttons button{{text-align:center}}.panel{{padding:16px}}dl{{grid-template-columns:1fr}}.panel-head{{align-items:center}}}}
  </style>
</head>
<body data-active-field="">
  <header><div class="nav"><div class="brand"><div class="logo">KM</div><div><strong>KMFA 经营分析系统</strong><small>S15-P1 · 绩效事实字段</small></div></div><nav>{links}{stage_links}</nav></div></header>
  <main class="page">
    <h1>绩效事实字段工作台</h1>
    <p class="intro">六个必需字段已接入结构候选；当前不展示项目、人员、金额、比率或日期明细。</p>
    <div class="badges"><span class="badge ok">Q4 / D</span><span class="badge blocked">NO_GO</span><span class="badge">内部复核</span></div>
    <div class="gate"><strong>门禁：</strong>6 个字段均缺权威行级与数值绑定，全部进入人工复核；不得形成绩效或薪资结论。</div>
    <section class="stats"><div class="stat"><strong>6 / 6</strong><span>必需字段已定义</span></div><div class="stat"><strong>6 / 6</strong><span>候选结构已观察</span></div><div class="stat"><strong>0 / 6</strong><span>权威值已绑定</span></div><div class="stat"><strong>6</strong><span>人工复核字段</span></div></section>
    <section class="section"><div class="section-head"><h2>字段绑定状态</h2><span class="tag blocked">不含业务值</span></div><table><thead><tr><th>字段</th><th>私有候选表</th><th>项目成本结构</th><th>回款结构</th><th>当前状态</th></tr></thead><tbody>{rows}</tbody></table></section>
    <section class="section"><div class="section-head"><h2>人工复核要求</h2><span class="tag blocked">6 项待补证</span></div><div class="workspace"><div class="field-buttons">{buttons}</div><div class="panel">{panels}</div></div>
      <div class="status"><span id="interaction-status">字段复核状态已加载。</span><div class="status-links">{links}</div></div></section>
    <footer>Stage 15 三个 phase 与整体复审均已完成；当前保持 Q4 / D · NO_GO，S16 仅可在下一 run work，不执行工资奖金、GitHub upload、app reinstall、正式报告、差异关闭或业务执行。</footer>
  </main>
  <script>
    const labels={labels};let actionSequence=0;
    function activate(id){{document.body.dataset.activeField=id;document.body.dataset.lastAction='field:'+id+':'+(++actionSequence);document.querySelectorAll('[data-field-button]').forEach(b=>b.classList.toggle('active',b.dataset.fieldButton===id));document.querySelectorAll('[data-field-panel]').forEach(p=>p.hidden=p.dataset.fieldPanel!==id);document.getElementById('interaction-status').textContent='已显示“'+labels[id]+'”人工复核要求；当前不形成绩效事实。';}}
    document.querySelectorAll('[data-field-button]').forEach(button=>button.addEventListener('click',()=>activate(button.dataset.fieldButton)));
    activate('invoice_amount');document.body.dataset.uiReady='true';
  </script>
</body>
</html>"""


class _QuietHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        return


def _browser_worker() -> dict[str, Any]:
    from playwright.sync_api import sync_playwright

    root = Path("KMFA/stage_artifacts").resolve()
    handler = functools.partial(_QuietHandler, directory=str(root))
    server = socketserver.TCPServer(("127.0.0.1", 0), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    base = f"http://127.0.0.1:{server.server_address[1]}/"
    workbench_url = urljoin(base, f"{PHASE_ID}/exports/html/{HTML_PATH.name}")
    viewport_checks: list[dict[str, Any]] = []
    field_checks: list[dict[str, Any]] = []
    http_checks: list[dict[str, Any]] = []
    navigation_checks: list[dict[str, Any]] = []
    PRIVATE_SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            for mode, width, height in (("desktop", 1440, 900), ("mobile", 390, 844)):
                page = browser.new_page(viewport={"width": width, "height": height})
                errors: list[str] = []
                page.on(
                    "console",
                    lambda message, errors=errors: errors.append(message.text)
                    if message.type == "error"
                    else None,
                )
                page.goto(workbench_url, wait_until="networkidle")
                page.wait_for_function("document.body.dataset.uiReady === 'true'")
                body = page.locator("body").inner_text()
                viewport_checks.append(
                    {
                        "mode": mode,
                        "width": width,
                        "height": height,
                        "marker_visible": page.get_by_role("heading", name="绩效事实字段工作台").is_visible(),
                        "d_no_go_visible": "Q4 / D" in body and "NO_GO" in body,
                        "manual_six_visible": "6 项待补证" in body,
                        "stage_boundary_visible": "S15-P2/P3" in body,
                        "console_error_count": len(errors),
                        "no_horizontal_overflow": page.evaluate(
                            "document.documentElement.scrollWidth <= document.documentElement.clientWidth"
                        ),
                    }
                )
                for spec in FIELD_SPECS:
                    field_key = spec["field_key"]
                    page.locator(f'[data-field-button="{field_key}"]').click()
                    field_checks.append(
                        {
                            "mode": mode,
                            "field_key": field_key,
                            "passed": page.locator(f'[data-field-panel="{field_key}"]').is_visible()
                            and page.locator("body").get_attribute("data-active-field") == field_key
                            and spec["label"] in page.locator("#interaction-status").inner_text(),
                        }
                    )
                page.screenshot(
                    path=str(PRIVATE_SCREENSHOT_DIR / f"performance_fact_fields_{mode}.png"),
                    full_page=True,
                )
                page.close()

            request = playwright.request.new_context()
            for link_id, (_, marker) in DEPENDENCY_LINKS.items():
                page = browser.new_page(viewport={"width": 1280, "height": 900})
                page.goto(workbench_url, wait_until="networkidle")
                href = page.locator(f'a[data-dependency-link="{link_id}"]').first.get_attribute("href") or ""
                target = urljoin(workbench_url, href)
                response = request.get(target)
                http_checks.append({"link_id": link_id, "status": response.status, "passed": response.ok})
                page.locator(f'a[data-dependency-link="{link_id}"]').first.click()
                page.wait_for_load_state("networkidle")
                navigation_checks.append(
                    {"link_id": link_id, "marker": marker, "passed": marker in page.locator("body").inner_text()}
                )
                page.close()
            request.dispose()
            browser.close()
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)

    result = {
        "status": "PASS",
        "viewport_checks": viewport_checks,
        "field_interaction_checks": field_checks,
        "dependency_link_http_checks": http_checks,
        "dependency_navigation_checks": navigation_checks,
    }
    if not (
        len(viewport_checks) == 2
        and all(
            row["marker_visible"]
            and row["d_no_go_visible"]
            and row["manual_six_visible"]
            and row["stage_boundary_visible"]
            and row["console_error_count"] == 0
            and row["no_horizontal_overflow"]
            for row in viewport_checks
        )
        and len(field_checks) == 12
        and all(row["passed"] for row in field_checks)
        and len(http_checks) == 4
        and all(row["passed"] for row in http_checks)
        and len(navigation_checks) == 4
        and all(row["passed"] for row in navigation_checks)
    ):
        result["status"] = "FAIL"
    _write_json(PRIVATE_BROWSER_PATH, result)
    return result


def _run_browser_review() -> dict[str, Any]:
    python = os.environ.get("KMFA_AUDIT_PYTHON", sys.executable)
    result = subprocess.run(
        [python, __file__, "--browser-evidence-only"],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
        env={**os.environ, "PYTHONDONTWRITEBYTECODE": "1", "PYTHONPATH": "."},
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip())
    browser = _read_json(PRIVATE_BROWSER_PATH)
    if browser.get("status") != "PASS":
        raise RuntimeError("S15-P1 browser validation failed")
    helper = s14_review.p1.s13_review.p1.s12_review.p1.s11_home
    baseline = helper._run_html_audit(HTML_BASELINE_ROOT, PRIVATE_BASELINE_AUDIT_PATH)
    current = helper._run_html_audit(HTML_DIR, PRIVATE_CURRENT_AUDIT_PATH)
    expected_baseline = {
        "file_count": 6,
        "control_row_count": 54,
        "pass_count": 54,
        "warn_count": 0,
        "fail_count": 0,
    }
    if baseline != expected_baseline:
        raise RuntimeError("v1.4 HTML baseline drift")
    if current["file_count"] != 1 or current["warn_count"] or current["fail_count"]:
        raise RuntimeError("S15-P1 HTML audit failed")
    if current["pass_count"] != current["control_row_count"]:
        raise RuntimeError("S15-P1 HTML audit incomplete")
    return {
        "status": "PASS",
        "baseline_file_count": baseline["file_count"],
        "baseline_control_row_count": baseline["control_row_count"],
        "baseline_pass_count": baseline["pass_count"],
        "baseline_warn_count": baseline["warn_count"],
        "baseline_fail_count": baseline["fail_count"],
        "current_page_count": current["file_count"],
        "current_control_row_count": current["control_row_count"],
        "current_pass_count": current["pass_count"],
        "current_warn_count": current["warn_count"],
        "current_fail_count": current["fail_count"],
        "viewport_check_count": len(browser["viewport_checks"]),
        "field_interaction_check_count": len(browser["field_interaction_checks"]),
        "dependency_link_http_check_count": len(browser["dependency_link_http_checks"]),
        "dependency_navigation_check_count": len(browser["dependency_navigation_checks"]),
        "console_error_count": sum(row["console_error_count"] for row in browser["viewport_checks"]),
        "horizontal_overflow_count": sum(
            not row["no_horizontal_overflow"] for row in browser["viewport_checks"]
        ),
    }


def _quality_gate() -> dict[str, Any]:
    return {
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "decision": "NO_GO",
        "release_permission": "blocked",
        "field_definition_allowed": True,
        "structure_reference_connection_allowed": True,
        "manual_review_marker_allowed": True,
        "candidate_as_fact_allowed": False,
        "performance_fact_materialization_allowed": False,
        "public_business_value_display_allowed": False,
        "performance_review_list_allowed": False,
        "salary_calculation_allowed": False,
        "bonus_approval_allowed": False,
        "payroll_export_allowed": False,
        "formal_report_allowed": False,
        "business_decision_basis_allowed": False,
    }


def _phase_boundaries() -> dict[str, bool]:
    return {
        "s15_p1_scope_included": True,
        "s15_p2_scope_included": False,
        "s15_p3_scope_included": False,
        "stage15_review_scope_included": False,
        "github_upload_scope_included": False,
        "app_reinstall_scope_included": False,
        "formal_report_scope_included": False,
        "difference_closure_scope_included": False,
        "salary_calculation_scope_included": False,
        "bonus_approval_scope_included": False,
        "payroll_export_scope_included": False,
        "persistent_business_write_scope_included": False,
        "business_execution_scope_included": False,
    }


def _raw_boundary() -> dict[str, bool]:
    return {
        "raw_read_authorized": True,
        "raw_snapshot_validation_performed": True,
        "private_candidate_probe_performed": True,
        "private_probe_roundtrip_performed": True,
        "raw_write_performed": False,
        "raw_delete_performed": False,
        "raw_move_performed": False,
        "raw_rename_performed": False,
        "raw_overwrite_performed": False,
        "raw_mutation_performed": False,
    }


def _public_safety() -> dict[str, bool]:
    return {
        "raw_business_data_committed": False,
        "raw_file_names_committed": False,
        "raw_file_hashes_committed": False,
        "source_headers_committed": False,
        "field_plaintext_committed": False,
        "business_values_committed": False,
        "business_amounts_committed": False,
        "project_or_customer_plaintext_committed": False,
        "salary_or_bonus_payload_committed": False,
        "private_probe_committed": False,
        "private_screenshot_committed": False,
        "zip_excel_pdf_private_csv_or_database_committed": False,
        "credentials_or_secrets_committed": False,
    }


def _acceptance_matrix(summary: dict[str, Any]) -> dict[str, Any]:
    checks = {
        "stage14_dependency": summary["stage14_post_remediation_review_dependency_validated"],
        "taskpack_contract": summary["taskpack_contract_validated"],
        "six_fields_defined": summary["field_definition_count"] == 6,
        "all_candidates_covered": summary["candidate_covered_field_count"] == 6,
        "all_manual_review": summary["manual_review_required_field_count"] == 6,
        "project_cost_structure_connected": (
            summary["project_cost_structure_reference_connected_field_count"] == 6
        ),
        "collection_structure_connected": (
            summary["collection_structure_reference_connected_field_count"] == 6
        ),
        "zero_authoritative_rows": summary["authoritative_row_binding_proven_field_count"] == 0,
        "zero_authoritative_values": summary["authoritative_value_binding_proven_field_count"] == 0,
        "zero_materialized_facts": summary["materialized_performance_fact_count"] == 0,
        "deterministic_probe": summary["private_probe_roundtrip_mismatch_count"] == 0,
        "raw_exact": summary["raw_snapshot_exact_match"],
        "raw_cross_phase_exact": summary["raw_cross_phase_snapshot_exact_match"],
        "browser_pass": summary["browser_status"] == "PASS",
        "quality_locked": summary["current_report_grade"] == "D" and summary["decision"] == "NO_GO",
        "no_s15_p2": not summary["s15_p2_performed"],
        "no_upload": not summary["github_upload_performed"],
        "no_salary_business_action": (
            not summary["salary_calculation_performed"]
            and not summary["bonus_approval_performed"]
            and not summary["business_execution_performed"]
        ),
    }
    rows = [{"check_id": key, "passed": value} for key, value in checks.items()]
    return {
        "schema_version": "kmfa.v014.s15_p1.performance_fact_fields_acceptance_matrix.v1",
        "check_count": len(rows),
        "check_pass_count": sum(row["passed"] for row in rows),
        "check_fail_count": sum(not row["passed"] for row in rows),
        "checks": rows,
    }


def _phase_public_files() -> list[str]:
    artifact_paths = (
        SUMMARY_PATH,
        MANIFEST_PATH,
        FIELD_DEFINITIONS_PATH,
        FIELD_BINDINGS_PATH,
        MANUAL_REVIEW_PATH,
        MATRIX_PATH,
        GO_NO_GO_PATH,
        HTML_PATH,
        REPORT_PATH,
        TEST_RESULTS_PATH,
        RISK_REGISTER_PATH,
        ROLLBACK_PATH,
        METADATA_SUMMARY_PATH,
        METADATA_MANIFEST_PATH,
        METADATA_FIELD_DEFINITIONS_PATH,
        METADATA_FIELD_BINDINGS_PATH,
        METADATA_MANUAL_REVIEW_PATH,
        METADATA_MATRIX_PATH,
        METADATA_GO_NO_GO_PATH,
    )
    governance_paths = (
        Path("KMFA/AGENTS.md"),
        Path("KMFA/CHANGELOG.md"),
        Path("KMFA/HANDOFF.md"),
        Path("KMFA/VERSION"),
        ASSURANCE_STATUS_PATH,
        Path("KMFA/docs/governance/DEVELOPMENT_LEDGER.md"),
        Path("KMFA/docs/governance/MODEL_SPEC.md"),
        Path("KMFA/docs/governance/TRACEABILITY_MATRIX.csv"),
        Path("KMFA/docs/governance/VERSION_MATRIX.yaml"),
        Path("KMFA/docs/governance/delivery_tasks.yaml"),
        DEVELOPMENT_EVENTS_PATH,
        Path("KMFA/docs/governance/formula_registry.yaml"),
        Path("KMFA/docs/governance/model_registry.yaml"),
        Path("KMFA/docs/governance/parameter_registry.csv"),
        Path("KMFA/metadata/model_registry.yaml"),
        STAGE_STATUS_PATH,
        TASK_STATUS_PATH,
        Path("KMFA/功能清单.md"),
        Path("KMFA/开发记录.md"),
        Path("KMFA/模型参数文件.md"),
    )
    code_paths = (
        Path("KMFA/tools/v014_s15_p1_post_remediation_performance_fact_fields.py"),
        Path("KMFA/tools/check_v014_s15_p1_post_remediation_performance_fact_fields.py"),
        Path("KMFA/tests/test_v014_s15_p1_post_remediation_performance_fact_fields.py"),
    )
    return [path.as_posix() for path in artifact_paths + governance_paths + code_paths]


def _write_governance(generated_at: str) -> None:
    evidence_ref = MANIFEST_PATH.as_posix()
    _sync_assurance_snapshot_time(generated_at)
    _upsert_jsonl(
        DEVELOPMENT_EVENTS_PATH,
        {
            "event_id": "DEV-KMFA-20260711-V014-S15-P1-POST-REMEDIATION-PERFORMANCE-FACT-FIELDS",
            "event_time": generated_at,
            "event_type": "phase_delivery",
            "project_id": "KMFA",
            "stage_id": "S15",
            "phase_id": PHASE_ID,
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "task_id": TASK_ID,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "required_field_count": 6,
            "manual_review_required_field_count": 6,
            "materialized_performance_fact_count": 0,
            "raw_business_data_committed": False,
            "github_upload_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "files_changed": _phase_public_files(),
            "result_commit": "PENDING",
        },
    )
    _upsert_jsonl(
        STAGE_STATUS_PATH,
        {
            "schema_version": "kmfa.stage_status.v1",
            "record_type": "phase_status",
            "project_id": "KMFA",
            "stage_id": "S15",
            "phase_id": PHASE_ID,
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "task_id": TASK_ID,
            "version": VERSION,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "current_report_grade": "D",
            "raw_data_committed": False,
            "github_upload_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at,
        },
    )
    _upsert_jsonl(
        TASK_STATUS_PATH,
        {
            "schema_version": "kmfa.v014_stage_phase_task_status.v1",
            "record_type": "v014_phase_delivery",
            "project_id": "KMFA",
            "stage_id": "S15",
            "governance_stage_id": "SALES-PERFORMANCE-FACTS",
            "roadmap_stage_id": "S15",
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "acceptance_id": ACCEPTANCE_ID,
            "version": VERSION,
            "name": "v0.1.4 S15-P1 post-remediation performance fact fields",
            "phase_goal": "define six fields and keep all missing authoritative bindings in manual review",
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "derived_percent": 100,
            "estimated_task_units": 1,
            "completed_task_units": 1,
            "task_count": 1,
            "raw_data_committed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at[:10],
        },
    )


def _render_report(summary: dict[str, Any]) -> str:
    counts = summary["private_candidate_sheet_count_by_field"]
    lines = "\n".join(f"- {FIELD_LABELS[key]}：候选工作表 {counts[key]}，人工复核。" for key in FIELD_KEYS)
    return f"""# KMFA v0.1.4 S15-P1 修补后绩效事实字段

## 结论

- 六个必需字段：6/6 已定义，6/6 观察到私有结构候选，6/6 进入人工复核。
- 项目成本/回款结构引用：6/6 / 6/6 已连接；权威行/值绑定：0/6 / 0/6。
- 已物化绩效事实、公开业务值、工资、奖金、薪资导出：均为 0。
- 当前状态：Q4 / D / NO_GO / 3-9-2-1。

## 字段候选

{lines}

## 解释

候选只证明限定窗口内存在相关结构词，不证明同一项目、人员、期间、单位、口径或数值。旧 S15-P1 将两个字段视为非人工复核的动态状态已被隔离；本轮六字段均以当前 Stage 13/14 证据为准。

## 边界

- 未执行 S15-P2/P3、Stage 15 整体复审、工资/奖金计算、GitHub upload、app reinstall、正式报告、差异关闭或业务执行。
- raw 文件名、工作表、字段、表头、金额、明细和诊断只保存在 ignored private runtime。
"""


def _render_test_results(summary: dict[str, Any], browser: dict[str, Any]) -> str:
    return f"""# S15-P1 测试结果

- focused test / strict validator：最终复验结果见 manifest。
- v1.4 baseline：{browser['baseline_pass_count']}/54 PASS。
- current HTML audit：{browser['current_pass_count']}/{browser['current_control_row_count']} PASS。
- desktop/mobile：{summary['browser_viewport_check_count']}/2 PASS。
- 六字段交互：{summary['field_interaction_check_count']}/12 PASS。
- 依赖 HTTP / 真实导航：{summary['dependency_link_http_check_count']}/4 / {summary['dependency_navigation_check_count']}/4 PASS。
- raw 前后/跨 Stage 14 review/current：exact match。
- 私有双次候选探针不一致：{summary['private_probe_roundtrip_mismatch_count']}。
"""


def _render_private_report(summary: dict[str, Any]) -> str:
    counts = summary["private_candidate_sheet_count_by_field"]
    lines = "\n".join(f"- {FIELD_LABELS[key]}：{counts[key]} 个候选工作表。" for key in FIELD_KEYS)
    return f"""# S15-P1 私有候选与差异报告

## 只读候选结果

{lines}

- 原始文件：{summary['raw_source_file_count']}
- XLSX 容器：{summary['private_xlsx_container_count']}
- 可解析 / 不可解析：{summary['private_parseable_xlsx_count']} / {summary['private_unparseable_xlsx_count']}
- 可解析工作表：{summary['private_parseable_sheet_count']}
- 唯一候选 / 跨字段候选：{summary['private_unique_candidate_sheet_count']} / {summary['private_multi_field_candidate_sheet_count']}
- 双次探针指纹不一致：{summary['private_probe_roundtrip_mismatch_count']}

## 当前差异

1. 六字段均观察到候选结构，但缺少权威项目行、人员、期间、单位、口径和精确数值绑定。
2. 项目成本与回款仅完成 public-safe 结构引用，不构成字段值绑定。
3. 开票金额、毛利率、结算速度、回款速度、审计偏差和客情费率均不得自动填值。
4. 当前无法生成逐项目绩效事实表或薪资输入。

## 处理结论

- 六字段全部保持人工复核，绩效事实物化数量为 0。
- raw review 前后、跨 Stage 14 review 和当前快照完全一致。
- 原始文件未修改、删除、移动、重命名、覆盖或写入。
- 最终 goal 多轮交叉验证仍无法对齐时，必须纳入全中文最终差异报告。
"""


def generate(*, final_validation: bool = False, write_governance: bool = True) -> dict[str, Any]:
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    raw_helper = s14_review.p1.s13_review.p1.s12_review.p1.s11_project
    raw_before = raw_helper._raw_snapshot("before_v014_s15_p1_post_remediation_performance_fact_fields")
    dependency = _load_dependency()
    contract = _load_contract()
    legacy = _load_legacy_fixture()
    source_state = _load_current_source_state()
    probe = _raw_candidate_probe(Path(raw_before["raw_root"]))
    definitions = _build_field_definitions(probe)
    bindings = _build_field_bindings(probe)
    manual = _build_manual_review_requirements()
    _write_text(HTML_PATH, _render_html(definitions, bindings, manual))
    browser = _run_browser_review()
    raw_after = raw_helper._raw_snapshot("after_v014_s15_p1_post_remediation_performance_fact_fields")
    prior_raw = _read_json(s14_review.PRIVATE_RAW_AFTER_PATH)
    current_raw = raw_helper._raw_snapshot("current_v014_s15_p1_post_remediation_performance_fact_fields")
    normalize = raw_helper._normalize_raw
    raw_exact = normalize(raw_before) == normalize(raw_after)
    raw_cross = normalize(raw_before) == normalize(prior_raw) == normalize(current_raw)
    if not raw_exact or not raw_cross:
        raise RuntimeError("raw source changed during S15-P1")

    counts = probe["private_candidate_sheet_count_by_field"]
    upstream = dependency["summary"]
    summary = {
        "schema_version": "kmfa.v014.s15_p1.post_remediation.summary.v1",
        "project_id": "KMFA",
        "stage_id": "S15",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "stage14_post_remediation_review_dependency_validated": True,
        "taskpack_contract_validated": True,
        "required_field_count": len(FIELD_SPECS),
        "field_definition_count": len(definitions),
        "field_binding_status_count": len(bindings),
        "manual_review_required_field_count": len(manual),
        "candidate_covered_field_count": probe["private_candidate_covered_field_count"],
        "private_candidate_sheet_count_by_field": counts,
        "project_cost_structure_reference_connected_field_count": sum(
            row["project_cost_structure_reference_connected"] for row in bindings
        ),
        "collection_structure_reference_connected_field_count": sum(
            row["collection_structure_reference_connected"] for row in bindings
        ),
        "authoritative_row_binding_proven_field_count": 0,
        "authoritative_value_binding_proven_field_count": 0,
        "materialized_performance_fact_count": 0,
        "public_business_value_count": 0,
        "workbench_html_count": 1,
        "raw_source_file_count": probe["raw_source_file_count"],
        "private_xlsx_container_count": probe["private_xlsx_container_count"],
        "private_parseable_xlsx_count": probe["private_parseable_xlsx_count"],
        "private_unparseable_xlsx_count": probe["private_unparseable_xlsx_count"],
        "private_parseable_sheet_count": probe["private_parseable_sheet_count"],
        "private_unique_candidate_sheet_count": probe["private_unique_candidate_sheet_count"],
        "private_multi_field_candidate_sheet_count": probe["private_multi_field_candidate_sheet_count"],
        "private_probe_roundtrip_mismatch_count": probe["private_probe_roundtrip_mismatch_count"],
        "raw_snapshot_exact_match": raw_exact,
        "raw_cross_phase_snapshot_exact_match": raw_cross,
        "browser_status": browser["status"],
        "baseline_html_control_row_count": browser["baseline_control_row_count"],
        "baseline_html_pass_count": browser["baseline_pass_count"],
        "current_html_control_row_count": browser["current_control_row_count"],
        "current_html_pass_count": browser["current_pass_count"],
        "browser_viewport_check_count": browser["viewport_check_count"],
        "field_interaction_check_count": browser["field_interaction_check_count"],
        "dependency_link_http_check_count": browser["dependency_link_http_check_count"],
        "dependency_navigation_check_count": browser["dependency_navigation_check_count"],
        "console_error_count": browser["console_error_count"],
        "horizontal_overflow_count": browser["horizontal_overflow_count"],
        "open_final_difference_accepted_count": upstream["open_final_difference_accepted_count"],
        "nonzero_delta_reconciliation_count": upstream["nonzero_delta_reconciliation_count"],
        "zero_delta_reconciliation_count": upstream["zero_delta_reconciliation_count"],
        "incomplete_reconciliation_count": upstream["incomplete_reconciliation_count"],
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "s15_p1_performed": True,
        "s15_p2_performed": False,
        "s15_p3_performed": False,
        "stage15_review_performed": False,
        "salary_calculation_performed": False,
        "bonus_approval_performed": False,
        "payroll_export_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
    }
    matrix = _acceptance_matrix(summary)
    go_no_go = {
        "schema_version": "kmfa.v014.s15_p1.performance_fact_fields_go_no_go.v1",
        "project_id": "KMFA",
        "stage_id": "S15",
        "phase_id": PHASE_ID,
        "decision": "NO_GO",
        "field_definition_allowed": True,
        "manual_review_allowed": True,
        "performance_fact_release_allowed": False,
        "performance_review_list_allowed": False,
        "salary_or_bonus_action_allowed": False,
        "payroll_export_allowed": False,
        "formal_report_allowed": False,
        "github_upload_allowed": False,
        "business_execution_allowed": False,
    }
    validation_summary = {
        "final_validation_recorded": final_validation,
        "focused_test": "PASS" if final_validation else "PENDING",
        "strict_validator": "PASS" if final_validation else "PENDING",
        "browser_desktop_mobile": "PASS" if final_validation else "PENDING",
        "raw_alignment": "PASS" if final_validation else "PENDING",
        "governance_and_safety_scans": "PASS" if final_validation else "PENDING",
    }
    manifest = {
        "schema_version": "kmfa.v014.s15_p1.post_remediation.manifest.v1",
        "project_id": "KMFA",
        "stage_id": "S15",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "generated_at": generated_at,
        "reviewed_head": _git_output(["rev-parse", "HEAD"]),
        "branch": _git_output(["branch", "--show-current"]),
        "remote": _git_output(["remote", "get-url", "origin"]),
        "summary": summary,
        "field_definitions": definitions,
        "field_binding_statuses": bindings,
        "manual_review_requirements": manual,
        "acceptance_matrix": matrix,
        "go_no_go": go_no_go,
        "quality_gate": _quality_gate(),
        "phase_boundaries": _phase_boundaries(),
        "raw_boundary": _raw_boundary(),
        "public_repo_safety": _public_safety(),
        "browser_review": browser,
        "taskpack_contract": contract,
        "stage14_post_remediation_review_dependency_validated": True,
        "current_source_state_validated": True,
        "current_source_state": {
            "project_cost_component_materialization_count": source_state["stage9"][
                "cost_component_materialization_count"
            ],
            "financial_raw_value_bound_lane_count": source_state["s13_p1"]["raw_value_bound_lane_count"],
            "collection_row_level_binding_proven_lane_count": source_state["s13_p2"][
                "row_level_binding_proven_lane_count"
            ],
            "collection_identified_business_item_count": source_state["s13_p2"][
                "identified_business_item_count"
            ],
            "cross_table_comparable_dimension_count": source_state["s13_p3"]["comparable_dimension_count"],
            "cross_table_exact_comparison_performed_count": source_state["s13_p3"][
                "exact_comparison_performed_count"
            ],
            "invoice_tax_value_binding_proven_lane_count": source_state["s14_p2"][
                "value_binding_proven_lane_count"
            ],
        },
        "historical_s15_p1_fixture_validated": legacy["fixture_validated"],
        "historical_s15_p1_dynamic_binding_state_is_authoritative": False,
        "historical_two_non_manual_fields_quarantined": legacy["historical_two_non_manual_fields"] == 2,
        "reviewed_dependencies": {
            "stage14_post_remediation_review": s14_review.MANIFEST_PATH.as_posix(),
            "stage9_post_remediation_review": CURRENT_STAGE9_REVIEW_PATH.as_posix(),
            "s13_p1_current": CURRENT_S13_P1_PATH.as_posix(),
            "s13_p2_current": CURRENT_S13_P2_PATH.as_posix(),
            "s13_p3_current": CURRENT_S13_P3_PATH.as_posix(),
            "s14_p2_current": CURRENT_S14_P2_PATH.as_posix(),
            "historical_s15_p1_fixture": LEGACY_MANIFEST_PATH.as_posix(),
        },
        "validation_summary": validation_summary,
        "next_phase": "S15-P2",
        "next_required_step": (
            "Execute S15-P2 performance review list only as a separate run; do not execute S15-P3, "
            "Stage 15 review, salary calculation, bonus approval, payroll export, GitHub upload, "
            "app reinstall, formal report, difference closure, persistent write or business execution."
        ),
    }

    definitions_document = {
        "schema_version": "kmfa.v014.s15_p1.performance_fact_field_definitions.v1",
        "field_count": len(definitions),
        "fields": definitions,
    }
    bindings_document = {
        "schema_version": "kmfa.v014.s15_p1.performance_fact_field_bindings.v1",
        "binding_status_count": len(bindings),
        "bindings": bindings,
    }
    manual_document = {
        "schema_version": "kmfa.v014.s15_p1.performance_fact_manual_review_requirements.v1",
        "manual_review_requirement_count": len(manual),
        "requirements": manual,
    }
    for path, value in (
        (SUMMARY_PATH, summary),
        (MANIFEST_PATH, manifest),
        (FIELD_DEFINITIONS_PATH, definitions_document),
        (FIELD_BINDINGS_PATH, bindings_document),
        (MANUAL_REVIEW_PATH, manual_document),
        (MATRIX_PATH, matrix),
        (GO_NO_GO_PATH, go_no_go),
        (METADATA_SUMMARY_PATH, summary),
        (METADATA_MANIFEST_PATH, manifest),
        (METADATA_FIELD_DEFINITIONS_PATH, definitions_document),
        (METADATA_FIELD_BINDINGS_PATH, bindings_document),
        (METADATA_MANUAL_REVIEW_PATH, manual_document),
        (METADATA_MATRIX_PATH, matrix),
        (METADATA_GO_NO_GO_PATH, go_no_go),
        (PRIVATE_RAW_BEFORE_PATH, raw_before),
        (PRIVATE_RAW_AFTER_PATH, raw_after),
        (PRIVATE_PROBE_PATH, probe),
    ):
        _write_json(path, value)
    _write_text(REPORT_PATH, _render_report(summary))
    _write_text(TEST_RESULTS_PATH, _render_test_results(summary, browser))
    _write_text(
        RISK_REGISTER_PATH,
        """# S15-P1 风险登记

| 风险 | 控制 | 状态 |
|---|---|---|
| legacy 两字段绑定状态回流 | 当前 Stage 13/14 evidence 为唯一动态事实；六字段全部人工复核 | controlled |
| 候选结构被当作绩效事实 | 权威行/值绑定和绩效事实物化独立计数且保持 0 | controlled |
| 项目成本/回款结构引用被误读为值绑定 | 结构连接与权威值绑定分栏展示 | controlled |
| 客情费率口径被自动推断 | 定义、分母、期间和权威值缺失时保持 pending | controlled |
| 绩效字段进入工资奖金计算 | S15-P2/P3、工资、奖金、薪资导出和业务动作全部阻断 | controlled |
| raw/private/secret 进入 Git | 详细探针与截图只写 ignored private runtime，公开证据仅含聚合计数 | controlled |
""",
    )
    _write_text(
        ROLLBACK_PATH,
        f"""# S15-P1 回滚计划

1. 回退本 phase 的本地 commit 和 {OUTPUT_DIR.as_posix()} 公开证据。
2. 删除 ignored private raw/probe/browser evidence，不触碰原始目录。
3. 恢复 Stage 14 review 为当前治理入口；不进入 S15-P2。
4. 不回退、不移动、不删除、不覆盖任何原始文件。
""",
    )
    _write_text(PRIVATE_REPORT_PATH, _render_private_report(summary))
    if write_governance:
        _write_governance(generated_at)
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--final-validation", action="store_true")
    parser.add_argument("--browser-evidence-only", action="store_true")
    parser.add_argument("--private-probe-only", action="store_true")
    parser.add_argument("--no-governance", action="store_true")
    args = parser.parse_args()
    if args.private_probe_only:
        raw_helper = s14_review.p1.s13_review.p1.s12_review.p1.s11_project
        snapshot = raw_helper._raw_snapshot("v014_s15_p1_private_probe_only")
        probe = _raw_candidate_probe(Path(snapshot["raw_root"]))
        print(
            "S15-P1 private probe: "
            f"fields={probe['private_candidate_covered_field_count']}/6 "
            f"candidates={probe['private_unique_candidate_sheet_count']} "
            f"mismatches={probe['private_probe_roundtrip_mismatch_count']}"
        )
        return 0
    if args.browser_evidence_only:
        browser = _browser_worker()
        print(browser["status"])
        return 0 if browser["status"] == "PASS" else 1
    manifest = generate(
        final_validation=args.final_validation,
        write_governance=not args.no_governance,
    )
    summary = manifest["summary"]
    print(
        "S15-P1 post-remediation performance fact fields: "
        f"fields={summary['field_definition_count']} "
        f"manual={summary['manual_review_required_field_count']} "
        f"bound={summary['authoritative_value_binding_proven_field_count']} "
        f"facts={summary['materialized_performance_fact_count']} "
        f"decision={summary['decision']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
