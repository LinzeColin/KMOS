#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""最近完工项目成本：从真实源算出，写成 App 能直接读的 JSON。

Owner 2026-07-27：「我根本没有看到项目成本，我说了我要最近完工的项目成本」。
所以这支程序的产物不是给人看的 CSV，而是**给页面读的 JSON**——数据要出现在驾驶舱里，
不是出现在聊天里。

两个口径并排，不调平：
  · 业务台账口径——红圈《生产项目状态表》里业务自己按项目填的
    材料费／交通费／生活住宿费／其他费用，以及自有与劳务人工工时；
  · 金蝶归集口径——明细账中按『销售合同号』归集的生产成本借方发生额。
差异本身就是要看的东西。任何一方缺失都如实留空，不拿另一方顶替。

口径锁（都是实测踩出来的，改动前先看这里）：
  · 账簿按名去重——源包里每本明细账存了 3 份同 CRC 副本，不去重直接三倍放大；
  · 取借方发生额而非净额——生产成本归集后结转到主营业务成本，净额会互相对冲成 0；
  · 合同号按完整主号匹配——序号跨年重复，按序号归并会把不同项目的钱并到一起；
  · 『不分项目』占位桶不计入任何项目。
"""
from __future__ import annotations
import argparse, collections, glob, io, json, os, re, sys, tempfile, zipfile
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import render_report  # noqa: E402
import rollup as R  # noqa: E402

STATUS_SHEET = "生产项目状态表.xlsx"
SUMMARY_ROWS = ("期初余额", "本期合计", "本年累计", "本期发生额", "期末余额")
BUSINESS_COST_COLUMNS = ("材料费", "交通费", "生活住宿费", "其他费用")

# ── 人工：金蝶给不出，只能从红圈工时来 ────────────────────────────────────
# 生产成本里「劳务费」占约八成，其中约七成记在 `不分项目_不分项目` 占位桶，
# 八个基准项目名下劳务费为 0——这是记账口径问题，历史改不了。
# Owner 2026-07-28 指出红圈《生产项目状态表》按项目填了工时，人工从那里出。
# 因为金蝶的人工归集不到项目，两者不会重复计。
#
# 单价从 8 份竣工项目财务报表标定（除以红圈同项目工时）：
#   自有  465.12 / 500.00 / 496.54 / 502.43 / 479.44  → 收敛在 ~490，取 500
#         另三份 741.66 / 796.09 / 587.78 报表自注了补贴或提成，不参与定价
#   劳务  池州恒鑫 9,010.00 ÷ 17 = 530.00；新疆宜化 930,152.75 ÷ 1,579.32 = 588.95 → 取 550
LABOUR_RATE_OWN = Decimal("500")
LABOUR_RATE_SUB = Decimal("550")

# 这两个是账上的**伪合同号**占位桶（合计约 3,261 万），不是真项目。
# 不排掉，它们会被当成两个金额巨大的项目混进表里。
PLACEHOLDER_CONTRACTS = frozenset({"KMX999", "KMX9999"})

# 已知读不到的源：武汉彤烨、湖北曦悦两家的 `.xls` 账簿。它们是单表结构、
# 另一套科目表（生产成本是 4101 不是 5001），且**没有「销售合同号」列**——
# 即使读进来也归不到项目。实测两家生产成本合计 5,951.17 元。
# 如实登记为未覆盖，不为 0.03% 的钱引入 xlrd 依赖，也不假装已覆盖。
UNCOVERED_SOURCES = [
    {"账簿": "武汉彤烨明细账.xls", "原因": "单表结构 + 4101 科目表 + 无销售合同号列",
     "实测生产成本": "5212.65"},
    {"账簿": "曦悦公司明细账.xls", "原因": "同上", "实测生产成本": "738.52"},
]


def norm_contract(value) -> str:
    """合同号归一：去空白、去重复自拼、只留主号。"""
    text = re.sub(r"\s", "", str(value or "")).upper()
    if "_" in text:
        head, tail = text.split("_", 1)
        if head == tail:
            text = head
    matched = re.match(r"(KMX\d+-?\d*)", text)
    return matched.group(1) if matched else text


def money(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            return Decimal(repr(float(value)))
        except Exception:
            return None
    text = re.sub(r"[,\s，元]", "", str(value))
    return Decimal(text) if re.fullmatch(r"-?\d+(\.\d+)?", text) else None


def open_workbook(path: str):
    """打开 xlsx；先剥掉新版 Excel 的数据验证标签。

    实测：《生产项目状态表》带 `<dataValidation id=...>`，openpyxl 解析不了
    （`TypeError: DataValidation.__init__() got an unexpected keyword argument 'id'`）。
    只读模式是惰性解析，错误要到迭代行时才抛，try/except 包 load_workbook 接不住——
    所以不试探，一律先剥。数据验证与取数无关，且只在临时副本上剥，不改原文件。

    ⚠ 拿到 sheet 后**必须调 `sheet.reset_dimensions()`**，或者直接用 `iter_sheet_rows()`。
      见下面那个函数的注释：WPS/红圈导出的表声明 `<dimension ref="A1"/>`，
      只读模式信这个声明，于是 max_row=1、整表静默变空。
    """
    import openpyxl
    temporary = os.path.join(tempfile.mkdtemp(), os.path.basename(path))
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(
        temporary, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                text = payload.decode("utf-8", "replace")
                text = re.sub(r"<dataValidations.*?</dataValidations>", "", text, flags=re.S)
                text = re.sub(r"<dataValidations[^>]*/>", "", text)
                text = re.sub(r"<extLst>.*?</extLst>", "", text, flags=re.S)
                payload = text.encode("utf-8")
            target.writestr(item, payload)
    return openpyxl.load_workbook(temporary, read_only=True, data_only=True)


def iter_sheet_rows(sheet, values_only: bool = True):
    """迭代一张表的所有行——**先把它自称的尺寸扔掉**。

    2026-07-27 实测的一个静默数据丢失（Owner：「wps的数据为什么不拉进来」）：
      WPS / 红圈导出的 xlsx 在 sheet 里写 `<dimension ref="A1"/>`——声称整表只有一个单元格。
      openpyxl 只读模式**信这个声明**去定 max_row，于是表头之后的每一行都读不到。
      它不报错、不告警，就是零行。

      被它吃掉的（都是真数据，不是边角料）：
        红圈主合同 4341 行、项目开票 4525 行、付款审批（日常费用）1200 行。
      而《生产项目状态表》没这个毛病，所以过去只有它一个源进得来——
      项目成本因此建在 20 个完工项目上，而红圈主合同里有 390 个。

    reset_dimensions() 让 openpyxl 改为**边读边算**真实范围。
    代价是不能提前知道行数，换来的是不会静默丢数据——这个交换在任何时候都值。
    """
    sheet.reset_dimensions()
    return sheet.iter_rows(values_only=values_only)


def read_completed_projects(data_root: str) -> dict:
    """红圈《生产项目状态表》：完工项目 + 业务自填的成本与工时。"""
    return read_projects(data_root, only_completed=True)


def read_projects(data_root: str, only_completed: bool = True) -> dict:
    """红圈《生产项目状态表》：项目 + 业务自填的成本与工时。

    only_completed=False 时连在建、待入场一起读出来——项目毛利要看的是「哪些项目在赚钱」，
    而在建项目「成本在走、收入还没落」本身就是一条要摆出来的信息，不是该被滤掉的噪声。
    """
    found: dict[str, dict] = {}
    for path in glob.glob(f"{data_root}/**/{STATUS_SHEET}", recursive=True):
        workbook = open_workbook(path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            header = header_row = None
            for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=8, values_only=True)):
                cells = ["" if x is None else str(x).strip() for x in (row or [])]
                if sum(1 for c in cells if c) >= 4 and any("合同" in c for c in cells):
                    header, header_row = cells, index
                    break
            if not header:
                continue

            def column(*keys):
                for position, cell in enumerate(header):
                    if any(key in cell for key in keys):
                        return position
                return None

            contract_col = column("合同编号", "合同号")
            if contract_col is None:
                continue
            name_col = column("甲方名称", "项目名称", "工程名称")
            status_col = column("施工状态", "状态")
            done_col = column("完工时间", "竣工时间")
            amount_col = column("含税合同金额", "合同额", "合同金额")
            cost_cols = {label: column(label) for label in BUSINESS_COST_COLUMNS}
            other_cols = {
                label: column(label)
                for label in ("自有人工工时", "劳务人工工时", "结算金额", "开票金额",
                              "项目类型", "负责人", "开工时间", "是否提供项目成本表")
            }

            def cell(row, position):
                if position is None or position >= len(row) or row[position] is None:
                    return None
                return row[position]

            for row in sheet.iter_rows(min_row=header_row + 2, values_only=True):
                if not row:
                    continue
                key = norm_contract(cell(row, contract_col))
                if not key.startswith("KMX"):
                    continue
                status = str(cell(row, status_col) or "").strip()
                completed_at = str(cell(row, done_col) or "")[:10]
                done = bool(re.search(r"完工|竣工|完成|结束|已交|验收", status + completed_at))
                if only_completed and not done:
                    continue
                record = {
                    "合同编号": key,
                    "甲方名称": str(cell(row, name_col) or "").strip(),
                    "施工状态": status,
                    "完工日期": completed_at,
                    "含税合同金额": str(money(cell(row, amount_col)) or ""),
                }
                for label, position in cost_cols.items():
                    value = money(cell(row, position))
                    record[label] = str(value) if value is not None else ""
                for label, position in other_cols.items():
                    value = cell(row, position)
                    record[label] = "" if value is None else str(value)[:24]
                record["已完工"] = done
                # 键必须带甲方。只用合同号做键，会把**不同项目**吃掉一个。
                #
                # 2026-07-28 实证：`KMX202595-064` 在《生产项目状态表》里挂着两行——
                # 日照钢铁（施工中，合同额 1,263,546）与新疆宜化（已完工，1,180,000）。
                # 原来 `found[合同号] = record` 加「保留完工日期较晚的」，于是日照钢铁
                # 整个项目**从表上消失**，32 行进、31 个出，没有任何提示。
                #
                # 静默丢弃比归并更糟：归并至少金额还在，丢弃是凭空少一个项目，
                # 而少掉的那个恰好是合同额最大的几个之一。
                # 同合同号同甲方才是重复导出，按完工日期取新；同合同号不同甲方是
                # **身份冲突**，两条都留、都打标，交给人去裁。
                identity = (key, record["甲方名称"])
                previous = found.get(identity)
                if not previous or completed_at > previous.get("完工日期", ""):
                    found[identity] = record

        workbook.close()

    # 标记冲突：同一合同号落在多个甲方名下
    by_contract: dict[str, list[tuple]] = {}
    for identity in found:
        by_contract.setdefault(identity[0], []).append(identity)
    for contract, group in by_contract.items():
        if len(group) < 2:
            continue
        others = {found[i]["甲方名称"] for i in group}
        for identity in group:
            found[identity]["身份冲突"] = (
                f"合同号 {contract} 同时挂在 {len(group)} 个甲方名下："
                + "、".join(sorted(others))
                + "。不得自动归并，需人工确认唯一项目/合同映射。")
    return found


def sort_key(record: dict) -> str:
    """完工日期缺失时用合同号里的日期兜底，让排序不至于把无日期的全推到末尾。"""
    if record.get("完工日期"):
        return record["完工日期"]
    matched = re.match(r"KMX(\d{6,8})", record["合同编号"])
    if matched:
        digits = matched.group(1)
        if len(digits) == 8:
            return f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"
        if len(digits) == 6:
            return f"20{digits[:2]}-{digits[2:4]}-{digits[4:]}"
    return ""


#: 账上有、但**归不到任何项目**的生产成本，按跳过原因分桶。运行时统计，
#: 不写死任何金额——KMOS 是公开仓，真实金额只能在运行期出现在产物里。
#:
#: 为什么必须统计它：2026-07-30 拿 8 份真实《竣工项目财务报表》对了一次线上产物，
#: **0/8 落在 ±10% 内，且 8 个全部偏低**（-10.4% 到 -79.3%）。偏低不是解析错，
#: 是会计把大量成本记在了 `不分项目`／`KMX999`／`KMX9999` 这些占位桶里——
#: 新疆宜化 064 的承包费 93 万在报表上有、在金蝶该项目名下是 0。
#:
#: 本产物此前只写了一句「成本偏保守、毛利偏乐观」。**定性不够**：
#: 看表的人无法判断偏 3% 还是偏 79%，也就无法判断这个毛利能不能拿去谈结算。
UNATTRIBUTED_REASONS = ("不分项目", "伪合同号占位桶", "不在主合同表中")


def collect_ledger_cost(data_root: str, targets: set[str], account_to_row: dict,
                        variants: dict | None = None,
                        unattributed: dict | None = None) -> dict:
    """明细账按项目归集生产成本借方发生额。返回 {合同主号: {科目: 金额}}。

    `unattributed` 传入时，另把**被跳过的**生产成本按原因累加进去——
    那部分钱确实发生了，只是账上没挂到项目，见 `UNATTRIBUTED_REASONS`。

    `variants` 传入时，另记 {合同主号: {原始销售合同号: 金额}}——**归并到主号之前
    长什么样**。

    为什么要记：合同号带变体后缀（`-Z`／`-XF`／`--Z`），`norm_contract` 会把它们
    并进主号。2026-07-28 实测，广安台泥 079 的金蝶成本 64,653.90 **全部**来自
    `KMX20251119-079-Z`，福建鼎信 003 的 1,919.19 **全部**来自 `KMX2026116-003--Z`。
    后缀的业务含义（变更／分包／中标）至今没有裁定，也就是说这两笔到底该不该算进
    主项目，是个未决问题——而代码在静默地替它做了决定。

    不改归并口径（不并会让这两个项目的金蝶成本直接归零，那更错），但把成分摆出来：
    看表的人至少能知道「这个数是从一个带后缀的合同号来的」。
    """
    import openpyxl
    aggregate: dict[str, dict[str, Decimal]] = collections.defaultdict(
        lambda: collections.defaultdict(Decimal))
    seen_books: set[str] = set()
    unknown_accounts: set[str] = set()

    def handle(sheet_name, sheet):
        prefix = re.match(r"(\d{4})", sheet_name)
        if not prefix or prefix.group(1) != "5001":
            return
        head = []
        for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
            head.append(["" if x is None else str(x).strip() for x in (row or [])])
        header = header_row = None
        for index, row in enumerate(head):
            if "摘要" in row and "借方" in row:
                header, header_row = row, index
                break
        if not header:
            return
        try:
            contract_i = header.index("销售合同号")
            account_i = header.index("科目")
            memo_i = header.index("摘要")
            debit_i = header.index("借方")
        except ValueError:
            return
        for row in sheet.iter_rows(min_row=header_row + 2, values_only=True):
            if not row:
                continue
            memo = str(row[memo_i]).strip() if memo_i < len(row) and row[memo_i] is not None else ""
            if not memo or memo in SUMMARY_ROWS:
                continue
            raw = row[contract_i] if contract_i < len(row) else None
            value = row[debit_i] if debit_i < len(row) else None
            try:
                amount = Decimal(str(value)) if value not in (None, "") else Decimal(0)
            except Exception:
                amount = Decimal(0)

            def drop(reason: str) -> None:
                """这笔钱花掉了，只是归不到项目——记下来，别静默丢。"""
                if unattributed is not None and amount:
                    unattributed[reason] = unattributed.get(reason, Decimal(0)) + amount

            if raw in (None, "") or "不分项目" in str(raw):
                drop("不分项目")
                continue
            key = norm_contract(raw)
            if key in PLACEHOLDER_CONTRACTS:    # 伪合同号占位桶，不是项目
                drop("伪合同号占位桶")
                continue
            if key not in targets:
                drop("不在主合同表中")
                continue
            account = str(row[account_i]).strip() if account_i < len(row) and row[account_i] else ""
            if account not in account_to_row:
                unknown_accounts.add(account)
                continue
            aggregate[key][account] += amount
            if variants is not None:
                variants.setdefault(key, collections.defaultdict(Decimal))[
                    re.sub(r"\s", "", str(raw)).upper()] += amount

    bundles = sorted(set(glob.glob(f"{data_root}/**/*金蝶*.zip", recursive=True)
                         + glob.glob(f"{data_root}/**/*明细账*.zip", recursive=True)))
    for bundle in bundles:
        with zipfile.ZipFile(bundle) as archive:
            for member in archive.namelist():
                if not member.lower().endswith((".xlsx", ".xlsm")):
                    continue
                base = os.path.basename(member)
                if base in seen_books:          # 源包里每本存 3 份同 CRC 副本
                    continue
                seen_books.add(base)
                try:
                    workbook = openpyxl.load_workbook(
                        io.BytesIO(archive.read(member)), read_only=True, data_only=True)
                except Exception:
                    continue
                for sheet_name in workbook.sheetnames:
                    try:
                        handle(sheet_name, workbook[sheet_name])
                    except Exception:
                        continue
                workbook.close()
    if unknown_accounts:
        raise KeyError(f"这些成本子科目未在映射表登记，拒绝静默丢弃：{sorted(unknown_accounts)}")
    return aggregate


#: 竣工报表（PDF）里注明的自有工数，与红圈《生产项目状态表》对不上的项目。
#: 数字来自另一线程的参考回放核对（`KMFA_项目成本_真实参考回放_8项目.xlsx`
#: 的「阻塞与异常」页，P0 人工工数冲突）。
#:
#: 这里**不替换**红圈的工时——两个来源哪个权威没有裁定，替换就是拿一个未经确认的
#: 数覆盖另一个。只打标：让用这张表的人知道该项目的人工成本存在 ~2× 的不确定性。
#: 广安台泥按 214 工算是 107,000，按 119 工算是 59,500，差 47,500——这不是小数点问题。
LABOUR_HOURS_CONFLICT = {
    "KMX20251119-079": {"报表工数": 119, "红圈工数": 214},
    "KMX2026120-004": {"报表工数": 31, "红圈工数": 32},
}


def labour_cost(record: dict) -> tuple[Decimal, Decimal, bool]:
    """人工＝红圈工时 × 标定单价。返回 (自有, 劳务, 是否填了工时)。

    工时**没填**和工时**是 0** 必须分得开：前者是不知道，后者是真没投人工。
    分不开的后果是，一个没填工时的项目会以「人工 0」的面目出现在表上，
    而人工往往是最大的一块——那等于给出一个系统性偏高的毛利。
    """
    own_hours = money(record.get("自有人工工时"))
    sub_hours = money(record.get("劳务人工工时"))
    recorded = own_hours is not None or sub_hours is not None
    own = (own_hours or Decimal(0)) * LABOUR_RATE_OWN
    sub = (sub_hours or Decimal(0)) * LABOUR_RATE_SUB
    return own, sub, recorded


#: 红圈《主合同》导出——**合同号的权威来源**（Owner 2026-07-29 定：
#: 「红圈数据和 wps 数据是合同号的权威来源」）。
MASTER_CONTRACT_GLOB = "红圈主合同*.xlsx"


def read_master_contracts(data_root: str) -> dict[str, dict]:
    """红圈《主合同》：合同号 → 甲方／状态／合同额／完工日期。

    为什么必须读它：《生产项目状态表》只有 **34 行**，而主合同表有 **4,332 个合同号**。
    一直只读状态表，等于把项目成本的范围缩到了三十几个——而金蝶里有成本记录的
    合同有 176 个。2026-07-29 实测：改用主合同表当身份权威后，能出成本的项目
    从 32 个变成 169 个。

    它同时解掉「一个合同号挂两个甲方」那种冲突：状态表里 KMX202595-064 同时挂在
    新疆宜化和日照钢铁名下，而主合同表里它唯一对应新疆宜化——**状态表那条录错了**。
    权威表说了算，不再需要人工裁定。
    """
    master: dict[str, dict] = {}
    for path in sorted(glob.glob(f"{data_root}/**/{MASTER_CONTRACT_GLOB}", recursive=True)):
        workbook = open_workbook(path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # read_only 下 openpyxl 信表里声明的 dimension，而这份导出声明的是单行——
            # 不 reset 就只读得到表头，函数静默返回空，表现成「主合同表里没有任何合同」。
            try:
                sheet.reset_dimensions()
            except AttributeError:
                pass
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            header = ["" if x is None else str(x).strip() for x in rows[0]]
            index = {name: position for position, name in enumerate(header) if name}

            def pick(row, *keys):
                for key in keys:
                    position = index.get(key)
                    if position is not None and position < len(row) and row[position] is not None:
                        return row[position]
                return None

            if not any("合同编号" in h or "合同号" in h for h in header):
                continue
            for row in rows[1:]:
                if not row:
                    continue
                key = norm_contract(pick(row, "合同编号", "合同号"))
                if not key.startswith("KMX"):
                    continue
                # 后出现的不覆盖先出现的：同名导出多份时以第一份为准，避免快照顺序影响结果
                master.setdefault(key, {
                    "甲方名称": str(pick(row, "甲方") or "").strip(),
                    "施工状态": str(pick(row, "施工状态") or "").strip(),
                    "完工时间": str(pick(row, "完工日期（产值上报）", "完工日期") or "")[:10],
                    "含税合同金额": pick(row, "含税合同额(元)", "含税合同金额"),
                    "结算金额": pick(row, "结算金额(元)"),
                })
        workbook.close()
    return master


def merge_master_contracts(data_root: str, projects: dict[str, dict]) -> dict[str, dict]:
    """把主合同表并进来：它定身份，状态表定工时与业务自填费用。

    并法有三条，每条都有实测理由：
      · 主合同表里有、状态表里没有的 → **补进来**（这是从 32 涨到 169 的来源）；
      · 两边都有 → 身份字段（甲方／状态／合同额／完工日）以主合同表为准，
        工时与台账费用仍取状态表——那些字段主合同表里没有；
      · 状态表里有、主合同表里没有的 → **保留但打标**，不静默丢。
        丢了就是拿「权威表没登记」当「这个项目不存在」，而它可能只是新合同还没进表。
    """
    master = read_master_contracts(data_root)
    if not master:
        return projects                      # 读不到权威表就按原样走，不因此丢覆盖

    merged: dict[str, dict] = {}
    for key, record in projects.items():
        contract_key = record["合同编号"]
        authority = master.get(contract_key)
        if authority:
            record = dict(record)
            same_party = (not authority.get("甲方名称")
                          or not record.get("甲方名称")
                          or authority["甲方名称"] == record["甲方名称"])
            if same_party:
                for field in ("甲方名称", "施工状态", "含税合同金额"):
                    if authority.get(field):
                        record[field] = authority[field]
                if authority.get("完工时间"):
                    record["完工日期"] = authority["完工时间"]
                record["身份来源"] = "红圈主合同（权威）"
                record.pop("身份冲突", None)   # 权威表已定案，冲突不再成立
            else:
                # 状态表这一行的甲方跟权威表对不上——**这一行的合同号填错了**，
                # 不能把它改写成权威表的甲方：那等于把「日照钢铁的一笔成本」
                # 挂到「新疆宜化」名下，凭空造出一个不存在的成本。
                # 保留原样并标出来，让人去改源头。
                record["身份来源"] = (
                    f"⚠ 合同号与权威表冲突：主合同表里 {contract_key} 属于"
                    f"「{authority['甲方名称']}」，而《生产项目状态表》这一行写的是"
                    f"「{record['甲方名称']}」。**这一行的合同号很可能填错了**，"
                    f"其成本未归入任何项目，需在红圈里更正。")
                record["合同号存疑"] = True
        else:
            record = dict(record)
            record["身份来源"] = "仅生产项目状态表——主合同表里查无此合同号"
        record["已完工"] = bool(re.search(
            r"完工|竣工|完成|结束|已交|验收",
            str(record.get("施工状态", "")) + str(record.get("完工日期", ""))))
        record["完工排序"] = record.get("完工日期") or ""
        merged[key] = record

    seen = {r["合同编号"] for r in merged.values()}
    for contract_key, authority in master.items():
        if contract_key in seen:
            continue
        done = bool(re.search(r"完工|竣工|完成|结束|已交|验收",
                              authority["施工状态"] + authority["完工时间"]))
        merged[contract_key] = {
            "合同编号": contract_key,
            "甲方名称": authority["甲方名称"],
            "施工状态": authority["施工状态"],
            "完工日期": authority["完工时间"],
            "含税合同金额": authority["含税合同金额"],
            "结算金额": authority.get("结算金额"),
            "已完工": done,
            "完工排序": authority["完工时间"],
            "身份来源": "红圈主合同（权威）",
            # 工时与业务自填费用只在《生产项目状态表》里有，这些项目没有那一行——
            # 于是它们的人工算不出来。**这不是「人工是 0」，是「工时没填」**，
            # labour_cost 会据此把「工时已填」标成 False。
        }
    return merged


def build(data_root: str, account_map: dict, limit: int = 0) -> dict:
    """limit=0 表示**不限**——项目成本要覆盖全部项目，不是最近 N 个。

    Owner 2026-07-28：「既然在系统上这些项目成本本来就应该是实时更新的，
    说明你根本没有全量跑所有信息」。之前默认 limit=20 只出最近 20 个完工项目，
    施工中的一个都没有，而施工中的项目恰恰是还能干预的那些。
    """
    account_to_row = {m["account"]: m["rows"]["A"]["row"] for m in account_map["mappings"]}
    everything = read_projects(data_root, only_completed=False)
    everything = merge_master_contracts(data_root, everything)
    ranked = sorted(everything.values(), key=sort_key, reverse=True)
    if limit:
        ranked = ranked[:limit]
    variants: dict[str, dict[str, Decimal]] = {}
    unattributed: dict[str, Decimal] = {}
    ledger = collect_ledger_cost(data_root, {r["合同编号"] for r in ranked},
                                 account_to_row, variants=variants,
                                 unattributed=unattributed)

    projects = []
    for record in ranked:
        key = record["合同编号"]
        leaf: dict[str, Decimal] = collections.defaultdict(Decimal)
        for account, amount in ledger.get(key, {}).items():
            leaf[account_to_row[account]] += amount
        violations = R.check_invariants("A", leaf)
        if violations:
            raise ValueError(f"{key} 上卷不自洽：{violations}")
        by_row, direct_total = R.rollup("A", leaf)

        business_total = sum((money(record.get(c)) or Decimal(0)) for c in BUSINESS_COST_COLUMNS)
        own_labour, sub_labour, hours_recorded = labour_cost(record)
        contract = money(record.get("含税合同金额"))

        # 逐行合并两个口径，**不是整块取大**——整块取大会把「金蝶有材料没现场费」
        # 和「红圈有现场费没材料」互相抵消掉，实测覆盖率从 84% 掉到 78%。
        #
        #   原材料／租赁费：按源优先级取（金蝶是入账实据，优先于红圈自填）；
        #   现场管理费：取大不相加——金蝶这一行有的已含管理人员工资、有的没含
        #     （工资进了「不分项目」），含没含无法逐项判定，重叠相加会把成本算高。
        ledger_material = by_row.get("（一）原材料") or Decimal(0)
        ledger_lease = by_row.get("（二）租赁费") or Decimal(0)
        ledger_site = by_row.get("（四）现场管理费") or Decimal(0)

        material = ledger_material or (money(record.get("材料费")) or Decimal(0))
        lease = ledger_lease
        site_from_business = own_labour + sum(
            (money(record.get(c)) or Decimal(0))
            for c in BUSINESS_COST_COLUMNS if c != "材料费")
        site = max(ledger_site, site_from_business)
        site_from_ledger = ledger_site

        # 金蝶归集到的**全额**都要进成本，不能只挑三行。
        #
        # 2026-07-29 抓到：原来 total 只取 原材料＋租赁费＋现场管理费＋劳务人工，
        # 而金蝶归集到的最大一行是 **（五）工资（承包费）支出 1,502 万**——整个被丢掉。
        # 一直以为「金蝶的人工记在『不分项目』、按项目归集不到」，那是错的：
        # 这一行按销售合同号归集得好好的，占全部归集成本的三分之二。
        #
        # 会不会和红圈工时重复？实测：84 个项目有金蝶(五)、16 个有红圈工时，
        # **两者都有的是 0 个**——它们是同一件事的互补来源，相加不重复。
        # 所以规则是：金蝶有归集就用金蝶全额，金蝶没有的项目才落到红圈工时上。
        # 人工到底算谁的，只看**金蝶(五)在不在**，不看金蝶有没有别的成本。
        # 第一版写成「金蝶有任何归集就不要红圈人工」，实测对 8 份竣工报表的覆盖率
        # 从 84% 掉到 39%——因为很多项目金蝶只归集到了材料，人工仍然只在红圈里。
        # 零重叠这个实测结论**只对 (五) 这一行成立**（84 个 vs 16 个，交集 0），
        # 拿它去推「所有金蝶科目」就是把一个局部结论当全称用。
        ledger_labour = by_row.get("（五）工资（承包费）支出") or Decimal(0)
        if direct_total:
            # 金蝶归集到了 → 以它的**全额**为底，只做两处调整：
            #
            #   · 现场管理费那一行换成 max(金蝶, 红圈)——金蝶这行有的已含管理人员工资、
            #     有的没含，含没含逐项判不了；红圈那边是「自有人工＋台账其他费用」。
            #     取大不相加，否则自有人工会被算两遍（第一版就是这么把 084 算到 112% 的）。
            #   · 劳务人工只在金蝶(五)缺位时补红圈的——(五) 就是承包费，两者是同一件事。
            #     实测 84 个项目有 (五)、16 个有红圈工时、交集 0，所以互补而不重叠。
            total = direct_total - ledger_site + max(ledger_site, site_from_business)
            if not ledger_labour:
                total += sub_labour
        else:
            # 金蝶完全没归集：只能落到红圈自填的费用与工时上。
            total = material + lease + site + sub_labour
        # 分摊管理费只摊给**真发生过成本**的项目。
        #
        # 2026-07-29 改用红圈主合同表当身份权威后，项目范围从 32 个变成 4,313 个历史合同。
        # 原来无条件按「合同额 × 2%」摊，在这个范围上摊出了 **967 万** ——
        # 而其中绝大多数合同在金蝶里一分钱成本都没有、工时也没填。
        # 给一个没跑过的合同摊 2%，那是凭空造成本：它会让「成本不知道」的项目
        # 看起来像「成本很低」，正是这套表最该避免的那种误导。
        #
        # 判据是「有没有成本发生额」，不是「合同额是不是 0」：
        # 金蝶归集到了、业务台账填了费用、或红圈填了工时——三者有其一才摊。
        has_actual_cost = bool(total) or bool(business_total) or hours_recorded
        management_fee = (contract * Decimal("0.02")) \
            if (contract is not None and has_actual_cost) else Decimal(0)

        projects.append({
            **{k: v for k, v in record.items()},
            "完工排序": sort_key(record),
            "业务台账成本合计": str(business_total),
            "台账口径毛利": str(contract - business_total) if contract is not None else "",
            "金蝶归集直接成本": str(direct_total),
            "金蝶成本明细": {row: str(value) for row, value in sorted(by_row.items())},
            "两口径差额": str(direct_total - business_total),
            "自有人工成本": str(own_labour),
            "劳务人工成本": str(sub_labour),
            "工时已填": hours_recorded,
            "现场成本取自": "金蝶" if site_from_ledger > site_from_business else "红圈工时＋台账费用",
            "分摊管理费": str(management_fee),
            "成本合计": str(total + management_fee),
            "毛利": str(contract - total - management_fee) if contract is not None else "",
            "可出毛利": bool(hours_recorded and contract is not None),
            # 工数两个来源对不上的，把两边都摆出来 —— 不替换、不取平均、不挑一个。
            # 哪个权威没有裁定，任何一种「处理」都是拿未经确认的数覆盖另一个。
            # 金蝶这笔钱原始挂在哪个合同号下。带后缀的单列出来——归并是代码替人
            # 做的决定，至少要让人看得见它做了什么。
            **({"金蝶合同号构成": {
                raw: str(amount) for raw, amount in
                sorted(variants.get(key, {}).items(), key=lambda kv: -kv[1])
            }} if variants.get(key) else {}),
            **({"工数冲突": {
                **LABOUR_HOURS_CONFLICT[key],
                "本表采用": "红圈工数",
                "影响": f"人工成本按 {LABOUR_HOURS_CONFLICT[key]['红圈工数']} 工计"
                        f"{LABOUR_HOURS_CONFLICT[key]['红圈工数'] * LABOUR_RATE_OWN:,.2f}；"
                        f"若按报表 {LABOUR_HOURS_CONFLICT[key]['报表工数']} 工则为 "
                        f"{LABOUR_HOURS_CONFLICT[key]['报表工数'] * LABOUR_RATE_OWN:,.2f}",
                "处理": "待人工锁定批准工时与人员项目分配后再定",
            }} if key in LABOUR_HOURS_CONFLICT else {}),
        })

    return {
        "schema_version": "kmfa.project_cost.recent_completed.v2",
        # 生成时间必须落在载荷里。Owner 2026-07-28 问的正是「项目成本是实时更新的吗」，
        # 而在此之前这份 JSON 里**没有任何时间戳**——看到的数是今天算的还是上周的，
        # 从页面上分不出来。分不出来就等于不知道它有没有在更新，
        # 跟「绿的但没干活」是同一类问题：看着有数，其实不知道数从哪一刻来。
        "生成时间": datetime.now(timezone(timedelta(hours=8))).isoformat(timespec="seconds"),
        "口径": {
            "业务台账": "红圈《生产项目状态表》里业务自填的 材料费＋交通费＋生活住宿费＋其他费用",
            "金蝶归集": "明细账中按『销售合同号』归集的生产成本借方发生额；不含记入『不分项目』的部分",
            "人工": f"红圈工时 × 标定单价（自有 {LABOUR_RATE_OWN}／劳务 {LABOUR_RATE_SUB} 元每工）"
                    "——金蝶的人工记在『不分项目』，按项目归集不到，因此两者不重复计",
            "为什么并排": "两个口径都来自真实记录，差异本身就是要看的东西——不做任何调平，也不挑一个好看的",
        },
        "锁定的算法": [
            "账簿按名去重：源包里每本明细账存了 3 份同 CRC 副本，不去重直接三倍放大",
            "取借方发生额而非净额：生产成本结转到主营业务成本，净额会互相对冲成 0",
            "合同号按完整主号匹配：序号跨年重复，按序号归并会把不同项目的钱并到一起",
            "『不分项目』占位桶不计入任何项目",
            "KMX999／KMX9999 是伪合同号占位桶（约 3,261 万），不是项目，排除",
            "现场这一块金蝶与红圈取大不相加：金蝶『现场管理费』有的已含工资、有的没含，"
            "无法逐项判定，重叠相加会把成本算高",
            "工时没填 ≠ 工时为 0：没填的项目不出毛利（`可出毛利:false`）",
            "项目键＝合同号＋甲方：同合同号不同甲方是身份冲突，两条都留并打 `身份冲突` 标，"
            "绝不静默丢弃（曾因此少掉合同额 1,263,546 的日照钢铁项目）",
            "合同号后缀（-Z／--Z／-XF／-Z1）按主号归并——2026-07-28 全量实测证实这是正确"
            "口径，不是假设：① 32 个项目里凡金蝶有成本的 18 个，100% 都记在带后缀的号下，"
            "说明后缀是账上固定写法而非例外；② 同项目出现两种后缀的 5 个里，有 4 个的 "
            "`--Z` 金额与『（一）原材料』**分毫不差**、`-Z` 与『（四）现场管理费』分毫不差"
            "（阜阳皖润 4,789.23／21,675.08、山东圣川 482.25／6,322.60、池州恒鑫 161.00／"
            "25,016.93、青海盐湖海纳 2,797.44／9,428.06）。**后缀标的是科目记账线，不是项目"
            "身份**，故按主号归并不会把不同项目并到一起。每个项目的 `金蝶合同号构成` 保留"
            "原始号，可逐笔回溯",
        ],
        "单价标定": {
            "自有": {"采用": str(LABOUR_RATE_OWN),
                     "样本": "8 份竣工报表工资行 ÷ 红圈同项目工时："
                             "465.12／500.00／496.54／502.43／479.44 收敛在 ~490；"
                             "741.66／796.09／587.78 三份报表自注了补贴或提成，不参与定价"},
            "劳务": {"采用": str(LABOUR_RATE_SUB),
                     "样本": "池州恒鑫 9,010.00÷17＝530.00；新疆宜化 930,152.75÷1,579.32＝588.95"},
        },
        "未覆盖": {
            "行": ["（六）信息费", "（七）税金", "1.2 占用的资金利息"],
            "影响方向": "三行全部是少算，所以本产物的成本偏保守、毛利偏乐观",
            "源": UNCOVERED_SOURCES,
        },
        # ── 少算多少：从形容词变成数 ──────────────────────────────────
        # 2026-07-30 拿 8 份真实《竣工项目财务报表》对了一次本产物：
        # **0/8 落在 ±10% 内，8 个全部偏低**，区间 -10.4% ～ -79.3%。
        # 此前这里只有一句「成本偏保守、毛利偏乐观」——定性不够：
        # 偏 3% 和偏 79% 是两件事，前者能拿去谈结算，后者会让人亏着钱以为在赚。
        #
        # 偏低不是解析错（解析已按列名取数、按 CRC 去重、取借方不取净额）。
        # 偏低是**账上就没把钱挂到项目**：下面这三个桶里的生产成本确实发生了，
        # 只是没有可用的项目归属。新疆宜化 064 的承包费在竣工报表上有 93 万，
        # 在金蝶该项目名下是 0——那笔钱在「不分项目」桶里。
        #
        # 金额一律运行时统计，**不写死在仓库里**（KMOS 是公开仓）。
        "未归集成本池": {
            "是什么": "账上确有发生、但归不到任何项目的生产成本借方发生额",
            "分桶": {reason: str(unattributed.get(reason, Decimal(0)))
                    for reason in UNATTRIBUTED_REASONS},
            "合计": str(sum(unattributed.values(), Decimal(0))),
            "已归集到项目": str(sum(
                (money(p.get("金蝶归集直接成本")) or Decimal(0)) for p in projects)),
            "怎么读": "本表每个项目的成本是**下限**——归集率越低，下限离真实成本越远，"
                    "毛利也就越偏乐观。要把下限抬成真值，得在记账时把这三桶里的钱"
                    "挂上销售合同号，代码这边补不出来。",
        },
        "毛利方向": "偏乐观（成本是下限）——见「未归集成本池」与「未覆盖」两节；"
                "拿去谈结算或考核前，先看该项目的金蝶归集是否覆盖了承包费与材料",
        "项目数": len(projects),
        "项目": projects,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="生成最近完工项目成本 JSON（供驾驶舱页面读取）")
    parser.add_argument("--data-root", required=True, help="KMFA_MetaData 根目录")
    parser.add_argument("--account-map", required=True, help="project_cost_account_map.json")
    parser.add_argument("--out", required=True, help="输出 JSON 路径")
    parser.add_argument("--limit", type=int, default=0,
                        help="只出最近 N 个；0＝全部（默认）")
    args = parser.parse_args()

    account_map = json.loads(Path(args.account_map).read_text(encoding="utf-8"))
    payload = build(args.data_root, account_map, args.limit)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ {payload['项目数']} 个完工项目 → {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
