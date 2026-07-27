#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""客户口径毛利：算出来写成 App 能直接读的 JSON。

为什么有客户口径这一页（2026-07-27 定案）：
  项目维度在账上大面积缺失——无归属成本里约八成连客户都没有，其余大头是关联公司往来，
  Owner 又明确底稿给不了、记账也改不了。但**同一批数据的客户维度是完整的**：
  收入与生产成本都带客户。所以「哪些客户在赚钱」这个问题现有数据能给出可信答案。
  能算准的才出，算不准的不装作算得准。

三个必须分开、绝不能合并成一个总数的东西（合并会严重失真）：
  · 关联方——集团自有公司（账套本身就是这些公司），它们之间的往来不是经营成果；
  · 疑似关联方——与账套间大额往来且成本近零，形态同为关联但未经确认，只标不断言；
  · 占位桶——『不分客户』，单列不摊。
看对外经营真实性，只读『外部客户』那一档。

另有四类数据提示必须逐条标注，而不是替读者算一个好看的数：
成本为负（红冲多于发生）、零成本有收入（成本走了别的主体）、毛利率超出 ±100%、
只有成本没有收入（在建或已完工未开票）。
"""
from __future__ import annotations
import argparse, collections, glob, io, json, os, re, sys, zipfile
from decimal import Decimal
from pathlib import Path

SUMMARY_ROWS = ("期初余额", "本期合计", "本年累计", "本期发生额", "期末余额")
GROUP_MARKERS = ("开明", "岚丹")          # 账套本身就是这些公司
SUSPECT_MARKERS = ("彤烨",)               # 形态同为关联，未经确认，只标不断言
REVENUE_PREFIX = "6001"
COST_PREFIX = "5001"


def classify(customer: str) -> str:
    if "不分客户" in customer or customer == "<空客户>":
        return "占位桶"
    if any(marker in customer for marker in GROUP_MARKERS):
        return "关联方"
    if any(marker in customer for marker in SUSPECT_MARKERS):
        return "疑似关联方"
    return "外部客户"


def data_flags(revenue: Decimal, cost: Decimal, gross: Decimal) -> list[str]:
    flags = []
    if cost < 0:
        flags.append("成本为负（红冲多于发生），毛利率不是真实经营结果")
    elif cost == 0 and revenue > 0:
        flags.append("零成本（成本未入账，或走了别的主体）")
    if revenue > 0 and abs(gross / revenue) > 1:
        flags.append("毛利率超出 ±100%，勿直接引用")
    if revenue == 0 and cost > 0:
        flags.append("只有成本没有收入（在建，或已完工未开票）")
    return flags


def collect(data_root: str) -> tuple[dict, dict, dict, dict, int]:
    import openpyxl
    revenue: dict[str, Decimal] = collections.defaultdict(Decimal)
    cost: dict[str, Decimal] = collections.defaultdict(Decimal)
    projects: dict[str, set] = collections.defaultdict(set)
    months: dict[str, set] = collections.defaultdict(set)
    seen_books: set[str] = set()

    def handle(sheet_name, sheet):
        prefix = re.match(r"(\d{4})", sheet_name)
        if not prefix or prefix.group(1) not in (COST_PREFIX, REVENUE_PREFIX):
            return
        kind = prefix.group(1)
        head = []
        for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
            head.append(["" if x is None else str(x).strip() for x in (row or [])])
        header = header_row = None
        for index, row in enumerate(head):
            if "摘要" in row and "借方" in row:
                header, header_row = row, index
                break
        if not header:
            return
        try:
            memo_i = header.index("摘要")
            debit_i = header.index("借方")
            credit_i = header.index("贷方")
            customer_i = header.index("客户")
            contract_i = header.index("销售合同号")
            date_i = header.index("日期")
        except ValueError:
            return
        for row in sheet.iter_rows(min_row=header_row + 2, values_only=True):
            if not row:
                continue
            memo = str(row[memo_i]).strip() if memo_i < len(row) and row[memo_i] is not None else ""
            if not memo or memo in SUMMARY_ROWS:
                continue
            customer = str(row[customer_i]).strip() if customer_i < len(row) and row[customer_i] else ""
            customer = customer or "<空客户>"
            contract = str(row[contract_i]).strip() if contract_i < len(row) and row[contract_i] else ""
            if contract and "不分项目" not in contract:
                projects[customer].add(contract[:40])
            when = row[date_i] if date_i < len(row) else None
            if when is not None:
                months[customer].add(str(when)[:7])

            def amount(position):
                value = row[position] if position < len(row) else None
                if value in (None, ""):
                    return Decimal(0)
                try:
                    return Decimal(str(value))
                except Exception:
                    return Decimal(0)

            if kind == REVENUE_PREFIX:
                revenue[customer] += amount(credit_i)
            else:
                cost[customer] += amount(debit_i)

    bundles = sorted(set(glob.glob(f"{data_root}/**/*金蝶*.zip", recursive=True)
                         + glob.glob(f"{data_root}/**/*明细账*.zip", recursive=True)))
    for bundle in bundles:
        with zipfile.ZipFile(bundle) as archive:
            for member in archive.namelist():
                if not member.lower().endswith((".xlsx", ".xlsm")):
                    continue
                base = os.path.basename(member)
                if base in seen_books:      # 源包里每本存 3 份同 CRC 副本
                    continue
                seen_books.add(base)
                try:
                    workbook = openpyxl.load_workbook(
                        io.BytesIO(archive.read(member)), read_only=True, data_only=True)
                except Exception:
                    continue
                for sheet_name in workbook.sheetnames:
                    try:
                        handle(sheet_name, workbook[sheet_name])
                    except Exception:
                        continue
                workbook.close()
    return revenue, cost, projects, months, len(seen_books)


def build(data_root: str) -> dict:
    revenue, cost, projects, months, book_count = collect(data_root)
    rows = []
    for customer in sorted(set(revenue) | set(cost)):
        r, c = revenue[customer], cost[customer]
        gross = r - c
        rows.append({
            "客户": customer,
            "类别": classify(customer),
            "收入": str(r),
            "已入账成本": str(c),
            "毛利": str(gross),
            "毛利率": (f"{float(gross / r) * 100:.1f}%" if r else ""),
            "涉及项目数": len(projects[customer]),
            "活跃月份数": len(months[customer]),
            "数据提示": data_flags(r, c, gross),
        })
    rows.sort(key=lambda x: -float(x["收入"]))

    by_kind: dict[str, dict] = {}
    for row in rows:
        bucket = by_kind.setdefault(row["类别"], {"家数": 0, "收入": Decimal(0), "成本": Decimal(0)})
        bucket["家数"] += 1
        bucket["收入"] += Decimal(row["收入"])
        bucket["成本"] += Decimal(row["已入账成本"])
    summary = {}
    for kind, bucket in by_kind.items():
        gross = bucket["收入"] - bucket["成本"]
        summary[kind] = {
            "家数": bucket["家数"], "收入": str(bucket["收入"]),
            "成本": str(bucket["成本"]), "毛利": str(gross),
            "毛利率": (f"{float(gross / bucket['收入']) * 100:.1f}%" if bucket["收入"] else ""),
        }

    return {
        "schema_version": "kmfa.customer_margin.v1",
        "口径": {
            "收入": "主营业务收入贷方发生额",
            "成本": "生产成本借方发生额（不取净额，净额会被归集/结转对冲）",
            "边界": "已入账口径；不含期间费用分摊，也不含未入账的在建成本",
        },
        "为什么是客户口径": "项目维度在账上大面积缺失（无归属成本里约八成连客户都没有，"
                            "其余大头是关联公司往来），而客户维度填充完整——能算准的才出",
        "分档说明": "关联方=集团自有公司；疑似关联方=与账套间大额往来且成本近零，未经确认只标不断言；"
                    "占位桶=不分客户，单列不摊。看对外经营真实性请只读『外部客户』那一档",
        "账簿数": book_count,
        "分档汇总": summary,
        "客户数": len(rows),
        "客户": rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="生成客户口径毛利 JSON（供驾驶舱页面读取）")
    parser.add_argument("--data-root", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    payload = build(args.data_root)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    external = payload["分档汇总"].get("外部客户", {})
    print(f"✓ {payload['客户数']} 个客户 → {out}；外部客户 {external.get('家数')} 家 "
          f"毛利率 {external.get('毛利率')}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
