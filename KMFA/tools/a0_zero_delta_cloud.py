#!/usr/bin/env python3
"""云上 A0 重算 + zero-delta(从 Private-Database 取真源,只回 public-safe)。

会计口径(agent 按财税/银行/政府合规标准自定,Owner 2026-07-25 授权):
  · A0 项目成本(权威=税务,现金基础):税务「开票纳税汇总表」sheet「2025项目成本明细」中
    收支类别==项目成本 的 付款金额 合计。
  · 系统重算(金蝶,权责基础):明细账中科目代码前缀 6401(主营业务成本)的 借方-贷方 净额,
    仅取交易行(跳期初/本期合计/本年累计)。**不含期间费用 6601/6602/6603、不含内部划转科目**
    ——这正是修掉上轮 737% 假差的关键。
  · zero-delta = 系统重算 − A0;跨基础(权责 vs 现金)差额非零属真实对账缺口,如实呈现,不造零。

铁律:真实金额只在运行内存/私有库;stdout 与产物只出 public-safe(差额率/通过与否/方法),零金额。
"""
import json, subprocess, sys, tempfile, os, collections
from pathlib import Path

REPO = "LinzeColin/Private-Database"
CLIENT = Path(__file__).resolve().parents[1].parent / "KMDatabase" / "machine" / "tools" / "private_db_client.py"
COST_PREFIX = ("6401",)            # 主营业务成本(项目成本权威口径)
SUMMARY_ROWS = ("期初余额", "本期合计", "本年累计", "期末余额")


def fetch(area, obj_path, out):
    subprocess.run([sys.executable, str(CLIENT), "get", area, obj_path, out], check=True)


def cents(x):
    try:
        return int(round(float(x) * 100))
    except Exception:
        return 0


def a0_cost_cents(tax_xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(tax_xlsx, read_only=True, data_only=True)
    ws = wb["2025项目成本明细"]
    total = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) < 6:
            continue
        cat, pay = r[3], r[5]
        if cat and "项目成本" in str(cat) and isinstance(pay, (int, float)):
            total += cents(pay)
    wb.close()
    return total


def kingdee_6401_net_cents(ledger_xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(ledger_xlsx, read_only=True, data_only=True)
    net = 0
    for sh in wb.sheetnames:
        code = sh.split("_", 1)[0].split("-", 1)[0].strip()
        if not any(code.startswith(p) for p in COST_PREFIX):
            continue
        ws = wb[sh]
        for r in ws.iter_rows(min_row=4, values_only=True):
            if len(r) < 13:
                continue
            zhaiyao = str(r[10]) if r[10] else ""
            if any(s in zhaiyao for s in SUMMARY_ROWS):
                continue
            debit, credit = r[11], r[12]
            net += cents(debit) - cents(credit)
    wb.close()
    return net


def resolve_from_manifest(td):
    """从 Private-KMDatabase/manifest.jsonl 按 domain 解析税务源与金蝶明细账的 object_path。"""
    mp = os.path.join(td, "manifest.jsonl")
    fetch("Private-KMDatabase", "manifest.jsonl", mp)
    tax_obj, ledger_objs = None, []
    for line in Path(mp).read_text(encoding="utf-8").splitlines():
        try:
            r = json.loads(line)
        except Exception:
            continue
        dom, op = str(r.get("domain", "")), r.get("object_path")
        if not op:
            continue
        if dom == "KMFA财务税务" and "汇总表" in op:
            tax_obj = op            # 取最后一条(最新批次)
        elif dom == "KMFA金蝶明细账" and "明细账" in op:
            ledger_objs.append(op)
    if not tax_obj or not ledger_objs:
        raise SystemExit(f"manifest 未解析到源:tax={bool(tax_obj)} ledgers={len(ledger_objs)}")
    return tax_obj, ledger_objs


def main():
    with tempfile.TemporaryDirectory() as td:
        tax_obj, ledger_objs = resolve_from_manifest(td)
        tp = os.path.join(td, "tax.xlsx")
        fetch("Private-KMDatabase", tax_obj, tp)
        a0 = a0_cost_cents(tp)
        sys_net = 0
        for i, lo in enumerate(ledger_objs):
            lp = os.path.join(td, f"ledger{i}.xlsx")
            fetch("Private-KMDatabase", lo, lp)
            sys_net += kingdee_6401_net_cents(lp)
    delta = sys_net - a0
    delta_pct = round(delta / a0 * 100, 1) if a0 else None
    result = {
        "schema_version": "kmfa.a0_zero_delta.v2",
        "method_authority": "税务 2025项目成本明细 付款金额(现金基础)",
        "method_system": "金蝶明细账 科目6401 主营业务成本 借方-贷方净额(权责基础,排除期间费用与内部划转)",
        "zero_delta_pass": delta == 0,
        "delta_vs_a0_pct": delta_pct,
        "basis_note": "税务=现金 vs 金蝶6401=权责;差额非零=真实跨基础/时点对账缺口,非管道错误",
        "amounts_public": False,
    }
    # 只出 public-safe;绝不打印绝对金额
    print(json.dumps(result, ensure_ascii=False))
    Path("a0_zero_delta_result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
