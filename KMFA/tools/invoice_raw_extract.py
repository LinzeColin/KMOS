#!/usr/bin/env python3
"""项目开票原始导出抽取 → `_staging.invoice_raw`（销项票逐票级，开票轴深化）。

来源：红圈「项目开票」导出两件（4,500+ 行/件）。要点：
- **坏维度标签**：导出件 `dimension` 声称 `A1:A1`，`read_only` 模式会只读出一列——必须 `reset_dimensions()`（新雷入册）。
- 逐票字段：发票号码/创建人/开票状态/开票日期/收票单位/合同名称/含税金额/税率；含税金额→整数分；
  税率原文字符串（如「13%」）；不含税/税额**不预计算**（除法舍入歧义留给对账时按票裁定）。
- 「发票回填」重复列组只取首列非空值。负数行=红字/冲减，如实入库。
- 注意：本源为**销项**票（对外开票）；材料轴 62 组需**进项**票判别——需求单第 6 项据此修正措辞。
用法：python3 KMFA/tools/invoice_raw_extract.py [--out <json>]
"""
from __future__ import annotations

import argparse
import hashlib
import json
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
PRIVATE = REPO / "KMFA" / ".codex_private_runtime"
DB_PATH = PRIVATE / "duckdb" / "kmfa_staging.duckdb"
INVENTORY = PRIVATE / "imports" / "excel_sheet_inventory" / "full_inventory.json"

VERSION = "invoice-raw-v1"
NAME_KEYWORD = "项目开票_导出文件"

COLUMNS = {
    "发票号码": ("invoice_no_raw", "raw"),
    "创建人": ("creator_ref", "raw"),
    "开票状态": ("status_ref", "raw"),
    "开票日期": ("invoice_date", "date"),
    "收票单位": ("customer_ref", "raw"),
    "合同名称": ("contract_name", "raw"),
    "本次开票含税金额(元)": ("gross_amount_cents", "cents"),
    "税率(%)": ("tax_rate_raw", "raw"),
    "发票回填": ("backfill_ref", "raw_first"),
}
SIGNATURE = ("invoice_no_raw", "customer_ref", "gross_amount_cents")
FIELDS = ["invoice_no_raw", "creator_ref", "status_ref", "invoice_date", "customer_ref",
          "contract_name", "gross_amount_cents", "tax_rate_raw", "backfill_ref"]

DDL = """CREATE TABLE IF NOT EXISTS _staging.invoice_raw (
    source_sha8 VARCHAR NOT NULL, sheet_hash VARCHAR NOT NULL, row_index INTEGER NOT NULL,
    invoice_no_raw VARCHAR, creator_ref VARCHAR, status_ref VARCHAR, invoice_date DATE,
    customer_ref VARCHAR, contract_name VARCHAR, gross_amount_cents BIGINT,
    tax_rate_raw VARCHAR, backfill_ref VARCHAR)"""


def norm(text) -> str:
    return unicodedata.normalize("NFKC", str(text)).strip().replace(" ", "").replace("\n", "")


NORM_COLUMNS = {norm(k): v for k, v in COLUMNS.items()}


class Money:
    def __init__(self) -> None:
        self.subcent = 0
        self.rejects = 0

    def to_cents(self, value) -> int | None:
        if value in (None, ""):
            return None
        try:
            dec = Decimal(str(value).replace(",", "").strip())
        except InvalidOperation:
            self.rejects += 1
            return None
        exact = dec * 100
        cents = exact.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        if exact != cents:
            self.subcent += 1
        return int(cents)


def to_iso(value):
    if value in (None, ""):
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = norm(value)
    for sep in ("-", "/", "."):
        parts = text.split(sep)
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            y, m, d = (int(p) for p in parts)
            if y < 100:
                y += 2000
            try:
                return date(y, m, d).isoformat()
            except ValueError:
                return None
    return None


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    import warnings
    warnings.filterwarnings("ignore")
    import duckdb
    import openpyxl

    inventory = json.loads(INVENTORY.read_text(encoding="utf-8"))
    manifest = {r["sha256"]: r for r in (json.loads(l) for l in
                (REPO / "KMDatabase/data/manifest.jsonl").read_text(encoding="utf-8").splitlines() if l.strip())}

    con = duckdb.connect(str(DB_PATH))
    con.execute("CREATE SCHEMA IF NOT EXISTS _staging")
    con.execute(DDL)

    money = Money()
    sheets_loaded, rows_loaded = 0, 0
    for f in sorted(inventory, key=lambda x: x["sha256"]):
        row = manifest[f["sha256"]]
        if NAME_KEYWORD not in row["original_name"]:
            continue
        path = REPO / "KMDatabase/data" / row["object_path"]
        sha8 = f["sha256"][:8]
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                sheet_hash = hashlib.sha256(sheet_name.encode()).hexdigest()[:12]
                idem = hashlib.sha256(f"{f['sha256']}|{sheet_name}|{VERSION}".encode()).hexdigest()
                if con.execute("SELECT 1 FROM _staging.extraction_manifest WHERE idempotency_key=?", [idem]).fetchone():
                    continue
                ws = wb[sheet_name]
                ws.reset_dimensions()  # 导出件维度标签坏（A1:A1），必须重置
                it = ws.iter_rows(values_only=True)
                header = [norm(c) if c is not None else "" for c in next(it)]
                mapping, backfill_cols = {}, []
                for col, cell in enumerate(header):
                    hit = NORM_COLUMNS.get(cell)
                    if not hit:
                        continue
                    if hit[1] == "raw_first":
                        backfill_cols.append(col)
                    elif hit[0] not in {v[0] for v in mapping.values()}:
                        mapping[col] = hit
                if not all(any(v[0] == s for v in mapping.values()) for s in SIGNATURE):
                    continue
                inserted = 0
                for r_i, data in enumerate(it, 2):
                    rec = dict.fromkeys(FIELDS)
                    for col, (canon, kind) in mapping.items():
                        v = data[col] if col < len(data) else None
                        if v in (None, ""):
                            continue
                        if kind == "cents":
                            rec[canon] = money.to_cents(v)
                        elif kind == "date":
                            rec[canon] = to_iso(v)
                        else:
                            rec[canon] = str(v).strip()
                    for col in backfill_cols:
                        v = data[col] if col < len(data) else None
                        if v not in (None, ""):
                            rec["backfill_ref"] = str(v).strip()
                            break
                    if not rec["invoice_no_raw"] or rec["gross_amount_cents"] is None:
                        continue
                    con.execute(f"INSERT INTO _staging.invoice_raw VALUES (?,?,?{',?' * len(FIELDS)})",
                                [sha8, sheet_hash, r_i, *[rec[k] for k in FIELDS]])
                    inserted += 1
                con.execute("INSERT INTO _staging.extraction_manifest VALUES (?,?,?,?,?,?,?,?,?,?)",
                            [f"sha256:{f['sha256']}", str(Path('KMDatabase/data') / row['object_path']), sheet_hash,
                             "_staging.invoice_raw", inserted, len(mapping) + (1 if backfill_cols else 0), 0,
                             datetime.now(), VERSION, idem])
                sheets_loaded += 1
                rows_loaded += inserted
        finally:
            wb.close()

    total, = con.execute("SELECT count(*) FROM _staging.invoice_raw").fetchone()
    con.close()
    summary = {
        "task_id": "TSK.KMFA.DATA.0007", "phase": "2_extract_invoice_raw",
        "extractor_version": VERSION, "sheets_loaded": sheets_loaded,
        "rows_loaded_this_run": rows_loaded, "table_total_rows": total,
        "amount_parse_rejections": money.rejects, "subcent_roundings": money.subcent,
        "amounts_integer_cents": True,
        "note": "销项票逐票级；不含税/税额不预计算（舍入歧义留待按票裁定）；维度标签坏须 reset_dimensions",
    }
    out = REPO / (args.out or "KMFA/stage_artifacts/DT5_DATA0047_invoice_raw/machine/extract_summary.json")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
