#!/usr/bin/env python3
"""Generate private candidate-fill draft for KMFA v0.1.4 source-map completion.

This phase uses the owner-authorized read-only raw inbox access to build a
private candidate draft for the 113 still-pending processed value source-map
completion items. Raw filenames, archive members, sheet names, headers, cell
values, PDF token context and diagnostics are written only to the ignored
private runtime. Public evidence contains aggregate counts and gate decisions
only.
"""

from __future__ import annotations

import argparse
import io
import json
import re
import subprocess
import sys
import zipfile
from copy import deepcopy
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from KMFA.tools.v014_s05_p1_a0_file_registration import (  # noqa: E402
    RAW_INBOX,
    is_hidden_zip_member,
    sha256_text,
    stat_snapshot,
)


PROJECT_ROOT = Path("KMFA")
PHASE_ID = "V014_PROCESSED_VALUE_SOURCE_MAP_COMPLETION_AUTO_CANDIDATE_DRAFT"
TASK_ID = "KMFA-V014-PROCESSED-VALUE-SOURCE-MAP-COMPLETION-AUTO-CANDIDATE-DRAFT-20260706"
ACCEPTANCE_ID = "ACC-V014-PROCESSED-VALUE-SOURCE-MAP-COMPLETION-AUTO-CANDIDATE-DRAFT"
VERSION = "0.1.4-processed-value-source-map-completion-auto-candidate-draft"
STATUS = "completed_validated_local_only_no_go_private_candidate_draft_owner_review_required"
NEXT_REQUIRED_INPUT = "owner_or_authorized_delegate_reviews_private_candidate_draft_and_confirms_authorized_completion_template"

ARTIFACT_DIR = PROJECT_ROOT / "stage_artifacts" / PHASE_ID
MACHINE_DIR = ARTIFACT_DIR / "machine"
HUMAN_DIR = ARTIFACT_DIR / "human"
SUMMARY_PATH = MACHINE_DIR / "processed_value_source_map_completion_auto_candidate_draft_summary.json"
MANIFEST_PATH = MACHINE_DIR / "processed_value_source_map_completion_auto_candidate_draft_manifest.json"
GO_NO_GO_PATH = MACHINE_DIR / "processed_value_source_map_completion_auto_candidate_draft_go_no_go_report.json"
REPORT_PATH = HUMAN_DIR / "processed_value_source_map_completion_auto_candidate_draft_report.md"
GO_NO_GO_RECORD_PATH = HUMAN_DIR / "go_no_go_record.md"
TEST_RESULTS_PATH = HUMAN_DIR / "test_results.md"
RISK_REGISTER_PATH = HUMAN_DIR / "risk_register.md"
ROLLBACK_PATH = HUMAN_DIR / "rollback_plan.md"

METADATA_SUMMARY_PATH = PROJECT_ROOT / "metadata/quality/v014_processed_value_source_map_completion_auto_candidate_draft_summary.json"
METADATA_MANIFEST_PATH = PROJECT_ROOT / "metadata/quality/v014_processed_value_source_map_completion_auto_candidate_draft_manifest.json"
METADATA_GO_NO_GO_PATH = PROJECT_ROOT / "metadata/quality/v014_processed_value_source_map_completion_auto_candidate_draft_go_no_go_report.json"
DEVELOPMENT_EVENTS_PATH = PROJECT_ROOT / "docs/governance/development_events.jsonl"

PRIVATE_OUTPUT_DIR = PROJECT_ROOT / ".codex_private_runtime/v014_processed_value_source_map_completion_auto_candidate_draft"
PRIVATE_RAW_INDEX_PATH = PRIVATE_OUTPUT_DIR / "private_raw_source_index.json"
PRIVATE_CANDIDATE_DRAFT_PATH = PRIVATE_OUTPUT_DIR / "auto_candidate_completion_template_draft.json"
PRIVATE_DIAGNOSTIC_PATH = PRIVATE_OUTPUT_DIR / "auto_match_diagnostic.json"
PRIVATE_QUESTION_LIST_PATH = PRIVATE_OUTPUT_DIR / "unmatched_question_list_zh.md"

PRIVATE_TEMPLATE_PATH = (
    PROJECT_ROOT
    / ".codex_private_runtime/v014_processed_value_source_map_completion_input_kit/owner_authorized_processed_value_source_map_completion_template.json"
)
PRIVATE_WORKLIST_PATH = (
    PROJECT_ROOT
    / ".codex_private_runtime/v014_private_processed_value_source_map_gap_resolution/private_owner_authorized_fill_worklist.json"
)
BLOCKER_SUMMARY_PATH = (
    PROJECT_ROOT
    / "stage_artifacts/V014_PROCESSED_VALUE_SOURCE_MAP_COMPLETION_BLOCKER_AUDIT/machine/processed_value_source_map_completion_blocker_audit_summary.json"
)

NUMERIC_TOKEN_RE = re.compile(r"(?<![A-Za-z0-9])[-+]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?%?(?![A-Za-z0-9])")
PENDING_VALUE = "PENDING_OWNER_OR_AUTHORIZED_DELEGATE_INPUT"
ACTION_SUPPLY = "supply_authorized_processed_value_fingerprint"
ACTION_KEEP_PENDING = "keep_pending"

CONTEXT_KEYWORDS: dict[str, tuple[str, ...]] = {
    "amount_value_private_ref": ("金额", "价款", "合计", "余额", "amount", "value", "total"),
    "collection_amount": ("回款", "收款", "已收", "到账", "collection", "receipt"),
    "revenue": ("收入", "营收", "产值", "销售", "revenue", "income"),
    "gross_margin_rate": ("毛利率", "利润率", "margin rate", "rate", "%"),
    "gross_profit": ("毛利", "利润", "gross profit", "profit"),
    "cash_paid_cost": ("已付", "付款", "支付", "现金", "成本", "paid", "cash", "cost"),
    "interest": ("利息", "interest"),
    "labor": ("人工", "工资", "劳务", "labor", "salary"),
    "machinery": ("机械", "设备", "machinery", "equipment"),
    "management_fee": ("管理费", "管理", "management"),
    "material": ("材料", "物料", "material"),
    "subcontract": ("分包", "外协", "subcontract"),
    "tax": ("税", "增值税", "tax"),
    "transport": ("运输", "运费", "物流", "transport", "freight"),
    "travel": ("差旅", "住宿", "交通", "travel"),
    "contract_amount": ("合同额", "合同金额", "合同价", "contract amount"),
    "cost_total": ("总成本", "成本合计", "成本", "total cost"),
    "invoice_amount": ("开票", "发票", "invoice"),
    "cash_gross_profit": ("现金毛利", "现金利润", "cash gross profit"),
    "management_project_cost": ("项目成本", "管理项目", "project cost"),
}
NON_NUMERIC_CONTEXTS = {"cost_category", "calculation_private_execution_ref"}


def _now(generated_at: str | None = None) -> str:
    return generated_at or datetime.now().astimezone().isoformat(timespec="seconds")


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"{path} must contain a JSON object")
    return value


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _git_output(args: list[str]) -> str:
    result = subprocess.run(
        ["git", "-c", "core.quotePath=false", *args],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if result.returncode != 0:
        return "UNKNOWN"
    return result.stdout.strip()


def _git_check_ignored(path: Path) -> bool:
    return subprocess.run(["git", "check-ignore", "-q", path.as_posix()], check=False).returncode == 0


def _decimal_text(value: Any) -> str | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            dec = Decimal(str(value).replace(",", ""))
        except (InvalidOperation, ValueError):
            return None
        return format(dec.normalize(), "f")
    if isinstance(value, str):
        normalized = value.strip().replace(",", "")
        if not normalized:
            return None
        if normalized.endswith("%"):
            normalized = normalized[:-1]
        try:
            dec = Decimal(normalized)
        except (InvalidOperation, ValueError):
            return None
        return format(dec.normalize(), "f")
    return None


def _decimal_fingerprint(value: Any) -> str | None:
    decimal_text = _decimal_text(value)
    return sha256_text(decimal_text) if decimal_text is not None else None


def _safe_cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def _source_ref(*parts: str) -> str:
    return sha256_text("||".join(parts))


def _public_raw_inventory() -> dict[str, int | bool]:
    files = sorted([path for path in RAW_INBOX.rglob("*") if path.is_file()]) if RAW_INBOX.exists() else []
    archive_count = sum(1 for path in files if path.suffix.lower() == ".zip")
    root_workbook_count = sum(1 for path in files if path.suffix.lower() in {".xlsx", ".xlsm"})
    other_count = len(files) - archive_count - root_workbook_count
    return {
        "raw_root_exists": RAW_INBOX.exists(),
        "raw_root_file_count": len(files),
        "raw_root_archive_count": archive_count,
        "raw_root_spreadsheet_count": root_workbook_count,
        "raw_root_other_file_count": other_count,
    }


def _fingerprint_workbook(binary: bytes, *, raw_file_name: str, archive_member_name: str | None, source_kind: str) -> dict[str, Any]:
    source_ref_hash = _source_ref(source_kind, raw_file_name, archive_member_name or "")
    source_record: dict[str, Any] = {
        "source_ref_hash": source_ref_hash,
        "source_kind": source_kind,
        "raw_file_name": raw_file_name,
        "archive_member_name": archive_member_name,
        "openable": False,
        "worksheet_count": 0,
        "non_empty_cell_count": 0,
        "numeric_cell_count": 0,
        "formula_cell_count": 0,
        "header_candidate_row_count": 0,
        "parse_error": "",
        "worksheets": [],
    }
    numeric_records: list[dict[str, Any]] = []
    try:
        formula_workbook = load_workbook(io.BytesIO(binary), data_only=False, read_only=True)
        value_workbook = load_workbook(io.BytesIO(binary), data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - depends on local workbook shape
        source_record["parse_error"] = type(exc).__name__
        return {"source_record": source_record, "numeric_records": numeric_records}

    source_record["openable"] = True
    source_record["worksheet_count"] = len(formula_workbook.worksheets)
    for formula_sheet, value_sheet in zip(formula_workbook.worksheets, value_workbook.worksheets):
        sheet_ref_hash = sha256_text(formula_sheet.title)
        sheet_record: dict[str, Any] = {
            "sheet_name": formula_sheet.title,
            "sheet_ref_hash": sheet_ref_hash,
            "header_candidate_rows": [],
            "numeric_record_count": 0,
        }
        header_by_col: dict[int, str] = {}
        for formula_row, value_row in zip(formula_sheet.iter_rows(), value_sheet.iter_rows()):
            row_idx = int(formula_row[0].row) if formula_row else 0
            row_strings: list[str] = []
            for formula_cell in formula_row:
                text = _safe_cell_text(formula_cell.value)
                if text:
                    source_record["non_empty_cell_count"] += 1
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                        source_record["formula_cell_count"] += 1
                    if not text.startswith("="):
                        row_strings.append(text)
                        if row_idx <= 30 and len(text) <= 80:
                            header_by_col[int(formula_cell.column)] = text
            if row_idx <= 20 and row_strings:
                sheet_record["header_candidate_rows"].append(
                    {
                        "row_index": row_idx,
                        "row_text_values": row_strings[:80],
                        "row_text_hash": sha256_text("||".join(row_strings)),
                    }
                )
            value_by_coordinate = {cell.coordinate: cell for cell in value_row}
            for formula_cell in formula_row:
                value_cell = value_by_coordinate.get(formula_cell.coordinate)
                value = value_cell.value if value_cell is not None else None
                fingerprint = _decimal_fingerprint(value)
                if fingerprint is None:
                    continue
                source_record["numeric_cell_count"] += 1
                sheet_record["numeric_record_count"] += 1
                decimal_text = _decimal_text(value)
                context_fragments = [
                    formula_sheet.title,
                    header_by_col.get(int(formula_cell.column), ""),
                    *row_strings[:12],
                ]
                context_text = " ".join(fragment for fragment in context_fragments if fragment)
                numeric_records.append(
                    {
                        "record_ref_hash": _source_ref(source_ref_hash, sheet_ref_hash, formula_cell.coordinate),
                        "record_kind": "workbook_cell",
                        "source_ref_hash": source_ref_hash,
                        "source_kind": source_kind,
                        "raw_file_name": raw_file_name,
                        "archive_member_name": archive_member_name,
                        "sheet_name": formula_sheet.title,
                        "sheet_ref_hash": sheet_ref_hash,
                        "cell_address": formula_cell.coordinate,
                        "cell_ref_hash": sha256_text(formula_cell.coordinate),
                        "raw_value": _safe_cell_text(value),
                        "normalized_decimal": decimal_text,
                        "numeric_value_fingerprint": fingerprint,
                        "is_percent_like": isinstance(value, str) and value.strip().endswith("%"),
                        "context_text": context_text,
                        "context_text_hash": sha256_text(context_text),
                    }
                )
        source_record["header_candidate_row_count"] += len(sheet_record["header_candidate_rows"])
        source_record["worksheets"].append(sheet_record)
    try:
        formula_workbook.close()
        value_workbook.close()
    except Exception:
        pass
    return {"source_record": source_record, "numeric_records": numeric_records}


def _fingerprint_pdf(binary: bytes, *, raw_file_name: str, archive_member_name: str | None, source_kind: str) -> dict[str, Any]:
    source_ref_hash = _source_ref(source_kind, raw_file_name, archive_member_name or "")
    source_record: dict[str, Any] = {
        "source_ref_hash": source_ref_hash,
        "source_kind": source_kind,
        "raw_file_name": raw_file_name,
        "archive_member_name": archive_member_name,
        "openable": False,
        "page_count": 0,
        "text_page_count": 0,
        "numeric_token_count": 0,
        "parse_error": "",
    }
    numeric_records: list[dict[str, Any]] = []
    try:
        with pdfplumber.open(io.BytesIO(binary)) as pdf:
            source_record["openable"] = True
            source_record["page_count"] = len(pdf.pages)
            for page_index, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    source_record["text_page_count"] += 1
                for token_index, match in enumerate(NUMERIC_TOKEN_RE.finditer(text), 1):
                    token = match.group(0)
                    fingerprint = _decimal_fingerprint(token)
                    if fingerprint is None:
                        continue
                    source_record["numeric_token_count"] += 1
                    start = max(match.start() - 60, 0)
                    end = min(match.end() + 60, len(text))
                    context_text = " ".join(text[start:end].split())
                    numeric_records.append(
                        {
                            "record_ref_hash": _source_ref(source_ref_hash, str(page_index), str(token_index)),
                            "record_kind": "pdf_numeric_token",
                            "source_ref_hash": source_ref_hash,
                            "source_kind": source_kind,
                            "raw_file_name": raw_file_name,
                            "archive_member_name": archive_member_name,
                            "page_index": page_index,
                            "token_index": token_index,
                            "raw_value": token,
                            "normalized_decimal": _decimal_text(token),
                            "numeric_value_fingerprint": fingerprint,
                            "is_percent_like": token.endswith("%"),
                            "context_text": context_text,
                            "context_text_hash": sha256_text(context_text),
                        }
                    )
    except Exception as exc:  # pragma: no cover - depends on local PDF shape
        source_record["parse_error"] = type(exc).__name__
    return {"source_record": source_record, "numeric_records": numeric_records}


def _scan_raw_sources() -> tuple[dict[str, Any], dict[str, Any]]:
    raw_root_before = stat_snapshot(RAW_INBOX) if RAW_INBOX.exists() else {}
    public = {
        **_public_raw_inventory(),
        "archive_member_count": 0,
        "archive_workbook_member_count": 0,
        "archive_pdf_member_count": 0,
        "openable_workbook_count": 0,
        "openable_pdf_count": 0,
        "workbook_parse_error_count": 0,
        "pdf_parse_error_count": 0,
        "raw_numeric_candidate_count": 0,
        "raw_unique_numeric_fingerprint_count": 0,
        "raw_value_fingerprints_generated": False,
        "raw_root_stat_unchanged_after_auto_candidate_draft": False,
    }
    private: dict[str, Any] = {
        "schema_version": "kmfa.private.v014_processed_value_source_map_completion_auto_candidate_raw_index.v1",
        "classification": "private_raw_source_index_do_not_commit",
        "raw_data_inbox": str(RAW_INBOX),
        "raw_root_before": raw_root_before,
        "source_records": [],
        "numeric_records": [],
        "parse_errors": [],
    }

    if RAW_INBOX.exists():
        for path in sorted(RAW_INBOX.rglob("*")):
            if not path.is_file():
                continue
            suffix = path.suffix.lower()
            if suffix in {".xlsx", ".xlsm"}:
                result = _fingerprint_workbook(
                    path.read_bytes(),
                    raw_file_name=path.name,
                    archive_member_name=None,
                    source_kind="raw_root_workbook",
                )
                private["source_records"].append(result["source_record"])
                private["numeric_records"].extend(result["numeric_records"])
                if result["source_record"]["openable"]:
                    public["openable_workbook_count"] += 1
                if result["source_record"]["parse_error"]:
                    public["workbook_parse_error_count"] += 1
                    private["parse_errors"].append(
                        {
                            "source_ref_hash": result["source_record"]["source_ref_hash"],
                            "raw_file_name": path.name,
                            "error_class": result["source_record"]["parse_error"],
                        }
                    )
            elif suffix == ".zip":
                try:
                    with zipfile.ZipFile(path) as archive:
                        for info in archive.infolist():
                            if info.is_dir() or is_hidden_zip_member(info.filename):
                                continue
                            public["archive_member_count"] += 1
                            member_suffix = Path(info.filename).suffix.lower()
                            if member_suffix in {".xlsx", ".xlsm"}:
                                public["archive_workbook_member_count"] += 1
                                result = _fingerprint_workbook(
                                    archive.read(info),
                                    raw_file_name=path.name,
                                    archive_member_name=info.filename,
                                    source_kind="raw_archive_workbook_member",
                                )
                                private["source_records"].append(result["source_record"])
                                private["numeric_records"].extend(result["numeric_records"])
                                if result["source_record"]["openable"]:
                                    public["openable_workbook_count"] += 1
                                if result["source_record"]["parse_error"]:
                                    public["workbook_parse_error_count"] += 1
                                    private["parse_errors"].append(
                                        {
                                            "source_ref_hash": result["source_record"]["source_ref_hash"],
                                            "raw_file_name": path.name,
                                            "archive_member_name": info.filename,
                                            "error_class": result["source_record"]["parse_error"],
                                        }
                                    )
                            elif member_suffix == ".pdf":
                                public["archive_pdf_member_count"] += 1
                                result = _fingerprint_pdf(
                                    archive.read(info),
                                    raw_file_name=path.name,
                                    archive_member_name=info.filename,
                                    source_kind="raw_archive_pdf_member",
                                )
                                private["source_records"].append(result["source_record"])
                                private["numeric_records"].extend(result["numeric_records"])
                                if result["source_record"]["openable"]:
                                    public["openable_pdf_count"] += 1
                                if result["source_record"]["parse_error"]:
                                    public["pdf_parse_error_count"] += 1
                                    private["parse_errors"].append(
                                        {
                                            "source_ref_hash": result["source_record"]["source_ref_hash"],
                                            "raw_file_name": path.name,
                                            "archive_member_name": info.filename,
                                            "error_class": result["source_record"]["parse_error"],
                                        }
                                    )
                except Exception as exc:  # pragma: no cover - depends on local archive shape
                    private["parse_errors"].append(
                        {
                            "source_ref_hash": _source_ref("raw_archive", path.name),
                            "raw_file_name": path.name,
                            "error_class": type(exc).__name__,
                        }
                    )

    raw_root_after = stat_snapshot(RAW_INBOX) if RAW_INBOX.exists() else {}
    private["raw_root_after"] = raw_root_after
    public["raw_root_stat_unchanged_after_auto_candidate_draft"] = raw_root_before == raw_root_after
    public["raw_numeric_candidate_count"] = len(private["numeric_records"])
    public["raw_unique_numeric_fingerprint_count"] = len(
        {record["numeric_value_fingerprint"] for record in private["numeric_records"]}
    )
    public["raw_value_fingerprints_generated"] = public["raw_numeric_candidate_count"] > 0
    return public, private


def _score_record(record: dict[str, Any], context_group: str) -> int:
    if context_group in NON_NUMERIC_CONTEXTS:
        return 0
    text = f"{record.get('context_text', '')} {record.get('raw_value', '')}".lower()
    keywords = CONTEXT_KEYWORDS.get(context_group, (context_group.replace("_", " "),))
    score = 0
    for keyword in keywords:
        if keyword.lower() in text:
            score += 3
    if context_group == "gross_margin_rate" and record.get("is_percent_like"):
        score += 4
    if context_group.endswith("_amount") and record.get("record_kind") == "workbook_cell":
        score += 1
    return score


def _build_context_candidates(numeric_records: list[dict[str, Any]], context_groups: list[str]) -> dict[str, dict[str, Any]]:
    context_candidates: dict[str, dict[str, Any]] = {}
    for group in sorted(set(context_groups)):
        scored: list[dict[str, Any]] = []
        for record in numeric_records:
            score = _score_record(record, group)
            if score <= 0:
                continue
            scored.append({"score": score, **record})
        scored.sort(key=lambda item: (-int(item["score"]), item.get("source_ref_hash", ""), item.get("record_ref_hash", "")))
        unique_fingerprints = sorted({item["numeric_value_fingerprint"] for item in scored})
        if group in NON_NUMERIC_CONTEXTS:
            status = "requires_non_numeric_owner_mapping"
        elif not scored:
            status = "auto_unmatched_requires_owner_review"
        elif len(unique_fingerprints) == 1:
            status = "auto_unique_candidate_requires_owner_confirmation"
        else:
            status = "auto_ambiguous_multiple_candidates_requires_owner_review"
        context_candidates[group] = {
            "context_group": group,
            "candidate_status": status,
            "candidate_record_count": len(scored),
            "candidate_unique_numeric_fingerprint_count": len(unique_fingerprints),
            "candidate_unique_numeric_fingerprints": unique_fingerprints[:50],
            "top_candidate_records": scored[:20],
        }
    return context_candidates


def _candidate_item(original_item: dict[str, Any], context_candidate: dict[str, Any]) -> dict[str, Any]:
    item = deepcopy(original_item)
    status = context_candidate["candidate_status"]
    top_records = context_candidate["top_candidate_records"]
    unique_fingerprints = context_candidate["candidate_unique_numeric_fingerprints"]
    if status == "auto_unique_candidate_requires_owner_confirmation" and unique_fingerprints:
        suggested_action = ACTION_SUPPLY
        suggested_fingerprint = unique_fingerprints[0]
        basis_ref = f"private://{PRIVATE_OUTPUT_DIR.name}/context/{item.get('context_group')}/unique-candidate"
    elif status == "requires_non_numeric_owner_mapping":
        suggested_action = ACTION_KEEP_PENDING
        suggested_fingerprint = PENDING_VALUE
        basis_ref = "OWNER_REVIEW_REQUIRED_NON_NUMERIC_OR_CALCULATION_CONTEXT"
    else:
        suggested_action = ACTION_KEEP_PENDING
        suggested_fingerprint = PENDING_VALUE
        basis_ref = "OWNER_REVIEW_REQUIRED_AMBIGUOUS_OR_UNMATCHED_RAW_CANDIDATE"
    item["auto_candidate_fill"] = {
        "draft_only_not_owner_authorization": True,
        "candidate_status": status,
        "candidate_suggested_action_code": suggested_action,
        "candidate_authorized_processed_value_fingerprint": suggested_fingerprint,
        "candidate_authorized_source_basis_reference": basis_ref,
        "candidate_record_count": context_candidate["candidate_record_count"],
        "candidate_unique_numeric_fingerprint_count": context_candidate["candidate_unique_numeric_fingerprint_count"],
        "private_top_candidate_records": top_records,
        "owner_review_required": True,
    }
    return item


def _build_candidate_draft(
    *,
    generated_at: str,
    template: dict[str, Any],
    worklist: dict[str, Any],
    raw_private_index: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any], str]:
    completion_items = template.get("completion_items", [])
    if not isinstance(completion_items, list):
        raise ValueError("completion_items must be a list")
    context_groups = [str(item.get("context_group", "")) for item in completion_items if isinstance(item, dict)]
    context_candidates = _build_context_candidates(raw_private_index["numeric_records"], context_groups)
    candidate_items = [
        _candidate_item(item, context_candidates.get(str(item.get("context_group", "")), {"candidate_status": "auto_unmatched_requires_owner_review", "candidate_record_count": 0, "candidate_unique_numeric_fingerprint_count": 0, "candidate_unique_numeric_fingerprints": [], "top_candidate_records": []}))
        for item in completion_items
        if isinstance(item, dict)
    ]
    status_counts: dict[str, int] = {}
    for item in candidate_items:
        status = item["auto_candidate_fill"]["candidate_status"]
        status_counts[status] = status_counts.get(status, 0) + 1
    unresolved_items = [
        item
        for item in candidate_items
        if item["auto_candidate_fill"]["candidate_status"] != "auto_unique_candidate_requires_owner_confirmation"
    ]
    draft = {
        "schema_version": "kmfa.private.v014_processed_value_source_map_completion_auto_candidate_draft.v1",
        "classification": "private_auto_candidate_completion_template_draft_do_not_commit",
        "record_type": "v014_processed_value_source_map_completion_auto_candidate_draft",
        "project_id": "KMFA",
        "version": VERSION,
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "generated_at": generated_at,
        "draft_only_not_active_owner_authorization": True,
        "source_completion_template_ref": "private_runtime_only",
        "source_owner_worklist_ref": "private_runtime_only",
        "completion_template_overwritten": False,
        "active_owner_authorized_fill_record_written": False,
        "candidate_draft_item_count": len(candidate_items),
        "candidate_status_counts": dict(sorted(status_counts.items())),
        "context_candidates": context_candidates,
        "candidate_completion_items": candidate_items,
        "source_worklist_summary": worklist.get("owner_worklist_summary", {}),
    }
    diagnostic = {
        "schema_version": "kmfa.private.v014_processed_value_source_map_completion_auto_match_diagnostic.v1",
        "classification": "private_auto_match_diagnostic_do_not_commit",
        "record_type": "v014_processed_value_source_map_completion_auto_match_diagnostic",
        "project_id": "KMFA",
        "version": VERSION,
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "generated_at": generated_at,
        "candidate_status_counts": dict(sorted(status_counts.items())),
        "unresolved_item_count": len(unresolved_items),
        "context_candidate_count": len(context_candidates),
        "raw_numeric_candidate_count": len(raw_private_index["numeric_records"]),
        "raw_unique_numeric_fingerprint_count": len(
            {record["numeric_value_fingerprint"] for record in raw_private_index["numeric_records"]}
        ),
        "unresolved_items": unresolved_items,
    }
    question_lines = [
        "# KMFA v0.1.4 private completion template 候选草案未匹配问题清单",
        "",
        "说明：本文件包含 raw 文件名、表头、金额、单元格/PDF 位置和诊断细节，只能保留在 `.codex_private_runtime/`，不得提交 GitHub。",
        "",
        f"- 候选草案条目数：{len(candidate_items)}",
        f"- 需要 owner/授权代理确认的条目数：{len(unresolved_items)}",
        f"- 可唯一候选但仍需确认的条目数：{status_counts.get('auto_unique_candidate_requires_owner_confirmation', 0)}",
        "",
        "## 需要确认的问题",
    ]
    for index, item in enumerate(unresolved_items, 1):
        fill = item["auto_candidate_fill"]
        question_lines.extend(
            [
                "",
                f"### Q{index:03d} target_slot_id={item.get('target_slot_id')}",
                f"- context_group：{item.get('context_group')}",
                f"- 当前状态：{fill['candidate_status']}",
                f"- 候选 raw 记录数：{fill['candidate_record_count']}",
                f"- 候选唯一数值指纹数：{fill['candidate_unique_numeric_fingerprint_count']}",
                "- 请确认：应选择哪个 private raw 候选作为授权 processed value fingerprint，或确认继续 keep_pending。",
            ]
        )
        for candidate_index, candidate in enumerate(fill.get("private_top_candidate_records", [])[:5], 1):
            question_lines.extend(
                [
                    f"  - 候选 {candidate_index}:",
                    f"    - raw_file_name: {candidate.get('raw_file_name')}",
                    f"    - archive_member_name: {candidate.get('archive_member_name')}",
                    f"    - record_kind: {candidate.get('record_kind')}",
                    f"    - sheet_name: {candidate.get('sheet_name')}",
                    f"    - cell_address/page_token: {candidate.get('cell_address') or (str(candidate.get('page_index')) + '/' + str(candidate.get('token_index')) if candidate.get('page_index') else '')}",
                    f"    - raw_value: {candidate.get('raw_value')}",
                    f"    - numeric_value_fingerprint: {candidate.get('numeric_value_fingerprint')}",
                    f"    - context_text: {candidate.get('context_text')}",
                ]
            )
    return draft, diagnostic, "\n".join(question_lines) + "\n"


def _raw_boundary(raw_public_summary: dict[str, Any]) -> dict[str, bool]:
    return {
        "user_authorized_raw_data_read_for_this_phase": True,
        "raw_data_root_readonly_policy_active": True,
        "raw_inbox_read_performed_by_this_phase": True,
        "raw_inbox_list_performed_by_this_phase": True,
        "raw_inbox_stat_performed_by_this_phase": True,
        "raw_inbox_hash_or_value_fingerprint_performed_by_this_phase": True,
        "raw_root_stat_unchanged_after_auto_candidate_draft": bool(
            raw_public_summary.get("raw_root_stat_unchanged_after_auto_candidate_draft")
        ),
        "raw_inbox_write_performed_by_this_phase": False,
        "raw_inbox_delete_performed_by_this_phase": False,
        "raw_inbox_move_performed_by_this_phase": False,
        "raw_inbox_rename_performed_by_this_phase": False,
        "raw_inbox_overwrite_performed_by_this_phase": False,
        "raw_inbox_copy_performed_by_this_phase": False,
        "raw_inbox_normalize_performed_by_this_phase": False,
        "raw_inbox_mutated_by_this_phase": False,
    }


def _public_safety() -> dict[str, bool]:
    return {
        "public_safe_aggregate_only": True,
        "private_raw_index_committed": False,
        "private_candidate_draft_committed": False,
        "private_match_diagnostic_committed": False,
        "private_question_list_committed": False,
        "raw_file_committed": False,
        "raw_filename_committed": False,
        "raw_archive_member_name_committed": False,
        "raw_digest_committed": False,
        "source_document_committed": False,
        "office_workbook_committed": False,
        "field_header_plaintext_committed": False,
        "row_value_committed": False,
        "business_value_committed": False,
        "credential_or_secret_committed": False,
    }


def _write_public_artifacts(
    *,
    generated_at: str,
    raw_public_summary: dict[str, Any],
    candidate_draft: dict[str, Any],
    diagnostic: dict[str, Any],
    blocker_summary: dict[str, Any],
) -> dict[str, Any]:
    counts = candidate_draft["candidate_status_counts"]
    unique_candidate_item_count = counts.get("auto_unique_candidate_requires_owner_confirmation", 0)
    unresolved_item_count = diagnostic["unresolved_item_count"]
    private_raw_index_gitignored = _git_check_ignored(PRIVATE_RAW_INDEX_PATH)
    private_candidate_draft_gitignored = _git_check_ignored(PRIVATE_CANDIDATE_DRAFT_PATH)
    private_diagnostic_gitignored = _git_check_ignored(PRIVATE_DIAGNOSTIC_PATH)
    private_question_list_gitignored = _git_check_ignored(PRIVATE_QUESTION_LIST_PATH)
    summary = {
        "schema_version": "kmfa.v014_processed_value_source_map_completion_auto_candidate_draft_summary.v1",
        "record_type": "v014_processed_value_source_map_completion_auto_candidate_draft_summary",
        "project_id": "KMFA",
        "version": VERSION,
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "status": STATUS,
        "generated_at": generated_at,
        "source_blocker_phase_id": blocker_summary.get("phase_id"),
        "source_blocker_decision": blocker_summary.get("decision"),
        "completion_template_item_count": candidate_draft["candidate_draft_item_count"],
        "candidate_draft_item_count": candidate_draft["candidate_draft_item_count"],
        "auto_unique_candidate_item_count": unique_candidate_item_count,
        "auto_ambiguous_candidate_item_count": counts.get("auto_ambiguous_multiple_candidates_requires_owner_review", 0),
        "auto_unmatched_item_count": counts.get("auto_unmatched_requires_owner_review", 0),
        "non_numeric_or_calculation_context_item_count": counts.get("requires_non_numeric_owner_mapping", 0),
        "owner_review_required_item_count": candidate_draft["candidate_draft_item_count"],
        "unresolved_question_item_count": unresolved_item_count,
        "private_raw_index_written": True,
        "private_candidate_draft_written": True,
        "private_match_diagnostic_written": True,
        "private_question_list_written": True,
        "private_raw_index_gitignored": private_raw_index_gitignored,
        "private_candidate_draft_gitignored": private_candidate_draft_gitignored,
        "private_match_diagnostic_gitignored": private_diagnostic_gitignored,
        "private_question_list_gitignored": private_question_list_gitignored,
        "completion_template_overwritten": False,
        "active_owner_authorized_fill_record_written": False,
        "authorized_completion_record_supplied": False,
        "source_map_completion_reapplication_ready": False,
        "source_map_completion_reapplication_performed": False,
        "source_map_records_applied_count": 0,
        "processed_value_materialization_replay_performed": False,
        "raw_to_processed_value_comparison_performed": False,
        "business_value_consistency_verified": False,
        "lineage_full_check_complete": False,
        "formal_report_allowed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
        "decision": "NO_GO",
        "diagnostic_conclusion": "private_candidate_draft_ready_owner_review_required",
        "next_required_input": NEXT_REQUIRED_INPUT,
        "raw_private_extraction_summary": raw_public_summary,
        "raw_boundary": _raw_boundary(raw_public_summary),
        "public_safety": _public_safety(),
    }
    go_no_go = {
        "schema_version": "kmfa.v014_processed_value_source_map_completion_auto_candidate_draft_go_no_go.v1",
        "record_type": "v014_processed_value_source_map_completion_auto_candidate_draft_go_no_go_report",
        "project_id": "KMFA",
        "version": VERSION,
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "generated_at": generated_at,
        "decision": "NO_GO",
        "status": STATUS,
        "diagnostic_conclusion": "private_candidate_draft_ready_owner_review_required",
        "candidate_draft_item_count": summary["candidate_draft_item_count"],
        "auto_unique_candidate_item_count": unique_candidate_item_count,
        "unresolved_question_item_count": unresolved_item_count,
        "source_map_completion_reapplication_ready": False,
        "raw_to_processed_value_comparison_performed": False,
        "business_value_consistency_verified": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
        "blocked_until": NEXT_REQUIRED_INPUT,
    }
    manifest = {
        "schema_version": "kmfa.v014_processed_value_source_map_completion_auto_candidate_draft_manifest.v1",
        "record_type": "v014_processed_value_source_map_completion_auto_candidate_draft_manifest",
        "project_id": "KMFA",
        "version": VERSION,
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_ids": [ACCEPTANCE_ID],
        "generated_at": generated_at,
        "status": STATUS,
        "summary": summary,
        "go_no_go": go_no_go,
        "artifact_refs": {
            "machine_summary": SUMMARY_PATH.as_posix(),
            "machine_manifest": MANIFEST_PATH.as_posix(),
            "machine_go_no_go": GO_NO_GO_PATH.as_posix(),
            "private_raw_index": "private_runtime_only",
            "private_candidate_draft": "private_runtime_only",
            "private_match_diagnostic": "private_runtime_only",
            "private_question_list": "private_runtime_only",
        },
        "validator": (
            "PYTHONDONTWRITEBYTECODE=1 PYTHONPATH=. "
            "python3 KMFA/tools/check_v014_processed_value_source_map_completion_auto_candidate_draft.py "
            "--require-private-draft"
        ),
        "git": {
            "branch": _git_output(["branch", "--show-current"]),
            "head": _git_output(["rev-parse", "HEAD"]),
            "status_short": _git_output(["status", "--short", "--branch"]),
        },
    }
    report = f"""# V014 Processed Value Source-map Completion Auto Candidate Draft

Decision: NO_GO

This phase generated a private candidate draft for the pending source-map completion template. It is not an active owner authorization record and cannot be applied without owner or authorized-delegate review.

## Public-safe aggregate result

- Completion template items: {summary["completion_template_item_count"]}
- Private candidate draft items: {summary["candidate_draft_item_count"]}
- Unique candidate items requiring confirmation: {summary["auto_unique_candidate_item_count"]}
- Ambiguous candidate items requiring review: {summary["auto_ambiguous_candidate_item_count"]}
- Unmatched items requiring questions: {summary["auto_unmatched_item_count"]}
- Non-numeric or calculation-context items requiring review: {summary["non_numeric_or_calculation_context_item_count"]}
- Raw numeric candidate count: {raw_public_summary["raw_numeric_candidate_count"]}
- Raw unique numeric fingerprint count: {raw_public_summary["raw_unique_numeric_fingerprint_count"]}
- Raw root stat unchanged: {raw_public_summary["raw_root_stat_unchanged_after_auto_candidate_draft"]}

## Boundary

- Raw source access was read-only and owner-authorized for this phase.
- Raw source mutation, copying into Git, normalization, overwrite, move, rename and delete were not performed.
- Raw filenames, archive member names, sheet names, headers, row values, business values and diagnostics were written only to ignored private runtime.
- Source-map reapplication, processed materialization replay, raw-to-processed value comparison, formal report, GitHub upload, app reinstall and business execution were not performed.

Next required input: `{NEXT_REQUIRED_INPUT}`.
"""
    go_no_go_record = f"""# Go/No-Go Record

- phase_id: `{PHASE_ID}`
- decision: `NO_GO`
- reason: private candidate draft is ready but owner/authorized-delegate review is required before any active completion template or source map application.
- source-map reapplication ready: `false`
- github upload performed: `false`
"""
    risk_register = """# Risk Register

- Risk: automatic context matching can be ambiguous when several raw numeric candidates share similar field context.
  Mitigation: the generated file is a private draft only and keeps every item owner-review-required.
- Risk: raw business detail leakage.
  Mitigation: public artifacts are aggregate-only and private files remain under ignored runtime.
- Risk: false reconciliation claim.
  Mitigation: this phase does not compare raw and processed values or verify business consistency.
"""
    rollback_plan = """# Rollback Plan

No raw file was modified. To roll back this phase, remove the generated public phase artifacts and metadata quality copies, then remove the ignored private candidate-draft directory if no longer needed.
"""
    test_results = f"""# Test Results

Pending until validator and focused tests are run after artifact generation.

Expected validator:
`PYTHONDONTWRITEBYTECODE=1 PYTHONPATH=. python3 KMFA/tools/check_v014_processed_value_source_map_completion_auto_candidate_draft.py --require-private-draft`
"""
    for path, payload in (
        (SUMMARY_PATH, summary),
        (MANIFEST_PATH, manifest),
        (GO_NO_GO_PATH, go_no_go),
        (METADATA_SUMMARY_PATH, summary),
        (METADATA_MANIFEST_PATH, manifest),
        (METADATA_GO_NO_GO_PATH, go_no_go),
    ):
        _write_json(path, payload)
    for path, text in (
        (REPORT_PATH, report),
        (GO_NO_GO_RECORD_PATH, go_no_go_record),
        (RISK_REGISTER_PATH, risk_register),
        (ROLLBACK_PATH, rollback_plan),
        (TEST_RESULTS_PATH, test_results),
    ):
        _write_text(path, text)
    return manifest


def _append_development_event(generated_at: str, manifest: dict[str, Any]) -> None:
    event_id = "DEV-KMFA-20260706-V014-PROCESSED-VALUE-SOURCE-MAP-COMPLETION-AUTO-CANDIDATE-DRAFT"
    event = {
        "event_id": event_id,
        "event_time": generated_at,
        "event_type": "development",
        "project_id": "KMFA",
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "version": VERSION,
        "iteration_id": "ITER-20260706-KMFA-V014-PROCESSED-VALUE-SOURCE-MAP-COMPLETION-AUTO-CANDIDATE-DRAFT",
        "fact_level": "EXTRACTED",
        "evidence_ref": MANIFEST_PATH.as_posix(),
        "go_no_go": manifest["go_no_go"]["decision"],
        "candidate_draft_item_count": manifest["summary"]["candidate_draft_item_count"],
        "unresolved_question_item_count": manifest["summary"]["unresolved_question_item_count"],
        "raw_inbox_read_performed": True,
        "raw_inbox_mutation_performed": False,
        "raw_to_processed_value_comparison_performed": False,
        "business_value_consistency_verified": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
        "result_commit": "PENDING",
        "summary": (
            "Generated a private-only auto candidate draft and Chinese question list for pending processed value "
            "source-map completion items. Public evidence is aggregate-only and downstream application remains blocked "
            "until owner or authorized-delegate confirmation."
        ),
    }
    DEVELOPMENT_EVENTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    existing = DEVELOPMENT_EVENTS_PATH.read_text(encoding="utf-8").splitlines() if DEVELOPMENT_EVENTS_PATH.exists() else []
    filtered = [line for line in existing if f'"event_id":"{event_id}"' not in line and f'"event_id": "{event_id}"' not in line]
    filtered.append(json.dumps(event, ensure_ascii=False, sort_keys=True))
    DEVELOPMENT_EVENTS_PATH.write_text("\n".join(filtered) + "\n", encoding="utf-8")


def generate(*, generated_at: str | None = None, write_governance_event: bool = True) -> dict[str, Any]:
    timestamp = _now(generated_at)
    template = _read_json(PRIVATE_TEMPLATE_PATH)
    worklist = _read_json(PRIVATE_WORKLIST_PATH)
    blocker_summary = _read_json(BLOCKER_SUMMARY_PATH)
    raw_public_summary, raw_private_index = _scan_raw_sources()

    raw_private_index.update(
        {
            "project_id": "KMFA",
            "version": VERSION,
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "generated_at": timestamp,
        }
    )
    candidate_draft, diagnostic, question_list = _build_candidate_draft(
        generated_at=timestamp,
        template=template,
        worklist=worklist,
        raw_private_index=raw_private_index,
    )

    _write_json(PRIVATE_RAW_INDEX_PATH, raw_private_index)
    _write_json(PRIVATE_CANDIDATE_DRAFT_PATH, candidate_draft)
    _write_json(PRIVATE_DIAGNOSTIC_PATH, diagnostic)
    _write_text(PRIVATE_QUESTION_LIST_PATH, question_list)

    manifest = _write_public_artifacts(
        generated_at=timestamp,
        raw_public_summary=raw_public_summary,
        candidate_draft=candidate_draft,
        diagnostic=diagnostic,
        blocker_summary=blocker_summary,
    )
    if write_governance_event:
        _append_development_event(timestamp, manifest)
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--generated-at")
    args = parser.parse_args()
    manifest = generate(generated_at=args.generated_at)
    print(
        "PASS: KMFA v0.1.4 private auto candidate draft generated "
        f"(decision={manifest['go_no_go']['decision']}, "
        f"candidate_items={manifest['summary']['candidate_draft_item_count']}, "
        f"owner_review_required={manifest['summary']['owner_review_required_item_count']})"
    )


if __name__ == "__main__":
    main()
