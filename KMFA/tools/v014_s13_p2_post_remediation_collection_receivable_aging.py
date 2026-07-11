#!/usr/bin/env python3
"""Generate current public-safe KMFA v0.1.4 S13-P2 receivable evidence."""

from __future__ import annotations

import argparse
import functools
import hashlib
import html
import http.server
import io
import json
import os
import socketserver
import subprocess
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

from KMFA.tools import v014_s13_p1_post_remediation_financial_operating_report as s13_p1
from KMFA.tools.check_v014_s13_p1_post_remediation_financial_operating_report import (
    validate_v014_s13_p1_post_remediation_financial_operating_report,
)


PHASE_ID = "V014_S13_P2_POST_REMEDIATION_COLLECTION_RECEIVABLE_AGING"
ROADMAP_PHASE_ID = "S13-P2"
TASK_ID = "KMFA-V014-S13-P2-POST-REMEDIATION-COLLECTION-RECEIVABLE-AGING-20260711"
ACCEPTANCE_ID = "ACC-V014-S13-P2-POST-REMEDIATION-COLLECTION-RECEIVABLE-AGING"
VERSION = "0.1.4-s13-p2-post-remediation-collection-receivable-aging"
STATUS = "completed_validated_local_only_s13_p2_method_locked_business_items_unproven_upload_deferred"
DECISION = "NO_GO"
FORMULA_ID = "FORM-KMFA-V014-S13P2-POST-REMEDIATION-COLLECTION-RECEIVABLE-AGING-001"
PARAMETER_IDS = (
    "PARAM-KMFA-1739",
    "PARAM-KMFA-1740",
    "PARAM-KMFA-1741",
    "PARAM-KMFA-1742",
)
MODEL_REGISTRY_KEY = "kmfa_v014_s13_p2_post_remediation_collection_receivable_aging"

OUTPUT_DIR = Path("KMFA/stage_artifacts") / PHASE_ID
MACHINE_DIR = OUTPUT_DIR / "machine"
HUMAN_DIR = OUTPUT_DIR / "human"
HTML_DIR = OUTPUT_DIR / "exports/html"
SUMMARY_PATH = MACHINE_DIR / "collection_receivable_aging_summary.json"
MANIFEST_PATH = MACHINE_DIR / "collection_receivable_aging_manifest.json"
LANES_PATH = MACHINE_DIR / "source_lane_status_public_safe.json"
ISSUES_PATH = MACHINE_DIR / "issue_definitions_public_safe.json"
MATRIX_PATH = MACHINE_DIR / "collection_receivable_aging_acceptance_matrix.json"
GO_NO_GO_PATH = MACHINE_DIR / "collection_receivable_aging_go_no_go.json"
HTML_PATH = HTML_DIR / "collection_receivable_aging_workbench.html"
REPORT_PATH = HUMAN_DIR / "collection_receivable_aging_report_zh.md"
TEST_RESULTS_PATH = HUMAN_DIR / "test_results_zh.md"
RISK_REGISTER_PATH = HUMAN_DIR / "risk_register_zh.md"
ROLLBACK_PATH = HUMAN_DIR / "rollback_plan_zh.md"

QUALITY_DIR = Path("KMFA/metadata/quality")
METADATA_SUMMARY_PATH = QUALITY_DIR / "v014_s13_p2_post_remediation_collection_receivable_aging_summary.json"
METADATA_MANIFEST_PATH = QUALITY_DIR / "v014_s13_p2_post_remediation_collection_receivable_aging_manifest.json"
METADATA_LANES_PATH = QUALITY_DIR / "v014_s13_p2_post_remediation_source_lane_status_public_safe.json"
METADATA_ISSUES_PATH = QUALITY_DIR / "v014_s13_p2_post_remediation_issue_definitions_public_safe.json"
METADATA_MATRIX_PATH = QUALITY_DIR / "v014_s13_p2_post_remediation_collection_receivable_aging_acceptance_matrix.json"
METADATA_GO_NO_GO_PATH = QUALITY_DIR / "v014_s13_p2_post_remediation_collection_receivable_aging_go_no_go.json"

PRIVATE_DIR = Path("KMFA/.codex_private_runtime/v014_s13_p2_post_remediation_collection_receivable_aging")
PRIVATE_RAW_BEFORE_PATH = PRIVATE_DIR / "raw_immutability_before.json"
PRIVATE_RAW_AFTER_PATH = PRIVATE_DIR / "raw_immutability_after.json"
PRIVATE_SOURCE_PROBE_PATH = PRIVATE_DIR / "private_raw_source_probe.json"
PRIVATE_WORKBOOK_PROFILE_PATH = PRIVATE_DIR / "private_workbook_structure_profile.json"
PRIVATE_ALIGNMENT_PATH = PRIVATE_DIR / "private_row_binding_diagnostic.json"
PRIVATE_DIFFERENCE_REPORT_PATH = PRIVATE_DIR / "s13_p2_private_difference_report_zh.md"
PRIVATE_BROWSER_PATH = PRIVATE_DIR / "browser_verification.json"
PRIVATE_BASELINE_AUDIT_PATH = PRIVATE_DIR / "human_flow_baseline_audit.csv"
PRIVATE_CURRENT_AUDIT_PATH = PRIVATE_DIR / "current_collection_receivable_aging_audit.csv"
PRIVATE_SCREENSHOT_DIR = PRIVATE_DIR / "screenshots"

V14_ROADMAP_PATH = Path("KMFA/taskpack/v1_4/roadmap/02_KMFA_Codex_Development_Roadmap_18_Stages_v1_4.md")
V14_TASKPACK_PATH = Path("KMFA/taskpack/v1_4/taskpack/01_KMFA_Codex_TaskPack_v1_4_public_safe.md")
V14_HTML_ROOT = Path("KMFA/taskpack/v1_4/html_uiux")
WPS_REGISTRY_PATH = Path("KMFA/metadata/imports/v014_s07_p2_wps_export_source_registry.json")
FINANCE_REGISTRY_PATH = Path("KMFA/metadata/imports/v014_s07_p1_finance_support_source_registry.json")
HISTORICAL_MANIFEST_PATH = Path(
    "KMFA/stage_artifacts/V014_S13_P2_COLLECTION_RECEIVABLE_AGING/machine/collection_receivable_aging_manifest.json"
)

DEVELOPMENT_EVENTS_PATH = Path("KMFA/docs/governance/development_events.jsonl")
STAGE_STATUS_PATH = Path("KMFA/metadata/stage_status.jsonl")
TASK_STATUS_PATH = Path("KMFA/metadata/traceability/v1_4_stage_phase_task_status.jsonl")

WEEKLY_HREF = (
    "../../../V014_S13_P1_POST_REMEDIATION_FINANCIAL_OPERATING_REPORT/exports/html/"
    "financial_operating_weekly_draft.html"
)
MONTHLY_HREF = (
    "../../../V014_S13_P1_POST_REMEDIATION_FINANCIAL_OPERATING_REPORT/exports/html/"
    "financial_operating_monthly_draft.html"
)

LANE_SPECS = (
    {
        "lane_id": "collection_table",
        "visible_lane_name": "回款表",
        "registry": "wps",
        "source_ref": "SRC-WPS-COLLECTION-001",
        "private_parse_policy": "native_wps_conversion_required",
    },
    {
        "lane_id": "receivable_aging",
        "visible_lane_name": "应收账龄",
        "registry": "wps",
        "source_ref": "SRC-WPS-AGING-001",
        "private_parse_policy": "native_wps_conversion_required",
    },
    {
        "lane_id": "customer_aging",
        "visible_lane_name": "客户账龄",
        "registry": "finance",
        "source_ref": "SRC-FIN-AGING-001",
        "private_parse_policy": "openable_archive_workbook_required",
    },
    {
        "lane_id": "journal",
        "visible_lane_name": "日记账",
        "registry": "finance",
        "source_ref": "SRC-FIN-JOURNAL-001",
        "private_parse_policy": "openable_archive_workbook_required",
    },
    {
        "lane_id": "invoice_plan",
        "visible_lane_name": "开票计划",
        "registry": "finance",
        "source_ref": "SRC-FIN-INVOICE-001",
        "private_parse_policy": "openable_archive_workbook_required",
    },
)

ISSUE_SPECS = (
    {
        "issue_type": "invoiced_not_collected",
        "visible_name": "已开票未回款",
        "required_lane_ids": ["invoice_plan", "collection_table", "journal"],
        "review_sequence": 1,
        "role_definition": "财务复核角色",
    },
    {
        "issue_type": "completed_not_settled",
        "visible_name": "完工未结算",
        "required_lane_ids": ["journal", "invoice_plan"],
        "review_sequence": 2,
        "role_definition": "项目与财务联合复核角色",
        "supplemental_structure_requirement": "production_project_status",
    },
    {
        "issue_type": "settled_not_invoiced",
        "visible_name": "结算未开票",
        "required_lane_ids": ["invoice_plan", "journal"],
        "review_sequence": 3,
        "role_definition": "开票复核角色",
    },
    {
        "issue_type": "overdue_receivable",
        "visible_name": "超期应收",
        "required_lane_ids": ["receivable_aging", "customer_aging", "collection_table"],
        "review_sequence": 4,
        "role_definition": "应收复核角色",
    },
)

PRIVATE_LANE_KEYWORDS = {
    "collection_table": ("回款", "收款", "到账", "银行"),
    "receivable_aging": ("应收", "账龄", "逾期", "超期"),
    "customer_aging": ("客户", "往来", "单位", "应收", "账龄"),
    "journal": ("日记账", "明细账", "凭证", "科目", "借方", "贷方"),
    "invoice_plan": ("开票", "发票", "税票"),
}
PRIVATE_KEY_CLASS_KEYWORDS = {
    "project_key": ("项目", "合同"),
    "customer_key": ("客户", "单位", "往来"),
    "invoice_key": ("发票", "开票"),
    "date_key": ("日期", "时间", "年月", "期间"),
    "amount_key": ("金额", "余额", "借方", "贷方", "应收", "回款"),
    "status_key": ("状态", "完工", "结算"),
}
PRIVATE_CANDIDATE_SIGNAL_TOKENS = (
    "无回款匹配",
    "未回款",
    "超期",
    "逾期",
    "完工未结算",
    "结算未开票",
)


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise RuntimeError(f"{path} must contain an object")
    return value


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8")


def _git_output(args: list[str]) -> str:
    result = subprocess.run(
        ["git", "-c", "core.quotePath=false", *args],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"git {' '.join(args)} failed: {result.stderr.strip()}")
    return result.stdout.strip()


def _upsert_jsonl(path: Path, row: dict[str, Any]) -> None:
    preserved: list[str] = []
    if path.is_file():
        for line in path.read_text(encoding="utf-8").splitlines():
            if line.strip() and json.loads(line).get("phase_id") != PHASE_ID:
                preserved.append(line)
    preserved.append(json.dumps(row, ensure_ascii=False, separators=(",", ":")))
    _write_text(path, "\n".join(preserved))


def _load_dependency() -> dict[str, Any]:
    dependency = validate_v014_s13_p1_post_remediation_financial_operating_report(
        require_private_evidence=True,
        require_browser_evidence=True,
        require_final_evidence=True,
    )
    summary = dependency.get("summary", {})
    expected = {
        "stage_id": "S13",
        "roadmap_phase_id": "S13-P1",
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "decision": "NO_GO",
        "open_final_difference_accepted_count": 3,
        "nonzero_delta_reconciliation_count": 9,
        "zero_delta_reconciliation_count": 2,
        "incomplete_reconciliation_count": 1,
        "raw_source_file_count": 5,
        "s13_p2_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
    }
    for key, value in expected.items():
        if summary.get(key) != value:
            raise RuntimeError(f"current S13-P1 dependency {key} mismatch")
    return dependency


def _load_contract() -> dict[str, Any]:
    roadmap = V14_ROADMAP_PATH.read_text(encoding="utf-8")
    taskpack = V14_TASKPACK_PATH.read_text(encoding="utf-8")
    for token in (
        "回款应收账龄",
        "回款表、应收账龄、客户账龄、日记账、开票计划",
        "已开票未回款、完工未结算、结算未开票、超期应收",
        "输出回款优先级和责任事项",
    ):
        if token not in roadmap:
            raise RuntimeError(f"v1.4 roadmap missing S13-P2 marker: {token}")
    for token in ("回款/应收/账龄线", "催收优先级", "银行回款 ↔ 应收账龄"):
        if token not in taskpack:
            raise RuntimeError(f"v1.4 task pack missing S13-P2 marker: {token}")
    return {
        "roadmap_read": True,
        "taskpack_read": True,
        "roadmap_s13_p2_contract_locked": True,
        "taskpack_receivable_line_locked": True,
        "source_refs": {
            "roadmap": V14_ROADMAP_PATH.as_posix(),
            "taskpack": V14_TASKPACK_PATH.as_posix(),
        },
    }


def _probe_ooxml(binary: bytes) -> tuple[bool, int]:
    try:
        with zipfile.ZipFile(io.BytesIO(binary)) as workbook:
            names = workbook.namelist()
    except (OSError, zipfile.BadZipFile):
        return False, 0
    worksheet_count = sum(
        name.startswith("xl/worksheets/") and name.endswith(".xml") for name in names
    )
    return "xl/workbook.xml" in names, worksheet_count


def _probe_raw_sources(raw_root: Path) -> dict[str, Any]:
    files = sorted(path for path in raw_root.iterdir() if path.is_file())
    details: list[dict[str, Any]] = []
    wps_private_container_count = 0
    archive_workbook_member_count = 0
    openable_archive_workbook_member_count = 0
    archive_pdf_member_count = 0
    openable_archive_pdf_member_count = 0
    for path in files:
        suffix = path.suffix.lower()
        detail: dict[str, Any] = {
            "private_filename": path.name,
            "private_path": str(path),
            "suffix": suffix,
            "size_bytes": path.stat().st_size,
            "read_only_probe": True,
        }
        with path.open("rb") as handle:
            magic = handle.read(8)
        if suffix in {".xlsx", ".xlsm"}:
            is_cdf = magic == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
            detail["container_kind"] = "wps_private_compound_document" if is_cdf else "ooxml_candidate"
            detail["native_wps_conversion_required"] = is_cdf
            if is_cdf:
                wps_private_container_count += 1
        elif suffix == ".zip":
            members: list[dict[str, Any]] = []
            with zipfile.ZipFile(path) as archive:
                for info in archive.infolist():
                    if info.is_dir() or info.filename.startswith("__MACOSX/"):
                        continue
                    member_suffix = Path(info.filename).suffix.lower()
                    if member_suffix in {".xlsx", ".xlsm"}:
                        archive_workbook_member_count += 1
                        binary = archive.read(info)
                        openable, worksheet_count = _probe_ooxml(binary)
                        openable_archive_workbook_member_count += int(openable)
                        members.append(
                            {
                                "private_member_name": info.filename,
                                "member_kind": "workbook",
                                "openable": openable,
                                "worksheet_xml_count": worksheet_count,
                                "size_bytes": info.file_size,
                            }
                        )
                    elif member_suffix == ".pdf":
                        archive_pdf_member_count += 1
                        with archive.open(info) as handle:
                            openable = handle.read(5) == b"%PDF-"
                        openable_archive_pdf_member_count += int(openable)
                        members.append(
                            {
                                "private_member_name": info.filename,
                                "member_kind": "pdf",
                                "openable": openable,
                                "size_bytes": info.file_size,
                            }
                        )
            detail["container_kind"] = "archive"
            detail["recognized_private_members"] = members
        else:
            detail["container_kind"] = "other"
        details.append(detail)

    if len(files) != 5:
        raise RuntimeError("S13-P2 requires the locked five-file raw snapshot")
    if wps_private_container_count != 2:
        raise RuntimeError("S13-P2 expected two WPS private compound-document sources")
    if openable_archive_workbook_member_count < 1:
        raise RuntimeError("S13-P2 requires at least one read-only openable archive workbook")

    differences = [
        {
            "difference_id": "S13P2-DIFF-PRIVATE-WPS-COLLECTION",
            "category": "native_wps_conversion_unavailable",
            "affected_lane_id": "collection_table",
            "status": "open_private_difference",
        },
        {
            "difference_id": "S13P2-DIFF-PRIVATE-WPS-AGING",
            "category": "native_wps_conversion_unavailable",
            "affected_lane_id": "receivable_aging",
            "status": "open_private_difference",
        },
        {
            "difference_id": "S13P2-DIFF-SHARED-ROW-KEY",
            "category": "shared_row_key_contract_unproven",
            "affected_lane_id": "all_required_lanes",
            "status": "open_private_difference",
        },
        {
            "difference_id": "S13P2-DIFF-PERIOD-ALIGNMENT",
            "category": "cross_source_period_alignment_unproven",
            "affected_lane_id": "all_required_lanes",
            "status": "open_private_difference",
        },
    ]
    return {
        "schema_version": "kmfa.private.v014.s13_p2.raw_source_probe.v1",
        "classification": "PRIVATE_RUNTIME_ONLY",
        "raw_root": str(raw_root),
        "raw_source_file_count": len(files),
        "wps_private_container_count": wps_private_container_count,
        "archive_workbook_member_count": archive_workbook_member_count,
        "openable_archive_workbook_member_count": openable_archive_workbook_member_count,
        "archive_pdf_member_count": archive_pdf_member_count,
        "openable_archive_pdf_member_count": openable_archive_pdf_member_count,
        "private_source_details": details,
        "private_differences": differences,
        "mutation_performed": False,
        "public_commit_allowed": False,
    }


def _private_workbook_profile_worker() -> dict[str, Any]:
    from openpyxl import load_workbook

    prior_raw = _read_json(s13_p1.PRIVATE_RAW_AFTER_PATH)
    raw_root = Path(prior_raw["raw_root"])
    profiles: list[dict[str, Any]] = []
    for archive_path in sorted(raw_root.glob("*.zip")):
        with zipfile.ZipFile(archive_path) as archive:
            for info in archive.infolist():
                if (
                    info.is_dir()
                    or info.filename.startswith("__MACOSX/")
                    or Path(info.filename).suffix.lower() not in {".xlsx", ".xlsm"}
                ):
                    continue
                binary = archive.read(info)
                try:
                    workbook = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
                except Exception as exc:
                    profiles.append(
                        {
                            "private_archive_name": archive_path.name,
                            "private_member_name": info.filename,
                            "openable": False,
                            "error_class": type(exc).__name__,
                        }
                    )
                    continue
                sheet_names = list(workbook.sheetnames)
                title_context = " ".join(sheet_names)
                selected: list[str] = []
                for index, title in enumerate(sheet_names):
                    if index < 3 or any(
                        token in title
                        for tokens in PRIVATE_LANE_KEYWORDS.values()
                        for token in tokens
                    ):
                        selected.append(title)
                    if len(selected) >= 60:
                        break
                observed_strings: list[str] = []
                for title in selected:
                    sheet = workbook[title]
                    for row in sheet.iter_rows(
                        min_row=1,
                        max_row=20,
                        max_col=40,
                        values_only=True,
                    ):
                        for value in row:
                            if isinstance(value, str):
                                cleaned = " ".join(value.split())
                                if cleaned and len(cleaned) <= 120:
                                    observed_strings.append(cleaned)
                workbook.close()
                context = " ".join([info.filename, title_context, *observed_strings])
                lane_ids = [
                    lane
                    for lane, tokens in PRIVATE_LANE_KEYWORDS.items()
                    if any(token in context for token in tokens)
                ]
                key_classes = [
                    key
                    for key, tokens in PRIVATE_KEY_CLASS_KEYWORDS.items()
                    if any(token in context for token in tokens)
                ]
                profiles.append(
                    {
                        "private_archive_name": archive_path.name,
                        "private_member_name": info.filename,
                        "openable": True,
                        "worksheet_count": len(sheet_names),
                        "private_sheet_names": sheet_names,
                        "sampled_sheet_count": len(selected),
                        "private_sampled_strings": observed_strings,
                        "candidate_lane_ids": lane_ids,
                        "candidate_key_classes": key_classes,
                        "private_candidate_signal_count": sum(
                            observed_strings.count(token)
                            for token in PRIVATE_CANDIDATE_SIGNAL_TOKENS
                        ),
                    }
                )

    lane_document_counts = {
        lane: sum(
            row.get("openable") is True
            and lane in row.get("candidate_lane_ids", [])
            for row in profiles
        )
        for lane in PRIVATE_LANE_KEYWORDS
    }
    lane_key_classes = {
        lane: sorted(
            {
                key
                for row in profiles
                if lane in row.get("candidate_lane_ids", [])
                for key in row.get("candidate_key_classes", [])
            }
        )
        for lane in PRIVATE_LANE_KEYWORDS
    }
    common_keys = (
        sorted(set.intersection(*(set(values) for values in lane_key_classes.values())))
        if all(lane_key_classes.values())
        else []
    )
    normalize = s13_p1.s12_review.p1.s11_project._normalize_raw
    raw_snapshot_hash = hashlib.sha256(
        json.dumps(
            normalize(prior_raw),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()
    payload = {
        "schema_version": "kmfa.private.v014.s13_p2.workbook_structure_profile.v1",
        "classification": "PRIVATE_RUNTIME_ONLY",
        "raw_snapshot_hash": raw_snapshot_hash,
        "raw_source_file_count": prior_raw["file_count"],
        "workbook_candidate_count": len(profiles),
        "openable_workbook_count": sum(row.get("openable") is True for row in profiles),
        "lane_document_counts": lane_document_counts,
        "lane_key_classes": lane_key_classes,
        "common_key_classes_across_all_lanes": common_keys,
        "private_candidate_signal_count": sum(
            row.get("private_candidate_signal_count", 0) for row in profiles
        ),
        "candidate_signal_is_authoritative_business_item": False,
        "row_level_binding_proven": False,
        "reason": (
            "keyword header and candidate-signal presence does not prove shared row identity "
            "period or business value equality"
        ),
        "private_workbook_profiles": profiles,
        "mutation_performed": False,
        "public_commit_allowed": False,
    }
    _write_json(PRIVATE_WORKBOOK_PROFILE_PATH, payload)
    return payload


def _run_private_workbook_profile() -> dict[str, Any]:
    helper = s13_p1.s12_review.p1.s11_home
    result = subprocess.run(
        [
            str(helper._audit_python()),
            "-m",
            "KMFA.tools.v014_s13_p2_post_remediation_collection_receivable_aging",
            "--private-profile-only",
        ],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=os.environ.copy(),
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"private workbook profile failed: {result.stdout}\n{result.stderr}")
    profile = _read_json(PRIVATE_WORKBOOK_PROFILE_PATH)
    if profile.get("row_level_binding_proven") is not False:
        raise RuntimeError("private workbook profile cannot claim row-level binding")
    return profile


def _build_source_lanes(probe: dict[str, Any]) -> list[dict[str, Any]]:
    wps_sources = {row["source_ref"]: row for row in _read_json(WPS_REGISTRY_PATH)["sources"]}
    finance_sources = {
        row["source_ref"]: row for row in _read_json(FINANCE_REGISTRY_PATH)["sources"]
    }
    archive_parseable = probe["openable_archive_workbook_member_count"] > 0
    lanes: list[dict[str, Any]] = []
    for spec in LANE_SPECS:
        registry = wps_sources if spec["registry"] == "wps" else finance_sources
        source = registry.get(spec["source_ref"])
        if not source:
            raise RuntimeError(f"missing current structure source {spec['source_ref']}")
        structure_connected = bool(source.get("parse_status"))
        private_parseable = (
            archive_parseable
            if spec["private_parse_policy"] == "openable_archive_workbook_required"
            else False
        )
        lanes.append(
            {
                "lane_id": spec["lane_id"],
                "visible_lane_name": spec["visible_lane_name"],
                "public_structure_source_ref": spec["source_ref"],
                "structure_connected": structure_connected,
                "private_raw_parseable": private_parseable,
                "row_level_binding_proven": False,
                "data_status": (
                    "structure_and_private_container_parseable_row_binding_unproven"
                    if private_parseable
                    else "structure_connected_native_wps_conversion_required"
                ),
                "contains_source_identity": False,
                "contains_field_plaintext": False,
                "contains_business_amounts": False,
                "business_priority_allowed": False,
                "responsibility_assignment_allowed": False,
                "business_decision_basis_allowed": False,
            }
        )
    if not all(row["structure_connected"] for row in lanes):
        raise RuntimeError("S13-P2 required public-safe structures are incomplete")
    return lanes


def _build_issue_definitions() -> list[dict[str, Any]]:
    return [
        {
            "issue_id": f"S13P2-ISSUE-{index:02d}",
            "issue_type": spec["issue_type"],
            "visible_name": spec["visible_name"],
            "required_lane_ids": spec["required_lane_ids"],
            "supplemental_structure_requirement": spec.get(
                "supplemental_structure_requirement", "none"
            ),
            "identification_status": "definition_locked_row_level_evidence_unproven",
            "identified_item_count": 0,
            "review_sequence": spec["review_sequence"],
            "priority_status": "method_only_not_business_priority",
            "responsibility_role_definition": spec["role_definition"],
            "responsibility_status": "role_definition_only_unassigned",
            "business_priority_allowed": False,
            "responsibility_assignment_allowed": False,
            "collection_action_allowed": False,
            "legal_collection_decision_allowed": False,
            "contains_source_identity": False,
            "contains_field_plaintext": False,
            "contains_business_amounts": False,
        }
        for index, spec in enumerate(ISSUE_SPECS, 1)
    ]


def _render_html(lanes: list[dict[str, Any]], issues: list[dict[str, Any]]) -> str:
    lane_rows = "".join(
        f"""<tr>
          <td>{html.escape(row['visible_lane_name'])}</td>
          <td><span class="state ok">结构已接入</span></td>
          <td><span class="state {'ok' if row['private_raw_parseable'] else 'blocked'}">{'可解析容器' if row['private_raw_parseable'] else '需原生 WPS 转换'}</span></td>
          <td><span class="state blocked">行级未证明</span></td>
        </tr>"""
        for row in lanes
    )
    buttons = "".join(
        f'<button type="button" data-issue-button="{row["issue_type"]}" class="issue-button{" active" if index == 0 else ""}">{html.escape(row["visible_name"])}</button>'
        for index, row in enumerate(issues)
    )
    panels = "".join(
        f"""<section data-issue-panel="{row['issue_type']}" class="issue-panel{' active' if index == 0 else ''}">
          <div class="panel-heading"><div><p class="kicker">复核序列 {row['review_sequence']}</p><h2>{html.escape(row['visible_name'])}</h2></div><span class="state blocked">0 项已证明</span></div>
          <dl><div><dt>识别状态</dt><dd>方法定义已锁定，行级证据未证明</dd></div><div><dt>优先级</dt><dd>仅方法顺序，不是业务催收优先级</dd></div><div><dt>责任事项</dt><dd>{html.escape(row['responsibility_role_definition'])}，未指派个人</dd></div></dl>
          <p class="limit">不得据此联系客户、催收、作法律判断、开票、付款或执行银行操作。</p>
        </section>"""
        for index, row in enumerate(issues)
    )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>KMFA 回款与应收账龄工作台</title>
  <style>
    :root{{--ink:#17211d;--muted:#5e6863;--line:#d9dfdc;--surface:#ffffff;--soft:#f4f7f5;--green:#176b4d;--green-soft:#e7f3ed;--red:#a53b36;--red-soft:#faecea;--blue:#285f88;--focus:#0b6e99}}
    *{{box-sizing:border-box;letter-spacing:0}}
    body{{margin:0;background:#eef2f0;color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC",sans-serif;font-size:14px}}
    a{{color:var(--blue)}} button{{font:inherit}}
    .appbar{{background:#18231f;color:#fff;border-bottom:3px solid #59a37d}}
    .appbar-inner{{max-width:1180px;margin:auto;padding:14px 22px;display:flex;align-items:center;justify-content:space-between;gap:16px}}
    .brand{{font-size:19px;font-weight:750}} .phase{{color:#c8d5cf;font-size:12px}}
    main{{max-width:1180px;margin:auto;padding:22px}}
    .headline{{display:flex;align-items:flex-start;justify-content:space-between;gap:20px;margin-bottom:18px}}
    h1{{font-size:28px;line-height:1.2;margin:0 0 8px}} h2{{font-size:18px;margin:0}}
    p{{line-height:1.6}} .subtitle{{margin:0;color:var(--muted)}}
    .status-strip{{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}}
    .badge,.state{{display:inline-flex;align-items:center;min-height:26px;padding:3px 8px;border:1px solid var(--line);border-radius:4px;font-weight:700;font-size:12px;white-space:nowrap}}
    .badge.danger,.state.blocked{{color:var(--red);background:var(--red-soft);border-color:#e7bbb7}}
    .badge.grade,.state.ok{{color:var(--green);background:var(--green-soft);border-color:#b9d9c8}}
    .metrics{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));border:1px solid var(--line);background:var(--surface);margin-bottom:18px}}
    .metric{{padding:15px;border-right:1px solid var(--line)}} .metric:last-child{{border-right:0}}
    .metric strong{{display:block;font-size:24px}} .metric span{{color:var(--muted);font-size:12px}}
    .band{{background:var(--surface);border:1px solid var(--line);margin-bottom:18px}}
    .band-header{{padding:14px 16px;border-bottom:1px solid var(--line);display:flex;justify-content:space-between;align-items:center;gap:12px}}
    .table-wrap{{overflow-x:auto}} table{{width:100%;border-collapse:collapse;min-width:680px}}
    th,td{{text-align:left;padding:11px 14px;border-bottom:1px solid #e8ecea}} th{{font-size:12px;color:var(--muted);background:var(--soft)}}
    .issue-layout{{display:grid;grid-template-columns:230px minmax(0,1fr)}}
    .issue-nav{{border-right:1px solid var(--line);padding:10px;display:grid;gap:6px;align-content:start;background:var(--soft)}}
    .issue-button{{border:1px solid transparent;background:transparent;text-align:left;padding:10px;border-radius:4px;color:var(--ink);cursor:pointer}}
    .issue-button:hover,.issue-button:focus-visible{{border-color:#9cb9aa;outline:2px solid transparent}}
    .issue-button.active{{background:#fff;border-color:#92b7a5;color:var(--green);font-weight:750}}
    .issue-panel{{display:none;padding:18px}} .issue-panel.active{{display:block}}
    .panel-heading{{display:flex;justify-content:space-between;align-items:flex-start;gap:12px}}
    .kicker{{font-size:12px;color:var(--muted);margin:0 0 4px}}
    dl{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;margin:18px 0}}
    dl div{{border-left:3px solid #8cab9c;padding:4px 10px}} dt{{font-size:12px;color:var(--muted)}} dd{{margin:5px 0 0;line-height:1.5}}
    .limit{{border:1px solid #e4beb9;background:var(--red-soft);padding:10px 12px;margin:0;color:#74302c}}
    .footer-band{{display:flex;justify-content:space-between;gap:16px;align-items:center;padding:14px 16px;background:#f8faf9}}
    .links{{display:flex;gap:14px;flex-wrap:wrap}} #interaction-status{{color:var(--muted);margin:0}}
    @media(max-width:760px){{main{{padding:14px}}.headline{{display:block}}.status-strip{{justify-content:flex-start;margin-top:12px}}.metrics{{grid-template-columns:repeat(2,minmax(0,1fr))}}.metric:nth-child(2){{border-right:0}}.metric:nth-child(-n+2){{border-bottom:1px solid var(--line)}}.issue-layout{{display:block}}.issue-nav{{border-right:0;border-bottom:1px solid var(--line);grid-template-columns:repeat(2,minmax(0,1fr))}}dl{{grid-template-columns:1fr}}.footer-band{{display:block}}.links{{margin-top:10px}}h1{{font-size:24px}}}}
  </style>
</head>
<body data-ui-ready="false" data-active-issue="{issues[0]['issue_type']}">
  <header class="appbar"><div class="appbar-inner"><div class="brand">KMFA</div><div class="phase">S13-P2 · 回款与应收账龄</div></div></header>
  <main>
    <section class="headline"><div><h1>回款与应收账龄工作台</h1><p class="subtitle">结构接入与私有只读探针已完成；业务明细仍受行级证据门禁限制。</p></div><div class="status-strip"><span class="badge grade">Q4 / D</span><span class="badge danger">NO_GO</span><span class="badge">内部复核</span></div></section>
    <section class="metrics" aria-label="状态摘要"><div class="metric"><strong>5 / 5</strong><span>结构主题已接入</span></div><div class="metric"><strong>3 / 5</strong><span>私有容器可解析</span></div><div class="metric"><strong>0 / 5</strong><span>行级绑定已证明</span></div><div class="metric"><strong>0</strong><span>可执行业务项</span></div></section>
    <section class="band"><div class="band-header"><h2>来源主题状态</h2><span class="state blocked">不含业务金额</span></div><div class="table-wrap"><table><thead><tr><th>主题</th><th>结构</th><th>私有解析</th><th>行级绑定</th></tr></thead><tbody>{lane_rows}</tbody></table></div></section>
    <section class="band"><div class="band-header"><h2>四类问题方法定义</h2><span class="state blocked">0 项已证明</span></div><div class="issue-layout"><nav class="issue-nav" aria-label="问题类型">{buttons}</nav><div>{panels}</div></div><div class="footer-band"><p id="interaction-status" aria-live="polite">已显示“{html.escape(issues[0]['visible_name'])}”方法定义；不是业务催收优先级。</p><div class="links"><a data-report-link="weekly" href="{WEEKLY_HREF}">经营周报初稿</a><a data-report-link="monthly" href="{MONTHLY_HREF}">经营月报初稿</a></div></div></section>
  </main>
  <script>
    const labels={{{','.join(json.dumps(row['issue_type'])+':'+json.dumps(row['visible_name'],ensure_ascii=False) for row in issues)}}};
    const buttons=[...document.querySelectorAll('[data-issue-button]')];
    const panels=[...document.querySelectorAll('[data-issue-panel]')];
    const status=document.getElementById('interaction-status');
    function activate(issue){{buttons.forEach(button=>button.classList.toggle('active',button.dataset.issueButton===issue));panels.forEach(panel=>panel.classList.toggle('active',panel.dataset.issuePanel===issue));document.body.dataset.activeIssue=issue;document.body.dataset.lastAction=`issue:${{issue}}:${{Date.now()}}`;status.textContent=`已切换至“${{labels[issue]}}”方法定义；仍无已证明业务项。`;}}
    buttons.forEach(button=>button.addEventListener('click',()=>activate(button.dataset.issueButton)));
    document.body.dataset.uiReady='true';
  </script>
</body>
</html>"""


class _QuietHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        return


def _workbench_url(base: str) -> str:
    return f"{base}/{HTML_PATH.as_posix().removeprefix('KMFA/stage_artifacts/')}"


def _browser_worker() -> dict[str, Any]:
    from playwright.sync_api import sync_playwright

    PRIVATE_SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    stage_root = Path("KMFA/stage_artifacts").resolve()
    handler = functools.partial(_QuietHandler, directory=str(stage_root))
    server = socketserver.TCPServer(("127.0.0.1", 0), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    base = f"http://127.0.0.1:{server.server_address[1]}"
    viewport_checks: list[dict[str, Any]] = []
    issue_checks: list[dict[str, Any]] = []
    http_checks: list[dict[str, Any]] = []
    navigation_checks: list[dict[str, Any]] = []
    helper = s13_p1.s12_review.p1.s11_home
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=True,
                executable_path=helper._chromium_path(),
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )
            for mode, viewport in (
                ("desktop", {"width": 1440, "height": 1000}),
                ("mobile", {"width": 390, "height": 844}),
            ):
                page = browser.new_page(viewport=viewport)
                console_errors: list[str] = []
                page.on(
                    "console",
                    lambda msg: console_errors.append(msg.text)
                    if msg.type == "error"
                    and helper._is_actionable_console_error(
                        f"{msg.text} {msg.location.get('url', '')}"
                    )
                    else None,
                )
                page.on("pageerror", lambda exc: console_errors.append(str(exc)))
                page.goto(_workbench_url(base), wait_until="networkidle")
                page.wait_for_function("document.body.dataset.uiReady === 'true'")
                body = page.locator("body").inner_text()
                if mode == "desktop":
                    for issue in ISSUE_SPECS:
                        issue_type = issue["issue_type"]
                        page.locator(f'[data-issue-button="{issue_type}"]').click()
                        issue_checks.append(
                            {
                                "issue_type": issue_type,
                                "passed": (
                                    page.locator("body").get_attribute("data-active-issue")
                                    == issue_type
                                    and page.locator(f'[data-issue-panel="{issue_type}"]').is_visible()
                                    and "方法定义" in page.locator("#interaction-status").inner_text()
                                ),
                            }
                        )
                dimensions = page.evaluate(
                    "({scrollWidth:document.documentElement.scrollWidth,innerWidth:window.innerWidth})"
                )
                viewport_checks.append(
                    {
                        "mode": mode,
                        "viewport": viewport,
                        "workbench_visible": "回款与应收账龄工作台" in body,
                        "d_no_go_visible": "Q4 / D" in body and "NO_GO" in body,
                        "zero_business_items_visible": "0" in body and "可执行业务项" in body,
                        "console_error_count": len(console_errors),
                        "no_horizontal_overflow": dimensions["scrollWidth"] <= dimensions["innerWidth"] + 1,
                    }
                )
                page.screenshot(
                    path=str(PRIVATE_SCREENSHOT_DIR / f"workbench_{mode}.png"),
                    full_page=True,
                )
                page.close()

            request = playwright.request.new_context()
            for link_id, marker in (("weekly", "经营周报初稿"), ("monthly", "经营月报初稿")):
                page = browser.new_page(viewport={"width": 1280, "height": 900})
                page.goto(_workbench_url(base), wait_until="networkidle")
                href = page.locator(f'a[data-report-link="{link_id}"]').get_attribute("href") or ""
                target_url = urljoin(page.url, href)
                response = request.get(target_url)
                http_checks.append(
                    {
                        "target": link_id,
                        "status": response.status,
                        "passed": response.ok and marker in response.text(),
                    }
                )
                page.locator(f'a[data-report-link="{link_id}"]').click()
                page.wait_for_load_state("networkidle")
                navigation_checks.append(
                    {"target": link_id, "passed": marker in page.locator("body").inner_text()}
                )
                page.close()
            request.dispose()
            browser.close()
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)

    passed = (
        len(viewport_checks) == 2
        and all(
            row["workbench_visible"]
            and row["d_no_go_visible"]
            and row["zero_business_items_visible"]
            and row["console_error_count"] == 0
            and row["no_horizontal_overflow"]
            for row in viewport_checks
        )
        and len(issue_checks) == 4
        and all(row["passed"] for row in issue_checks)
        and len(http_checks) == 2
        and all(row["passed"] for row in http_checks)
        and len(navigation_checks) == 2
        and all(row["passed"] for row in navigation_checks)
    )
    result = {
        "status": "PASS" if passed else "FAIL",
        "viewport_checks": viewport_checks,
        "issue_interaction_checks": issue_checks,
        "dependency_link_http_checks": http_checks,
        "dependency_navigation_checks": navigation_checks,
    }
    _write_json(PRIVATE_BROWSER_PATH, result)
    if not passed:
        raise RuntimeError("S13-P2 desktop/mobile browser evidence failed")
    return result


def _run_browser_review() -> dict[str, Any]:
    helper = s13_p1.s12_review.p1.s11_home
    baseline = helper._run_html_audit(V14_HTML_ROOT, PRIVATE_BASELINE_AUDIT_PATH)
    current = helper._run_html_audit(HTML_DIR, PRIVATE_CURRENT_AUDIT_PATH)
    env = os.environ.copy()
    env["KMFA_CHROMIUM"] = helper._chromium_path()
    result = subprocess.run(
        [
            str(helper._audit_python()),
            "-m",
            "KMFA.tools.v014_s13_p2_post_remediation_collection_receivable_aging",
            "--browser-evidence-only",
        ],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"browser evidence failed: {result.stdout}\n{result.stderr}")
    browser = _read_json(PRIVATE_BROWSER_PATH)
    return {
        "status": browser["status"],
        "baseline_file_count": baseline["file_count"],
        "baseline_control_row_count": baseline["control_row_count"],
        "baseline_pass_count": baseline["pass_count"],
        "baseline_warn_count": baseline["warn_count"],
        "baseline_fail_count": baseline["fail_count"],
        "current_page_count": current["file_count"],
        "current_control_row_count": current["control_row_count"],
        "current_pass_count": current["pass_count"],
        "current_warn_count": current["warn_count"],
        "current_fail_count": current["fail_count"],
        "viewport_check_count": len(browser["viewport_checks"]),
        "issue_interaction_check_count": len(browser["issue_interaction_checks"]),
        "dependency_link_http_check_count": len(browser["dependency_link_http_checks"]),
        "dependency_navigation_check_count": len(browser["dependency_navigation_checks"]),
        "console_error_count": sum(row["console_error_count"] for row in browser["viewport_checks"]),
        "horizontal_overflow_count": sum(not row["no_horizontal_overflow"] for row in browser["viewport_checks"]),
    }


def _quality_gate() -> dict[str, bool]:
    return {
        "method_definition_allowed": True,
        "public_safe_status_workbench_allowed": True,
        "private_readonly_probe_allowed": True,
        "business_priority_allowed": False,
        "responsibility_assignment_allowed": False,
        "collection_action_allowed": False,
        "legal_collection_decision_allowed": False,
        "invoice_issuance_allowed": False,
        "payment_or_bank_operation_allowed": False,
        "formal_report_allowed": False,
        "business_decision_basis_allowed": False,
        "s13_p3_allowed": False,
        "stage13_review_allowed": False,
        "github_upload_allowed": False,
        "app_reinstall_allowed": False,
        "business_execution_allowed": False,
    }


def _phase_boundaries() -> dict[str, bool]:
    return {
        "s13_p1_performed": True,
        "s13_p2_performed": True,
        "s13_p3_performed": False,
        "stage13_review_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "formal_report_release_performed": False,
        "business_execution_performed": False,
        "persistent_business_write_performed": False,
    }


def _public_repo_safety() -> dict[str, bool]:
    return {
        "raw_business_data_committed": False,
        "raw_file_names_committed": False,
        "raw_file_hashes_committed": False,
        "field_plaintext_committed": False,
        "source_header_plaintext_committed": False,
        "business_amount_values_committed": False,
        "project_or_customer_plaintext_committed": False,
        "zip_committed": False,
        "excel_committed": False,
        "pdf_committed": False,
        "private_csv_committed": False,
        "credential_or_secret_committed": False,
        "bank_statement_committed": False,
        "contract_committed": False,
        "salary_material_committed": False,
        "tax_filing_material_committed": False,
    }


def _matrix(summary: dict[str, Any]) -> dict[str, Any]:
    checks = {
        "dependency_validated": summary["s13_p1_dependency_validated"],
        "grade_q4": summary["current_data_quality_grade"] == "Q4",
        "report_grade_d": summary["current_report_grade"] == "D",
        "decision_no_go": summary["decision"] == "NO_GO",
        "five_lanes": summary["source_lane_count"] == 5,
        "five_structures": summary["structure_connected_lane_count"] == 5,
        "three_private_parseable": summary["private_raw_parseable_lane_count"] == 3,
        "zero_row_bindings": summary["row_level_binding_proven_lane_count"] == 0,
        "four_issue_definitions": summary["issue_definition_count"] == 4,
        "zero_identified_items": summary["identified_business_item_count"] == 0,
        "zero_actionable_priorities": summary["actionable_collection_priority_item_count"] == 0,
        "zero_assigned_responsibilities": summary["assigned_responsibility_item_count"] == 0,
        "five_raw_sources": summary["raw_source_file_count"] == 5,
        "two_wps_private_containers": summary["wps_private_container_count"] == 2,
        "raw_exact": summary["raw_snapshot_exact_match"],
        "raw_cross_phase_exact": summary["raw_cross_phase_snapshot_exact_match"],
        "private_differences_recorded": summary["private_difference_item_count"] >= 2,
        "browser_pass": summary["browser_status"] == "PASS",
        "no_console_errors": summary["console_error_count"] == 0,
        "no_overflow": summary["horizontal_overflow_count"] == 0,
        "no_formal_report": summary["formal_report_count"] == 0,
        "no_decision_basis": summary["business_decision_basis_count"] == 0,
        "no_s13_p3": not summary["s13_p3_performed"],
        "no_upload": not summary["github_upload_performed"],
        "no_reinstall": not summary["app_reinstall_performed"],
        "no_business_execution": not summary["business_execution_performed"],
    }
    rows = [{"check_id": key, "passed": value} for key, value in sorted(checks.items())]
    return {
        "schema_version": "kmfa.v014.s13_p2_post_remediation.acceptance_matrix.v1",
        "check_count": len(rows),
        "check_pass_count": sum(row["passed"] for row in rows),
        "check_fail_count": sum(not row["passed"] for row in rows),
        "checks": rows,
    }


def _phase_public_files() -> list[str]:
    paths = (
        SUMMARY_PATH,
        MANIFEST_PATH,
        LANES_PATH,
        ISSUES_PATH,
        MATRIX_PATH,
        GO_NO_GO_PATH,
        HTML_PATH,
        REPORT_PATH,
        TEST_RESULTS_PATH,
        RISK_REGISTER_PATH,
        ROLLBACK_PATH,
        METADATA_SUMMARY_PATH,
        METADATA_MANIFEST_PATH,
        METADATA_LANES_PATH,
        METADATA_ISSUES_PATH,
        METADATA_MATRIX_PATH,
        METADATA_GO_NO_GO_PATH,
    )
    governance_paths = (
        Path("KMFA/AGENTS.md"),
        Path("KMFA/CHANGELOG.md"),
        Path("KMFA/HANDOFF.md"),
        Path("KMFA/VERSION"),
        Path("KMFA/docs/governance/ASSURANCE_STATUS.yaml"),
        Path("KMFA/docs/governance/DEVELOPMENT_LEDGER.md"),
        Path("KMFA/docs/governance/MODEL_SPEC.md"),
        Path("KMFA/docs/governance/TRACEABILITY_MATRIX.csv"),
        Path("KMFA/docs/governance/VERSION_MATRIX.yaml"),
        Path("KMFA/docs/governance/delivery_tasks.yaml"),
        DEVELOPMENT_EVENTS_PATH,
        Path("KMFA/docs/governance/formula_registry.yaml"),
        Path("KMFA/docs/governance/model_registry.yaml"),
        Path("KMFA/docs/governance/parameter_registry.csv"),
        Path("KMFA/metadata/model_registry.yaml"),
        STAGE_STATUS_PATH,
        TASK_STATUS_PATH,
        Path("KMFA/功能清单.md"),
        Path("KMFA/开发记录.md"),
        Path("KMFA/模型参数文件.md"),
    )
    return [path.as_posix() for path in paths + governance_paths] + [
        "KMFA/tools/v014_s13_p2_post_remediation_collection_receivable_aging.py",
        "KMFA/tools/check_v014_s13_p2_post_remediation_collection_receivable_aging.py",
        "KMFA/tests/test_v014_s13_p2_post_remediation_collection_receivable_aging.py",
    ]


def _write_governance(generated_at: str) -> None:
    evidence_ref = MANIFEST_PATH.as_posix()
    _upsert_jsonl(
        DEVELOPMENT_EVENTS_PATH,
        {
            "event_id": "DEV-KMFA-20260711-V014-S13-P2-POST-REMEDIATION-COLLECTION-RECEIVABLE-AGING",
            "event_time": generated_at,
            "event_type": "phase_completion",
            "project_id": "KMFA",
            "stage_id": "S13",
            "phase_id": PHASE_ID,
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "task_id": TASK_ID,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "raw_business_data_committed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "files_changed": _phase_public_files(),
            "result_commit": "PENDING",
        },
    )
    _upsert_jsonl(
        STAGE_STATUS_PATH,
        {
            "schema_version": "kmfa.stage_status.v1",
            "record_type": "v014_phase_status",
            "project_id": "KMFA",
            "stage_id": "S13",
            "phase_id": PHASE_ID,
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "task_id": TASK_ID,
            "version": VERSION,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "derived_percent": "66.67",
            "raw_data_committed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at,
        },
    )
    _upsert_jsonl(
        TASK_STATUS_PATH,
        {
            "schema_version": "kmfa.v014_stage_phase_task_status.v1",
            "record_type": "v014_phase",
            "project_id": "KMFA",
            "stage_id": "S13",
            "governance_stage_id": "FINANCIAL-OPERATING-REPORTING",
            "roadmap_stage_id": "S13",
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "acceptance_id": ACCEPTANCE_ID,
            "version": VERSION,
            "name": "v0.1.4 S13-P2 post-remediation collection and receivable aging",
            "phase_goal": "lock five source-lane structures and four issue methods without inventing business items",
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "derived_percent": 100,
            "estimated_task_units": 3,
            "completed_task_units": 3,
            "task_count": 3,
            "raw_data_committed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at[:10],
        },
    )


def _render_report(summary: dict[str, Any]) -> str:
    return f"""# KMFA v0.1.4 S13-P2 回款与应收账龄

## 结论

- 当前状态：`Q4 / D / NO_GO / 3-9-2-1`
- 五类主题：结构 `5/5`；私有容器可解析 `{summary['private_raw_parseable_lane_count']}/5`；行级绑定 `0/5`
- 四类问题：方法定义 `4/4`；已证明业务项 `0`；可执行回款优先级 `0`；已指派责任事项 `0`
- 原始文件：`{summary['raw_source_file_count']}` 个，phase 前后及跨 S13-P1 快照完全一致
- 私有差异：`{summary['private_difference_item_count']}` 类，仅保存在 ignored private runtime

## 验收解释

T1 的五类来源已完成 public-safe 结构接入，并用只读私有探针区分容器可解析与行级绑定。T2/T3 已锁定四类识别方法、复核顺序与责任角色定义；由于原生 WPS 转换、共享行键和期间对齐尚未被证明，不生成项目、客户、金额、催收优先级或责任人明细。

## 边界

- 不是正式回款报告、催收清单、法律意见、开票指令、付款指令或银行操作依据。
- 未执行 S13-P3、Stage 13 review、GitHub upload、app reinstall、持久业务写入或 business execution。
- 原始身份、文件名、字段、表头、金额与诊断仅保留在 ignored private runtime。
"""


def _render_private_difference_report(
    probe: dict[str, Any], profile: dict[str, Any]
) -> str:
    source_lines = "\n".join(
        f"- {row['private_filename']}：{row['container_kind']}，只读探针已执行。"
        for row in probe["private_source_details"]
    )
    return f"""# S13-P2 私有差异报告

## 原始来源只读核对

{source_lines}

## 当前差异

1. 回款表对应来源为 WPS 私密复合文档，当前环境缺少原生 WPS 安全转换能力。
2. 应收账龄对应来源为 WPS 私密复合文档，当前环境缺少原生 WPS 安全转换能力。
3. 五类来源之间没有被证明的共享行级主键，不能把项目、客户、发票、回款和账龄明细强行拼接。
4. 跨来源期间口径没有被证明一致，不能生成逾期或回款优先级业务结论。

## 处理结论

- 原始文件数：{probe['raw_source_file_count']}
- WPS 私密容器数：{probe['wps_private_container_count']}
- 可打开归档工作簿数：{probe['openable_archive_workbook_member_count']}
- 私有结构抽样工作簿：{profile['openable_workbook_count']}/{profile['workbook_candidate_count']}
- 五类主题均观察到候选文档；公共键标签类数：{len(profile['common_key_classes_across_all_lanes'])}
- 私有候选信号数：{profile['private_candidate_signal_count']}（仅候选，不是已识别业务事项）
- 差异类别数：{len(probe['private_differences'])}
- 原始文件未修改、未删除、未移动、未重命名、未覆盖。
- 本 phase 保留 4 类方法定义并将业务项锁定为 0；最终 goal 若多轮交叉验证仍未解除，纳入全中文最终差异报告。
"""


def generate(*, final_validation: bool = False, write_governance: bool = True) -> dict[str, Any]:
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    raw_helper = s13_p1.s12_review.p1.s11_project
    raw_before = raw_helper._raw_snapshot("before_v014_s13_p2_post_remediation_collection_receivable_aging")
    dependency = _load_dependency()
    contract = _load_contract()
    historical = _read_json(HISTORICAL_MANIFEST_PATH)
    probe = _probe_raw_sources(Path(raw_before["raw_root"]))
    profile = _run_private_workbook_profile()
    if profile["openable_workbook_count"] != probe["openable_archive_workbook_member_count"]:
        raise RuntimeError("private workbook profile openable count drift")
    lanes = _build_source_lanes(probe)
    issues = _build_issue_definitions()
    _write_text(HTML_PATH, _render_html(lanes, issues))
    browser = _run_browser_review()
    raw_after = raw_helper._raw_snapshot("after_v014_s13_p2_post_remediation_collection_receivable_aging")
    prior_raw = _read_json(s13_p1.PRIVATE_RAW_AFTER_PATH)
    current_raw = raw_helper._raw_snapshot("current_v014_s13_p2_post_remediation_collection_receivable_aging")
    normalize = raw_helper._normalize_raw
    raw_exact = normalize(raw_before) == normalize(raw_after)
    raw_cross = normalize(raw_before) == normalize(prior_raw) == normalize(current_raw)
    if not raw_exact or not raw_cross:
        raise RuntimeError("raw source changed during S13-P2")

    upstream = dependency["summary"]
    parseable_count = sum(row["private_raw_parseable"] for row in lanes)
    summary = {
        "schema_version": "kmfa.v014.s13_p2_post_remediation.summary.v1",
        "project_id": "KMFA",
        "stage_id": "S13",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "s13_p1_dependency_validated": True,
        "source_lane_count": len(lanes),
        "structure_connected_lane_count": sum(row["structure_connected"] for row in lanes),
        "private_raw_parseable_lane_count": parseable_count,
        "row_level_binding_proven_lane_count": sum(row["row_level_binding_proven"] for row in lanes),
        "required_issue_type_count": len(ISSUE_SPECS),
        "issue_definition_count": len(issues),
        "identified_business_item_count": sum(row["identified_item_count"] for row in issues),
        "priority_review_definition_count": len(issues),
        "actionable_collection_priority_item_count": 0,
        "responsibility_role_definition_count": len(issues),
        "assigned_responsibility_item_count": 0,
        "workbench_html_count": 1,
        "open_final_difference_accepted_count": upstream["open_final_difference_accepted_count"],
        "nonzero_delta_reconciliation_count": upstream["nonzero_delta_reconciliation_count"],
        "zero_delta_reconciliation_count": upstream["zero_delta_reconciliation_count"],
        "incomplete_reconciliation_count": upstream["incomplete_reconciliation_count"],
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "formal_report_count": 0,
        "business_decision_basis_count": 0,
        "raw_source_file_count": raw_before["file_count"],
        "wps_private_container_count": probe["wps_private_container_count"],
        "archive_workbook_member_count": probe["archive_workbook_member_count"],
        "openable_archive_workbook_member_count": probe["openable_archive_workbook_member_count"],
        "archive_pdf_member_count": probe["archive_pdf_member_count"],
        "openable_archive_pdf_member_count": probe["openable_archive_pdf_member_count"],
        "private_difference_item_count": len(probe["private_differences"]),
        "private_workbook_structure_profile_performed": True,
        "raw_snapshot_exact_match": raw_exact,
        "raw_cross_phase_snapshot_exact_match": raw_cross,
        "browser_status": browser["status"],
        "browser_viewport_check_count": browser["viewport_check_count"],
        "issue_interaction_check_count": browser["issue_interaction_check_count"],
        "dependency_link_http_check_count": browser["dependency_link_http_check_count"],
        "dependency_navigation_check_count": browser["dependency_navigation_check_count"],
        "console_error_count": browser["console_error_count"],
        "horizontal_overflow_count": browser["horizontal_overflow_count"],
        "s13_p3_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
    }
    matrix = _matrix(summary)
    validation_summary = {
        "final_validation_recorded": final_validation,
        "focused_test": "PASS" if final_validation else "PENDING",
        "strict_validator": "PASS" if final_validation else "PENDING",
        "browser_desktop_mobile": "PASS" if final_validation else "PENDING",
        "raw_alignment": "PASS" if final_validation else "PENDING",
        "governance_and_safety_scans": "PASS" if final_validation else "PENDING",
    }
    historical_summary = historical.get("collection_receivable_summary", {})
    manifest = {
        "schema_version": "kmfa.v014.s13_p2_post_remediation.manifest.v1",
        "project_id": "KMFA",
        "stage_id": "S13",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "generated_at": generated_at,
        "reviewed_head": _git_output(["rev-parse", "HEAD"]),
        "branch": _git_output(["branch", "--show-current"]),
        "remote": _git_output(["remote", "get-url", "origin"]),
        "summary": summary,
        "source_lane_status": lanes,
        "issue_definitions": issues,
        "acceptance_matrix": matrix,
        "quality_gate": _quality_gate(),
        "phase_boundaries": _phase_boundaries(),
        "raw_boundary": {
            "raw_read_authorized": True,
            "raw_snapshot_validation_performed": True,
            "private_readonly_probe_performed": True,
            "private_workbook_structure_profile_performed": True,
            "private_difference_report_generated": True,
            "raw_write_performed": False,
            "raw_delete_performed": False,
            "raw_move_performed": False,
            "raw_rename_performed": False,
            "raw_overwrite_performed": False,
            "raw_mutation_performed": False,
        },
        "public_repo_safety": _public_repo_safety(),
        "browser_review": browser,
        "s13_p1_dependency_validated": True,
        "historical_s13_p2_policy_fixture_validated": (
            historical.get("stage_id") == "S13"
            and historical_summary.get("source_lane_count") == 5
            and historical_summary.get("required_issue_type_count") == 4
        ),
        "historical_s13_p2_dynamic_state_is_authoritative": False,
        "historical_pending_twelve_quarantined": (
            historical_summary.get("pending_reconciliation_count") == 12
        ),
        "historical_static_priority_and_responsibility_quarantined": (
            historical_summary.get("priority_item_count") == 4
            and historical_summary.get("responsibility_item_count") == 4
        ),
        "taskpack_contract": contract,
        "reviewed_dependencies": {
            "current_s13_p1": s13_p1.MANIFEST_PATH.as_posix(),
            "current_wps_structure_registry": WPS_REGISTRY_PATH.as_posix(),
            "current_finance_structure_registry": FINANCE_REGISTRY_PATH.as_posix(),
            "historical_s13_p2_policy_fixture": HISTORICAL_MANIFEST_PATH.as_posix(),
        },
        "next_phase": "S13-P3",
        "next_required_step": "Execute S13-P3 cross-table review as a separate run after S13-P2 local validation and commit.",
        "validation_summary": validation_summary,
    }
    go_no_go = {
        "schema_version": "kmfa.v014.s13_p2_post_remediation.go_no_go.v1",
        "project_id": "KMFA",
        "stage_id": "S13",
        "phase_id": PHASE_ID,
        "decision": "NO_GO",
        "method_definition_allowed": True,
        "business_priority_allowed": False,
        "responsibility_assignment_allowed": False,
        "collection_action_allowed": False,
        "formal_report_allowed": False,
        "business_decision_basis_allowed": False,
        "s13_p3_performed": False,
        "stage13_review_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
    }
    alignment = {
        "schema_version": "kmfa.private.v014.s13_p2.row_binding_diagnostic.v1",
        "classification": "PRIVATE_RUNTIME_ONLY",
        "raw_source_file_count": raw_before["file_count"],
        "source_lane_count": len(lanes),
        "structure_connected_lane_count": summary["structure_connected_lane_count"],
        "private_raw_parseable_lane_count": parseable_count,
        "row_level_binding_proven_lane_count": 0,
        "identified_business_item_count": 0,
        "actionable_collection_priority_item_count": 0,
        "assigned_responsibility_item_count": 0,
        "private_difference_item_count": len(probe["private_differences"]),
        "private_differences": probe["private_differences"],
        "private_workbook_structure_profile_performed": True,
        "private_workbook_candidate_count": profile["workbook_candidate_count"],
        "private_openable_workbook_count": profile["openable_workbook_count"],
        "private_common_key_class_count": len(
            profile["common_key_classes_across_all_lanes"]
        ),
        "private_candidate_signal_count": profile["private_candidate_signal_count"],
        "private_candidate_signal_is_authoritative_business_item": False,
        "raw_snapshot_exact_match": raw_exact,
        "raw_cross_phase_snapshot_exact_match": raw_cross,
        "business_value_comparison_performed": False,
        "business_value_comparison_blocked_by_unproven_row_binding": True,
        "difference_report_required_for_this_phase": True,
        "final_goal_difference_report_required_if_unresolved": True,
    }

    for path, value in (
        (SUMMARY_PATH, summary),
        (MANIFEST_PATH, manifest),
        (LANES_PATH, {"schema_version": "kmfa.v014.s13_p2.source_lanes.v1", "lanes": lanes}),
        (ISSUES_PATH, {"schema_version": "kmfa.v014.s13_p2.issue_definitions.v1", "issues": issues}),
        (MATRIX_PATH, matrix),
        (GO_NO_GO_PATH, go_no_go),
        (METADATA_SUMMARY_PATH, summary),
        (METADATA_MANIFEST_PATH, manifest),
        (METADATA_LANES_PATH, {"schema_version": "kmfa.v014.s13_p2.source_lanes.v1", "lanes": lanes}),
        (METADATA_ISSUES_PATH, {"schema_version": "kmfa.v014.s13_p2.issue_definitions.v1", "issues": issues}),
        (METADATA_MATRIX_PATH, matrix),
        (METADATA_GO_NO_GO_PATH, go_no_go),
        (PRIVATE_RAW_BEFORE_PATH, raw_before),
        (PRIVATE_RAW_AFTER_PATH, raw_after),
        (PRIVATE_SOURCE_PROBE_PATH, probe),
        (PRIVATE_ALIGNMENT_PATH, alignment),
    ):
        _write_json(path, value)
    _write_text(REPORT_PATH, _render_report(summary))
    _write_text(
        TEST_RESULTS_PATH,
        f"""# S13-P2 测试结果

- focused test / strict validator：最终复验结果见 manifest
- v1.4 baseline：`{browser['baseline_pass_count']} PASS / {browser['baseline_warn_count']} WARN / {browser['baseline_fail_count']} FAIL`
- current HTML audit：`{browser['current_pass_count']} PASS / {browser['current_warn_count']} WARN / {browser['current_fail_count']} FAIL`
- desktop/mobile：`{browser['viewport_check_count']}/2 PASS`
- 四类问题交互：`{browser['issue_interaction_check_count']}/4 PASS`
- S13-P1 链接 HTTP / 真实导航：`{browser['dependency_link_http_check_count']}/2 / {browser['dependency_navigation_check_count']}/2 PASS`
- raw 前后/跨 S13-P1/current：exact match
""",
    )
    _write_text(
        RISK_REGISTER_PATH,
        """# S13-P2 风险登记

| 风险 | 控制 | 状态 |
|---|---|---|
| WPS 私密容器被错误当作普通 Excel | 只读魔数探针；明确要求原生 WPS 转换；不对原文件另存或修改 | controlled |
| 结构接入被误读为行级绑定 | 结构、私有可解析、行级已证明三层状态分开显示 | controlled |
| 历史 12 pending 或静态优先级回流 | 历史产物只作 policy fixture，动态状态不具权威性 | controlled |
| 复核顺序被误用为催收优先级 | 优先级标记为 method-only；可执行业务项固定为 0 | controlled |
| 责任角色被误用为个人指派 | 只保留角色定义；责任人指派固定为 0 | controlled |
| raw/private/secret 进入 Git | 原始详情和差异只写 ignored private runtime；公开证据只含聚合计数 | controlled |
""",
    )
    _write_text(
        ROLLBACK_PATH,
        f"""# S13-P2 回滚计划

1. 回退本 phase 的本地 commit 和 `{OUTPUT_DIR.as_posix()}` 公开证据。
2. 删除 ignored private browser/raw/probe/difference evidence，不触碰原始目录。
3. 恢复 S13-P1 为当前治理入口；不进入 S13-P3 或 Stage 13 review。
4. 不回退、不移动、不删除、不覆盖任何原始文件。
""",
    )
    _write_text(
        PRIVATE_DIFFERENCE_REPORT_PATH,
        _render_private_difference_report(probe, profile),
    )
    if write_governance:
        _write_governance(generated_at)
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--final-validation", action="store_true")
    parser.add_argument("--browser-evidence-only", action="store_true")
    parser.add_argument("--private-profile-only", action="store_true")
    args = parser.parse_args()
    if args.private_profile_only:
        result = _private_workbook_profile_worker()
        print(
            "S13-P2 private workbook profile: "
            f"candidates={result['workbook_candidate_count']} "
            f"openable={result['openable_workbook_count']} "
            f"signals={result['private_candidate_signal_count']}"
        )
        return 0
    if args.browser_evidence_only:
        result = _browser_worker()
        print(
            "S13-P2 browser evidence: "
            f"viewports={len(result['viewport_checks'])} "
            f"issues={len(result['issue_interaction_checks'])} status={result['status']}"
        )
        return 0
    manifest = generate(final_validation=args.final_validation)
    summary = manifest["summary"]
    print(
        "S13-P2 post-remediation collection receivable aging: "
        f"lanes={summary['source_lane_count']} parseable={summary['private_raw_parseable_lane_count']} "
        f"bound={summary['row_level_binding_proven_lane_count']} "
        f"items={summary['identified_business_item_count']} decision={summary['decision']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
