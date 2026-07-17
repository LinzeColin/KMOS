#!/usr/bin/env python3
"""40 个 Excel 的全量 sheet 盘点与 S07 适配器覆盖映射（TSK.KMFA.DATA.0007 阶段一）。

只读扫描 KMDatabase/data 里已登记的 xlsx/xls：
  - 公开面（stdout/证据）：每文件 = 指纹前8位/域/格式/sheet 数/最大行列/匹配到的适配器类别——**零单元格值、零 sheet 名明文**
  - 私有面（.codex_private_runtime/imports/excel_sheet_inventory/）：完整 sheet 名清单

类别关键词映射（文件名+sheet 名 → v0.1.4 S07 规格类别）在本文件维护；
未匹配的显式标 unknown_deferred（任务包允许"未映射显式登记 deferred"）。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
PRIVATE_DIR = REPO / "KMFA" / ".codex_private_runtime" / "imports" / "excel_sheet_inventory"

# 类别 → (规格来源, 关键词)。规格类别对齐 finance_file_adapter.FINANCE_CATEGORY_SPECS
# 与 wps_file_adapter.WPS_EXPORT_SPECS 的类别名；kingdee_ledger/tax/loan/performance 为 DT5 新增待建规格。
CATEGORY_KEYWORDS = [
    ("journal", "finance", ("日记账",)),
    ("expense_lines", "finance", ("报表系列一",)),
    ("operating_analysis", "finance", ("费用", "税金", "资产", "经营分析", "经营情况")),
    ("collection", "wps", ("回款",)),
    ("receivable_aging", "wps", ("账龄", "应收")),
    ("invoicing", "wps", ("开票",)),
    ("payment_approval", "wps", ("付款审批", "付款申请")),
    ("contract", "wps", ("合同",)),
    ("kingdee_ledger", "pending_spec", ("明细账", "金蝶", "序时账", "科目")),
    ("tax_filing", "pending_spec", ("税", "申报", "纳税")),
    ("loan", "pending_spec", ("贷款", "借款")),
    ("performance", "pending_spec", ("绩效", "考核")),
    ("primary_data", "pending_spec", ("一级", "二级", "数据支撑")),
]


def classify(name_pool: str) -> tuple[str, str]:
    for category, source, keywords in CATEGORY_KEYWORDS:
        if any(k in name_pool for k in keywords):
            return category, source
    return "unknown_deferred", "none"


def sheets_of(path: Path, fmt: str) -> list[dict]:
    out = []
    if fmt == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            out.append({"name": ws.title, "max_row": ws.max_row or 0, "max_col": ws.max_column or 0})
        wb.close()
    else:
        import xlrd
        wb = xlrd.open_workbook(str(path), on_demand=True)
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            out.append({"name": sheet_name, "max_row": ws.nrows, "max_col": ws.ncols})
        wb.release_resources()
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--kmdb-manifest", default="KMDatabase/data/manifest.jsonl")
    parser.add_argument("--out", default="KMFA/stage_artifacts/DT5_DATA0007_excel_inventory/machine/excel_inventory.json")
    args = parser.parse_args()

    rows = [json.loads(l) for l in (REPO / args.kmdb_manifest).read_text(encoding="utf-8").splitlines() if l.strip()]
    targets = [r for r in rows if r["original_name"].lower().endswith((".xlsx", ".xls"))]

    public_files, private_files, category_counts, errors = [], [], {}, []
    for row in sorted(targets, key=lambda r: r["original_name"]):
        name = unicodedata.normalize("NFC", row["original_name"])
        fmt = "xlsx" if name.lower().endswith(".xlsx") else "xls"
        obj = REPO / "KMDatabase" / "data" / row["object_path"]
        try:
            sheets = sheets_of(obj, fmt)
        except Exception as exc:  # noqa: BLE001 - 盘点必须报全不中断
            errors.append({"sha8": row["sha256"][:8], "error": exc.__class__.__name__})
            continue
        pool = name + "|" + "|".join(s["name"] for s in sheets)
        category, spec_source = classify(pool)
        category_counts[category] = category_counts.get(category, 0) + 1
        public_files.append({
            "sha8": row["sha256"][:8],
            "domain": row["domain"],
            "format": fmt,
            "size_bytes": row["size_bytes"],
            "sheet_count": len(sheets),
            "max_rows": max((s["max_row"] for s in sheets), default=0),
            "matched_category": category,
            "spec_source": spec_source,
        })
        private_files.append({"sha256": row["sha256"], "original_name": name, "sheets": sheets,
                              "matched_category": category})

    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)
    (PRIVATE_DIR / "full_inventory.json").write_text(
        json.dumps(private_files, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "task_id": "TSK.KMFA.DATA.0007",
        "phase": "1_inventory_and_adapter_mapping",
        "excel_total": len(targets),
        "inventoried": len(public_files),
        "open_errors": errors,
        "category_counts": dict(sorted(category_counts.items())),
        "spec_coverage": {
            "finance_or_wps_spec_ready": sum(1 for f in public_files if f["spec_source"] in ("finance", "wps")),
            "pending_new_spec": sum(1 for f in public_files if f["spec_source"] == "pending_spec"),
            "unknown_deferred": sum(1 for f in public_files if f["spec_source"] == "none"),
        },
        "sheet_total": sum(f["sheet_count"] for f in public_files),
        "private_full_inventory_ref": str((PRIVATE_DIR / "full_inventory.json").relative_to(REPO)),
        "files": public_files,
    }
    out_path = REPO / args.out
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: v for k, v in summary.items() if k != "files"}, ensure_ascii=False))
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
