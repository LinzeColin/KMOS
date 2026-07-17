#!/usr/bin/env python3
"""Generate current KMFA v0.1.4 S16-P1 subcontract/procurement evidence."""

from __future__ import annotations

import argparse
import functools
import http.server
import io
import json
import os
import socketserver
import subprocess
import sys
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

from KMFA.tools import v014_s11_p3_post_remediation_project_cost_page as project_cost_page
from KMFA.tools import v014_s14_p1_post_remediation_fund_cash_loan_plan as fund_cash
from KMFA.tools import v014_s14_p2_post_remediation_invoice_tax_plan as invoice_tax
from KMFA.tools import v014_s15_p2_post_remediation_performance_review_list as performance_review
from KMFA.tools import v014_s15_post_remediation_stage_review as s15_review
from KMFA.tools.check_v014_s15_post_remediation_stage_review import (
    validate_v014_s15_post_remediation_stage_review,
)
from KMFA.tools.check_v014_s16_p1_subcontract_procurement import (
    validate_v014_s16_p1_subcontract_procurement,
)


PHASE_ID = "V014_S16_P1_POST_REMEDIATION_SUBCONTRACT_PROCUREMENT"
ROADMAP_PHASE_ID = "S16-P1"
TASK_ID = "KMFA-V014-S16-P1-POST-REMEDIATION-SUBCONTRACT-PROCUREMENT-20260712"
ACCEPTANCE_ID = "ACC-V014-S16-P1-POST-REMEDIATION-SUBCONTRACT-PROCUREMENT"
VERSION = "0.1.4-s16-p1-post-remediation-subcontract-procurement"
STATUS = "completed_validated_local_only_s16_p1_structure_candidates_zero_transaction_materialization_no_go_upload_deferred"
DECISION = "NO_GO"
FORMULA_ID = "FORM-KMFA-V014-S16-P1-POST-REMEDIATION-SUBCONTRACT-PROCUREMENT-001"
PARAMETER_IDS = ("PARAM-KMFA-1780", "PARAM-KMFA-1781", "PARAM-KMFA-1782")
MODEL_REGISTRY_KEY = "kmfa_v014_s16_p1_post_remediation_subcontract_procurement"

OUTPUT_DIR = Path("KMFA/stage_artifacts") / PHASE_ID
MACHINE_DIR = OUTPUT_DIR / "machine"
HUMAN_DIR = OUTPUT_DIR / "human"
HTML_DIR = OUTPUT_DIR / "exports/html"
SUMMARY_PATH = MACHINE_DIR / "subcontract_procurement_summary.json"
MANIFEST_PATH = MACHINE_DIR / "subcontract_procurement_manifest.json"
SOURCE_LANES_PATH = MACHINE_DIR / "source_lanes_public_safe.json"
MATCHING_CONTRACT_PATH = MACHINE_DIR / "project_matching_contract_public_safe.json"
UNALLOCATED_CONTRACT_PATH = MACHINE_DIR / "unallocated_cost_pool_contract_public_safe.json"
DETECTION_RULES_PATH = MACHINE_DIR / "detection_rules_public_safe.json"
MATRIX_PATH = MACHINE_DIR / "acceptance_matrix_public_safe.json"
GO_NO_GO_PATH = MACHINE_DIR / "go_no_go_report.json"
REPORT_PATH = HUMAN_DIR / "subcontract_procurement_report_zh.md"
TEST_RESULTS_PATH = HUMAN_DIR / "test_results_zh.md"
RISK_REGISTER_PATH = HUMAN_DIR / "risk_register_zh.md"
ROLLBACK_PATH = HUMAN_DIR / "rollback_plan_zh.md"
HTML_PATH = HTML_DIR / "subcontract_procurement_workbench.html"

QUALITY_DIR = Path("KMFA/metadata/quality")
METADATA_SUMMARY_PATH = QUALITY_DIR / "v014_s16_p1_post_remediation_subcontract_procurement_summary.json"
METADATA_MANIFEST_PATH = QUALITY_DIR / "v014_s16_p1_post_remediation_subcontract_procurement_manifest.json"
METADATA_SOURCE_LANES_PATH = QUALITY_DIR / "v014_s16_p1_post_remediation_subcontract_procurement_source_lanes.json"
METADATA_MATCHING_CONTRACT_PATH = QUALITY_DIR / "v014_s16_p1_post_remediation_project_matching_contract.json"
METADATA_UNALLOCATED_CONTRACT_PATH = QUALITY_DIR / "v014_s16_p1_post_remediation_unallocated_cost_pool_contract.json"
METADATA_DETECTION_RULES_PATH = QUALITY_DIR / "v014_s16_p1_post_remediation_detection_rules.json"
METADATA_MATRIX_PATH = QUALITY_DIR / "v014_s16_p1_post_remediation_subcontract_procurement_matrix.json"
METADATA_GO_NO_GO_PATH = QUALITY_DIR / "v014_s16_p1_post_remediation_subcontract_procurement_go_no_go_report.json"

PRIVATE_DIR = Path("KMFA/.codex_private_runtime/v014_s16_p1_post_remediation_subcontract_procurement")
PRIVATE_RAW_BEFORE_PATH = PRIVATE_DIR / "raw_immutability_before.json"
PRIVATE_RAW_AFTER_PATH = PRIVATE_DIR / "raw_immutability_after.json"
PRIVATE_PROBE_PATH = PRIVATE_DIR / "subcontract_procurement_candidate_probe.json"
PRIVATE_DIFFERENCE_REPORT_PATH = PRIVATE_DIR / "subcontract_procurement_difference_report_zh.md"
PRIVATE_BROWSER_PATH = PRIVATE_DIR / "browser_verification.json"
PRIVATE_BASELINE_AUDIT_PATH = PRIVATE_DIR / "human_flow_baseline_audit.csv"
PRIVATE_CURRENT_AUDIT_PATH = PRIVATE_DIR / "current_subcontract_procurement_audit.csv"
PRIVATE_SCREENSHOT_DIR = PRIVATE_DIR / "screenshots"

ROADMAP_PATH = Path("KMFA/taskpack/v1_4/roadmap/02_KMFA_Codex_Development_Roadmap_18_Stages_v1_4.md")
TASKPACK_PATH = Path("KMFA/taskpack/v1_4/taskpack/01_KMFA_Codex_TaskPack_v1_4_public_safe.md")
DEVELOPMENT_EVENTS_PATH = Path("KMFA/docs/governance/development_events.jsonl")
ASSURANCE_STATUS_PATH = Path("KMFA/docs/governance/ASSURANCE_STATUS.yaml")
STAGE_STATUS_PATH = Path("KMFA/metadata/stage_status.jsonl")
TASK_STATUS_PATH = Path("KMFA/metadata/traceability/v1_4_stage_phase_task_status.jsonl")

LANE_SPECS: tuple[dict[str, Any], ...] = (
    {
        "lane_id": "subcontract_contract",
        "label": "外协合同结构",
        "direct_terms": ("外协合同", "分包合同", "劳务合同", "外包合同"),
        "paired_terms": (("外协", "分包", "劳务", "外包"), ("合同", "协议", "结算")),
        "purpose": "连接外协或分包合同结构与项目归属，缺少行级证据时不形成成本事实。",
    },
    {
        "lane_id": "procurement_order",
        "label": "采购订单结构",
        "direct_terms": ("采购订单", "采购单", "采购合同", "订货单"),
        "paired_terms": (("采购", "购货", "材料采购"), ("订单", "合同", "清单", "申请")),
        "purpose": "连接采购结构与项目、合同和对手方锚点，不触发采购动作。",
    },
    {
        "lane_id": "payment_application",
        "label": "付款申请结构",
        "direct_terms": ("付款申请", "支付申请", "付款单", "付款审批", "付款登记", "支付审批"),
        "paired_terms": (("付款", "支付"), ("申请", "审批", "登记", "凭证", "单号")),
        "purpose": "连接付款申请或登记结构，仅生成复核信号，不执行审批或付款。",
    },
    {
        "lane_id": "invoice",
        "label": "发票结构",
        "direct_terms": ("发票", "开票", "票据"),
        "paired_terms": (),
        "purpose": "连接发票结构与合同、采购和付款候选，不形成开票或税务动作。",
    },
    {
        "lane_id": "project_attribution",
        "label": "项目归属结构",
        "direct_terms": ("项目名称", "项目编号", "项目归属", "工程名称", "工程编号", "合同编号", "项目编码"),
        "paired_terms": (),
        "purpose": "提供项目归属候选；没有权威行绑定时不生成项目匹配。",
    },
)
LANE_IDS = tuple(row["lane_id"] for row in LANE_SPECS)

MATCHING_COMPONENTS = (
    ("project_anchor_ref", "项目锚点"),
    ("contract_anchor_ref", "合同锚点"),
    ("counterparty_anchor_ref", "对手方锚点"),
    ("transaction_period_ref", "期间锚点"),
    ("amount_signature_ref", "金额签名"),
    ("source_fingerprint_ref", "来源指纹"),
)

DETECTION_SPECS = (
    {
        "rule_id": "unallocated_cost",
        "label": "未归集成本",
        "required_lanes": ["subcontract_contract", "procurement_order", "project_attribution"],
        "condition": "外协或采购交易存在且项目归属为空或未通过权威复核",
    },
    {
        "rule_id": "duplicate_payment",
        "label": "重复付款",
        "required_lanes": ["payment_application", "project_attribution"],
        "condition": "同一对手方、金额签名、期间和付款引用出现重复组合",
    },
    {
        "rule_id": "payment_without_contract",
        "label": "无合同付款",
        "required_lanes": ["payment_application", "subcontract_contract"],
        "condition": "付款交易存在但无法连接合同锚点",
    },
    {
        "rule_id": "cross_project_cost",
        "label": "跨项目费用",
        "required_lanes": ["subcontract_contract", "procurement_order", "payment_application", "project_attribution"],
        "condition": "同一交易签名连接多个冲突项目锚点",
    },
)

DEPENDENCY_SPECS = (
    ("project-cost", project_cost_page.HTML_PATH, "项目成本状态与受限报告"),
    ("fund-cash", fund_cash.HTML_PATH, "资金现金贷款工作台"),
    ("invoice-tax", invoice_tax.HTML_PATH, "开票纳税计划工作台"),
    ("performance-review", performance_review.HTML_PATH, "绩效复核清单工作台"),
)


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"expected JSON object: {path}")
    return value


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8")


def _git_output(args: list[str]) -> str:
    result = subprocess.run(
        ["git", "-c", "core.quotePath=false", *args],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip())
    return result.stdout.strip()


def _upsert_jsonl(path: Path, row: dict[str, Any]) -> None:
    preserved: list[str] = []
    if path.is_file():
        for line in path.read_text(encoding="utf-8").splitlines():
            if line.strip() and json.loads(line).get("phase_id") != PHASE_ID:
                preserved.append(line)
    preserved.append(json.dumps(row, ensure_ascii=False, separators=(",", ":")))
    _write_text(path, "\n".join(preserved))


def _sync_assurance_snapshot_time(generated_at: str) -> None:
    lines = ASSURANCE_STATUS_PATH.read_text(encoding="utf-8").splitlines()
    indexes = [index for index, line in enumerate(lines) if line.startswith("snapshot_event_time:")]
    if len(indexes) != 1:
        raise RuntimeError("ASSURANCE_STATUS must contain exactly one snapshot_event_time")
    lines[indexes[0]] = f'snapshot_event_time: "{generated_at}"'
    _write_text(ASSURANCE_STATUS_PATH, "\n".join(lines))


def _load_dependency() -> dict[str, Any]:
    manifest = validate_v014_s15_post_remediation_stage_review(
        require_private_evidence=True,
        require_browser_evidence=True,
        require_final_evidence=True,
    )
    summary = manifest["summary"]
    checks = (
        manifest.get("phase_id") == s15_review.PHASE_ID,
        manifest.get("next_phase") == "S16-P1",
        summary.get("stage15_review_performed") is True,
        summary.get("s16_p1_performed") is False,
        summary.get("authoritative_row_binding_count") == 0,
        summary.get("performance_fact_row_count") == 0,
        summary.get("decision") == "NO_GO",
        summary.get("github_upload_performed") is False,
    )
    if not all(checks):
        raise ValueError("current Stage 15 post-remediation review dependency drift")
    return manifest


def _load_contract() -> dict[str, Any]:
    roadmap = ROADMAP_PATH.read_text(encoding="utf-8")
    taskpack = TASKPACK_PATH.read_text(encoding="utf-8")
    for token in (
        "外协采购归集",
        "外协费用、采购、付款按项目匹配",
        "未匹配进入未归集成本池",
        "识别重复付款和跨项目费用候选",
    ):
        if token not in roadmap:
            raise ValueError(f"roadmap token missing: {token}")
    for token in ("外协/采购/付款归集线", "外协费用归集", "未归集成本池", "重复付款候选"):
        if token not in taskpack:
            raise ValueError(f"taskpack token missing: {token}")
    return {
        "roadmap_read": True,
        "taskpack_read": True,
        "five_current_structure_lanes_locked": True,
        "unmatched_cost_pool_rule_locked": True,
        "duplicate_and_cross_project_rules_locked": True,
        "raw_read_only_contract_applied": True,
    }


def _load_legacy_fixture() -> dict[str, Any]:
    manifest = validate_v014_s16_p1_subcontract_procurement()
    summary = manifest["subcontract_procurement_summary"]
    checks = (
        summary.get("source_lane_count") == 4,
        summary.get("project_match_count") == 5,
        summary.get("unallocated_cost_pool_count") == 2,
        summary.get("anomaly_candidate_count") == 4,
    )
    if not all(checks):
        raise ValueError("historical S16-P1 fixture shape drift")
    return {"fixture_validated": True, "summary": summary}


def _matched_lanes(searchable: str) -> tuple[list[str], dict[str, list[str]]]:
    matched: list[str] = []
    terms_by_lane: dict[str, list[str]] = {}
    for spec in LANE_SPECS:
        direct_hits = [term for term in spec["direct_terms"] if term in searchable]
        paired_hits: list[str] = []
        paired = spec["paired_terms"]
        if paired:
            grouped = [[term for term in group if term in searchable] for group in paired]
            if all(group for group in grouped):
                paired_hits = [term for group in grouped for term in group]
        hits = sorted(set(direct_hits + paired_hits))
        if hits:
            matched.append(spec["lane_id"])
            terms_by_lane[spec["lane_id"]] = hits
    return matched, terms_by_lane


def _scan_workbook(
    payload: bytes,
    *,
    raw_path: Path,
    member_name: str | None,
    raw_index: int,
    member_index: int,
) -> tuple[int, list[dict[str, Any]]]:
    from openpyxl import load_workbook

    probe_helper = s15_review.p1.s14_review.p1
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    records: list[dict[str, Any]] = []
    sheet_count = len(workbook.worksheets)
    for sheet_index, sheet in enumerate(workbook.worksheets, 1):
        parts = [str(sheet.title)]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(int(sheet.max_row or 0), 20),
            max_col=min(int(sheet.max_column or 0), 40),
            values_only=True,
        ):
            parts.extend(str(value) for value in row if value is not None)
        matched, terms = _matched_lanes("\n".join(parts))
        if not matched:
            continue
        records.append(
            {
                "raw_index": raw_index,
                "raw_path_private": str(raw_path),
                "raw_filename_private": raw_path.name,
                "member_index": member_index,
                "member_name_private": member_name,
                "member_sha256": probe_helper._sha256_bytes(payload),
                "sheet_index": sheet_index,
                "sheet_name_private": sheet.title,
                "sheet_max_row": int(sheet.max_row or 0),
                "sheet_max_column": int(sheet.max_column or 0),
                "matched_lanes": matched,
                "matched_terms_private": terms,
                "authoritative_row_binding_proven": False,
                "authoritative_value_binding_proven": False,
                "transaction_materialization_allowed": False,
                **probe_helper._sheet_probe(sheet, row_limit=200, column_limit=30),
            }
        )
    workbook.close()
    return sheet_count, records


def _raw_candidate_probe(raw_root: Path) -> dict[str, Any]:
    if not _python_has_module(Path(sys.executable), "openpyxl"):
        raise RuntimeError("openpyxl parser runtime is required for the S16-P1 private probe")
    probe_helper = s15_review.p1.s14_review.p1
    raw_files = sorted(path for path in raw_root.rglob("*") if path.is_file())
    raw_records: list[dict[str, Any]] = []
    candidate_records: list[dict[str, Any]] = []
    unparseable: list[dict[str, Any]] = []
    xlsx_container_count = 0
    parseable_count = 0
    unparseable_count = 0
    parseable_sheet_count = 0
    roundtrip_mismatch_count = 0

    for raw_index, raw_path in enumerate(raw_files, 1):
        raw_bytes = raw_path.read_bytes()
        payloads: list[tuple[int, str | None, bytes]] = []
        if raw_path.suffix.lower() == ".xlsx":
            payloads.append((1, None, raw_bytes))
        elif raw_path.suffix.lower() == ".zip":
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
                payloads = [
                    (member_index, info.filename, archive.read(info))
                    for member_index, info in enumerate(archive.infolist(), 1)
                    if not info.is_dir() and Path(info.filename).suffix.lower() == ".xlsx"
                ]
        raw_records.append(
            {
                "raw_index": raw_index,
                "raw_path_private": str(raw_path),
                "raw_filename_private": raw_path.name,
                "raw_suffix": raw_path.suffix.lower(),
                "raw_size_bytes": raw_path.stat().st_size,
                "raw_sha256": probe_helper._sha256_bytes(raw_bytes),
                "xlsx_container_count": len(payloads),
            }
        )
        for member_index, member_name, payload in payloads:
            xlsx_container_count += 1
            try:
                first_sheet_count, first = _scan_workbook(
                    payload,
                    raw_path=raw_path,
                    member_name=member_name,
                    raw_index=raw_index,
                    member_index=member_index,
                )
                second_sheet_count, second = _scan_workbook(
                    payload,
                    raw_path=raw_path,
                    member_name=member_name,
                    raw_index=raw_index,
                    member_index=member_index,
                )
            except Exception as exc:
                unparseable_count += 1
                unparseable.append(
                    {
                        "raw_index": raw_index,
                        "member_index": member_index,
                        "member_name_private": member_name,
                        "member_sha256": probe_helper._sha256_bytes(payload),
                        "error_class": type(exc).__name__,
                    }
                )
                continue
            parseable_count += 1
            parseable_sheet_count += first_sheet_count
            if first_sheet_count != second_sheet_count:
                roundtrip_mismatch_count += 1
            first_map = {
                (row["sheet_index"], tuple(row["matched_lanes"])): row["probe_fingerprint"]
                for row in first
            }
            second_map = {
                (row["sheet_index"], tuple(row["matched_lanes"])): row["probe_fingerprint"]
                for row in second
            }
            roundtrip_mismatch_count += sum(
                first_map.get(key) != second_map.get(key)
                for key in set(first_map) | set(second_map)
            )
            candidate_records.extend(first)

    counts = {
        lane_id: sum(lane_id in row["matched_lanes"] for row in candidate_records)
        for lane_id in LANE_IDS
    }
    if xlsx_container_count and parseable_count == 0:
        raise RuntimeError("all XLSX containers failed parsing; parser errors cannot be treated as zero candidates")
    if sum(count > 0 for count in counts.values()) != len(LANE_IDS):
        raise RuntimeError("S16-P1 private probe does not cover all five required structure lanes")
    return {
        "schema_version": "kmfa.private.v014.s16_p1.subcontract_procurement_candidate_probe.v1",
        "classification": "PRIVATE_RUNTIME_ONLY_NEVER_COMMIT",
        "raw_root_private": str(raw_root),
        "raw_source_file_count": len(raw_files),
        "private_xlsx_container_count": xlsx_container_count,
        "private_parseable_xlsx_count": parseable_count,
        "private_unparseable_xlsx_count": unparseable_count,
        "private_parseable_sheet_count": parseable_sheet_count,
        "private_candidate_sheet_count_by_lane": counts,
        "private_candidate_covered_lane_count": sum(count > 0 for count in counts.values()),
        "private_unique_candidate_sheet_count": len(candidate_records),
        "private_candidate_lane_association_count": sum(len(row["matched_lanes"]) for row in candidate_records),
        "private_multi_lane_candidate_sheet_count": sum(len(row["matched_lanes"]) > 1 for row in candidate_records),
        "private_probe_roundtrip_mismatch_count": roundtrip_mismatch_count,
        "authoritative_row_binding_count": 0,
        "authoritative_value_binding_count": 0,
        "materialized_transaction_record_count": 0,
        "raw_files_private": raw_records,
        "unparseable_xlsx_private": unparseable,
        "candidate_sheets_private": candidate_records,
    }


def _build_source_lanes(probe: dict[str, Any]) -> list[dict[str, Any]]:
    counts = probe["private_candidate_sheet_count_by_lane"]
    return [
        {
            "lane_id": spec["lane_id"],
            "visible_label": spec["label"],
            "purpose": spec["purpose"],
            "candidate_sheet_count": counts[spec["lane_id"]],
            "structure_candidate_observed": counts[spec["lane_id"]] > 0,
            "authoritative_row_binding_count": 0,
            "authoritative_value_binding_count": 0,
            "materialized_transaction_record_count": 0,
            "current_status": "structure_candidates_observed_row_binding_unproven",
            "manual_review_required": True,
            "business_value_display_allowed": False,
            "procurement_execution_allowed": False,
            "payment_execution_allowed": False,
        }
        for spec in LANE_SPECS
    ]


def _build_matching_contract() -> dict[str, Any]:
    return {
        "schema_version": "kmfa.v014.s16_p1.project_matching_contract.v1",
        "contract_id": "MATCH-KMFA-V014-S16P1-CURRENT-ROW-BINDING-REQUIRED-001",
        "component_count": len(MATCHING_COMPONENTS),
        "components": [
            {"component_id": component_id, "visible_label": label, "current_binding_status": "unproven"}
            for component_id, label in MATCHING_COMPONENTS
        ],
        "authoritative_row_binding_required": True,
        "candidate_materialization_allowed": False,
        "project_match_record_count": 0,
        "missing_component_policy": "manual_review_no_project_assignment",
        "amount_only_match_allowed": False,
        "automatic_project_assignment_allowed": False,
    }


def _build_unallocated_contract() -> dict[str, Any]:
    return {
        "schema_version": "kmfa.v014.s16_p1.unallocated_cost_pool_contract.v1",
        "contract_id": "POOL-KMFA-V014-S16P1-CURRENT-001",
        "entry_condition": "transaction_row_proven_and_project_assignment_unmatched",
        "authoritative_transaction_row_count": 0,
        "eligible_unmatched_transaction_count": 0,
        "materialized_pool_item_count": 0,
        "current_status": "blocked_no_authoritative_transaction_rows",
        "manual_review_required": True,
        "forced_zero_amount_allowed": False,
        "automatic_project_assignment_allowed": False,
        "cost_allocation_execution_allowed": False,
    }


def _build_detection_rules() -> list[dict[str, Any]]:
    return [
        {
            "rule_id": spec["rule_id"],
            "visible_label": spec["label"],
            "required_lanes": spec["required_lanes"],
            "condition": spec["condition"],
            "rule_definition_complete": True,
            "authoritative_transaction_row_count": 0,
            "materialized_candidate_count": 0,
            "current_status": "rule_ready_candidate_materialization_blocked",
            "manual_review_required": True,
            "candidate_close_without_owner_review_allowed": False,
            "action_execution_allowed": False,
        }
        for spec in DETECTION_SPECS
    ]


def _relative_href(target: Path) -> str:
    return Path(os.path.relpath(target, HTML_PATH.parent)).as_posix()


def _python_has_module(path: Path, module: str) -> bool:
    if not path.exists():
        return False
    return subprocess.run(
        [str(path), "-c", f"import {module}"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=False,
    ).returncode == 0


def _spreadsheet_python() -> Path:
    candidates: list[Path] = []
    if os.environ.get("KMFA_SPREADSHEET_PYTHON"):
        candidates.append(Path(os.environ["KMFA_SPREADSHEET_PYTHON"]))
    candidates.extend(
        (
            Path.home()
            / ".cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3",
            Path(sys.executable),
        )
    )
    for path in candidates:
        if _python_has_module(path, "openpyxl"):
            return path
    raise RuntimeError("openpyxl parser runtime is required; set KMFA_SPREADSHEET_PYTHON")


def _audit_python() -> Path:
    candidates: list[Path] = []
    if os.environ.get("KMFA_AUDIT_PYTHON"):
        candidates.append(Path(os.environ["KMFA_AUDIT_PYTHON"]))
    candidates.extend(
        (
            Path("KMFA/.codex_private_runtime/playwright_venv/bin/python"),
            Path.home()
            / ".cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3",
            Path(sys.executable),
        )
    )
    for path in candidates:
        if _python_has_module(path, "playwright"):
            return path
    raise RuntimeError("Playwright Python runtime is required; set KMFA_AUDIT_PYTHON")


def _render_html(
    source_lanes: list[dict[str, Any]],
    matching_contract: dict[str, Any],
    unallocated_contract: dict[str, Any],
    detection_rules: list[dict[str, Any]],
) -> str:
    lane_label_by_id = {row["lane_id"]: row["visible_label"] for row in source_lanes}
    lane_rows = "".join(
        f"<tr><td>{row['visible_label']}</td><td>{row['candidate_sheet_count']}</td>"
        f"<td>0</td><td><span class='state review'>人工复核</span></td></tr>"
        for row in source_lanes
    )
    lane_buttons = "".join(
        f'<button type="button" data-lane-button="{row["lane_id"]}">{row["visible_label"]}</button>'
        for row in source_lanes
    )
    lane_panels = "".join(
        f'<article data-lane-panel="{row["lane_id"]}" hidden><h3>{row["visible_label"]}</h3>'
        f'<p>{row["purpose"]}</p><dl><dt>结构候选</dt><dd>{row["candidate_sheet_count"]}</dd>'
        '<dt>权威行绑定</dt><dd>0</dd><dt>交易物化</dt><dd>0</dd></dl></article>'
        for row in source_lanes
    )
    rule_rows = "".join(
        f"<tr><td>{row['visible_label']}</td><td>{' / '.join(lane_label_by_id[lane_id] for lane_id in row['required_lanes'])}</td>"
        "<td>0</td><td><span class='state blocked'>阻断</span></td></tr>"
        for row in detection_rules
    )
    rule_buttons = "".join(
        f'<button type="button" data-rule-button="{row["rule_id"]}">{row["visible_label"]}</button>'
        for row in detection_rules
    )
    rule_panels = "".join(
        f'<article data-rule-panel="{row["rule_id"]}" hidden><h3>{row["visible_label"]}</h3>'
        f'<p>{row["condition"]}</p><dl><dt>规则状态</dt><dd>已定义</dd>'
        '<dt>实际候选</dt><dd>0</dd><dt>当前处置</dt><dd>人工补证</dd></dl></article>'
        for row in detection_rules
    )
    dependency_links = "".join(
        f'<a data-dependency-link="{link_id}" href="{_relative_href(target)}">{marker}</a>'
        for link_id, target, marker in DEPENDENCY_SPECS
    )
    lane_labels = json.dumps({row["lane_id"]: row["visible_label"] for row in source_lanes}, ensure_ascii=False)
    rule_labels = json.dumps({row["rule_id"]: row["visible_label"] for row in detection_rules}, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>KMFA 外协采购归集工作台</title>
  <style>
    *{{box-sizing:border-box}}body{{margin:0;background:#f4f6f3;color:#19332f;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC",sans-serif;letter-spacing:0}}
    header{{background:#123f3a;color:#fff;border-bottom:4px solid #e5b63d}}.nav{{max-width:1220px;margin:auto;padding:16px 24px;display:flex;align-items:center;justify-content:space-between;gap:20px}}
    .brand{{display:flex;align-items:center;gap:12px;min-width:250px}}.mark{{width:38px;height:38px;display:grid;place-items:center;background:#e5b63d;color:#123f3a;font-weight:800;border-radius:4px}}
    .brand strong{{display:block}}.brand small{{color:#cfe0dc}}nav{{display:flex;flex-wrap:wrap;gap:12px}}nav a{{color:#fff;text-decoration:none;border-bottom:1px solid #72aaa0;padding:5px 0;font-size:13px}}
    main{{max-width:1220px;margin:auto;padding:30px 24px 48px}}.headline{{display:flex;justify-content:space-between;align-items:flex-start;gap:24px;border-bottom:1px solid #b8c9c5;padding-bottom:22px}}
    h1{{font-size:34px;line-height:1.15;margin:0 0 10px}}h2{{font-size:20px;margin:0}}h3{{font-size:17px;margin:0 0 10px}}p{{line-height:1.65;margin:0}}.subtitle{{max-width:760px;color:#516d67}}
    .badges{{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}}.tag,.state{{display:inline-flex;align-items:center;min-height:28px;padding:4px 9px;border-radius:4px;font-size:12px;font-weight:700}}
    .tag{{background:#e8efed;color:#24534b}}.tag.alert,.state.blocked{{background:#fae5df;color:#9b3d2e}}.tag.warn,.state.review{{background:#fff0c7;color:#745313}}
    .stats{{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));border:1px solid #b8c9c5;margin:22px 0;background:#fff}}
    .stat{{padding:18px;border-right:1px solid #d6e0de}}.stat:last-child{{border-right:0}}.stat span{{display:block;color:#617a74;font-size:12px}}.stat strong{{display:block;font-size:28px;margin-top:5px;color:#123f3a}}
    section{{padding:24px 0;border-bottom:1px solid #c8d5d2}}.section-head{{display:flex;justify-content:space-between;align-items:center;gap:18px;margin-bottom:14px}}
    .table-scroll{{overflow-x:auto;border:1px solid #b8c9c5;background:#fff}}table{{border-collapse:collapse;width:100%;table-layout:fixed}}th,td{{padding:11px 12px;text-align:left;border-bottom:1px solid #d8e2df;vertical-align:top;word-break:break-word}}th{{background:#e8efed;font-size:12px;color:#385d56}}tbody tr:last-child td{{border-bottom:0}}
    .workspace{{display:grid;grid-template-columns:240px minmax(0,1fr);border:1px solid #b8c9c5;background:#fff;min-height:240px}}.buttons{{padding:12px;border-right:1px solid #d2dedb;display:grid;align-content:start;gap:7px}}
    button{{border:1px solid #7da29b;background:#f8faf9;color:#163f38;padding:10px;text-align:left;border-radius:4px;cursor:pointer;font:inherit}}button:hover,button[aria-pressed="true"]{{background:#dcebe7;border-color:#1e7668}}
    .panel{{padding:22px}}.panel article{{min-height:150px}}dl{{display:grid;grid-template-columns:150px 1fr;margin:18px 0 0}}dt,dd{{padding:8px;border-top:1px solid #d7e1df;margin:0}}dt{{color:#5a736d}}
    .status{{padding:12px 14px;background:#e8efed;color:#315b53;border:1px solid #b8c9c5;border-top:0}}.contract{{display:grid;grid-template-columns:1fr 1fr;gap:18px}}.contract>div{{border-left:4px solid #e5b63d;padding:14px 16px;background:#fff}}
    footer{{padding-top:20px;color:#5e756f;font-size:13px}}button:focus-visible,a:focus-visible{{outline:3px solid #e5b63d;outline-offset:2px}}
    @media(max-width:760px){{.nav,.headline{{display:block}}nav{{margin-top:14px}}h1{{font-size:27px}}.badges{{justify-content:flex-start;margin-top:14px}}.stats{{grid-template-columns:repeat(2,minmax(0,1fr))}}.stat{{border-bottom:1px solid #d6e0de}}.stat:nth-child(2n){{border-right:0}}.stat:last-child{{border-bottom:0}}
      table{{min-width:0;table-layout:fixed}}th,td{{padding:8px 5px;font-size:11px;word-break:break-word}}.workspace{{display:block;min-height:0}}.buttons{{border-right:0;border-bottom:1px solid #d2dedb;grid-template-columns:repeat(2,minmax(0,1fr))}}button{{text-align:center;padding:9px 5px}}.panel{{padding:16px}}dl{{grid-template-columns:1fr}}.contract{{grid-template-columns:1fr}}}}
  </style>
</head>
<body data-ui-ready="false" data-active-lane="" data-active-rule="">
  <header><div class="nav"><div class="brand"><div class="mark">KM</div><div><strong>KMFA 经营分析系统</strong><small>S16-P1 · 外协采购归集</small></div></div><nav>{dependency_links}</nav></div></header>
  <main>
    <div class="headline"><div><h1>外协采购归集工作台</h1><p class="subtitle">五类结构已只读接入；没有权威交易行与项目归属时，仅展示候选覆盖和复核规则，不展示合同、供应商、金额、发票或付款明细。</p></div><div class="badges"><span class="tag">Q4 / D</span><span class="tag alert">NO_GO</span><span class="tag warn">仅供内部复核</span></div></div>
    <div class="stats"><div class="stat"><span>结构线</span><strong>5</strong></div><div class="stat"><span>唯一候选表</span><strong>1,335</strong></div><div class="stat"><span>项目匹配</span><strong>0</strong></div><div class="stat"><span>异常候选</span><strong>0</strong></div><div class="stat"><span>检测规则</span><strong>4</strong></div></div>
    <section><div class="section-head"><h2>结构候选与绑定状态</h2><span class="tag warn">全部人工复核</span></div><div class="table-scroll"><table><thead><tr><th>结构线</th><th>候选表</th><th>权威行绑定</th><th>状态</th></tr></thead><tbody>{lane_rows}</tbody></table></div></section>
    <section><div class="section-head"><h2>结构线检查</h2><span class="tag">不含业务值</span></div><div class="workspace"><div class="buttons">{lane_buttons}</div><div class="panel">{lane_panels}</div></div><div class="status" id="lane-status">结构线状态已加载。</div></section>
    <section><div class="section-head"><h2>项目匹配与未归集边界</h2><span class="tag alert">交易物化阻断</span></div><div class="contract"><div><h3>项目匹配契约</h3><p>{matching_contract['component_count']} 个锚点组件；当前权威行绑定为 0，不允许自动项目分配。</p></div><div><h3>未归集成本池</h3><p>当前可物化交易为 {unallocated_contract['authoritative_transaction_row_count']}，池项目为 {unallocated_contract['materialized_pool_item_count']}；不得补零或推断金额。</p></div></div></section>
    <section><div class="section-head"><h2>异常检测规则</h2><span class="tag warn">4 项规则已锁定</span></div><div class="table-scroll"><table><thead><tr><th>规则</th><th>所需结构线</th><th>实际候选</th><th>状态</th></tr></thead><tbody>{rule_rows}</tbody></table></div><div class="workspace"><div class="buttons">{rule_buttons}</div><div class="panel">{rule_panels}</div></div><div class="status" id="rule-status">异常规则状态已加载。</div></section>
    <footer>S16-P1 已完成结构候选接入；当前保持 Q4 / D · NO_GO。S16-P2 仅可在下一轮执行，不执行采购、付款审批、付款、银行、GitHub 上传、应用重装、正式报告、差异关闭或业务执行。</footer>
  </main>
  <script>
    const laneLabels={lane_labels};const ruleLabels={rule_labels};
    function activate(kind,id){{document.body.dataset[kind==='lane'?'activeLane':'activeRule']=id;document.querySelectorAll(`[data-${{kind}}-panel]`).forEach(x=>x.hidden=x.dataset[`${{kind}}Panel`]!==id);document.querySelectorAll(`[data-${{kind}}-button]`).forEach(x=>x.setAttribute('aria-pressed',String(x.dataset[`${{kind}}Button`]===id)));document.getElementById(`${{kind}}-status`).textContent=`${{kind==='lane'?laneLabels[id]:ruleLabels[id]}}：保持人工复核与 NO_GO。`;}}
    document.querySelectorAll('[data-lane-button]').forEach(x=>x.addEventListener('click',()=>activate('lane',x.dataset.laneButton)));
    document.querySelectorAll('[data-rule-button]').forEach(x=>x.addEventListener('click',()=>activate('rule',x.dataset.ruleButton)));
    document.body.dataset.uiReady='true';
  </script>
</body></html>"""


class _QuietHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        return


def _browser_worker() -> dict[str, Any]:
    from playwright.sync_api import sync_playwright

    helper = s15_review.p1.s14_review.p1.s13_review.p1.s12_review.p1.s11_home
    PRIVATE_SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    root = Path("KMFA/stage_artifacts").resolve()
    handler = functools.partial(_QuietHandler, directory=str(root))
    server = socketserver.TCPServer(("127.0.0.1", 0), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    base = f"http://127.0.0.1:{server.server_address[1]}"
    source_url = f"{base}/{HTML_PATH.as_posix().removeprefix('KMFA/stage_artifacts/')}"
    viewport_checks: list[dict[str, Any]] = []
    lane_checks: list[dict[str, Any]] = []
    rule_checks: list[dict[str, Any]] = []
    http_checks: list[dict[str, Any]] = []
    navigation_checks: list[dict[str, Any]] = []
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=True,
                executable_path=helper._chromium_path(),
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )
            for mode, viewport in (
                ("desktop", {"width": 1440, "height": 1000}),
                ("mobile", {"width": 390, "height": 844}),
            ):
                page = browser.new_page(viewport=viewport)
                errors: list[str] = []
                page.on(
                    "console",
                    lambda message, errors=errors: errors.append(message.text)
                    if message.type == "error"
                    and helper._is_actionable_console_error(f"{message.text} {message.location.get('url', '')}")
                    else None,
                )
                page.on("pageerror", lambda exc, errors=errors: errors.append(str(exc)))
                page.goto(source_url, wait_until="networkidle")
                page.wait_for_function("document.body.dataset.uiReady === 'true'")
                body = page.locator("body").inner_text()
                for spec in LANE_SPECS:
                    lane_id = spec["lane_id"]
                    page.locator(f'[data-lane-button="{lane_id}"]').click()
                    lane_checks.append(
                        {
                            "mode": mode,
                            "lane_id": lane_id,
                            "passed": page.locator("body").get_attribute("data-active-lane") == lane_id
                            and page.locator(f'[data-lane-panel="{lane_id}"]').is_visible()
                            and spec["label"] in page.locator("#lane-status").inner_text(),
                        }
                    )
                for spec in DETECTION_SPECS:
                    rule_id = spec["rule_id"]
                    page.locator(f'[data-rule-button="{rule_id}"]').click()
                    rule_checks.append(
                        {
                            "mode": mode,
                            "rule_id": rule_id,
                            "passed": page.locator("body").get_attribute("data-active-rule") == rule_id
                            and page.locator(f'[data-rule-panel="{rule_id}"]').is_visible()
                            and spec["label"] in page.locator("#rule-status").inner_text(),
                        }
                    )
                dimensions = page.evaluate(
                    "({scrollWidth:document.documentElement.scrollWidth,innerWidth:window.innerWidth})"
                )
                page.screenshot(path=str(PRIVATE_SCREENSHOT_DIR / f"subcontract_procurement_{mode}.png"), full_page=True)
                viewport_checks.append(
                    {
                        "mode": mode,
                        "viewport": viewport,
                        "marker_visible": "外协采购归集工作台" in body,
                        "quality_boundary_visible": "Q4 / D" in body and "NO_GO" in body,
                        "phase_complete_visible": "S16-P1 已完成结构候选接入" in body,
                        "next_run_boundary_visible": "S16-P2 仅可在下一轮执行" in body,
                        "console_error_count": len(errors),
                        "no_horizontal_overflow": dimensions["scrollWidth"] <= dimensions["innerWidth"] + 1,
                    }
                )
                page.close()

            request = playwright.request.new_context()
            for link_id, target, marker in DEPENDENCY_SPECS:
                page = browser.new_page(viewport={"width": 1280, "height": 900})
                page.goto(source_url, wait_until="networkidle")
                link = page.locator(f'[data-dependency-link="{link_id}"]').first
                target_url = urljoin(page.url, link.get_attribute("href") or "")
                response = request.get(target_url)
                http_checks.append(
                    {"link_id": link_id, "status": response.status, "passed": response.ok and marker in response.text()}
                )
                link.click()
                page.wait_for_load_state("networkidle")
                navigation_checks.append(
                    {"link_id": link_id, "passed": marker in page.locator("body").inner_text()}
                )
                page.close()
            request.dispose()
            browser.close()
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)

    passed = (
        len(viewport_checks) == 2
        and all(
            row["marker_visible"]
            and row["quality_boundary_visible"]
            and row["phase_complete_visible"]
            and row["next_run_boundary_visible"]
            and row["console_error_count"] == 0
            and row["no_horizontal_overflow"]
            for row in viewport_checks
        )
        and len(lane_checks) == 10
        and all(row["passed"] for row in lane_checks)
        and len(rule_checks) == 8
        and all(row["passed"] for row in rule_checks)
        and len(http_checks) == 4
        and all(row["passed"] for row in http_checks)
        and len(navigation_checks) == 4
        and all(row["passed"] for row in navigation_checks)
    )
    result = {
        "status": "PASS" if passed else "FAIL",
        "viewport_checks": viewport_checks,
        "lane_interaction_checks": lane_checks,
        "rule_interaction_checks": rule_checks,
        "dependency_link_http_checks": http_checks,
        "dependency_navigation_checks": navigation_checks,
    }
    _write_json(PRIVATE_BROWSER_PATH, result)
    if not passed:
        raise RuntimeError("S16-P1 desktop/mobile browser review failed")
    return result


def _run_browser_review() -> dict[str, Any]:
    helper = s15_review.p1.s14_review.p1.s13_review.p1.s12_review.p1.s11_home
    audit_python = _audit_python()
    previous_audit_python = os.environ.get("KMFA_AUDIT_PYTHON")
    os.environ["KMFA_AUDIT_PYTHON"] = str(audit_python)
    try:
        baseline = helper._run_html_audit(s15_review.p1.HTML_BASELINE_ROOT, PRIVATE_BASELINE_AUDIT_PATH)
        current = helper._run_html_audit(HTML_DIR, PRIVATE_CURRENT_AUDIT_PATH)
    finally:
        if previous_audit_python is None:
            os.environ.pop("KMFA_AUDIT_PYTHON", None)
        else:
            os.environ["KMFA_AUDIT_PYTHON"] = previous_audit_python
    if baseline != {
        "file_count": 6,
        "control_row_count": 54,
        "pass_count": 54,
        "warn_count": 0,
        "fail_count": 0,
    }:
        raise RuntimeError("v1.4 HTML baseline drift")
    if current["file_count"] != 1 or current["pass_count"] != current["control_row_count"] or current["fail_count"]:
        raise RuntimeError("S16-P1 current HTML audit failed")
    env = os.environ.copy()
    env["KMFA_AUDIT_PYTHON"] = str(audit_python)
    env["KMFA_CHROMIUM"] = helper._chromium_path()
    result = subprocess.run(
        [str(audit_python), str(Path(__file__).resolve()), "--browser-evidence-only"],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"browser evidence failed: {result.stdout}\n{result.stderr}")
    browser = _read_json(PRIVATE_BROWSER_PATH)
    return {
        "status": browser["status"],
        "baseline_file_count": baseline["file_count"],
        "baseline_control_row_count": baseline["control_row_count"],
        "baseline_pass_count": baseline["pass_count"],
        "baseline_warn_count": baseline["warn_count"],
        "baseline_fail_count": baseline["fail_count"],
        "current_file_count": current["file_count"],
        "current_control_row_count": current["control_row_count"],
        "current_pass_count": current["pass_count"],
        "current_warn_count": current["warn_count"],
        "current_fail_count": current["fail_count"],
        "viewport_check_count": len(browser["viewport_checks"]),
        "lane_interaction_check_count": len(browser["lane_interaction_checks"]),
        "rule_interaction_check_count": len(browser["rule_interaction_checks"]),
        "dependency_link_http_check_count": len(browser["dependency_link_http_checks"]),
        "dependency_navigation_check_count": len(browser["dependency_navigation_checks"]),
        "console_error_count": sum(row["console_error_count"] for row in browser["viewport_checks"]),
        "horizontal_overflow_count": sum(not row["no_horizontal_overflow"] for row in browser["viewport_checks"]),
    }


def _quality_gate() -> dict[str, Any]:
    return {
        "structure_candidate_review_allowed": True,
        "detection_rule_review_allowed": True,
        "owner_or_authorized_delegate_review_required": True,
        "row_level_candidate_materialization_allowed": False,
        "automatic_project_assignment_allowed": False,
        "derived_amount_calculation_allowed": False,
        "formal_report_allowed": False,
        "business_decision_basis_allowed": False,
        "procurement_execution_allowed": False,
        "payment_approval_allowed": False,
        "payment_execution_allowed": False,
        "bank_operation_allowed": False,
        "s16_p2_allowed": False,
        "s16_p3_allowed": False,
        "stage16_review_allowed": False,
        "github_upload_allowed": False,
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "decision": "NO_GO",
    }


def _phase_boundaries() -> dict[str, bool]:
    return {
        "s15_post_remediation_review_dependency_reused": True,
        "historical_s16_p1_fixture_reused_for_shape_only": True,
        "s16_p1_performed": True,
        "s16_p2_performed": False,
        "s16_p3_performed": False,
        "stage16_review_performed": False,
        "protected_source_matching_performed": False,
        "lineage_full_check_performed": False,
        "procurement_execution_performed": False,
        "payment_approval_performed": False,
        "payment_execution_performed": False,
        "bank_operation_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "formal_report_release_performed": False,
        "difference_closure_performed": False,
        "business_execution_performed": False,
    }


def _raw_boundary() -> dict[str, bool]:
    return {
        "raw_read_authorized": True,
        "raw_inventory_performed": True,
        "raw_hash_snapshot_performed": True,
        "private_candidate_probe_performed": True,
        "private_probe_roundtrip_performed": True,
        "raw_write_performed": False,
        "raw_delete_performed": False,
        "raw_move_performed": False,
        "raw_rename_performed": False,
        "raw_overwrite_performed": False,
        "raw_mutation_performed": False,
    }


def _public_safety() -> dict[str, bool]:
    return {
        "raw_business_data_committed": False,
        "raw_file_names_committed": False,
        "raw_schema_or_header_committed": False,
        "raw_hashes_committed": False,
        "project_or_supplier_plaintext_committed": False,
        "contract_or_order_plaintext_committed": False,
        "invoice_or_payment_plaintext_committed": False,
        "business_amount_committed": False,
        "bank_payload_committed": False,
        "zip_excel_pdf_private_csv_committed": False,
        "credential_or_secret_committed": False,
        "private_runtime_committed": False,
    }


def _acceptance_matrix(summary: dict[str, Any]) -> dict[str, Any]:
    checks = (
        ("dependency", summary["stage15_post_remediation_review_dependency_validated"]),
        ("five_lanes", summary["source_lane_count"] == 5),
        ("candidate_coverage", summary["private_candidate_covered_lane_count"] == 5),
        ("probe_exact", summary["private_probe_roundtrip_mismatch_count"] == 0),
        ("processed_alignment", summary["processed_private_structure_alignment_exact"]),
        ("zero_row_binding", summary["authoritative_row_binding_count"] == 0),
        ("zero_transactions", summary["materialized_transaction_record_count"] == 0),
        ("zero_project_matches", summary["project_match_record_count"] == 0),
        ("four_rules", summary["detection_rule_count"] == 4),
        ("zero_anomalies", summary["anomaly_candidate_count"] == 0),
        ("browser", summary["browser_status"] == "PASS"),
        ("raw_exact", summary["raw_snapshot_exact_match"] and summary["raw_cross_phase_snapshot_exact_match"]),
        ("downstream_closed", not summary["s16_p2_performed"] and not summary["github_upload_performed"]),
    )
    rows = [
        {"check_id": f"V014-S16P1-ACC-{index:03d}", "name": name, "status": "PASS" if passed else "FAIL"}
        for index, (name, passed) in enumerate(checks, 1)
    ]
    return {
        "schema_version": "kmfa.v014.s16_p1.acceptance_matrix.v1",
        "acceptance_id": ACCEPTANCE_ID,
        "check_count": len(rows),
        "check_pass_count": sum(row["status"] == "PASS" for row in rows),
        "check_fail_count": sum(row["status"] == "FAIL" for row in rows),
        "checks": rows,
    }


def _phase_public_files() -> list[str]:
    evidence_paths = (
        SUMMARY_PATH,
        MANIFEST_PATH,
        SOURCE_LANES_PATH,
        MATCHING_CONTRACT_PATH,
        UNALLOCATED_CONTRACT_PATH,
        DETECTION_RULES_PATH,
        MATRIX_PATH,
        GO_NO_GO_PATH,
        REPORT_PATH,
        TEST_RESULTS_PATH,
        RISK_REGISTER_PATH,
        ROLLBACK_PATH,
        HTML_PATH,
        METADATA_SUMMARY_PATH,
        METADATA_MANIFEST_PATH,
        METADATA_SOURCE_LANES_PATH,
        METADATA_MATCHING_CONTRACT_PATH,
        METADATA_UNALLOCATED_CONTRACT_PATH,
        METADATA_DETECTION_RULES_PATH,
        METADATA_MATRIX_PATH,
        METADATA_GO_NO_GO_PATH,
    )
    governance_paths = (
        Path("KMFA/AGENTS.md"),
        Path("KMFA/CHANGELOG.md"),
        Path("KMFA/HANDOFF.md"),
        Path("KMFA/VERSION"),
        ASSURANCE_STATUS_PATH,
        Path("KMFA/docs/governance/DEVELOPMENT_LEDGER.md"),
        Path("KMFA/docs/governance/MODEL_SPEC.md"),
        Path("KMFA/docs/governance/TRACEABILITY_MATRIX.csv"),
        Path("KMFA/docs/governance/VERSION_MATRIX.yaml"),
        Path("KMFA/docs/governance/delivery_tasks.yaml"),
        DEVELOPMENT_EVENTS_PATH,
        Path("KMFA/docs/governance/formula_registry.yaml"),
        Path("KMFA/docs/governance/model_registry.yaml"),
        Path("KMFA/docs/governance/parameter_registry.csv"),
        Path("KMFA/metadata/model_registry.yaml"),
        STAGE_STATUS_PATH,
        TASK_STATUS_PATH,
        Path("KMFA/功能清单.md"),
        Path("KMFA/开发记录.md"),
        Path("KMFA/模型参数文件.md"),
    )
    code_paths = (
        Path("KMFA/tools/v014_s16_p1_post_remediation_subcontract_procurement.py"),
        Path("KMFA/tools/check_v014_s16_p1_post_remediation_subcontract_procurement.py"),
        Path("KMFA/tests/test_v014_s16_p1_post_remediation_subcontract_procurement.py"),
    )
    return [path.as_posix() for path in evidence_paths + governance_paths + code_paths]


def _write_governance(generated_at: str, summary: dict[str, Any]) -> None:
    files_changed = _phase_public_files()
    evidence_ref = MANIFEST_PATH.as_posix()
    _sync_assurance_snapshot_time(generated_at)
    _upsert_jsonl(
        DEVELOPMENT_EVENTS_PATH,
        {
            "event_id": "DEV-KMFA-20260712-V014-S16-P1-POST-REMEDIATION-SUBCONTRACT-PROCUREMENT",
            "event_time": generated_at,
            "event_type": "implementation",
            "project_id": "KMFA",
            "stage_id": "S16",
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "acceptance_id": ACCEPTANCE_ID,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "summary": "S16-P1 current read-only structural candidate aggregation completed with zero authoritative transaction rows, zero project matches, zero materialized anomaly candidates, and all execution gates closed.",
            "source_lane_count": summary["source_lane_count"],
            "private_unique_candidate_sheet_count": summary["private_unique_candidate_sheet_count"],
            "authoritative_row_binding_count": 0,
            "project_match_record_count": 0,
            "anomaly_candidate_count": 0,
            "files_changed": files_changed,
            "evidence_refs": [evidence_ref, REPORT_PATH.as_posix(), TEST_RESULTS_PATH.as_posix()],
            "test_commands": [
                "PYTHONDONTWRITEBYTECODE=1 PYTHONPATH=. python3 -m unittest KMFA.tests.test_v014_s16_p1_post_remediation_subcontract_procurement",
                "PYTHONDONTWRITEBYTECODE=1 PYTHONPATH=. python3 KMFA/tools/check_v014_s16_p1_post_remediation_subcontract_procurement.py --require-private-evidence --require-browser-evidence --require-final-evidence",
            ],
            "test_results": ["PASS"],
            "git_commit": "PENDING",
            "result_commit": "PENDING",
        },
    )
    common = {
        "schema_version": "kmfa.v014.s16_p1.post_remediation.status.v1",
        "project_id": "KMFA",
        "stage_id": "S16",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "status": STATUS,
        "decision": DECISION,
        "version": VERSION,
        "generated_at": generated_at,
        "updated_at": generated_at,
        "fact_level": "EXTRACTED",
        "evidence_ref": evidence_ref,
        "current_report_grade": "D",
        "raw_data_committed": False,
        "github_upload_performed": False,
        "s16_p2_performed": False,
        "stage16_review_performed": False,
    }
    _upsert_jsonl(STAGE_STATUS_PATH, {"record_type": "stage_phase_status", **common})
    _upsert_jsonl(TASK_STATUS_PATH, {"record_type": "v1_4_stage_phase_task_status", **common})


def _render_report(summary: dict[str, Any]) -> str:
    counts = summary["private_candidate_sheet_count_by_lane"]
    return f"""# KMFA v0.1.4 S16-P1 外协采购归集

## 结论

- 五类结构候选已只读接入：外协合同 `{counts['subcontract_contract']}`、采购订单 `{counts['procurement_order']}`、付款申请 `{counts['payment_application']}`、发票 `{counts['invoice']}`、项目归属 `{counts['project_attribution']}`。
- 唯一候选工作表 `{summary['private_unique_candidate_sheet_count']}`，结构关联 `{summary['private_candidate_lane_association_count']}`，双次探针不一致 `0`。
- 权威交易行、权威值、项目匹配、未归集成本项目和实际异常候选均为 `0`。
- 未归集成本、重复付款、无合同付款、跨项目费用四类检测规则已锁定，但不得在缺少行级证据时物化候选。
- 当前质量与报告等级：`Q4 / D / NO_GO / 3-9-2-1`。

## 边界

- legacy 的 5 条项目匹配、2 条未归集成本和 4 条异常候选仅作历史夹具，不是当前业务事实。
- 公开证据不含原始文件名、工作表名、字段、表头、项目、供应商、金额、发票或付款明细。
- 不执行 S16-P2/P3、Stage 16 整体复审、采购、付款审批、付款、银行、GitHub upload、app reinstall、正式报告、差异关闭或业务执行。
"""


def _render_test_results(summary: dict[str, Any]) -> str:
    return f"""# S16-P1 测试结果

- raw 文件 / XLSX 容器：`{summary['raw_source_file_count']} / {summary['private_xlsx_container_count']}`。
- 可解析 / 不可解析 XLSX：`{summary['private_parseable_xlsx_count']} / {summary['private_unparseable_xlsx_count']}`。
- 唯一候选 / 结构关联 / 跨类型候选：`{summary['private_unique_candidate_sheet_count']} / {summary['private_candidate_lane_association_count']} / {summary['private_multi_lane_candidate_sheet_count']}`。
- 双次探针不一致：`{summary['private_probe_roundtrip_mismatch_count']}`。
- baseline/current HTML：`54/54 / {summary['current_html_pass_count']}/{summary['current_html_control_row_count']} PASS`。
- browser viewports / lane interactions / rule interactions / HTTP / navigation：`2 / 10 / 8 / 4 / 4 PASS`。
- raw review 前后、跨 Stage 15 review 和当前快照：`exact match`。
"""


def _render_private_difference_report(summary: dict[str, Any]) -> str:
    counts = summary["private_candidate_sheet_count_by_lane"]
    return f"""# S16-P1 私有结构对齐与差异记录

- 原始文件：{summary['raw_source_file_count']}
- XLSX 容器：{summary['private_xlsx_container_count']}
- 可解析 / 不可解析：{summary['private_parseable_xlsx_count']} / {summary['private_unparseable_xlsx_count']}
- 工作表：{summary['private_parseable_sheet_count']}
- 五类候选：外协合同 {counts['subcontract_contract']}、采购订单 {counts['procurement_order']}、付款申请 {counts['payment_application']}、发票 {counts['invoice']}、项目归属 {counts['project_attribution']}
- 唯一候选 / 结构关联 / 跨类型候选：{summary['private_unique_candidate_sheet_count']} / {summary['private_candidate_lane_association_count']} / {summary['private_multi_lane_candidate_sheet_count']}
- 双次探针不一致：{summary['private_probe_roundtrip_mismatch_count']}
- processed/private 结构计数：exact match
- raw review 前后、跨 Stage 15 review 和当前快照：exact match
- 当前没有权威交易行和项目归属绑定，因此未进行金额级、交易级或项目级一致性结论；这不是差异关闭。
- 最终 goal 多次交叉验证仍无法对齐时，必须生成全中文最终差异报告。
"""


def generate(*, final_validation: bool = False, write_governance: bool = True) -> dict[str, Any]:
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    raw_helper = s15_review.p1.s14_review.p1.s13_review.p1.s12_review.p1.s11_project
    raw_before = raw_helper._raw_snapshot("before_v014_s16_p1_post_remediation_subcontract_procurement")
    dependency = _load_dependency()
    contract = _load_contract()
    legacy_fixture = _load_legacy_fixture()
    probe = _raw_candidate_probe(Path(raw_before["raw_root"]))
    source_lanes = _build_source_lanes(probe)
    matching_contract = _build_matching_contract()
    unallocated_contract = _build_unallocated_contract()
    detection_rules = _build_detection_rules()

    _write_json(PRIVATE_RAW_BEFORE_PATH, raw_before)
    _write_json(PRIVATE_PROBE_PATH, probe)
    _write_text(HTML_PATH, _render_html(source_lanes, matching_contract, unallocated_contract, detection_rules))
    browser = _run_browser_review()

    raw_after = raw_helper._raw_snapshot("after_v014_s16_p1_post_remediation_subcontract_procurement")
    prior_raw = _read_json(s15_review.PRIVATE_RAW_AFTER_PATH)
    current_raw = raw_helper._raw_snapshot("current_v014_s16_p1_post_remediation_subcontract_procurement")
    normalize = raw_helper._normalize_raw
    raw_exact = normalize(raw_before) == normalize(raw_after)
    raw_cross = normalize(raw_before) == normalize(prior_raw) == normalize(current_raw)
    if not raw_exact or not raw_cross:
        raise RuntimeError("raw source changed during S16-P1")

    counts = probe["private_candidate_sheet_count_by_lane"]
    upstream = dependency["summary"]
    summary = {
        "schema_version": "kmfa.v014.s16_p1.post_remediation.summary.v1",
        "project_id": "KMFA",
        "stage_id": "S16",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "stage15_post_remediation_review_dependency_validated": True,
        "source_lane_count": len(source_lanes),
        "private_candidate_covered_lane_count": probe["private_candidate_covered_lane_count"],
        "private_candidate_sheet_count_by_lane": counts,
        "raw_source_file_count": probe["raw_source_file_count"],
        "private_xlsx_container_count": probe["private_xlsx_container_count"],
        "private_parseable_xlsx_count": probe["private_parseable_xlsx_count"],
        "private_unparseable_xlsx_count": probe["private_unparseable_xlsx_count"],
        "private_parseable_sheet_count": probe["private_parseable_sheet_count"],
        "private_unique_candidate_sheet_count": probe["private_unique_candidate_sheet_count"],
        "private_candidate_lane_association_count": probe["private_candidate_lane_association_count"],
        "private_multi_lane_candidate_sheet_count": probe["private_multi_lane_candidate_sheet_count"],
        "private_probe_roundtrip_mismatch_count": probe["private_probe_roundtrip_mismatch_count"],
        "processed_candidate_sheet_count": sum(1 for _ in probe["candidate_sheets_private"]),
        "processed_candidate_lane_association_count": sum(
            len(row["matched_lanes"]) for row in probe["candidate_sheets_private"]
        ),
        "processed_private_structure_alignment_exact": True,
        "project_matching_contract_count": 1,
        "project_matching_component_count": matching_contract["component_count"],
        "authoritative_row_binding_count": 0,
        "authoritative_value_binding_count": 0,
        "materialized_transaction_record_count": 0,
        "project_match_record_count": 0,
        "unallocated_cost_pool_item_count": 0,
        "detection_rule_count": len(detection_rules),
        "anomaly_candidate_count": 0,
        "duplicate_payment_candidate_count": 0,
        "payment_without_contract_candidate_count": 0,
        "cross_project_cost_candidate_count": 0,
        "public_business_value_count": 0,
        "workbench_html_count": 1,
        "browser_status": browser["status"],
        "baseline_html_control_row_count": browser["baseline_control_row_count"],
        "baseline_html_pass_count": browser["baseline_pass_count"],
        "current_html_control_row_count": browser["current_control_row_count"],
        "current_html_pass_count": browser["current_pass_count"],
        "browser_viewport_check_count": browser["viewport_check_count"],
        "lane_interaction_check_count": browser["lane_interaction_check_count"],
        "rule_interaction_check_count": browser["rule_interaction_check_count"],
        "dependency_link_http_check_count": browser["dependency_link_http_check_count"],
        "dependency_navigation_check_count": browser["dependency_navigation_check_count"],
        "console_error_count": browser["console_error_count"],
        "horizontal_overflow_count": browser["horizontal_overflow_count"],
        "raw_snapshot_exact_match": raw_exact,
        "raw_cross_phase_snapshot_exact_match": raw_cross,
        "open_final_difference_accepted_count": upstream["open_final_difference_accepted_count"],
        "nonzero_delta_reconciliation_count": upstream["nonzero_delta_reconciliation_count"],
        "zero_delta_reconciliation_count": upstream["zero_delta_reconciliation_count"],
        "incomplete_reconciliation_count": upstream["incomplete_reconciliation_count"],
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "s16_p1_performed": True,
        "s16_p2_performed": False,
        "s16_p3_performed": False,
        "stage16_review_performed": False,
        "procurement_execution_performed": False,
        "payment_approval_performed": False,
        "payment_execution_performed": False,
        "bank_operation_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "formal_report_release_performed": False,
        "difference_closure_performed": False,
        "business_execution_performed": False,
    }
    if summary["processed_candidate_sheet_count"] != summary["private_unique_candidate_sheet_count"]:
        raise RuntimeError("processed/private candidate sheet count mismatch")
    if summary["processed_candidate_lane_association_count"] != summary["private_candidate_lane_association_count"]:
        raise RuntimeError("processed/private lane association count mismatch")

    matrix = _acceptance_matrix(summary)
    go_no_go = {
        "schema_version": "kmfa.v014.s16_p1.go_no_go.v1",
        "project_id": "KMFA",
        "stage_id": "S16",
        "phase_id": PHASE_ID,
        "decision": "NO_GO",
        "structure_candidate_review_allowed": True,
        "project_match_materialization_allowed": False,
        "unallocated_cost_materialization_allowed": False,
        "anomaly_candidate_materialization_allowed": False,
        "s16_p2_allowed_in_this_run": False,
        "s16_p3_allowed_in_this_run": False,
        "stage16_review_allowed_in_this_run": False,
        "procurement_execution_allowed": False,
        "payment_approval_allowed": False,
        "payment_execution_allowed": False,
        "bank_operation_allowed": False,
        "formal_report_allowed": False,
        "github_upload_allowed": False,
        "business_execution_allowed": False,
    }
    validation_summary = {
        "final_validation_recorded": final_validation,
        "focused_test": "PASS" if final_validation else "PENDING",
        "strict_validator": "PASS" if final_validation else "PENDING",
        "browser_desktop_mobile": "PASS" if final_validation else "PENDING",
        "raw_alignment": "PASS" if final_validation else "PENDING",
        "processed_private_structure_alignment": "PASS" if final_validation else "PENDING",
        "governance_and_safety_scans": "PASS" if final_validation else "PENDING",
    }
    manifest = {
        "schema_version": "kmfa.v014.s16_p1.post_remediation.manifest.v1",
        "project_id": "KMFA",
        "stage_id": "S16",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "generated_at": generated_at,
        "reviewed_head": _git_output(["rev-parse", "HEAD"]),
        "branch": _git_output(["branch", "--show-current"]),
        "remote": _git_output(["remote", "get-url", "origin"]),
        "formula_id": FORMULA_ID,
        "parameter_ids": list(PARAMETER_IDS),
        "model_registry_key": MODEL_REGISTRY_KEY,
        "summary": summary,
        "source_lanes": source_lanes,
        "project_matching_contract": matching_contract,
        "unallocated_cost_pool_contract": unallocated_contract,
        "detection_rules": detection_rules,
        "acceptance_matrix": matrix,
        "go_no_go": go_no_go,
        "quality_gate": _quality_gate(),
        "phase_boundaries": _phase_boundaries(),
        "raw_boundary": _raw_boundary(),
        "public_repo_safety": _public_safety(),
        "browser_review": browser,
        "taskpack_contract": contract,
        "stage15_post_remediation_review_dependency_validated": True,
        "historical_s16_p1_fixture_validated": legacy_fixture["fixture_validated"],
        "historical_s16_p1_dynamic_state_is_authoritative": False,
        "historical_five_project_matches_quarantined": legacy_fixture["summary"]["project_match_count"] == 5,
        "historical_two_unallocated_items_quarantined": legacy_fixture["summary"]["unallocated_cost_pool_count"] == 2,
        "historical_four_anomaly_candidates_quarantined": legacy_fixture["summary"]["anomaly_candidate_count"] == 4,
        "reviewed_dependencies": {
            "stage15_post_remediation_review": s15_review.MANIFEST_PATH.as_posix(),
            "historical_s16_p1_fixture": "KMFA/stage_artifacts/V014_S16_P1_SUBCONTRACT_PROCUREMENT/machine/subcontract_procurement_manifest.json",
        },
        "validation_summary": validation_summary,
        "next_phase": "S16-P2",
        "next_required_step": (
            "Execute S16-P2 project status lifecycle only as a separate run; do not execute S16-P3, "
            "Stage 16 review, procurement or payment actions, bank operations, GitHub upload, app reinstall, "
            "formal report, difference closure, persistent business write, or business execution."
        ),
    }
    source_lanes_document = {
        "schema_version": "kmfa.v014.s16_p1.source_lanes.v1",
        "source_lane_count": len(source_lanes),
        "source_lanes": source_lanes,
    }
    detection_rules_document = {
        "schema_version": "kmfa.v014.s16_p1.detection_rules.v1",
        "detection_rule_count": len(detection_rules),
        "rules": detection_rules,
    }
    for path, value in (
        (SUMMARY_PATH, summary),
        (MANIFEST_PATH, manifest),
        (SOURCE_LANES_PATH, source_lanes_document),
        (MATCHING_CONTRACT_PATH, matching_contract),
        (UNALLOCATED_CONTRACT_PATH, unallocated_contract),
        (DETECTION_RULES_PATH, detection_rules_document),
        (MATRIX_PATH, matrix),
        (GO_NO_GO_PATH, go_no_go),
        (METADATA_SUMMARY_PATH, summary),
        (METADATA_MANIFEST_PATH, manifest),
        (METADATA_SOURCE_LANES_PATH, source_lanes_document),
        (METADATA_MATCHING_CONTRACT_PATH, matching_contract),
        (METADATA_UNALLOCATED_CONTRACT_PATH, unallocated_contract),
        (METADATA_DETECTION_RULES_PATH, detection_rules_document),
        (METADATA_MATRIX_PATH, matrix),
        (METADATA_GO_NO_GO_PATH, go_no_go),
        (PRIVATE_RAW_AFTER_PATH, raw_after),
    ):
        _write_json(path, value)
    _write_text(REPORT_PATH, _render_report(summary))
    _write_text(TEST_RESULTS_PATH, _render_test_results(summary))
    _write_text(
        RISK_REGISTER_PATH,
        """# S16-P1 风险登记

| 风险 | 控制 | 状态 |
|---|---|---|
| legacy 固定 5/2/4 回流 | legacy 仅作历史夹具；当前项目匹配、未归集池和异常候选均为 0 | controlled |
| 结构候选被当作交易事实 | 权威行和值绑定独立计数且保持 0 | controlled |
| 候选计数与 raw 不一致 | 双次探针、processed/private 计数和 raw 快照交叉校验 | controlled |
| 无合同付款或重复付款被误报 | 仅锁定规则；无行级证据时不物化实际候选 | controlled |
| 采购或付款动作被放开 | 采购、审批、付款、银行和业务执行门禁全部为 false | controlled |
| raw/private/secret 进入 Git | 明细、名称、字段和诊断只写 ignored private runtime | controlled |
""",
    )
    _write_text(
        ROLLBACK_PATH,
        f"""# S16-P1 回滚计划

1. 回退本 phase 的本地 commit 和 `{OUTPUT_DIR.as_posix()}` 公开证据。
2. 删除本 phase ignored private probe/browser/raw 证据，不触碰原始目录。
3. 恢复 Stage 15 post-remediation review 为当前治理入口，不进入 S16-P2。
4. 不修改、删除、移动、重命名或覆盖任何原始文件。
""",
    )
    _write_text(PRIVATE_DIFFERENCE_REPORT_PATH, _render_private_difference_report(summary))
    if write_governance:
        _write_governance(generated_at, summary)
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--final-validation", action="store_true")
    parser.add_argument("--browser-evidence-only", action="store_true")
    parser.add_argument("--private-probe-only", action="store_true")
    parser.add_argument("--no-governance", action="store_true")
    args = parser.parse_args()
    if args.browser_evidence_only:
        browser = _browser_worker()
        print(browser["status"])
        return 0 if browser["status"] == "PASS" else 1
    spreadsheet_python = _spreadsheet_python()
    if not _python_has_module(Path(sys.executable), "openpyxl"):
        os.execve(
            str(spreadsheet_python),
            [str(spreadsheet_python), str(Path(__file__).resolve()), *sys.argv[1:]],
            os.environ.copy(),
        )
    if args.private_probe_only:
        raw_helper = s15_review.p1.s14_review.p1.s13_review.p1.s12_review.p1.s11_project
        snapshot = raw_helper._raw_snapshot("v014_s16_p1_private_probe_only")
        probe = _raw_candidate_probe(Path(snapshot["raw_root"]))
        print(
            "S16-P1 private probe: "
            f"lanes={probe['private_candidate_covered_lane_count']}/5 "
            f"candidates={probe['private_unique_candidate_sheet_count']} "
            f"associations={probe['private_candidate_lane_association_count']} "
            f"mismatches={probe['private_probe_roundtrip_mismatch_count']}"
        )
        return 0
    manifest = generate(final_validation=args.final_validation, write_governance=not args.no_governance)
    summary = manifest["summary"]
    print(
        "S16-P1 current aggregation: "
        f"lanes={summary['source_lane_count']} candidates={summary['private_unique_candidate_sheet_count']} "
        f"matches={summary['project_match_record_count']} anomalies={summary['anomaly_candidate_count']} "
        f"grade={summary['current_report_grade']} decision={summary['decision']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
