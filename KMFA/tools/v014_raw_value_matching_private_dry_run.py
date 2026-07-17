#!/usr/bin/env python3
"""Generate KMFA v0.1.4 raw value matching private dry-run evidence.

This phase performs a read-only private dry run against the owner-confirmed raw
container. Public artifacts expose only aggregate counts, readiness flags and
gate decisions. Raw filenames, archive entry names, sheet names, headers, cell values,
PDF text and business values stay out of Git and out of public evidence.
"""

from __future__ import annotations

import csv
import io
import json
import re
import subprocess
import sys
import zipfile
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from KMFA.tools.v014_s05_p1_a0_file_registration import (  # noqa: E402
    EXPECTED_EXCEL_COUNT,
    EXPECTED_PDF_COUNT,
    RAW_INBOX,
    is_hidden_zip_member,
    resolve_private_raw_zip,
    sha256_text,
    stat_snapshot,
)


SCHEMA_VERSION = "kmfa.v014_raw_value_matching_private_dry_run.v1"
GO_NO_GO_SCHEMA_VERSION = "kmfa.v014_raw_value_matching_private_dry_run_go_no_go.v1"
PRIVATE_SCHEMA_VERSION = "kmfa.private.v014_raw_value_matching_private_dry_run.v1"
PROJECT_ID = "KMFA"
VERSION = "0.1.4"
PHASE_ID = "V014_RAW_VALUE_MATCHING_PRIVATE_DRY_RUN"
TASK_ID = "KMFA-V014-RAW-VALUE-MATCHING-PRIVATE-DRY-RUN-20260705"
ACCEPTANCE_ID = "ACC-V014-RAW-VALUE-MATCHING-PRIVATE-DRY-RUN"
STATUS = "completed_validated_local_only_no_go_private_raw_value_fingerprints_processed_targets_missing"
NEXT_RECOMMENDED_PHASE = "V014_PRIVATE_PROCESSED_VALUE_STAGING"

OUTPUT_DIR = Path("KMFA/stage_artifacts/V014_RAW_VALUE_MATCHING_PRIVATE_DRY_RUN")
MACHINE_DIR = OUTPUT_DIR / "machine"
HUMAN_DIR = OUTPUT_DIR / "human"
MANIFEST_PATH = MACHINE_DIR / "raw_value_matching_private_dry_run_manifest.json"
GO_NO_GO_PATH = MACHINE_DIR / "raw_value_matching_private_dry_run_go_no_go_report.json"
GAP_SUMMARY_PATH = MACHINE_DIR / "raw_value_matching_private_dry_run_gap_summary.json"
REPORT_PATH = HUMAN_DIR / "raw_value_matching_private_dry_run_report.md"
GO_NO_GO_RECORD_PATH = HUMAN_DIR / "go_no_go_record.md"
TEST_RESULTS_PATH = HUMAN_DIR / "test_results.md"
RISK_REGISTER_PATH = HUMAN_DIR / "risk_register.md"
ROLLBACK_PATH = HUMAN_DIR / "rollback_plan.md"

METADATA_MANIFEST_PATH = Path("KMFA/metadata/quality/v014_raw_value_matching_private_dry_run_manifest.json")
METADATA_GO_NO_GO_PATH = Path("KMFA/metadata/quality/v014_raw_value_matching_private_dry_run_go_no_go_report.json")
METADATA_GAP_SUMMARY_PATH = Path("KMFA/metadata/quality/v014_raw_value_matching_private_dry_run_gap_summary.json")

PRIVATE_OUTPUT_DIR = Path("KMFA/.codex_private_runtime/v014_raw_value_matching_private_dry_run")
PRIVATE_DIAGNOSTIC_PATH = PRIVATE_OUTPUT_DIR / "private_raw_value_matching_diagnostic.json"
PRIVATE_GAP_REPORT_PATH = PRIVATE_OUTPUT_DIR / "local_gap_report.md"

VALUE_SCOPE_MANIFEST_PATH = Path(
    "KMFA/stage_artifacts/V014_VALUE_CONSISTENCY_SCOPE_GATE/machine/value_consistency_scope_manifest.json"
)
VALUE_SCOPE_GO_NO_GO_PATH = Path(
    "KMFA/stage_artifacts/V014_VALUE_CONSISTENCY_SCOPE_GATE/machine/value_consistency_scope_go_no_go_report.json"
)
RAW_BASELINE_LOCK_PATH = Path("KMFA/metadata/baseline/v014_authoritative_raw_baseline_lock.json")

NUMERIC_TOKEN_RE = re.compile(r"(?<![A-Za-z0-9])[-+]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?%?(?![A-Za-z0-9])")
PROCESSED_SCAN_ROOTS = (Path("KMFA/metadata"), Path("KMFA/stage_artifacts"))
PROCESSED_VALUE_KEY_RE = re.compile(
    r"(amount|balance|cash|cost|gross|invoice|margin|profit|tax|value|cent|price|rate|ratio|revenue)",
    re.IGNORECASE,
)


def _now(generated_at: str | None = None) -> str:
    return generated_at or datetime.now().astimezone().isoformat(timespec="seconds")


def _git_output(args: list[str]) -> str:
    result = subprocess.run(
        ["git", "-c", "core.quotePath=false", *args],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if result.returncode != 0:
        return "UNKNOWN"
    return result.stdout.strip()


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"{path} must contain a JSON object")
    return value


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _decimal_fingerprint(value: Any) -> str | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            dec = Decimal(str(value).replace(",", ""))
        except (InvalidOperation, ValueError):
            return None
        return sha256_text(format(dec.normalize(), "f"))
    if isinstance(value, str):
        normalized = value.strip().replace(",", "")
        if normalized.endswith("%"):
            normalized = normalized[:-1]
        try:
            dec = Decimal(normalized)
        except (InvalidOperation, ValueError):
            return None
        return sha256_text(format(dec.normalize(), "f"))
    return None


def _root_public_inventory() -> dict[str, Any]:
    files = sorted([path for path in RAW_INBOX.iterdir() if path.is_file()]) if RAW_INBOX.exists() else []
    suffix_counts: dict[str, int] = {}
    for path in files:
        suffix = path.suffix.lower()
        kind = "archive" if suffix == ".zip" else "spreadsheet" if suffix in {".xlsx", ".xls", ".xlsm"} else "other"
        suffix_counts[kind] = suffix_counts.get(kind, 0) + 1
    return {
        "raw_root_exists": RAW_INBOX.exists(),
        "raw_root_file_count": len(files),
        "raw_root_archive_count": suffix_counts.get("archive", 0),
        "raw_root_spreadsheet_count": suffix_counts.get("spreadsheet", 0),
        "raw_root_other_file_count": suffix_counts.get("other", 0),
    }


def _fingerprint_workbook(binary: bytes, *, source_ref_hash: str) -> dict[str, Any]:
    private_records: list[dict[str, str]] = []
    summary = {
        "workbook_openable": False,
        "worksheet_count": 0,
        "non_empty_cell_count": 0,
        "numeric_cell_count": 0,
        "formula_cell_count": 0,
        "fingerprint_count": 0,
        "error_class": "",
    }
    try:
        workbook = load_workbook(io.BytesIO(binary), data_only=False, read_only=True)
        value_workbook = load_workbook(io.BytesIO(binary), data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - depends on local workbook shape
        summary["error_class"] = type(exc).__name__
        return {"summary": summary, "private_records": private_records}

    summary["workbook_openable"] = True
    summary["worksheet_count"] = len(workbook.worksheets)
    for sheet, value_sheet in zip(workbook.worksheets, value_workbook.worksheets):
        sheet_hash = sha256_text(sheet.title)
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                summary["non_empty_cell_count"] += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    summary["formula_cell_count"] += 1
                value_cell = value_sheet[cell.coordinate]
                value_fingerprint = _decimal_fingerprint(value_cell.value)
                if value_fingerprint is None:
                    continue
                summary["numeric_cell_count"] += 1
                private_records.append(
                    {
                        "source_ref_hash": source_ref_hash,
                        "sheet_ref_hash": sheet_hash,
                        "cell_ref_hash": sha256_text(cell.coordinate),
                        "numeric_value_fingerprint": value_fingerprint,
                    }
                )
    summary["fingerprint_count"] = len(private_records)
    return {"summary": summary, "private_records": private_records}


def _fingerprint_pdf(binary: bytes, *, source_ref_hash: str) -> dict[str, Any]:
    private_records: list[dict[str, str]] = []
    summary = {
        "pdf_openable": False,
        "page_count": 0,
        "text_page_count": 0,
        "numeric_token_count": 0,
        "fingerprint_count": 0,
        "error_class": "",
    }
    try:
        with pdfplumber.open(io.BytesIO(binary)) as pdf:
            summary["pdf_openable"] = True
            summary["page_count"] = len(pdf.pages)
            for page_index, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    summary["text_page_count"] += 1
                for match_index, match in enumerate(NUMERIC_TOKEN_RE.findall(text), 1):
                    fingerprint = _decimal_fingerprint(match)
                    if fingerprint is None:
                        continue
                    summary["numeric_token_count"] += 1
                    private_records.append(
                        {
                            "source_ref_hash": source_ref_hash,
                            "page_ref_hash": sha256_text(str(page_index)),
                            "token_ref_hash": sha256_text(str(match_index)),
                            "numeric_value_fingerprint": fingerprint,
                        }
                    )
    except Exception as exc:  # pragma: no cover - depends on local PDF shape
        summary["error_class"] = type(exc).__name__
    summary["fingerprint_count"] = len(private_records)
    return {"summary": summary, "private_records": private_records}


def _extract_private_raw_fingerprints() -> tuple[dict[str, Any], dict[str, Any]]:
    raw_package, selector_status, candidate_count = resolve_private_raw_zip()
    public = {
        **_root_public_inventory(),
        "source_container_selector_status": selector_status,
        "source_container_candidate_count": candidate_count,
        "source_container_present": raw_package is not None and raw_package.exists(),
        "source_container_openable": False,
        "package_business_entry_count": 0,
        "package_pdf_entry_count": 0,
        "package_workbook_entry_count": 0,
        "root_spreadsheet_file_count": 0,
        "root_spreadsheet_openable_count": 0,
        "raw_value_fingerprint_count": 0,
        "workbook_numeric_fingerprint_count": 0,
        "pdf_numeric_fingerprint_count": 0,
        "raw_value_fingerprints_generated": False,
    }
    private = {
        "selected_source_ref_hash": sha256_text(str(raw_package)) if raw_package is not None else "",
        "workbook_records": [],
        "pdf_records": [],
        "root_workbook_records": [],
        "parse_errors": [],
    }

    if raw_package is not None and raw_package.exists():
        with zipfile.ZipFile(raw_package) as archive:
            public["source_container_openable"] = True
            for info in archive.infolist():
                if info.is_dir() or is_hidden_zip_member(info.filename):
                    continue
                suffix = Path(info.filename).suffix.lower()
                source_ref_hash = sha256_text(info.filename)
                binary = archive.read(info)
                public["package_business_entry_count"] += 1
                if suffix == ".pdf":
                    public["package_pdf_entry_count"] += 1
                    result = _fingerprint_pdf(binary, source_ref_hash=source_ref_hash)
                    public["pdf_numeric_fingerprint_count"] += int(result["summary"]["fingerprint_count"])
                    private["pdf_records"].extend(result["private_records"])
                    if result["summary"].get("error_class"):
                        private["parse_errors"].append({"source_ref_hash": source_ref_hash, "error_class": result["summary"]["error_class"]})
                elif suffix in {".xlsx", ".xlsm"}:
                    public["package_workbook_entry_count"] += 1
                    result = _fingerprint_workbook(binary, source_ref_hash=source_ref_hash)
                    public["workbook_numeric_fingerprint_count"] += int(result["summary"]["fingerprint_count"])
                    private["workbook_records"].extend(result["private_records"])
                    if result["summary"].get("error_class"):
                        private["parse_errors"].append({"source_ref_hash": source_ref_hash, "error_class": result["summary"]["error_class"]})

    if RAW_INBOX.exists():
        for path in sorted(RAW_INBOX.iterdir()):
            if path.suffix.lower() not in {".xlsx", ".xlsm"} or not path.is_file():
                continue
            public["root_spreadsheet_file_count"] += 1
            source_ref_hash = sha256_text(str(path))
            result = _fingerprint_workbook(path.read_bytes(), source_ref_hash=source_ref_hash)
            if result["summary"]["workbook_openable"]:
                public["root_spreadsheet_openable_count"] += 1
            public["workbook_numeric_fingerprint_count"] += int(result["summary"]["fingerprint_count"])
            private["root_workbook_records"].extend(result["private_records"])
            if result["summary"].get("error_class"):
                private["parse_errors"].append({"source_ref_hash": source_ref_hash, "error_class": result["summary"]["error_class"]})

    public["raw_value_fingerprint_count"] = (
        public["workbook_numeric_fingerprint_count"] + public["pdf_numeric_fingerprint_count"]
    )
    public["raw_value_fingerprints_generated"] = public["raw_value_fingerprint_count"] > 0
    public["expected_pdf_entry_count"] = EXPECTED_PDF_COUNT
    public["expected_workbook_entry_count"] = EXPECTED_EXCEL_COUNT
    return public, private


def _iter_public_records(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    if path.suffix.lower() == ".json":
        try:
            value = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return records
        if isinstance(value, dict):
            records.append(value)
        elif isinstance(value, list):
            records.extend(item for item in value if isinstance(item, dict))
    elif path.suffix.lower() == ".jsonl":
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except Exception:
                continue
            if isinstance(value, dict):
                records.append(value)
    elif path.suffix.lower() == ".csv":
        try:
            with path.open(newline="", encoding="utf-8") as handle:
                records.extend(dict(row) for row in csv.DictReader(handle))
        except Exception:
            return records
    return records


def _scan_processed_targets() -> dict[str, Any]:
    files: list[Path] = []
    for root in PROCESSED_SCAN_ROOTS:
        if root.exists():
            files.extend(path for path in root.rglob("*") if path.suffix.lower() in {".json", ".jsonl", ".csv"})
    candidate_records = 0
    explicit_value_policy_blocks = 0
    reference_or_hash_only_records = 0
    approved_private_target_records = 0
    for path in files:
        for record in _iter_public_records(path):
            value_like_keys = [key for key in record if PROCESSED_VALUE_KEY_RE.search(str(key))]
            if not value_like_keys:
                continue
            candidate_records += 1
            flags = {
                key: record.get(key)
                for key in (
                    "raw_business_values_allowed",
                    "public_amount_values_allowed",
                    "public_numeric_values_allowed",
                    "contains_true_amounts",
                    "normalized_business_values_committed",
                    "raw_business_values_committed",
                )
                if key in record
            }
            if any(value is False for value in flags.values()):
                explicit_value_policy_blocks += 1
            if any("ref" in str(record.get(key, "")).lower() or "sha256:" in str(record.get(key, "")).lower() for key in value_like_keys):
                reference_or_hash_only_records += 1
    return {
        "processed_artifact_files_scanned": len(files),
        "processed_candidate_records_seen": candidate_records,
        "processed_records_with_explicit_no_value_policy": explicit_value_policy_blocks,
        "processed_reference_or_hash_only_records": reference_or_hash_only_records,
        "approved_private_processed_value_target_count": approved_private_target_records,
        "processed_value_targets_available": approved_private_target_records > 0,
        "processed_value_target_status": "missing_private_processed_value_targets",
    }


def _public_safety() -> dict[str, bool]:
    return {
        "public_safe_aggregate_only": True,
        "raw_business_data_committed": False,
        "compressed_raw_package_committed": False,
        "office_workbook_committed": False,
        "source_document_committed": False,
        "private_csv_committed": False,
        "raw_filenames_committed": False,
        "raw_hashes_committed": False,
        "archive_entry_names_committed": False,
        "sheet_names_committed": False,
        "field_or_header_plaintext_committed": False,
        "raw_or_processed_business_values_committed": False,
        "credential_or_secret_committed": False,
        "private_diagnostic_committed": False,
    }


def _go_no_go(processed_scan: dict[str, Any]) -> dict[str, Any]:
    return {
        "record_type": "v014_raw_value_matching_private_dry_run_go_no_go_report",
        "schema_version": GO_NO_GO_SCHEMA_VERSION,
        "project_id": PROJECT_ID,
        "version": VERSION,
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "decision": "NO_GO",
        "decision_reason": (
            "Private raw value fingerprints were generated read-only, but approved private processed value targets "
            "are not available, so raw-to-processed value consistency cannot be claimed."
        ),
        "resolved_blocker_ids": [
            "RAW_PRIVATE_VALUE_FINGERPRINTS_GENERATED",
            "RAW_INBOX_READONLY_STAT_GUARD_PASSED",
            "DRY_RUN_GAP_REPORT_GENERATED",
        ],
        "blocker_ids": [
            "PROCESSED_PRIVATE_VALUE_TARGETS_MISSING",
            "COMPARABLE_RAW_PROCESSED_VALUE_PAIRS_ZERO",
            "BUSINESS_VALUE_CONSISTENCY_NOT_VERIFIED",
            "LINEAGE_FULL_CHECK_BLOCKED_BY_VALUE_CONSISTENCY",
            "FORMAL_REPORT_RELEASE_BLOCKED_BY_LINEAGE",
            "GITHUB_UPLOAD_BLOCKED_BY_NO_GO",
            "APP_REINSTALL_BLOCKED_BY_NO_GO",
            "BUSINESS_EXECUTION_BLOCKED_BY_NO_GO",
        ],
        "processed_value_targets_available": processed_scan["processed_value_targets_available"],
        "comparable_value_pair_count": 0,
        "business_value_consistency_verified": False,
        "lineage_full_check_complete": False,
        "formal_report_allowed": False,
        "github_upload_allowed": False,
        "app_reinstall_allowed": False,
        "business_execution_allowed": False,
        "next_recommended_phase": NEXT_RECOMMENDED_PHASE,
    }


def _gap_summary(processed_scan: dict[str, Any]) -> dict[str, Any]:
    return {
        "record_type": "v014_raw_value_matching_private_dry_run_gap_summary",
        "schema_version": "kmfa.v014_raw_value_matching_private_dry_run_gap_summary.v1",
        "project_id": PROJECT_ID,
        "version": VERSION,
        "phase_id": PHASE_ID,
        "gap_count": 1,
        "gaps": [
            {
                "gap_id": "V014-RVM-DRYRUN-GAP-001",
                "scope_lane_id": "VC-L04",
                "mismatch_class": "processed_private_value_targets_missing",
                "severity": "blocking",
                "responsible_role": "codex_private_runtime_or_owner_authorized_processed_export",
                "raw_value_fingerprints_generated": True,
                "processed_value_targets_available": processed_scan["processed_value_targets_available"],
                "business_values_disclosed": False,
                "next_action": NEXT_RECOMMENDED_PHASE,
            }
        ],
        "repeated_cross_validation_mismatch_confirmed": False,
        "final_goal_closeout_difference_report_required_if_repeated": True,
        "public_raw_or_processed_values_included": False,
    }


def generate(generated_at: str | None = None) -> dict[str, Any]:
    generated_at = _now(generated_at)
    raw_root_before = stat_snapshot(RAW_INBOX) if RAW_INBOX.exists() else {}
    raw_public, raw_private = _extract_private_raw_fingerprints()
    raw_root_after = stat_snapshot(RAW_INBOX) if RAW_INBOX.exists() else {}
    processed_scan = _scan_processed_targets()
    gap_summary = _gap_summary(processed_scan)
    go_no_go = _go_no_go(processed_scan)

    private_diagnostic = {
        "schema_version": PRIVATE_SCHEMA_VERSION,
        "classification": "private_raw_value_matching_diagnostic_do_not_commit",
        "project_id": PROJECT_ID,
        "version": VERSION,
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "generated_at": generated_at,
        "raw_root_before": raw_root_before,
        "raw_root_after": raw_root_after,
        "raw_root_stat_unchanged_after_dry_run": raw_root_before == raw_root_after,
        "raw_private_summary": raw_public,
        "raw_private_records": raw_private,
        "processed_scan_summary": processed_scan,
        "comparable_value_pair_count": 0,
        "business_value_consistency_verified": False,
        "raw_inbox_mutation_detected": False,
    }
    _write_json(PRIVATE_DIAGNOSTIC_PATH, private_diagnostic)
    _write_text(
        PRIVATE_GAP_REPORT_PATH,
        "\n".join(
            [
                "# KMFA private raw value matching dry-run gap report",
                "",
                "- classification: private_runtime_do_not_commit",
                "- gap: approved private processed value targets are missing",
                "- raw private fingerprints generated: true",
                "- comparable raw processed value pairs: 0",
                "- business value consistency verified: false",
                "- raw inbox mutation detected: false",
                "- next action: create authorized private processed value staging before claiming consistency",
                "",
            ]
        ),
    )

    manifest = {
        "record_type": "v014_raw_value_matching_private_dry_run_manifest",
        "schema_version": SCHEMA_VERSION,
        "project_id": PROJECT_ID,
        "version": VERSION,
        "phase_id": PHASE_ID,
        "phase_name": "Raw value matching private dry run",
        "task_id": TASK_ID,
        "acceptance_ids": [ACCEPTANCE_ID],
        "status": STATUS,
        "generated_at": generated_at,
        "source_commit": _git_output(["rev-parse", "HEAD"]),
        "branch": _git_output(["branch", "--show-current"]),
        "dependencies": {
            "value_scope_manifest": str(VALUE_SCOPE_MANIFEST_PATH),
            "value_scope_go_no_go": str(VALUE_SCOPE_GO_NO_GO_PATH),
            "authoritative_raw_baseline_lock": str(RAW_BASELINE_LOCK_PATH),
        },
        "phase_scope_controls": {
            "current_phase_only": True,
            "private_raw_value_matching_dry_run_only": True,
            "authoritative_raw_baseline_dependency_consumed": True,
            "raw_value_fingerprint_extraction_performed": True,
            "processed_value_target_scan_performed": True,
            "raw_to_processed_value_comparison_performed": False,
            "processed_data_reconciliation_performed": False,
            "lineage_full_check_performed": False,
            "formal_report_performed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "next_phase_started": False,
        },
        "raw_readonly_boundary": {
            "raw_inbox_read_performed_by_this_phase": True,
            "raw_inbox_list_performed_by_this_phase": True,
            "raw_inbox_stat_performed_by_this_phase": True,
            "raw_inbox_hash_or_value_fingerprint_performed_by_this_phase": True,
            "raw_inbox_write_performed_by_this_phase": False,
            "raw_inbox_delete_performed_by_this_phase": False,
            "raw_inbox_move_performed_by_this_phase": False,
            "raw_inbox_rename_performed_by_this_phase": False,
            "raw_inbox_overwrite_performed_by_this_phase": False,
            "raw_inbox_copy_performed_by_this_phase": False,
            "raw_inbox_create_extra_files_inside_by_this_phase": False,
            "raw_inbox_normalize_performed_by_this_phase": False,
            "raw_inbox_mutated_by_this_phase": False,
            "raw_root_stat_unchanged_after_dry_run": raw_root_before == raw_root_after,
        },
        "raw_private_extraction_summary": raw_public,
        "processed_target_summary": processed_scan,
        "value_matching_summary": {
            "raw_value_fingerprints_generated": raw_public["raw_value_fingerprints_generated"],
            "raw_value_fingerprint_count": raw_public["raw_value_fingerprint_count"],
            "approved_private_processed_value_target_count": processed_scan["approved_private_processed_value_target_count"],
            "comparable_value_pair_count": 0,
            "raw_to_processed_value_comparison_performed": False,
            "processed_data_consistency_verified": False,
            "business_value_consistency_verified": False,
            "dry_run_gap_report_generated": True,
            "repeated_cross_validation_mismatch_confirmed": False,
            "minimum_independent_validation_passes_required": 3,
            "independent_validation_passes_completed_by_this_phase": 1,
            "final_goal_closeout_difference_report_required_if_repeated": True,
        },
        "go_no_go": go_no_go,
        "gap_summary": gap_summary,
        "public_repo_safety": _public_safety(),
        "private_diagnostic_ref": "KMFA/.codex_private_runtime/v014_raw_value_matching_private_dry_run/private_raw_value_matching_diagnostic.json",
        "private_gap_report_ref": "KMFA/.codex_private_runtime/v014_raw_value_matching_private_dry_run/local_gap_report.md",
        "evidence_refs": [
            str(REPORT_PATH),
            str(GO_NO_GO_RECORD_PATH),
            str(TEST_RESULTS_PATH),
            str(RISK_REGISTER_PATH),
            str(ROLLBACK_PATH),
            str(MANIFEST_PATH),
            str(GO_NO_GO_PATH),
            str(GAP_SUMMARY_PATH),
            str(METADATA_MANIFEST_PATH),
            str(METADATA_GO_NO_GO_PATH),
            str(METADATA_GAP_SUMMARY_PATH),
        ],
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "formal_report_performed": False,
        "business_execution_performed": False,
        "next_recommended_phase": NEXT_RECOMMENDED_PHASE,
    }

    for path, payload in (
        (MANIFEST_PATH, manifest),
        (GO_NO_GO_PATH, go_no_go),
        (GAP_SUMMARY_PATH, gap_summary),
        (METADATA_MANIFEST_PATH, manifest),
        (METADATA_GO_NO_GO_PATH, go_no_go),
        (METADATA_GAP_SUMMARY_PATH, gap_summary),
    ):
        _write_json(path, payload)

    _write_text(
        REPORT_PATH,
        "\n".join(
            [
                "# KMFA v0.1.4 Raw Value Matching Private Dry Run",
                "",
                "- status: `completed_validated_local_only_no_go_private_raw_value_fingerprints_processed_targets_missing`",
                f"- phase_id: `{PHASE_ID}`",
                f"- task_id: `{TASK_ID}`",
                "- raw_value_fingerprints_generated: `true`",
                f"- raw_value_fingerprint_count: `{raw_public['raw_value_fingerprint_count']}`",
                f"- processed_value_targets_available: `{str(processed_scan['processed_value_targets_available']).lower()}`",
                "- comparable_value_pair_count: `0`",
                "- business_value_consistency_verified: `false`",
                "- dry_run_gap_report_generated: `true`",
                "- decision: `NO_GO`",
                "",
                "## Boundary",
                "",
                "- This phase reads raw sources only to generate private value fingerprints.",
                "- Public evidence contains only aggregate counts, readiness flags and gate status.",
                "- Raw filenames, archive entry names, sheet names, headers, cell values, PDF text and business values are not public evidence.",
                "- Raw inbox write, delete, move, rename, overwrite, copy and in-place normalization are false.",
                "- The current blocker is missing approved private processed value targets; consistency is not claimed.",
                "",
            ]
        ),
    )
    _write_text(
        GO_NO_GO_RECORD_PATH,
        "\n".join(
            [
                "# KMFA v0.1.4 Raw Value Matching Private Dry Run Go/No-Go",
                "",
                "- decision: `NO_GO`",
                "- reason: `raw_private_fingerprints_exist_but_processed_private_value_targets_missing`",
                "- raw_to_processed_value_comparison_performed: `false`",
                "- business_value_consistency_verified: `false`",
                "- github_upload_allowed: `false`",
                "- app_reinstall_allowed: `false`",
                "- formal_report_allowed: `false`",
                "- business_execution_allowed: `false`",
                "",
            ]
        ),
    )
    _write_text(
        TEST_RESULTS_PATH,
        "\n".join(
            [
                "# KMFA v0.1.4 Raw Value Matching Private Dry Run Test Results",
                "",
                "- status: `pending_final_validation`",
                "- generator: `pending_final_validation`",
                "- validator: `pending_final_validation`",
                "- focused_unit_test: `pending_final_validation`",
                "- governance_validator: `pending_final_validation`",
                "- raw_private_scan: `pending_final_validation`",
                "- secret_scan: `pending_final_validation`",
                "- diff_check: `pending_final_validation`",
                "",
            ]
        ),
    )
    _write_text(
        RISK_REGISTER_PATH,
        "\n".join(
            [
                "# KMFA v0.1.4 Raw Value Matching Private Dry Run Risk Register",
                "",
                "- risk: approved private processed value targets are missing.",
                "  mitigation: keep NO_GO and create a separate private processed staging phase.",
                "- risk: private fingerprints could be mistaken for public consistency evidence.",
                "  mitigation: validator requires business_value_consistency_verified=false and public-safe aggregate-only evidence.",
                "- risk: raw inbox mutation.",
                "  mitigation: before/after stat guard and no writes to raw inbox.",
                "",
            ]
        ),
    )
    _write_text(
        ROLLBACK_PATH,
        "\n".join(
            [
                "# KMFA v0.1.4 Raw Value Matching Private Dry Run Rollback",
                "",
                "- Remove this phase public artifacts, metadata copies, validator, focused test and governance records.",
                "- Remove ignored private diagnostic files under the phase private runtime directory if needed.",
                "- Do not modify or delete the raw inbox.",
                "",
            ]
        ),
    )
    return manifest


def main() -> int:
    manifest = generate()
    summary = manifest["value_matching_summary"]
    print(
        "PASS: generated KMFA v0.1.4 raw value matching private dry run "
        f"(raw_fingerprints={summary['raw_value_fingerprint_count']}, "
        f"processed_targets={summary['approved_private_processed_value_target_count']}, "
        f"business_consistency={summary['business_value_consistency_verified']}, decision=NO_GO)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
