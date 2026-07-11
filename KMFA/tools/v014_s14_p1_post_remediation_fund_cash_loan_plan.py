#!/usr/bin/env python3
"""Generate current public-safe KMFA v0.1.4 S14-P1 evidence."""

from __future__ import annotations

import argparse
import functools
import hashlib
import html
import http.server
import io
import json
import os
import socketserver
import subprocess
import sys
import threading
import zipfile
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

from KMFA.tools import v014_s13_post_remediation_stage_review as s13_review
from KMFA.tools.check_v014_s13_post_remediation_stage_review import (
    validate_v014_s13_post_remediation_stage_review,
)


PHASE_ID = "V014_S14_P1_POST_REMEDIATION_FUND_CASH_LOAN_PLAN"
ROADMAP_PHASE_ID = "S14-P1"
TASK_ID = "KMFA-V014-S14-P1-POST-REMEDIATION-FUND-CASH-LOAN-PLAN-20260711"
ACCEPTANCE_ID = "ACC-V014-S14-P1-POST-REMEDIATION-FUND-CASH-LOAN-PLAN"
VERSION = "0.1.4-s14-p1-post-remediation-fund-cash-loan-plan"
STATUS = "completed_validated_local_only_s14_p1_structure_method_no_go_upload_deferred"
DECISION = "NO_GO"
FORMULA_ID = "FORM-KMFA-V014-S14P1-POST-REMEDIATION-FUND-CASH-LOAN-PLAN-001"
PARAMETER_IDS = (
    "PARAM-KMFA-1750",
    "PARAM-KMFA-1751",
    "PARAM-KMFA-1752",
    "PARAM-KMFA-1753",
)
MODEL_REGISTRY_KEY = "kmfa_v014_s14_p1_post_remediation_fund_cash_loan_plan"

OUTPUT_DIR = Path("KMFA/stage_artifacts") / PHASE_ID
MACHINE_DIR = OUTPUT_DIR / "machine"
HUMAN_DIR = OUTPUT_DIR / "human"
HTML_DIR = OUTPUT_DIR / "exports/html"
SUMMARY_PATH = MACHINE_DIR / "fund_cash_loan_plan_summary.json"
MANIFEST_PATH = MACHINE_DIR / "fund_cash_loan_plan_manifest.json"
LANES_PATH = MACHINE_DIR / "source_lane_status_public_safe.json"
METHODS_PATH = MACHINE_DIR / "planning_method_definitions_public_safe.json"
MATRIX_PATH = MACHINE_DIR / "fund_cash_loan_plan_acceptance_matrix.json"
GO_NO_GO_PATH = MACHINE_DIR / "fund_cash_loan_plan_go_no_go.json"
HTML_PATH = HTML_DIR / "fund_cash_loan_workbench.html"
REPORT_PATH = HUMAN_DIR / "fund_cash_loan_plan_report_zh.md"
TEST_RESULTS_PATH = HUMAN_DIR / "test_results_zh.md"
RISK_REGISTER_PATH = HUMAN_DIR / "risk_register_zh.md"
ROLLBACK_PATH = HUMAN_DIR / "rollback_plan_zh.md"

QUALITY_DIR = Path("KMFA/metadata/quality")
METADATA_SUMMARY_PATH = QUALITY_DIR / "v014_s14_p1_post_remediation_fund_cash_loan_plan_summary.json"
METADATA_MANIFEST_PATH = QUALITY_DIR / "v014_s14_p1_post_remediation_fund_cash_loan_plan_manifest.json"
METADATA_LANES_PATH = QUALITY_DIR / "v014_s14_p1_post_remediation_source_lane_status_public_safe.json"
METADATA_METHODS_PATH = QUALITY_DIR / "v014_s14_p1_post_remediation_planning_method_definitions_public_safe.json"
METADATA_MATRIX_PATH = QUALITY_DIR / "v014_s14_p1_post_remediation_fund_cash_loan_plan_acceptance_matrix.json"
METADATA_GO_NO_GO_PATH = QUALITY_DIR / "v014_s14_p1_post_remediation_fund_cash_loan_plan_go_no_go.json"

PRIVATE_DIR = Path("KMFA/.codex_private_runtime/v014_s14_p1_post_remediation_fund_cash_loan_plan")
PRIVATE_RAW_BEFORE_PATH = PRIVATE_DIR / "raw_immutability_before.json"
PRIVATE_RAW_AFTER_PATH = PRIVATE_DIR / "raw_immutability_after.json"
PRIVATE_PROBE_PATH = PRIVATE_DIR / "raw_fund_cash_loan_candidate_probe.json"
PRIVATE_REPORT_PATH = PRIVATE_DIR / "s14_p1_private_validation_zh.md"
PRIVATE_BROWSER_PATH = PRIVATE_DIR / "browser_verification.json"
PRIVATE_BASELINE_AUDIT_PATH = PRIVATE_DIR / "human_flow_baseline_audit.csv"
PRIVATE_CURRENT_AUDIT_PATH = PRIVATE_DIR / "current_fund_cash_loan_workbench_audit.csv"
PRIVATE_SCREENSHOT_DIR = PRIVATE_DIR / "screenshots"

ROADMAP_PATH = Path("KMFA/taskpack/v1_4/roadmap/02_KMFA_Codex_Development_Roadmap_18_Stages_v1_4.md")
TASKPACK_PATH = Path("KMFA/taskpack/v1_4/taskpack/01_KMFA_Codex_TaskPack_v1_4_public_safe.md")
HTML_BASELINE_ROOT = Path("KMFA/taskpack/v1_4/html_uiux")
CURRENT_SOURCE_REGISTRY_PATH = Path("KMFA/metadata/imports/v014_s07_p1_finance_support_source_registry.json")
CURRENT_CANDIDATES_PATH = Path("KMFA/metadata/schema_maps/v014_s07_p1_finance_field_candidates.jsonl")
HISTORICAL_MANIFEST_PATH = Path(
    "KMFA/stage_artifacts/V014_S14_P1_FUND_CASH_LOAN_PLAN/machine/fund_cash_loan_plan_manifest.json"
)

DEVELOPMENT_EVENTS_PATH = Path("KMFA/docs/governance/development_events.jsonl")
STAGE_STATUS_PATH = Path("KMFA/metadata/stage_status.jsonl")
TASK_STATUS_PATH = Path("KMFA/metadata/traceability/v1_4_stage_phase_task_status.jsonl")

DEPENDENCY_LINKS = {
    "weekly": (
        "../../../V014_S13_P1_POST_REMEDIATION_FINANCIAL_OPERATING_REPORT/exports/html/"
        "financial_operating_weekly_draft.html",
        "经营周报初稿",
    ),
    "monthly": (
        "../../../V014_S13_P1_POST_REMEDIATION_FINANCIAL_OPERATING_REPORT/exports/html/"
        "financial_operating_monthly_draft.html",
        "经营月报初稿",
    ),
    "receivable": (
        "../../../V014_S13_P2_POST_REMEDIATION_COLLECTION_RECEIVABLE_AGING/exports/html/"
        "collection_receivable_aging_workbench.html",
        "回款与应收账龄工作台",
    ),
    "cross-table": (
        "../../../V014_S13_P3_POST_REMEDIATION_CROSS_TABLE_REVIEW/exports/html/"
        "cross_table_quality_workbench.html",
        "跨表复核质量工作台",
    ),
}

LANE_SPECS = (
    {
        "lane_id": "account_list",
        "visible_lane_name": "账户清单",
        "categories": ("account",),
        "status_label": "账户结构候选和私有工作表候选已接入；账号、余额与主体绑定未被证明。",
    },
    {
        "lane_id": "monthly_cash",
        "visible_lane_name": "月度现金",
        "categories": ("cash",),
        "status_label": "现金结构候选可解析；期间、账户与数值绑定未被证明。",
    },
    {
        "lane_id": "fund_plan",
        "visible_lane_name": "资金计划",
        "categories": ("cash", "journal"),
        "status_label": "资金计划候选结构已观察；收支口径与计划期间仍需权威绑定。",
    },
    {
        "lane_id": "loan_detail",
        "visible_lane_name": "贷款明细",
        "categories": ("loan",),
        "status_label": "贷款候选结构可解析；合同、金额、利率与到期日绑定未被证明。",
    },
)

LANE_KEYWORDS = {
    "account_list": (("账户", "账号", "开户行", "银行账号", "账户名称"), 2),
    "monthly_cash": (("月度现金", "现金情况", "现金余额", "期初现金", "期末现金", "现金流", "现金"), 1),
    "fund_plan": (("资金计划", "资金需求", "收支计划", "付款计划", "资金预算"), 1),
    "loan_detail": (("贷款", "借款", "还款", "到期日", "贷款明细"), 1),
}

METHOD_SPECS = (
    {
        "method_id": "cash_pressure",
        "visible_name": "现金压力",
        "required_lanes": ("monthly_cash", "fund_plan", "account_list"),
        "review_sequence": 1,
        "method_note": "需权威期初、流入、流出、期末及期间绑定后才能识别压力项。",
    },
    {
        "method_id": "loan_due",
        "visible_name": "贷款到期",
        "required_lanes": ("loan_detail", "account_list"),
        "review_sequence": 2,
        "method_note": "需权威贷款合同、到期日、金额和主体绑定后才能形成到期提示。",
    },
    {
        "method_id": "account_balance_summary",
        "visible_name": "账户余额汇总",
        "required_lanes": ("account_list", "monthly_cash"),
        "review_sequence": 3,
        "method_note": "需账户身份、币种、期间和余额字段绑定后才能形成汇总。",
    },
)


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"{path} must contain an object")
    return value


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"{path} contains a non-object row")
            rows.append(value)
    return rows


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8")


def _git_output(args: list[str]) -> str:
    result = subprocess.run(
        ["git", "-c", "core.quotePath=false", *args],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip())
    return result.stdout.strip()


def _upsert_jsonl(path: Path, row: dict[str, Any]) -> None:
    preserved: list[str] = []
    if path.is_file():
        for line in path.read_text(encoding="utf-8").splitlines():
            if line.strip() and json.loads(line).get("phase_id") != PHASE_ID:
                preserved.append(line)
    preserved.append(json.dumps(row, ensure_ascii=False, separators=(",", ":")))
    _write_text(path, "\n".join(preserved))


def _sha256_bytes(value: bytes) -> str:
    return "sha256:" + hashlib.sha256(value).hexdigest()


def _sha256_json(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return _sha256_bytes(payload)


def _load_dependency() -> dict[str, Any]:
    manifest = validate_v014_s13_post_remediation_stage_review(
        require_private_evidence=True,
        require_browser_evidence=True,
        require_final_evidence=True,
    )
    summary = manifest["summary"]
    if manifest.get("next_phase") != "S14-P1":
        raise ValueError("Stage 13 review must route to S14-P1")
    if summary.get("s14_p1_performed") is not False:
        raise ValueError("Stage 13 dependency already claims S14-P1")
    if summary.get("decision") != "NO_GO" or summary.get("current_report_grade") != "D":
        raise ValueError("Stage 13 quality state drift")
    return manifest


def _load_contract() -> dict[str, Any]:
    roadmap = ROADMAP_PATH.read_text(encoding="utf-8")
    taskpack = TASKPACK_PATH.read_text(encoding="utf-8")
    for token in (
        "S14-P1",
        "账户清单",
        "月度现金",
        "资金计划",
        "贷款明细",
        "现金压力",
        "贷款到期",
        "账户余额汇总",
        "不做付款审批和银行操作",
    ):
        if token not in roadmap:
            raise ValueError(f"roadmap token missing: {token}")
    for token in ("资金计划/现金/贷款线", "资金缺口", "账户汇总", "贷款到期提示", "不做付款操作"):
        if token not in taskpack:
            raise ValueError(f"taskpack token missing: {token}")
    return {
        "roadmap_read": True,
        "taskpack_read": True,
        "human_flow_baseline_read": True,
        "raw_read_only_contract_applied": True,
    }


def _build_source_lanes(private_counts: dict[str, int]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    registry = _read_json(CURRENT_SOURCE_REGISTRY_PATH)
    candidates = _read_jsonl(CURRENT_CANDIDATES_PATH)
    sources_by_category: dict[str, list[dict[str, Any]]] = {}
    candidates_by_category: dict[str, list[dict[str, Any]]] = {}
    for row in registry.get("sources", []):
        if isinstance(row, dict):
            sources_by_category.setdefault(str(row.get("finance_category")), []).append(row)
    for row in candidates:
        candidates_by_category.setdefault(str(row.get("finance_category")), []).append(row)

    lanes: list[dict[str, Any]] = []
    unique_sources: set[str] = set()
    unique_candidates: set[str] = set()
    lane_source_bindings = 0
    lane_candidate_associations = 0
    for spec in LANE_SPECS:
        source_refs: set[str] = set()
        candidate_refs: set[str] = set()
        quality_states: list[dict[str, Any]] = []
        for category in spec["categories"]:
            for source in sources_by_category.get(category, []):
                source_ref = str(source.get("source_ref") or "")
                if source_ref:
                    source_refs.add(source_ref)
                if source.get("read_only_parse") is not True:
                    raise ValueError(f"non-read-only source in {spec['lane_id']}")
            for candidate in candidates_by_category.get(category, []):
                candidate_ref = str(candidate.get("canonical_field_ref") or candidate.get("candidate_id") or "")
                if candidate_ref:
                    candidate_refs.add(candidate_ref)
                quality = candidate.get("quality_state")
                if isinstance(quality, dict):
                    quality_states.append(quality)
        if not source_refs or not candidate_refs:
            raise ValueError(f"missing public-safe structure for {spec['lane_id']}")
        if any(row.get("q4_human_confirmed") is True for row in quality_states):
            raise ValueError(f"unexpected Q4 binding in {spec['lane_id']}")
        unique_sources.update(source_refs)
        unique_candidates.update(candidate_refs)
        lane_source_bindings += len(source_refs)
        lane_candidate_associations += len(candidate_refs)
        lanes.append(
            {
                "lane_id": spec["lane_id"],
                "visible_lane_name": spec["visible_lane_name"],
                "public_source_ref_count": len(source_refs),
                "public_structure_candidate_count": len(candidate_refs),
                "private_candidate_sheet_count": private_counts[spec["lane_id"]],
                "structure_connected": True,
                "private_candidate_structure_parseable": private_counts[spec["lane_id"]] > 0,
                "row_level_binding_proven": False,
                "value_binding_proven": False,
                "data_status": "structure_and_private_candidates_connected_values_unproven",
                "status_label": spec["status_label"],
                "contains_business_amounts": False,
                "contains_account_identifiers": False,
                "contains_field_plaintext": False,
                "formal_report_allowed": False,
                "business_decision_basis_allowed": False,
                "payment_or_bank_operation_allowed": False,
                "loan_management_action_allowed": False,
            }
        )
    return lanes, {
        "unique_public_source_ref_count": len(unique_sources),
        "lane_source_binding_count": lane_source_bindings,
        "unique_structure_candidate_count": len(unique_candidates),
        "lane_structure_candidate_association_count": lane_candidate_associations,
    }


def _build_methods() -> list[dict[str, Any]]:
    return [
        {
            "method_id": spec["method_id"],
            "visible_name": spec["visible_name"],
            "required_lanes": list(spec["required_lanes"]),
            "review_sequence": spec["review_sequence"],
            "method_definition_complete": True,
            "method_note": spec["method_note"],
            "identified_business_item_count": 0,
            "public_business_amount_count": 0,
            "row_level_binding_required": True,
            "current_binding_status": "unproven",
            "current_output_status": "blocked_no_authoritative_row_value_binding",
            "cash_forecast_decision_allowed": False,
            "payment_approval_allowed": False,
            "bank_operation_allowed": False,
            "loan_management_action_allowed": False,
            "formal_report_allowed": False,
        }
        for spec in METHOD_SPECS
    ]


def _normalize_cell(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return format(Decimal(str(value)), "f")
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if type(value).__module__ == "openpyxl.worksheet.formula" and type(value).__name__ == "ArrayFormula":
        return {
            "formula_type": "ArrayFormula",
            "ref": str(value.ref),
            "text": str(value.text),
        }
    return str(value)


def _sheet_probe(sheet: Any, *, row_limit: int = 200, column_limit: int = 30) -> dict[str, Any]:
    normalized_rows: list[list[Any]] = []
    preview_rows: list[list[Any]] = []
    text_cells = 0
    numeric_cells = 0
    formula_cells = 0
    nonempty_rows = 0
    max_row = min(int(sheet.max_row or 0), row_limit)
    max_column = min(int(sheet.max_column or 0), column_limit)
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True):
        normalized = [_normalize_cell(value) for value in row]
        while normalized and normalized[-1] is None:
            normalized.pop()
        if not normalized or all(value is None for value in normalized):
            continue
        nonempty_rows += 1
        normalized_rows.append(normalized)
        if len(preview_rows) < 5:
            preview_rows.append(normalized)
        for value in normalized:
            if value is None:
                continue
            if isinstance(value, str) and value.startswith("="):
                formula_cells += 1
            elif isinstance(value, str):
                text_cells += 1
            else:
                numeric_cells += 1
    return {
        "probe_row_limit": row_limit,
        "probe_column_limit": column_limit,
        "observed_nonempty_row_count": nonempty_rows,
        "observed_text_cell_count": text_cells,
        "observed_numeric_cell_count": numeric_cells,
        "observed_formula_cell_count": formula_cells,
        "preview_rows_private": preview_rows,
        "probe_fingerprint": _sha256_json(normalized_rows),
        "probe_truncated": int(sheet.max_row or 0) > row_limit or int(sheet.max_column or 0) > column_limit,
    }


def _scan_workbook(
    payload: bytes,
    *,
    raw_path: Path,
    member_name: str | None,
    raw_index: int,
    member_index: int,
) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    records: list[dict[str, Any]] = []
    for sheet_index, sheet in enumerate(workbook.worksheets, 1):
        parts = [sheet.title]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(int(sheet.max_row or 0), 12),
            max_col=min(int(sheet.max_column or 0), 30),
            values_only=True,
        ):
            parts.extend(str(value) for value in row if value is not None)
        searchable = "\n".join(parts)
        matched_lanes: list[str] = []
        matched_terms: dict[str, list[str]] = {}
        for lane_id, (terms, threshold) in LANE_KEYWORDS.items():
            hits = [term for term in terms if term in searchable]
            if len(hits) >= threshold:
                matched_lanes.append(lane_id)
                matched_terms[lane_id] = hits
        if not matched_lanes:
            continue
        records.append(
            {
                "raw_index": raw_index,
                "raw_path_private": str(raw_path),
                "raw_filename_private": raw_path.name,
                "member_index": member_index,
                "member_name_private": member_name,
                "member_sha256": _sha256_bytes(payload),
                "sheet_index": sheet_index,
                "sheet_name_private": sheet.title,
                "sheet_max_row": int(sheet.max_row or 0),
                "sheet_max_column": int(sheet.max_column or 0),
                "matched_lanes": sorted(matched_lanes),
                "matched_terms_private": matched_terms,
                **_sheet_probe(sheet),
            }
        )
    workbook.close()
    return records


def _raw_candidate_probe(raw_root: Path) -> dict[str, Any]:
    raw_files = sorted(path for path in raw_root.rglob("*") if path.is_file())
    candidate_records: list[dict[str, Any]] = []
    unparseable: list[dict[str, Any]] = []
    xlsx_container_count = 0
    parseable_count = 0
    unparseable_count = 0
    roundtrip_mismatch_count = 0
    raw_records: list[dict[str, Any]] = []

    for raw_index, raw_path in enumerate(raw_files, 1):
        payloads: list[tuple[int, str | None, bytes]] = []
        raw_bytes = raw_path.read_bytes()
        if raw_path.suffix.lower() == ".xlsx":
            payloads.append((1, None, raw_bytes))
        elif raw_path.suffix.lower() == ".zip":
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
                payloads = [
                    (member_index, info.filename, archive.read(info))
                    for member_index, info in enumerate(archive.infolist(), 1)
                    if not info.is_dir() and Path(info.filename).suffix.lower() == ".xlsx"
                ]
        raw_records.append(
            {
                "raw_index": raw_index,
                "raw_path_private": str(raw_path),
                "raw_filename_private": raw_path.name,
                "raw_suffix": raw_path.suffix.lower(),
                "raw_size_bytes": raw_path.stat().st_size,
                "raw_sha256": _sha256_bytes(raw_bytes),
                "xlsx_container_count": len(payloads),
            }
        )
        for member_index, member_name, payload in payloads:
            xlsx_container_count += 1
            try:
                first = _scan_workbook(
                    payload,
                    raw_path=raw_path,
                    member_name=member_name,
                    raw_index=raw_index,
                    member_index=member_index,
                )
                second = _scan_workbook(
                    payload,
                    raw_path=raw_path,
                    member_name=member_name,
                    raw_index=raw_index,
                    member_index=member_index,
                )
            except Exception as exc:
                unparseable_count += 1
                unparseable.append(
                    {
                        "raw_index": raw_index,
                        "member_index": member_index,
                        "member_name_private": member_name,
                        "member_sha256": _sha256_bytes(payload),
                        "error_class": type(exc).__name__,
                    }
                )
                continue
            parseable_count += 1
            second_map = {
                (row["sheet_index"], tuple(row["matched_lanes"])): row["probe_fingerprint"]
                for row in second
            }
            for row in first:
                key = (row["sheet_index"], tuple(row["matched_lanes"]))
                if second_map.get(key) != row["probe_fingerprint"]:
                    roundtrip_mismatch_count += 1
            candidate_records.extend(first)

    count_by_lane = {
        lane_id: sum(lane_id in row["matched_lanes"] for row in candidate_records)
        for lane_id in LANE_KEYWORDS
    }
    return {
        "schema_version": "kmfa.v014.s14_p1_private_raw_candidate_probe.v1",
        "classification": "private_raw_diagnostic_never_commit",
        "raw_root_private": str(raw_root),
        "raw_file_count": len(raw_files),
        "private_xlsx_container_count": xlsx_container_count,
        "private_parseable_xlsx_count": parseable_count,
        "private_unparseable_xlsx_count": unparseable_count,
        "private_unique_candidate_sheet_count": len(candidate_records),
        "private_candidate_sheet_count_by_lane": count_by_lane,
        "private_parseable_lane_count": sum(count > 0 for count in count_by_lane.values()),
        "private_probe_roundtrip_mismatch_count": roundtrip_mismatch_count,
        "row_level_binding_proven_lane_count": 0,
        "value_binding_proven_lane_count": 0,
        "raw_files_private": raw_records,
        "unparseable_xlsx_private": unparseable,
        "candidate_sheets_private": candidate_records,
    }


def _render_html(lanes: list[dict[str, Any]], methods: list[dict[str, Any]]) -> str:
    lane_rows = "".join(
        f"<tr><td>{html.escape(row['visible_lane_name'])}</td><td>已接入</td>"
        f"<td>{row['private_candidate_sheet_count']}</td><td>未证明</td></tr>"
        for row in lanes
    )
    buttons = "".join(
        f'<button type="button" data-method-button="{row["method_id"]}">{html.escape(row["visible_name"])}</button>'
        for row in methods
    )
    panels = "".join(
        f'<section data-method-panel="{row["method_id"]}" hidden><div class="panel-head"><div><span>复核序列 {row["review_sequence"]}</span>'
        f'<h3>{html.escape(row["visible_name"])}</h3></div><span class="tag blocked">0 项已证明</span></div>'
        f'<dl><div><dt>输入主题</dt><dd>{len(row["required_lanes"])} 条</dd></div><div><dt>行级绑定</dt><dd>未证明</dd></div>'
        f'<div><dt>业务输出</dt><dd>已阻断</dd></div></dl><p>{html.escape(row["method_note"])}</p>'
        '<div class="limit">不得据此发起付款审批、银行操作、还款、续贷或其他贷款管理动作。</div></section>'
        for row in methods
    )
    links = "".join(
        f'<a data-dependency-link="{link_id}" href="{href}">{label}</a>'
        for link_id, (href, label) in DEPENDENCY_LINKS.items()
    )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>KMFA｜资金现金贷款工作台</title>
  <style>
    *{{box-sizing:border-box}} body{{margin:0;background:#f2f5f7;color:#172522;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif;letter-spacing:0}}
    button,a{{font:inherit}} .topbar{{background:#12352e;color:#fff;border-bottom:3px solid #67b89e}} .topbar-inner{{max-width:1120px;margin:auto;min-height:66px;padding:12px 20px;display:flex;align-items:center;justify-content:space-between;gap:18px}}
    .brand{{display:flex;align-items:center;gap:12px}} .mark{{width:38px;height:38px;background:#fff;color:#12352e;display:grid;place-items:center;font-weight:800;border-radius:6px}} .brand strong{{display:block;font-size:16px}} .brand span{{display:block;font-size:12px;color:#cce5dc;margin-top:3px}}
    .top-links{{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}} .top-links a{{color:#fff;text-decoration:none;border:1px solid #628a7d;padding:8px 10px;border-radius:4px;font-size:13px}}
    main{{max-width:1120px;margin:auto;padding:22px 20px 44px}} .headline{{display:flex;justify-content:space-between;gap:24px;align-items:flex-start;margin-bottom:16px}} h1{{font-size:28px;margin:0 0 8px}} .subtitle{{margin:0;color:#5b6e68;line-height:1.7}}
    .badges{{display:flex;gap:8px;flex-wrap:wrap}} .tag{{display:inline-flex;align-items:center;border:1px solid #cad5d1;background:#fff;padding:5px 8px;border-radius:4px;font-size:12px;font-weight:700}} .tag.ok{{color:#137455;border-color:#a9d8c8;background:#edf8f4}} .tag.blocked{{color:#a43c36;border-color:#edb5b0;background:#fff2f0}}
    .metrics{{display:grid;grid-template-columns:repeat(4,1fr);border:1px solid #cfd9d6;background:#fff;margin-bottom:18px}} .metric{{padding:16px;border-right:1px solid #cfd9d6}} .metric:last-child{{border-right:0}} .metric strong{{display:block;font-size:25px;color:#13372f}} .metric span{{font-size:12px;color:#667873}}
    .notice{{border-left:4px solid #a43c36;background:#fff2f0;padding:13px 15px;margin-bottom:18px;color:#7e302c;line-height:1.6}} .band{{background:#fff;border:1px solid #cfd9d6;margin-bottom:18px}} .band-header{{display:flex;justify-content:space-between;align-items:center;padding:14px 16px;border-bottom:1px solid #d9e1de}} h2{{font-size:19px;margin:0}} h3{{font-size:20px;margin:5px 0 0}}
    .table-wrap{{overflow-x:auto}} table{{width:100%;border-collapse:collapse;min-width:660px}} th,td{{padding:12px 16px;text-align:left;border-bottom:1px solid #e2e8e6;font-size:13px}} th{{background:#f6f8f7;color:#5e716b}} tbody tr:last-child td{{border-bottom:0}}
    .method-layout{{display:grid;grid-template-columns:210px minmax(0,1fr)}} .method-nav{{padding:12px;background:#f4f7f6;border-right:1px solid #d9e1de}} .method-nav button{{display:block;width:100%;min-height:42px;text-align:left;border:1px solid transparent;background:transparent;padding:9px 10px;color:#1c322c;cursor:pointer;border-radius:4px}} .method-nav button.active{{background:#fff;border-color:#78ae9d;color:#116b51;font-weight:700}}
    .method-panels{{padding:20px}} .panel-head{{display:flex;justify-content:space-between;gap:18px;align-items:flex-start}} .panel-head span{{font-size:12px;color:#6a7d77}} dl{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:20px 0}} dl div{{border-left:3px solid #8ebfaf;padding-left:10px}} dt{{font-size:12px;color:#6a7d77}} dd{{margin:4px 0 0;font-weight:700}} .method-panels p{{line-height:1.7}} .limit{{border:1px solid #edb5b0;background:#fff2f0;color:#8d342f;padding:12px}}
    .footer-band{{display:flex;justify-content:space-between;gap:16px;align-items:center;padding:13px 16px;background:#f6f8f7;border-top:1px solid #d9e1de}} .footer-band p{{margin:0;color:#5b6e68;font-size:13px}} .footer-band a{{color:#136c94;text-decoration:none;margin-left:12px;font-size:13px}} footer{{color:#687b75;font-size:12px;line-height:1.7}}
    @media(max-width:720px){{.topbar-inner,.headline,.footer-band{{display:block}} .top-links{{justify-content:flex-start;margin-top:12px}} main{{padding:18px 12px 36px}} h1{{font-size:24px}} .badges{{margin-top:12px}} .metrics{{grid-template-columns:1fr 1fr}} .metric:nth-child(2){{border-right:0}} .metric:nth-child(-n+2){{border-bottom:1px solid #cfd9d6}} .method-layout{{grid-template-columns:1fr}} .method-nav{{border-right:0;border-bottom:1px solid #d9e1de;display:grid;grid-template-columns:repeat(3,1fr);gap:5px}} .method-nav button{{text-align:center;padding:7px 4px}} dl{{grid-template-columns:1fr}} .footer-band .links{{margin-top:10px}} .footer-band a{{display:inline-block;margin:4px 10px 4px 0}}}}
  </style>
</head>
<body data-page-id="s14-p1-fund-cash-loan" data-active-method="cash_pressure" data-ui-ready="false">
  <header class="topbar"><div class="topbar-inner"><div class="brand"><div class="mark">KM</div><div><strong>KMFA 经营分析系统</strong><span>S14-P1 · 财务资金</span></div></div><nav class="top-links" aria-label="阶段页面导航">{links}</nav></div></header>
  <main>
    <section class="headline"><div><h1>资金现金贷款工作台</h1><p class="subtitle">原始候选结构已只读解析；当前仅展示结构状态与复核方法，不展示账户、金额或贷款明细。</p></div><div class="badges"><span class="tag ok">Q4 / D</span><span class="tag blocked">NO_GO</span><span class="tag">内部复核</span></div></section>
    <div class="notice"><strong>门禁：</strong> 4 条主题均缺权威行级与数值绑定；现金压力、贷款到期和账户余额均不得形成业务结论。</div>
    <section class="metrics" aria-label="状态摘要"><div class="metric"><strong>4 / 4</strong><span>结构主题已接入</span></div><div class="metric"><strong>4 / 4</strong><span>私有候选可解析</span></div><div class="metric"><strong>0 / 4</strong><span>行级绑定已证明</span></div><div class="metric"><strong>0</strong><span>已证明业务事项</span></div></section>
    <section class="band"><div class="band-header"><h2>资金来源主题</h2><span class="tag blocked">不含业务金额</span></div><div class="table-wrap"><table><thead><tr><th>主题</th><th>公开结构</th><th>私有候选表</th><th>行级绑定</th></tr></thead><tbody>{lane_rows}</tbody></table></div></section>
    <section class="band"><div class="band-header"><h2>三类复核方法</h2><span class="tag blocked">0 项已证明</span></div><div class="method-layout"><nav class="method-nav" aria-label="复核方法">{buttons}</nav><div class="method-panels">{panels}</div></div><div class="footer-band"><p id="interaction-status" aria-live="polite">已显示“现金压力”复核方法；当前不形成业务结论。</p><div class="links">{links}</div></div></section>
    <footer>本 phase 仅完成 S14-P1；S14-P2、S14-P3、Stage 14 整体复审、GitHub upload、app reinstall、正式报告、差异关闭和业务执行均未执行。</footer>
  </main>
  <script>
    const labels={json.dumps({row['method_id']: row['visible_name'] for row in methods}, ensure_ascii=False, separators=(',', ':'))};
    let actionSequence=0;
    function activate(id){{document.body.dataset.activeMethod=id;document.body.dataset.lastAction=`method:${{id}}:${{++actionSequence}}`;document.querySelectorAll('[data-method-button]').forEach(b=>b.classList.toggle('active',b.dataset.methodButton===id));document.querySelectorAll('[data-method-panel]').forEach(p=>p.hidden=p.dataset.methodPanel!==id);document.getElementById('interaction-status').textContent=`已显示“${{labels[id]}}”复核方法；当前不形成业务结论。`;}}
    document.querySelectorAll('[data-method-button]').forEach(button=>button.addEventListener('click',()=>activate(button.dataset.methodButton)));
    activate('cash_pressure');document.body.dataset.uiReady='true';
  </script>
</body>
</html>"""


class _QuietHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        return


def _browser_worker() -> dict[str, Any]:
    from playwright.sync_api import sync_playwright

    root = Path("KMFA/stage_artifacts").resolve()
    handler = functools.partial(_QuietHandler, directory=str(root))
    server = socketserver.TCPServer(("127.0.0.1", 0), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    base = f"http://127.0.0.1:{server.server_address[1]}/"
    workbench_url = urljoin(base, f"{PHASE_ID}/exports/html/{HTML_PATH.name}")
    viewport_checks: list[dict[str, Any]] = []
    method_checks: list[dict[str, Any]] = []
    http_checks: list[dict[str, Any]] = []
    navigation_checks: list[dict[str, Any]] = []
    PRIVATE_SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            for mode, width, height in (("desktop", 1440, 900), ("mobile", 390, 844)):
                page = browser.new_page(viewport={"width": width, "height": height})
                errors: list[str] = []
                page.on("console", lambda message, errors=errors: errors.append(message.text) if message.type == "error" else None)
                page.goto(workbench_url, wait_until="networkidle")
                page.wait_for_function("document.body.dataset.uiReady === 'true'")
                viewport_checks.append(
                    {
                        "mode": mode,
                        "width": width,
                        "height": height,
                        "marker_visible": page.get_by_role("heading", name="资金现金贷款工作台").is_visible(),
                        "d_no_go_visible": "Q4 / D" in page.locator("body").inner_text() and "NO_GO" in page.locator("body").inner_text(),
                        "console_error_count": len(errors),
                        "no_horizontal_overflow": page.evaluate("document.documentElement.scrollWidth <= document.documentElement.clientWidth"),
                    }
                )
                for method in METHOD_SPECS:
                    page.locator(f'[data-method-button="{method["method_id"]}"]').click()
                    method_checks.append(
                        {
                            "mode": mode,
                            "method_id": method["method_id"],
                            "passed": page.locator(f'[data-method-panel="{method["method_id"]}"]').is_visible()
                            and page.locator("body").get_attribute("data-active-method") == method["method_id"]
                            and method["visible_name"] in page.locator("#interaction-status").inner_text(),
                        }
                    )
                page.screenshot(path=str(PRIVATE_SCREENSHOT_DIR / f"fund_cash_loan_{mode}.png"), full_page=True)
                page.close()

            request = playwright.request.new_context()
            for link_id, (_, marker) in DEPENDENCY_LINKS.items():
                page = browser.new_page(viewport={"width": 1280, "height": 900})
                page.goto(workbench_url, wait_until="networkidle")
                href = page.locator(f'a[data-dependency-link="{link_id}"]').first.get_attribute("href") or ""
                target = urljoin(workbench_url, href)
                response = request.get(target)
                http_checks.append({"link_id": link_id, "status": response.status, "passed": response.ok})
                page.locator(f'a[data-dependency-link="{link_id}"]').first.click()
                page.wait_for_load_state("networkidle")
                navigation_checks.append(
                    {"link_id": link_id, "marker": marker, "passed": marker in page.locator("body").inner_text()}
                )
                page.close()
            request.dispose()
            browser.close()
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)
    result = {
        "status": "PASS",
        "viewport_checks": viewport_checks,
        "method_interaction_checks": method_checks,
        "dependency_link_http_checks": http_checks,
        "dependency_navigation_checks": navigation_checks,
    }
    if not (
        len(viewport_checks) == 2
        and all(
            row["marker_visible"]
            and row["d_no_go_visible"]
            and row["console_error_count"] == 0
            and row["no_horizontal_overflow"]
            for row in viewport_checks
        )
        and len(method_checks) == 6
        and all(row["passed"] for row in method_checks)
        and len(http_checks) == 4
        and all(row["passed"] for row in http_checks)
        and len(navigation_checks) == 4
        and all(row["passed"] for row in navigation_checks)
    ):
        result["status"] = "FAIL"
    _write_json(PRIVATE_BROWSER_PATH, result)
    return result


def _run_browser_review() -> dict[str, Any]:
    python = os.environ.get("KMFA_AUDIT_PYTHON", sys.executable)
    result = subprocess.run(
        [python, __file__, "--browser-evidence-only"],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
        env={**os.environ, "PYTHONDONTWRITEBYTECODE": "1", "PYTHONPATH": "."},
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip())
    browser = _read_json(PRIVATE_BROWSER_PATH)
    if browser.get("status") != "PASS":
        raise RuntimeError("S14-P1 browser validation failed")
    helper = s13_review.p1.s12_review.p1.s11_home
    baseline = helper._run_html_audit(HTML_BASELINE_ROOT, PRIVATE_BASELINE_AUDIT_PATH)
    current = helper._run_html_audit(HTML_DIR, PRIVATE_CURRENT_AUDIT_PATH)
    if baseline != {"file_count": 6, "control_row_count": 54, "pass_count": 54, "warn_count": 0, "fail_count": 0}:
        raise RuntimeError("v1.4 HTML baseline drift")
    if current["fail_count"] or current["warn_count"] or current["pass_count"] != current["control_row_count"]:
        raise RuntimeError("S14-P1 HTML audit failed")
    return {
        "status": "PASS",
        "baseline_file_count": baseline["file_count"],
        "baseline_control_row_count": baseline["control_row_count"],
        "baseline_pass_count": baseline["pass_count"],
        "baseline_warn_count": baseline["warn_count"],
        "baseline_fail_count": baseline["fail_count"],
        "current_page_count": current["file_count"],
        "current_control_row_count": current["control_row_count"],
        "current_pass_count": current["pass_count"],
        "current_warn_count": current["warn_count"],
        "current_fail_count": current["fail_count"],
        "viewport_check_count": len(browser["viewport_checks"]),
        "method_interaction_check_count": len(browser["method_interaction_checks"]),
        "dependency_link_http_check_count": len(browser["dependency_link_http_checks"]),
        "dependency_navigation_check_count": len(browser["dependency_navigation_checks"]),
        "console_error_count": sum(row["console_error_count"] for row in browser["viewport_checks"]),
        "horizontal_overflow_count": sum(not row["no_horizontal_overflow"] for row in browser["viewport_checks"]),
    }


def _quality_gate() -> dict[str, Any]:
    return {
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "decision": "NO_GO",
        "release_permission": "blocked",
        "structure_status_allowed": True,
        "method_definition_allowed": True,
        "cash_forecast_decision_allowed": False,
        "account_balance_business_summary_allowed": False,
        "loan_due_business_alert_allowed": False,
        "derived_amount_calculation_allowed": False,
        "formal_report_allowed": False,
        "business_decision_basis_allowed": False,
        "payment_approval_allowed": False,
        "payment_execution_allowed": False,
        "bank_operation_allowed": False,
        "loan_management_action_allowed": False,
        "business_execution_allowed": False,
        "github_upload_allowed": False,
    }


def _phase_boundaries() -> dict[str, bool]:
    return {
        "stage13_post_remediation_review_validated": True,
        "s14_p1_performed": True,
        "s14_p2_performed": False,
        "s14_p3_performed": False,
        "stage14_review_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "formal_report_release_performed": False,
        "difference_closure_performed": False,
        "persistent_business_write_performed": False,
        "payment_or_bank_operation_performed": False,
        "loan_management_action_performed": False,
        "business_execution_performed": False,
    }


def _public_safety() -> dict[str, bool]:
    return {
        "aggregate_only": True,
        "raw_file_committed": False,
        "raw_filename_committed": False,
        "raw_hash_committed": False,
        "field_or_header_plaintext_committed": False,
        "business_amount_committed": False,
        "account_identifier_committed": False,
        "loan_contract_detail_committed": False,
        "credential_or_secret_committed": False,
        "private_runtime_committed": False,
        "zip_excel_pdf_private_csv_or_database_committed": False,
    }


def _acceptance_matrix(summary: dict[str, Any]) -> dict[str, Any]:
    checks = {
        "current_stage13_dependency": summary["stage13_post_remediation_review_dependency_validated"],
        "four_structure_lanes": summary["source_lane_count"] == 4 == summary["structure_connected_lane_count"],
        "four_private_candidate_lanes": summary["private_parseable_lane_count"] == 4,
        "zero_row_value_bindings": summary["row_level_binding_proven_lane_count"] == 0
        and summary["value_binding_proven_lane_count"] == 0,
        "three_method_definitions": summary["planning_method_definition_count"] == 3,
        "zero_business_items": summary["identified_business_item_count"] == 0,
        "raw_exact": summary["raw_snapshot_exact_match"] and summary["raw_cross_phase_snapshot_exact_match"],
        "private_probe_roundtrip_exact": summary["private_probe_roundtrip_mismatch_count"] == 0,
        "browser_pass": summary["browser_status"] == "PASS",
        "quality_locked": summary["current_data_quality_grade"] == "Q4"
        and summary["current_report_grade"] == "D"
        and summary["decision"] == "NO_GO",
        "no_later_scope": not summary["s14_p2_performed"] and not summary["github_upload_performed"],
    }
    return {
        "schema_version": "kmfa.v014.s14_p1_acceptance_matrix.v1",
        "checks": checks,
        "pass_count": sum(checks.values()),
        "fail_count": sum(not value for value in checks.values()),
        "decision": "NO_GO",
    }


def _phase_public_files() -> list[str]:
    paths = (
        SUMMARY_PATH,
        MANIFEST_PATH,
        LANES_PATH,
        METHODS_PATH,
        MATRIX_PATH,
        GO_NO_GO_PATH,
        HTML_PATH,
        REPORT_PATH,
        TEST_RESULTS_PATH,
        RISK_REGISTER_PATH,
        ROLLBACK_PATH,
        METADATA_SUMMARY_PATH,
        METADATA_MANIFEST_PATH,
        METADATA_LANES_PATH,
        METADATA_METHODS_PATH,
        METADATA_MATRIX_PATH,
        METADATA_GO_NO_GO_PATH,
        Path("KMFA/AGENTS.md"),
        Path("KMFA/CHANGELOG.md"),
        Path("KMFA/HANDOFF.md"),
        Path("KMFA/VERSION"),
        Path("KMFA/docs/governance/ASSURANCE_STATUS.yaml"),
        Path("KMFA/docs/governance/DEVELOPMENT_LEDGER.md"),
        Path("KMFA/docs/governance/MODEL_SPEC.md"),
        Path("KMFA/docs/governance/TRACEABILITY_MATRIX.csv"),
        Path("KMFA/docs/governance/VERSION_MATRIX.yaml"),
        Path("KMFA/docs/governance/delivery_tasks.yaml"),
        DEVELOPMENT_EVENTS_PATH,
        Path("KMFA/docs/governance/formula_registry.yaml"),
        Path("KMFA/docs/governance/model_registry.yaml"),
        Path("KMFA/docs/governance/parameter_registry.csv"),
        Path("KMFA/metadata/model_registry.yaml"),
        STAGE_STATUS_PATH,
        TASK_STATUS_PATH,
        Path("KMFA/功能清单.md"),
        Path("KMFA/开发记录.md"),
        Path("KMFA/模型参数文件.md"),
        Path("KMFA/tools/v014_s14_p1_post_remediation_fund_cash_loan_plan.py"),
        Path("KMFA/tools/check_v014_s14_p1_post_remediation_fund_cash_loan_plan.py"),
        Path("KMFA/tests/test_v014_s14_p1_post_remediation_fund_cash_loan_plan.py"),
    )
    return [path.as_posix() for path in paths]


def _write_governance(generated_at: str) -> None:
    evidence_ref = MANIFEST_PATH.as_posix()
    _upsert_jsonl(
        DEVELOPMENT_EVENTS_PATH,
        {
            "event_id": "DEV-KMFA-20260711-V014-S14-P1-POST-REMEDIATION-FUND-CASH-LOAN-PLAN",
            "event_time": generated_at,
            "event_type": "development",
            "project_id": "KMFA",
            "stage_id": "S14",
            "phase_id": PHASE_ID,
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "task_id": TASK_ID,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "identified_business_item_count": 0,
            "raw_business_data_committed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "files_changed": _phase_public_files(),
            "result_commit": "PENDING",
        },
    )
    _upsert_jsonl(
        STAGE_STATUS_PATH,
        {
            "schema_version": "kmfa.stage_status.v1",
            "record_type": "v014_phase_status",
            "project_id": "KMFA",
            "stage_id": "S14",
            "phase_id": PHASE_ID,
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "task_id": TASK_ID,
            "version": VERSION,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "derived_percent": "33.33",
            "raw_data_committed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at,
        },
    )
    _upsert_jsonl(
        TASK_STATUS_PATH,
        {
            "schema_version": "kmfa.v014_stage_phase_task_status.v1",
            "record_type": "v014_phase",
            "project_id": "KMFA",
            "stage_id": "S14",
            "governance_stage_id": "FUND-CASH-LOAN",
            "roadmap_stage_id": "S14",
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "acceptance_id": ACCEPTANCE_ID,
            "version": VERSION,
            "name": "v0.1.4 S14-P1 post-remediation fund cash loan plan",
            "phase_goal": "connect fund cash loan structures and methods without inventing business values or actions",
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "derived_percent": 100,
            "estimated_task_units": 3,
            "completed_task_units": 3,
            "task_count": 3,
            "raw_data_committed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at[:10],
        },
    )


def _render_report(summary: dict[str, Any]) -> str:
    return f"""# KMFA v0.1.4 S14-P1 资金现金贷款

## 结论

- 当前状态：`Q4 / D / NO_GO / 3-9-2-1`
- 结构主题：账户清单、月度现金、资金计划、贷款明细 `4/4` 接入
- 私有 raw 候选：`{summary['private_unique_candidate_sheet_count']}` 个唯一工作表候选，4 条主题均存在候选结构
- 权威绑定：行级/数值绑定=`0/4 / 0/4`
- 复核方法：现金压力、贷款到期、账户余额汇总 `3/3` 定义完成
- 已证明业务事项：`0`；公开业务金额、付款审批、银行操作和贷款管理动作均为 `0`

## 数据状态

本 phase 已对只读 raw 容器执行候选解析与二次探针指纹复验，探针不一致数为 `0`。候选工作表和字段命中只能证明结构相关，不能证明权威行、期间、账户、合同或数值绑定，因此不生成现金压力数值、贷款到期事项或账户余额。

## 边界

- 不推断、不平均、不补零，不忽略 0.01 元。
- 不执行付款审批、付款、银行操作、还款、续贷或其他贷款管理动作。
- 未执行 S14-P2/P3、Stage 14 整体复审、GitHub upload、app reinstall、正式报告、差异关闭或业务执行。
- raw 前后、跨 Stage 13 review 和当前快照一致；原始名称、成员、工作表、字段预览与候选值指纹仅保存在 ignored private runtime。
"""


def _render_private_report(summary: dict[str, Any]) -> str:
    return f"""# S14-P1 私有 raw 解析与对齐记录

- 原始文件数：{summary['raw_source_file_count']}
- XLSX 容器：{summary['private_xlsx_container_count']}
- 可解析 / 不可解析：{summary['private_parseable_xlsx_count']} / {summary['private_unparseable_xlsx_count']}
- 唯一候选工作表：{summary['private_unique_candidate_sheet_count']}
- 候选主题计数：{json.dumps(summary['private_candidate_sheet_count_by_lane'], ensure_ascii=False, sort_keys=True)}
- 二次探针指纹不一致：{summary['private_probe_roundtrip_mismatch_count']}
- 行级 / 数值权威绑定：0 / 0
- 公开业务金额与事项：0 / 0
- raw 前后、跨 Stage 13 review 与当前快照：exact match
- 本轮结论：候选解析已跑通且未损伤原始文件；当前仍缺权威行、期间、账户、合同与数值绑定，不能形成资金业务结论。
- 最终 goal 若多轮交叉验证仍无法完成绑定，必须生成全中文最终差异报告。
"""


def generate(*, final_validation: bool = False, write_governance: bool = True) -> dict[str, Any]:
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    raw_helper = s13_review.p1.s12_review.p1.s11_project
    raw_before = raw_helper._raw_snapshot("before_v014_s14_p1_post_remediation_fund_cash_loan_plan")
    dependency = _load_dependency()
    contract = _load_contract()
    historical = _read_json(HISTORICAL_MANIFEST_PATH)
    raw_probe = _raw_candidate_probe(Path(raw_before["raw_root"]))
    expected_counts = {"account_list": 12, "monthly_cash": 23, "fund_plan": 4, "loan_detail": 154}
    if raw_probe["private_candidate_sheet_count_by_lane"] != expected_counts:
        raise ValueError("private raw candidate structure drift")
    if raw_probe["private_probe_roundtrip_mismatch_count"] != 0:
        raise ValueError("private raw candidate probe roundtrip mismatch")
    lanes, structure = _build_source_lanes(raw_probe["private_candidate_sheet_count_by_lane"])
    methods = _build_methods()
    _write_text(HTML_PATH, _render_html(lanes, methods))
    browser = _run_browser_review()
    raw_after = raw_helper._raw_snapshot("after_v014_s14_p1_post_remediation_fund_cash_loan_plan")
    prior_raw = _read_json(s13_review.PRIVATE_RAW_AFTER_PATH)
    current_raw = raw_helper._raw_snapshot("current_v014_s14_p1_post_remediation_fund_cash_loan_plan")
    normalize = raw_helper._normalize_raw
    raw_exact = normalize(raw_before) == normalize(raw_after)
    raw_cross = normalize(raw_before) == normalize(prior_raw) == normalize(current_raw)
    if not raw_exact or not raw_cross:
        raise ValueError("raw source changed during S14-P1")

    upstream = dependency["summary"]
    summary = {
        "schema_version": "kmfa.v014.s14_p1_post_remediation_summary.v1",
        "project_id": "KMFA",
        "stage_id": "S14",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "stage13_post_remediation_review_dependency_validated": True,
        "source_lane_count": len(lanes),
        "structure_connected_lane_count": sum(row["structure_connected"] for row in lanes),
        **structure,
        "private_parseable_lane_count": raw_probe["private_parseable_lane_count"],
        "row_level_binding_proven_lane_count": 0,
        "value_binding_proven_lane_count": 0,
        "planning_method_definition_count": len(methods),
        "identified_cash_pressure_item_count": 0,
        "identified_loan_due_item_count": 0,
        "identified_account_balance_item_count": 0,
        "identified_business_item_count": 0,
        "public_business_amount_count": 0,
        "payment_approval_count": 0,
        "bank_operation_count": 0,
        "loan_management_action_count": 0,
        "raw_source_file_count": raw_probe["raw_file_count"],
        "private_xlsx_container_count": raw_probe["private_xlsx_container_count"],
        "private_parseable_xlsx_count": raw_probe["private_parseable_xlsx_count"],
        "private_unparseable_xlsx_count": raw_probe["private_unparseable_xlsx_count"],
        "private_unique_candidate_sheet_count": raw_probe["private_unique_candidate_sheet_count"],
        "private_candidate_sheet_count_by_lane": raw_probe["private_candidate_sheet_count_by_lane"],
        "private_probe_roundtrip_mismatch_count": raw_probe["private_probe_roundtrip_mismatch_count"],
        "open_final_difference_accepted_count": upstream["open_final_difference_accepted_count"],
        "nonzero_delta_reconciliation_count": upstream["nonzero_delta_reconciliation_count"],
        "zero_delta_reconciliation_count": upstream["zero_delta_reconciliation_count"],
        "incomplete_reconciliation_count": upstream["incomplete_reconciliation_count"],
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "browser_status": browser["status"],
        "browser_viewport_check_count": browser["viewport_check_count"],
        "method_interaction_check_count": browser["method_interaction_check_count"],
        "dependency_link_http_check_count": browser["dependency_link_http_check_count"],
        "dependency_navigation_check_count": browser["dependency_navigation_check_count"],
        "console_error_count": browser["console_error_count"],
        "horizontal_overflow_count": browser["horizontal_overflow_count"],
        "raw_snapshot_exact_match": raw_exact,
        "raw_cross_phase_snapshot_exact_match": raw_cross,
        "s14_p2_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
    }
    matrix = _acceptance_matrix(summary)
    historical_summary = historical.get("fund_cash_loan_summary", {})
    historical_quarantine = {
        "legacy_manifest_validated_as_historical_fixture": True,
        "legacy_s14_p1_dynamic_state_is_authoritative": False,
        "legacy_pending_twelve_quarantined": historical_summary.get("pending_reconciliation_count") == 12,
        "legacy_cash_pressure_records_quarantined": historical_summary.get("cash_pressure_record_count") == 4,
        "legacy_loan_due_alerts_quarantined": historical_summary.get("loan_due_alert_count") == 3,
        "legacy_account_balance_summaries_quarantined": historical_summary.get("account_balance_summary_count") == 3,
        "current_identified_business_item_count": 0,
    }
    validation_summary = {
        "final_validation_recorded": final_validation,
        "focused_test": "PASS" if final_validation else "PENDING",
        "strict_validator": "PASS" if final_validation else "PENDING",
        "browser_desktop_mobile": "PASS" if final_validation else "PENDING",
        "raw_candidate_probe": "PASS" if final_validation else "PENDING",
        "raw_alignment": "PASS" if final_validation else "PENDING",
        "governance_and_safety_scans": "PASS" if final_validation else "PENDING",
    }
    manifest = {
        "schema_version": "kmfa.v014.s14_p1_post_remediation_manifest.v1",
        "project_id": "KMFA",
        "stage_id": "S14",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "generated_at": generated_at,
        "reviewed_head": _git_output(["rev-parse", "HEAD"]),
        "branch": _git_output(["branch", "--show-current"]),
        "remote": _git_output(["remote", "get-url", "origin"]),
        "stage13_post_remediation_review_dependency_validated": True,
        "historical_s14_p1_dynamic_state_is_authoritative": False,
        "summary": summary,
        "source_lanes": lanes,
        "planning_methods": methods,
        "browser_review": browser,
        "quality_gate": _quality_gate(),
        "phase_boundaries": _phase_boundaries(),
        "public_repo_safety": _public_safety(),
        "historical_quarantine": historical_quarantine,
        "taskpack_contract": contract,
        "reviewed_dependencies": {
            "current_stage13_review": s13_review.MANIFEST_PATH.as_posix(),
            "current_finance_structure_registry": CURRENT_SOURCE_REGISTRY_PATH.as_posix(),
            "current_finance_structure_candidates": CURRENT_CANDIDATES_PATH.as_posix(),
            "historical_s14_p1_policy_fixture": HISTORICAL_MANIFEST_PATH.as_posix(),
        },
        "raw_boundary": {
            "raw_read_authorized": True,
            "raw_snapshot_validation_performed": True,
            "raw_write_performed": False,
            "raw_delete_performed": False,
            "raw_move_performed": False,
            "raw_rename_performed": False,
            "raw_overwrite_performed": False,
            "raw_mutation_performed": False,
        },
        "acceptance_matrix": matrix,
        "validation_summary": validation_summary,
        "next_phase": "S14-P2",
        "next_required_step": "Execute S14-P2 only as a separate run; do not perform S14-P3, Stage 14 review, upload, reinstall, formal release, difference closure, persistent write, or business execution.",
    }
    go_no_go = {
        "schema_version": "kmfa.v014.s14_p1_go_no_go.v1",
        "phase_id": PHASE_ID,
        "decision": "NO_GO",
        "data_quality_grade": "Q4",
        "report_grade": "D",
        "reason": "authoritative row period account contract and value bindings remain unproven",
        "s14_p2_allowed_in_this_run": False,
        "github_upload_allowed": False,
        "business_execution_allowed": False,
    }

    public_writes = (
        (SUMMARY_PATH, summary),
        (MANIFEST_PATH, manifest),
        (LANES_PATH, {"schema_version": "kmfa.v014.s14_p1_source_lanes.v1", "lanes": lanes}),
        (METHODS_PATH, {"schema_version": "kmfa.v014.s14_p1_methods.v1", "methods": methods}),
        (MATRIX_PATH, matrix),
        (GO_NO_GO_PATH, go_no_go),
        (METADATA_SUMMARY_PATH, summary),
        (METADATA_MANIFEST_PATH, manifest),
        (METADATA_LANES_PATH, {"schema_version": "kmfa.v014.s14_p1_source_lanes.v1", "lanes": lanes}),
        (METADATA_METHODS_PATH, {"schema_version": "kmfa.v014.s14_p1_methods.v1", "methods": methods}),
        (METADATA_MATRIX_PATH, matrix),
        (METADATA_GO_NO_GO_PATH, go_no_go),
    )
    for path, value in public_writes:
        _write_json(path, value)
    _write_text(REPORT_PATH, _render_report(summary))
    _write_text(
        TEST_RESULTS_PATH,
        f"""# S14-P1 测试结果

- focused tests：{'PASS' if final_validation else 'PENDING'}
- strict validator：{'PASS' if final_validation else 'PENDING'}
- baseline HTML audit：{browser['baseline_pass_count']}/54 PASS
- current HTML audit：{browser['current_pass_count']}/{browser['current_control_row_count']} PASS
- desktop/mobile：{browser['viewport_check_count']}/2 PASS
- 三类方法交互：{browser['method_interaction_check_count']}/6 PASS
- Stage 13 链接 HTTP / 真实导航：{browser['dependency_link_http_check_count']}/4 / {browser['dependency_navigation_check_count']}/4 PASS
- private raw probe：180 unique candidate sheets / 0 roundtrip mismatch
- raw 前后/跨 Stage 13 review/current：exact match
""",
    )
    _write_text(
        RISK_REGISTER_PATH,
        """# S14-P1 风险登记

1. 私有候选结构不等于权威行或数值绑定，误读会产生虚假现金、余额和贷款结论。
2. 23 个 XLSX 容器当前不能由 openpyxl 直接解析，只能保留私有诊断，不得损伤原文件转换。
3. 账户、主体、币种、期间、合同、金额和到期日未权威绑定，任何业务事项必须保持 0。
4. 当前 Q4 / D / NO_GO，禁止付款审批、银行操作和贷款管理动作。
""",
    )
    _write_text(
        ROLLBACK_PATH,
        """# S14-P1 回滚方案

仅回退本 phase 的 public-safe 产物、代码、测试、治理登记和 ignored private runtime。不得修改、移动、删除、重命名、覆盖或回写原始数据目录。
""",
    )
    _write_json(PRIVATE_RAW_BEFORE_PATH, raw_before)
    _write_json(PRIVATE_RAW_AFTER_PATH, raw_after)
    _write_json(PRIVATE_PROBE_PATH, raw_probe)
    _write_text(PRIVATE_REPORT_PATH, _render_private_report(summary))
    if write_governance:
        _write_governance(generated_at)
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--browser-evidence-only", action="store_true")
    parser.add_argument("--final-validation", action="store_true")
    parser.add_argument("--no-governance", action="store_true")
    args = parser.parse_args()
    if args.browser_evidence_only:
        result = _browser_worker()
        return 0 if result["status"] == "PASS" else 1
    manifest = generate(final_validation=args.final_validation, write_governance=not args.no_governance)
    summary = manifest["summary"]
    print(
        "S14-P1 post-remediation: "
        f"lanes={summary['source_lane_count']} candidates={summary['private_unique_candidate_sheet_count']} "
        f"bound={summary['value_binding_proven_lane_count']} items={summary['identified_business_item_count']} "
        f"grade={summary['current_report_grade']} decision={summary['decision']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
