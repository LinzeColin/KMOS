#!/usr/bin/env python3
"""Generate current public-safe KMFA v0.1.4 S14-P3 evidence."""

from __future__ import annotations

import argparse
import functools
import html
import io
import json
import os
import socketserver
import subprocess
import sys
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

from KMFA.tools import v014_s14_p2_post_remediation_invoice_tax_plan as p2
from KMFA.tools.check_v014_s14_p2_post_remediation_invoice_tax_plan import (
    validate_v014_s14_p2_post_remediation_invoice_tax_plan,
)


PHASE_ID = "V014_S14_P3_POST_REMEDIATION_POLICY_EVIDENCE_PLAN"
ROADMAP_PHASE_ID = "S14-P3"
TASK_ID = "KMFA-V014-S14-P3-POST-REMEDIATION-POLICY-EVIDENCE-PLAN-20260711"
ACCEPTANCE_ID = "ACC-V014-S14-P3-POST-REMEDIATION-POLICY-EVIDENCE-PLAN"
VERSION = "0.1.4-s14-p3-post-remediation-policy-evidence-plan"
STATUS = "completed_validated_local_only_s14_p3_evidence_gap_risk_no_go_upload_deferred"
DECISION = "NO_GO"
FORMULA_ID = "FORM-KMFA-V014-S14P3-POST-REMEDIATION-POLICY-EVIDENCE-PLAN-001"
PARAMETER_IDS = (
    "PARAM-KMFA-1758",
    "PARAM-KMFA-1759",
    "PARAM-KMFA-1760",
    "PARAM-KMFA-1761",
)
MODEL_REGISTRY_KEY = "kmfa_v014_s14_p3_post_remediation_policy_evidence_plan"

OUTPUT_DIR = Path("KMFA/stage_artifacts") / PHASE_ID
MACHINE_DIR = OUTPUT_DIR / "machine"
HUMAN_DIR = OUTPUT_DIR / "human"
HTML_DIR = OUTPUT_DIR / "exports/html"
SUMMARY_PATH = MACHINE_DIR / "policy_evidence_plan_summary.json"
MANIFEST_PATH = MACHINE_DIR / "policy_evidence_plan_manifest.json"
DIRECTORIES_PATH = MACHINE_DIR / "policy_evidence_directories_public_safe.json"
GAPS_PATH = MACHINE_DIR / "policy_evidence_gaps_public_safe.json"
RISKS_PATH = MACHINE_DIR / "policy_risk_tips_public_safe.json"
MATRIX_PATH = MACHINE_DIR / "policy_evidence_plan_acceptance_matrix.json"
GO_NO_GO_PATH = MACHINE_DIR / "policy_evidence_plan_go_no_go.json"
HTML_PATH = HTML_DIR / "policy_evidence_workbench.html"
REPORT_PATH = HUMAN_DIR / "policy_evidence_plan_report_zh.md"
TEST_RESULTS_PATH = HUMAN_DIR / "test_results_zh.md"
RISK_REGISTER_PATH = HUMAN_DIR / "risk_register_zh.md"
ROLLBACK_PATH = HUMAN_DIR / "rollback_plan_zh.md"

QUALITY_DIR = Path("KMFA/metadata/quality")
METADATA_SUMMARY_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_evidence_plan_summary.json"
METADATA_MANIFEST_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_evidence_plan_manifest.json"
METADATA_DIRECTORIES_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_evidence_directories_public_safe.json"
METADATA_GAPS_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_evidence_gaps_public_safe.json"
METADATA_RISKS_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_risk_tips_public_safe.json"
METADATA_MATRIX_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_evidence_plan_acceptance_matrix.json"
METADATA_GO_NO_GO_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_evidence_plan_go_no_go.json"

PRIVATE_DIR = Path("KMFA/.codex_private_runtime/v014_s14_p3_post_remediation_policy_evidence_plan")
PRIVATE_RAW_BEFORE_PATH = PRIVATE_DIR / "raw_immutability_before.json"
PRIVATE_RAW_AFTER_PATH = PRIVATE_DIR / "raw_immutability_after.json"
PRIVATE_PROBE_PATH = PRIVATE_DIR / "raw_policy_lexical_candidate_probe.json"
PRIVATE_REPORT_PATH = PRIVATE_DIR / "s14_p3_private_difference_report_zh.md"
PRIVATE_BROWSER_PATH = PRIVATE_DIR / "browser_verification.json"
PRIVATE_BASELINE_AUDIT_PATH = PRIVATE_DIR / "human_flow_baseline_audit.csv"
PRIVATE_CURRENT_AUDIT_PATH = PRIVATE_DIR / "current_policy_evidence_workbench_audit.csv"
PRIVATE_SCREENSHOT_DIR = PRIVATE_DIR / "screenshots"

ROADMAP_PATH = Path("KMFA/taskpack/v1_4/roadmap/02_KMFA_Codex_Development_Roadmap_18_Stages_v1_4.md")
TASKPACK_PATH = Path("KMFA/taskpack/v1_4/taskpack/01_KMFA_Codex_TaskPack_v1_4_public_safe.md")
HTML_BASELINE_ROOT = Path("KMFA/taskpack/v1_4/html_uiux")
CURRENT_SOURCE_REGISTRY_PATH = Path("KMFA/metadata/imports/v014_s07_p1_finance_support_source_registry.json")
CURRENT_CANDIDATES_PATH = Path("KMFA/metadata/schema_maps/v014_s07_p1_finance_field_candidates.jsonl")
HISTORICAL_MANIFEST_PATH = Path(
    "KMFA/stage_artifacts/V014_S14_P3_POLICY_EVIDENCE_PLAN/machine/policy_evidence_plan_manifest.json"
)

DEVELOPMENT_EVENTS_PATH = Path("KMFA/docs/governance/development_events.jsonl")
STAGE_STATUS_PATH = Path("KMFA/metadata/stage_status.jsonl")
TASK_STATUS_PATH = Path("KMFA/metadata/traceability/v1_4_stage_phase_task_status.jsonl")

DEPENDENCY_LINKS = {
    "fund-cash-loan": (
        "../../../V014_S14_P1_POST_REMEDIATION_FUND_CASH_LOAN_PLAN/exports/html/"
        "fund_cash_loan_workbench.html",
        "资金现金贷款工作台",
    ),
    "invoice-tax": (
        "../../../V014_S14_P2_POST_REMEDIATION_INVOICE_TAX_PLAN/exports/html/"
        "invoice_tax_plan_workbench.html",
        "开票纳税计划工作台",
    ),
    "monthly": (
        "../../../V014_S13_P1_POST_REMEDIATION_FINANCIAL_OPERATING_REPORT/exports/html/"
        "financial_operating_monthly_draft.html",
        "经营月报初稿",
    ),
    "cross-table": (
        "../../../V014_S13_P3_POST_REMEDIATION_CROSS_TABLE_REVIEW/exports/html/"
        "cross_table_quality_workbench.html",
        "跨表复核质量工作台",
    ),
}

PROGRAM_SPECS = (
    {
        "program_id": "small_tech_company",
        "visible_name": "科小",
        "finance_categories": ("r_and_d_expense", "tax"),
        "required_evidence_categories": (
            "enterprise_basic_qualification",
            "technical_activity_evidence",
            "r_and_d_project_record",
            "personnel_or_team_evidence",
        ),
        "gap_summary": "企业基础、研发活动、项目和人员证据均需权威索引与人工复核。",
        "risk_level": "medium",
        "risk_tip": "结构性目录不得解释为科小资格判断。",
    },
    {
        "program_id": "high_tech_enterprise",
        "visible_name": "高新",
        "finance_categories": ("r_and_d_expense", "tax", "invoice"),
        "required_evidence_categories": (
            "intellectual_property_evidence",
            "r_and_d_expense_collection",
            "technology_field_mapping",
            "science_technology_personnel_evidence",
            "high_tech_revenue_evidence",
        ),
        "gap_summary": "知识产权、研发费用、技术领域、科技人员和高新收入证据链未闭合。",
        "risk_level": "high",
        "risk_tip": "证据链未闭合时不得输出高新资格或准备度肯定结论。",
    },
    {
        "program_id": "specialized_refined_innovative",
        "visible_name": "专精特新",
        "finance_categories": ("operating_analysis", "r_and_d_expense"),
        "required_evidence_categories": (
            "specialization_evidence",
            "refinement_management_evidence",
            "innovation_output_evidence",
            "industry_chain_position_evidence",
        ),
        "gap_summary": "专业化、精细化、创新输出和产业链位置材料缺少权威闭环。",
        "risk_level": "medium",
        "risk_tip": "风险提示只能指向材料缺口与复核动作，不得替代资格判断。",
    },
    {
        "program_id": "little_giant",
        "visible_name": "小巨人",
        "finance_categories": ("operating_analysis", "r_and_d_expense"),
        "required_evidence_categories": (
            "specialized_refined_innovative_basis",
            "market_segment_evidence",
            "innovation_capability_evidence",
            "quality_efficiency_evidence",
            "compliance_and_growth_evidence",
        ),
        "gap_summary": "专精特新基础、细分市场、创新能力、质量效益和成长合规证据未闭合。",
        "risk_level": "high",
        "risk_tip": "当前只能提示材料准备风险，不得形成小巨人资格结论。",
    },
    {
        "program_id": "r_and_d_expense",
        "visible_name": "研发费用",
        "finance_categories": ("r_and_d_expense", "tax", "invoice"),
        "required_evidence_categories": (
            "r_and_d_project_ledger",
            "r_and_d_expense_detail",
            "personnel_time_evidence",
            "invoice_tax_support_evidence",
            "project_outcome_evidence",
        ),
        "gap_summary": "研发项目台账、费用明细、人员工时、票税支持和成果证据没有权威逐项绑定。",
        "risk_level": "high",
        "risk_tip": "研发费用结构候选不得替代税务申报、加计扣除或政策结论。",
    },
)

RAW_PROGRAM_TERMS = {
    "small_tech_company": ("科技型中小企业", "科小"),
    "high_tech_enterprise": ("高新技术企业认定", "高新技术企业", "高企认定"),
    "specialized_refined_innovative": ("专精特新",),
    "little_giant": ("小巨人",),
    "r_and_d_expense": ("研发费用", "研究开发费用", "研发项目", "研究开发项目", "研发人员"),
}


def _load_dependency() -> dict[str, Any]:
    manifest = validate_v014_s14_p2_post_remediation_invoice_tax_plan(
        require_private_evidence=True,
        require_browser_evidence=True,
        require_final_evidence=True,
    )
    summary = manifest["summary"]
    if manifest.get("next_phase") != "S14-P3":
        raise ValueError("S14-P2 must route to S14-P3")
    if summary.get("s14_p3_performed") is not False:
        raise ValueError("S14-P2 dependency already claims S14-P3")
    if summary.get("decision") != "NO_GO" or summary.get("current_report_grade") != "D":
        raise ValueError("S14-P2 quality state drift")
    return manifest


def _load_contract() -> dict[str, Any]:
    roadmap = ROADMAP_PATH.read_text(encoding="utf-8")
    taskpack = TASKPACK_PATH.read_text(encoding="utf-8")
    for token in (
        "S14-P3",
        "政策证据",
        "科小",
        "高新",
        "专精特新",
        "小巨人",
        "研发费用证据目录",
        "只输出证据缺口和风险提示",
        "不输出正式政策资格结论",
    ):
        if token not in roadmap:
            raise ValueError(f"roadmap token missing: {token}")
    for token in ("开票/纳税/税务政策线", "证据缺口和风险提示", "不做正式纳税申报", "研发费用 ↔ 项目/人员/成果证据"):
        if token not in taskpack:
            raise ValueError(f"taskpack token missing: {token}")
    return {
        "roadmap_read": True,
        "taskpack_read": True,
        "human_flow_baseline_read": True,
        "raw_read_only_contract_applied": True,
    }


def _build_directories(probe: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    registry = p2.p1._read_json(CURRENT_SOURCE_REGISTRY_PATH)
    candidates = p2.p1._read_jsonl(CURRENT_CANDIDATES_PATH)
    sources_by_category: dict[str, list[dict[str, Any]]] = {}
    candidates_by_category: dict[str, list[dict[str, Any]]] = {}
    for row in registry.get("sources", []):
        if isinstance(row, dict):
            sources_by_category.setdefault(str(row.get("finance_category")), []).append(row)
    for row in candidates:
        candidates_by_category.setdefault(str(row.get("finance_category")), []).append(row)

    private_counts = probe["private_policy_lexical_candidate_sheet_count_by_program"]
    directories: list[dict[str, Any]] = []
    unique_sources: set[str] = set()
    unique_candidates: set[str] = set()
    source_associations = 0
    candidate_associations = 0
    for sequence, spec in enumerate(PROGRAM_SPECS, 1):
        source_refs: set[str] = set()
        candidate_refs: set[str] = set()
        quality_states: list[dict[str, Any]] = []
        for category in spec["finance_categories"]:
            for source in sources_by_category.get(category, []):
                source_ref = str(source.get("source_ref") or "")
                if source_ref:
                    source_refs.add(source_ref)
                if source.get("read_only_parse") is not True:
                    raise ValueError(f"non-read-only source in {spec['program_id']}")
            for candidate in candidates_by_category.get(category, []):
                candidate_ref = str(candidate.get("canonical_field_ref") or candidate.get("candidate_id") or "")
                if candidate_ref:
                    candidate_refs.add(candidate_ref)
                quality = candidate.get("quality_state")
                if isinstance(quality, dict):
                    quality_states.append(quality)
        if not source_refs or not candidate_refs:
            raise ValueError(f"missing public-safe structure for {spec['program_id']}")
        if any(row.get("q4_human_confirmed") is True for row in quality_states):
            raise ValueError(f"unexpected Q4 evidence binding in {spec['program_id']}")
        unique_sources.update(source_refs)
        unique_candidates.update(candidate_refs)
        source_associations += len(source_refs)
        candidate_associations += len(candidate_refs)
        directories.append(
            {
                "program_id": spec["program_id"],
                "visible_name": spec["visible_name"],
                "directory_sequence": sequence,
                "directory_definition_registered": True,
                "required_evidence_categories": list(spec["required_evidence_categories"]),
                "required_evidence_category_count": len(spec["required_evidence_categories"]),
                "public_source_ref_count": len(source_refs),
                "public_structure_candidate_count": len(candidate_refs),
                "private_lexical_candidate_sheet_count": private_counts[spec["program_id"]],
                "lexical_candidate_observed": private_counts[spec["program_id"]] > 0,
                "authoritative_evidence_bound": False,
                "evidence_complete": False,
                "formal_policy_qualification_conclusion_allowed": False,
                "policy_score_allowed": False,
                "policy_application_submission_allowed": False,
                "subsidy_application_allowed": False,
                "formal_report_allowed": False,
                "business_decision_basis_allowed": False,
                "contains_business_amounts": False,
                "contains_field_or_header_plaintext": False,
                "directory_status": "definition_registered_authoritative_evidence_unbound",
            }
        )
    return directories, {
        "unique_public_source_ref_count": len(unique_sources),
        "program_source_association_count": source_associations,
        "unique_structure_candidate_count": len(unique_candidates),
        "program_structure_candidate_association_count": candidate_associations,
    }


def _build_gaps() -> list[dict[str, Any]]:
    return [
        {
            "program_id": spec["program_id"],
            "visible_name": spec["visible_name"],
            "gap_sequence": sequence,
            "gap_status": "authoritative_evidence_not_bound",
            "gap_summary": spec["gap_summary"],
            "evidence_gap_only": True,
            "authoritative_evidence_bound": False,
            "evidence_complete": False,
            "formal_policy_qualification_conclusion_allowed": False,
            "policy_score_allowed": False,
            "policy_application_submission_allowed": False,
            "subsidy_application_allowed": False,
            "formal_report_allowed": False,
            "business_decision_basis_allowed": False,
            "contains_business_amounts": False,
        }
        for sequence, spec in enumerate(PROGRAM_SPECS, 1)
    ]


def _build_risks() -> list[dict[str, Any]]:
    return [
        {
            "program_id": spec["program_id"],
            "visible_name": spec["visible_name"],
            "risk_sequence": sequence,
            "risk_level": spec["risk_level"],
            "risk_tip": spec["risk_tip"],
            "risk_tip_only": True,
            "required_control": "manual_authoritative_policy_evidence_review_before_any_conclusion",
            "formal_policy_qualification_conclusion_allowed": False,
            "policy_application_submission_allowed": False,
            "subsidy_application_allowed": False,
            "formal_report_allowed": False,
            "business_decision_basis_allowed": False,
        }
        for sequence, spec in enumerate(PROGRAM_SPECS, 1)
    ]


def _scan_workbook(
    payload: bytes,
    *,
    raw_path: Path,
    member_name: str | None,
    raw_index: int,
    member_index: int,
) -> tuple[int, list[dict[str, Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    records: list[dict[str, Any]] = []
    for sheet_index, sheet in enumerate(workbook.worksheets, 1):
        normalized_rows: list[list[Any]] = []
        searchable_parts = [str(sheet.title)]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(int(sheet.max_row or 0), 12),
            max_col=min(int(sheet.max_column or 0), 30),
            values_only=True,
        ):
            normalized = [p2.p1._normalize_cell(value) for value in row]
            while normalized and normalized[-1] is None:
                normalized.pop()
            normalized_rows.append(normalized)
            searchable_parts.extend(str(value) for value in normalized if value is not None)
        searchable = "\n".join(searchable_parts)
        matched_programs: list[str] = []
        matched_terms: dict[str, list[str]] = {}
        for program_id, terms in RAW_PROGRAM_TERMS.items():
            hits = [term for term in terms if term in searchable]
            if hits:
                matched_programs.append(program_id)
                matched_terms[program_id] = hits
        if not matched_programs:
            continue
        records.append(
            {
                "raw_index": raw_index,
                "raw_path_private": str(raw_path),
                "raw_filename_private": raw_path.name,
                "member_index": member_index,
                "member_name_private": member_name,
                "member_sha256": p2.p1._sha256_bytes(payload),
                "sheet_index": sheet_index,
                "sheet_name_private": sheet.title,
                "sheet_max_row": int(sheet.max_row or 0),
                "sheet_max_column": int(sheet.max_column or 0),
                "matched_programs": sorted(matched_programs),
                "matched_terms_private": matched_terms,
                "classification_fingerprint": p2.p1._sha256_json(
                    {"sheet_title": p2.p1._normalize_cell(sheet.title), "rows": normalized_rows}
                ),
            }
        )
    sheet_count = len(workbook.worksheets)
    workbook.close()
    return sheet_count, records


def _raw_candidate_probe(raw_root: Path) -> dict[str, Any]:
    raw_files = sorted(path for path in raw_root.rglob("*") if path.is_file())
    raw_records: list[dict[str, Any]] = []
    unparseable: list[dict[str, Any]] = []
    candidate_records: list[dict[str, Any]] = []
    xlsx_container_count = 0
    parseable_count = 0
    unparseable_count = 0
    parseable_sheet_count = 0
    roundtrip_mismatch_count = 0

    for raw_index, raw_path in enumerate(raw_files, 1):
        raw_bytes = raw_path.read_bytes()
        payloads: list[tuple[int, str | None, bytes]] = []
        if raw_path.suffix.lower() == ".xlsx":
            payloads.append((1, None, raw_bytes))
        elif raw_path.suffix.lower() == ".zip":
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
                payloads = [
                    (member_index, info.filename, archive.read(info))
                    for member_index, info in enumerate(archive.infolist(), 1)
                    if not info.is_dir() and Path(info.filename).suffix.lower() == ".xlsx"
                ]
        raw_records.append(
            {
                "raw_index": raw_index,
                "raw_path_private": str(raw_path),
                "raw_filename_private": raw_path.name,
                "raw_suffix": raw_path.suffix.lower(),
                "raw_size_bytes": raw_path.stat().st_size,
                "raw_sha256": p2.p1._sha256_bytes(raw_bytes),
                "xlsx_container_count": len(payloads),
            }
        )
        for member_index, member_name, payload in payloads:
            xlsx_container_count += 1
            try:
                first_sheet_count, first = _scan_workbook(
                    payload,
                    raw_path=raw_path,
                    member_name=member_name,
                    raw_index=raw_index,
                    member_index=member_index,
                )
                second_sheet_count, second = _scan_workbook(
                    payload,
                    raw_path=raw_path,
                    member_name=member_name,
                    raw_index=raw_index,
                    member_index=member_index,
                )
            except Exception as exc:
                unparseable_count += 1
                unparseable.append(
                    {
                        "raw_index": raw_index,
                        "member_index": member_index,
                        "member_name_private": member_name,
                        "member_sha256": p2.p1._sha256_bytes(payload),
                        "error_class": type(exc).__name__,
                    }
                )
                continue
            parseable_count += 1
            parseable_sheet_count += first_sheet_count
            if first_sheet_count != second_sheet_count:
                roundtrip_mismatch_count += 1
            first_map = {
                (row["sheet_index"], tuple(row["matched_programs"])): row["classification_fingerprint"]
                for row in first
            }
            second_map = {
                (row["sheet_index"], tuple(row["matched_programs"])): row["classification_fingerprint"]
                for row in second
            }
            all_keys = set(first_map) | set(second_map)
            roundtrip_mismatch_count += sum(first_map.get(key) != second_map.get(key) for key in all_keys)
            candidate_records.extend(first)

    counts = {
        program_id: sum(program_id in row["matched_programs"] for row in candidate_records)
        for program_id in RAW_PROGRAM_TERMS
    }
    return {
        "schema_version": "kmfa.v014.s14_p3_private_raw_policy_lexical_probe.v1",
        "classification": "private_raw_diagnostic_never_commit",
        "raw_root_private": str(raw_root),
        "raw_file_count": len(raw_files),
        "private_xlsx_container_count": xlsx_container_count,
        "private_parseable_xlsx_count": parseable_count,
        "private_unparseable_xlsx_count": unparseable_count,
        "private_parseable_sheet_count": parseable_sheet_count,
        "private_policy_lexical_candidate_sheet_count_by_program": counts,
        "private_unique_policy_lexical_candidate_sheet_count": len(candidate_records),
        "private_multi_program_candidate_sheet_count": sum(
            len(row["matched_programs"]) > 1 for row in candidate_records
        ),
        "private_lexical_candidate_covered_program_count": sum(count > 0 for count in counts.values()),
        "private_probe_roundtrip_mismatch_count": roundtrip_mismatch_count,
        "authoritative_evidence_bound_program_count": 0,
        "evidence_complete_program_count": 0,
        "raw_files_private": raw_records,
        "unparseable_xlsx_private": unparseable,
        "candidate_sheets_private": candidate_records,
    }


def _render_html(
    directories: list[dict[str, Any]],
    gaps: list[dict[str, Any]],
    risks: list[dict[str, Any]],
) -> str:
    gap_by_program = {row["program_id"]: row for row in gaps}
    risk_by_program = {row["program_id"]: row for row in risks}
    directory_rows = "".join(
        f"<tr><td>{html.escape(row['visible_name'])}</td><td>{row['required_evidence_category_count']}</td>"
        f"<td>{row['private_lexical_candidate_sheet_count']}</td><td>0</td></tr>"
        for row in directories
    )
    buttons = "".join(
        f'<button type="button" data-program-button="{row["program_id"]}">{html.escape(row["visible_name"])}</button>'
        for row in directories
    )
    panels = "".join(
        f'<section data-program-panel="{row["program_id"]}" hidden><div class="panel-head"><div><span>目录序列 {row["directory_sequence"]}</span>'
        f'<h3>{html.escape(row["visible_name"])}</h3></div><span class="tag blocked">资格结论 0</span></div>'
        f'<dl><div><dt>证据类别</dt><dd>{row["required_evidence_category_count"]} 类</dd></div>'
        f'<div><dt>词法候选表</dt><dd>{row["private_lexical_candidate_sheet_count"]}</dd></div>'
        '<div><dt>权威证据绑定</dt><dd>0</dd></div></dl>'
        f'<div class="gap"><strong>证据缺口</strong><p>{html.escape(gap_by_program[row["program_id"]]["gap_summary"])}</p></div>'
        f'<div class="risk"><strong>{html.escape(risk_by_program[row["program_id"]]["risk_level"].upper())} 风险</strong>'
        f'<p>{html.escape(risk_by_program[row["program_id"]]["risk_tip"])}</p></div>'
        '<div class="limit">不得据此输出政策资格、评分、申报或补贴申请结论。</div></section>'
        for row in directories
    )
    links = "".join(
        f'<a data-dependency-link="{link_id}" href="{href}">{label}</a>'
        for link_id, (href, label) in DEPENDENCY_LINKS.items()
    )
    labels = json.dumps(
        {row["program_id"]: row["visible_name"] for row in directories},
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>KMFA｜政策证据工作台</title>
  <style>
    *{{box-sizing:border-box}} body{{margin:0;background:#f2f5f7;color:#172522;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif;letter-spacing:0}} button,a{{font:inherit}}
    .topbar{{background:#12352e;color:#fff;border-bottom:3px solid #67b89e}} .topbar-inner{{max-width:1120px;margin:auto;min-height:66px;padding:12px 20px;display:flex;align-items:center;justify-content:space-between;gap:18px}} .brand{{display:flex;align-items:center;gap:12px}} .mark{{width:38px;height:38px;background:#fff;color:#12352e;display:grid;place-items:center;font-weight:800;border-radius:6px}} .brand strong{{display:block;font-size:16px}} .brand span{{display:block;font-size:12px;color:#cce5dc;margin-top:3px}}
    .top-links{{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}} .top-links a{{color:#fff;text-decoration:none;border:1px solid #628a7d;padding:8px 10px;border-radius:4px;font-size:13px}} main{{max-width:1120px;margin:auto;padding:22px 20px 44px}} .headline{{display:flex;justify-content:space-between;gap:24px;align-items:flex-start;margin-bottom:16px}} h1{{font-size:28px;margin:0 0 8px}} .subtitle{{margin:0;color:#5b6e68;line-height:1.7}}
    .badges{{display:flex;gap:8px;flex-wrap:wrap}} .tag{{display:inline-flex;align-items:center;border:1px solid #cad5d1;background:#fff;padding:5px 8px;border-radius:4px;font-size:12px;font-weight:700}} .tag.ok{{color:#137455;border-color:#a9d8c8;background:#edf8f4}} .tag.blocked{{color:#a43c36;border-color:#edb5b0;background:#fff2f0}} .notice{{border-left:4px solid #a43c36;background:#fff2f0;padding:13px 15px;margin-bottom:18px;color:#7e302c;line-height:1.6}}
    .metrics{{display:grid;grid-template-columns:repeat(4,1fr);border:1px solid #cfd9d6;background:#fff;margin-bottom:18px}} .metric{{padding:16px;border-right:1px solid #cfd9d6}} .metric:last-child{{border-right:0}} .metric strong{{display:block;font-size:25px;color:#13372f}} .metric span{{font-size:12px;color:#667873}} .band{{background:#fff;border:1px solid #cfd9d6;margin-bottom:18px}} .band-header{{display:flex;justify-content:space-between;align-items:center;padding:14px 16px;border-bottom:1px solid #d9e1de}} h2{{font-size:19px;margin:0}} h3{{font-size:20px;margin:5px 0 0}}
    .table-wrap{{overflow-x:auto}} table{{width:100%;border-collapse:collapse;min-width:700px}} th,td{{padding:12px 16px;text-align:left;border-bottom:1px solid #e2e8e6;font-size:13px}} th{{background:#f6f8f7;color:#5e716b}} tbody tr:last-child td{{border-bottom:0}} .program-layout{{display:grid;grid-template-columns:210px minmax(0,1fr)}} .program-nav{{padding:12px;background:#f4f7f6;border-right:1px solid #d9e1de}} .program-nav button{{display:block;width:100%;min-height:42px;text-align:left;border:1px solid transparent;background:transparent;padding:9px 10px;color:#1c322c;cursor:pointer;border-radius:4px}} .program-nav button.active{{background:#fff;border-color:#78ae9d;color:#116b51;font-weight:700}}
    .program-panels{{padding:20px}} .panel-head{{display:flex;justify-content:space-between;gap:18px;align-items:flex-start}} .panel-head span{{font-size:12px;color:#6a7d77}} dl{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:20px 0}} dl div{{border-left:3px solid #8ebfaf;padding-left:10px}} dt{{font-size:12px;color:#6a7d77}} dd{{margin:4px 0 0;font-weight:700}} .gap,.risk{{padding:12px 14px;margin:10px 0;border:1px solid #d6dfdc;background:#f7f9f8}} .gap p,.risk p{{margin:5px 0 0;line-height:1.6}} .risk{{border-color:#e4c68e;background:#fff8e9}} .limit{{border:1px solid #edb5b0;background:#fff2f0;color:#8d342f;padding:12px}}
    .footer-band{{display:flex;justify-content:space-between;gap:16px;align-items:center;padding:13px 16px;background:#f6f8f7;border-top:1px solid #d9e1de}} .footer-band p{{margin:0;color:#5b6e68;font-size:13px}} .footer-band a{{color:#136c94;text-decoration:none;margin-left:12px;font-size:13px}} footer{{color:#687b75;font-size:12px;line-height:1.7}} @media(max-width:720px){{.topbar-inner,.headline,.footer-band{{display:block}} .top-links{{justify-content:flex-start;margin-top:12px}} main{{padding:18px 12px 36px}} h1{{font-size:24px}} .badges{{margin-top:12px}} .metrics{{grid-template-columns:1fr 1fr}} .metric:nth-child(2){{border-right:0}} .metric:nth-child(-n+2){{border-bottom:1px solid #cfd9d6}} table{{min-width:0;table-layout:fixed}} th,td{{padding:9px 6px;font-size:11px;word-break:break-word}} .program-layout{{grid-template-columns:1fr}} .program-nav{{border-right:0;border-bottom:1px solid #d9e1de;display:grid;grid-template-columns:repeat(2,1fr);gap:5px}} .program-nav button{{text-align:center;padding:7px 4px}} dl{{grid-template-columns:1fr}} .footer-band .links{{margin-top:10px}} .footer-band a{{display:inline-block;margin:4px 10px 4px 0}}}}
  </style>
</head>
<body data-page-id="s14-p3-policy-evidence" data-active-program="small_tech_company" data-ui-ready="false">
  <header class="topbar"><div class="topbar-inner"><div class="brand"><div class="mark">KM</div><div><strong>KMFA 经营分析系统</strong><span>S14-P3 · 政策证据</span></div></div><nav class="top-links" aria-label="阶段页面导航">{links}</nav></div></header>
  <main>
    <section class="headline"><div><h1>政策证据工作台</h1><p class="subtitle">仅登记证据目录、缺口和风险；词法候选不代表材料完整、适用条件或政策资格。</p></div><div class="badges"><span class="tag ok">Q4 / D</span><span class="tag blocked">NO_GO</span><span class="tag">内部复核</span></div></section>
    <div class="notice"><strong>门禁：</strong> 5 类政策目录均未完成权威证据绑定；不得输出资格、评分、申报或补贴申请结论。</div>
    <section class="metrics" aria-label="状态摘要"><div class="metric"><strong>5 / 5</strong><span>目录定义已登记</span></div><div class="metric"><strong>2 / 5</strong><span>存在私有词法候选</span></div><div class="metric"><strong>0 / 5</strong><span>权威证据已绑定</span></div><div class="metric"><strong>0</strong><span>正式资格结论</span></div></section>
    <section class="band"><div class="band-header"><h2>政策证据目录</h2><span class="tag blocked">只登记缺口</span></div><div class="table-wrap"><table><thead><tr><th>目录</th><th>证据类别</th><th>私有词法候选表</th><th>权威绑定</th></tr></thead><tbody>{directory_rows}</tbody></table></div></section>
    <section class="band"><div class="band-header"><h2>缺口与风险复核</h2><span class="tag blocked">5 缺口 · 5 风险</span></div><div class="program-layout"><nav class="program-nav" aria-label="政策目录">{buttons}</nav><div class="program-panels">{panels}</div></div><div class="footer-band"><p id="interaction-status" aria-live="polite">已显示“科小”证据缺口；当前不形成资格结论。</p><div class="links">{links}</div></div></section>
    <footer>本 phase 仅完成 S14-P3；Stage 14 整体复审、政策资格结论、政策申报、补贴申请、GitHub upload、app reinstall、正式报告、差异关闭和业务执行均未执行。</footer>
  </main>
  <script>
    const labels={labels};
    let actionSequence=0;
    function activate(id){{document.body.dataset.activeProgram=id;document.body.dataset.lastAction="program:"+id+":"+String(++actionSequence);document.querySelectorAll("[data-program-button]").forEach(function(button){{button.classList.toggle("active",button.dataset.programButton===id);}});document.querySelectorAll("[data-program-panel]").forEach(function(panel){{panel.hidden=panel.dataset.programPanel!==id;}});document.getElementById("interaction-status").textContent="已显示“"+labels[id]+"”证据缺口；当前不形成资格结论。";}}
    document.querySelectorAll("[data-program-button]").forEach(function(button){{button.addEventListener("click",function(){{activate(button.dataset.programButton);}});}});
    activate("small_tech_company");document.body.dataset.uiReady="true";
  </script>
</body>
</html>"""


def _browser_worker() -> dict[str, Any]:
    from playwright.sync_api import sync_playwright

    root = Path("KMFA/stage_artifacts").resolve()
    handler = functools.partial(p2.p1._QuietHandler, directory=str(root))
    server = socketserver.TCPServer(("127.0.0.1", 0), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    base = f"http://127.0.0.1:{server.server_address[1]}/"
    workbench_url = urljoin(base, f"{PHASE_ID}/exports/html/{HTML_PATH.name}")
    viewport_checks: list[dict[str, Any]] = []
    program_checks: list[dict[str, Any]] = []
    http_checks: list[dict[str, Any]] = []
    navigation_checks: list[dict[str, Any]] = []
    PRIVATE_SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            for mode, width, height in (("desktop", 1440, 900), ("mobile", 390, 844)):
                page = browser.new_page(viewport={"width": width, "height": height})
                errors: list[str] = []
                page.on(
                    "console",
                    lambda message, errors=errors: errors.append(message.text) if message.type == "error" else None,
                )
                page.goto(workbench_url, wait_until="networkidle")
                page.wait_for_function("document.body.dataset.uiReady === 'true'")
                body_text = page.locator("body").inner_text()
                viewport_checks.append(
                    {
                        "mode": mode,
                        "width": width,
                        "height": height,
                        "marker_visible": page.get_by_role("heading", name="政策证据工作台").is_visible(),
                        "d_no_go_visible": "Q4 / D" in body_text and "NO_GO" in body_text,
                        "console_error_count": len(errors),
                        "no_horizontal_overflow": page.evaluate(
                            "document.documentElement.scrollWidth <= document.documentElement.clientWidth"
                        ),
                    }
                )
                for program in PROGRAM_SPECS:
                    page.locator(f'[data-program-button="{program["program_id"]}"]').click()
                    program_checks.append(
                        {
                            "mode": mode,
                            "program_id": program["program_id"],
                            "passed": page.locator(f'[data-program-panel="{program["program_id"]}"]').is_visible()
                            and page.locator("body").get_attribute("data-active-program") == program["program_id"]
                            and program["visible_name"] in page.locator("#interaction-status").inner_text(),
                        }
                    )
                page.screenshot(path=str(PRIVATE_SCREENSHOT_DIR / f"policy_evidence_{mode}.png"), full_page=True)
                page.close()

            request = playwright.request.new_context()
            for link_id, (_, marker) in DEPENDENCY_LINKS.items():
                page = browser.new_page(viewport={"width": 1280, "height": 900})
                page.goto(workbench_url, wait_until="networkidle")
                href = page.locator(f'a[data-dependency-link="{link_id}"]').first.get_attribute("href") or ""
                target = urljoin(workbench_url, href)
                response = request.get(target)
                http_checks.append({"link_id": link_id, "status": response.status, "passed": response.ok})
                page.locator(f'a[data-dependency-link="{link_id}"]').first.click()
                page.wait_for_load_state("networkidle")
                navigation_checks.append(
                    {"link_id": link_id, "marker": marker, "passed": marker in page.locator("body").inner_text()}
                )
                page.close()
            request.dispose()
            browser.close()
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)
    result = {
        "status": "PASS",
        "viewport_checks": viewport_checks,
        "program_interaction_checks": program_checks,
        "dependency_link_http_checks": http_checks,
        "dependency_navigation_checks": navigation_checks,
    }
    if not (
        len(viewport_checks) == 2
        and all(
            row["marker_visible"]
            and row["d_no_go_visible"]
            and row["console_error_count"] == 0
            and row["no_horizontal_overflow"]
            for row in viewport_checks
        )
        and len(program_checks) == 10
        and all(row["passed"] for row in program_checks)
        and len(http_checks) == 4
        and all(row["passed"] for row in http_checks)
        and len(navigation_checks) == 4
        and all(row["passed"] for row in navigation_checks)
    ):
        result["status"] = "FAIL"
    p2.p1._write_json(PRIVATE_BROWSER_PATH, result)
    return result


def _run_browser_review() -> dict[str, Any]:
    python = os.environ.get("KMFA_AUDIT_PYTHON", sys.executable)
    result = subprocess.run(
        [python, __file__, "--browser-evidence-only"],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
        env={**os.environ, "PYTHONDONTWRITEBYTECODE": "1", "PYTHONPATH": "."},
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip())
    browser = p2.p1._read_json(PRIVATE_BROWSER_PATH)
    if browser.get("status") != "PASS":
        raise RuntimeError("S14-P3 browser validation failed")
    helper = p2.p1.s13_review.p1.s12_review.p1.s11_home
    baseline = helper._run_html_audit(HTML_BASELINE_ROOT, PRIVATE_BASELINE_AUDIT_PATH)
    current = helper._run_html_audit(HTML_DIR, PRIVATE_CURRENT_AUDIT_PATH)
    if baseline != {
        "file_count": 6,
        "control_row_count": 54,
        "pass_count": 54,
        "warn_count": 0,
        "fail_count": 0,
    }:
        raise RuntimeError("v1.4 HTML baseline drift")
    if current["fail_count"] or current["warn_count"] or current["pass_count"] != current["control_row_count"]:
        raise RuntimeError("S14-P3 HTML audit failed")
    return {
        "status": "PASS",
        "baseline_file_count": baseline["file_count"],
        "baseline_control_row_count": baseline["control_row_count"],
        "baseline_pass_count": baseline["pass_count"],
        "baseline_warn_count": baseline["warn_count"],
        "baseline_fail_count": baseline["fail_count"],
        "current_page_count": current["file_count"],
        "current_control_row_count": current["control_row_count"],
        "current_pass_count": current["pass_count"],
        "current_warn_count": current["warn_count"],
        "current_fail_count": current["fail_count"],
        "viewport_check_count": len(browser["viewport_checks"]),
        "program_interaction_check_count": len(browser["program_interaction_checks"]),
        "dependency_link_http_check_count": len(browser["dependency_link_http_checks"]),
        "dependency_navigation_check_count": len(browser["dependency_navigation_checks"]),
        "console_error_count": sum(row["console_error_count"] for row in browser["viewport_checks"]),
        "horizontal_overflow_count": sum(not row["no_horizontal_overflow"] for row in browser["viewport_checks"]),
    }


def _quality_gate() -> dict[str, Any]:
    return {
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "decision": "NO_GO",
        "release_permission": "blocked",
        "policy_evidence_directory_definition_allowed": True,
        "evidence_gap_signal_allowed": True,
        "risk_tip_allowed": True,
        "authoritative_evidence_bound_program_use_allowed": False,
        "evidence_complete_program_use_allowed": False,
        "formal_policy_qualification_conclusion_allowed": False,
        "policy_score_allowed": False,
        "policy_application_submission_allowed": False,
        "policy_filing_allowed": False,
        "subsidy_application_allowed": False,
        "derived_amount_calculation_allowed": False,
        "formal_report_allowed": False,
        "business_decision_basis_allowed": False,
        "business_execution_allowed": False,
        "github_upload_allowed": False,
    }


def _phase_boundaries() -> dict[str, bool]:
    return {
        "s14_p1_post_remediation_validated": True,
        "s14_p2_post_remediation_validated": True,
        "s14_p3_performed": True,
        "stage14_review_performed": False,
        "formal_policy_qualification_conclusion_performed": False,
        "policy_application_submission_performed": False,
        "subsidy_application_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "formal_report_release_performed": False,
        "difference_closure_performed": False,
        "persistent_business_write_performed": False,
        "business_execution_performed": False,
    }


def _public_safety() -> dict[str, bool]:
    return {
        "aggregate_only": True,
        "raw_file_committed": False,
        "raw_filename_committed": False,
        "raw_hash_committed": False,
        "field_or_header_plaintext_committed": False,
        "business_amount_committed": False,
        "policy_evidence_material_committed": False,
        "policy_application_detail_committed": False,
        "project_personnel_outcome_plaintext_committed": False,
        "credential_or_secret_committed": False,
        "private_runtime_committed": False,
        "zip_excel_pdf_private_csv_or_database_committed": False,
    }


def _acceptance_matrix(summary: dict[str, Any]) -> dict[str, Any]:
    checks = {
        "current_s14_p2_dependency": summary["s14_p2_post_remediation_dependency_validated"],
        "five_directory_definitions": summary["policy_program_count"] == 5
        == summary["evidence_directory_definition_count"],
        "twenty_three_evidence_categories": summary["required_evidence_category_total_count"] == 23,
        "zero_authoritative_evidence_binding": summary["authoritative_evidence_bound_program_count"] == 0
        and summary["evidence_complete_program_count"] == 0,
        "five_gap_records": summary["evidence_gap_count"] == 5,
        "five_risk_tips": summary["risk_tip_count"] == 5,
        "zero_policy_conclusions_actions": summary["formal_policy_qualification_conclusion_count"] == 0
        and summary["policy_application_submission_count"] == 0
        and summary["subsidy_application_count"] == 0,
        "raw_exact": summary["raw_snapshot_exact_match"] and summary["raw_cross_phase_snapshot_exact_match"],
        "private_probe_roundtrip_exact": summary["private_probe_roundtrip_mismatch_count"] == 0,
        "browser_pass": summary["browser_status"] == "PASS",
        "quality_locked": summary["current_data_quality_grade"] == "Q4"
        and summary["current_report_grade"] == "D"
        and summary["decision"] == "NO_GO",
        "no_review_release_execution": not summary["stage14_review_performed"]
        and not summary["github_upload_performed"]
        and not summary["business_execution_performed"],
    }
    return {
        "schema_version": "kmfa.v014.s14_p3_acceptance_matrix.v1",
        "checks": checks,
        "pass_count": sum(checks.values()),
        "fail_count": sum(not value for value in checks.values()),
        "decision": "NO_GO",
    }


def _phase_public_files() -> list[str]:
    paths = (
        SUMMARY_PATH,
        MANIFEST_PATH,
        DIRECTORIES_PATH,
        GAPS_PATH,
        RISKS_PATH,
        MATRIX_PATH,
        GO_NO_GO_PATH,
        HTML_PATH,
        REPORT_PATH,
        TEST_RESULTS_PATH,
        RISK_REGISTER_PATH,
        ROLLBACK_PATH,
        METADATA_SUMMARY_PATH,
        METADATA_MANIFEST_PATH,
        METADATA_DIRECTORIES_PATH,
        METADATA_GAPS_PATH,
        METADATA_RISKS_PATH,
        METADATA_MATRIX_PATH,
        METADATA_GO_NO_GO_PATH,
        Path("KMFA/AGENTS.md"),
        Path("KMFA/CHANGELOG.md"),
        Path("KMFA/HANDOFF.md"),
        Path("KMFA/VERSION"),
        Path("KMFA/docs/governance/ASSURANCE_STATUS.yaml"),
        Path("KMFA/docs/governance/DEVELOPMENT_LEDGER.md"),
        Path("KMFA/docs/governance/MODEL_SPEC.md"),
        Path("KMFA/docs/governance/TRACEABILITY_MATRIX.csv"),
        Path("KMFA/docs/governance/VERSION_MATRIX.yaml"),
        Path("KMFA/docs/governance/delivery_tasks.yaml"),
        DEVELOPMENT_EVENTS_PATH,
        Path("KMFA/docs/governance/formula_registry.yaml"),
        Path("KMFA/docs/governance/model_registry.yaml"),
        Path("KMFA/docs/governance/parameter_registry.csv"),
        Path("KMFA/metadata/model_registry.yaml"),
        STAGE_STATUS_PATH,
        TASK_STATUS_PATH,
        Path("KMFA/功能清单.md"),
        Path("KMFA/开发记录.md"),
        Path("KMFA/模型参数文件.md"),
        Path("KMFA/tools/v014_s14_p3_post_remediation_policy_evidence_plan.py"),
        Path("KMFA/tools/check_v014_s14_p3_post_remediation_policy_evidence_plan.py"),
        Path("KMFA/tests/test_v014_s14_p3_post_remediation_policy_evidence_plan.py"),
    )
    return [path.as_posix() for path in paths]


def _write_governance(generated_at: str) -> None:
    evidence_ref = MANIFEST_PATH.as_posix()
    p2._upsert_jsonl(
        DEVELOPMENT_EVENTS_PATH,
        {
            "event_id": "DEV-KMFA-20260711-V014-S14-P3-POST-REMEDIATION-POLICY-EVIDENCE-PLAN",
            "event_time": generated_at,
            "event_type": "development",
            "project_id": "KMFA",
            "stage_id": "S14",
            "phase_id": PHASE_ID,
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "task_id": TASK_ID,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "evidence_gap_count": 5,
            "risk_tip_count": 5,
            "formal_policy_qualification_conclusion_count": 0,
            "identified_business_item_count": 0,
            "raw_business_data_committed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "files_changed": _phase_public_files(),
            "result_commit": "PENDING",
        },
    )
    p2._upsert_jsonl(
        STAGE_STATUS_PATH,
        {
            "schema_version": "kmfa.stage_status.v1",
            "record_type": "v014_phase_status",
            "project_id": "KMFA",
            "stage_id": "S14",
            "phase_id": PHASE_ID,
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "task_id": TASK_ID,
            "version": VERSION,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "derived_percent": "100.00",
            "raw_data_committed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at,
        },
    )
    p2._upsert_jsonl(
        TASK_STATUS_PATH,
        {
            "schema_version": "kmfa.v014_stage_phase_task_status.v1",
            "record_type": "v014_phase",
            "project_id": "KMFA",
            "stage_id": "S14",
            "governance_stage_id": "POLICY-EVIDENCE",
            "roadmap_stage_id": "S14",
            "roadmap_phase_id": ROADMAP_PHASE_ID,
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "acceptance_id": ACCEPTANCE_ID,
            "version": VERSION,
            "name": "v0.1.4 S14-P3 post-remediation policy evidence plan",
            "phase_goal": "register policy evidence directories gaps and risks without qualification conclusions or actions",
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "derived_percent": 100,
            "estimated_task_units": 3,
            "completed_task_units": 3,
            "task_count": 3,
            "raw_data_committed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at[:10],
        },
    )


def _render_report(summary: dict[str, Any]) -> str:
    return f"""# KMFA v0.1.4 S14-P3 政策证据

## 结论

- 当前状态：Q4 / D / NO_GO / 3-9-2-1
- 目录定义：科小、高新、专精特新、小巨人、研发费用 5/5
- 必需证据类别：23 类
- 私有词法候选覆盖：2/5 个目录；唯一候选工作表 {summary['private_unique_policy_lexical_candidate_sheet_count']}
- 权威证据绑定与证据完整目录：0/5 / 0/5
- 证据缺口与风险提示：5 / 5
- 正式资格结论、评分、政策申报和补贴申请：0 / 0 / 0 / 0

## 数据状态

本 phase 对只读 raw 执行精确词组候选探针与二次指纹复验，探针不一致数为 {summary['private_probe_roundtrip_mismatch_count']}。词法候选仅说明工作表前 12 行存在政策相关词组，不证明材料身份、有效期、主体、项目、人员、成果、金额、适用条件或资格。

## 边界

- 只输出证据目录、缺口和风险提示，不输出正式政策资格结论。
- 不计算政策评分，不执行政策申报、补贴申请、纳税申报或其他外部动作。
- 未执行 Stage 14 整体复审、GitHub upload、app reinstall、正式报告、差异关闭或业务执行。
- raw 前后、跨 S14-P2 和当前快照一致；原始名称、工作表、字段、命中词和值指纹仅保存在 ignored private runtime。
"""


def _render_private_report(summary: dict[str, Any]) -> str:
    return f"""# S14-P3 私有 raw 政策证据差异记录

- 原始文件数：{summary['raw_source_file_count']}
- XLSX 容器：{summary['private_xlsx_container_count']}
- 可解析 / 不可解析：{summary['private_parseable_xlsx_count']} / {summary['private_unparseable_xlsx_count']}
- 可解析工作表：{summary['private_parseable_sheet_count']}
- 各目录词法候选：{json.dumps(summary['private_policy_lexical_candidate_sheet_count_by_program'], ensure_ascii=False, sort_keys=True)}
- 唯一候选 / 跨目录候选 / 覆盖目录：{summary['private_unique_policy_lexical_candidate_sheet_count']} / {summary['private_multi_program_candidate_sheet_count']} / {summary['private_lexical_candidate_covered_program_count']}
- 二次探针指纹不一致：{summary['private_probe_roundtrip_mismatch_count']}
- 权威证据绑定 / 证据完整目录：0 / 0
- 正式资格结论 / 评分 / 申报 / 补贴申请：0 / 0 / 0 / 0
- raw 前后、跨 S14-P2 与当前快照：exact match
- 当前差异原因：词法命中无法证明材料身份、完整性、有效期、主体、项目、人员、成果、金额、适用条件或政策资格。
- 本轮结论：目录、缺口与风险流程已跑通且未损伤原始文件；不得把词法候选转换为政策资格或申报结论。
- 最终 goal 若多轮交叉验证仍无法完成权威证据绑定，必须生成全中文最终差异报告。
"""


def generate(*, final_validation: bool = False, write_governance: bool = True) -> dict[str, Any]:
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    raw_helper = p2.p1.s13_review.p1.s12_review.p1.s11_project
    raw_before = raw_helper._raw_snapshot("before_v014_s14_p3_post_remediation_policy_evidence_plan")
    dependency = _load_dependency()
    contract = _load_contract()
    historical = p2.p1._read_json(HISTORICAL_MANIFEST_PATH)
    raw_probe = _raw_candidate_probe(Path(raw_before["raw_root"]))
    expected_probe = {
        "raw_file_count": 5,
        "private_xlsx_container_count": 48,
        "private_parseable_xlsx_count": 25,
        "private_unparseable_xlsx_count": 23,
        "private_parseable_sheet_count": 4198,
        "private_policy_lexical_candidate_sheet_count_by_program": {
            "small_tech_company": 0,
            "high_tech_enterprise": 1,
            "specialized_refined_innovative": 0,
            "little_giant": 0,
            "r_and_d_expense": 3830,
        },
        "private_unique_policy_lexical_candidate_sheet_count": 3830,
        "private_multi_program_candidate_sheet_count": 1,
        "private_lexical_candidate_covered_program_count": 2,
        "private_probe_roundtrip_mismatch_count": 0,
        "authoritative_evidence_bound_program_count": 0,
        "evidence_complete_program_count": 0,
    }
    for key, expected in expected_probe.items():
        if raw_probe.get(key) != expected:
            raise ValueError(f"private raw policy lexical structure drift: {key}")
    directories, structure = _build_directories(raw_probe)
    gaps = _build_gaps()
    risks = _build_risks()
    p2.p1._write_text(HTML_PATH, _render_html(directories, gaps, risks))
    browser = _run_browser_review()
    raw_after = raw_helper._raw_snapshot("after_v014_s14_p3_post_remediation_policy_evidence_plan")
    prior_raw = p2.p1._read_json(p2.PRIVATE_RAW_AFTER_PATH)
    current_raw = raw_helper._raw_snapshot("current_v014_s14_p3_post_remediation_policy_evidence_plan")
    normalize = raw_helper._normalize_raw
    raw_exact = normalize(raw_before) == normalize(raw_after)
    raw_cross = normalize(raw_before) == normalize(prior_raw) == normalize(current_raw)
    if not raw_exact or not raw_cross:
        raise ValueError("raw source changed during S14-P3")

    upstream = dependency["summary"]
    summary = {
        "schema_version": "kmfa.v014.s14_p3_post_remediation_summary.v1",
        "project_id": "KMFA",
        "stage_id": "S14",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "s14_p2_post_remediation_dependency_validated": True,
        "policy_program_count": len(PROGRAM_SPECS),
        "evidence_directory_definition_count": len(directories),
        "required_evidence_category_total_count": sum(
            row["required_evidence_category_count"] for row in directories
        ),
        **structure,
        "authoritative_evidence_bound_program_count": 0,
        "evidence_complete_program_count": 0,
        "evidence_gap_count": len(gaps),
        "risk_tip_count": len(risks),
        "formal_policy_qualification_conclusion_count": 0,
        "policy_score_count": 0,
        "policy_application_submission_count": 0,
        "subsidy_application_count": 0,
        "identified_business_item_count": 0,
        "public_business_amount_count": 0,
        "raw_source_file_count": raw_probe["raw_file_count"],
        "private_xlsx_container_count": raw_probe["private_xlsx_container_count"],
        "private_parseable_xlsx_count": raw_probe["private_parseable_xlsx_count"],
        "private_unparseable_xlsx_count": raw_probe["private_unparseable_xlsx_count"],
        "private_parseable_sheet_count": raw_probe["private_parseable_sheet_count"],
        "private_policy_lexical_candidate_sheet_count_by_program": raw_probe[
            "private_policy_lexical_candidate_sheet_count_by_program"
        ],
        "private_unique_policy_lexical_candidate_sheet_count": raw_probe[
            "private_unique_policy_lexical_candidate_sheet_count"
        ],
        "private_multi_program_candidate_sheet_count": raw_probe[
            "private_multi_program_candidate_sheet_count"
        ],
        "private_lexical_candidate_covered_program_count": raw_probe[
            "private_lexical_candidate_covered_program_count"
        ],
        "private_probe_roundtrip_mismatch_count": raw_probe["private_probe_roundtrip_mismatch_count"],
        "open_final_difference_accepted_count": upstream["open_final_difference_accepted_count"],
        "nonzero_delta_reconciliation_count": upstream["nonzero_delta_reconciliation_count"],
        "zero_delta_reconciliation_count": upstream["zero_delta_reconciliation_count"],
        "incomplete_reconciliation_count": upstream["incomplete_reconciliation_count"],
        "current_data_quality_grade": "Q4",
        "current_report_grade": "D",
        "browser_status": browser["status"],
        "browser_viewport_check_count": browser["viewport_check_count"],
        "program_interaction_check_count": browser["program_interaction_check_count"],
        "dependency_link_http_check_count": browser["dependency_link_http_check_count"],
        "dependency_navigation_check_count": browser["dependency_navigation_check_count"],
        "console_error_count": browser["console_error_count"],
        "horizontal_overflow_count": browser["horizontal_overflow_count"],
        "raw_snapshot_exact_match": raw_exact,
        "raw_cross_phase_snapshot_exact_match": raw_cross,
        "stage14_review_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
    }
    matrix = _acceptance_matrix(summary)
    historical_summary = historical.get("policy_evidence_summary", {})
    historical_quarantine = {
        "legacy_manifest_validated_as_historical_fixture": True,
        "legacy_s14_p3_dynamic_state_is_authoritative": False,
        "legacy_pending_twelve_quarantined": historical_summary.get("pending_reconciliation_count") == 12,
        "legacy_five_gap_statuses_quarantined": historical_summary.get("evidence_gap_count") == 5,
        "legacy_five_risk_statuses_quarantined": historical_summary.get("risk_tip_count") == 5,
        "current_authoritative_evidence_bound_program_count": 0,
        "current_formal_policy_qualification_conclusion_count": 0,
    }
    validation_summary = {
        "final_validation_recorded": final_validation,
        "focused_test": "PASS" if final_validation else "PENDING",
        "strict_validator": "PASS" if final_validation else "PENDING",
        "browser_desktop_mobile": "PASS" if final_validation else "PENDING",
        "raw_candidate_probe": "PASS" if final_validation else "PENDING",
        "raw_alignment": "PASS" if final_validation else "PENDING",
        "governance_and_safety_scans": "PASS" if final_validation else "PENDING",
    }
    manifest = {
        "schema_version": "kmfa.v014.s14_p3_post_remediation_manifest.v1",
        "project_id": "KMFA",
        "stage_id": "S14",
        "phase_id": PHASE_ID,
        "roadmap_phase_id": ROADMAP_PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "generated_at": generated_at,
        "reviewed_head": p2.p1._git_output(["rev-parse", "HEAD"]),
        "branch": p2.p1._git_output(["branch", "--show-current"]),
        "remote": p2.p1._git_output(["remote", "get-url", "origin"]),
        "s14_p2_post_remediation_dependency_validated": True,
        "historical_s14_p3_dynamic_state_is_authoritative": False,
        "summary": summary,
        "policy_evidence_directories": directories,
        "policy_evidence_gaps": gaps,
        "policy_risk_tips": risks,
        "browser_review": browser,
        "quality_gate": _quality_gate(),
        "phase_boundaries": _phase_boundaries(),
        "public_repo_safety": _public_safety(),
        "historical_quarantine": historical_quarantine,
        "taskpack_contract": contract,
        "reviewed_dependencies": {
            "current_s14_p2": p2.MANIFEST_PATH.as_posix(),
            "current_finance_structure_registry": CURRENT_SOURCE_REGISTRY_PATH.as_posix(),
            "current_finance_structure_candidates": CURRENT_CANDIDATES_PATH.as_posix(),
            "historical_s14_p3_policy_fixture": HISTORICAL_MANIFEST_PATH.as_posix(),
        },
        "raw_boundary": {
            "raw_read_authorized": True,
            "raw_snapshot_validation_performed": True,
            "raw_write_performed": False,
            "raw_delete_performed": False,
            "raw_move_performed": False,
            "raw_rename_performed": False,
            "raw_overwrite_performed": False,
            "raw_mutation_performed": False,
        },
        "acceptance_matrix": matrix,
        "validation_summary": validation_summary,
        "next_phase": "S14-REVIEW",
        "next_required_step": (
            "Execute Stage 14 overall review as a separate run; do not execute S15, policy conclusions, "
            "applications, upload, reinstall, formal release, difference closure, persistent write, or business execution."
        ),
    }
    go_no_go = {
        "schema_version": "kmfa.v014.s14_p3_go_no_go.v1",
        "phase_id": PHASE_ID,
        "decision": "NO_GO",
        "data_quality_grade": "Q4",
        "report_grade": "D",
        "reason": "authoritative policy evidence identity completeness validity applicability and binding remain unproven",
        "stage14_review_allowed_in_this_run": False,
        "formal_policy_qualification_conclusion_allowed": False,
        "policy_application_submission_allowed": False,
        "subsidy_application_allowed": False,
        "github_upload_allowed": False,
        "business_execution_allowed": False,
    }

    public_writes = (
        (SUMMARY_PATH, summary),
        (MANIFEST_PATH, manifest),
        (
            DIRECTORIES_PATH,
            {"schema_version": "kmfa.v014.s14_p3_directories.v1", "directories": directories},
        ),
        (GAPS_PATH, {"schema_version": "kmfa.v014.s14_p3_gaps.v1", "gaps": gaps}),
        (RISKS_PATH, {"schema_version": "kmfa.v014.s14_p3_risks.v1", "risks": risks}),
        (MATRIX_PATH, matrix),
        (GO_NO_GO_PATH, go_no_go),
        (METADATA_SUMMARY_PATH, summary),
        (METADATA_MANIFEST_PATH, manifest),
        (
            METADATA_DIRECTORIES_PATH,
            {"schema_version": "kmfa.v014.s14_p3_directories.v1", "directories": directories},
        ),
        (METADATA_GAPS_PATH, {"schema_version": "kmfa.v014.s14_p3_gaps.v1", "gaps": gaps}),
        (METADATA_RISKS_PATH, {"schema_version": "kmfa.v014.s14_p3_risks.v1", "risks": risks}),
        (METADATA_MATRIX_PATH, matrix),
        (METADATA_GO_NO_GO_PATH, go_no_go),
    )
    for path, value in public_writes:
        p2.p1._write_json(path, value)
    p2.p1._write_text(REPORT_PATH, _render_report(summary))
    p2.p1._write_text(
        TEST_RESULTS_PATH,
        f"""# S14-P3 测试结果

- focused tests：{'PASS' if final_validation else 'PENDING'}
- strict validator：{'PASS' if final_validation else 'PENDING'}
- baseline HTML audit：{browser['baseline_pass_count']}/54 PASS
- current HTML audit：{browser['current_pass_count']}/{browser['current_control_row_count']} PASS
- desktop/mobile：{browser['viewport_check_count']}/2 PASS
- 五类目录交互：{browser['program_interaction_check_count']}/10 PASS
- 上游链接 HTTP / 真实导航：{browser['dependency_link_http_check_count']}/4 / {browser['dependency_navigation_check_count']}/4 PASS
- private raw probe：3,830 unique policy lexical candidate sheets / 0 roundtrip mismatch
- raw 前后/跨 S14-P2/current：exact match
""",
    )
    p2.p1._write_text(
        RISK_REGISTER_PATH,
        """# S14-P3 风险登记

1. 私有词法候选不等于政策证据材料，更不等于资格、评分或申报结论。
2. 23 个 XLSX 容器当前不能由 openpyxl 直接解析，只能保留私有诊断，不得损伤原文件转换。
3. 材料身份、完整性、有效期、主体、项目、人员、成果、金额和适用条件均未权威绑定。
4. 当前 Q4 / D / NO_GO，禁止政策资格结论、政策申报、补贴申请和正式报告。
""",
    )
    p2.p1._write_text(
        ROLLBACK_PATH,
        """# S14-P3 回滚方案

仅回退本 phase 的 public-safe 产物、代码、测试、治理登记和 ignored private runtime。不得修改、移动、删除、重命名、覆盖或回写原始数据目录。
""",
    )
    p2.p1._write_json(PRIVATE_RAW_BEFORE_PATH, raw_before)
    p2.p1._write_json(PRIVATE_RAW_AFTER_PATH, raw_after)
    p2.p1._write_json(PRIVATE_PROBE_PATH, raw_probe)
    p2.p1._write_text(PRIVATE_REPORT_PATH, _render_private_report(summary))
    if write_governance:
        _write_governance(generated_at)
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--browser-evidence-only", action="store_true")
    parser.add_argument("--final-validation", action="store_true")
    parser.add_argument("--no-governance", action="store_true")
    args = parser.parse_args()
    if args.browser_evidence_only:
        result = _browser_worker()
        return 0 if result["status"] == "PASS" else 1
    manifest = generate(final_validation=args.final_validation, write_governance=not args.no_governance)
    summary = manifest["summary"]
    print(
        "S14-P3 post-remediation: "
        f"programs={summary['policy_program_count']} lexical={summary['private_unique_policy_lexical_candidate_sheet_count']} "
        f"bound={summary['authoritative_evidence_bound_program_count']} conclusions={summary['formal_policy_qualification_conclusion_count']} "
        f"grade={summary['current_report_grade']} decision={summary['decision']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
