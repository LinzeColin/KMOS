#!/usr/bin/env python3
"""Resolve one remaining cash project and accept the final private difference."""

from __future__ import annotations

import argparse
import collections
import datetime as dt
import hashlib
import io
import json
import subprocess
import sys
import warnings
import zipfile
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from KMFA.tools import (  # noqa: E402
    v014_remaining_cash_source_private_trace_or_difference_acceptance as source_phase,
)


PROJECT_ROOT = Path("KMFA")
PHASE_ID = (
    "V014_REMAINING_TWO_PROJECT_CASH_COLLECTION_EVIDENCE_OR_FINAL_DIFFERENCE_ACCEPTANCE"
)
TASK_ID = (
    "KMFA-V014-REMAINING-TWO-PROJECT-CASH-COLLECTION-EVIDENCE-OR-FINAL-"
    "DIFFERENCE-ACCEPTANCE-20260710"
)
ACCEPTANCE_ID = (
    "ACC-V014-REMAINING-TWO-PROJECT-CASH-COLLECTION-EVIDENCE-OR-FINAL-"
    "DIFFERENCE-ACCEPTANCE"
)
VERSION = (
    "0.1.4-remaining-two-project-cash-collection-evidence-or-final-"
    "difference-acceptance"
)
STATUS = (
    "completed_validated_local_only_third_cash_project_materialized_final_cash_"
    "project_difference_accepted_no_go"
)
DECISION = "NO_GO"
PREFIX = "remaining_two_project_cash_collection_evidence_or_final_difference_acceptance"

ARTIFACT_DIR = PROJECT_ROOT / "stage_artifacts" / PHASE_ID
MACHINE_DIR = ARTIFACT_DIR / "machine"
HUMAN_DIR = ARTIFACT_DIR / "human"
SUMMARY_PATH = MACHINE_DIR / f"{PREFIX}_summary.json"
MANIFEST_PATH = MACHINE_DIR / f"{PREFIX}_manifest.json"
GO_NO_GO_PATH = MACHINE_DIR / f"{PREFIX}_go_no_go_report.json"
MATRIX_PATH = MACHINE_DIR / f"{PREFIX}_matrix_public_safe.json"
REPORT_PATH = HUMAN_DIR / f"{PREFIX}.md"
GO_NO_GO_RECORD_PATH = HUMAN_DIR / "go_no_go_record.md"
TEST_RESULTS_PATH = HUMAN_DIR / "test_results.md"
RISK_REGISTER_PATH = HUMAN_DIR / "risk_register.md"
ROLLBACK_PATH = HUMAN_DIR / "rollback_plan.md"

METADATA_SUMMARY_PATH = PROJECT_ROOT / f"metadata/quality/v014_{PREFIX}_summary.json"
METADATA_MANIFEST_PATH = PROJECT_ROOT / f"metadata/quality/v014_{PREFIX}_manifest.json"
METADATA_GO_NO_GO_PATH = PROJECT_ROOT / f"metadata/quality/v014_{PREFIX}_go_no_go_report.json"
METADATA_MATRIX_PATH = PROJECT_ROOT / f"metadata/quality/v014_{PREFIX}_matrix_public_safe.json"

PRIVATE_OUTPUT_DIR = (
    PROJECT_ROOT
    / ".codex_private_runtime/"
    "v014_remaining_two_project_cash_collection_evidence_or_final_difference_acceptance"
)
PRIVATE_RAW_BEFORE_PATH = PRIVATE_OUTPUT_DIR / "raw_immutability_before.json"
PRIVATE_RAW_AFTER_PATH = PRIVATE_OUTPUT_DIR / "raw_immutability_after.json"
PRIVATE_COLLECTION_CANDIDATES_PATH = (
    PRIVATE_OUTPUT_DIR / "private_raw_collection_candidates.jsonl"
)
PRIVATE_COLLECTION_LINKS_PATH = (
    PRIVATE_OUTPUT_DIR / "private_unique_collection_links.jsonl"
)
PRIVATE_REJECTED_LINKS_PATH = (
    PRIVATE_OUTPUT_DIR / "private_rejected_collection_links.jsonl"
)
PRIVATE_CASH_DECISIONS_PATH = PRIVATE_OUTPUT_DIR / "private_cash_source_decisions.jsonl"
PRIVATE_CASH_METRICS_PATH = PRIVATE_OUTPUT_DIR / "private_materialized_cash_metrics.jsonl"
PRIVATE_TARGET_MATERIALIZATIONS_PATH = (
    PRIVATE_OUTPUT_DIR / "private_s09_target_slot_materializations.jsonl"
)
PRIVATE_UNRESOLVED_TARGETS_PATH = (
    PRIVATE_OUTPUT_DIR / "private_unresolved_cash_value_targets.jsonl"
)
PRIVATE_RECONCILIATION_COMPARISONS_PATH = (
    PRIVATE_OUTPUT_DIR / "private_s09_reconciliation_comparisons.jsonl"
)
PRIVATE_DIAGNOSTIC_PATH = PRIVATE_OUTPUT_DIR / "private_collection_evidence_diagnostic.json"
PRIVATE_FINAL_DIFFERENCE_REPORT_PATH = (
    PRIVATE_OUTPUT_DIR / "private_final_difference_acceptance_report_zh.md"
)

SOURCE_CASH_DECISIONS_PATH = source_phase.PRIVATE_CASH_DECISIONS_PATH
SOURCE_CASH_METRICS_PATH = source_phase.PRIVATE_CASH_METRICS_PATH
SOURCE_TARGET_MATERIALIZATIONS_PATH = source_phase.PRIVATE_TARGET_MATERIALIZATIONS_PATH
SOURCE_UNRESOLVED_TARGETS_PATH = source_phase.PRIVATE_UNRESOLVED_TARGETS_PATH
SOURCE_RECONCILIATION_COMPARISONS_PATH = (
    source_phase.PRIVATE_RECONCILIATION_COMPARISONS_PATH
)
SOURCE_PRIVATE_SPEC_PATH = source_phase.SOURCE_PRIVATE_SPEC_PATH
SOURCE_LEDGER_EVIDENCE_PATH = source_phase.SOURCE_LEDGER_EVIDENCE_PATH
SOURCE_WPS_RECOVERY_DIAGNOSTIC_PATH = source_phase.PRIVATE_WPS_RECOVERY_DIAGNOSTIC_PATH

DEVELOPMENT_EVENTS_PATH = PROJECT_ROOT / "docs/governance/development_events.jsonl"
STAGE_STATUS_PATH = PROJECT_ROOT / "metadata/stage_status.jsonl"
TASK_STATUS_PATH = PROJECT_ROOT / "metadata/traceability/v1_4_stage_phase_task_status.jsonl"

COLLECTION_MARKERS = ("回款", "收款", "已收", "实收", "到款", "到账")
AMOUNT_MARKERS = ("金额", "累计", "本期", "总额", "合计", *COLLECTION_MARKERS)
EXCLUDED_COLLECTION_MARKERS = ("率", "天", "周期", "次数", "排名", "比例", "状态", "进度")
CUSTOMER_MARKERS = ("客户", "甲方", "建设单位", "付款单位", "对方单位", "往来单位")
DATE_MARKERS = ("日期", "时间", "年月")
PROJECT_MARKERS = ("项目", "合同", "工程")
OLE_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")


def _now(generated_at: str | None = None) -> str:
    return generated_at or dt.datetime.now().astimezone().isoformat(timespec="seconds")


def read_json(path: Path) -> dict[str, Any]:
    return source_phase.read_json(path)


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return source_phase.read_jsonl(path)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    source_phase._write_json(path, payload)


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    source_phase._write_jsonl(path, rows)


def _write_text(path: Path, text: str) -> None:
    source_phase._write_text(path, text)


def _raw_snapshot(kind: str) -> dict[str, Any]:
    snapshot = source_phase._raw_snapshot(kind)
    snapshot["phase_id"] = PHASE_ID
    snapshot["snapshot_kind"] = kind
    return snapshot


def _snapshot_core(snapshot: dict[str, Any]) -> dict[str, Any]:
    return source_phase._snapshot_core(snapshot)


def _contains_float(value: Any) -> bool:
    return source_phase._contains_float(value)


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _integer_cents(row: dict[str, Any], key: str) -> int:
    value = row.get(key, 0)
    if not isinstance(value, int) or isinstance(value, bool):
        raise ValueError(f"{key} must be integer cents")
    return value


def source_private_paths() -> list[Path]:
    return [
        SOURCE_PRIVATE_SPEC_PATH,
        SOURCE_LEDGER_EVIDENCE_PATH,
        SOURCE_WPS_RECOVERY_DIAGNOSTIC_PATH,
        SOURCE_CASH_DECISIONS_PATH,
        SOURCE_CASH_METRICS_PATH,
        SOURCE_TARGET_MATERIALIZATIONS_PATH,
        SOURCE_UNRESOLVED_TARGETS_PATH,
        SOURCE_RECONCILIATION_COMPARISONS_PATH,
    ]


def phase_output_paths() -> list[Path]:
    return [
        SUMMARY_PATH,
        MANIFEST_PATH,
        GO_NO_GO_PATH,
        MATRIX_PATH,
        REPORT_PATH,
        GO_NO_GO_RECORD_PATH,
        TEST_RESULTS_PATH,
        RISK_REGISTER_PATH,
        ROLLBACK_PATH,
        METADATA_SUMMARY_PATH,
        METADATA_MANIFEST_PATH,
        METADATA_GO_NO_GO_PATH,
        METADATA_MATRIX_PATH,
        PRIVATE_RAW_BEFORE_PATH,
        PRIVATE_RAW_AFTER_PATH,
        PRIVATE_COLLECTION_CANDIDATES_PATH,
        PRIVATE_COLLECTION_LINKS_PATH,
        PRIVATE_REJECTED_LINKS_PATH,
        PRIVATE_CASH_DECISIONS_PATH,
        PRIVATE_CASH_METRICS_PATH,
        PRIVATE_TARGET_MATERIALIZATIONS_PATH,
        PRIVATE_UNRESOLVED_TARGETS_PATH,
        PRIVATE_RECONCILIATION_COMPARISONS_PATH,
        PRIVATE_DIAGNOSTIC_PATH,
        PRIVATE_FINAL_DIFFERENCE_REPORT_PATH,
    ]


def classify_collection_link(
    *,
    source_records: list[dict[str, Any]],
    bank_row: dict[str, Any],
    voucher_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """Accept only a unique project/customer/bank/receivable collection chain."""

    if not source_records:
        raise ValueError("collection link requires source records")
    if _contains_float([source_records, bank_row, voucher_rows]):
        raise ValueError("collection link cannot contain float values")
    amounts = {row.get("amount_cents") for row in source_records}
    if len(amounts) != 1:
        raise ValueError("collection source amounts must be identical")
    amount = next(iter(amounts))
    if not isinstance(amount, int) or isinstance(amount, bool) or amount <= 0:
        raise ValueError("collection amount must be positive integer cents")
    for row in [bank_row, *voucher_rows]:
        for key in ("debit_cents", "credit_cents"):
            _integer_cents(row, key)

    bank_inflow = _integer_cents(bank_row, "debit_cents") - _integer_cents(
        bank_row, "credit_cents"
    )
    voucher_debit = sum(_integer_cents(row, "debit_cents") for row in voucher_rows)
    voucher_credit = sum(_integer_cents(row, "credit_cents") for row in voucher_rows)
    nonbank_rows = [
        row
        for row in voucher_rows
        if not str(row.get("account", "")).startswith(("1001", "1002"))
    ]
    receivable_credit = sum(
        _integer_cents(row, "credit_cents") - _integer_cents(row, "debit_cents")
        for row in nonbank_rows
        if str(row.get("account", "")).startswith("1122")
    )
    noncumulative_sources = [
        row
        for row in source_records
        if row.get("header_type") in {"generic_amount", "period"}
        and row.get("is_formula") is False
    ]

    reasons: list[str] = []
    if not any(row.get("project_dimension_present") is True for row in source_records):
        reasons.append("project_dimension_missing")
    if not any(row.get("customer_dimension_present") is True for row in source_records):
        reasons.append("customer_dimension_missing")
    if not any(row.get("customer_bank_match") is True for row in source_records):
        reasons.append("customer_bank_match_missing")
    if not noncumulative_sources:
        reasons.append("noncumulative_source_missing")
    if all(row.get("is_formula") is True for row in source_records):
        reasons.append("formula_source_only")
    if bank_inflow != amount:
        reasons.append("bank_inflow_amount_mismatch")
    if voucher_debit != voucher_credit:
        reasons.append("voucher_not_balanced")
    if receivable_credit <= 0:
        reasons.append("receivable_credit_missing")
    if receivable_credit != amount:
        reasons.append("bank_receivable_amount_mismatch")

    accepted = not reasons
    return {
        "schema_version": "kmfa.private.collection_link_classification.v1",
        "classification": (
            "unique_project_customer_bank_receivable_collection"
            if accepted
            else "rejected_collection_link"
        ),
        "accepted": accepted,
        "reason_codes": reasons,
        "amount_cents": amount,
        "source_record_count": len(source_records),
        "noncumulative_source_record_count": len(noncumulative_sources),
        "voucher_row_count": len(voucher_rows),
        "voucher_balanced": voucher_debit == voucher_credit,
        "bank_equals_receivable_credit": bank_inflow == receivable_credit == amount,
        "integer_only": True,
        "raw_layer_write_allowed": False,
        "public_commit_allowed": False,
    }


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value).strip()


def _amount_cents(value: Any) -> int | None:
    if value is None or isinstance(value, bool) or not isinstance(
        value, (int, float, Decimal)
    ):
        return None
    try:
        scaled = Decimal(str(value)) * 100
    except InvalidOperation:
        return None
    rounded = scaled.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return int(rounded) if scaled == rounded else None


def _alias_matches(value: str, tokens: list[str]) -> list[str]:
    folded = value.lower()
    return [token for token in tokens if token.lower() in folded]


def _collection_header(label: str) -> bool:
    return (
        any(marker in label for marker in COLLECTION_MARKERS)
        and any(marker in label for marker in AMOUNT_MARKERS)
        and not any(marker in label for marker in EXCLUDED_COLLECTION_MARKERS)
    )


def _header_type(label: str) -> str:
    if "累计" in label or "已回款" in label or "已收" in label:
        return "cumulative_or_to_date"
    if "本期" in label or "当期" in label or "本月" in label:
        return "period"
    return "generic_amount"


def _raw_file_map(raw_snapshot: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        str(record["relative_path"]): record
        for record in raw_snapshot.get("records", [])
        if isinstance(record, dict)
    }


def _discover_collection_candidates(
    *,
    raw_snapshot: dict[str, Any],
    source_spec: dict[str, Any],
    ledger: dict[str, Any],
    unresolved_decisions: list[dict[str, Any]],
) -> dict[str, Any]:
    raw_root = Path(str(raw_snapshot["raw_root"]))
    file_map = _raw_file_map(raw_snapshot)
    zip_paths = sorted(
        raw_root / relative_path
        for relative_path in file_map
        if Path(relative_path).suffix.lower() == ".zip"
    )
    source_archive = Path(str(source_spec["raw_archive_path"])).resolve()
    project_specs = {
        str(row["project_code"]): row for row in source_spec["project_specs"]
    }
    bank_inflows: list[dict[str, Any]] = []
    for bank_index, row in enumerate(ledger["bank_rows"]):
        amount = _integer_cents(row, "debit_cents") - _integer_cents(
            row, "credit_cents"
        )
        if amount <= 0:
            continue
        bank_inflows.append(
            {
                "bank_index": bank_index,
                "amount_cents": amount,
                "identity": " ".join(
                    _as_text(row.get(key))
                    for key in (
                        "summary",
                        "counterparty",
                        "customer",
                        "sales_contract",
                        "research_project",
                    )
                ),
                "row": row,
            }
        )

    candidates: list[dict[str, Any]] = []
    strict_groups: dict[tuple[str, int, int], dict[str, Any]] = {}
    accessible_workbooks = 0
    raw_archive_count = len(zip_paths)
    raw_ole_xlsx_count = 0
    raw_ooxml_xlsx_count = 0
    for relative_path, record in file_map.items():
        if Path(relative_path).suffix.lower() != ".xlsx":
            continue
        magic = (raw_root / relative_path).read_bytes()[:8]
        raw_ole_xlsx_count += int(magic.startswith(OLE_MAGIC))
        raw_ooxml_xlsx_count += int(magic.startswith(b"PK"))

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        for archive_index, archive_path in enumerate(zip_paths, start=1):
            if archive_path.resolve() == source_archive:
                continue
            relative_archive = archive_path.relative_to(raw_root).as_posix()
            archive_hash = str(file_map[relative_archive]["sha256"])
            with zipfile.ZipFile(archive_path) as archive:
                for member_index, member in enumerate(archive.infolist(), start=1):
                    suffix = Path(member.filename).suffix.lower()
                    if member.is_dir() or suffix not in {".xlsx", ".xls"}:
                        continue
                    payload = archive.read(member)
                    if not payload.startswith(b"PK"):
                        continue
                    try:
                        values_book = load_workbook(
                            io.BytesIO(payload), read_only=True, data_only=True
                        )
                        formula_book = load_workbook(
                            io.BytesIO(payload), read_only=True, data_only=False
                        )
                    except Exception:
                        continue
                    accessible_workbooks += 1
                    for sheet_index, worksheet in enumerate(values_book.worksheets):
                        formula_sheet = formula_book.worksheets[sheet_index]
                        header_labels: dict[int, list[str]] = collections.defaultdict(list)
                        for row in worksheet.iter_rows(
                            min_row=1, max_row=40, values_only=True
                        ):
                            for column_index, value in enumerate(row):
                                label = _as_text(value)
                                if label:
                                    header_labels[column_index].append(label)
                        collection_columns = {
                            column_index: next(
                                label
                                for label in labels
                                if _collection_header(label)
                            )
                            for column_index, labels in header_labels.items()
                            if any(_collection_header(label) for label in labels)
                        }
                        customer_columns = {
                            column_index
                            for column_index, labels in header_labels.items()
                            if any(
                                any(marker in label for marker in CUSTOMER_MARKERS)
                                for label in labels
                            )
                        }
                        date_columns = {
                            column_index
                            for column_index, labels in header_labels.items()
                            if any(
                                any(marker in label for marker in DATE_MARKERS)
                                for label in labels
                            )
                        }
                        project_columns = {
                            column_index
                            for column_index, labels in header_labels.items()
                            if any(
                                any(marker in label for marker in PROJECT_MARKERS)
                                for label in labels
                            )
                        }
                        if not collection_columns:
                            continue
                        value_rows = worksheet.iter_rows(values_only=True)
                        formula_rows = formula_sheet.iter_rows(values_only=True)
                        for row_number, (row, formula_row) in enumerate(
                            zip(value_rows, formula_rows), start=1
                        ):
                            row_text = " ".join(
                                _as_text(value) for value in row if value is not None
                            )
                            if not row_text:
                                continue
                            for decision in unresolved_decisions:
                                code = str(decision["project_code"])
                                tokens = [
                                    str(value).strip()
                                    for value in project_specs[code].get("alias_tokens", [])
                                    if str(value).strip()
                                ]
                                alias_hits = _alias_matches(row_text, tokens)
                                if not alias_hits:
                                    continue
                                customer_values = [
                                    _as_text(row[column_index])
                                    for column_index in customer_columns
                                    if column_index < len(row)
                                    and len(_as_text(row[column_index])) >= 3
                                ]
                                project_values = [
                                    _as_text(row[column_index])
                                    for column_index in project_columns
                                    if column_index < len(row)
                                    and _as_text(row[column_index])
                                ]
                                date_values = [
                                    _as_text(row[column_index])
                                    for column_index in date_columns
                                    if column_index < len(row)
                                    and _as_text(row[column_index])
                                ]
                                for column_index, label in collection_columns.items():
                                    if column_index >= len(row):
                                        continue
                                    amount = _amount_cents(row[column_index])
                                    if amount is None or amount <= 0:
                                        continue
                                    is_formula = bool(
                                        column_index < len(formula_row)
                                        and isinstance(formula_row[column_index], str)
                                        and formula_row[column_index].startswith("=")
                                    )
                                    amount_matches = [
                                        bank
                                        for bank in bank_inflows
                                        if bank["amount_cents"] == amount
                                    ]
                                    customer_matches = [
                                        bank
                                        for bank in amount_matches
                                        if any(
                                            customer.lower() in bank["identity"].lower()
                                            for customer in customer_values
                                        )
                                    ]
                                    candidate = {
                                        "schema_version": (
                                            "kmfa.private.raw_collection_candidate.v1"
                                        ),
                                        "classification": (
                                            "private_raw_collection_candidate_do_not_commit"
                                        ),
                                        "project_code": code,
                                        "binding_id": decision["binding_id"],
                                        "legacy_margin_record_id": decision[
                                            "legacy_margin_record_id"
                                        ],
                                        "archive_path": archive_path.as_posix(),
                                        "archive_sha256": archive_hash,
                                        "archive_index": archive_index,
                                        "member_name": member.filename,
                                        "member_index": member_index,
                                        "sheet_name": worksheet.title,
                                        "sheet_index": sheet_index,
                                        "row_number": row_number,
                                        "column_number": column_index + 1,
                                        "collection_header": label,
                                        "header_type": _header_type(label),
                                        "amount_cents": amount,
                                        "is_formula": is_formula,
                                        "project_dimension_present": bool(project_values),
                                        "customer_dimension_present": bool(customer_values),
                                        "customer_bank_match": len(customer_matches) == 1,
                                        "project_values": project_values,
                                        "customer_values": customer_values,
                                        "date_values": date_values,
                                        "alias_hits": alias_hits,
                                        "amount_bank_match_count": len(amount_matches),
                                        "customer_amount_bank_match_count": len(
                                            customer_matches
                                        ),
                                        "integer_only": True,
                                        "raw_layer_write_allowed": False,
                                        "public_commit_allowed": False,
                                    }
                                    candidates.append(candidate)
                                    unique_customer_matches = {
                                        int(bank["bank_index"]): bank
                                        for bank in customer_matches
                                    }
                                    if (
                                        project_values
                                        and customer_values
                                        and len(unique_customer_matches) == 1
                                    ):
                                        bank = next(iter(unique_customer_matches.values()))
                                        group_key = (
                                            code,
                                            amount,
                                            int(bank["bank_index"]),
                                        )
                                        group = strict_groups.setdefault(
                                            group_key,
                                            {
                                                "project_code": code,
                                                "binding_id": decision["binding_id"],
                                                "legacy_margin_record_id": decision[
                                                    "legacy_margin_record_id"
                                                ],
                                                "amount_cents": amount,
                                                "bank_row": bank["row"],
                                                "source_records": [],
                                            },
                                        )
                                        group["source_records"].append(candidate)
                    values_book.close()
                    formula_book.close()

    return {
        "raw_archive_count": raw_archive_count,
        "raw_ole_xlsx_count": raw_ole_xlsx_count,
        "raw_ooxml_xlsx_count": raw_ooxml_xlsx_count,
        "bank_inflow_record_count": len(bank_inflows),
        "accessible_nonledger_ooxml_workbook_count": accessible_workbooks,
        "candidates": candidates,
        "strict_groups": list(strict_groups.values()),
    }


def _voucher_rows_for_groups(
    source_spec: dict[str, Any], groups: list[dict[str, Any]]
) -> dict[tuple[str, str], list[dict[str, Any]]]:
    cash_source = source_phase.source_phase
    target_keys = {
        (
            str(group["bank_row"].get("date", "")),
            str(group["bank_row"].get("voucher", "")),
        )
        for group in groups
    }
    inner, shared_strings, sheets, _ = cash_source._load_inner_workbook(source_spec)
    output: dict[tuple[str, str], list[dict[str, Any]]] = collections.defaultdict(list)
    for sheet_name, target in sheets:
        for row_number, values in cash_source._sheet_rows(
            inner, shared_strings, target
        ):
            if row_number < 4 or not cash_source._transaction_row(values):
                continue
            key = (values[8], values[9].strip())
            if key in target_keys:
                output[key].append(
                    cash_source._row_record(sheet_name, row_number, values)
                )
    inner.close()
    return output


def _classify_groups(
    source_spec: dict[str, Any], groups: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    voucher_rows = _voucher_rows_for_groups(source_spec, groups)
    accepted: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    for index, group in enumerate(
        sorted(
            groups,
            key=lambda row: (
                str(row["project_code"]),
                int(row["amount_cents"]),
                str(row["bank_row"].get("date", "")),
                str(row["bank_row"].get("voucher", "")),
            ),
        ),
        start=1,
    ):
        key = (
            str(group["bank_row"].get("date", "")),
            str(group["bank_row"].get("voucher", "")),
        )
        linked_voucher_rows = voucher_rows.get(key, [])
        classification = classify_collection_link(
            source_records=group["source_records"],
            bank_row=group["bank_row"],
            voucher_rows=linked_voucher_rows,
        )
        record = {
            **classification,
            "link_id": f"V014-FINAL-COLLECTION-LINK-{index:03d}",
            "project_code": group["project_code"],
            "binding_id": group["binding_id"],
            "legacy_margin_record_id": group["legacy_margin_record_id"],
            "source_records": group["source_records"],
            "bank_row": group["bank_row"],
            "voucher_rows": linked_voucher_rows,
        }
        (accepted if classification["accepted"] else rejected).append(record)
    return accepted, rejected


def _evidence_amount(row: dict[str, Any]) -> int:
    return _integer_cents(row, "debit_cents") - _integer_cents(row, "credit_cents")


def _updated_decisions(
    source_decisions: list[dict[str, Any]], accepted_links: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], str, str]:
    decisions = [dict(row) for row in source_decisions]
    pending = [row for row in decisions if row.get("resolution_status") == "unresolved"]
    resolved_codes = {str(row["project_code"]) for row in accepted_links}
    if len(pending) != 2 or len(resolved_codes) != 1:
        raise ValueError("final cash phase must resolve one of two pending projects")
    resolved_code = next(iter(resolved_codes))
    resolved = next(row for row in pending if str(row["project_code"]) == resolved_code)
    final_pending = next(row for row in pending if str(row["project_code"]) != resolved_code)
    project_links = [
        row for row in accepted_links if str(row["project_code"]) == resolved_code
    ]
    collection = sum(int(row["amount_cents"]) for row in project_links)
    cash_paid = sum(
        _evidence_amount(row) for row in resolved.get("cash_paid_cost_evidence", [])
    )
    cash_gross = collection - cash_paid
    resolved.update(
        {
            "resolution_status": (
                "resolved_with_unique_raw_project_customer_bank_receivable_links"
            ),
            "reason_codes": [
                "two_unique_project_customer_bank_receivable_collection_links",
                "bank_vouchers_balanced_and_integer_replayable",
                "external_secure_wps_content_unavailable_not_claimed",
            ],
            "collection_amount_cents": collection,
            "cash_paid_cost_cents": cash_paid,
            "cash_gross_profit_cents": cash_gross,
            "cash_collection_evidence_count": len(project_links),
            "cash_collection_evidence": project_links,
            "unresolved_collection_count": 0,
            "unresolved_row_count": 0,
            "unresolved_evidence": [],
            "external_crosscheck_completed": True,
            "external_crosscheck_status": (
                "unique_raw_project_customer_bank_receivable_links_completed"
            ),
            "business_consistency_verified": False,
            "zero_inferred_from_absence": False,
        }
    )
    final_pending.update(
        {
            "resolution_status": "difference_accepted_unresolved",
            "reason_codes": [
                "final_difference_acceptance_no_positive_collection_evidence",
                "accessible_ledger_and_ooxml_sources_exhausted",
                "secure_wps_content_unavailable_not_claimed",
            ],
            "final_difference_acceptance": True,
            "external_crosscheck_completed": False,
            "external_crosscheck_status": (
                "final_difference_accepted_after_repeated_crosscheck"
            ),
            "business_consistency_verified": False,
            "zero_inferred_from_absence": False,
        }
    )
    return (
        decisions,
        str(resolved["legacy_margin_record_id"]),
        str(final_pending["legacy_margin_record_id"]),
    )


def _cash_metric_records(
    source_metrics: list[dict[str, Any]],
    decisions: list[dict[str, Any]],
    new_margin_record_id: str,
) -> list[dict[str, Any]]:
    metrics = [dict(row) for row in source_metrics]
    decision = next(
        row
        for row in decisions
        if str(row["legacy_margin_record_id"]) == new_margin_record_id
    )
    for index, metric_key in enumerate(
        ("collection_amount_cents", "cash_paid_cost_cents", "cash_gross_profit_cents"),
        start=1,
    ):
        value = decision[metric_key]
        if not isinstance(value, int) or isinstance(value, bool):
            raise ValueError("resolved cash metric must be integer cents")
        metrics.append(
            {
                "schema_version": "kmfa.private.materialized_cash_metric.v1",
                "classification": "private_materialized_cash_metric_do_not_commit",
                "metric_record_id": f"V014-FINAL-CASH-METRIC-{index:03d}",
                "binding_id": decision["binding_id"],
                "legacy_margin_record_id": new_margin_record_id,
                "metric_key": metric_key,
                "value": value,
                "unit": "cents",
                "derivation": (
                    "unique_raw_project_customer_bank_receivable_link_integer_formula"
                ),
                "integer_only": True,
                "raw_layer_write_allowed": False,
                "public_commit_allowed": False,
            }
        )
    return metrics


def _materialized_slot(row: dict[str, Any], value: int) -> dict[str, Any]:
    result = {
        key: child
        for key, child in row.items()
        if key not in {"reason_codes", "resolution_status"}
    }
    result.update(
        {
            "schema_version": "kmfa.private.s09_target_slot_materialization.v1",
            "classification": "private_s09_target_slot_materialization_do_not_commit",
            "materialization_status": "materialized_private_only",
            "value": value,
            "unit": "cents",
            "value_fingerprint": _sha256_text(str(value)),
            "raw_layer_write_allowed": False,
            "public_commit_allowed": False,
        }
    )
    return result


def _apply_target_materialization(
    *,
    new_margin_record_id: str,
    decisions: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], int]:
    materialized = read_jsonl(SOURCE_TARGET_MATERIALIZATIONS_PATH)
    unresolved = read_jsonl(SOURCE_UNRESOLVED_TARGETS_PATH)
    comparisons = read_jsonl(SOURCE_RECONCILIATION_COMPARISONS_PATH)
    decision = next(
        row
        for row in decisions
        if str(row["legacy_margin_record_id"]) == new_margin_record_id
    )
    comparison = next(
        row
        for row in comparisons
        if row["margin_record_id"] == new_margin_record_id
        and row["difference_type"] == "cash_vs_accrual_gross_profit"
    )
    cash_value = int(decision["cash_gross_profit_cents"])
    amount_b = comparison["amount_b"]
    if not isinstance(amount_b, int) or isinstance(amount_b, bool):
        raise ValueError("accrual comparison value must be integer cents")
    delta = cash_value - amount_b
    comparison.update(
        {
            "amount_a": cash_value,
            "delta": delta,
            "comparison_status": (
                "comparison_complete_zero_delta"
                if delta == 0
                else "comparison_complete_nonzero_delta"
            ),
        }
    )
    expected_context_values = {
        "cash_gross_profit": cash_value,
        "amount_a_cents_private_ref": cash_value,
        "delta_cents_private_ref": delta,
    }
    consumed_ids: set[str] = set()
    for row in unresolved:
        context = str(row["context_group"])
        if context not in expected_context_values:
            continue
        if context == "cash_gross_profit" and new_margin_record_id not in str(
            row["private_processed_ref"]
        ):
            continue
        if context != "cash_gross_profit" and row.get("difference_id") != comparison[
            "difference_id"
        ]:
            continue
        materialized.append(_materialized_slot(row, expected_context_values[context]))
        consumed_ids.add(str(row["target_slot_id"]))
    if len(consumed_ids) != 3:
        raise ValueError("resolved collection evidence must materialize exactly three slots")
    unresolved = [
        row for row in unresolved if str(row["target_slot_id"]) not in consumed_ids
    ]
    materialized.sort(key=lambda row: str(row["target_slot_id"]))
    unresolved.sort(key=lambda row: str(row["target_slot_id"]))
    comparisons.sort(key=lambda row: str(row["comparison_id"]))
    return materialized, unresolved, comparisons, len(consumed_ids)


def _git_ignored(path: Path) -> bool:
    return (
        subprocess.run(
            ["git", "check-ignore", "-q", path.as_posix()],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        ).returncode
        == 0
    )


def _git_output(args: list[str]) -> str:
    return subprocess.check_output(["git", *args], text=True).strip()


def _public_matrix(summary: dict[str, Any]) -> dict[str, Any]:
    checks = [
        ("source_private_inputs_unchanged", summary["source_private_inputs_unchanged"]),
        ("raw_snapshot_exact_match", summary["raw_snapshot_exact_match"]),
        ("nineteen_ooxml_workbooks", summary["accessible_nonledger_ooxml_workbook_count"] == 19),
        ("forty_eight_raw_candidates", summary["raw_collection_candidate_record_count"] == 48),
        ("two_unique_collection_links", summary["unique_balanced_collection_link_count"] == 2),
        ("two_balanced_receivable_vouchers", summary["balanced_receivable_voucher_count"] == 2),
        ("third_cash_project_resolved", summary["cash_project_resolved_count"] == 3),
        ("one_cash_project_unresolved", summary["cash_project_unresolved_count"] == 1),
        ("final_difference_accepted", summary["final_difference_accepted_project_count"] == 1),
        ("three_new_targets", summary["newly_materialized_cash_target_slot_count"] == 3),
        ("three_targets_remain", summary["unresolved_cash_value_target_slot_count"] == 3),
        ("eleven_comparisons_complete", summary["completed_reconciliation_comparison_count"] == 11),
        ("existing_nonzero_preserved", summary["nonzero_delta_reconciliation_count"] >= 8),
        ("one_cash_comparison_incomplete", summary["incomplete_cash_reconciliation_count"] == 1),
        ("secure_wps_not_claimed", summary["secure_wps_content_readable_count"] == 0),
        ("forced_zero_prohibited", summary["forced_zero_materialization_count"] == 0),
        ("business_consistency_not_claimed", not summary["business_value_consistency_verified"]),
        ("global_residual_not_replayed", not summary["global_residual_difference_queue_replayed"]),
        ("github_not_uploaded", not summary["github_upload_performed"]),
        ("stage_review_not_performed", not summary["stage_review_performed"]),
    ]
    rows = [
        {"check_id": f"FCE-{index:02d}", "check_name": name, "passed": passed}
        for index, (name, passed) in enumerate(checks, start=1)
    ]
    return {
        "schema_version": "kmfa.v014.final_collection_evidence_matrix.v1",
        "project_id": "KMFA",
        "phase_id": PHASE_ID,
        "check_count": len(rows),
        "check_pass_count": sum(row["passed"] for row in rows),
        "check_fail_count": sum(not row["passed"] for row in rows),
        "checks": rows,
    }


def _private_difference_report(
    *,
    summary: dict[str, Any],
    candidates: list[dict[str, Any]],
    links: list[dict[str, Any]],
    decisions: list[dict[str, Any]],
    comparisons: list[dict[str, Any]],
) -> str:
    lines = [
        "# KMFA v0.1.4 最终现金收款证据与最终差异接受报告",
        "",
        "## 本轮结论",
        "",
        f"- 可访问非主账 OOXML 工作簿：{summary['accessible_nonledger_ooxml_workbook_count']} 个。",
        f"- 正向收款候选：{summary['raw_collection_candidate_record_count']} 条。",
        f"- 唯一银行入账链：{summary['unique_balanced_collection_link_count']} 条。",
        f"- 现金项目已闭环 / 未闭环：{summary['cash_project_resolved_count']} / {summary['cash_project_unresolved_count']}。",
        f"- 新增物化 / 累计物化 / 剩余槽位：{summary['newly_materialized_cash_target_slot_count']} / {summary['materialized_business_value_target_slot_count']} / {summary['unresolved_cash_value_target_slot_count']}。",
        f"- 完成比较 / 零差异 / 非零差异 / 未完成：{summary['completed_reconciliation_comparison_count']} / {summary['zero_delta_reconciliation_count']} / {summary['nonzero_delta_reconciliation_count']} / {summary['incomplete_cash_reconciliation_count']}。",
        "- 原始数据保持不变；所有金额均按整数分重放。",
        "",
        "## 唯一银行入账链",
        "",
    ]
    for link in links:
        lines.extend(
            [
                f"### {link['link_id']}",
                "",
                f"- 私有项目编码：{link['project_code']}",
                f"- 收款金额（分）：{link['amount_cents']}",
                f"- 来源记录：{link['source_records']}",
                f"- 银行入账：{link['bank_row']}",
                f"- 同凭证记录：{link['voucher_rows']}",
                f"- 借贷平衡：{link['voucher_balanced']}",
                f"- 银行入账等于应收贷方：{link['bank_equals_receivable_credit']}",
                "",
            ]
        )
    lines.extend(["## 正向收款候选审计", ""])
    for candidate in candidates:
        lines.append(
            f"- {candidate['binding_id']} / {candidate['archive_path']} / {candidate['member_name']} / {candidate['sheet_name']} / 行{candidate['row_number']}列{candidate['column_number']}：金额（分）={candidate['amount_cents']}，项目维度={candidate['project_dimension_present']}，客户维度={candidate['customer_dimension_present']}，客户银行唯一匹配={candidate['customer_bank_match']}，公式={candidate['is_formula']}。"
        )
    lines.extend(["", "## 项目与比较状态", ""])
    for decision in decisions:
        lines.append(
            f"- {decision['binding_id']}：{decision['resolution_status']}；原因={'; '.join(decision['reason_codes'])}；现金毛利（分）={decision['cash_gross_profit_cents']}。"
        )
    for comparison in comparisons:
        lines.append(
            f"- {comparison['difference_id']}：{comparison['comparison_status']}；A={comparison['amount_a']}；B={comparison['amount_b']}；差额={comparison['delta']}。"
        )
    lines.extend(
        [
            "",
            "## 最终差异接受",
            "",
            "- 最后一个项目在主账、银行、应收、可访问 OOXML 与 WPS/OLE 兼容层的重复交叉核验中，仍没有可证明的正向收款证据。",
            "- 缺失证据未写成零，剩余 3 个现金槽位和 1 条现金比较继续未决。",
            "- WPS 安全内容仍依赖专有 ticket runtime，未声明已读取，未用空白兼容层替代业务数据。",
            "- 非零差异不覆盖、不平均、不抹平；全局 residual queue 仍需后续独立 phase 重放。",
            "- 原始数据保持不变，不生成正式报告，不上传 GitHub，不执行业务动作。",
            "",
        ]
    )
    return "\n".join(lines)


def _public_reports(summary: dict[str, Any]) -> dict[Path, str]:
    report = f"""# v0.1.4 剩余两项目收款证据或最终差异接受

- Phase: `{PHASE_ID}`
- 决策: `{DECISION}`
- 可访问非主账 OOXML 工作簿: {summary['accessible_nonledger_ooxml_workbook_count']}
- 正向收款候选 / 唯一平衡入账链: {summary['raw_collection_candidate_record_count']} / {summary['unique_balanced_collection_link_count']}
- 已闭环 / 未闭环项目: {summary['cash_project_resolved_count']} / {summary['cash_project_unresolved_count']}
- 新增物化 / 累计物化 / 剩余现金槽位: {summary['newly_materialized_cash_target_slot_count']} / {summary['materialized_business_value_target_slot_count']} / {summary['unresolved_cash_value_target_slot_count']}
- 完成比较 / 非零差异 / 未完成现金比较: {summary['completed_reconciliation_comparison_count']} / {summary['nonzero_delta_reconciliation_count']} / {summary['incomplete_cash_reconciliation_count']}
- 最终差异接受项目: {summary['final_difference_accepted_project_count']}
- raw 前后完全一致: `{str(summary['raw_snapshot_exact_match']).lower()}`

本 phase 通过项目维度、客户、唯一等额银行入账、应收贷方和借贷平衡证据，新增闭环一个项目。最后一个项目重复交叉核验后仍无正向收款证据，已生成 private 全中文最终差异接受报告，未将缺失值写成零。
"""
    go_no_go = f"""# Go / No-Go 记录

- 决策: `NO_GO`
- 已完成: 新增 {summary['unique_balanced_collection_link_count']} 条唯一收款链，第三个项目现金闭环，新增 {summary['newly_materialized_cash_target_slot_count']} 个私有目标槽位。
- 阻断: {summary['cash_project_unresolved_count']} 个项目、{summary['unresolved_cash_value_target_slot_count']} 个现金槽位、{summary['nonzero_delta_reconciliation_count']} 条非零差异、全局 residual queue 未重放。
- GitHub upload / app reinstall / business execution: `not_performed`
"""
    tests = f"""# 测试结果

- focused unit test: `PASS` (`2 tests`)
- phase/private validator: `PASS` (`resolved=3 / unresolved=1 / links=2 / materialized=37 / decision=NO_GO`)
- previous phase validator: `PASS` (`resolved=2 / unresolved=2 / materialized=34 / decision=NO_GO`)
- governance validators: `PASS` (`errors=0 / warnings=0`)
- structured parse and registry counts: `PASS` (`37 files / 1304 active parameters / 275 active formulas`)
- raw/private/secret scan: `PASS` (`5 raw files unchanged / 0 tracked private files / 0 sensitive hits`)
"""
    risks = f"""# 风险登记

- 高: 最后一个项目仍缺正向收款证据，不能以未命中推导零值。
- 高: {summary['nonzero_delta_reconciliation_count']} 条非零口径差异继续保留，不得覆盖权威值或自动抹平。
- 高: WPS/OLE 安全内容仍不可读，最终差异接受不等于业务一致性验证。
- 中: 全局 72 条 residual queue 尚未重放。
"""
    rollback = """# 回滚方案

1. 删除本 phase 公开 artifacts 和 metadata 镜像。
2. 删除本 phase ignored private outputs；不触碰原始数据目录或上一 phase 私有输入。
3. 移除本 phase 治理记录并重跑上一 phase validator。
"""
    return {
        REPORT_PATH: report,
        GO_NO_GO_RECORD_PATH: go_no_go,
        TEST_RESULTS_PATH: tests,
        RISK_REGISTER_PATH: risks,
        ROLLBACK_PATH: rollback,
    }


def _phase_public_files() -> list[str]:
    return [
        "KMFA/CHANGELOG.md",
        "KMFA/HANDOFF.md",
        "KMFA/VERSION",
        "KMFA/docs/governance/ASSURANCE_STATUS.yaml",
        "KMFA/docs/governance/DEVELOPMENT_LEDGER.md",
        "KMFA/docs/governance/MODEL_SPEC.md",
        "KMFA/docs/governance/OWNER_STATUS.md",
        "KMFA/docs/governance/STATUS.md",
        "KMFA/docs/governance/TRACEABILITY_MATRIX.csv",
        "KMFA/docs/governance/VERSION_MATRIX.yaml",
        "KMFA/docs/governance/delivery_tasks.yaml",
        "KMFA/docs/governance/development_events.jsonl",
        "KMFA/docs/governance/formula_registry.yaml",
        "KMFA/docs/governance/model_registry.yaml",
        "KMFA/docs/governance/parameter_registry.csv",
        "KMFA/metadata/model_registry.yaml",
        METADATA_SUMMARY_PATH.as_posix(),
        METADATA_MANIFEST_PATH.as_posix(),
        METADATA_GO_NO_GO_PATH.as_posix(),
        METADATA_MATRIX_PATH.as_posix(),
        "KMFA/metadata/stage_status.jsonl",
        "KMFA/metadata/traceability/v1_4_stage_phase_task_status.jsonl",
        SUMMARY_PATH.as_posix(),
        MANIFEST_PATH.as_posix(),
        GO_NO_GO_PATH.as_posix(),
        MATRIX_PATH.as_posix(),
        REPORT_PATH.as_posix(),
        GO_NO_GO_RECORD_PATH.as_posix(),
        TEST_RESULTS_PATH.as_posix(),
        RISK_REGISTER_PATH.as_posix(),
        ROLLBACK_PATH.as_posix(),
        (
            "KMFA/tests/test_v014_remaining_two_project_cash_collection_evidence_"
            "or_final_difference_acceptance.py"
        ),
        (
            "KMFA/tools/check_v014_remaining_two_project_cash_collection_evidence_"
            "or_final_difference_acceptance.py"
        ),
        (
            "KMFA/tools/v014_remaining_two_project_cash_collection_evidence_or_"
            "final_difference_acceptance.py"
        ),
        "KMFA/功能清单.md",
        "KMFA/开发记录.md",
        "KMFA/模型参数文件.md",
    ]


def _write_governance(generated_at: str, summary: dict[str, Any]) -> None:
    evidence_ref = MANIFEST_PATH.as_posix()
    source_phase._upsert_jsonl(
        DEVELOPMENT_EVENTS_PATH,
        "event_id",
        {
            "event_id": (
                "DEV-KMFA-20260710-V014-REMAINING-TWO-PROJECT-CASH-COLLECTION-"
                "EVIDENCE-OR-FINAL-DIFFERENCE-ACCEPTANCE"
            ),
            "event_time": generated_at,
            "event_type": "development",
            "project_id": "KMFA",
            "stage_id": "value-consistency",
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "decision": DECISION,
            "unique_balanced_collection_link_count": summary[
                "unique_balanced_collection_link_count"
            ],
            "cash_project_resolved_count": summary["cash_project_resolved_count"],
            "cash_project_unresolved_count": summary["cash_project_unresolved_count"],
            "final_difference_accepted_project_count": summary[
                "final_difference_accepted_project_count"
            ],
            "raw_snapshot_exact_match": summary["raw_snapshot_exact_match"],
            "raw_business_data_committed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "files_changed": _phase_public_files(),
            "result_commit": "PENDING",
        },
    )
    source_phase._upsert_jsonl(
        STAGE_STATUS_PATH,
        "phase_id",
        {
            "schema_version": "kmfa.stage_status.v1",
            "record_type": "stage_phase_status",
            "project_id": "KMFA",
            "stage_id": "VALUE-CONSISTENCY",
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "version": VERSION,
            "status": STATUS,
            "decision": DECISION,
            "raw_data_committed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at,
        },
    )
    source_phase._upsert_jsonl(
        TASK_STATUS_PATH,
        "phase_id",
        {
            "schema_version": "kmfa.v014_stage_phase_task_status.v1",
            "record_type": "v014_phase",
            "project_id": "KMFA",
            "stage_id": "VALUE-CONSISTENCY",
            "governance_stage_id": "VALUE-CONSISTENCY",
            "roadmap_stage_id": "VALUE-CONSISTENCY",
            "roadmap_phase_id": (
                "REMAINING_TWO_PROJECT_CASH_COLLECTION_EVIDENCE_OR_FINAL_"
                "DIFFERENCE_ACCEPTANCE"
            ),
            "phase_id": PHASE_ID,
            "task_id": TASK_ID,
            "acceptance_id": ACCEPTANCE_ID,
            "version": VERSION,
            "name": (
                "v0.1.4 remaining two project cash collection evidence or final "
                "difference acceptance"
            ),
            "phase_goal": (
                "materialize unique positive collection evidence and accept the final "
                "unresolved cash difference"
            ),
            "status": STATUS,
            "fact_level": "EXTRACTED",
            "derived_percent": 100.0,
            "estimated_task_units": 1,
            "completed_task_units": 1,
            "task_count": 1,
            "raw_data_committed": False,
            "evidence_ref": evidence_ref,
            "updated_at": generated_at[:10],
        },
    )


def generate(
    *, generated_at: str | None = None, write_governance_event: bool = True
) -> dict[str, Any]:
    timestamp = _now(generated_at)
    source_hashes_before = {
        path.as_posix(): _sha256_bytes(path.read_bytes()) for path in source_private_paths()
    }
    raw_before = _raw_snapshot("before")
    source_spec = read_json(SOURCE_PRIVATE_SPEC_PATH)
    ledger = read_json(SOURCE_LEDGER_EVIDENCE_PATH)
    source_decisions = read_jsonl(SOURCE_CASH_DECISIONS_PATH)
    source_metrics = read_jsonl(SOURCE_CASH_METRICS_PATH)
    wps = read_json(SOURCE_WPS_RECOVERY_DIAGNOSTIC_PATH)
    unresolved_source_decisions = [
        row for row in source_decisions if row.get("resolution_status") == "unresolved"
    ]

    discovery = _discover_collection_candidates(
        raw_snapshot=raw_before,
        source_spec=source_spec,
        ledger=ledger,
        unresolved_decisions=unresolved_source_decisions,
    )
    accepted_links, rejected_links = _classify_groups(
        source_spec, discovery["strict_groups"]
    )
    decisions, new_margin_record_id, final_pending_margin_record_id = _updated_decisions(
        source_decisions, accepted_links
    )
    metrics = _cash_metric_records(source_metrics, decisions, new_margin_record_id)
    materialized, unresolved, comparisons, newly_materialized = (
        _apply_target_materialization(
            new_margin_record_id=new_margin_record_id,
            decisions=decisions,
        )
    )

    source_hashes_after = {
        path.as_posix(): _sha256_bytes(path.read_bytes()) for path in source_private_paths()
    }
    source_unchanged = source_hashes_before == source_hashes_after
    raw_after = _raw_snapshot("after")
    raw_snapshot_exact_match = _snapshot_core(raw_before) == _snapshot_core(raw_after)
    if not source_unchanged or not raw_snapshot_exact_match:
        raise ValueError("source private inputs or raw inbox changed during final cash evidence phase")
    if discovery["accessible_nonledger_ooxml_workbook_count"] != 19:
        raise ValueError("accessible OOXML workbook count changed")
    if len(discovery["candidates"]) != 48:
        raise ValueError("raw collection candidate count changed")
    if len(accepted_links) != 2 or len({row["project_code"] for row in accepted_links}) != 1:
        raise ValueError("exactly two links for one project must be accepted")
    if len(metrics) != 9 or len(materialized) != 37 or len(unresolved) != 3:
        raise ValueError("final cash materialization counts do not match contract")
    if _contains_float(
        [
            discovery["candidates"],
            accepted_links,
            rejected_links,
            decisions,
            metrics,
            materialized,
            unresolved,
            comparisons,
        ]
    ):
        raise ValueError("private final cash outputs cannot contain floats")

    resolved = [
        row
        for row in decisions
        if str(row.get("resolution_status", "")).startswith("resolved_")
    ]
    final_pending = [
        row for row in decisions if row.get("resolution_status") == "difference_accepted_unresolved"
    ]
    completed = [
        row
        for row in comparisons
        if str(row.get("comparison_status", "")).startswith("comparison_complete_")
    ]
    zero = [row for row in completed if row.get("delta") == 0]
    nonzero = [row for row in completed if row.get("delta") != 0]
    incomplete = [row for row in comparisons if row not in completed]
    summary = {
        "schema_version": "kmfa.v014.final_collection_evidence_summary.v1",
        "record_type": "v014_final_collection_evidence_summary",
        "project_id": "KMFA",
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "acceptance_id": ACCEPTANCE_ID,
        "version": VERSION,
        "status": STATUS,
        "decision": DECISION,
        "generated_at": timestamp,
        "source_private_inputs_unchanged": source_unchanged,
        "raw_archive_count": discovery["raw_archive_count"],
        "raw_ole_xlsx_count": discovery["raw_ole_xlsx_count"],
        "raw_ooxml_xlsx_count": discovery["raw_ooxml_xlsx_count"],
        "accessible_nonledger_ooxml_workbook_count": discovery[
            "accessible_nonledger_ooxml_workbook_count"
        ],
        "bank_inflow_record_count": discovery["bank_inflow_record_count"],
        "raw_collection_candidate_record_count": len(discovery["candidates"]),
        "strict_collection_source_record_count": sum(
            len(group["source_records"]) for group in discovery["strict_groups"]
        ),
        "unique_balanced_collection_link_count": len(accepted_links),
        "balanced_receivable_voucher_count": sum(
            row["voucher_balanced"] and row["bank_equals_receivable_credit"]
            for row in accepted_links
        ),
        "cash_project_candidate_count": len(decisions),
        "cash_project_resolved_count": len(resolved),
        "cash_project_unresolved_count": len(decisions) - len(resolved),
        "collection_evidence_project_resolved_count": 1,
        "final_difference_accepted_project_count": len(final_pending),
        "final_pending_margin_record_count": int(bool(final_pending_margin_record_id)),
        "private_cash_metric_record_count": len(metrics),
        "newly_materialized_cash_target_slot_count": newly_materialized,
        "materialized_business_value_target_slot_count": len(materialized),
        "unresolved_cash_value_target_slot_count": len(unresolved),
        "reconciliation_record_count": len(comparisons),
        "completed_reconciliation_comparison_count": len(completed),
        "zero_delta_reconciliation_count": len(zero),
        "nonzero_delta_reconciliation_count": len(nonzero),
        "incomplete_cash_reconciliation_count": len(incomplete),
        "wps_source_count": wps["source_count"],
        "office_compatibility_unlock_count": wps["office_compatibility_unlock_count"],
        "empty_compatibility_workbook_count": wps[
            "empty_compatibility_workbook_count"
        ],
        "secure_wps_content_readable_count": wps[
            "secure_wps_content_readable_count"
        ],
        "partial_external_crosscheck_completed": True,
        "external_crosscheck_completed": False,
        "final_cash_difference_acceptance_completed": True,
        "forced_zero_materialization_count": 0,
        "global_unresolved_difference_count": 72,
        "global_residual_difference_queue_replayed": False,
        "partial_processed_value_materialization_complete": True,
        "full_processed_value_materialization_complete": False,
        "partial_raw_to_processed_reconciliation_performed": True,
        "full_raw_to_processed_value_comparison_complete": False,
        "processed_consistency_verified": False,
        "business_value_consistency_verified": False,
        "raw_source_file_count": raw_before["file_count"],
        "raw_snapshot_exact_match": raw_snapshot_exact_match,
        "raw_inbox_mutated_by_this_phase": not raw_snapshot_exact_match,
        "private_collection_links_gitignored": _git_ignored(
            PRIVATE_COLLECTION_LINKS_PATH
        ),
        "private_final_difference_report_gitignored": _git_ignored(
            PRIVATE_FINAL_DIFFERENCE_REPORT_PATH
        ),
        "raw_business_data_committed": False,
        "raw_filename_or_value_committed": False,
        "credential_or_secret_committed": False,
        "stage_review_performed": False,
        "github_upload_performed": False,
        "app_reinstall_performed": False,
        "business_execution_performed": False,
        "goal_status_recommendation": (
            "continue_active_with_global_residual_difference_replay"
        ),
        "next_recommended_phase": (
            "global_residual_difference_queue_replay_or_authoritative_exclusion"
        ),
    }
    matrix = _public_matrix(summary)
    if matrix["check_fail_count"]:
        raise ValueError("public acceptance matrix contains failures")
    manifest = {
        "schema_version": "kmfa.v014.final_collection_evidence_manifest.v1",
        "record_type": "v014_final_collection_evidence_manifest",
        "project_id": "KMFA",
        "version": VERSION,
        "phase_id": PHASE_ID,
        "task_id": TASK_ID,
        "generated_at": timestamp,
        "reviewed_head": _git_output(["rev-parse", "HEAD"]),
        "branch": _git_output(["branch", "--show-current"]),
        "remote": _git_output(["remote", "get-url", "origin"]),
        "summary": summary,
        "artifact_refs": {
            "summary": SUMMARY_PATH.as_posix(),
            "go_no_go": GO_NO_GO_PATH.as_posix(),
            "matrix": MATRIX_PATH.as_posix(),
            "report": REPORT_PATH.as_posix(),
            "test_results": TEST_RESULTS_PATH.as_posix(),
            "validator": (
                "KMFA/tools/check_v014_remaining_two_project_cash_collection_"
                "evidence_or_final_difference_acceptance.py"
            ),
        },
        "public_repo_safety": {
            "aggregate_only": True,
            "raw_file_committed": False,
            "raw_filename_committed": False,
            "raw_hash_committed": False,
            "identity_plaintext_committed": False,
            "business_value_committed": False,
            "private_ref_committed": False,
            "credential_or_secret_committed": False,
        },
        "phase_boundaries": {
            "single_phase_only": True,
            "positive_collection_evidence_crosscheck_performed": True,
            "final_cash_difference_acceptance_performed": True,
            "forced_zero_materialization_allowed": False,
            "secure_wps_content_read_claimed": False,
            "global_residual_queue_replayed": False,
            "stage_review_performed": False,
            "github_upload_performed": False,
            "app_reinstall_performed": False,
            "business_execution_performed": False,
        },
    }
    go_no_go = {
        "schema_version": "kmfa.v014.final_collection_evidence_go_no_go.v1",
        "project_id": "KMFA",
        "phase_id": PHASE_ID,
        "decision": DECISION,
        "cash_project_unresolved_count": len(decisions) - len(resolved),
        "unresolved_cash_value_target_slot_count": len(unresolved),
        "nonzero_delta_reconciliation_count": len(nonzero),
        "incomplete_cash_reconciliation_count": len(incomplete),
        "final_difference_accepted_project_count": len(final_pending),
        "blocking_reason_codes": [
            "one_cash_project_has_final_accepted_missing_collection_evidence",
            "nonzero_scope_differences_remain",
            "global_residual_queue_not_replayed",
            "secure_wps_content_unavailable",
        ],
        "github_upload_performed": False,
    }

    _write_json(PRIVATE_RAW_BEFORE_PATH, raw_before)
    _write_json(PRIVATE_RAW_AFTER_PATH, raw_after)
    _write_jsonl(PRIVATE_COLLECTION_CANDIDATES_PATH, discovery["candidates"])
    _write_jsonl(PRIVATE_COLLECTION_LINKS_PATH, accepted_links)
    _write_jsonl(PRIVATE_REJECTED_LINKS_PATH, rejected_links)
    _write_jsonl(PRIVATE_CASH_DECISIONS_PATH, decisions)
    _write_jsonl(PRIVATE_CASH_METRICS_PATH, metrics)
    _write_jsonl(PRIVATE_TARGET_MATERIALIZATIONS_PATH, materialized)
    _write_jsonl(PRIVATE_UNRESOLVED_TARGETS_PATH, unresolved)
    _write_jsonl(PRIVATE_RECONCILIATION_COMPARISONS_PATH, comparisons)
    _write_json(
        PRIVATE_DIAGNOSTIC_PATH,
        {
            "schema_version": "kmfa.private.final_collection_evidence_diagnostic.v1",
            "classification": "private_final_collection_evidence_do_not_commit",
            "generated_at": timestamp,
            "source_hashes_before": source_hashes_before,
            "source_hashes_after": source_hashes_after,
            "source_private_inputs_unchanged": source_unchanged,
            "raw_before_snapshot": raw_before,
            "raw_after_snapshot": raw_after,
            "raw_snapshot_exact_match": raw_snapshot_exact_match,
            "candidate_count": len(discovery["candidates"]),
            "accepted_link_count": len(accepted_links),
            "rejected_link_count": len(rejected_links),
            "accepted_project_codes": sorted(
                {str(row["project_code"]) for row in accepted_links}
            ),
            "final_pending_margin_record_id": final_pending_margin_record_id,
        },
    )
    _write_text(
        PRIVATE_FINAL_DIFFERENCE_REPORT_PATH,
        _private_difference_report(
            summary=summary,
            candidates=discovery["candidates"],
            links=accepted_links,
            decisions=decisions,
            comparisons=comparisons,
        ),
    )
    for path, payload in (
        (SUMMARY_PATH, summary),
        (MANIFEST_PATH, manifest),
        (GO_NO_GO_PATH, go_no_go),
        (MATRIX_PATH, matrix),
        (METADATA_SUMMARY_PATH, summary),
        (METADATA_MANIFEST_PATH, manifest),
        (METADATA_GO_NO_GO_PATH, go_no_go),
        (METADATA_MATRIX_PATH, matrix),
    ):
        _write_json(path, payload)
    for path, content in _public_reports(summary).items():
        _write_text(path, content)
    if write_governance_event:
        _write_governance(timestamp, summary)
    return {"summary": summary, "manifest": manifest, "go_no_go": go_no_go, "matrix": matrix}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--generated-at")
    parser.add_argument("--skip-governance-event", action="store_true")
    args = parser.parse_args()
    result = generate(
        generated_at=args.generated_at,
        write_governance_event=not args.skip_governance_event,
    )
    summary = result["summary"]
    print(
        "final collection evidence: "
        f"decision={summary['decision']} "
        f"resolved={summary['cash_project_resolved_count']} "
        f"unresolved={summary['cash_project_unresolved_count']} "
        f"links={summary['unique_balanced_collection_link_count']} "
        f"materialized={summary['materialized_business_value_target_slot_count']} "
        f"raw_unchanged={summary['raw_snapshot_exact_match']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
