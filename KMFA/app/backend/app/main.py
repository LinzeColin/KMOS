#!/usr/bin/env python3
"""KMFA App 后端骨架（TSK.KMFA.PROD.0001，D2=A：KMIDS 同栈 FastAPI）。

只读吃机器面：machine/facts（状态/数据管线）+ metadata/quality/assertions.jsonl（对账断言）。
页眉三元组（质量等级/报告等级/GO 状态）由 /api/状态 直给——DoD 第 1 条的页眉数据源。
私有派生层（DuckDB）不经本服务暴露明细；App 只出 public-safe 聚合。
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import secrets
import tempfile
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from html import escape as html_escape
from urllib.parse import quote, urlsplit
from pathlib import Path
from typing import Any

import yaml
from fastapi import Body, FastAPI, HTTPException, Query, Request
from fastapi.exception_handlers import request_validation_exception_handler
from fastapi.exceptions import RequestValidationError
from fastapi.responses import (
    FileResponse,
    HTMLResponse,
    JSONResponse,
    RedirectResponse,
    Response,
)

from .anti_abuse import AntiAbuseMiddleware
from .anti_abuse import ops_router as anti_abuse_ops_router
from .private_access import PrivateOperationsAccessMiddleware, load_access_settings
from .public_indexing import (
    PublicIndexBoundaryMiddleware,
    control_headers,
    public_indexing_enabled,
    robots_body,
    sitemap_body,
)
from .secret_hygiene import SecretHygieneMiddleware, install_secret_redaction
from .walking_skeleton import API_PREFIX as WALKING_API_PREFIX
from .walking_skeleton import router as walking_skeleton_router

REPO = Path(__file__).resolve().parents[4]
KMFA = REPO / "KMFA"
FACTS = KMFA / "machine" / "facts"
LINEAGE_PATH = KMFA / "machine" / "lineage.yaml"
STAGE_ARTIFACTS = KMFA / "stage_artifacts"
ASSERTIONS_PATH = KMFA / "metadata" / "quality" / "assertions.jsonl"
# PROD.0002 铁律：API 永不直读 raw inbox（KMDatabase/data）——只吃 machine/facts、
# machine/lineage.yaml、metadata/quality、stage_artifacts 这些已治理的派生/证据面。
FORBIDDEN_READ_ROOT = REPO / "KMDatabase" / "data"

app = FastAPI(
    title="KMFA App",
    version="0.2.0-prod0002",
    openapi_url="/ops/openapi.json",
    docs_url="/ops/docs",
    redoc_url=None,
)
app.add_middleware(PrivateOperationsAccessMiddleware)
# Keep this outermost so even 403/503 responses produced by the private guard
# receive the crawler/cache boundary.
app.add_middleware(PublicIndexBoundaryMiddleware)
# Keep the abuse gate inside secret hygiene so challenged/limited responses
# inherit the same no-leak browser boundary. It holds concurrency leases until
# the final ASGI response body, including streamed artifact downloads.
app.add_middleware(AntiAbuseMiddleware)
# Outermost: redact request/error records and attach the browser boundary even
# to responses short-circuited by the indexing or private-operations guards.
install_secret_redaction()
app.add_middleware(SecretHygieneMiddleware)
app.include_router(walking_skeleton_router)
app.include_router(anti_abuse_ops_router)


@app.exception_handler(RequestValidationError)
async def sanitize_walking_validation_error(
    request: Request,
    error: RequestValidationError,
) -> Response:
    """Keep capability-shaped request values out of public validation errors."""

    path = request.url.path
    if path == WALKING_API_PREFIX or path.startswith(f"{WALKING_API_PREFIX}/"):
        return JSONResponse(
            {"detail": "request_validation_failed"},
            status_code=422,
        )
    return await request_validation_exception_handler(request, error)


def _paginate(rows: list[Any], page: int, size: int) -> tuple[list[Any], dict[str, int]]:
    size = max(1, min(int(size), 500))
    page = max(1, int(page))
    total = len(rows)
    pages = max(1, (total + size - 1) // size)
    start = (page - 1) * size
    return rows[start : start + size], {"page": page, "size": size, "total": total, "pages": pages}
FRONTEND_DIST = Path(__file__).resolve().parents[2] / "frontend" / "dist"
PUBLIC_SHELL_MODULE_RE = re.compile(
    r'<script\b[^>]*\bid=["\']kmfa-app-module["\'][^>]*>\s*</script>',
    flags=re.IGNORECASE,
)
PUBLIC_INDEX_META = '<meta name="robots" content="index,follow,max-snippet:-1">'
PUBLIC_HOLD_META = '<meta name="robots" content="noindex,nofollow,noarchive">'


def _public_shell_enabled() -> bool:
    """默认启用公共增强壳；仅识别明确的关闭值，避免配置笔误静默回滚。"""
    return os.environ.get("KMFA_PUBLIC_SHELL_ENABLED", "1").strip().lower() not in {
        "0",
        "false",
        "no",
        "off",
    }


def _frontend_index() -> Path:
    index_path = FRONTEND_DIST / "index.html"
    if not index_path.exists():
        raise HTTPException(status_code=503, detail="前端未构建（KMFA/app/frontend: npm run build）")
    return index_path


@app.api_route("/", methods=["GET", "HEAD"], include_in_schema=False)
def index():
    index_path = _frontend_index()
    # 入口 HTML 禁缓存且禁止边缘层改写，避免代理自动注入第三方脚本；
    # 内容哈希资产仍由 /assets 单独永久缓存。
    headers = {
        "Cache-Control": "no-cache, must-revalidate, no-transform",
        "X-KMFA-Shell-Mode": "public-app",
    }
    shell_enabled = _public_shell_enabled()
    indexing_enabled = public_indexing_enabled()
    if not shell_enabled or not indexing_enabled:
        html = index_path.read_text(encoding="utf-8")
        if not indexing_enabled:
            replacements = html.count(PUBLIC_INDEX_META)
            if replacements != 1:
                raise HTTPException(status_code=503, detail="索引回退入口不可用")
            html = html.replace(PUBLIC_INDEX_META, PUBLIC_HOLD_META)
    if not shell_enabled:
        # 快速回滚只关闭增强 JS，不改路由、不动数据、不打开私有面。稳定静态壳仍含六个入口
        # 与清晰状态；marker 缺失时 fail-closed，避免误以为已经回滚。
        stable_html, replacements = PUBLIC_SHELL_MODULE_RE.subn("", html)
        if replacements != 1:
            raise HTTPException(status_code=503, detail="稳定静态入口不可用")
        headers["X-KMFA-Shell-Mode"] = "stable-static"
        return Response(stable_html, media_type="text/html", headers=headers)
    if not indexing_enabled:
        return Response(html, media_type="text/html", headers=headers)
    return FileResponse(
        index_path,
        headers=headers,
    )


@app.api_route("/ops/app/{app_path:path}", methods=["GET", "HEAD"], include_in_schema=False)
@app.api_route("/ops/app", methods=["GET", "HEAD"], include_in_schema=False)
@app.api_route("/ops/app/", methods=["GET", "HEAD"], include_in_schema=False)
def private_operations_app(app_path: str | None = None):
    """保留既有经营仪表盘的数据访问路径，并置于已守卫的 /ops 私有面。"""
    return FileResponse(
        _frontend_index(),
        headers={
            "Cache-Control": "no-cache, must-revalidate, no-transform",
            "X-KMFA-App-Mode": "private-operations",
        },
    )


@app.get("/ops/daily-funds", include_in_schema=False)
def daily_funds_private_entry():
    """Stable private deep link; the actual UI remains the shared KMFA app."""
    return RedirectResponse(url="/ops/app?tab=%E6%AF%8F%E6%97%A5%E8%B5%84%E9%87%91", status_code=307)


@app.api_route("/ui/{legacy_path:path}", methods=["GET", "HEAD"], include_in_schema=False)
@app.api_route("/ui", methods=["GET", "HEAD"], include_in_schema=False)
@app.api_route("/ui/", methods=["GET", "HEAD"], include_in_schema=False)
def legacy_ui_redirect(legacy_path: str | None = None):
    # 兼容旧书签/旧深链，但整个旧 UI 子树只允许单跳永久归一到根路径。
    return RedirectResponse(url="/", status_code=308)


@app.api_route("/assets/{asset_path:path}", methods=["GET", "HEAD"], include_in_schema=False)
def frontend_assets(asset_path: str):
    target = (FRONTEND_DIST / "assets" / asset_path).resolve()
    asset_root = (FRONTEND_DIST / "assets").resolve()
    if not target.is_relative_to(asset_root) or not target.is_file():
        raise HTTPException(status_code=404)
    # 资产文件名含内容哈希，内容一变名字必变——可放心永久缓存
    return FileResponse(target, headers={"Cache-Control": "public, max-age=31536000, immutable"})


#: 服务端自己渲染的页面（`/项目成本`）用的脚本。**必须外链**：CSP 是
#: `script-src 'self'`，内联 <script> 一律被拒——而 `style-src` 带 'unsafe-inline'，
#: 于是页面样式照常、交互全哑，看上去毫无异样。详见 project_cost.js 顶部。
SERVER_PAGE_STATIC = Path(__file__).resolve().parent / "static"


@app.api_route("/static/{name}", methods=["GET", "HEAD"], include_in_schema=False)
def server_page_static(name: str):
    target = (SERVER_PAGE_STATIC / name).resolve()
    if (
        not target.is_relative_to(SERVER_PAGE_STATIC.resolve())
        or not target.is_file()
        or target.suffix != ".js"
    ):
        raise HTTPException(status_code=404)
    # 文件名里**没有**内容哈希，所以不能 immutable：那会让一次改错的
    # JS 永久钉在用户浏览器里，改好了也下不来。必须每次回源核对。
    return FileResponse(
        target,
        media_type="text/javascript; charset=utf-8",
        headers={"Cache-Control": "no-cache"},
    )


def load_json(path: Path):
    if not path.exists():
        raise HTTPException(status_code=503, detail=f"事实文件缺失: {path.name}")
    return json.loads(path.read_text(encoding="utf-8"))


@app.get("/healthz")
def healthz():
    # 公共浅健康不得泄露 facts、数据库、队列、扫描器或部署细节。
    return {"status": "ok"}


@app.get("/ops/healthz")
def operations_healthz():
    return {"status": "ok", "facts_dir_present": FACTS.is_dir()}


@app.api_route("/robots.txt", methods=["GET", "HEAD"], include_in_schema=False)
def robots():
    return Response(
        robots_body(),
        media_type="text/plain",
        headers=control_headers(),
    )


@app.api_route("/sitemap.xml", methods=["GET", "HEAD"], include_in_schema=False)
def sitemap():
    return Response(
        sitemap_body(),
        media_type="application/xml",
        headers=control_headers(),
    )


def _quality_grade_short() -> str:
    """页眉质量等级取自 data_pipeline 事实，而非硬编码。

    PROD.0004 要求「数据来自 facts」——原实现把 "Q3" 写死在代码里，
    facts 一旦升级（如 Q3→Q4）页眉会静默说谎。
    """
    pipeline = json.loads((FACTS / "data_pipeline.json").read_text(encoding="utf-8"))
    full = str(pipeline.get("quality_grade_current") or "").strip()
    return full.split("（")[0].strip() or full or "未知"


def _quality_plain() -> str:
    """质量档位的人话：取 quality_grade_current 括号内的说明。

    Owner 2026-07-25 反馈「三徽章看不懂」——徽章只给 Q3/D/NO_GO 代码，非技术用户读不懂。
    这里把 facts 里本就写好的人话说明抽出来给前端，仍是 facts 单一真源，不硬编码。
    """
    pipeline = json.loads((FACTS / "data_pipeline.json").read_text(encoding="utf-8"))
    full = str(pipeline.get("quality_grade_current") or "").strip()
    if "（" in full and "）" in full:
        return full.split("（", 1)[1].rsplit("）", 1)[0].strip()
    return ""


@app.get("/api/状态")
def status():
    s = load_json(FACTS / "status.json")
    report_grade_full = str(s.get("report_grade") or "").strip()
    # "D —— 缺失、过期、失败或关键差异未闭合" → 破折号后的人话原因
    report_plain = report_grade_full.split("——", 1)[1].strip() if "——" in report_grade_full else ""
    verdict = str(s.get("business_verdict") or "")
    可对外 = "NO_GO" not in verdict.upper()
    return {
        "版本": s.get("version"), "阶段": s.get("stage"), "当前任务": s.get("task"),
        "真实进度": s.get("real_progress"),
        "页眉": {
            "质量等级": _quality_grade_short(),
            "质量人话": _quality_plain(),
            "报告等级": s.get("report_grade"),
            "报告人话": report_plain,
            "GO状态": s.get("business_verdict"),
            "可对外": 可对外,
            "交付人话": "可对外使用" if 可对外 else "暂不能对外",
        },
    }


@app.get("/api/我在哪")
def where_am_i():
    """首页「我在哪」（PROD.0004）——与 `文档/00_我在哪.md` 渲染件**同源**。

    同吃 machine/facts 的 status.json / blockers.json / roadmap.json（渲染件的事实源
    在其文件头已声明），确保页面与渲染件字字对得上；验收即以该渲染件为基准。
    """
    status_facts = load_json(FACTS / "status.json")
    blockers = load_json(FACTS / "blockers.json")
    roadmap = load_json(FACTS / "roadmap.json")
    pipeline = load_json(FACTS / "data_pipeline.json")
    stages = roadmap.get("stages", []) if isinstance(roadmap, dict) else []
    blocker_rows = blockers if isinstance(blockers, list) else []
    return {
        "更新于": status_facts.get("rendered_at"),
        "当前状态": {
            "版本": status_facts.get("version"),
            "阶段": status_facts.get("stage"),
            "分期": status_facts.get("phase"),
            "任务": status_facts.get("task"),
            "进度": status_facts.get("real_progress"),
            "报告可信度": status_facts.get("report_grade"),
            "业务结论": status_facts.get("business_verdict"),
            "证据状态": status_facts.get("evidence_status"),
            "卡住件数": len(blocker_rows),
        },
        "卡住的事": blocker_rows,
        "路线图": {"合计": len(stages), "阶段": stages},
        "数据面": {
            "质量等级": pipeline.get("quality_grade_current"),
            "截止批次": pipeline.get("data_as_of_batch"),
        },
        "同源": "machine/facts/{status,blockers,roadmap}.json —— 与 文档/00_我在哪.md 同源",
    }


@app.get("/api/数据管线")
def data_pipeline():
    return load_json(FACTS / "data_pipeline.json")


def _load_assertions() -> list[dict[str, Any]]:
    if not ASSERTIONS_PATH.exists():
        raise HTTPException(status_code=503, detail="断言表缺失")
    return [json.loads(l) for l in ASSERTIONS_PATH.read_text(encoding="utf-8").splitlines() if l.strip()]


@app.get("/api/断言")
def assertions(status: str | None = None, domain: str | None = None, page: int = 1, size: int = 50):
    """断言表（支持按状态/域过滤 + 分页）。

    顶层 total/closed/analyzed_open/items 保持既有契约不变（前端概览页依赖），
    过滤与分页为 PROD.0002 新增能力。
    """
    rows = _load_assertions()
    closed = sum(1 for r in rows if str(r.get("status", "")).startswith("closed"))
    selected = rows
    if status:
        selected = [r for r in selected if str(r.get("status", "")) == status]
    if domain:
        selected = [r for r in selected if str(r.get("domain", "")) == domain]
    items, meta = _paginate(selected, page, size)
    return {
        "total": len(rows),
        "closed": closed,
        "analyzed_open": len(rows) - closed,
        "筛选": {"status": status, "domain": domain, "命中": len(selected)},
        "分页": meta,
        "状态清单": sorted({str(r.get("status")) for r in rows if r.get("status")}),
        "域清单": sorted({str(r.get("domain")) for r in rows if r.get("domain")}),
        "items": items,
    }


@app.get("/api/技能")
def skills():
    import re

    def clean(raw: str | None) -> str | None:
        if raw is None:
            return None
        text = re.sub(r"\s+#.*$", "", raw).strip()
        return text.strip('"') or None

    reg = (KMFA / "skills" / "registry.yaml").read_text(encoding="utf-8")
    skills_block = reg.split("\nschedules:")[0]
    items = []
    for chunk in re.split(r"^  - id: ", skills_block, flags=re.M)[1:]:
        def field(name: str) -> str | None:
            m = re.search(rf"^    {name}: (.+)$", chunk, re.M)
            return clean(m.group(1)) if m else None

        deps = field("external_deps") or "[]"
        scheds = field("schedules") or "[]"
        items.append({
            "id": chunk.split("\n", 1)[0].strip(),
            "名称": field("name_zh"),
            "用途": field("purpose_zh"),
            "登记状态": field("status"),
            "排程": [s.strip(' "') for s in scheds.strip("[]").split(",") if s.strip(' "')],
            "外部依赖": [s.strip(' "') for s in deps.strip("[]").split(",") if s.strip(' "')],
            "本地路径硬编码": int(field("hardcoded_local_paths") or 0),
        })
    return {"count": len(items), "skills": items}


# ── PROD.0002 新增：事实八件套 / 血缘图 / 报表清单 ──────────────────────────────

@app.get("/api/事实")
def facts_index():
    """机器面事实文件清单（machine/facts）——页面据此发现可用事实，不硬编码文件名。"""
    if not FACTS.is_dir():
        raise HTTPException(status_code=503, detail="事实目录缺失")
    items = [
        {"名": p.stem, "格式": p.suffix.lstrip("."), "字节": p.stat().st_size}
        for p in sorted(FACTS.iterdir())
        if p.is_file() and p.suffix in (".json", ".yaml", ".yml")
    ]
    return {"count": len(items), "items": items}


@app.get("/api/事实/{name}")
def facts_one(name: str):
    """按名取单个事实文件。白名单式解析 + 目录逃逸防护（永不越出 machine/facts）。

    路径参数名刻意用 ASCII `name`：非 ASCII 参数名会被 Starlette 编成非法的正则命名
    捕获组，导致该路由在真实 HTTP 下永远不匹配（TestClient 却能过——本单元实测踩到，
    正是「只跑单测不真起服务」会漏掉的那类缺陷）。
    """
    for suffix in (".json", ".yaml", ".yml"):
        candidate = (FACTS / f"{name}{suffix}").resolve()
        if candidate.is_file() and candidate.parent == FACTS.resolve():
            text = candidate.read_text(encoding="utf-8")
            return json.loads(text) if suffix == ".json" else yaml.safe_load(text)
    raise HTTPException(status_code=404, detail=f"无此事实文件: {name}")


@app.get("/api/血缘")
def lineage(include_graph: bool = False, page: int = 1, size: int = 100):
    """血缘图：覆盖统计 + 派生表 + 完备判定；include_graph=true 时附节点/边（分页）。"""
    if not LINEAGE_PATH.exists():
        raise HTTPException(status_code=503, detail="血缘图缺失（跑 KMFA/tools/lineage_graph.py build）")
    graph = yaml.safe_load(LINEAGE_PATH.read_text(encoding="utf-8")) or {}
    nodes = graph.get("nodes") or []
    edges = graph.get("edges") or []
    out: dict[str, Any] = {
        "schema": graph.get("schema"),
        "生成自": graph.get("generated_from", []),
        "覆盖类别": graph.get("covered_categories", []),
        "覆盖": {
            "原始资产": graph.get("raw_assets"),
            "有派生边": graph.get("raw_with_staging_edges"),
            "整表延后": graph.get("raw_deferred_all_sheets"),
            "未抽取": graph.get("raw_not_yet_extracted"),
        },
        "派生表": graph.get("staging_tables", []),
        "完备": {"v1": graph.get("lineage_complete_v1"), "说明": graph.get("lineage_complete_note")},
        "规模": {"节点": len(nodes), "边": len(edges)},
    }
    if include_graph:
        node_items, node_meta = _paginate(nodes, page, size)
        edge_items, edge_meta = _paginate(edges, page, size)
        out["图"] = {"节点": node_items, "节点分页": node_meta, "边": edge_items, "边分页": edge_meta}
    return out


SOURCE_MATRIX_PATH = KMFA / "metadata" / "sources" / "source_check_matrix.jsonl"
AGING_MANIFEST_PATH = KMFA / "metadata" / "reports" / "collection_receivable_aging_manifest.json"
AGING_LANES_PATH = KMFA / "metadata" / "reports" / "collection_receivable_aging_source_lanes.jsonl"
AGING_ITEMS_PATH = KMFA / "metadata" / "reports" / "collection_receivable_aging_priority_items.jsonl"
AGING_STAGING_TABLES = ("receivable_aging", "collection", "v_collection_authoritative")


def _cents_to_yuan(cents: Any) -> str | None:
    """整数分 → 元字符串。**全程整数运算，禁用浮点**（金额纪律：恒整数分）。"""
    if cents is None:
        return None
    value = int(cents)
    sign = "-" if value < 0 else ""
    abs_value = abs(value)
    return f"{sign}{abs_value // 100:,}.{abs_value % 100:02d}"


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return [json.loads(l) for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]


@app.get("/api/账龄回款")
def receivable_aging():
    """应收账龄与回款视图（PROD.0010）——`collection_receivable_aging` 真数据化。

    数据分两层，**如实区分**：
    · **对账层（真数字）**：断言表 collection 域逐月 delta_cents 与 receivable_aging 恒等式，
      皆为已核到分的真实结果，直接呈现。
    · **v014 账龄结构层（值被阻断）**：source_lanes 全部 `data_status=structure_available_values_blocked`，
      priority_items 只有 `public_aging_bucket_ref_00x` 匿名指针、`collection_action_allowed=false`
      ——**故本页不产出账龄分桶金额**，只报结构与阻断状态。
    """
    rows = _load_assertions()
    collection_rows = [r for r in rows if str(r.get("domain")) == "collection"]
    aging_rows = [r for r in rows if str(r.get("domain")) == "receivable_aging"]

    monthly = []
    for r in sorted(collection_rows, key=lambda x: str(x.get("period"))):
        delta = r.get("delta_cents")
        monthly.append({
            "断言": r.get("assertion_id"),
            "口径": r.get("metric"),
            "期间": r.get("period"),
            "差异分": delta,
            "差异元": _cents_to_yuan(delta),
            "状态": r.get("status"),
            "证据": r.get("evidence_ref"),
        })
    zero = [m for m in monthly if m["差异分"] == 0]
    open_rows = [m for m in monthly if str(m["状态"]) == "analyzed_open"]
    with_delta = [m for m in monthly if isinstance(m["差异分"], int) and m["差异分"] != 0]
    largest = max(with_delta, key=lambda m: abs(m["差异分"]), default=None)

    manifest = json.loads(AGING_MANIFEST_PATH.read_text(encoding="utf-8")) if AGING_MANIFEST_PATH.exists() else {}
    lanes = _read_jsonl(AGING_LANES_PATH)
    items = _read_jsonl(AGING_ITEMS_PATH)
    pipeline = load_json(FACTS / "data_pipeline.json")
    staging = pipeline.get("staging_tables") or {}

    return {
        "回款对账": {
            "月数": len(monthly),
            "零分差月数": len(zero),
            "未闭月数": len(open_rows),
            "最大差异": ({"期间": largest["期间"], "差异分": largest["差异分"], "差异元": largest["差异元"]}
                         if largest else None),
            "逐月": monthly,
        },
        "账龄恒等式": [
            {"断言": r.get("assertion_id"), "口径": r.get("metric"), "快照": r.get("period"),
             "差异分": r.get("delta_cents"), "状态": r.get("status"), "证据": r.get("evidence_ref")}
            for r in aging_rows
        ],
        "账龄结构层": {
            "公式版本": manifest.get("formula_version"),
            "报告版本": manifest.get("report_version"),
            "生成于": manifest.get("generated_at"),
            "源泳道数": len(lanes),
            "泳道数据状态": sorted({str(l.get("data_status")) for l in lanes}),
            "优先事项数": len(items),
            "允许作经营依据": bool((manifest.get("quality_gate") or {}).get("business_decision_basis_allowed")),
            "允许催收动作": all(not i.get("collection_action_allowed") for i in items) is False,
            "限制": manifest.get("limitations", []),
        },
        "派生层规模": [
            {"表": name, "行数": (staging.get(name) or {}).get("rows")}
            for name in AGING_STAGING_TABLES if name in staging
        ],
        "诚实边界": ("对账层为已核到分的真实结果；账龄分桶金额在 v014 结构层仍 values_blocked，"
                     "本页不产出分桶金额，亦不构成催收依据。"),
    }


QUALITY_DIR = KMFA / "metadata" / "quality"
# v014 S14 三段：P1 资金/现金/贷款计划、P2 开票/纳税计划、P3 税务政策证据
S14_P1 = QUALITY_DIR / "v014_s14_p1_post_remediation_fund_cash_loan_plan"
S14_P2 = QUALITY_DIR / "v014_s14_p2_post_remediation_invoice_tax_plan"
S14_P3 = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_evidence_plan"
POLICY_RISKS_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_risk_tips_public_safe.json"
POLICY_GAPS_PATH = QUALITY_DIR / "v014_s14_p3_post_remediation_policy_evidence_gaps_public_safe.json"
# 开票/纳税/贷款三域会吃的派生层表（行数取自 data_pipeline 事实）
INVOICE_STAGING_TABLES = ("invoice_raw", "invoice_lines", "tax_composition", "loan_register")


def _assertion_row(r: dict[str, Any]) -> dict[str, Any]:
    """断言 → 展示行。差异分**原样透传**，元值走整数换算，绝不另造一套数。"""
    delta = r.get("delta_cents")
    return {
        "断言": r.get("assertion_id"),
        "口径": r.get("metric"),
        "期间": r.get("period"),
        "差异分": delta,
        "差异元": _cents_to_yuan(delta),
        "状态": r.get("status"),
        "对账方": r.get("expect_source"),
        "我方源": r.get("our_source"),
        "结论": r.get("finding"),
        "证据": r.get("evidence_ref"),
    }


def _s14_lanes_and_methods(manifest_path: Path, method_keys: tuple[str, ...]) -> dict[str, Any]:
    """读 v014 S14 manifest 的车道与方法定义——只报结构与阻断状态，不取任何金额。"""
    if not manifest_path.exists():
        return {"车道": [], "方法": [], "缺失": manifest_path.name}
    m = json.loads(manifest_path.read_text(encoding="utf-8"))
    lanes = [
        {
            "车道": l.get("lane_id"),
            "数据状态": l.get("data_status"),
            "私有候选表数": (l.get("private_candidate_sheet_count")
                             or l.get("private_direct_candidate_sheet_count")),
            "含业务金额": bool(l.get("contains_business_amounts")),
            "允许作经营依据": bool(l.get("business_decision_basis_allowed")),
        }
        for l in (m.get("source_lanes") or [])
    ]
    methods = []
    for key in method_keys:
        for d in (m.get(key) or []):
            lanes_needed = d.get("required_lanes") or []
            # 三个 cash_summary 方法事实里没有 method_note；兜底句由 required_lanes 推出，
            # 且必须落在 API 而不是某个页面——否则导出/自动化拿到的仍是 null。
            note = d.get("method_note") or (
                f"需 {'、'.join(lanes_needed)} 车道的权威期间值绑定后才能出汇总。"
                if lanes_needed else None)
            methods.append({
                "方法组": key,
                "方法": d.get("method_id"),
                # 事实里每个方法都带中文 visible_name 与 required_lanes；只显英文 id、
                # 让三个无 method_note 的方法露出"—"是把已有事实丢了——真开页面时实测到。
                "名称": d.get("visible_name"),
                "依赖车道": lanes_needed,
                "定义完备": bool(d.get("method_definition_complete")),
                "产出状态": d.get("current_output_status"),
                "绑定状态": d.get("current_binding_status"),
                "说明": note,
            })
    return {"车道": lanes, "方法": methods}


@app.get("/api/开票纳税")
def invoice_tax_fund():
    """开票纳税与资金贷款视图（PROD.0011）——`invoice_tax_plan`/`fund_cash_loan_plan` 真数据化。

    与账龄页同构，**数据分两层如实区分**：
    · **对账层（真数字）**：断言表 invoicing / tax / loan 三域，皆已核到分，直接呈现。
      其中 `AST-LOAN-ZHONGLI-ADHERENCE` 为 0 分差已闭，`AST-TAX-AXIS-HBKM-2025` 39 格
      逐月逐税种仅差 1 分。
    · **v014 S14 结构层（值被阻断）**：P1 四车道 / P2 三车道全部
      `values_unproven`，九个方法（3 计划 + 3 现金汇总 + 3 问题复核）全部
      `blocked_no_authoritative_*_value_binding`——**故本页不产出计划金额、不列到期提示**。

    任务包红线（P1 §212 / P2 §213）：**不做付款操作、不做正式纳税申报**。红线计数直接读
    v014 summary 事实（非硬编码），任何一项非零都会被契约测试打回。
    """
    rows = _load_assertions()
    by_domain = {d: [r for r in rows if str(r.get("domain")) == d]
                 for d in ("invoicing", "tax", "loan")}

    def block(domain: str) -> dict[str, Any]:
        rs = [_assertion_row(r) for r in sorted(by_domain[domain], key=lambda x: str(x.get("assertion_id")))]
        return {
            "条数": len(rs),
            "零分差条数": len([r for r in rs if r["差异分"] == 0]),
            "未闭条数": len([r for r in rs if str(r["状态"]) == "analyzed_open"]),
            "逐条": rs,
        }

    p1 = load_json(Path(f"{S14_P1}_summary.json"))
    p2 = load_json(Path(f"{S14_P2}_summary.json"))
    p3 = load_json(Path(f"{S14_P3}_summary.json"))

    risks = {r.get("program_id"): r for r in
             (json.loads(POLICY_RISKS_PATH.read_text(encoding="utf-8")).get("risks") or []
              if POLICY_RISKS_PATH.exists() else [])}
    gaps = (json.loads(POLICY_GAPS_PATH.read_text(encoding="utf-8")).get("gaps") or []
            if POLICY_GAPS_PATH.exists() else [])
    policy = []
    for g in sorted(gaps, key=lambda x: x.get("gap_sequence") or 0):
        r = risks.get(g.get("program_id")) or {}
        policy.append({
            "项目": g.get("visible_name"),
            "风险等级": r.get("risk_level"),
            "风险提示": r.get("risk_tip"),
            "证据缺口": g.get("gap_summary"),
            "缺口状态": g.get("gap_status"),
            "证据完备": bool(g.get("evidence_complete")),
            "允许出资格结论": bool(g.get("formal_policy_qualification_conclusion_allowed")),
        })

    pipeline = load_json(FACTS / "data_pipeline.json")
    staging = pipeline.get("staging_tables") or {}

    return {
        "开票对账": block("invoicing"),
        "税务对账": block("tax"),
        "贷款对账": block("loan"),
        "派生层规模": [
            {"表": name, "行数": (staging.get(name) or {}).get("rows")}
            for name in INVOICE_STAGING_TABLES if name in staging
        ],
        "结构层": {
            "开票纳税计划（S14-P2）": {
                "决策": p2.get("decision"),
                "已证值绑定车道数": p2.get("value_binding_proven_lane_count"),
                "公开业务金额数": p2.get("public_business_amount_count"),
                **_s14_lanes_and_methods(Path(f"{S14_P2}_manifest.json"),
                                         ("cash_summary_methods", "issue_review_methods")),
            },
            "资金贷款计划（S14-P1）": {
                "决策": p1.get("decision"),
                "已证值绑定车道数": p1.get("value_binding_proven_lane_count"),
                "公开业务金额数": p1.get("public_business_amount_count"),
                **_s14_lanes_and_methods(Path(f"{S14_P1}_manifest.json"), ("planning_methods",)),
            },
        },
        "税务政策证据": {
            "项目数": p3.get("policy_program_count"),
            "证据完备项目数": p3.get("evidence_complete_program_count"),
            "证据缺口数": p3.get("evidence_gap_count"),
            "要求证据类目数": p3.get("required_evidence_category_total_count"),
            "逐项": policy,
        },
        "红线": {
            "开票次数": p2.get("invoice_issuance_count"),
            "纳税申报次数": p2.get("tax_filing_count"),
            "付款或动账次数": p2.get("payment_or_bank_operation_count"),
            "银行操作次数": p1.get("bank_operation_count"),
            "付款审批次数": p1.get("payment_approval_count"),
            "贷款管理动作数": p1.get("loan_management_action_count"),
            "政策申报提交次数": p3.get("policy_application_submission_count"),
            "补贴申请次数": p3.get("subsidy_application_count"),
        },
        "诚实边界": ("对账层为已核到分的真实结果（仲利摊销 0 分差已闭、税负率 39 格仅差 1 分）；"
                     "计划层在 v014 S14 仍 values_unproven，本页**不产出计划金额、不列贷款到期提示**。"
                     "税务政策部分只出证据缺口与风险提示，**不构成资格判断**，"
                     "且全线不开票、不申报、不付款、不动账。"),
    }


COST_MANIFEST_PATH = KMFA / "metadata" / "reports" / "project_cost_fact_layer_manifest.json"
COST_RECORDS_PATH = KMFA / "metadata" / "lineage" / "project_cost_fact_records.jsonl"
# 成本归集会吃的派生层表（下钻用；行数取自 data_pipeline 事实）
COST_STAGING_TABLES = ("expense_lines", "kingdee_ledger", "kingdee_voucher", "goods_movement",
                       "invoice_lines", "collection", "receivable_aging")


# 最近完工项目成本：由 skills 容器算好写进共享卷，App 只读不算。
# 这条路是验证过的——公开技能健康端点就是这么读到真实台账的。
RECENT_COST_PATH = Path(os.environ.get(
    "KMFA_RECENT_COST", "/var/log/kmfa/project_cost/recent_completed.json"))


DWS_CANDIDATE_PATH = Path(os.environ.get(
    "KMFA_DWS_CANDIDATES", "/var/log/kmfa/dws/candidate_groups.json"))
DWS_SELECTED_PATH = Path(os.environ.get(
    "KMFA_DWS_SELECTED", "/var/log/kmfa/dws/selected_groups.json"))


@app.get("/api/归档目标群")
def archive_target_groups():
    """上游归档的候选群与当前勾选。

    Owner 2026-07-27：「dws 上游存档是增量存档，他也需要前端控制器筛选目标群」。
    此前链路缺这一环：自举产出候选清单，归档要已确认清单，中间没人勾选——归档因此长期 rc=4。

    群名与群 ID 是敏感信息：本接口在 Access 之后的私有面，公开面永远不出这些。
    """
    if not DWS_CANDIDATE_PATH.exists():
        return {"可读": False,
                "原因": "候选群清单尚未产出：技能 dws-bootstrap-groups 未成功跑完一次"
                        "（它需要容器内 dws 处于已登录态）",
                "候选": [], "已选": []}
    try:
        candidates = json.loads(DWS_CANDIDATE_PATH.read_text(encoding="utf-8")).get("群", [])
    except (OSError, json.JSONDecodeError) as exc:
        return {"可读": False, "原因": f"候选清单无法解析：{type(exc).__name__}",
                "候选": [], "已选": []}
    selected = []
    if DWS_SELECTED_PATH.exists():
        try:
            selected = json.loads(DWS_SELECTED_PATH.read_text(encoding="utf-8")).get("已选群", [])
        except (OSError, json.JSONDecodeError):
            selected = []
    return {
        "可读": True, "候选": candidates, "已选": selected,
        "说明": "归档是增量的：只拉勾选的群。没有勾选时归档会停下并如实报告，不会去猜该拉哪些群。",
        "更新时间": datetime.fromtimestamp(DWS_CANDIDATE_PATH.stat().st_mtime, BEIJING).isoformat(),
    }


@app.post("/api/归档目标群")
async def save_archive_target_groups(request: Request):
    """保存勾选。只接受候选清单里存在的群 ID——不允许从请求里凭空引入新群。"""
    if not DWS_CANDIDATE_PATH.exists():
        raise HTTPException(status_code=409, detail="候选群清单尚未产出，无法保存勾选")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="请求体不是合法 JSON")
    wanted = body.get("已选群")
    if not isinstance(wanted, list) or not all(isinstance(x, str) for x in wanted):
        raise HTTPException(status_code=400, detail="已选群必须是字符串数组")
    known = {g.get("id") for g in
             json.loads(DWS_CANDIDATE_PATH.read_text(encoding="utf-8")).get("群", [])}
    unknown = [x for x in wanted if x not in known]
    if unknown:
        # 不接受请求里凭空出现的群：那等于让调用方指定归档去拉任意会话。
        raise HTTPException(status_code=400, detail=f"这些群不在候选清单里，拒绝保存：{unknown[:5]}")
    DWS_SELECTED_PATH.parent.mkdir(parents=True, exist_ok=True)
    DWS_SELECTED_PATH.write_text(json.dumps(
        {"schema_version": "kmfa.dws.selected_groups.v1", "已选群": wanted},
        ensure_ascii=False, indent=2), encoding="utf-8")
    return {"已保存": len(wanted), "生效时机": "下一次 upstream-archive 运行（每天 11:00，或部署后立即重试）"}


CUSTOMER_MARGIN_PATH = Path(os.environ.get(
    "KMFA_CUSTOMER_MARGIN", "/var/log/kmfa/project_cost/customer_margin.json"))


@app.get("/api/客户毛利")
def customer_margin():
    """客户口径毛利——分四档，绝不合并成一个总数。

    项目维度在账上大面积缺失，客户维度却是完整的，所以「哪些客户在赚钱」这个问题
    现有数据能给出可信答案。但关联方（集团自有公司之间的往来）与外部客户混算会严重失真，
    故分档呈现；数据异常（成本为负、零成本有收入等）逐条标注，不替读者算一个好看的数。

    读不到就说读不到，不拿空列表冒充『没有客户』。
    """
    if not CUSTOMER_MARGIN_PATH.exists():
        return {"可读": False,
                "原因": "刷新作业尚未产出：技能 project-cost-refresh 从未成功跑完一次",
                "客户": [], "诚实边界": "读不到就说读不到，不拿空列表冒充『没有客户』。"}
    try:
        payload = json.loads(CUSTOMER_MARGIN_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return {"可读": False, "原因": f"产物无法解析：{type(exc).__name__}", "客户": []}
    payload["可读"] = True
    payload["产出时间"] = datetime.fromtimestamp(
        CUSTOMER_MARGIN_PATH.stat().st_mtime, BEIJING).isoformat()
    return payload


DATA_SOURCE_MATRIX_PATH = Path(os.environ.get(
    "KMFA_DATA_SOURCE_MATRIX", "/var/log/kmfa/project_cost/data_source_matrix.json"))
DATA_SOURCE_MATRIX_CSV = Path(os.environ.get(
    "KMFA_DATA_SOURCE_MATRIX_CSV", "/var/log/kmfa/project_cost/data_source_matrix.csv"))


@app.get("/api/数据源矩阵")
def data_source_matrix():
    """四个平台各自的固定输入——**声明该有什么，实测实际有什么**。

    Owner 2026-07-27：「我根本看不到你的数据源矩阵」+「钉钉红圈金蝶WPS四个平台，
    每个平台都有自己的固定输入，你全部都需要让系统能做到自己定时定期收集上传整理」。

    这一页要能回答三个问题，缺一个都不算数：
      · 每个平台该给哪几个输入？（声明）
      · 实际到位了没有、多少行、数据截到哪天？（实测，不信文件自称的尺寸）
      · 哪些是系统自己在收，哪些还靠人工放文件？（采集现状——这才是「自动化到哪一步」）

    行数为什么必须实测：WPS/红圈导出的 xlsx 谎报尺寸，只查「文件在不在」的话
    那几个文件全是绿的——文件确实在，只是一行都读不出来。
    """
    if not DATA_SOURCE_MATRIX_PATH.exists():
        return {"可读": False,
                "原因": "刷新作业尚未产出：技能 project-cost-refresh 从未成功跑完一次",
                "平台": [], "诚实边界": "读不到就说读不到，不拿空列表冒充『没有数据源』。"}
    try:
        payload = json.loads(DATA_SOURCE_MATRIX_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return {"可读": False, "原因": f"产物无法解析：{type(exc).__name__}", "平台": []}
    payload["可读"] = True
    payload["产出时间"] = datetime.fromtimestamp(
        DATA_SOURCE_MATRIX_PATH.stat().st_mtime, BEIJING).isoformat()
    payload["可下载"] = DATA_SOURCE_MATRIX_CSV.exists()
    return payload


@app.get("/api/数据源矩阵/下载")
def data_source_matrix_csv():
    """把矩阵下载成 CSV。Owner 要能拿走，不是只能在页面上看。"""
    if not DATA_SOURCE_MATRIX_CSV.exists():
        raise HTTPException(status_code=404, detail="矩阵 CSV 尚未产出——技能 project-cost-refresh 未成功跑完")
    return FileResponse(DATA_SOURCE_MATRIX_CSV, media_type="text/csv; charset=utf-8",
                        filename="KMFA_数据源矩阵.csv")


@app.get("/api/项目毛利")
def project_margin():
    """Retire the legacy lower-bound-cost / upper-bound-margin surface."""

    raise HTTPException(
        status_code=410,
        detail=(
            "旧项目毛利上限接口已下线；请使用 /public-api/项目成本。"
            "新接口仅在成本完整性闭合后发布毛利与毛利率。"
        ),
    )


@app.get("/api/项目成本/完工")
def recent_completed_cost():
    """最近完工项目的成本——两个口径并排，不调平。

    Owner 2026-07-27：「我根本没有看到项目成本，我说了我要最近完工的项目成本」。
    所以数要出现在页面上，不是出现在导出的文件里。

    读不到就说读不到：产物缺失时返回 `可读:false` 与原因，**不返回空数组冒充"没有完工项目"**
    ——空数组和读不到在页面上长得一样，但意思完全相反。
    """
    if not RECENT_COST_PATH.exists():
        parent = RECENT_COST_PATH.parent
        if not parent.exists():
            reason = f"{parent} 不存在——app 容器没挂 kmfa-logs 卷（部署配置问题）"
        else:
            reason = "刷新作业尚未产出：技能 project-cost-refresh 从未成功跑完一次"
        return {"可读": False, "原因": reason, "项目": [],
                "诚实边界": "读不到就说读不到，不拿空列表冒充『没有完工项目』。"}
    try:
        payload = json.loads(RECENT_COST_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return {"可读": False, "原因": f"产物无法解析：{type(exc).__name__}", "项目": []}
    try:
        _assert_current_project_cost_runtime(payload)
    except ValueError as exc:
        return {
            "可读": False,
            "原因": f"项目成本运行态版本不兼容或不完整：{exc}",
            "项目": [],
        }
    payload["可读"] = True
    payload["产出时间"] = datetime.fromtimestamp(
        RECENT_COST_PATH.stat().st_mtime, BEIJING).isoformat()
    return payload


@app.get("/api/项目成本")
def project_cost():
    """项目成本页（PROD.0006）——与 `project_cost_fact_layer` 输出一致。

    **诚实边界（务必保留）**：事实层当前
    `fact_layer_status = structural_fact_layer_blocked_for_formal_calculation`，
    全部记录 `amount_calculation_performed=false`、`calculation_status=blocked_pending_quality_resolution`
    ——**金额从未计算**，因其依赖 A0 权威基准，而 A0 被 BLK-001（Owner 约 273 行字段确认）阻塞。
    故本接口**不产出任何毛利/现金毛利数字**，只如实报结构、槽位与阻塞链。

    公开面只出 sha256 与 `private_ref://` 指针：明文值不在本仓
    （`metric_values_public_committed=false`），符合任务包「私面数据只在本机 App 显示、不入导出默认」。
    """
    if not COST_MANIFEST_PATH.exists() or not COST_RECORDS_PATH.exists():
        raise HTTPException(status_code=503, detail="项目成本事实层产物缺失")
    manifest = json.loads(COST_MANIFEST_PATH.read_text(encoding="utf-8"))
    records = [json.loads(l) for l in COST_RECORDS_PATH.read_text(encoding="utf-8").splitlines() if l.strip()]
    blockers = load_json(FACTS / "blockers.json")
    pipeline = load_json(FACTS / "data_pipeline.json")
    staging = pipeline.get("staging_tables") or {}

    rows = []
    for rec in records:
        cost_slots = rec.get("cost_category_slots") or []
        metric_slots = rec.get("metric_slots") or []
        rows.append({
            "记录号": rec.get("fact_record_id"),
            "项目实体": rec.get("project_entity_ref"),
            "计算状态": rec.get("calculation_status"),
            "金额已计算": bool(rec.get("amount_calculation_performed")),
            "允许正式计算": bool(rec.get("formal_calculation_allowed")),
            "成本槽位": len(cost_slots),
            "指标槽位": len(metric_slots),
            "已登记哈希": len(rec.get("cost_category_hash_refs") or {}) + len(rec.get("metric_hash_refs") or {}),
            "明文已公开": bool(rec.get("metric_values_public_committed")),
            "证据": rec.get("evidence_ref"),
        })

    calculated = sum(1 for r in rows if r["金额已计算"])
    return {
        "事实层": {
            "状态": manifest.get("fact_layer_status"),
            "公式版本": manifest.get("formula_version"),
            "映射版本": manifest.get("mapping_version"),
            "生成于": manifest.get("generated_at"),
            "记录数": len(rows),
            "已算金额记录数": calculated,
        },
        "必需结构": {
            "成本类别": manifest.get("required_cost_categories", []),
            "事实指标": manifest.get("required_fact_metrics", []),
        },
        "记录": rows,
        "阻塞链": {
            "直接原因": "A0 权威基准未从真实来源生成 → 事实层不允许正式计算",
            "基准引用": (records[0].get("authority_baseline_ref") if records else None),
            "根阻塞": [
                {"编号": b.get("id"), "内容": b.get("内容"), "只有Owner可解": b.get("owner_only"),
                 "已卡": b.get("首次登记")}
                for b in (blockers if isinstance(blockers, list) else [])
            ],
        },
        "可下钻派生层": [
            {"表": name, "行数": (staging.get(name) or {}).get("rows")}
            for name in COST_STAGING_TABLES if name in staging
        ],
        "诚实边界": ("毛利/现金毛利在 A0 就位前无法计算，本页不产出任何金额数字；"
                     "明文值仅存私有面（private_ref），公开面只有 sha256 指纹。"),
    }


@app.get("/api/源检查")
def source_check():
    """源检查板（PROD.0005）：矩阵协议状态 + 真实源覆盖矩阵 + 新鲜度 stale 提示。

    诚实边界：正式源检查矩阵 `metadata/sources/source_check_matrix.jsonl` 目前只有
    protocol_header、**零已提交源行**（S03-P2 协议定义态，源行由 file_import_register 产出）。
    本接口如实报出该状态，**不编造 entity_ref / account_ref 等取不到的维度值充数**。
    覆盖矩阵取自血缘图 + data_pipeline 事实（皆为机械生成面）；新鲜度由
    `data_as_of_batch` 与血缘节点批次比对得出——全程不读 raw inbox。
    """
    if not LINEAGE_PATH.exists():
        raise HTTPException(status_code=503, detail="血缘图缺失")
    graph = yaml.safe_load(LINEAGE_PATH.read_text(encoding="utf-8")) or {}
    pipeline = json.loads((FACTS / "data_pipeline.json").read_text(encoding="utf-8"))

    header: dict[str, Any] = {}
    committed_rows = 0
    if SOURCE_MATRIX_PATH.exists():
        for line in SOURCE_MATRIX_PATH.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            row = json.loads(line)
            if row.get("record_type") == "protocol_header":
                header = row
            else:
                committed_rows += 1

    nodes = graph.get("nodes") or []
    matrix: dict[str, dict[str, int]] = {}
    batches: set[str] = set()
    for node in nodes:
        source = str(node.get("domain") or "未标注")
        state = str(node.get("status") or "已抽取")
        matrix.setdefault(source, {})
        matrix[source][state] = matrix[source].get(state, 0) + 1
        if node.get("batch"):
            batches.add(str(node["batch"]))

    states = sorted({s for row in matrix.values() for s in row})
    as_of = str(pipeline.get("data_as_of_batch") or "")
    newer = sorted(b for b in batches if as_of and b > as_of)

    return {
        "矩阵协议": {
            "schema": header.get("schema_version"),
            "阶段": header.get("stage_phase"),
            "状态": header.get("status"),
            "必需维度": header.get("required_dimensions", []),
            "允许状态": header.get("allowed_statuses", []),
            "已提交源行": committed_rows,
            "说明": "协议已定义；源行待 file_import_register 产出后提交（当前为零行，如实报出）",
        },
        "覆盖矩阵": {
            "源": sorted(matrix),
            "状态列": states,
            "行": [
                {"源": src, "合计": sum(matrix[src].values()), **{st: matrix[src].get(st, 0) for st in states}}
                for src in sorted(matrix)
            ],
            "资产合计": len(nodes),
        },
        "新鲜度": {
            "数据批次": as_of,
            "血缘批次": sorted(batches),
            "stale": bool(newer),
            "更新的批次": newer,
            "提示": ("发现比 data_as_of_batch 更新的批次，需重跑抽取→血缘→facts"
                     if newer else "无更新批次，覆盖面与事实批次一致"),
        },
        "派生层": {
            "表数": len(pipeline.get("staging_tables") or {}),
            "行合计": pipeline.get("staging_rows_total"),
            "质量等级": pipeline.get("quality_grade_current"),
        },
    }


def _report_title(report_dir: Path) -> str | None:
    human = report_dir / "human"
    if not human.is_dir():
        return None
    for md in sorted(human.glob("*.md")):
        for line in md.read_text(encoding="utf-8").splitlines():
            if line.strip().startswith("#"):
                return line.lstrip("#").strip()
    return None


@app.get("/api/报表")
def reports(q: str | None = None, page: int = 1, size: int = 50):
    """报表清单：《一致性证明与差异分析报告》各号，标题取自 human 正文首个标题。"""
    if not STAGE_ARTIFACTS.is_dir():
        raise HTTPException(status_code=503, detail="证据目录缺失")
    items = []
    for d in STAGE_ARTIFACTS.glob("DT5_DATA0019_report_no*"):
        if not d.is_dir():
            continue
        match = re.search(r"report_no(\d+)", d.name)
        files = sorted(p.relative_to(d).as_posix() for p in d.rglob("*") if p.is_file())
        items.append({
            "编号": int(match.group(1)) if match else 0,
            "目录": d.name,
            "标题": _report_title(d),
            "文件数": len(files),
            "文件": files[:20],
        })
    items.sort(key=lambda r: r["编号"])
    if q:
        items = [r for r in items if q in (r["标题"] or "") or q in r["目录"]]
    page_items, meta = _paginate(items, page, size)
    return {"count": len(items), "分页": meta, "items": page_items}


# ── PROD.0007 差异工作台：三选一决策 + append-only 回写 ────────────────────────
# 契约来自既有 KMFA/tools/manual_resolution_events.py（任务包原文：「走既有
# manual_resolution_events 契约」），不另造一套：
#   · 24 个必填字段、event_type=resolution_event、manual_action_kind=difference_handling
#   · append_only=true；**改主意不是改记录**——silent_update_allowed=false、
#     reversal_required_for_change=true，要改就追加一条 reverses_event_id 的冲正事件
#     （既有 MANEVT-S12P1-005 冲正 -003 即此模式）
#   · raw/source 层一律不可写：断言表 assertions.jsonl 是治理数据面，App 只读不改
APPROVALS_DIR = KMFA / "metadata" / "approvals"
REPO_EVENTS_PATH = APPROVALS_DIR / "manual_resolution_events.jsonl"
# 应用状态面与数据面分离（PROD.0001 的 D2=A 约定）：App 写的事件落**可写状态目录**，
# 绝不写进治理数据面。SQLite 状态面待 PROD.0001 建成后接管，契约与此处完全一致。
APP_STATE_DIR = Path(os.environ.get("KMFA_APP_STATE_DIR", "/var/lib/kmfa/state"))
# PROD.0001：应用状态面走 SQLite（D2=A）。append-only 由触发器在**库层**强制，
# 不再依赖"我们只用 'a' 模式打开文件"这种君子协定。
APP_DB_PATH = APP_STATE_DIR / "kmfa_app_state.sqlite3"
from app import app_state as _st  # noqa: E402

APP_EVENTS_PATH = APP_STATE_DIR / "manual_resolution_events.jsonl"

BEIJING = timezone(timedelta(hours=8))  # 业务锚 +0800，与技能容器挂钟一致

# 三选一决策 → 断言状态流转（open / closed / excluded）
DECISIONS: dict[str, dict[str, str]] = {
    "闭案": {"到状态": "closed", "reason_code": "DIFF_ACCEPTED_CLOSED",
             "event_action": "close_difference"},
    "排除": {"到状态": "excluded", "reason_code": "DIFF_OUT_OF_SCOPE_EXCLUDED",
             "event_action": "exclude_difference"},
    "保持未闭": {"到状态": "open", "reason_code": "DIFF_REMAIN_OPEN_PENDING_INPUT",
                 "event_action": "keep_difference_open"},
}
REQUIRED_EVENT_FIELDS = (
    "event_id", "schema_version", "record_type", "stage_phase", "event_type",
    "manual_action_kind", "actor_ref", "actor_role", "event_time", "reason_code",
    "reason_summary", "impact_scope", "event_version", "target_layer", "target_ref",
    "status", "append_only", "raw_layer_write_allowed", "raw_source_mutation_allowed",
    "source_layer_write_allowed", "business_plaintext_committed", "forbidden_plaintext",
    "evidence_refs",
)
# 公开面禁词（取自既有契约 FORBIDDEN_PUBLIC_KEYS 的金额/明文子集）——事件里只准放
# 断言号与理由，绝不准把金额或业务明文塞进来
FORBIDDEN_EVENT_KEYS = frozenset({
    "amount_cents", "amount_yuan", "raw_value", "normalized_value", "original_value",
    "plaintext_value", "source_header_text", "bank_account_number", "account_number",
    "identity_document_number", "project_name_plaintext", "customer_name_plaintext",
    "counterparty_plaintext",
})
APP_EVENT_ID_RE = re.compile(r"^MANEVT-APP-[0-9]{4}$")


def _read_events(path: Path) -> list[dict[str, Any]]:
    """仓内事件读 JSONL（治理数据面，只读）；App 自己写的读 SQLite 状态面。"""
    rows = _st.read(APP_DB_PATH, "resolution_events") if path == APP_EVENTS_PATH else _read_jsonl(path)
    return [r for r in rows if r.get("record_type") != "protocol_header"]


def _all_events() -> list[dict[str, Any]]:
    """仓内既有事件（只读）+ App 写的事件（应用状态面）。"""
    return _read_events(REPO_EVENTS_PATH) + _read_events(APP_EVENTS_PATH)


def _assertion_state(row: dict[str, Any]) -> str:
    """断言原始三态：open / closed / excluded（任务包要求的可视化分组）。"""
    status = str(row.get("status") or "")
    if status.startswith("closed"):
        return "closed"
    if "exclud" in status:
        return "excluded"
    return "open"


def _scan_forbidden(node: Any) -> list[str]:
    hits: list[str] = []
    if isinstance(node, dict):
        for k, v in node.items():
            if str(k) in FORBIDDEN_EVENT_KEYS:
                hits.append(str(k))
            hits.extend(_scan_forbidden(v))
    elif isinstance(node, list):
        for v in node:
            hits.extend(_scan_forbidden(v))
    return hits


def _content_hash(event: dict[str, Any]) -> str:
    payload = {k: v for k, v in event.items() if k != "content_hash"}
    blob = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return "sha256:" + hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _append_event(event: dict[str, Any]) -> dict[str, Any]:
    """写前全量校验，**只以追加方式落盘**，绝不改写既有行。"""
    missing = [f for f in REQUIRED_EVENT_FIELDS if f not in event]
    if missing:
        raise HTTPException(status_code=500, detail=f"事件缺必填字段：{missing}")
    forbidden = _scan_forbidden(event)
    if forbidden:
        raise HTTPException(status_code=400, detail=f"事件含禁写字段：{sorted(set(forbidden))}")
    for flag in ("raw_layer_write_allowed", "raw_source_mutation_allowed",
                 "source_layer_write_allowed", "business_plaintext_committed",
                 "forbidden_plaintext", "silent_update_allowed"):
        if event.get(flag) is not False:
            raise HTTPException(status_code=500, detail=f"{flag} 必须为 false")
    if event.get("append_only") is not True:
        raise HTTPException(status_code=500, detail="append_only 必须为 true")
    event["content_hash"] = _content_hash(event)

    _st.append(APP_DB_PATH, "resolution_events", event)
    return event


def _next_app_event_id() -> str:
    existing = [e.get("event_id", "") for e in _read_events(APP_EVENTS_PATH)]
    nums = [int(e.rsplit("-", 1)[1]) for e in existing if APP_EVENT_ID_RE.match(str(e))]
    return f"MANEVT-APP-{max(nums, default=0) + 1:04d}"


def _base_event(assertion_id: str, reason: str, actor: str) -> dict[str, Any]:
    return {
        "event_id": _next_app_event_id(),
        "schema_version": "kmfa.manual_resolution_event.v1",
        "record_type": "manual_resolution_event",
        "stage_phase": "DT6-PROD0007",
        "event_type": "resolution_event",
        "manual_action_kind": "difference_handling",
        "actor_ref": f"actor_ref://owner_or_authorized_delegate/{actor}",
        "actor_role": "owner_or_authorized_delegate",
        "event_time": datetime.now(BEIJING).isoformat(timespec="seconds"),
        "reason_summary": reason,
        "impact_scope": ["assertion_status_flow", "difference_workbench"],
        "event_version": "MANUAL-EVENT-KMFA-DT6-PROD0007-001",
        "target_layer": "quality",
        "target_ref": assertion_id,
        "project_id": "KMFA",
        "system_name": "KMFA 经营分析系统",
        "append_only": True,
        "raw_layer_write_allowed": False,
        "raw_source_mutation_allowed": False,
        "source_layer_write_allowed": False,
        "business_plaintext_committed": False,
        "forbidden_plaintext": False,
        "silent_update_allowed": False,
        "reversal_required_for_change": True,
        "approved_event_immutable": True,
        "evidence_refs": [
            "KMFA/metadata/quality/assertions.jsonl",
            "KMFA/tools/manual_resolution_events.py",
        ],
    }


@app.get("/api/差异工作台")
def difference_workbench():
    """差异工作台（PROD.0007）——断言 open/closed/excluded 可视化 + 决策留痕。

    「与 assertions.jsonl 双向一致」的落法：
    · 正向——每条断言挂出针对它的全部决策事件（含冲正）；
    · 反向——每条事件的 target_ref 必须能解析到真实断言，孤儿事件计数须为 0。
    断言表本身**只读**：状态流转由事件表达，App 绝不改写治理数据面。
    """
    rows = _load_assertions()
    events = _all_events()
    by_target: dict[str, list[dict[str, Any]]] = {}
    for e in events:
        by_target.setdefault(str(e.get("target_ref")), []).append(e)

    reversed_ids = {str(e.get("reverses_event_id")) for e in events if e.get("reverses_event_id")}

    def view(e: dict[str, Any]) -> dict[str, Any]:
        return {
            "事件号": e.get("event_id"),
            "动作": e.get("event_action"),
            "决策": e.get("decision_label"),
            "到状态": e.get("target_state"),
            "理由": e.get("reason_summary"),
            "理由码": e.get("reason_code"),
            "时间": e.get("event_time"),
            "操作人": e.get("actor_ref"),
            "已被冲正": e.get("event_id") in reversed_ids,
            "冲正的是": e.get("reverses_event_id"),
            "内容哈希": e.get("content_hash"),
            "来源": "仓内治理面（只读）" if e.get("stage_phase") != "DT6-PROD0007" else "App 应用状态面",
        }

    items = []
    for r in sorted(rows, key=lambda x: str(x.get("assertion_id"))):
        aid = str(r.get("assertion_id"))
        evs = sorted(by_target.get(aid, []), key=lambda e: str(e.get("event_time")))
        live = [e for e in evs if e.get("event_id") not in reversed_ids
                and not e.get("reverses_event_id")]
        items.append({
            "断言": aid,
            "域": r.get("domain"),
            "口径": r.get("metric"),
            "期间": r.get("period"),
            "原始状态": r.get("status"),
            "分组": _assertion_state(r),
            "差异分": r.get("delta_cents"),
            "差异元": _cents_to_yuan(r.get("delta_cents")),
            "结论": r.get("finding"),
            "决策事件": [view(e) for e in evs],
            "现行决策": view(live[-1]) if live else None,
        })

    known = {str(r.get("assertion_id")) for r in rows}
    # 「双向一致」的诚实口径：本台写入的事件**必须**条条解析到真实断言（写入时已 404 拦住）；
    # 仓内既有的 S12-P1 事件指向 S09P3-REC-001 等**治理记录号**而非断言号，它们不是孤儿、
    # 也不该被算成一致性缺陷——如实单列为「未挂载到断言层」。
    app_events = _read_events(APP_EVENTS_PATH)
    app_orphans = sorted({str(e.get("target_ref")) for e in app_events
                          if str(e.get("target_ref")) not in known})
    unmounted = sorted({str(e.get("target_ref")) for e in _read_events(REPO_EVENTS_PATH)
                        if str(e.get("target_ref")) not in known})
    groups = {g: len([i for i in items if i["分组"] == g]) for g in ("open", "closed", "excluded")}

    return {
        "分组计数": groups,
        "断言总数": len(items),
        "断言明细": items,
        "决策入口": [
            {"决策": k, "到状态": v["到状态"], "理由码": v["reason_code"]} for k, v in DECISIONS.items()
        ],
        "事件": {
            "总数": len(events),
            "仓内既有": len(_read_events(REPO_EVENTS_PATH)),
            "App 写入": len(_read_events(APP_EVENTS_PATH)),
            "已被冲正": len(reversed_ids),
            "写入位置": str(APP_EVENTS_PATH),
        },
        "双向一致": {
            "本台孤儿事件数": len(app_orphans),
            "本台孤儿事件": app_orphans,
            "一致": not app_orphans,
            "仓内未挂载事件数": len(unmounted),
            "仓内未挂载事件": unmounted,
            "说明": ("正向：每条断言挂出针对它的全部决策事件；反向：本台写入的事件条条解析到"
                     "真实断言（写入时即 404 拦截孤儿）。仓内既有 S12-P1 事件指向 "
                     "S09P3-REC-001 等治理记录号而非断言号，如实单列为未挂载，不算一致性缺陷。"),
        },
        "写入纪律": {
            "append_only": True,
            "允许静默改写": False,
            "改主意的做法": "追加一条 reverses_event_id 的冲正事件，绝不编辑既有行",
            "断言表可写": False,
        },
    }


@app.post("/api/差异工作台/决策")
def workbench_decide(payload: dict[str, Any] = Body(...)):
    """三选一决策入口——append-only 回写一条 difference_handling 事件。"""
    assertion_id = str(payload.get("断言") or "").strip()
    decision = str(payload.get("决策") or "").strip()
    reason = str(payload.get("理由") or "").strip()
    actor = str(payload.get("操作人") or "owner").strip() or "owner"

    if decision not in DECISIONS:
        raise HTTPException(status_code=400, detail=f"决策须为三选一：{list(DECISIONS)}")
    if not reason:
        raise HTTPException(status_code=400, detail="必须写明理由——无理由的决策不留痕等于没决策")
    known = {str(r.get("assertion_id")) for r in _load_assertions()}
    if assertion_id not in known:
        raise HTTPException(status_code=404, detail=f"断言不存在：{assertion_id}（拒绝写孤儿事件）")

    spec = DECISIONS[decision]
    event = _base_event(assertion_id, reason, actor)
    event.update({
        "event_action": spec["event_action"],
        "reason_code": spec["reason_code"],
        "decision_label": decision,
        "target_state": spec["到状态"],
        "status": "recorded_pending_approval",
        "approval_state": "draft",
    })
    written = _append_event(event)
    _audit("processing", subject_ref=assertion_id, result_status="OK",
           evidence_ref=str(APP_EVENTS_PATH), event_ref=written["event_id"],
           decision=decision)
    return {"已写入": written, "写入位置": str(APP_EVENTS_PATH)}


@app.post("/api/差异工作台/冲正")
def workbench_reverse(payload: dict[str, Any] = Body(...)):
    """改主意的唯一合法做法：追加冲正事件（silent_update_allowed=false）。"""
    target_event_id = str(payload.get("冲正事件号") or "").strip()
    reason = str(payload.get("理由") or "").strip()
    actor = str(payload.get("操作人") or "owner").strip() or "owner"
    if not reason:
        raise HTTPException(status_code=400, detail="冲正必须写明理由")

    events = {str(e.get("event_id")): e for e in _all_events()}
    origin = events.get(target_event_id)
    if origin is None:
        raise HTTPException(status_code=404, detail=f"被冲正事件不存在：{target_event_id}")
    if any(str(e.get("reverses_event_id")) == target_event_id for e in events.values()):
        raise HTTPException(status_code=409, detail=f"{target_event_id} 已被冲正过，不得重复冲正")

    event = _base_event(str(origin.get("target_ref")), reason, actor)
    event.update({
        "event_action": "reverse_difference_decision",
        "reason_code": "DECISION_REVERSED_BY_HUMAN",
        "decision_label": "冲正",
        "target_state": None,
        "status": "reverse_event_recorded",
        "approval_state": "reversal_recorded",
        "reverses_event_id": target_event_id,
    })
    written = _append_event(event)
    _audit("processing", subject_ref=str(origin.get("target_ref")), result_status="REVERSED",
           evidence_ref=str(APP_EVENTS_PATH), event_ref=written["event_id"],
           reverses=target_event_id)
    return {"已写入": written, "被冲正": target_event_id}


# ── PROD.0009 报告中心：HTML/CSV/PDF 三格式导出 + 不可去除的 D 级水印 ──────────
# 权威任务包第 18 行：HTML/CSV 导出承接既有 runtime；新增 PDF（复用 KMIDS PDF 管线）；
# 报告页眉强制显示等级/Q 级/delivery 状态。
# 验收：**三格式导出 hash 登记；D 级水印在解锁前不可去除**。
#
# 既有 runtime 契约（KMFA/tools/report_export_runtime.py）：
#   · HTML/CSV = public-safe 可提交；PDF = enabled_private_runtime_only，
#     committed_artifact_path 恒为 null——**公开仓永不提交 PDF 文件**。
#   · FORBIDDEN_PUBLIC_SUFFIXES 含 .pdf，故 PDF 只在运行时生成、只走响应流。
# KMIDS 管线本机不可得（按需 clone 铁律，未 clone），故 PDF 用 reportlab 内置
# STSong-Light CID 字体渲染中文——不装任何字体文件，策略与既有契约完全一致。
GRADE_RECORDS_PATH = KMFA / "metadata" / "reports" / "report_grade_runtime_records.jsonl"
DELIVERY_GATE_PATH = KMFA / "metadata" / "quality" / "v014_s18_p2_go_no_go_report.json"
EXPORT_REGISTRY_PATH = APP_STATE_DIR / "report_export_records.jsonl"
EXPORT_FORMATS = ("html", "csv", "pdf")


def _delivery_state() -> dict[str, Any]:
    """页眉三元组 + 水印判据——**全部取自事实**，没有任何请求参数能左右它。"""
    grade_rows = _read_jsonl(GRADE_RECORDS_PATH)
    grade_row = grade_rows[0] if grade_rows else {}
    inputs = grade_row.get("grade_inputs") or {}
    gate = load_json(DELIVERY_GATE_PATH) if DELIVERY_GATE_PATH.exists() else {}
    delivery_allowed = bool(gate.get("delivery_allowed"))
    return {
        "报告等级": grade_row.get("computed_report_grade") or "未知",
        "质量等级": inputs.get("source_quality_grade") or _quality_grade_short(),
        "delivery_allowed": delivery_allowed,
        "delivery状态": "已解锁" if delivery_allowed else "未解锁（NO_GO）",
        "正式报告可出": bool(grade_row.get("formal_report_allowed")),
        "可作经营依据": bool(grade_row.get("business_decision_basis_allowed")),
        "等级政策版本": grade_row.get("grade_policy_version"),
        "判据来源": [
            "KMFA/metadata/reports/report_grade_runtime_records.jsonl",
            "KMFA/metadata/quality/v014_s18_p2_go_no_go_report.json",
        ],
    }


def _watermark_text() -> str | None:
    """水印文案由事实推出。delivery 解锁前**恒非空**，且无参数可关。"""
    state = _delivery_state()
    if state["delivery_allowed"]:
        return None
    # 分隔符用全角竖线：U+00B7（·）不在 STSong-Light 的 UniGB-UCS2-H 映射里，
    # PDF 里会渲成黑三角豆腐块——真把 PDF 渲成图看才发现的。
    return (f"{state['报告等级']} 级 ｜ 未解锁不可作经营依据 ｜ "
            f"质量 {state['质量等级']} ｜ delivery_allowed=false")


def _report_dirs() -> list[Path]:
    return sorted((d for d in STAGE_ARTIFACTS.glob("DT5_DATA0019_report_no*") if d.is_dir()),
                  key=lambda d: int(re.search(r"report_no(\d+)", d.name).group(1)))


def _report_dir(no: int) -> Path:
    for d in _report_dirs():
        if int(re.search(r"report_no(\d+)", d.name).group(1)) == no:
            return d
    raise HTTPException(status_code=404, detail=f"报告不存在：第 {no} 号")


def _report_body(d: Path) -> str:
    docs = sorted((d / "human").glob("*.md"))
    if not docs:
        raise HTTPException(status_code=503, detail=f"{d.name} 缺 human 正文")
    return docs[0].read_text(encoding="utf-8")


def _export_html(no: int, title: str, body: str, header: dict[str, Any], mark: str | None) -> bytes:
    import html as _html

    # 水印用 ::before 伪元素 + 固定定位铺满，**不提供任何开关**；
    # 页眉三元组同样硬渲染进文档，不是可选装饰。
    band = "" if mark is None else f"""
  <div class="wm" aria-label="水印">{_html.escape(mark)}</div>
  <style>
    .wm {{ position: fixed; inset: 0; display: flex; align-items: center; justify-content: center;
          transform: rotate(-28deg); font-size: 2.2rem; font-weight: 800; color: rgba(192,57,43,.18);
          pointer-events: none; z-index: 9999; white-space: pre-wrap; text-align: center; }}
    @media print {{ .wm {{ display: flex !important; }} }}
  </style>"""
    rows = "\n".join(
        f"<tr><td>{_html.escape(str(k))}</td><td><b>{_html.escape(str(v))}</b></td></tr>"
        for k, v in (("报告等级", header["报告等级"]), ("质量等级", header["质量等级"]),
                     ("delivery 状态", header["delivery状态"]))
    )
    esc_body = _html.escape(body)
    return f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8">
<title>{_html.escape(title)}</title>
<style>
 body{{font-family:-apple-system,"PingFang SC",sans-serif;max-width:60rem;margin:0 auto;padding:2rem;line-height:1.7}}
 table{{border-collapse:collapse;margin:.8rem 0}} td{{border:1px solid #ccc;padding:.35rem .6rem}}
 pre{{white-space:pre-wrap;word-wrap:break-word}}
 header{{border-bottom:2px solid #174a7c;padding-bottom:.6rem}}
</style></head><body>{band}
<header><h1>{_html.escape(title)}</h1><table><tbody>{rows}</tbody></table></header>
<pre>{esc_body}</pre>
</body></html>""".encode("utf-8")


def _export_csv(no: int, d: Path, header: dict[str, Any], mark: str | None) -> bytes:
    import csv as _csv
    import io as _io

    disp_path = d / "machine" / "dispositions.json"
    disp = json.loads(disp_path.read_text(encoding="utf-8")) if disp_path.exists() else {}
    buf = _io.StringIO()
    w = _csv.writer(buf)
    # 水印与页眉三元组写成 CSV 前置行——任何打开方式都看得到，删不掉才算不可去除
    if mark:
        w.writerow(["水印", mark])
    w.writerow(["报告等级", header["报告等级"]])
    w.writerow(["质量等级", header["质量等级"]])
    w.writerow(["delivery 状态", header["delivery状态"]])
    w.writerow([])
    w.writerow(["条目", "状态", "差异分", "差异元", "结论"])
    for item in (disp.get("dispositions") or []):
        cents = item.get("delta_cents")
        w.writerow([item.get("item"), item.get("status"), cents,
                    _cents_to_yuan(cents) or "", item.get("finding") or ""])
    return buf.getvalue().encode("utf-8-sig")  # BOM：Excel 直接双击不乱码


# STSong-Light（UniGB-UCS2-H）渲不出的字符 → 可渲替身。
# 表是**实测出来的**：扫 8 份报告正文找 GBK 编不出的字符，再加上 U+00B7——
# 它 GBK 能编但实测仍渲成黑三角（把 PDF 渲成 PNG 看才发现）。
PDF_GLYPH_FALLBACKS = {
    "\u2705": "\u221a",   # ✅ → √
    "\u00a5": "\uffe5",   # ¥ → ￥（半角日元符渲不出，全角人民币符可以）
    "\u2194": "<->",      # ↔
    "\u2212": "-",        # − 数学减号
    "\u2213": "-/+",      # ∓
    "\u00b7": " - ",      # · 间隔号
    "\u2022": "-",        # •
    "\u2013": "-", "\u2014": "-",
}


def _pdf_safe(text: str) -> str:
    """把渲不出的字符换成可渲替身。**PDF 里所有落笔的文字都要过这一道。**"""
    for bad, good in PDF_GLYPH_FALLBACKS.items():
        text = text.replace(bad, good)
    return text


def _markdown_to_plain(body: str) -> list[str]:
    """把报告 markdown 压成可读纯文本行——PDF 里倒 `###`/`**`/`|---|` 原文没法看。

    表格保留内容、去掉分隔线；标题降为「N、」式；强调标记剥掉。数字一律不动。
    """
    lines: list[str] = []
    for raw in body.splitlines():
        line = raw.rstrip()
        if re.fullmatch(r"\s*\|?[\s\|:–—-]*\|[\s\|:–—-]*\|?\s*", line) and "-" in line:
            continue  # 表格分隔线
        line = re.sub(r"^\s{0,3}(#{1,6})\s*", "", line)
        line = re.sub(r"^\s{0,3}>\s?", "", line)
        line = line.replace("**", "").replace("`", "")
        if line.startswith("|"):
            cells = [c.strip() for c in line.strip("|").split("|")]
            line = "    ".join(c for c in cells if c)
        lines.append(line)
    return lines


def _export_pdf(no: int, title: str, body: str, header: dict[str, Any], mark: str | None) -> bytes:
    import io as _io

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfgen import canvas as _canvas
    except ImportError as exc:  # 缺依赖就明说，绝不悄悄产出一个假 PDF
        raise HTTPException(status_code=503, detail=f"PDF 管线不可用：{exc}") from exc

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    buf = _io.BytesIO()
    # invariant=1：固定 /CreationDate 与文档 ID。默认会写入挂钟时间，导致同一份报告
    # 每次导出字节都不同——那样登记的 hash 事后验不了任何东西（契约测试当场逮到）。
    c = _canvas.Canvas(buf, pagesize=A4, invariant=1)
    width, height = A4

    def stamp() -> None:
        """每一页都盖水印——翻页删不掉、抽页也删不掉。"""
        if not mark:
            return
        c.saveState()
        c.setFont("STSong-Light", 20)
        c.setFillColorRGB(0.75, 0.22, 0.17, alpha=0.18)
        c.translate(width / 2, height / 2)
        c.rotate(28)
        c.drawCentredString(0, 0, _pdf_safe(mark))
        c.restoreState()

    safe_title = _pdf_safe(title)
    c.setTitle(safe_title)

    def head(y: float) -> float:
        c.setFont("STSong-Light", 15)
        c.drawString(45, y, safe_title)
        c.setFont("STSong-Light", 9.5)
        c.drawString(45, y - 18, _pdf_safe(
            f"报告等级 {header['报告等级']} ｜ 质量等级 {header['质量等级']} "
            f"｜ delivery {header['delivery状态']}"))
        c.line(45, y - 26, width - 45, y - 26)
        return y - 44

    stamp()
    y = head(height - 50)
    body_size, usable = 9, width - 90
    c.setFont("STSong-Light", body_size)

    def wrap(text: str) -> list[str]:
        """按**实际字宽**折行。原来按字数切 46，会把 1,462,802.90 劈成两截——
        财务报告折断金额是硬伤，真把 PDF 渲成图看才发现。"""
        out, cur = [], ""
        for ch in text:
            trial = cur + ch
            if pdfmetrics.stringWidth(trial, "STSong-Light", body_size) > usable:
                out.append(cur)
                cur = ch
            else:
                cur = trial
        out.append(cur)
        return out or [""]

    for raw in (_pdf_safe(l) for l in _markdown_to_plain(body)):
        for chunk in wrap(raw):
            c.drawString(45, y, chunk)
            y -= 13
            if y < 50:
                c.showPage()
                stamp()
                y = head(height - 50)
                c.setFont("STSong-Light", body_size)
    c.save()
    return buf.getvalue()


def _register_export(record: dict[str, Any]) -> dict[str, Any]:
    """导出 hash 登记——与 PROD.0007 同一条纪律：只追加，不改写。"""
    return _st.append(APP_DB_PATH, "export_records", record)


@app.get("/api/报告中心")
def report_center():
    """报告中心（PROD.0009）——八份报告 × 三格式，页眉三元组与水印状态如实呈现。"""
    header = _delivery_state()
    mark = _watermark_text()
    registered = _st.read(APP_DB_PATH, "export_records")
    by_key: dict[str, dict[str, Any]] = {}
    for r in registered:
        by_key[f"{r.get('报告')}|{r.get('格式')}"] = r

    items = []
    for d in _report_dirs():
        no = int(re.search(r"report_no(\d+)", d.name).group(1))
        docs = sorted((d / "human").glob("*.md"))
        items.append({
            "编号": no,
            "标题": _report_title(d),
            "目录": d.name,
            "正文字数": len(docs[0].read_text(encoding="utf-8")) if docs else 0,
            "格式": [
                {
                    "格式": fmt,
                    # **不再给可导航的 GET 地址**（S07/T-S07-03）。
                    # 一个 `<a href>` 就是一次可被预取的 GET：浏览器、链接预览、
                    # 爬虫都会替用户按下它，而导出会写业务记录。
                    # 这里只留一个页内锚点标记（保留 `格式=` 供前端与流程审计定位），
                    # 真正的导出由前端 POST 受控任务发起。
                    "下载": f"#导出?报告={no}&格式={fmt}",
                    "导出任务": {
                        "方法": "POST", "路径": "/api/导出任务",
                        "请求体": {"报告": no, "格式": fmt},
                        "需要头": "Idempotency-Key",
                    },
                    "可提交公开仓": fmt != "pdf",
                    "已登记": by_key.get(f"{no}|{fmt}", {}).get("sha256"),
                }
                for fmt in EXPORT_FORMATS
            ],
        })

    return {
        "页眉": {"报告等级": header["报告等级"], "质量等级": header["质量等级"],
                 "delivery状态": header["delivery状态"]},
        "交付判据": header,
        "水印": {
            "文案": mark,
            "生效中": mark is not None,
            "可关闭": False,
            "去除条件": "delivery_allowed 由 false 转 true（Owner 签 GO），非任何前端/参数开关",
            "覆盖格式": list(EXPORT_FORMATS),
        },
        "报告": items,
        "导出登记": {
            "条数": len(registered),
            "位置": str(EXPORT_REGISTRY_PATH),
            "追加式": True,
            "记录": registered[-20:],
        },
        "PDF策略": {
            "运行时生成": True,
            "提交进公开仓": False,
            "说明": ("既有 runtime 契约 committed_artifact_path 恒 null、"
                     "FORBIDDEN_PUBLIC_SUFFIXES 含 .pdf；本 App 只走响应流，不落仓。"),
            "中文渲染": "reportlab 内置 STSong-Light CID 字体，不依赖系统字体文件",
        },
    }


@app.get("/api/报告中心/导出")
def report_export(报告: int, 格式: str = "html"):
    """**已停用**（S07/T-S07-03）——导出改走受控命令。

    停用而不是「留着但不再登记」，是因为登记本身是这条端点存在的理由之一：
    导出 hash 登记册是交付事实的依据。一个不登记的导出端点看起来还能用，
    却在悄悄制造「发出去了但没记录」的报告——比停用危险得多。

    为什么原来那样不行：GET 会被浏览器预取、链接预览、爬虫、代理预热替你按下，
    每一次都往登记册里落一条「导出过」。登记册被噪声灌满之后就不再是依据。
    加上渲染 PDF 不便宜，把它挂在 GET 上等于把昂贵操作放在人人可无限触发的位置。

    **410 而不是静默改行为**：迁移的已知风险是「旧客户端无提示失败」。
    410 是响亮的失败，且响应体里直说替代路径怎么调——
    让调用方当场知道该改什么，而不是过几周才发现自己的导出全丢了。
    """
    raise HTTPException(status_code=410, detail={
        "code": "export_get_retired",
        "message": "GET 导出已停用：它会写业务记录，而 GET 会被预取、爬虫和代理"
                   "替你触发，把导出登记册灌成噪声。改用受控导出任务。",
        "replacement": {
            "创建": "POST /api/导出任务（需 Idempotency-Key 头）",
            "查状态": "GET /api/导出任务/{job_id}",
            "取制品": "GET /api/导出任务/{job_id}/制品",
            "取消": "POST /api/导出任务/{job_id}/取消",
        },
        "示例请求体": {"报告": 1, "格式": "html"},
    })


def _render_export(报告: int, 格式: str) -> tuple[bytes, str, dict[str, Any]]:
    """渲染一份导出并登记。**只由导出任务调用**，不再挂在任何 GET 上。

    水印不接受任何参数控制——只认 delivery 事实。这一条从旧实现原样保留：
    水印是给「这份报告能不能对外」下的结论，能被请求参数左右就等于没有。
    """
    fmt = str(格式).lower().strip()
    if fmt not in EXPORT_FORMATS:
        raise HTTPException(status_code=400, detail=f"格式须为 {list(EXPORT_FORMATS)}")

    d = _report_dir(int(报告))
    title = _report_title(d) or f"一致性证明与差异分析报告 第 {报告} 号"
    header = _delivery_state()
    mark = _watermark_text()
    body = _report_body(d)

    if fmt == "html":
        data, media = _export_html(报告, title, body, header, mark), "text/html; charset=utf-8"
    elif fmt == "csv":
        data, media = _export_csv(报告, d, header, mark), "text/csv; charset=utf-8"
    else:
        data, media = _export_pdf(报告, title, body, header, mark), "application/pdf"

    digest = "sha256:" + hashlib.sha256(data).hexdigest()
    _register_export({
        "报告": int(报告),
        "标题": title,
        "格式": fmt,
        "sha256": digest,
        "字节": len(data),
        "水印已加": mark is not None,
        "水印文案": mark,
        "报告等级": header["报告等级"],
        "质量等级": header["质量等级"],
        "delivery_allowed": header["delivery_allowed"],
        "提交进公开仓": False if fmt == "pdf" else True,
        "导出时间": datetime.now(BEIJING).isoformat(timespec="seconds"),
    })

    _audit("export", subject_ref=f"report_no{报告}:{fmt}", result_status="OK",
           evidence_ref=str(EXPORT_REGISTRY_PATH), sha256=digest, bytes=len(data),
           report_grade=header["报告等级"], quality_grade=header["质量等级"],
           delivery_allowed=header["delivery_allowed"], watermark_applied=mark is not None)

    return data, media, {
        "filename": f"kmfa_report_{报告}.{fmt}",
        "报告等级": header["报告等级"],
        "质量等级": header["质量等级"],
        "delivery_allowed": header["delivery_allowed"],
        "水印已加": mark is not None,
        "sha256": digest,
        "字节": len(data),
    }


# ── S07/T-S07-03 受控导出任务 ──────────────────────────────────────────────────
#
# 命令与查询分开：POST 创建（带幂等键），GET 只读。判定逻辑全在 `export_jobs`，
# 是纯函数；这一层只负责存取与 HTTP 形状。分开是为了让「同键不同请求要 409」
# 这类规则能被单独测，而不必每次都起一个 HTTP 客户端。

from app import export_jobs as _ej  # noqa: E402


def _job_events() -> list[dict[str, Any]]:
    return _st.read(APP_DB_PATH, "export_jobs")


def _load_job(job_id: str) -> dict[str, Any] | None:
    return _ej.fold_job([e for e in _job_events() if e.get("job_id") == job_id])


def _all_jobs() -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for event in _job_events():
        grouped.setdefault(str(event.get("job_id")), []).append(event)
    return [job for job in (_ej.fold_job(v) for v in grouped.values()) if job]


def _now_epoch() -> int:
    return int(datetime.now(BEIJING).timestamp())


def _job_error(error: _ej.ExportJobError):
    raise HTTPException(status_code=error.status_code,
                        detail={"code": error.code, "message": error.message})


@app.post("/api/导出任务")
def create_export_job(request: Request, 载荷: dict[str, Any] = Body(...)):
    """创建导出任务。**POST 而非 GET**——它产生业务记录。

    幂等键必填。没有它，一次网络抖动引发的重试就会变成两条导出登记，
    而登记册是交付事实的依据，多出来的那条没人能事后分辨真假。
    """
    key = request.headers.get("Idempotency-Key")
    owner = "management"  # 本 App 目前单租户；owner 进 id 是为将来多租户留位
    payload = {"报告": int(载荷.get("报告", 0)), "格式": str(载荷.get("格式", "html")).lower()}

    jobs = _all_jobs()
    now = _now_epoch()
    job_id = _ej.job_id_for(owner, key) if key else ""
    existing = next((j for j in jobs if j["job_id"] == job_id), None) if job_id else None
    running = sum(1 for j in jobs if j["state"] in {"queued", "running"})

    try:
        decision = _ej.admit(
            owner=owner, key=key or "", request=payload, existing=existing,
            running_count=running, owner_job_count=len(jobs))
    except _ej.ExportJobError as error:
        _job_error(error)

    if decision["action"] == "reuse":
        # 同键同请求：**一个字节都不重新渲染**。这正是幂等键的意义——
        # 重试不该花第二份 CPU，也不该多出第二条登记。
        return {"复用": True, "任务": _ej.project(decision["job"], now)}

    job_id = decision["job_id"]
    stamp = datetime.now(BEIJING).isoformat(timespec="seconds")
    _st.append(APP_DB_PATH, "export_jobs", {
        "event": "created", "job_id": job_id, "owner": owner,
        "idempotency_key": key, "fingerprint": decision["fingerprint"],
        "request": payload, "at": stamp})
    _st.append(APP_DB_PATH, "export_jobs",
               {"event": "started", "job_id": job_id, "at": stamp})

    try:
        data, media, meta = _render_export(payload["报告"], payload["格式"])
    except HTTPException as error:
        # 失败要**落成事件**，不是抛完就算：失败的任务和不存在的任务
        # 对调用方意味着完全不同的下一步，而只有落了事件才分得开。
        _st.append(APP_DB_PATH, "export_jobs", {
            "event": "failed", "job_id": job_id,
            "failure": str(error.detail)[:500], "at": stamp})
        raise
    if len(data) > _ej.MAX_ARTIFACT_BYTES:
        _st.append(APP_DB_PATH, "export_jobs", {
            "event": "failed", "job_id": job_id,
            "failure": f"制品 {len(data)} 字节超过上限 {_ej.MAX_ARTIFACT_BYTES}",
            "at": stamp})
        raise HTTPException(status_code=413, detail={
            "code": "export_artifact_too_large",
            "message": f"制品超过 {_ej.MAX_ARTIFACT_BYTES} 字节上限。"})

    artifact_path = APP_STATE_DIR / "export-artifacts" / f"{job_id}.bin"
    artifact_path.parent.mkdir(parents=True, exist_ok=True)
    artifact_path.write_bytes(data)
    _st.append(APP_DB_PATH, "export_jobs", {
        "event": "succeeded", "job_id": job_id, "at": stamp,
        "artifact": {**meta, "media_type": media, "path": str(artifact_path),
                     "produced_at_epoch": now}})
    return {"复用": False, "任务": _ej.project(_load_job(job_id), now)}


@app.get("/api/导出任务/{job_id}")
def get_export_job(job_id: str):
    """查状态。**纯读**——不推进任何状态机。

    过期是读的时候算出来的，不是靠定时任务改状态：定时任务没跑的那段时间里，
    改状态的做法会拿着过期制品当有效的发。
    """
    job = _load_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail={
            "code": "export_job_not_found", "message": "没有这个导出任务。"})
    return _ej.project(job, _now_epoch())


@app.get("/api/导出任务/{job_id}/制品")
def get_export_artifact(job_id: str):
    """取制品。**纯读**——制品在任务成功时就已产出，这里只是搬运。"""
    from fastapi.responses import Response

    try:
        plan = _ej.artifact_response_plan(_load_job(job_id), _now_epoch())
    except _ej.ExportJobError as error:
        _job_error(error)
    artifact = plan["artifact"]
    path = Path(str(artifact["path"]))
    if not path.exists():
        raise HTTPException(status_code=410, detail={
            "code": "export_artifact_missing",
            "message": "任务记录为成功，但制品文件已不在。"
                       "这是存储侧的问题，不是任务没跑——重新创建任务。"})
    return Response(content=path.read_bytes(), media_type=str(artifact["media_type"]),
                    headers={
                        "Content-Disposition":
                            f'attachment; filename="{artifact["filename"]}"',
                        "Cache-Control": "private, no-store",
                        "X-Content-Type-Options": "nosniff",
                        "X-KMFA-Report-Grade": str(artifact["报告等级"]),
                        "X-KMFA-Quality-Grade": str(artifact["质量等级"]),
                        "X-KMFA-Delivery-Allowed":
                            str(artifact["delivery_allowed"]).lower(),
                        "X-KMFA-Watermark":
                            "applied" if artifact["水印已加"] else "none",
                        "X-KMFA-Sha256": str(artifact["sha256"]),
                    })


@app.post("/api/导出任务/{job_id}/取消")
def cancel_export_job(job_id: str):
    """取消。终态不可取消，且必须报出当前是什么状态——
    只说「不能取消」而不说现在是什么，调用方只能猜，而猜错的方向通常是「再试一次」。"""
    try:
        plan = _ej.plan_cancel(_load_job(job_id), _now_epoch())
    except _ej.ExportJobError as error:
        _job_error(error)
    _st.append(APP_DB_PATH, "export_jobs", {
        "event": "cancelled", "job_id": plan["job_id"],
        "at": datetime.now(BEIJING).isoformat(timespec="seconds")})
    return _ej.project(_load_job(job_id), _now_epoch())


# ── PROD.0008 影响预览与重跑 ────────────────────────────────────────────────────
# 权威任务包第 17 行：血缘图可视化；选中资产→显示下游影响面；手动触发重跑
# （承接 v014 manual rerun 机制），进度与结果留痕。
# **验收：一次真实重跑从页面发起并完成。**
#
# 承接 KMFA/tools/manual_rerun_mechanism.py 的既有契约：
#   · 重跑链恒为四层 field_mapping → fact_layer → derived_metric → report_reference
#   · overwrite_old_version_allowed=false、old_version_status_after_rerun=retained_not_overwritten
#     —— 重跑**造新版本**，绝不覆盖旧版本
#   · raw/source 层一律不可写；report_grade_upgrade_allowed=false（重跑不许偷偷升等级）
RERUN_STEPS_PATH = KMFA / "metadata" / "lineage" / "manual_rerun_steps.jsonl"
IMPACT_PREVIEWS_PATH = APPROVALS_DIR / "manual_impact_previews.jsonl"
APP_PREVIEWS_PATH = APP_STATE_DIR / "manual_impact_previews.jsonl"
APP_RERUN_STEPS_PATH = APP_STATE_DIR / "manual_rerun_steps.jsonl"
APP_RERUN_CONSISTENCY_PATH = APP_STATE_DIR / "manual_rerun_consistency_checks.jsonl"

RERUN_CHAIN = (
    ("field_mapping", "字段映射"),
    ("fact_layer", "事实层"),
    ("derived_metric", "派生指标"),
    ("report_reference", "报告引用"),
)
# 派生表 → 消费它的 App 视图（下游影响面靠这张表算，不是猜的）
TABLE_TO_VIEWS: dict[str, tuple[str, ...]] = {
    "collection": ("账龄回款",), "receivable_aging": ("账龄回款", "项目成本"),
    "v_collection_authoritative": ("账龄回款",),
    "invoice_raw": ("开票纳税",), "invoice_lines": ("开票纳税", "项目成本"),
    "tax_composition": ("开票纳税",), "loan_register": ("开票纳税",),
    "expense_lines": ("项目成本",), "kingdee_ledger": ("项目成本",),
    "kingdee_voucher": ("项目成本",), "goods_movement": ("项目成本",),
    "bank_journal": ("账龄回款",), "personal_advance": ("项目成本",),
    "op_monthly": ("我在哪",), "op_key_indicators": ("我在哪",),
    "row_matches": ("源检查板",), "subject_code_map": ("源检查板",),
}
VIEW_ENDPOINTS = {
    "账龄回款": "/api/账龄回款", "开票纳税": "/api/开票纳税",
    "项目成本": "/api/项目成本", "我在哪": "/api/我在哪", "源检查板": "/api/源检查",
}


def _lineage_graph() -> dict[str, Any]:
    if not LINEAGE_PATH.exists():
        raise HTTPException(status_code=503, detail="血缘图缺失")
    return yaml.safe_load(LINEAGE_PATH.read_text(encoding="utf-8")) or {}


def _downstream(asset: str) -> dict[str, Any]:
    """选中资产 → 下游影响面：派生表 → App 视图 → 报告引用。全部由血缘边算出。"""
    graph = _lineage_graph()
    edges = [e for e in (graph.get("edges") or []) if str(e.get("from")) == asset]
    tables = sorted({str(e.get("to")).removeprefix("_staging.") for e in edges})
    views: set[str] = set()
    for t in tables:
        views.update(TABLE_TO_VIEWS.get(t, ()))
    # our_source 的真实长相带括号与 via：
    #   "_staging.expense_lines(6403) via _staging.tax_composition"
    #   "_staging.kingdee_ledger（book=<受控法人主体>）"
    # 原来用精确相等匹配，这些**全都漏掉**，影响面少报——真开页面看到"受影响断言域 —"才发现。
    domains = sorted({str(r.get("domain")) for r in _load_assertions()
                      if any(f"_staging.{t}" in str(r.get("our_source", "")) for t in tables)})
    return {
        "资产": asset,
        "派生表": [{"表": t, "行数": sum(int(e.get("rows") or 0) for e in edges
                                       if str(e.get("to")).removeprefix("_staging.") == t),
                    "版本": next((e.get("version") for e in edges
                                  if str(e.get("to")).removeprefix("_staging.") == t), None)}
                   for t in tables],
        "受影响视图": sorted(views),
        "受影响断言域": domains,
        "受影响报告": sorted({f"report_no{r['编号']}" for r in _reports_touching(domains)}),
        "边数": len(edges),
    }


def _reports_touching(domains: list[str]) -> list[dict[str, Any]]:
    """报告与域的对应：读报告正文首标题里的域名，不硬编码映射。"""
    out = []
    domain_zh = {"collection": "回款", "invoicing": "开票", "tax": "税费", "loan": "借款",
                 "expense": "费用", "material": "材料", "advance": "个人借支",
                 "kingdee": "账套", "receivable_aging": "回款", "pipeline": "回款"}
    wanted = {domain_zh.get(d, d) for d in domains}
    for d in _report_dirs():
        title = _report_title(d) or ""
        if any(w and w in title for w in wanted):
            out.append({"编号": int(re.search(r"report_no(\d+)", d.name).group(1)), "标题": title})
    return out


def _view_payload_hash(view: str) -> dict[str, Any]:
    """真算一遍该视图的输出并取内容哈希——重跑要**真跑**，不是写条记录了事。"""
    fn = {
        "账龄回款": receivable_aging, "开票纳税": invoice_tax_fund,
        "项目成本": project_cost, "我在哪": where_am_i, "源检查板": source_check,
    }.get(view)
    if fn is None:
        return {"视图": view, "状态": "no_recompute_binding"}
    payload = fn()
    blob = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return {"视图": view, "端点": VIEW_ENDPOINTS.get(view),
            "内容哈希": "sha256:" + hashlib.sha256(blob.encode("utf-8")).hexdigest(),
            "字节": len(blob.encode("utf-8")), "状态": "recomputed"}


def _append_state(path: Path, record: dict[str, Any]) -> dict[str, Any]:
    table = {APP_RERUN_STEPS_PATH: "rerun_steps",
             APP_RERUN_CONSISTENCY_PATH: "rerun_consistency"}[path]
    return _st.append(APP_DB_PATH, table, record)


@app.get("/api/影响重跑")
def impact_and_rerun(asset: str | None = None):
    """影响预览与重跑页（PROD.0008）——血缘可视化 + 下游影响面 + 重跑留痕。"""
    graph = _lineage_graph()
    edges = graph.get("edges") or []
    nodes = graph.get("nodes") or []
    with_edges = sorted({str(e.get("from")) for e in edges})
    app_steps = _st.read(APP_DB_PATH, "rerun_steps")
    runs: dict[str, list[dict[str, Any]]] = {}
    for s in app_steps:
        runs.setdefault(str(s.get("rerun_run_id")), []).append(s)

    return {
        "血缘": {
            "节点数": len(nodes), "边数": len(edges),
            "可选资产数": len(with_edges),
            "派生表": sorted({str(e.get("to")).removeprefix("_staging.") for e in edges}),
            "资产": [
                {"资产": a,
                 "域": next((n.get("domain") for n in nodes if str(n.get("asset")) == a), None),
                 "派生表数": len({str(e.get("to")) for e in edges if str(e.get("from")) == a})}
                for a in with_edges
            ],
        },
        "选中": _downstream(asset) if asset else None,
        "重跑链": [{"层": k, "名称": zh, "序": i + 1} for i, (k, zh) in enumerate(RERUN_CHAIN)],
        "重跑纪律": {
            "覆盖旧版本": False,
            "旧版本处置": "retained_not_overwritten",
            "raw层可写": False,
            "允许借重跑升报告等级": False,
            "留痕": "每层一条 manual_rerun_step + 一条一致性检查，皆追加式",
        },
        "既有仓内留痕": {
            "重跑步骤": len(_read_jsonl(RERUN_STEPS_PATH)),
            "影响预览": len(_read_jsonl(IMPACT_PREVIEWS_PATH)),
        },
        "本机重跑记录": {
            "轮次": len(runs),
            "步骤数": len(app_steps),
            "位置": str(APP_RERUN_STEPS_PATH),
            "最近": [
                {"轮次号": rid, "步骤": len(steps),
                 "起于": min(str(s.get("rerun_at")) for s in steps),
                 "止于": max(str(s.get("rerun_at")) for s in steps),
                 "状态": "completed" if len(steps) == len(RERUN_CHAIN) else "incomplete",
                 "各层": [{"层": s.get("chain_layer"), "新版本": s.get("new_derived_version_ref"),
                           "旧版本": s.get("old_derived_version_ref"),
                           "旧版本状态": s.get("old_version_status_after_rerun"),
                           "结果": s.get("rerun_result")} for s in
                          sorted(steps, key=lambda x: x.get("chain_order") or 0)]}
                for rid, steps in sorted(runs.items())[-5:]
            ],
        },
    }


@app.post("/api/影响重跑/重跑")
def trigger_rerun(payload: dict[str, Any] = Body(...)):
    """从页面发起一次**真实重跑**：四层链逐层真算，每层留痕，旧版本保留。"""
    asset = str(payload.get("资产") or "").strip()
    reason = str(payload.get("理由") or "").strip()
    actor = str(payload.get("操作人") or "owner").strip() or "owner"
    if not reason:
        raise HTTPException(status_code=400, detail="必须写明重跑理由")

    graph = _lineage_graph()
    if asset not in {str(e.get("from")) for e in (graph.get("edges") or [])}:
        raise HTTPException(status_code=404, detail=f"资产不在血缘图内或无派生边：{asset}")

    down = _downstream(asset)
    started = datetime.now(BEIJING)
    # 轮次号按**已有轮次递增**，不拿挂钟拼：同一秒内对同一资产连点两次（Owner 手快双击）
    # 会撞出同一个 id，两轮留痕被并成一轮——契约测试当场逮到。时间戳留在 rerun_at 字段里。
    seq = len({str(s.get("rerun_run_id")) for s in _st.read(APP_DB_PATH, "rerun_steps")}) + 1
    run_id = f"RERUN-APP-{seq:04d}-{hashlib.sha256(asset.encode()).hexdigest()[:6]}"

    steps: list[dict[str, Any]] = []
    for order, (layer, layer_zh) in enumerate(RERUN_CHAIN, start=1):
        if layer == "field_mapping":
            detail = {"派生表": down["派生表"], "边数": down["边数"]}
        elif layer == "fact_layer":
            pipeline = load_json(FACTS / "data_pipeline.json")
            staging = pipeline.get("staging_tables") or {}
            detail = {"表行数": {t["表"]: (staging.get(t["表"]) or {}).get("rows")
                                 for t in down["派生表"]}}
        elif layer == "derived_metric":
            detail = {"视图": [_view_payload_hash(v) for v in down["受影响视图"]]}
        else:
            detail = {"报告": down["受影响报告"], "断言域": down["受影响断言域"]}

        blob = json.dumps(detail, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        digest = hashlib.sha256(blob.encode("utf-8")).hexdigest()
        step = {
            "record_type": "manual_rerun_step",
            "schema_version": "kmfa.manual_rerun_step.v1",
            "stage_phase": "DT6-PROD0008",
            "rerun_run_id": run_id,
            "rerun_step_id": f"{run_id}-{order:02d}",
            "rerun_version": f"MANUAL-RERUN-KMFA-DT6-PROD0008-001.{order:02d}",
            "chain_layer": layer,
            "chain_layer_label": layer_zh,
            "chain_order": order,
            "source_asset": asset,
            "actor_ref": f"actor_ref://owner_or_authorized_delegate/{actor}",
            "reason_summary": reason,
            "rerun_at": datetime.now(BEIJING).isoformat(timespec="seconds"),
            "rerun_status": "completed_public_safe_metadata_only",
            "rerun_result": detail,
            "content_hash": "sha256:" + digest,
            # 造新版本、保留旧版本——既有契约的硬要求
            "new_derived_version_ref": f"version_ref://KMFA/DT6-PROD0008/{run_id}/{layer}/new-{digest[:12]}",
            "old_derived_version_ref": f"version_ref://KMFA/DT6-PROD0008/{run_id}/{layer}/old-retained",
            "old_version_status_after_rerun": "retained_not_overwritten",
            "overwrite_old_version_allowed": False,
            "append_only_version_record_required": True,
            "raw_layer_write_allowed": False,
            "raw_source_mutation_allowed": False,
            "source_layer_write_allowed": False,
            "business_plaintext_committed": False,
            "forbidden_plaintext": False,
            "formal_report_generated": False,
            "report_grade_upgrade_allowed": False,
            "business_decision_basis_allowed": False,
            "project_id": "KMFA",
            "system_name": "KMFA 经营分析系统",
            "evidence_refs": ["KMFA/machine/lineage.yaml",
                              "KMFA/tools/manual_rerun_mechanism.py"],
        }
        steps.append(_append_state(APP_RERUN_STEPS_PATH, step))

    finished = datetime.now(BEIJING)
    consistency = _append_state(APP_RERUN_CONSISTENCY_PATH, {
        "record_type": "manual_rerun_consistency_check",
        "schema_version": "kmfa.manual_rerun_consistency_check.v1",
        "stage_phase": "DT6-PROD0008",
        "rerun_run_id": run_id,
        "consistency_id": f"CONS-APP-{run_id[-6:]}",
        "checked_at": finished.isoformat(timespec="seconds"),
        "chain_layers_expected": [k for k, _ in RERUN_CHAIN],
        "chain_layers_completed": [s["chain_layer"] for s in steps],
        "chain_complete": [s["chain_layer"] for s in steps] == [k for k, _ in RERUN_CHAIN],
        "report_grade_unchanged": True,
        "old_versions_retained": all(
            s["old_version_status_after_rerun"] == "retained_not_overwritten" for s in steps),
        "raw_layer_untouched": True,
    })

    _audit("processing", subject_ref=asset, result_status="COMPLETED",
           evidence_ref=str(APP_RERUN_STEPS_PATH), run_id=run_id,
           layers=len(steps), chain_complete=consistency["chain_complete"])

    return {
        "轮次号": run_id,
        "资产": asset,
        "耗时秒": round((finished - started).total_seconds(), 3),
        "步骤数": len(steps),
        "链完整": consistency["chain_complete"],
        "旧版本全保留": consistency["old_versions_retained"],
        "各层": [{"序": s["chain_order"], "层": s["chain_layer"], "名称": s["chain_layer_label"],
                  "新版本": s["new_derived_version_ref"], "哈希": s["content_hash"],
                  "结果": s["rerun_result"]} for s in steps],
        "一致性检查": consistency,
        "留痕位置": str(APP_RERUN_STEPS_PATH),
    }


# ── PROD.0003 访问安全承接 S17：审计日志 append-only ──────────────────────────
# 权威任务包第 12 行：本机单用户模式；导出水印（等级/Q 级/delivery 永远印在页眉）；
# 审计日志 append-only。验收＝安全走查单过（承接 v014 S17 access security 口径）。
#
# 契约取自既有 KMFA/metadata/security/：
#   audit_log_policy.jsonl —— 7 个必填字段、append_only、五种 action_type、
#     raw_payload_allowed=false / business_value_plaintext_allowed=false
#   v014_s17_p1_..._audit_event_contract.jsonl —— persistent_event_write_enabled=false
#     （S17-P1 只定契约不落盘）；**本单元即是把它真正落盘的那一步**
SECURITY_DIR = KMFA / "metadata" / "security"
AUDIT_POLICY_PATH = SECURITY_DIR / "audit_log_policy.jsonl"
ACCESS_POLICY_PATH = SECURITY_DIR / "access_security_policy_manifest.json"
APP_AUDIT_PATH = APP_STATE_DIR / "audit_events.jsonl"

AUDIT_REQUIRED_FIELDS = (
    "event_id", "event_time", "actor_role", "action_type",
    "subject_ref", "evidence_ref", "result_status",
)
AUDIT_ACTION_TYPES = ("import", "processing", "report", "export", "notification")
# 审计事件里绝不允许出现的东西——记「谁在什么时候对什么做了什么」，
# 不记业务明文与原始载荷（契约：raw_payload_allowed=false）
AUDIT_FORBIDDEN_KEYS = FORBIDDEN_EVENT_KEYS | frozenset({
    "payload", "raw_payload", "body", "report_body", "text", "content",
})


def _audit(action_type: str, subject_ref: str, result_status: str,
           evidence_ref: str, actor_role: str = "management",
           **extra: Any) -> dict[str, Any]:
    """写一条审计事件。**只追加，且写失败不能拖垮业务动作**。

    审计是旁证不是主流程：日志写不进去时业务不该跟着挂，但也不能悄悄吞掉——
    失败会以 audit_write_failed 记进返回值，调用方可见。
    """
    if action_type not in AUDIT_ACTION_TYPES:
        raise HTTPException(status_code=500, detail=f"非法 action_type：{action_type}")
    event = {
        "event_id": f"AUD-APP-{hashlib.sha256(f'{action_type}{subject_ref}{datetime.now(BEIJING).isoformat()}'.encode()).hexdigest()[:16]}",
        "event_time": datetime.now(BEIJING).isoformat(timespec="seconds"),
        "actor_role": actor_role,
        "action_type": action_type,
        "subject_ref": subject_ref,
        "evidence_ref": evidence_ref,
        "result_status": result_status,
        "record_type": "audit_event",
        "policy_version": "AUD-KMFA-S17P1-ACTION-LOG-001",
        "stage_phase": "DT6-PROD0003",
        "append_only": True,
        "raw_payload_committed": False,
        "business_value_plaintext_committed": False,
        **extra,
    }
    leaked = sorted(set(event) & AUDIT_FORBIDDEN_KEYS)
    if leaked:
        raise HTTPException(status_code=500, detail=f"审计事件含禁写字段：{leaked}")
    try:
        _st.append(APP_DB_PATH, "audit_events", event)
    except Exception as exc:  # 审计是旁证：写不进去也不该拖垮业务动作
        event["audit_write_failed"] = str(exc)
    return event


@app.get("/api/审计日志")
def audit_log(action_type: str | None = None, page: int = 1, size: int = 50):
    """审计日志（PROD.0003）——append-only，只记动作不记业务明文。"""
    rows = _st.read(APP_DB_PATH, "audit_events")
    selected = [r for r in rows if not action_type or r.get("action_type") == action_type]
    items, meta = _paginate(list(reversed(selected)), page, size)
    policy = _read_jsonl(AUDIT_POLICY_PATH)
    access = load_json(ACCESS_POLICY_PATH) if ACCESS_POLICY_PATH.exists() else {}
    by_type: dict[str, int] = {}
    for r in rows:
        key = str(r.get("action_type"))
        by_type[key] = by_type.get(key, 0) + 1
    return {
        "总数": len(rows),
        "按动作": by_type,
        "分页": meta,
        "事件": items,
        "契约": {
            "必填字段": list(AUDIT_REQUIRED_FIELDS),
            "动作类型": list(AUDIT_ACTION_TYPES),
            "append_only": True,
            "允许记原始载荷": False,
            "允许记业务明文": False,
            "政策版本": (policy[0].get("policy_version") if policy else None),
            "落盘位置": str(APP_AUDIT_PATH),
            "契约来源": [
                "KMFA/metadata/security/audit_log_policy.jsonl",
                "KMFA/metadata/security/access_security_policy_manifest.json",
            ],
        },
        "访问模式": {
            "模式": "本机单用户",
            "应用内登录": False,
            "生产鉴权": "Cloudflare Access 仅保护 /api* 与 /ops*；源站校验签名 JWT",
            "角色口径": (access.get("required_roles") or []),
            "说明": "根域名公开；本机不做多用户与角色分权，私有运维面继续由 Access 控制",
        },
    }


# ── 排程健康：让「排程到底跑没跑」这件事在页面上一眼可见 ──────────────────────
# 起因：2026-07-20 Owner 说「不可能每次结果都让我给你们反馈啊，你自己不会复审检查吗」。
# 他是对的。在此之前 app 容器一个卷都不挂，排程状态只能靠人登服务器 `cat` 日志——
# 于是每次都要 Owner 亲自查、再回报给开发侧。这个来回本身就是设计缺陷，不是沟通问题。
# 现在 app 只读挂 kmfa-logs，本接口直接读 skills 写的 ledger。
SKILL_LEDGER_PATH = Path(os.environ.get("KMFA_SKILL_LEDGER", "/var/log/kmfa/ledger.jsonl"))
LEDGER_UPLINK_STATUS_PATH = Path(os.environ.get(
    "KMFA_LEDGER_UPLINK_STATUS", "/var/log/kmfa/ledger_uplink_status.json"))

#: 考勤投递回执的归档根。技能按 `<root>/<YYYYMM>/*.dispatch.json` 写。
ATTENDANCE_ARCHIVE_ROOT = Path(os.environ.get(
    "KMFA_ATTENDANCE_ARCHIVE_ROOT", "/var/log/kmfa/dingtalk_attendance"))

#: 回执里允许出现在公开面的字段。**白名单，不是黑名单**——回执里还有
#: `management_report`／`hr_report`／`notification_template_text`，那是全员考勤正文，
#: 一个都不能出去。用黑名单的话，将来回执加字段就会默认泄露。
_DISPATCH_PUBLIC_FIELDS = (
    "notification_status", "channel", "run_type", "work_date", "failure_reason")

#: 公开失败码的形状校验。台账文件是**另一个容器**写的，这里不拿它当可信输入——
#: 那边写坏了、被塞进业务文本，也绝不能顺着流到公开端点上。
_PUBLIC_CODE_UPPER_SNAKE = re.compile(r"^[A-Z][A-Z0-9]*(?:_[A-Z0-9]+)*$")
_PUBLIC_CODE_CAMEL = re.compile(r"^(?:[A-Z]{1,6}[a-z0-9]+)+$")
_PUBLIC_CODE_CREDENTIAL = re.compile(r"^(?:gh[pousr]_|sk-|xox[bap]-)|^[A-Fa-f0-9]{24,}$")


def _ledger_uplink_state() -> dict[str, Any]:
    """台账回传通没通——**回传按设计静默失败**，没有这个就没人会发现它断了。

    实测：私有库里压根没有 skill-ledger 目录，而回传每次都"成功"返回 0（不该拖垮技能），
    于是断了几十次也无人察觉。这里把它摆到台面上。
    """
    if not LEDGER_UPLINK_STATUS_PATH.exists():
        return {"成功": False, "情况": "回传从未留下过结果——技能容器里还没跑到回传这一步，或日志卷没挂上"}
    try:
        raw = json.loads(LEDGER_UPLINK_STATUS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return {"成功": False, "情况": f"回传留痕无法解析：{type(exc).__name__}"}
    if not isinstance(raw, dict):
        return {"成功": False, "情况": "回传留痕格式不对"}
    # 只透传三个已知字段，不整包回显——留痕文件是别的容器写的，不当可信输入。
    return {
        "成功": bool(raw.get("成功")),
        "情况": str(raw.get("情况") or "")[:120],
        "时间": str(raw.get("时间") or "")[:40],
    }


def _public_failure_code(raw: object) -> str | None:
    """把台账里的 code 变成可公开的失败码；任何不合形状的一律丢弃（返回 None）。

    与 KMFA/tools/skill_failure_code.py 的 is_public_safe 同规则，**故意重写一遍**：
    这两侧分属两个容器、两条部署链，一侧被改坏时另一侧仍然拦得住。
    """
    if not isinstance(raw, str):
        return None
    code = raw.strip()
    if not code or not (3 <= len(code) <= 60) or _PUBLIC_CODE_CREDENTIAL.search(code):
        return None
    if _PUBLIC_CODE_CAMEL.match(code):
        return code
    if _PUBLIC_CODE_UPPER_SNAKE.match(code) and ("_" in code or len(code) <= 12):
        return code
    return None
# 排程契约（与 deploy/skills-runtime/crontab.txt 一致；北京时间）
SCHEDULE_CONTRACT = {
    "attendance-morning": "每天 08:01",
    "attendance-evening": "每天 17:31",
    "work-check-morning": "每天 11:35",
    "work-check-evening": "每天 17:05",
    "fund-weekly": "周一/周六 11:00",
    "mgmt-monthly": "每月 1 日 09:00",
    "upstream-archive": "每天 11:00",
    "self-audit": "周日 01:00",
    "daily-backup": "每天 00:30",
    "dws-keepalive": "每 4 小时 :20",
    # 排程表里有它、台账与健康端点里却没有——于是「群清单自举到底跑没跑」长期无人可见，
    # 而上游归档正是卡在它的产出上。漏登记本身就是一种假绿：看不见的排程等于没有排程。
    # 无固定钟点：缺目标文件时由 entrypoint 冷启动自举。它不该有排程——
    # 解析一次就够，反复探测等于反复真发消息。
    "attendance-bootstrap-targets": "缺目标文件时自举",
    "dws-bootstrap-groups": "周日 10:30",
    "project-cost-refresh": "每天 05:45",
    # 2026-07-29：**这条漏登记害我查错了方向。** 健康端点是
    # `for skill in sorted(SCHEDULE_CONTRACT)`——只输出契约里有的技能。
    # dws-data-auth 不在契约里，于是它**哪怕一直在跑，端点也永远不会显示它**。
    # 我据此连着两轮判定「技能没跑」，还为此改了两版触发方式。
    # 判据用了一个结构上就不可能显示该技能的面——看不见 ≠ 没发生。
    "dws-data-auth": "每 15 分钟（闸在技能内：未卡住不请求、请求后静默 6 小时）",
    # 独立容器的状态只经受控 projection 卷进入本页；不借现有 skills 的台账、DWS
    # profile、登录态或任何本机排程。没有状态文件不是健康，而是「未跑通」。
    "daily-funds": "历史轮询每 15 分钟；授权每分钟；保活每小时；R2 零费用守卫每 6 小时；回填/冷备/观察每日；非生产恢复演练每月",
}
# 技能归属业务模块（Owner 2026-07-21：「所有 skills 都需要整合进 kmfa 功能模块」）
SKILL_MODULE = {
    "attendance-morning": "考勤与日检", "attendance-evening": "考勤与日检",
    "work-check-morning": "考勤与日检", "work-check-evening": "考勤与日检",
    "fund-weekly": "资金与经营报告", "mgmt-monthly": "资金与经营报告",
    "upstream-archive": "数据接入",
    "self-audit": "系统底座", "daily-backup": "系统底座", "dws-data-auth": "上游归档", "dws-keepalive": "系统底座",
    "attendance-bootstrap-targets": "钉钉考勤",
    "dws-bootstrap-groups": "系统底座",
    "project-cost-refresh": "成本与利润",
    "daily-funds": "每日资金",
}


# ── 每日资金：只读已验证 projection，绝不读群消息、原始附件或私有 Git ──────────────
# 这个卷由独立 daily-funds 容器写，app 只读（control 卷除外），因此前端无法绕过
# D1/Git/R2 的发布门直接接触原始数据。
DAILY_FUNDS_PUBLICATION_DIR = Path(os.environ.get(
    "DAILY_FUNDS_PUBLICATION_DIR", "/var/lib/kmfa/daily-funds"))
DAILY_FUNDS_CONTROL_DIR = Path(os.environ.get(
    "DAILY_FUNDS_CONTROL_DIR", "/var/lib/kmfa/daily-funds-control"))
DAILY_FUNDS_APP_BUILD_SOURCE_COMMIT_FILE = Path(__file__).with_name(".kmfa-source-commit")
DAILY_FUNDS_AUTH_REQUEST_SCHEMA = "kmfa.daily_funds.dws_auth_request.v1"
DAILY_FUNDS_AUTH_SESSION_SCHEMA = "kmfa.daily_funds.dws_auth_session.v1"
DAILY_FUNDS_AUTH_REQUEST_FILE = "dws_auth_request.json"
DAILY_FUNDS_AUTH_SESSION_FILE = "dws_auth_session.json"
DAILY_FUNDS_AUTH_ACTOR = "kmfa_private_owner_ui"
DAILY_FUNDS_AUTH_LIVE_STATES = {"REQUESTED", "AWAITING_APPROVAL", "CANCELLING"}
DAILY_FUNDS_AUTH_TERMINAL_STATES = {"SUCCEEDED", "FAILED", "EXPIRED", "CANCELLED"}
DAILY_FUNDS_AUTH_CODE_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9-]{2,63}$")
DAILY_FUNDS_HISTORY_PROBE_REQUEST_SCHEMA = "kmfa.daily_funds.dws_history_probe_request.v1"
DAILY_FUNDS_HISTORY_PROBE_SESSION_SCHEMA = "kmfa.daily_funds.dws_history_probe_session.v2"
DAILY_FUNDS_HISTORY_PROBE_LEGACY_SESSION_SCHEMA = "kmfa.daily_funds.dws_history_probe_session.v1"
DAILY_FUNDS_HISTORY_PROBE_REQUEST_FILE = "dws_history_probe_request.json"
DAILY_FUNDS_HISTORY_PROBE_SESSION_FILE = "dws_history_probe_session.json"
DAILY_FUNDS_HISTORY_PROBE_ACTOR = "kmfa_private_owner_ui"
DAILY_FUNDS_HISTORY_PROBE_ENDPOINT_HEADER = "X-KMFA-Daily-Funds-Probe"
DAILY_FUNDS_HISTORY_PROBE_ENDPOINT_VALUE = "v1"
DAILY_FUNDS_HISTORY_PROBE_LIVE_STATES = {"REQUESTED", "RUNNING"}
DAILY_FUNDS_HISTORY_PROBE_TERMINAL_STATES = {"COMPLETED", "FAILED", "EXPIRED"}
DAILY_FUNDS_HISTORY_PROBE_CONTINUATION_STATES = {
    "NOT_STARTED", "FIRST_PAGE_TERMINAL", "SECOND_PAGE_TERMINAL", "SECOND_PAGE_CONTINUES",
    "GROUP_HISTORY_FALLBACK_FIRST_PAGE_TERMINAL",
    "GROUP_HISTORY_FALLBACK_SECOND_PAGE_TERMINAL",
    "GROUP_HISTORY_FALLBACK_SECOND_PAGE_CONTINUES",
    "GROUP_HISTORY_V2_FIRST_PAGE_TERMINAL",
    "GROUP_HISTORY_V2_SECOND_PAGE_TERMINAL",
    "GROUP_HISTORY_V2_SECOND_PAGE_CONTINUES",
}
DAILY_FUNDS_HISTORY_PROBE_CURSOR_TRANSCRIPTS = {
    "NOT_STARTED": "NOT_STARTED",
    "FIRST_PAGE_TERMINAL": "FIRST_PAGE_TERMINAL",
    "SECOND_PAGE_TERMINAL": "OPAQUE_CURSOR_REUSED_SECOND_PAGE_TERMINAL",
    "SECOND_PAGE_CONTINUES": "OPAQUE_CURSOR_REUSED_SECOND_PAGE_CONTINUES",
    "GROUP_HISTORY_FALLBACK_FIRST_PAGE_TERMINAL": "GROUP_HISTORY_FALLBACK_FIRST_PAGE_TERMINAL",
    "GROUP_HISTORY_FALLBACK_SECOND_PAGE_TERMINAL": "GROUP_HISTORY_FALLBACK_OPAQUE_CURSOR_REUSED_SECOND_PAGE_TERMINAL",
    "GROUP_HISTORY_FALLBACK_SECOND_PAGE_CONTINUES": "GROUP_HISTORY_FALLBACK_OPAQUE_CURSOR_REUSED_SECOND_PAGE_CONTINUES",
    "GROUP_HISTORY_V2_FIRST_PAGE_TERMINAL": "GROUP_HISTORY_V2_FIRST_PAGE_TERMINAL",
    "GROUP_HISTORY_V2_SECOND_PAGE_TERMINAL": "GROUP_HISTORY_V2_PROVIDER_MILLISECOND_CURSOR_REUSED_SECOND_PAGE_TERMINAL",
    "GROUP_HISTORY_V2_SECOND_PAGE_CONTINUES": "GROUP_HISTORY_V2_PROVIDER_MILLISECOND_CURSOR_REUSED_SECOND_PAGE_CONTINUES",
}
DAILY_FUNDS_HISTORY_PROBE_RECORD_LIST_SHAPES = {
    "NOT_OBSERVED", "NO_DIRECT_LIST", "UNRECOGNIZED_DIRECT_LIST",
}
DAILY_FUNDS_ALLOWED_RANGES = {"1d": 1, "7d": 7, "30d": 30, "90d": 90, "180d": 180, "360d": 360}
DAILY_FUNDS_HUMAN_STATUSES = {"已更新", "处理中", "需处理"}
DAILY_FUNDS_HARD_THRESHOLD_FEN = 60_000_000
DAILY_FUNDS_SOFT_THRESHOLD_FEN = 120_000_000
DAILY_FUNDS_FIXED_RISKS = {"正常", "关注", "高风险"}
DAILY_FUNDS_DYNAMIC_FLAGS = {"动态偏低", "动态明显偏低"}
DAILY_FUNDS_FLOATING_LINE_ORDER = (
    "three_month",
    "six_month",
    "custom_date_range",
    "custom_numeric",
)
DAILY_FUNDS_FLOATING_LINE_NAMES = set(DAILY_FUNDS_FLOATING_LINE_ORDER)
# This is an operational enum, not a failure code.  In particular, ``OK`` is
# deliberately two characters and must not be lost through the generic
# public failure-code sanitizer.
DAILY_FUNDS_BACKUP_STATES = {"OK", "LAG", "PENDING", "UNKNOWN"}
DAILY_FUNDS_STATUS_SCHEMA = "kmfa.daily_funds.status.v1"
DAILY_FUNDS_CASHFLOW_OBSERVATION_SCHEMA = "kmfa.daily_funds.cashflow_observation.v2"
DAILY_FUNDS_CASHFLOW_OBSERVATION_STATUSES = {"VERIFIED", "NEEDS_REVIEW", "NOT_AVAILABLE"}
DAILY_FUNDS_CASHFLOW_REJECTION_CATEGORY_LABELS = {
    "HEADER_LAYOUT": "表头布局",
    "OCR_CONFIDENCE": "文字清晰度",
    "FOOTER_RECONCILIATION": "合计勾稽",
    "DATE_FIELD": "日期字段",
    "ROW_AMOUNT": "收支行金额",
    "OCR_FORMAT": "表格识别",
    "OTHER_REVIEW": "其他确定性复核",
}
DAILY_FUNDS_CASHFLOW_OBSERVATION_FIELDS = frozenset({
    "schema_version", "generated_at", "parser_version", "source_coverage",
    "rejection_categories", "evidence_version", "points", "status", "machine_code",
})
# This is a read-side schema allowlist, not a second scheduler or health
# authority.  The daily-funds worker remains the sole writer; the app only
# displays the fixed worker contract after checking its exact shape so a
# shared-volume extension cannot disclose an identifier or masquerade as an
# alternative schedule.
DAILY_FUNDS_STATUS_SCHEDULES = {
    "history_poll": "*/15 * * * * Asia/Shanghai",
    "auth_probe": "* * * * * Asia/Shanghai",
    "keepalive": "0 * * * * Asia/Shanghai",
    "backfill": "5,20,35,50 * * * * Asia/Shanghai",
    "observer": "30 3 * * * Asia/Shanghai",
    "r2_guard": "0 */6 * * * Asia/Shanghai",
    "cold_backup": "10 4 * * * Asia/Shanghai",
    "raw_archive_audit": "20 5 * * * Asia/Shanghai",
    "runtime_audit": "45 5 * * * Asia/Shanghai",
    "restore_drill": "0 5 1 * * Asia/Shanghai",
}
DAILY_FUNDS_STATUS_FIELDS = frozenset({
    "schema_version", "human_status", "machine_code", "effective_business_date",
    "last_verified_at", "publication_id", "updated_at", "schedules", "backup_state",
})
DAILY_FUNDS_CAPABILITY_FAMILIES = {"资金账户明细表", "资金流水明细", "资金明细", "UNCLASSIFIED"}
DAILY_FUNDS_CAPABILITY_SUFFIXES = {
    ".csv", ".txt", ".xlsx", ".xlsm", ".xls", ".pdf", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp",
    "UNKNOWN_SUFFIX",
}
DAILY_FUNDS_CAPABILITY_MAGICS = {"TEXT", "ZIP", "OLE", "PDF", "PNG", "JPEG", "GIF", "BMP", "WEBP", "BINARY", "EMPTY"}
DAILY_FUNDS_CAPABILITY_OUTCOMES = {"SUPPORTED", "NEEDS_REVIEW"}
# The worker journal needs precise machine codes to support a protected
# incident investigation.  The browser does not: raw codes make it too easy
# to accidentally turn a future parser implementation detail into a public
# source-data channel.  These fixed labels retain the one decision-relevant
# distinction (what kind of work is blocked) without forwarding a code,
# filename, document field, source ID, or amount.
DAILY_FUNDS_CAPABILITY_DIAGNOSTIC_LABELS = {
    "OCR_GENERIC_HEADER_SCHEMA_MISSING": "图片表头未形成余额或流水完整结构",
    "OCR_GENERIC_ROW_SCHEMA_MISSING": "图片数据行未形成余额或流水完整结构",
    "OCR_GENERIC_CONFIDENCE_BLOCKED": "图片关键字段置信度不足",
    "OCR_GENERIC_FAMILY_UNRESOLVED": "图片无法确定为余额或流水",
    "OCR_GENERIC_FAMILY_AMBIGUOUS": "图片同时匹配余额与流水",
    "OCR_PROFILE_CALIBRATING": "图片版式校准中",
    "OCR_LOW_CONFIDENCE": "图片关键字段置信度不足",
    "OCR_CONFIDENCE_INVALID": "图片关键字段置信度不足",
    "OCR_CONFIDENCE_THRESHOLD_INVALID": "图片关键字段置信度不足",
    "OCR_RUNTIME_UNAVAILABLE": "云端图片解析运行环境不可用",
    "UNSUPPORTED_ATTACHMENT": "文件格式或表格结构未通过确定性校验",
    "DOCUMENT_FAMILY_UNSUPPORTED": "文件格式或表格结构未通过确定性校验",
    "CORRUPT_ATTACHMENT": "文件格式或表格结构未通过确定性校验",
    "FORMAT_MAGIC_MISMATCH": "文件格式或表格结构未通过确定性校验",
    "MIME_DECLARATION_INVALID": "文件格式或表格结构未通过确定性校验",
    "MIME_SUFFIX_MISMATCH": "文件格式或表格结构未通过确定性校验",
}
DAILY_FUNDS_CAPABILITY_DIAGNOSTIC_ORDER = {
    "图片表头未形成余额或流水完整结构": 1,
    "图片数据行未形成余额或流水完整结构": 2,
    "图片无法确定为余额或流水": 3,
    "图片同时匹配余额与流水": 4,
    "图片版式校准中": 5,
    "图片关键字段置信度不足": 6,
    "云端图片解析运行环境不可用": 7,
    "图片表格结构未通过确定性校验": 8,
    "文件格式或表格结构未通过确定性校验": 9,
    "来源或字节校验未通过": 10,
    "表格字段或业务规则未通过确定性校验": 11,
    "其他确定性解析门未通过": 12,
}
DAILY_FUNDS_SOURCE_DISCOVERY_STATES = {
    "UNKNOWN",
    "HISTORY_EMPTY",
    "TARGET_DOCUMENT_NOT_FOUND",
    "TARGET_ATTACHMENT_MISSING",
    "ATTACHMENT_ACQUIRED",
    "DOCUMENT_PAIR_MISSING",
    "ACCOUNT_SNAPSHOT_MISSING",
    "TRANSACTION_FACT_MISSING",
    "SOURCE_FACT_DATE_MISMATCH",
    "COMPLETE_PAIR_READY",
}
DAILY_FUNDS_BUSINESS_FLOW_STAGES = {
    "RUNTIME_AUDITED", "RUNTIME_NEEDS_ATTENTION", "WAITING_FOR_VALID_PUBLICATION",
    "PARSER_NEEDS_REVIEW", "POLL_NEEDS_ATTENTION", "POLL_PUBLISHED",
    # Historical raw-first discovery may prove a source attachment before a
    # deterministic parser supports it.  These are explicit non-publication
    # states, not aliases for a successful reconciliation.
    "BACKFILL_EMPTY_WINDOW", "BACKFILL_ARCHIVED", "BACKFILL_ARCHIVED_NEEDS_REVIEW",
    "BACKFILLING", "BACKFILLING_NEEDS_REVIEW",
    "BACKFILL_COMPLETE", "BACKFILL_COMPLETE_NEEDS_REVIEW",
    "RAW_ARCHIVE_AUDITED", "RAW_ARCHIVE_AUDIT_NEEDS_REVIEW",
    "OBSERVER_NEEDS_ATTENTION", "OBSERVER_WAITING_FOR_PUBLICATION_LOCK",
    "OBSERVER_BASELINE_CAPTURED", "OBSERVER_WAITING_FOR_NEXT_BUSINESS_DATE",
    "POST_DEPLOY_OBSERVING", "POST_DEPLOY_OBSERVATION_COMPLETE",
    "RESTORE_DRILL", "RESTORE_DRILL_NEEDS_ATTENTION", "UNKNOWN",
}
DAILY_FUNDS_OPERATION_RECEIPTS = (
    ("poll", "历史轮询"),
    ("auth-probe", "认证探测"),
    ("keepalive", "认证保活"),
    ("backfill", "历史回填"),
    ("r2-guard", "R2 零费用守卫"),
    ("raw-archive-audit", "归档原件复核"),
    ("cold-backup", "OCI 冷备"),
    ("observer", "上线后观察"),
    ("restore-drill", "恢复演练"),
)


def _read_daily_funds_json(name: str) -> dict[str, Any] | None:
    """Read a bounded, local projection file only; no raw source path is valid."""
    target = (DAILY_FUNDS_PUBLICATION_DIR / name).resolve()
    root = DAILY_FUNDS_PUBLICATION_DIR.resolve()
    if not str(target).startswith(str(root) + "/") or not target.is_file():
        return None
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return payload if isinstance(payload, dict) else None


def _daily_funds_status() -> dict[str, Any]:
    def unavailable(code: str) -> dict[str, Any]:
        return {
            "human_status": "需处理",
            "machine_code": code,
            "effective_business_date": None,
            "last_verified_at": None,
            "publication_id": None,
            "updated_at": None,
            "schedules": {},
            "backup_state": "UNKNOWN",
        }

    payload = _read_daily_funds_json("status.json")
    if payload is None:
        return unavailable("UNKNOWN")
    # ``status.json`` is the worker's one hand-off into the existing KMFA
    # status centre.  It is a versioned values-free record, not an opaque
    # dict: fail closed on either schema drift or a schedule extension rather
    # than reflecting arbitrary shared-volume strings in the owner UI.
    if (
        set(payload) != DAILY_FUNDS_STATUS_FIELDS
        or payload.get("schema_version") != DAILY_FUNDS_STATUS_SCHEMA
        or payload.get("schedules") != DAILY_FUNDS_STATUS_SCHEDULES
    ):
        return unavailable("STATUS_INVALID")

    status = payload.get("human_status")
    machine_code = _public_failure_code(payload.get("machine_code"))
    effective_business_date = payload.get("effective_business_date")
    last_verified_at = payload.get("last_verified_at")
    publication_id = payload.get("publication_id")
    updated_at = payload.get("updated_at")
    backup_state = payload.get("backup_state")
    if (
        not isinstance(status, str)
        or status not in DAILY_FUNDS_HUMAN_STATUSES
        or machine_code is None
        or (effective_business_date is not None and _daily_funds_date(effective_business_date) is None)
        or (last_verified_at is not None and _daily_funds_timestamp(last_verified_at) is None)
        or (publication_id is not None and not _daily_funds_lower_hex(publication_id, 64))
        or _daily_funds_timestamp(updated_at) is None
        or not isinstance(backup_state, str)
        or backup_state not in DAILY_FUNDS_BACKUP_STATES
    ):
        return unavailable("STATUS_INVALID")
    return {
        "human_status": status,
        "machine_code": machine_code,
        "effective_business_date": effective_business_date,
        "last_verified_at": last_verified_at,
        "publication_id": publication_id,
        "updated_at": updated_at,
        "schedules": dict(DAILY_FUNDS_STATUS_SCHEDULES),
        "backup_state": backup_state,
    }


def _daily_funds_flow_token(value: object, *, allowed: set[str], default: str) -> str:
    if not isinstance(value, str):
        return default
    token = value.strip().upper()
    return token if token in allowed else default


def _daily_funds_app_build_source_fingerprint() -> str | None:
    """Read a strictly validated build-layer source fact without exposing it."""

    try:
        raw = DAILY_FUNDS_APP_BUILD_SOURCE_COMMIT_FILE.read_text(encoding="ascii")
    except (OSError, UnicodeError):
        return None
    if len(raw) != 41 or not raw.endswith("\n"):
        return None
    commit = raw[:-1]
    if not _daily_funds_lower_hex(commit, 40):
        return None
    return hashlib.sha256(commit.encode("ascii")).hexdigest()


def _daily_funds_embedded_source_identity(deployment: object) -> str:
    """Compare two image-layer source fingerprints without leaking either.

    Matching fingerprints establish only that the app and worker were built
    from the same validated source commit.  A deployment record/image digest
    is intentionally outside this local comparison and remains unknown.
    """

    row = deployment if isinstance(deployment, dict) else {}
    worker_state = _daily_funds_flow_token(
        row.get("identity_state"),
        allowed={"BUILD_SOURCE_COMMIT_EMBEDDED"},
        default="UNKNOWN",
    )
    worker_fingerprint = row.get("source_commit_fingerprint")
    app_fingerprint = _daily_funds_app_build_source_fingerprint()
    if (
        worker_state != "BUILD_SOURCE_COMMIT_EMBEDDED"
        or not _daily_funds_lower_hex(worker_fingerprint, 64)
        or app_fingerprint is None
    ):
        return "UNKNOWN"
    return (
        "SOURCE_COMMIT_MATCHED_IMAGE_DIGEST_UNKNOWN"
        if worker_fingerprint == app_fingerprint
        else "SOURCE_COMMIT_FINGERPRINT_MISMATCH"
    )


def _daily_funds_source_discovery(value: object) -> dict[str, str]:
    """Expose only the poll gate reached, never source content or identity."""

    row = value if isinstance(value, dict) else {}
    state = _daily_funds_flow_token(
        row.get("state"),
        allowed=DAILY_FUNDS_SOURCE_DISCOVERY_STATES,
        default="UNKNOWN",
    )
    labels = {
        "UNKNOWN": "未验证",
        "HISTORY_EMPTY": "历史窗口无消息",
        "TARGET_DOCUMENT_NOT_FOUND": "历史已读取，目标文件未命中",
        "TARGET_ATTACHMENT_MISSING": "目标文件缺少附件",
        "ATTACHMENT_ACQUIRED": "附件已取得，等待确定性解析与勾稽",
        "DOCUMENT_PAIR_MISSING": "附件已取得，账户/流水尚未成对",
        "ACCOUNT_SNAPSHOT_MISSING": "附件已取得，缺少账户余额事实",
        "TRANSACTION_FACT_MISSING": "附件已取得，缺少资金流水事实",
        "SOURCE_FACT_DATE_MISMATCH": "附件已取得，但账户与流水业务日期未成对",
        "COMPLETE_PAIR_READY": "账户与流水已成对，等待后续勾稽与发布",
    }
    return {"状态": state, "说明": labels[state]}


def _daily_funds_attachment_capability_summary(rows: object) -> dict[str, Any]:
    """Reduce worker receipts to a browser-safe, fail-closed capability state.

    The worker's SQLite matrix intentionally retains format detail so that an
    operator can diagnose a parser gate on the protected worker volume.  The
    KMFA app needs only the small answer required for the product status:
    whether real private-Git readback samples were supported or need review.
    It therefore does not forward MIME, filename, hash, parser version,
    document text, attachment bytes, or parser failure details.
    """

    unobserved = {
        "状态": "未观测",
        "已支持附件数": 0,
        "待复核附件数": 0,
        "待复核原因": [],
        "最近观测": None,
    }
    unknown = {
        "状态": "UNKNOWN",
        "已支持附件数": 0,
        "待复核附件数": 0,
        "待复核原因": [],
        "最近观测": None,
    }
    if rows is None:
        return unobserved
    if not isinstance(rows, list) or len(rows) > 64:
        return unknown

    supported = 0
    needs_review = 0
    review_reasons: dict[str, int] = {}
    latest: tuple[datetime, str] | None = None
    for row in rows:
        if not isinstance(row, dict):
            return unknown
        family = row.get("family")
        suffix = row.get("suffix")
        magic = row.get("magic")
        outcome = row.get("outcome")
        code = row.get("code")
        count = row.get("count")
        observed_at = _daily_funds_timestamp(row.get("last_observed_at"))
        declared_mime = row.get("declared_mime")
        parser_version = row.get("parser_version")
        if (
            not isinstance(family, str)
            or not isinstance(suffix, str)
            or not isinstance(magic, str)
            or not isinstance(outcome, str)
            or family not in DAILY_FUNDS_CAPABILITY_FAMILIES
            or suffix not in DAILY_FUNDS_CAPABILITY_SUFFIXES
            or magic not in DAILY_FUNDS_CAPABILITY_MAGICS
            or outcome not in DAILY_FUNDS_CAPABILITY_OUTCOMES
            or not _daily_funds_is_integer(count)
            or count < 1
            or count > 100_000
            or observed_at is None
            or not isinstance(code, str)
            or not code
            or len(code) > 80
            or not isinstance(parser_version, str)
            or not parser_version
            or len(parser_version) > 128
            or (declared_mime is not None and (
                not isinstance(declared_mime, str)
                or not declared_mime.isascii()
                or not declared_mime
                or len(declared_mime) > 128
            ))
        ):
            return unknown
        if outcome == "SUPPORTED" and code != "PARSER_OPEN_OK":
            return unknown
        # A review result remains an explicit non-pass even if the worker has
        # added a newer parser failure code than this app knows about.
        if outcome == "NEEDS_REVIEW" and _public_failure_code(code) is None:
            return unknown
        parsed_at = datetime.fromisoformat(observed_at.replace("Z", "+00:00"))
        if latest is None or parsed_at > latest[0]:
            latest = (parsed_at, observed_at)
        if outcome == "SUPPORTED":
            supported += count
        else:
            needs_review += count
            label = _daily_funds_capability_diagnostic_label(code)
            review_reasons[label] = review_reasons.get(label, 0) + count

    if not rows:
        return unobserved
    return {
        "状态": "待复核" if needs_review else "已支持",
        "已支持附件数": supported,
        "待复核附件数": needs_review,
        "待复核原因": [
            {"类别": label, "数量": review_reasons[label]}
            for label in sorted(
                review_reasons,
                key=lambda item: (DAILY_FUNDS_CAPABILITY_DIAGNOSTIC_ORDER.get(item, 99), item),
            )
        ],
        "最近观测": latest[1] if latest is not None else None,
    }


def _daily_funds_capability_diagnostic_label(code: str) -> str:
    """Return a fixed public-safe category for a protected parser code."""

    if code in DAILY_FUNDS_CAPABILITY_DIAGNOSTIC_LABELS:
        return DAILY_FUNDS_CAPABILITY_DIAGNOSTIC_LABELS[code]
    if code.startswith("OCR_"):
        return "图片表格结构未通过确定性校验"
    if code.startswith("SOURCE_"):
        return "来源或字节校验未通过"
    if code.startswith(("CSV_", "XLSX_", "COLUMN_", "BUSINESS_DATE_", "TRANSACTION_", "ACCOUNT_", "CURRENCY_", "INTERNAL_TRANSFER_")):
        return "表格字段或业务规则未通过确定性校验"
    return "其他确定性解析门未通过"


def _daily_funds_operation_receipts(rows: object) -> dict[str, dict[str, object]]:
    """Project one values-free receipt per independently scheduled operation.

    ``status.json`` describes publication truth.  A successful auth probe is
    useful evidence about the isolated DWS session, but it is neither a money
    publication nor a substitute for the source poll.  The worker writes the
    two facts separately and this reducer keeps malformed entries explicit as
    UNKNOWN rather than allowing an arbitrary shared-volume string into the
    owner UI.
    """

    receipts: dict[str, dict[str, object]] = {
        label: {"状态": "UNKNOWN", "结果": "UNKNOWN", "最近一次": None}
        for _, label in DAILY_FUNDS_OPERATION_RECEIPTS
    }
    if not isinstance(rows, dict):
        return receipts
    for job, label in DAILY_FUNDS_OPERATION_RECEIPTS:
        row = rows.get(job)
        if not isinstance(row, dict):
            continue
        state = row.get("state")
        code = _public_failure_code(row.get("code"))
        finished_at = _daily_funds_timestamp(row.get("finished_at"))
        started_at = _daily_funds_timestamp(row.get("started_at"))
        if state == "RUNNING" and code is not None and started_at is not None:
            receipts[label] = {
                "状态": "处理中",
                "结果": code,
                "最近一次": started_at,
            }
            continue
        if state not in {"SUCCEEDED", "FAILED"} or code is None or finished_at is None:
            continue
        receipts[label] = {
            "状态": "成功" if state == "SUCCEEDED" else "失败",
            "结果": code,
            "最近一次": finished_at,
        }
    return receipts


def _daily_funds_historical_backfill(
    value: object,
    receipt: dict[str, object],
) -> dict[str, object]:
    """Expose only values-free historical coverage and its latest job receipt.

    The worker keeps its next business-date cursor in a private SQLite volume.
    This boundary deliberately projects aggregate coverage only; neither a
    source date nor an attachment/message identifier reaches the app.
    """

    unknown = {
        "状态": "UNKNOWN",
        "窗口天数": 360,
        "已覆盖天数": None,
        "待覆盖天数": None,
        "最近作业": dict(receipt),
    }
    if not isinstance(value, dict):
        return unknown
    state = value.get("state")
    window_days = value.get("window_days")
    completed_days = value.get("completed_days")
    remaining_days = value.get("remaining_days")
    if (
        state not in {"NOT_STARTED", "IN_PROGRESS", "COMPLETE", "NEEDS_ATTENTION"}
        or not _daily_funds_is_integer(window_days)
        or window_days != 360
        or not _daily_funds_is_integer(completed_days)
        or not _daily_funds_is_integer(remaining_days)
        or completed_days < 0
        or completed_days > window_days
        or remaining_days < 0
        or remaining_days > window_days
        or completed_days + remaining_days != window_days
        or (state == "NOT_STARTED" and completed_days != 0)
        or (state == "COMPLETE" and remaining_days != 0)
    ):
        return unknown
    status = {
        "NOT_STARTED": "未开始",
        "IN_PROGRESS": "进行中",
        "COMPLETE": "已完成",
        "NEEDS_ATTENTION": "需处理",
    }[state]
    return {
        "状态": status,
        "窗口天数": window_days,
        "已覆盖天数": completed_days,
        "待覆盖天数": remaining_days,
        "最近作业": dict(receipt),
    }


def _daily_funds_flow_state() -> dict[str, Any]:
    """Safely fold the worker's flow record into the existing status center.

    ``flow_state.json`` is not a dashboard and cannot become a source of
    authority: it is a values-free, worker-written hand-off.  The canonical
    human status remains ``status.json`` and malformed/extended input is
    reduced to an explicit unknown state before it reaches an API response.
    """

    status = _daily_funds_status()
    observer_schedule = str(status["schedules"].get("observer") or "")
    if observer_schedule != "30 3 * * * Asia/Shanghai":
        observer_schedule = "30 3 * * * Asia/Shanghai"
    default_receipts = _daily_funds_operation_receipts(None)
    default = {
        "部署": {
            "运行": "UNKNOWN",
            "实例": "UNKNOWN",
            "身份": "UNKNOWN",
            "最近运行审计": None,
        },
        "业务流": {
            "阶段": "UNKNOWN",
            "状态": status["human_status"],
            "有效业务日期": status["effective_business_date"],
            "最近验证": status["last_verified_at"],
            "已验证发布": False,
        },
        "运行回执": default_receipts,
        "历史回填": _daily_funds_historical_backfill(None, default_receipts["历史回填"]),
        "来源诊断": _daily_funds_source_discovery(None),
        "附件能力": _daily_funds_attachment_capability_summary(None),
        "自愈": {
            "状态": "UNKNOWN",
            "重启恢复": "UNKNOWN",
            "恢复演练": "NOT_YET_RUN",
            "最近恢复演练": None,
        },
        "上线后观察": {
            "排程": observer_schedule,
            "状态": "NOT_STARTED",
            "最近对照": "NOT_STARTED",
            "需验证业务日": 5,
            "已完成业务日": 0,
            "基准业务日期": None,
            "开始时间": None,
            "最近观察": None,
            "每日对照": [],
        },
    }
    payload = _read_daily_funds_json("flow_state.json")
    if not isinstance(payload, dict) or payload.get("schema_version") != "kmfa.daily_funds.flow_state.v1":
        return default
    deployment = payload.get("deployment") if isinstance(payload.get("deployment"), dict) else {}
    business = payload.get("business_flow") if isinstance(payload.get("business_flow"), dict) else {}
    operation_receipts = _daily_funds_operation_receipts(payload.get("operations"))
    historical_backfill = _daily_funds_historical_backfill(
        payload.get("historical_backfill"),
        operation_receipts["历史回填"],
    )
    source_discovery = _daily_funds_source_discovery(payload.get("source_discovery"))
    attachment_capabilities = _daily_funds_attachment_capability_summary(payload.get("attachment_capabilities"))
    healing = payload.get("self_healing") if isinstance(payload.get("self_healing"), dict) else {}
    observer = payload.get("post_deploy_observer") if isinstance(payload.get("post_deploy_observer"), dict) else {}
    comparison_rows = observer.get("comparisons")
    comparisons: list[dict[str, Any]] = []
    if isinstance(comparison_rows, list):
        for row in comparison_rows[:5]:
            if not isinstance(row, dict):
                continue
            business_date = _daily_funds_date(row.get("business_date"))
            observed_at = _daily_funds_timestamp(row.get("observed_at"))
            latency = row.get("latency_minutes")
            if (
                business_date is None
                or observed_at is None
                or (latency is not None and (not _daily_funds_is_integer(latency) or latency < 0 or latency > 60 * 24 * 31))
            ):
                continue
            comparisons.append({
                "业务日期": business_date.isoformat(),
                "观察时间": observed_at,
                "对照": _daily_funds_flow_token(
                    row.get("comparison_state"),
                    allowed={"D1_AND_POINTER_VERIFIED"},
                    default="UNKNOWN",
                ),
                "覆盖": _daily_funds_flow_token(
                    row.get("coverage_state"),
                    allowed={"DIRECT_OBSERVATION"},
                    default="UNKNOWN",
                ),
                "金额": _daily_funds_flow_token(
                    row.get("amount_state"),
                    allowed={"ZERO_FEN"},
                    default="UNKNOWN",
                ),
                "阈值": _daily_funds_flow_token(
                    row.get("threshold_state"),
                    allowed={"VALID"},
                    default="UNKNOWN",
                ),
                "取数": _daily_funds_flow_token(
                    row.get("retrieval_state"),
                    allowed={"COMPLETE_PAIR"},
                    default="UNKNOWN",
                ),
                "重复": _daily_funds_flow_token(
                    row.get("duplicate_state"),
                    allowed={"SOURCE_VERSION_UNIQUE"},
                    default="UNKNOWN",
                ),
                "备份": _daily_funds_flow_token(
                    row.get("backup_state"),
                    allowed=DAILY_FUNDS_BACKUP_STATES,
                    default="UNKNOWN",
                ),
                "恢复": _daily_funds_flow_token(
                    row.get("restore_state"),
                    allowed={"OK", "IN_PROGRESS", "NEEDS_ATTENTION", "NOT_YET_RUN", "UNKNOWN"},
                    default="UNKNOWN",
                ),
                "延迟分钟": latency,
            })
    required_days = observer.get("required_business_days")
    completed_days = observer.get("completed_business_days")
    if not _daily_funds_is_integer(required_days) or required_days != 5:
        required_days = 5
    if not _daily_funds_is_integer(completed_days) or completed_days < 0 or completed_days > required_days:
        completed_days = 0
    return {
        "部署": {
            "运行": _daily_funds_flow_token(
                deployment.get("runtime_state"),
                allowed={"RUNTIME_AUDITED", "RUNTIME_NEEDS_ATTENTION", "UNKNOWN"},
                default="UNKNOWN",
            ),
            "实例": _daily_funds_flow_token(
                deployment.get("instance_state"),
                allowed={"OBSERVED", "UNKNOWN"},
                default="UNKNOWN",
            ),
            "身份": _daily_funds_embedded_source_identity(deployment),
            "最近运行审计": _daily_funds_timestamp(deployment.get("runtime_audit_at")),
        },
        "业务流": {
            "阶段": _daily_funds_flow_token(
                business.get("stage"),
                allowed=DAILY_FUNDS_BUSINESS_FLOW_STAGES,
                default="UNKNOWN",
            ),
            # The main status writer remains the only primary human status.
            "状态": status["human_status"],
            "有效业务日期": status["effective_business_date"],
            "最近验证": status["last_verified_at"],
            "已验证发布": business.get("publication_present") is True,
        },
        "运行回执": operation_receipts,
        "历史回填": historical_backfill,
        "来源诊断": source_discovery,
        "附件能力": attachment_capabilities,
        "自愈": {
            "状态": _daily_funds_flow_token(
                healing.get("state"), allowed={"JOURNAL_READY", "UNKNOWN"}, default="UNKNOWN",
            ),
            "重启恢复": _daily_funds_flow_token(
                healing.get("restart_recovery"), allowed={"CURSOR_INBOX_LEASES", "UNKNOWN"}, default="UNKNOWN",
            ),
            "恢复演练": _daily_funds_flow_token(
                healing.get("restore_drill"),
                allowed={"OK", "IN_PROGRESS", "NEEDS_ATTENTION", "NOT_YET_RUN", "UNKNOWN"},
                default="UNKNOWN",
            ),
            "最近恢复演练": _daily_funds_timestamp(healing.get("restore_drill_at")),
        },
        "上线后观察": {
            "排程": observer_schedule,
            "状态": _daily_funds_flow_token(
                observer.get("state"),
                allowed={
                    "NOT_STARTED", "WAITING_FOR_VALID_PUBLICATION", "WAITING_FOR_LOCK",
                    "BASELINE_CAPTURED", "WAITING_FOR_NEXT_BUSINESS_DATE", "OBSERVING",
                    "COMPLETE", "NEEDS_ATTENTION", "UNKNOWN",
                },
                default="NOT_STARTED",
            ),
            "最近对照": _daily_funds_flow_token(
                observer.get("last_comparison"),
                allowed={
                    "NOT_STARTED", "SOURCE_MISSING", "CONFIG_INVALID", "DEPLOYMENT_MARKER_UNAVAILABLE",
                    "OBSERVATION_CLOCK_INVALID", "POINTER_OR_HISTORY_INVALID", "STALE",
                    "D1_ORACLE_FAILED", "D1_AND_POINTER_VERIFIED", "PUBLISHER_LOCK_HELD",
                    "OBSERVER_LOCK_HELD", "POINTER_BEFORE_DEPLOYMENT_BASELINE", "NON_WORKING_DAY", "OBSERVER_FAILED",
                    "UNKNOWN",
                },
                default="UNKNOWN",
            ),
            "需验证业务日": required_days,
            "已完成业务日": completed_days,
            "基准业务日期": (
                _daily_funds_date(observer.get("baseline_business_date")).isoformat()
                if _daily_funds_date(observer.get("baseline_business_date")) is not None else None
            ),
            "开始时间": _daily_funds_timestamp(observer.get("started_at")),
            "最近观察": _daily_funds_timestamp(observer.get("last_observed_at")),
            "每日对照": comparisons,
        },
    }


def _daily_funds_schedule_row() -> dict[str, Any]:
    status = _daily_funds_status()
    flow = _daily_funds_flow_state()
    poll = flow["运行回执"]["历史轮询"]
    poll_state = poll["状态"]
    poll_ran = poll_state in {"成功", "失败", "处理中"} and poll["最近一次"] is not None
    poll_terminal = poll_state in {"成功", "失败"} and poll["最近一次"] is not None
    poll_succeeded = poll_state == "成功"
    return {
        "技能": "daily-funds",
        "业务模块": "每日资金",
        "约定时刻": SCHEDULE_CONTRACT["daily-funds"],
        # Publication truth and scheduler truth are deliberately separate:
        # a minute-level AUTH_OK is not a 15-minute source-poll success, while
        # a failed poll must not disappear behind a later successful probe.
        "跑过": poll_ran,
        "最近一次": poll["最近一次"] if poll_ran else None,
        "距今小时": None,
        "退出码": 0 if poll_succeeded else 1 if poll_state == "失败" else None,
        "成功": poll_succeeded if poll_terminal else None,
        "运行中": poll_state == "处理中",
        "失败码": poll["结果"] if poll_state == "失败" else None,
        # This isolated worker has no delivery switch; ``None`` renders as an
        # em dash rather than the misleading shared-skill "空跑" label.
        "投递开关": None,
        "次数": None,
        "失败次数": None,
        "成功率": None,
        "连续失败": 0 if poll_succeeded else 1 if poll_state == "失败" else None,
        "历史": [],
        "压测": None,
        "每日资金状态": {
            "状态": status["human_status"],
            "有效业务日期": status["effective_business_date"],
            "最近验证": status["last_verified_at"],
            "备份": status["backup_state"],
            "排程": status["schedules"],
            "业务流": flow,
        },
    }


def _daily_funds_public_skill_health_row(now: datetime) -> dict[str, Any]:
    """Expose the isolated worker's latest *values-free* poll/backfill receipts.

    ``daily-funds`` deliberately does not write the shared ``kmfa-logs``
    ledger: sharing it would also share the DWS profile and state that its
    isolation contract forbids.  The public health endpoint must therefore
    never manufacture a shared-ledger zero for this worker.  It may expose
    only the already schema-checked latest poll and historical-backfill
    receipts -- no business date, amount, source metadata, raw log, or
    identifier crosses this boundary.

    The worker retains one latest receipt rather than an append-only public
    history.  ``运行次数`` below is consequently the count of *verifiable
    retained poll receipts* (0 or 1), explicitly not a lifetime run count.
    """

    row = _daily_funds_schedule_row()
    ran = bool(row["跑过"])
    last = row["最近一次"] if ran else None
    elapsed_hours = None
    if last is not None:
        try:
            elapsed_hours = round(
                (now - datetime.fromisoformat(str(last))).total_seconds() / 3600,
                1,
            )
        except (TypeError, ValueError):
            elapsed_hours = None

    flow = row["每日资金状态"]["业务流"]
    poll = flow["运行回执"]["历史轮询"]
    outcome = _public_failure_code(poll.get("结果")) or "UNKNOWN"
    backfill = flow["运行回执"]["历史回填"]
    backfill_state = backfill["状态"]
    backfill_ran = (
        backfill_state in {"成功", "失败", "处理中"}
        and backfill["最近一次"] is not None
    )
    backfill_terminal = (
        backfill_state in {"成功", "失败"}
        and backfill["最近一次"] is not None
    )
    backfill_outcome = _public_failure_code(backfill.get("结果")) or "UNKNOWN"
    return {
        "技能": "daily-funds",
        "最近一次": last,
        "距今小时": elapsed_hours,
        "退出码": row["退出码"],
        "成功": row["成功"],
        "运行次数": 1 if ran else 0,
        "运行计数口径": "仅保留最近一次历史轮询回执，非累计历史次数",
        "失败码": row["失败码"],
        "本次状态": outcome,
        "运行中": bool(row["运行中"]),
        # Coolify currently exposes neither a usable container exec endpoint
        # nor stdout cron logs.  This independently schema-checked receipt is
        # therefore the values-free evidence that the cloud-native backfill
        # scheduler actually ran; it is never publication or money evidence.
        "历史回填": {
            "最近一次": backfill["最近一次"] if backfill_ran else None,
            "退出码": 0 if backfill_state == "成功" else 1 if backfill_state == "失败" else None,
            "成功": backfill_state == "成功" if backfill_terminal else None,
            "运行次数": 1 if backfill_ran else 0,
            "运行计数口径": "仅保留最近一次历史回填回执，非累计历史次数",
            "失败码": backfill_outcome if backfill_state == "失败" else None,
            "本次状态": backfill_outcome,
            "运行中": backfill_state == "处理中",
        },
    }


def _daily_funds_is_integer(value: object) -> bool:
    return isinstance(value, int) and not isinstance(value, bool)


def _daily_funds_date(value: object) -> date | None:
    if not isinstance(value, str):
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _daily_funds_timestamp(value: object) -> str | None:
    if not isinstance(value, str) or len(value) > 40:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    return value if parsed.tzinfo is not None else None


def _daily_funds_lower_hex(value: object, length: int) -> bool:
    return (
        isinstance(value, str)
        and len(value) == length
        and all(character in "0123456789abcdef" for character in value)
    )


def _daily_funds_current() -> dict[str, Any]:
    """Return only a structurally valid, already-published projection.

    The daily-funds worker is the sole producer.  The application must not
    make a partially written, malformed, or merely plausible JSON file look
    like a trusted financial publication just because it happens to be on the
    shared read-only volume.
    """

    target = (DAILY_FUNDS_PUBLICATION_DIR / "current.json").resolve()
    root = DAILY_FUNDS_PUBLICATION_DIR.resolve()
    # A missing projection is the normal pre-publication state.  It is not a
    # malformed publication and the owner page must be able to render its
    # values-free waiting view without turning that state into a transport
    # failure.  Anything present but unreadable/structurally wrong remains a
    # hard failure below.
    if target.parent != root or not target.is_file():
        raise HTTPException(status_code=503, detail="daily_funds_projection_unavailable")
    payload = _read_daily_funds_json("current.json")
    if not isinstance(payload, dict) or not isinstance(payload.get("publication"), dict):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    expected_snapshot_fields = {
        "schema_version", "publication", "summary", "daily_balances", "transactions",
    }
    snapshot_fields = set(payload)
    if (
        snapshot_fields != expected_snapshot_fields
        and snapshot_fields != expected_snapshot_fields | {"runtime"}
    ) or payload.get("schema_version") != "kmfa.daily_funds.current_projection.v1":
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    publication = payload["publication"]
    source_versions = publication.get("source_versions")
    expected_fields = {
        "publication_id", "business_date", "status", "source_versions", "reconciliation_difference_fen",
        "threshold_snapshot", "created_at", "git_commit_sha", "d1_projection_version", "r2_manifest_sha256",
        "oci_backup_state",
    }
    if (
        set(publication) != expected_fields
        or publication.get("status") != "VALID"
        or not _daily_funds_is_integer(publication.get("reconciliation_difference_fen"))
        or publication.get("reconciliation_difference_fen") != 0
        or not _daily_funds_lower_hex(publication.get("publication_id"), 64)
        or _daily_funds_date(publication.get("business_date")) is None
        or _daily_funds_timestamp(publication.get("created_at")) is None
        or not _daily_funds_lower_hex(publication.get("git_commit_sha"), 40)
        or not _daily_funds_lower_hex(publication.get("r2_manifest_sha256"), 64)
        or publication.get("d1_projection_version") != "kmfa.daily_funds.d1.v1"
        or not isinstance(source_versions, list)
        or len(source_versions) != 2
        or len({item.get("source_version") for item in source_versions if isinstance(item, dict)}) != len(source_versions)
        or any(
            not isinstance(item, dict)
            or set(item) != {"source_version"}
            or not _daily_funds_lower_hex(item.get("source_version"), 64)
            for item in source_versions
        )
    ):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    return payload


def _daily_funds_range_days(
    range_value: str,
    from_date: str | None,
    to_date: str | None,
    *,
    publication_day: date,
) -> tuple[str, str]:
    # A trusted publication, not wall-clock time, anchors preset ranges.  If
    # the worker is delayed, using "today" would create an empty chart and
    # falsely imply that a valid prior snapshot has no data.
    if range_value in DAILY_FUNDS_ALLOWED_RANGES:
        start = publication_day - timedelta(days=DAILY_FUNDS_ALLOWED_RANGES[range_value] - 1)
        return start.isoformat(), publication_day.isoformat()
    if range_value != "custom" or not from_date or not to_date:
        raise HTTPException(status_code=422, detail="daily_funds_range_invalid")
    start = _daily_funds_date(from_date)
    end = _daily_funds_date(to_date)
    if start is None or end is None or end < start or (end - start).days + 1 < 7:
        raise HTTPException(status_code=422, detail="daily_funds_custom_range_invalid")
    return start.isoformat(), end.isoformat()


def _daily_funds_unpublished_range_days(
    range_value: str,
    from_date: str | None,
    to_date: str | None,
) -> tuple[str | None, str | None]:
    """Validate a UI range before a first publication exists.

    Preset ranges are normally anchored to the last trusted business date.
    That anchor does not exist before the first publication, so returning
    dates would make an empty chart look like a zero-balance interval.  A
    valid custom range may still be echoed as an operator-selected filter;
    no source or monetary fact is implied by it.
    """

    if range_value in DAILY_FUNDS_ALLOWED_RANGES:
        return None, None
    if range_value != "custom" or not from_date or not to_date:
        raise HTTPException(status_code=422, detail="daily_funds_range_invalid")
    start = _daily_funds_date(from_date)
    end = _daily_funds_date(to_date)
    if start is None or end is None or end < start or (end - start).days + 1 < 7:
        raise HTTPException(status_code=422, detail="daily_funds_custom_range_invalid")
    return start.isoformat(), end.isoformat()


def _daily_funds_filtered_timeseries(payload: dict[str, Any], start: str, end: str) -> list[dict[str, Any]]:
    rows = payload.get("daily_balances")
    if not isinstance(rows, list):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    safe: list[dict[str, Any]] = []
    seen_days: set[str] = set()
    for row in rows:
        if not isinstance(row, dict):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        observed = _daily_funds_date(row.get("business_date"))
        amount = row.get("ending_available_fen")
        direct = row.get("direct_observation")
        gap = row.get("coverage_gap")
        carried = row.get("carried_forward")
        if (
            observed is None
            or not _daily_funds_is_integer(amount)
            or not all(isinstance(flag, bool) for flag in (direct, gap, carried))
            or (direct and (gap or carried))
            or (gap and carried)
            or (not direct and not gap and not carried)
        ):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        business_date = observed.isoformat()
        if business_date in seen_days:
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        seen_days.add(business_date)
        if start <= business_date <= end:
            safe.append({
                "business_date": business_date,
                "ending_available_fen": amount,
                "direct_observation": direct,
                "coverage_gap": gap,
                "carried_forward": carried,
            })
    return sorted(safe, key=lambda row: row["business_date"])


def _daily_funds_range_health(points: list[dict[str, Any]], start: str, end: str) -> dict[str, Any]:
    start_day = _daily_funds_date(start)
    end_day = _daily_funds_date(end)
    if start_day is None or end_day is None:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    expected_dates = [
        (start_day + timedelta(days=offset)).isoformat()
        for offset in range((end_day - start_day).days + 1)
    ]
    present = {str(point["business_date"]) for point in points}
    coverage_gaps = [str(point["business_date"]) for point in points if point["coverage_gap"]]
    return {
        "expected_days": len(expected_dates),
        "published_days": len(points),
        "expected_dates": expected_dates,
        "missing_dates": [business_date for business_date in expected_dates if business_date not in present],
        "coverage_gap_dates": coverage_gaps,
    }


def _daily_funds_projection_day(payload: dict[str, Any]) -> date:
    publication = payload["publication"]
    business_day = _daily_funds_date(publication.get("business_date"))
    if business_day is None:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    return business_day


def _daily_funds_safe_label(value: object) -> str:
    """Return a bounded display label, never an arbitrary JSON value."""
    if not isinstance(value, str):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    label = value.strip()
    if not label or len(label) > 80 or any(ord(character) < 32 for character in label):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    return label


def _daily_funds_safe_breakdown(value: object, *, total: int) -> list[dict[str, Any]]:
    if not isinstance(value, dict):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    safe: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw_label, amount in value.items():
        label = _daily_funds_safe_label(raw_label)
        if label in seen or not _daily_funds_is_integer(amount):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        seen.add(label)
        safe.append({"label": label, "ending_available_fen": amount})
    if not safe or sum(int(row["ending_available_fen"]) for row in safe) != total:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    return sorted(safe, key=lambda row: (-abs(int(row["ending_available_fen"])), str(row["label"])))


def _daily_funds_safe_accounts(value: object, *, total: int) -> list[dict[str, Any]]:
    if not isinstance(value, dict):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    safe: list[dict[str, Any]] = []
    for account_hash, amount in value.items():
        if not _daily_funds_lower_hex(account_hash, 64) or not _daily_funds_is_integer(amount):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        safe.append({"account_alias": f"••••{str(account_hash)[-4:]}", "ending_available_fen": amount})
    if not safe or sum(int(row["ending_available_fen"]) for row in safe) != total:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    return sorted(safe, key=lambda row: (-abs(int(row["ending_available_fen"])), str(row["account_alias"])))


def _daily_funds_safe_transactions(
    payload: dict[str, Any],
    *,
    publication_day: date,
    declared_source_versions: set[str],
) -> list[dict[str, Any]]:
    rows = payload.get("transactions")
    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    expected = {
        "transaction_key_hash", "business_date", "inflow_fen", "outflow_fen", "adjustment_fen",
        "internal_transfer", "source_version", "message_id_hash",
    }
    safe: list[dict[str, Any]] = []
    seen: set[str] = set()
    transaction_versions: set[str] = set()
    for row in rows:
        if not isinstance(row, dict) or set(row) != expected:
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        key = row.get("transaction_key_hash")
        observed = _daily_funds_date(row.get("business_date"))
        inflow = row.get("inflow_fen")
        outflow = row.get("outflow_fen")
        adjustment = row.get("adjustment_fen")
        internal_transfer = row.get("internal_transfer")
        if (
            not _daily_funds_lower_hex(key, 64)
            or key in seen
            or observed != publication_day
            or not all(_daily_funds_is_integer(amount) for amount in (inflow, outflow, adjustment))
            or int(inflow) < 0
            or int(outflow) < 0
            or (int(inflow) and int(outflow))
            or not isinstance(internal_transfer, bool)
            or not _daily_funds_lower_hex(row.get("source_version"), 64)
            or row["source_version"] not in declared_source_versions
            or not _daily_funds_lower_hex(row.get("message_id_hash"), 64)
        ):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        seen.add(str(key))
        transaction_versions.add(str(row["source_version"]))
        safe.append({
            "transaction_ref": str(key)[-8:],
            "business_date": observed.isoformat(),
            "inflow_fen": inflow,
            "outflow_fen": outflow,
            "adjustment_fen": adjustment,
            "internal_transfer": internal_transfer,
        })
    # Publication is a strictly paired account/transaction fact set.  The
    # upstream D1 gate already verifies the complementary account source; the
    # browser projection must at least reject a shared-volume snapshot that
    # tries to mix several transaction source versions or cite a third source.
    if len(transaction_versions) != 1:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    return sorted(safe, key=lambda row: (str(row["business_date"]), str(row["transaction_ref"])))


def _daily_funds_runtime_backup_state(payload: dict[str, Any]) -> str:
    """Validate the operational hand-off without exposing its identifiers.

    The first pointer write is allowed to have no ``runtime`` member while
    OCI is pending.  Final publications and restored pointers must otherwise
    carry one of the exact writer-produced shapes; arbitrary fields are never
    ignored on the shared projection volume.
    """

    if "runtime" not in payload:
        return "PENDING"
    runtime = payload.get("runtime")
    if not isinstance(runtime, dict):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    state = runtime.get("oci_backup_state")
    if state not in {"OK", "LAG"}:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    keys = set(runtime)
    if keys == {"oci_backup_state", "git_publication_commit_sha"}:
        if state != "LAG" or not _daily_funds_lower_hex(runtime.get("git_publication_commit_sha"), 40):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    elif keys == {"oci_backup_state", "git_publication_commit_sha", "oci_restore_manifest_sha"}:
        if (
            state != "OK"
            or not _daily_funds_lower_hex(runtime.get("git_publication_commit_sha"), 40)
            or not _daily_funds_lower_hex(runtime.get("oci_restore_manifest_sha"), 64)
        ):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    elif keys == {"oci_backup_state", "restored_at"}:
        if state != "OK" or _daily_funds_timestamp(runtime.get("restored_at")) is None:
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    else:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    return state


def _daily_funds_safe_thresholds(publication: dict[str, Any], *, total_available_fen: int) -> dict[str, Any]:
    """Validate the frozen threshold decision before it reaches the owner UI.

    The App reads a shared projection volume rather than the worker's private
    Git/D1 authority.  It must therefore independently reject an extension or
    semantic drift in ``threshold_snapshot`` instead of displaying otherwise
    plausible money under a different decision rule.
    """

    snapshot = publication.get("threshold_snapshot")
    expected_snapshot_fields = {"currency", "fixed", "floating", "fixed_risk", "dynamic_flag"}
    if not isinstance(snapshot, dict) or set(snapshot) != expected_snapshot_fields:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    fixed = snapshot.get("fixed")
    floating = snapshot.get("floating")
    if (
        snapshot.get("currency") != "CNY"
        or
        not isinstance(fixed, dict)
        or set(fixed) != {"hard_fen", "soft_fen"}
        or fixed.get("hard_fen") != DAILY_FUNDS_HARD_THRESHOLD_FEN
        or fixed.get("soft_fen") != DAILY_FUNDS_SOFT_THRESHOLD_FEN
        or not isinstance(floating, list)
        or snapshot.get("fixed_risk") not in DAILY_FUNDS_FIXED_RISKS
        or (snapshot.get("dynamic_flag") is not None and snapshot.get("dynamic_flag") not in DAILY_FUNDS_DYNAMIC_FLAGS)
    ):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    required = {
        "name", "threshold_fen", "start", "end", "days", "direct_observations", "covered_days",
        "carried_forward_days", "coverage", "active", "reason",
    }
    lines: list[dict[str, Any]] = []
    seen: set[str] = set()
    for line in floating:
        if not isinstance(line, dict) or set(line) != required:
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        name = _daily_funds_safe_label(line.get("name"))
        threshold = line.get("threshold_fen")
        start = _daily_funds_date(line.get("start"))
        end = _daily_funds_date(line.get("end"))
        fields = (line.get("days"), line.get("direct_observations"), line.get("covered_days"), line.get("carried_forward_days"))
        coverage = line.get("coverage")
        reason = line.get("reason")
        if (
            name in seen
            or name not in DAILY_FUNDS_FLOATING_LINE_NAMES
            or start is None
            or end is None
            or end < start
            or not all(_daily_funds_is_integer(value) and int(value) >= 0 for value in fields)
            or not isinstance(coverage, str)
            or not coverage
            or coverage != coverage.strip()
            or len(coverage) > 24
            or not isinstance(line.get("active"), bool)
            or (threshold is not None and not _daily_funds_is_integer(threshold))
            or (reason is not None and (not isinstance(reason, str) or len(reason) > 120))
        ):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        days, direct, covered, carried = (int(value) for value in fields)
        if (
            days <= 0
            or days != (end - start).days + 1
            or direct + carried != covered
            or covered > days
        ):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        try:
            coverage_value = Decimal(coverage)
        except InvalidOperation as exc:
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid") from exc
        if (
            not coverage_value.is_finite()
            or coverage_value < 0
            or coverage_value > 1
            or coverage_value != Decimal(covered) / Decimal(days)
        ):
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        if line["active"]:
            if not _daily_funds_is_integer(threshold) or int(threshold) < 0 or reason is not None:
                raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        elif threshold is not None or not isinstance(reason, str) or not reason.strip():
            raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
        seen.add(name)
        lines.append({
            "name": name,
            "threshold_fen": threshold,
            "start": start.isoformat(),
            "end": end.isoformat(),
            "days": line["days"],
            "direct_observations": line["direct_observations"],
            "covered_days": line["covered_days"],
            "carried_forward_days": line["carried_forward_days"],
            "coverage": coverage,
            "active": line["active"],
            "reason": reason,
        })
    fixed_risk = (
        "高风险" if total_available_fen <= DAILY_FUNDS_HARD_THRESHOLD_FEN
        else "关注" if total_available_fen <= DAILY_FUNDS_SOFT_THRESHOLD_FEN
        else "正常"
    )
    active_thresholds = [int(line["threshold_fen"]) for line in lines if line["active"]]
    dynamic_flag: str | None
    if not active_thresholds:
        dynamic_flag = None
    elif all(total_available_fen <= threshold for threshold in active_thresholds):
        dynamic_flag = "动态明显偏低"
    elif any(total_available_fen <= threshold for threshold in active_thresholds):
        dynamic_flag = "动态偏低"
    else:
        dynamic_flag = None
    if snapshot["fixed_risk"] != fixed_risk or snapshot["dynamic_flag"] != dynamic_flag:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    return {
        "fixed": {"hard_fen": DAILY_FUNDS_HARD_THRESHOLD_FEN, "soft_fen": DAILY_FUNDS_SOFT_THRESHOLD_FEN},
        "floating": lines,
        "fixed_risk": fixed_risk,
        "dynamic_flag": dynamic_flag,
    }


def _daily_funds_unpublished_thresholds() -> dict[str, Any]:
    """Expose the frozen policy without implying that a money projection exists.

    The fixed 60/120 万 lines are task-pack configuration, rather than an
    observation from an account or transaction.  They remain useful to orient
    the UI while the publication gate is closed.  Dynamic lines intentionally
    stay inactive: deriving one requires verified daily balances and must never
    be guessed from an incomplete source chain.
    """

    return {
        "fixed": {
            "hard_fen": DAILY_FUNDS_HARD_THRESHOLD_FEN,
            "soft_fen": DAILY_FUNDS_SOFT_THRESHOLD_FEN,
        },
        "floating": [
            {
                "name": name,
                "threshold_fen": None,
                "start": None,
                "end": None,
                "days": 0,
                "direct_observations": 0,
                "covered_days": 0,
                "carried_forward_days": 0,
                "coverage": "0",
                "active": False,
                "reason": "尚无足够已验证日余额",
            }
            for name in DAILY_FUNDS_FLOATING_LINE_ORDER
        ],
        "fixed_risk": None,
        "dynamic_flag": None,
    }


def _daily_funds_unpublished_summary(
    *,
    range_value: str,
    from_date: str | None,
    to_date: str | None,
) -> dict[str, Any]:
    """Return the stable, values-free contract used before first publication.

    This is deliberately a successful API response, not a synthetic zero
    projection.  The page needs enough shape to render its frozen policy and
    publication gates, while every financial field stays absent/empty until
    the worker has written a fully validated integer-fen publication.
    """

    start, end = _daily_funds_unpublished_range_days(range_value, from_date, to_date)
    return {
        "data_available": False,
        "range": range_value,
        "from": start,
        "to": end,
        "scope": "global",
        "granularity": "daily",
        "range_health": {
            "expected_days": 0,
            "published_days": 0,
            "expected_dates": [],
            "missing_dates": [],
            "coverage_gap_dates": [],
        },
        "publication": None,
        "total_available_fen": None,
        "risk_label": None,
        "dynamic_flag": None,
        "by_company_ending_fen": [],
        "by_bank_ending_fen": [],
        "account_breakdown": [],
        "today": {},
        "top_inflows": [],
        "top_outflows": [],
        "points": [],
    }


def _daily_funds_unpublished_timeseries(
    *,
    range_value: str,
    from_date: str | None,
    to_date: str | None,
) -> dict[str, Any]:
    start, end = _daily_funds_unpublished_range_days(range_value, from_date, to_date)
    return {
        "data_available": False,
        "range": range_value,
        "from": start,
        "to": end,
        "granularity": "daily",
        "range_health": {
            "expected_days": 0,
            "published_days": 0,
            "expected_dates": [],
            "missing_dates": [],
            "coverage_gap_dates": [],
        },
        "points": [],
        "thresholds": _daily_funds_unpublished_thresholds(),
    }


def _daily_funds_projection_view(payload: dict[str, Any]) -> dict[str, Any]:
    """Validate the entire browser projection before exposing any financial value."""
    publication = payload["publication"]
    publication_day = _daily_funds_projection_day(payload)
    all_points = _daily_funds_filtered_timeseries(payload, "0001-01-01", "9999-12-31")
    if any(point["business_date"] > publication_day.isoformat() for point in all_points):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    points = [point for point in all_points if point["business_date"] <= publication_day.isoformat()]
    current_point = next((point for point in points if point["business_date"] == publication_day.isoformat()), None)
    summary = payload.get("summary")
    if not isinstance(summary, dict) or not _daily_funds_is_integer(summary.get("total_available_fen")):
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    total = int(summary["total_available_fen"])
    if current_point is None or not current_point["direct_observation"] or current_point["ending_available_fen"] != total:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    thresholds = _daily_funds_safe_thresholds(publication, total_available_fen=total)
    dynamic_flag = thresholds["dynamic_flag"]
    expected_risk_label = thresholds["fixed_risk"] if thresholds["fixed_risk"] != "正常" else (dynamic_flag or "正常")
    if summary.get("risk_label") != expected_risk_label or summary.get("dynamic_flag") != dynamic_flag:
        raise HTTPException(status_code=503, detail="daily_funds_projection_not_valid")
    declared_source_versions = {
        str(row["source_version"])
        for row in publication["source_versions"]
    }
    backup_state = _daily_funds_runtime_backup_state(payload)
    return {
        "publication_day": publication_day,
        "points": points,
        "transactions": _daily_funds_safe_transactions(
            payload,
            publication_day=publication_day,
            declared_source_versions=declared_source_versions,
        ),
        "thresholds": thresholds,
        "total_available_fen": total,
        "risk_label": expected_risk_label,
        "dynamic_flag": dynamic_flag,
        "by_company_ending_fen": _daily_funds_safe_breakdown(summary.get("by_company_ending_fen"), total=total),
        "by_bank_ending_fen": _daily_funds_safe_breakdown(summary.get("by_bank_ending_fen"), total=total),
        "account_breakdown": _daily_funds_safe_accounts(summary.get("account_ending_by_hash"), total=total),
        "backup_state": backup_state,
    }


def _daily_funds_activity(rows: list[dict[str, Any]], *, publication_day: date) -> dict[str, Any]:
    today_rows = [row for row in rows if row["business_date"] == publication_day.isoformat()]
    today = {
        "inflow_fen": sum(int(row["inflow_fen"]) for row in today_rows),
        "outflow_fen": sum(int(row["outflow_fen"]) for row in today_rows),
        "adjustment_fen": sum(int(row["adjustment_fen"]) for row in today_rows),
        "internal_transfer_count": sum(bool(row["internal_transfer"]) for row in today_rows),
    }
    today["net_change_fen"] = int(today["inflow_fen"]) - int(today["outflow_fen"]) + int(today["adjustment_fen"])
    inflows = sorted((row for row in today_rows if int(row["inflow_fen"]) > 0), key=lambda row: (-int(row["inflow_fen"]), str(row["transaction_ref"])))[:5]
    outflows = sorted((row for row in today_rows if int(row["outflow_fen"]) > 0), key=lambda row: (-int(row["outflow_fen"]), str(row["transaction_ref"])))[:5]
    return {"today": today, "top_inflows": inflows, "top_outflows": outflows}


def _daily_funds_public_control_value(control: object) -> dict[str, Any] | None:
    if not isinstance(control, dict):
        return None
    mode = control.get("mode")
    revision = control.get("revision")
    expected = {"schema_version", "mode", "revision", "applied_at", "actor", "reason"}
    if mode == "numeric":
        expected.add("amount_fen")
    elif mode == "date_range":
        expected.update({"from", "to"})
    if (
        set(control) != expected
        or control.get("schema_version") != "kmfa.daily_funds.threshold_control.v1"
        or mode not in {"disabled", "date_range", "numeric"}
        or not _daily_funds_lower_hex(revision, 64)
        or _daily_funds_timestamp(control.get("applied_at")) is None
        or control.get("actor") != "kmfa_private_owner_ui"
    ):
        return None
    reason = control.get("reason")
    if not isinstance(reason, str) or not reason.strip() or len(reason) > 500 or any(ord(character) < 32 for character in reason):
        return None
    public: dict[str, Any] = {
        "mode": mode,
        "revision": revision,
        "applied_at": control["applied_at"],
        "actor": "kmfa_private_owner_ui",
        "reason": reason.strip(),
    }
    if mode == "numeric" and _daily_funds_is_integer(control.get("amount_fen")):
        public["amount_fen"] = control["amount_fen"]
    elif mode == "date_range":
        start = _daily_funds_date(control.get("from"))
        end = _daily_funds_date(control.get("to"))
        if start is not None and end is not None and end >= start:
            public["from"] = start.isoformat()
            public["to"] = end.isoformat()
        else:
            return None
    elif mode != "disabled":
        return None
    return public


def _daily_funds_public_control() -> dict[str, Any] | None:
    return _daily_funds_public_control_value(_read_daily_funds_control())


def _daily_funds_source_health_view() -> dict[str, Any]:
    status = _daily_funds_status()
    flow = _daily_funds_flow_state()
    parser_capability = flow["附件能力"]
    source_discovery = flow["来源诊断"]
    historical_backfill = flow["历史回填"]
    view: dict[str, Any] = {
        "human_status": status["human_status"],
        "effective_business_date": status["effective_business_date"],
        "last_verified_at": status["last_verified_at"],
        "updated_at": status["updated_at"],
        "backup_state": status["backup_state"],
        # This is an operational parser receipt, not an attachment field:
        # no raw attachment metadata crosses the app boundary.
        "parser_capability": parser_capability,
        "source_discovery": source_discovery,
        "historical_backfill": historical_backfill,
        "has_trusted_publication": False,
        "message": "尚无可展示的已验证资金数据。",
    }
    try:
        payload = _daily_funds_current()
        projection = _daily_funds_projection_view(payload)
    except HTTPException:
        # A stale status hand-off must not make malformed or absent money look
        # like a fresh success in the owner UI.  Preserve an in-flight state,
        # but downgrade a claimed update to the explicit actionable state.
        if view["human_status"] == "已更新":
            view["human_status"] = "需处理"
        view["backup_state"] = "UNKNOWN"
        if parser_capability["状态"] == "待复核":
            view["message"] = "已归档附件待确定性解析复核；在账户余额与资金流水成对勾稽并完成发布前，不展示或推断金额。"
        else:
            view["message"] = (
                "正在等待可验证的正式资金数据。" if view["human_status"] == "处理中"
                else "正式资金 projection 未通过完整性校验，需处理。"
            )
        return view
    publication = payload["publication"]
    view.update({
        "has_trusted_publication": True,
        "effective_business_date": publication["business_date"],
        "last_verified_at": publication["created_at"],
        "backup_state": projection["backup_state"],
        "source_families": {"required": 2, "published": len(publication["source_versions"])},
        "reconciliation_difference_fen": 0,
        "git_evidence_available": True,
        "r2_mirror_available": True,
        "evidence_version": str(publication["publication_id"])[-12:],
    })
    view["message"] = (
        "当前显示已验证的正式资金数据。" if view["human_status"] == "已更新"
        else "最新运行尚未覆盖；页面保留上一份已验证资金数据。" if view["human_status"] == "处理中"
        else "最新运行需处理；页面仅保留上一份已验证资金数据。"
    )
    return view


def _daily_funds_cashflow_observation_view() -> dict[str, Any]:
    """Read the independent, chart-only receipt/payment projection safely.

    It is never a fallback for ``current.json``.  The point series is exposed
    only after every current eligible screenshot passed its own footer
    reconciliation, and the response deliberately excludes raw source IDs,
    attachment metadata, banks, counterparties and parser internals.
    """

    unavailable = {
        "status": "NOT_AVAILABLE",
        "message": "尚未形成已采集收支流水观察。",
        "generated_at": None,
        "evidence_version": None,
        "source_coverage": {
            "eligible_documents": 0,
            "parsed_documents": 0,
            "rejected_documents": 0,
            "distinct_business_days": 0,
        },
        "rejection_categories": {},
        "points": [],
    }
    needs_review = {
        **unavailable,
        "status": "NEEDS_REVIEW",
        "message": "收支截图尚未完整通过逐行与合计复核；不显示金额。",
    }
    payload = _read_daily_funds_json("cashflow_observation.json")
    if not isinstance(payload, dict) or set(payload) != DAILY_FUNDS_CASHFLOW_OBSERVATION_FIELDS:
        return unavailable if payload is None else needs_review
    if payload.get("schema_version") != DAILY_FUNDS_CASHFLOW_OBSERVATION_SCHEMA:
        return needs_review
    status = payload.get("status")
    generated_at = _daily_funds_timestamp(payload.get("generated_at"))
    evidence_version = payload.get("evidence_version")
    coverage = payload.get("source_coverage")
    rejection_categories = payload.get("rejection_categories")
    if (
        status not in DAILY_FUNDS_CASHFLOW_OBSERVATION_STATUSES
        or generated_at is None
        or not isinstance(evidence_version, (str, type(None)))
        or (isinstance(evidence_version, str) and not _daily_funds_lower_hex(evidence_version, 12))
        or not isinstance(coverage, dict)
        or set(coverage) != {"eligible_documents", "parsed_documents", "rejected_documents", "distinct_business_days"}
        or not all(_daily_funds_is_integer(coverage.get(key)) and 0 <= coverage[key] <= 100_000 for key in coverage)
        or coverage["parsed_documents"] + coverage["rejected_documents"] != coverage["eligible_documents"]
        or not isinstance(rejection_categories, dict)
        or not set(rejection_categories) <= set(DAILY_FUNDS_CASHFLOW_REJECTION_CATEGORY_LABELS)
        or not all(_daily_funds_is_integer(value) and value > 0 for value in rejection_categories.values())
        or sum(rejection_categories.values()) != coverage["rejected_documents"]
        or not isinstance(payload.get("parser_version"), str)
        or not payload["parser_version"].startswith("kmfa.daily_funds.cashflow_observation.")
        or not isinstance(payload.get("machine_code"), str)
        or not payload["machine_code"].startswith("CASHFLOW_OBSERVATION_")
        or not isinstance(payload.get("points"), list)
    ):
        return needs_review
    public_base = {
        "status": status,
        "generated_at": generated_at,
        "evidence_version": evidence_version,
        "source_coverage": dict(coverage),
        "rejection_categories": {
            DAILY_FUNDS_CASHFLOW_REJECTION_CATEGORY_LABELS[key]: value
            for key, value in rejection_categories.items()
        },
        "points": [],
    }
    if status != "VERIFIED":
        return {
            **public_base,
            "message": "收支截图尚未完整通过逐行与合计复核；不显示金额。"
            if status == "NEEDS_REVIEW" else "尚未形成已采集收支流水观察。",
        }
    if (
        payload["machine_code"] != "CASHFLOW_OBSERVATION_VERIFIED"
        or coverage["eligible_documents"] < 1
        or coverage["parsed_documents"] != coverage["eligible_documents"]
        or coverage["rejected_documents"] != 0
        or coverage["distinct_business_days"] < 2
        or rejection_categories
        or len(payload["points"]) != coverage["distinct_business_days"]
        or len(payload["points"]) > 366
    ):
        return needs_review
    points: list[dict[str, Any]] = []
    prior: date | None = None
    for row in payload["points"]:
        if not isinstance(row, dict) or set(row) != {
            "business_date", "inflow_fen", "outflow_fen", "net_change_fen",
        }:
            return needs_review
        business_date = _daily_funds_date(row.get("business_date"))
        inflow = row.get("inflow_fen")
        outflow = row.get("outflow_fen")
        net = row.get("net_change_fen")
        if (
            business_date is None
            or prior is not None and business_date <= prior
            or not all(_daily_funds_is_integer(value) for value in (inflow, outflow, net))
            or inflow < 0
            or outflow < 0
            or net != inflow - outflow
        ):
            return needs_review
        prior = business_date
        points.append({
            "business_date": business_date.isoformat(),
            "inflow_fen": inflow,
            "outflow_fen": outflow,
            "net_change_fen": net,
        })
    return {
        **public_base,
        "points": points,
        "message": "已按截图逐行与合计复核的收支流水；它不代表可用资金或账户余额。",
    }


def _daily_funds_cashflow_observation_range(
    *,
    range_value: str,
    from_date: str | None,
    to_date: str | None,
) -> dict[str, Any]:
    view = _daily_funds_cashflow_observation_view()
    if view["status"] != "VERIFIED":
        # Preserve range validation semantics even while no money is exposed.
        if range_value not in set(DAILY_FUNDS_ALLOWED_RANGES) | {"custom"}:
            raise HTTPException(status_code=422, detail="daily_funds_range_invalid")
        if range_value == "custom":
            start = _daily_funds_date(from_date)
            end = _daily_funds_date(to_date)
            if start is None or end is None or end < start:
                raise HTTPException(status_code=422, detail="daily_funds_custom_range_invalid")
            return {"range": range_value, "from": start.isoformat(), "to": end.isoformat(), **view}
        return {"range": range_value, "from": None, "to": None, **view}
    latest = _daily_funds_date(view["points"][-1]["business_date"])
    if latest is None:
        return {
            "range": range_value,
            "from": None,
            "to": None,
            "status": "NEEDS_REVIEW",
            "message": "收支截图尚未完整通过逐行与合计复核；不显示金额。",
            "generated_at": view["generated_at"],
            "evidence_version": view["evidence_version"],
            "source_coverage": view["source_coverage"],
            "rejection_categories": view["rejection_categories"],
            "points": [],
        }
    start, end = _daily_funds_range_days(range_value, from_date, to_date, publication_day=latest)
    return {
        "range": range_value,
        "from": start,
        "to": end,
        **view,
        "points": [
            point for point in view["points"]
            if start <= point["business_date"] <= end
        ],
    }


@app.get("/ops/api/daily-funds/summary")
@app.get("/api/daily-funds/summary")
def daily_funds_summary(range: str = "30d", from_: str | None = Query(None, alias="from"), to: str | None = None, scope: str = "global"):
    if scope != "global":
        raise HTTPException(status_code=422, detail="daily_funds_scope_invalid")
    try:
        payload = _daily_funds_current()
    except HTTPException as exc:
        if exc.status_code == 503 and exc.detail == "daily_funds_projection_unavailable":
            return _daily_funds_unpublished_summary(
                range_value=range,
                from_date=from_,
                to_date=to,
            )
        raise
    projection = _daily_funds_projection_view(payload)
    start, end = _daily_funds_range_days(range, from_, to, publication_day=projection["publication_day"])
    points = [point for point in projection["points"] if start <= point["business_date"] <= end]
    activity = _daily_funds_activity(projection["transactions"], publication_day=projection["publication_day"])
    publication = payload["publication"]
    return {
        "range": range,
        "from": start,
        "to": end,
        "scope": "global",
        "granularity": "daily",
        "range_health": _daily_funds_range_health(points, start, end),
        "publication": {
            "business_date": publication["business_date"],
            "created_at": publication["created_at"],
            "evidence_version": str(publication["publication_id"])[-12:],
            "reconciliation_difference_fen": 0,
            "oci_backup_state": projection["backup_state"],
            "git_evidence_available": True,
            "r2_mirror_available": True,
        },
        "total_available_fen": projection["total_available_fen"],
        "risk_label": projection["risk_label"],
        "dynamic_flag": projection["dynamic_flag"],
        "by_company_ending_fen": projection["by_company_ending_fen"],
        "by_bank_ending_fen": projection["by_bank_ending_fen"],
        "account_breakdown": projection["account_breakdown"],
        **activity,
        "points": points,
    }


@app.get("/ops/api/daily-funds/timeseries")
@app.get("/api/daily-funds/timeseries")
def daily_funds_timeseries(range: str = "30d", from_: str | None = Query(None, alias="from"), to: str | None = None):
    try:
        payload = _daily_funds_current()
    except HTTPException as exc:
        if exc.status_code == 503 and exc.detail == "daily_funds_projection_unavailable":
            return _daily_funds_unpublished_timeseries(
                range_value=range,
                from_date=from_,
                to_date=to,
            )
        raise
    projection = _daily_funds_projection_view(payload)
    start, end = _daily_funds_range_days(range, from_, to, publication_day=projection["publication_day"])
    points = [point for point in projection["points"] if start <= point["business_date"] <= end]
    return {
        "range": range,
        "from": start,
        "to": end,
        "granularity": "daily",
        "range_health": _daily_funds_range_health(points, start, end),
        "points": points,
        "thresholds": projection["thresholds"],
    }


@app.get("/ops/api/daily-funds/transactions")
@app.get("/api/daily-funds/transactions")
def daily_funds_transactions(page: int = 1, size: int = 100):
    payload = _daily_funds_current()
    projection = _daily_funds_projection_view(payload)
    items, meta = _paginate(projection["transactions"], page, size)
    return {"items": items, "pagination": meta}


@app.get("/ops/api/daily-funds/source-health")
@app.get("/api/daily-funds/source-health")
def daily_funds_source_health():
    return _daily_funds_source_health_view()


@app.get("/ops/api/daily-funds/cashflow-observations")
@app.get("/api/daily-funds/cashflow-observations")
def daily_funds_cashflow_observations(
    range: str = "30d",
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
):
    return _daily_funds_cashflow_observation_range(
        range_value=range,
        from_date=from_,
        to_date=to,
    )


def _daily_funds_auth_timestamp(value: object) -> datetime | None:
    if not isinstance(value, str) or len(value) > 40:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed if parsed.tzinfo is not None else None


def _daily_funds_auth_iso(value: datetime) -> str:
    return value.astimezone(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _daily_funds_auth_code(value: object) -> str:
    token = "".join(
        character
        for character in str(value or "UNKNOWN").strip().upper()
        if character.isascii() and (character.isupper() or character.isdigit() or character == "_")
    )
    return token[:80] or "UNKNOWN"


def _daily_funds_auth_url(value: object) -> str | None:
    if not isinstance(value, str) or not value or len(value) > 2_048:
        return None
    try:
        parsed = urlsplit(value)
        hostname = (parsed.hostname or "").lower()
    except ValueError:
        return None
    if (
        parsed.scheme != "https"
        or not hostname
        or parsed.username is not None
        or parsed.password is not None
        or not (hostname == "dingtalk.com" or hostname.endswith(".dingtalk.com")
                or hostname == "dingtalk.cn" or hostname.endswith(".dingtalk.cn"))
    ):
        return None
    return value


def _daily_funds_auth_control_path(name: str) -> Path:
    if name not in {DAILY_FUNDS_AUTH_REQUEST_FILE, DAILY_FUNDS_AUTH_SESSION_FILE}:
        raise ValueError("daily funds auth control filename invalid")
    root = DAILY_FUNDS_CONTROL_DIR.resolve()
    target = (root / name).resolve()
    if target.parent != root:
        raise ValueError("daily funds auth control path invalid")
    return target


def _daily_funds_auth_read_object(name: str) -> dict[str, Any] | None:
    try:
        target = _daily_funds_auth_control_path(name)
    except (OSError, ValueError):
        return None
    if target.is_symlink() or not target.is_file():
        return None
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return payload if isinstance(payload, dict) else None


def _daily_funds_auth_read_request(now: datetime) -> dict[str, Any] | None:
    payload = _daily_funds_auth_read_object(DAILY_FUNDS_AUTH_REQUEST_FILE)
    if not isinstance(payload, dict) or set(payload) != {
        "schema_version", "request_id", "action", "actor", "requested_at", "expires_at",
    }:
        return None
    request_id = payload.get("request_id")
    action = payload.get("action")
    requested_at = _daily_funds_auth_timestamp(payload.get("requested_at"))
    expires_at = _daily_funds_auth_timestamp(payload.get("expires_at"))
    if (
        payload.get("schema_version") != DAILY_FUNDS_AUTH_REQUEST_SCHEMA
        or not isinstance(request_id, str)
        or re.fullmatch(r"[0-9a-f]{64}", request_id) is None
        or action not in {"START", "CANCEL"}
        or payload.get("actor") != DAILY_FUNDS_AUTH_ACTOR
        or requested_at is None
        or expires_at is None
        or expires_at <= requested_at
        or (expires_at - requested_at).total_seconds() > 660
    ):
        return None
    return {
        "request_id": request_id,
        "action": action,
        "requested_at": requested_at,
        "expires_at": expires_at,
        "expired": expires_at.astimezone(timezone.utc) <= now.astimezone(timezone.utc),
    }


def _daily_funds_auth_read_session(now: datetime) -> dict[str, Any] | None:
    payload = _daily_funds_auth_read_object(DAILY_FUNDS_AUTH_SESSION_FILE)
    if not isinstance(payload, dict) or set(payload) != {
        "schema_version", "request_id", "state", "machine_code", "created_at", "updated_at", "expires_at",
        "authorization_url", "user_code",
    }:
        return None
    request_id = payload.get("request_id")
    state = payload.get("state")
    created_at = _daily_funds_auth_timestamp(payload.get("created_at"))
    updated_at = _daily_funds_auth_timestamp(payload.get("updated_at"))
    expires_at = _daily_funds_auth_timestamp(payload.get("expires_at"))
    if (
        payload.get("schema_version") != DAILY_FUNDS_AUTH_SESSION_SCHEMA
        or not isinstance(request_id, str)
        or re.fullmatch(r"[0-9a-f]{64}", request_id) is None
        or state not in DAILY_FUNDS_AUTH_LIVE_STATES | DAILY_FUNDS_AUTH_TERMINAL_STATES
        or created_at is None
        or updated_at is None
        or expires_at is None
    ):
        return None
    authorization_url = _daily_funds_auth_url(payload.get("authorization_url"))
    user_code = payload.get("user_code")
    if state == "AWAITING_APPROVAL":
        if authorization_url is None or not isinstance(user_code, str) or DAILY_FUNDS_AUTH_CODE_RE.fullmatch(user_code) is None:
            return None
    elif authorization_url is not None or user_code is not None:
        return None
    if state in DAILY_FUNDS_AUTH_LIVE_STATES and expires_at.astimezone(timezone.utc) <= now.astimezone(timezone.utc):
        return {
            "state": "EXPIRED",
            "machine_code": "DWS_AUTH_BOOTSTRAP_EXPIRED",
            "updated_at": _daily_funds_auth_iso(now),
            "expires_at": _daily_funds_auth_iso(expires_at),
            "authorization_url": None,
            "user_code": None,
        }
    return {
        "state": state,
        "machine_code": _daily_funds_auth_code(payload.get("machine_code")),
        "updated_at": _daily_funds_auth_iso(updated_at),
        "expires_at": _daily_funds_auth_iso(expires_at),
        "authorization_url": authorization_url if state == "AWAITING_APPROVAL" else None,
        "user_code": user_code if state == "AWAITING_APPROVAL" else None,
    }


def _daily_funds_auth_write_request(payload: dict[str, Any]) -> None:
    DAILY_FUNDS_CONTROL_DIR.mkdir(parents=True, exist_ok=True)
    target = _daily_funds_auth_control_path(DAILY_FUNDS_AUTH_REQUEST_FILE)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=DAILY_FUNDS_CONTROL_DIR, delete=False) as handle:
        json.dump(payload, handle, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        handle.write("\n")
        temporary = Path(handle.name)
    try:
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()


def _daily_funds_auth_same_origin(request: Request) -> bool:
    origin = request.headers.get("origin", "").strip()
    host = request.headers.get("host", "").strip().lower()
    if not origin or not host:
        return False
    try:
        parsed = urlsplit(origin)
    except ValueError:
        return False
    return parsed.scheme in {"https", "http"} and parsed.netloc.lower() == host


def _daily_funds_auth_response(payload: dict[str, Any], *, status_code: int = 200) -> JSONResponse:
    return JSONResponse(
        payload,
        status_code=status_code,
        headers={
            "Cache-Control": "private, no-store",
            "Pragma": "no-cache",
            "X-Robots-Tag": "noindex, nofollow, noarchive",
        },
    )


def _daily_funds_history_probe_response(payload: dict[str, Any], *, status_code: int = 200) -> JSONResponse:
    """Mark a values-free history-probe reply as originating from this app.

    The Access bridge retains response bodies only in a mode-0700 temporary
    directory, but a generic edge 503 has no reliable body contract.  This
    fixed marker lets it distinguish a deliberate app control-volume failure
    from an upstream/edge failure without disclosing a path, request ID,
    source value, or credential.
    """

    response = _daily_funds_auth_response(payload, status_code=status_code)
    response.headers[DAILY_FUNDS_HISTORY_PROBE_ENDPOINT_HEADER] = DAILY_FUNDS_HISTORY_PROBE_ENDPOINT_VALUE
    return response


def _daily_funds_history_probe_error(*, status_code: int, detail: str) -> JSONResponse:
    """Return a fixed, non-cacheable error from the narrow history-probe API."""

    response = _daily_funds_auth_response({"detail": detail}, status_code=status_code)
    response.headers[DAILY_FUNDS_HISTORY_PROBE_ENDPOINT_HEADER] = DAILY_FUNDS_HISTORY_PROBE_ENDPOINT_VALUE
    return response


@app.get("/ops/api/daily-funds/auth-session")
def daily_funds_auth_session():
    """Read only the short-lived owner-facing device authorization state."""

    now = datetime.now(timezone.utc)
    # A cancellation request must win over a stale AWAITING_APPROVAL record.
    # The broker sees the shared-volume request asynchronously; returning the
    # old prompt in that small interval would keep a device code visible after
    # the owner explicitly chose to revoke it.
    request = _daily_funds_auth_read_request(now)
    if request is not None and not request["expired"] and request["action"] == "CANCEL":
        return _daily_funds_auth_response({
            "state": "CANCELLING",
            "machine_code": "DWS_AUTH_BOOTSTRAP_CANCELLING",
            "updated_at": _daily_funds_auth_iso(now),
            "expires_at": _daily_funds_auth_iso(request["expires_at"]),
            "authorization_url": None,
            "user_code": None,
        })
    session = _daily_funds_auth_read_session(now)
    if session is not None:
        return _daily_funds_auth_response(session)
    if request is not None and not request["expired"]:
        return _daily_funds_auth_response({
            "state": "REQUESTED",
            "machine_code": "DWS_AUTH_BOOTSTRAP_STARTING",
            "updated_at": _daily_funds_auth_iso(now),
            "expires_at": _daily_funds_auth_iso(request["expires_at"]),
            "authorization_url": None,
            "user_code": None,
        })
    return _daily_funds_auth_response({
        "state": "NOT_REQUESTED",
        "machine_code": "DWS_BOOTSTRAP_REQUIRED",
        "updated_at": _daily_funds_auth_iso(now),
        "expires_at": None,
        "authorization_url": None,
        "user_code": None,
    })


@app.post("/ops/api/daily-funds/auth-session")
def start_daily_funds_auth_session(request: Request):
    """Queue exactly one Access-gated DWS device authorization; never a shell command."""

    if not _daily_funds_auth_same_origin(request):
        raise HTTPException(status_code=403, detail="daily_funds_auth_same_origin_required")
    now = datetime.now(timezone.utc)
    existing = _daily_funds_auth_read_session(now)
    pending = _daily_funds_auth_read_request(now)
    if (
        (existing is not None and existing["state"] in DAILY_FUNDS_AUTH_LIVE_STATES)
        or (pending is not None and not pending["expired"])
    ):
        raise HTTPException(status_code=409, detail="daily_funds_auth_already_pending")
    expires_at = now + timedelta(minutes=10)
    payload = {
        "schema_version": DAILY_FUNDS_AUTH_REQUEST_SCHEMA,
        "request_id": secrets.token_hex(32),
        "action": "START",
        "actor": DAILY_FUNDS_AUTH_ACTOR,
        "requested_at": _daily_funds_auth_iso(now),
        "expires_at": _daily_funds_auth_iso(expires_at),
    }
    try:
        _daily_funds_auth_write_request(payload)
    except OSError as exc:
        raise HTTPException(status_code=503, detail="daily_funds_auth_control_unavailable") from exc
    return _daily_funds_auth_response({
        "state": "REQUESTED",
        "machine_code": "DWS_AUTH_BOOTSTRAP_STARTING",
        "updated_at": _daily_funds_auth_iso(now),
        "expires_at": _daily_funds_auth_iso(expires_at),
        "authorization_url": None,
        "user_code": None,
    }, status_code=202)


@app.delete("/ops/api/daily-funds/auth-session")
def cancel_daily_funds_auth_session(request: Request):
    """Cancel the one permitted device authorization and wipe its UI prompt."""

    if not _daily_funds_auth_same_origin(request):
        raise HTTPException(status_code=403, detail="daily_funds_auth_same_origin_required")
    now = datetime.now(timezone.utc)
    session = _daily_funds_auth_read_session(now)
    pending = _daily_funds_auth_read_request(now)
    request_id: str | None = None
    if session is not None and session["state"] in DAILY_FUNDS_AUTH_LIVE_STATES:
        raw = _daily_funds_auth_read_object(DAILY_FUNDS_AUTH_SESSION_FILE) or {}
        candidate = raw.get("request_id")
        if isinstance(candidate, str) and re.fullmatch(r"[0-9a-f]{64}", candidate):
            request_id = candidate
    if request_id is None and pending is not None and not pending["expired"]:
        request_id = str(pending["request_id"])
    if request_id is None:
        return _daily_funds_auth_response({
            "state": "NOT_REQUESTED",
            "machine_code": "DWS_BOOTSTRAP_REQUIRED",
            "updated_at": _daily_funds_auth_iso(now),
            "expires_at": None,
            "authorization_url": None,
            "user_code": None,
        })
    expires_at = now + timedelta(minutes=2)
    payload = {
        "schema_version": DAILY_FUNDS_AUTH_REQUEST_SCHEMA,
        "request_id": request_id,
        "action": "CANCEL",
        "actor": DAILY_FUNDS_AUTH_ACTOR,
        "requested_at": _daily_funds_auth_iso(now),
        "expires_at": _daily_funds_auth_iso(expires_at),
    }
    try:
        _daily_funds_auth_write_request(payload)
    except OSError as exc:
        raise HTTPException(status_code=503, detail="daily_funds_auth_control_unavailable") from exc
    return _daily_funds_auth_response({
        "state": "CANCELLING",
        "machine_code": "DWS_AUTH_BOOTSTRAP_CANCELLING",
        "updated_at": _daily_funds_auth_iso(now),
        "expires_at": _daily_funds_auth_iso(expires_at),
        "authorization_url": None,
        "user_code": None,
    }, status_code=202)


def _daily_funds_history_probe_control_path(name: str) -> Path:
    if name not in {DAILY_FUNDS_HISTORY_PROBE_REQUEST_FILE, DAILY_FUNDS_HISTORY_PROBE_SESSION_FILE}:
        raise ValueError("daily funds history probe control filename invalid")
    root = DAILY_FUNDS_CONTROL_DIR.resolve()
    target = (root / name).resolve()
    if target.parent != root:
        raise ValueError("daily funds history probe control path invalid")
    return target


def _daily_funds_history_probe_read_object(name: str) -> dict[str, Any] | None:
    try:
        target = _daily_funds_history_probe_control_path(name)
    except (OSError, ValueError):
        return None
    if target.is_symlink() or not target.is_file():
        return None
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return payload if isinstance(payload, dict) else None


def _daily_funds_history_probe_read_request(now: datetime) -> dict[str, Any] | None:
    payload = _daily_funds_history_probe_read_object(DAILY_FUNDS_HISTORY_PROBE_REQUEST_FILE)
    if not isinstance(payload, dict) or set(payload) != {
        "schema_version", "request_id", "action", "actor", "requested_at", "expires_at",
    }:
        return None
    request_id = payload.get("request_id")
    requested_at = _daily_funds_auth_timestamp(payload.get("requested_at"))
    expires_at = _daily_funds_auth_timestamp(payload.get("expires_at"))
    if (
        payload.get("schema_version") != DAILY_FUNDS_HISTORY_PROBE_REQUEST_SCHEMA
        or not isinstance(request_id, str)
        or re.fullmatch(r"[0-9a-f]{64}", request_id) is None
        or payload.get("action") != "PROBE"
        or payload.get("actor") != DAILY_FUNDS_HISTORY_PROBE_ACTOR
        or requested_at is None
        or expires_at is None
        or requested_at > now + timedelta(minutes=2)
        or requested_at < now - timedelta(hours=1)
        or expires_at <= requested_at
        or (expires_at - requested_at).total_seconds() > 660
    ):
        return None
    return {
        "request_id": request_id,
        "requested_at": requested_at,
        "expires_at": expires_at,
        "expired": expires_at.astimezone(timezone.utc) <= now.astimezone(timezone.utc),
    }


def _daily_funds_history_probe_read_session(now: datetime) -> dict[str, Any] | None:
    payload = _daily_funds_history_probe_read_object(DAILY_FUNDS_HISTORY_PROBE_SESSION_FILE)
    session_keys = {
        "schema_version", "request_id", "state", "machine_code", "created_at", "updated_at", "expires_at",
        "continuation_state", "cursor_transcript", "record_list_shape",
    }
    legacy_session_keys = session_keys - {"record_list_shape"}
    if not isinstance(payload, dict):
        return None
    payload_keys = set(payload)
    if payload_keys != session_keys and payload_keys != legacy_session_keys:
        return None
    schema_version = payload.get("schema_version")
    if schema_version == DAILY_FUNDS_HISTORY_PROBE_SESSION_SCHEMA:
        if set(payload) != session_keys:
            return None
        record_list_shape = payload.get("record_list_shape")
    elif schema_version == DAILY_FUNDS_HISTORY_PROBE_LEGACY_SESSION_SCHEMA:
        if set(payload) != legacy_session_keys:
            return None
        record_list_shape = "NOT_OBSERVED"
    else:
        return None
    request_id = payload.get("request_id")
    state = payload.get("state")
    created_at = _daily_funds_auth_timestamp(payload.get("created_at"))
    updated_at = _daily_funds_auth_timestamp(payload.get("updated_at"))
    expires_at = _daily_funds_auth_timestamp(payload.get("expires_at"))
    continuation_state = payload.get("continuation_state")
    cursor_transcript = payload.get("cursor_transcript")
    if (
        not isinstance(request_id, str)
        or re.fullmatch(r"[0-9a-f]{64}", request_id) is None
        or state not in DAILY_FUNDS_HISTORY_PROBE_LIVE_STATES | DAILY_FUNDS_HISTORY_PROBE_TERMINAL_STATES
        or created_at is None
        or updated_at is None
        or expires_at is None
        or expires_at <= created_at
        or continuation_state not in DAILY_FUNDS_HISTORY_PROBE_CONTINUATION_STATES
        or cursor_transcript != DAILY_FUNDS_HISTORY_PROBE_CURSOR_TRANSCRIPTS.get(continuation_state)
        or record_list_shape not in DAILY_FUNDS_HISTORY_PROBE_RECORD_LIST_SHAPES
    ):
        return None
    if state in DAILY_FUNDS_HISTORY_PROBE_LIVE_STATES and expires_at.astimezone(timezone.utc) <= now.astimezone(timezone.utc):
        return {
            "state": "EXPIRED",
            "machine_code": "DWS_HISTORY_PROBE_EXPIRED",
            "updated_at": _daily_funds_auth_iso(now),
            "expires_at": _daily_funds_auth_iso(expires_at),
            "continuation_state": "NOT_STARTED",
            "cursor_transcript": "NOT_STARTED",
            "record_list_shape": "NOT_OBSERVED",
        }
    return {
        "state": state,
        "machine_code": _daily_funds_auth_code(payload.get("machine_code")),
        "updated_at": _daily_funds_auth_iso(updated_at),
        "expires_at": _daily_funds_auth_iso(expires_at),
        "continuation_state": continuation_state,
        "cursor_transcript": cursor_transcript,
        "record_list_shape": record_list_shape,
    }


def _daily_funds_history_probe_write_request(payload: dict[str, Any]) -> None:
    DAILY_FUNDS_CONTROL_DIR.mkdir(parents=True, exist_ok=True)
    target = _daily_funds_history_probe_control_path(DAILY_FUNDS_HISTORY_PROBE_REQUEST_FILE)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=DAILY_FUNDS_CONTROL_DIR, delete=False) as handle:
        json.dump(payload, handle, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        handle.write("\n")
        temporary = Path(handle.name)
    try:
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()


@app.get("/ops/api/daily-funds/history-probe")
def daily_funds_history_probe():
    """Return only the latest fixed history-probe state, never source values."""

    now = datetime.now(timezone.utc)
    session = _daily_funds_history_probe_read_session(now)
    if session is not None:
        return _daily_funds_history_probe_response(session)
    request = _daily_funds_history_probe_read_request(now)
    if request is not None and not request["expired"]:
        return _daily_funds_history_probe_response({
            "state": "REQUESTED",
            "machine_code": "DWS_HISTORY_PROBE_QUEUED",
            "updated_at": _daily_funds_auth_iso(now),
            "expires_at": _daily_funds_auth_iso(request["expires_at"]),
            "continuation_state": "NOT_STARTED",
            "cursor_transcript": "NOT_STARTED",
            "record_list_shape": "NOT_OBSERVED",
        })
    return _daily_funds_history_probe_response({
        "state": "NOT_REQUESTED",
        "machine_code": "DWS_HISTORY_PROBE_NOT_REQUESTED",
        "updated_at": _daily_funds_auth_iso(now),
        "expires_at": None,
        "continuation_state": "NOT_STARTED",
        "cursor_transcript": "NOT_STARTED",
        "record_list_shape": "NOT_OBSERVED",
    })


@app.post("/ops/api/daily-funds/history-probe")
async def start_daily_funds_history_probe(request: Request):
    """Queue one pre-defined cloud history probe; this endpoint accepts no input."""

    if not _daily_funds_auth_same_origin(request):
        return _daily_funds_history_probe_error(
            status_code=403,
            detail="daily_funds_history_probe_same_origin_required",
        )
    # Reject payload-bearing requests without reading their body.  This keeps
    # the route from becoming an accidental transport for a command, group ID,
    # source payload, or financial value.  The sole allowed action is encoded
    # below in the exact shared-volume schema.
    if request.headers.get("content-length", "").strip() not in {"", "0"} or request.headers.get("transfer-encoding", "").strip():
        return _daily_funds_history_probe_error(
            status_code=422,
            detail="daily_funds_history_probe_body_forbidden",
        )
    now = datetime.now(timezone.utc)
    existing = _daily_funds_history_probe_read_session(now)
    pending = _daily_funds_history_probe_read_request(now)
    if (
        (existing is not None and existing["state"] in DAILY_FUNDS_HISTORY_PROBE_LIVE_STATES)
        or (pending is not None and not pending["expired"])
    ):
        return _daily_funds_history_probe_error(
            status_code=409,
            detail="daily_funds_history_probe_already_pending",
        )
    expires_at = now + timedelta(minutes=10)
    payload = {
        "schema_version": DAILY_FUNDS_HISTORY_PROBE_REQUEST_SCHEMA,
        "request_id": secrets.token_hex(32),
        "action": "PROBE",
        "actor": DAILY_FUNDS_HISTORY_PROBE_ACTOR,
        "requested_at": _daily_funds_auth_iso(now),
        "expires_at": _daily_funds_auth_iso(expires_at),
    }
    try:
        _daily_funds_history_probe_write_request(payload)
    except OSError:
        return _daily_funds_history_probe_error(
            status_code=503,
            detail="daily_funds_history_probe_control_unavailable",
        )
    return _daily_funds_history_probe_response({
        "state": "REQUESTED",
        "machine_code": "DWS_HISTORY_PROBE_QUEUED",
        "updated_at": _daily_funds_auth_iso(now),
        "expires_at": _daily_funds_auth_iso(expires_at),
        "continuation_state": "NOT_STARTED",
        "cursor_transcript": "NOT_STARTED",
        "record_list_shape": "NOT_OBSERVED",
    }, status_code=202)


def _read_daily_funds_control() -> dict[str, Any] | None:
    target = (DAILY_FUNDS_CONTROL_DIR / "active_threshold.json").resolve()
    root = DAILY_FUNDS_CONTROL_DIR.resolve()
    if not str(target).startswith(str(root) + "/") or not target.is_file():
        return None
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return payload if isinstance(payload, dict) else None


def _daily_funds_control_audit() -> dict[str, Any]:
    """Expose a tiny, redacted audit projection for the Owner-only page.

    Audit corruption is never converted into a healthy result.  It also must
    not make a valid financial publication disappear, so this secondary view
    explicitly reports itself unavailable instead of forwarding raw JSONL.
    """
    target = (DAILY_FUNDS_CONTROL_DIR / "threshold_audit.jsonl").resolve()
    root = DAILY_FUNDS_CONTROL_DIR.resolve()
    if not str(target).startswith(str(root) + "/") or not target.is_file():
        return {"available": False, "entries": []}
    try:
        raw = target.read_bytes()
        if len(raw) > 65_536:
            raw = raw[-65_536:]
            if b"\n" not in raw:
                return {"available": False, "entries": []}
            raw = raw.split(b"\n", 1)[1]
        lines = raw.decode("utf-8").splitlines()
    except (OSError, UnicodeDecodeError):
        return {"available": False, "entries": []}
    entries: list[dict[str, Any]] = []
    for line in lines:
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            return {"available": False, "entries": []}
        if not isinstance(row, dict) or set(row) != {
            "schema_version", "revision", "actor", "changed_at", "old_value", "new_value", "reason", "rollback_version",
        }:
            return {"available": False, "entries": []}
        old_value = row.get("old_value")
        old = None if old_value is None else _daily_funds_public_control_value(old_value)
        new = _daily_funds_public_control_value(row.get("new_value"))
        rollback_version = row.get("rollback_version")
        if (
            row.get("schema_version") != "kmfa.daily_funds.threshold_audit.v1"
            or not _daily_funds_lower_hex(row.get("revision"), 64)
            or row.get("actor") != "kmfa_private_owner_ui"
            or _daily_funds_timestamp(row.get("changed_at")) is None
            or not isinstance(row.get("reason"), str)
            or len(row["reason"]) > 500
            or any(ord(character) < 32 for character in row["reason"])
            or new is None
            or str(row["revision"]) != new["revision"]
            or (old_value is not None and old is None)
            or (rollback_version is not None and not _daily_funds_lower_hex(rollback_version, 64))
        ):
            return {"available": False, "entries": []}
        entries.append({
            "revision": row["revision"],
            "actor": "kmfa_private_owner_ui",
            "changed_at": row["changed_at"],
            "old_value": old,
            "new_value": new,
            "reason": row["reason"],
            "rollback_version": rollback_version,
        })
    return {"available": True, "entries": entries[-10:]}


@app.get("/ops/api/daily-funds/thresholds")
@app.get("/api/daily-funds/thresholds")
def daily_funds_thresholds():
    try:
        payload = _daily_funds_current()
        projection = _daily_funds_projection_view(payload)
        thresholds = projection["thresholds"]
        data_available = True
    except HTTPException as exc:
        # A missing or invalid projection must still not conceal the two frozen
        # policy lines.  Return no account, transaction, risk or dynamic amount
        # here; the summary and timeseries endpoints remain fail-closed.
        if exc.status_code != 503 or exc.detail not in {
            "daily_funds_projection_unavailable",
            "daily_funds_projection_not_valid",
        }:
            raise
        thresholds = _daily_funds_unpublished_thresholds()
        data_available = False
    return {
        "active": thresholds,
        "control": _daily_funds_public_control(),
        "control_audit": _daily_funds_control_audit(),
        "fixed_editable": False,
        "data_available": data_available,
    }


@app.put("/ops/api/daily-funds/thresholds")
@app.put("/api/daily-funds/thresholds")
def update_daily_funds_thresholds(body: dict[str, Any] = Body(...)):
    """Queue a versioned custom-threshold request for the isolated worker.

    The app never edits a publication or D1 itself.  The worker validates and
    applies the request on its next controlled run, preserving the single
    compute authority and an append-only control audit.
    """
    mode = str(body.get("mode") or "")
    if mode not in {"disabled", "date_range", "numeric"}:
        raise HTTPException(status_code=422, detail="daily_funds_threshold_mode_invalid")
    reason = str(body.get("reason") or "").strip()
    if not reason:
        raise HTTPException(status_code=422, detail="daily_funds_threshold_reason_required")
    existing = _read_daily_funds_control()
    expected_revision = body.get("expected_revision")
    if existing is not None and expected_revision != existing.get("revision"):
        raise HTTPException(status_code=409, detail="daily_funds_threshold_revision_conflict")
    # The private Access boundary supplies the actor class.  Do not accept an
    # arbitrary client-supplied actor string, which would make the control
    # audit forgeable even though the endpoint itself is owner-protected.
    request: dict[str, Any] = {
        "mode": mode,
        "reason": reason[:500],
        "actor": "kmfa_private_owner_ui",
        "submitted_at": datetime.now(BEIJING).isoformat(),
    }
    if mode == "date_range":
        try:
            start = datetime.fromisoformat(str(body.get("from") or "")).date()
            end = datetime.fromisoformat(str(body.get("to") or "")).date()
        except ValueError as exc:
            raise HTTPException(status_code=422, detail="daily_funds_threshold_date_invalid") from exc
        if end < start or (end - start).days + 1 < 7:
            raise HTTPException(status_code=422, detail="daily_funds_threshold_date_invalid")
        request.update({"from": start.isoformat(), "to": end.isoformat()})
    if mode == "numeric":
        amount = body.get("amount_fen")
        if isinstance(amount, bool) or not isinstance(amount, int) or not 0 <= amount <= 999_999_999_999_999:
            raise HTTPException(status_code=422, detail="daily_funds_threshold_amount_invalid")
        request["amount_fen"] = amount
    request["scope"] = "global"
    request["revision"] = hashlib.sha256(json.dumps(request, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
    try:
        DAILY_FUNDS_CONTROL_DIR.mkdir(parents=True, exist_ok=True)
        target = DAILY_FUNDS_CONTROL_DIR / "threshold_request.json"
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=DAILY_FUNDS_CONTROL_DIR, delete=False) as handle:
            json.dump(request, handle, ensure_ascii=False, sort_keys=True)
            handle.write("\n")
            temporary = Path(handle.name)
        os.replace(temporary, target)
    except OSError as exc:
        raise HTTPException(status_code=503, detail="daily_funds_control_unavailable") from exc
    return {"accepted": True, "revision": request["revision"], "state": "pending_worker_validation"}


def _log_tail_line(log_path) -> str | None:
    """取当次运行日志的最后一行实质输出（跳过 run_skill 的「结束 rc=」包装行）。"""
    if not log_path:
        return None
    try:
        target = Path(str(log_path)).resolve()
        root = SKILL_LEDGER_PATH.parent.resolve()
        if not str(target).startswith(str(root) + "/") or not target.is_file():
            return None
        lines = [l.strip() for l in target.read_bytes()[-8192:].decode("utf-8", "replace").splitlines() if l.strip()]
        for l in reversed(lines):
            if ": 结束 rc=" in l or ": 开始" in l:
                continue
            return l[:160]
        return lines[-1][:160] if lines else None
    except OSError:
        return None


@app.get("/public-api/登录入口")
def public_login_entry(request: Request):
    """给未登录的人一个**能用**的登录入口。**本身必须是公开的** ——
    一个需要登录才能拿到的登录入口没有意义。

    为什么不直接在前端写死一个链接：
      · 写死 team domain，它一变链接就断，而断了的表现是「点了没反应」；
      · 更要紧的是 **redirect**。之前前端指向 `/api/状态` 触发登录，
        Cloudflare 就把用户在登录后送回 `/api/状态` —— 那里返回一坨 JSON。
        人看到那个，得到的结论是「登录了但没用」，而不是「登录成功了」。
        所以这里显式把落点定成 `/`，登录完直接回驾驶舱。

    没配 team domain 时**如实说没有**，不编一个链接出来：
    一个点了没反应的按钮比没有按钮更难查。
    """
    try:
        settings = load_access_settings()
    except Exception as error:
        return {"可用": False,
                "原因": f"未配置 Cloudflare Access team domain（{type(error).__name__}）",
                "登录地址": None}
    host = (request.headers.get("host") or "").split(":")[0]
    try:
        return {"可用": True, "登录地址": settings.login_url(host, "/"),
                "落点": "/", "说明": "登录完直接回驾驶舱，不会停在 JSON 接口上。"}
    except Exception as error:
        return {"可用": False, "原因": str(error), "登录地址": None}


#: 逐目标结果里表示「这份报告发没发成」的键。产线上有两种形状：
#:   · `_target_send_result()` 出的是 `management_status` / `hr_status`（按报告分开）；
#:   · 解析目标那一段出的是单个 `status`。
#: 2026-07-28 实测踩过：端点只读 `status`，而真实投递走的是前一种，于是**把成功的
#: 投递显示成失败**——顶层写着 SENT，目标行却是 成功=False。这正是本端点要消灭的
#: 那类误导信号，不能由端点自己制造。
_TARGET_STATUS_KEYS = ("management_status", "hr_status", "status")


def _dispatch_target(item: dict[str, Any]) -> dict[str, Any]:
    """把一条逐目标结果裁成公开安全的形状。

    成功的判据是**所有实际出现的报告状态都为 SENT**，而不是任一为 SENT：
    管理报表发了、HR 报表没发，是「没发全」，不能算成功。
    `SKIPPED` 表示这份报告本就不在该目标的订阅里，不参与判定。
    """
    statuses = {key: str(item.get(key) or "") for key in _TARGET_STATUS_KEYS
                if item.get(key) not in (None, "")}
    effective = [v for v in statuses.values() if v != "SKIPPED"]
    return {
        "对象": str(item.get("label") or "?"),
        # `SENT_UNVERIFIED` 明确**不算成功**：它表示 dws 命令跑完了，但返回体里
        # 没有钉钉侧的投递凭据。把它算成功，就是这次「绿了一个月、一条没收到」
        # 事故的复现。
        "成功": bool(effective) and all(v == "SENT" for v in effective),
        "查无投递凭据": any(v == "SENT_UNVERIFIED" for v in effective),
        "各报告状态": statuses or {"（回执里没有状态字段）": ""},
        "通道": str(item.get("channel") or item.get("resolved_channel") or ""),
        "失败原因": str(item.get("failure_reason") or "") or None,
        # trace_id 本体不出（是钉钉侧的追踪标识），只出「有没有」——
        # 没有 trace 的「成功」值得怀疑，这一位就够判断了。
        "有回执追踪号": bool(item.get("trace_id_present") or item.get("trace_id")),
    }


def _dispatch_receipts(limit: int = 8) -> list[dict[str, Any]]:
    """把最近若干份投递回执裁成公开安全的形状。

    只出白名单字段，外加每个目标的「谁 / 成没成 / 走的哪条通道」。
    **不出**报表正文、模板文本、user_id——那是全员考勤数据和员工标识。
    """
    if not ATTENDANCE_ARCHIVE_ROOT.is_dir():
        return []
    files = sorted(ATTENDANCE_ARCHIVE_ROOT.glob("*/*.dispatch.json"),
                   key=lambda p: p.stat().st_mtime, reverse=True)[:limit]
    out = []
    for path in files:
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            out.append({"回执": path.name, "可读": False,
                        "原因": f"{type(exc).__name__}"})
            continue
        if not isinstance(raw, dict):
            out.append({"回执": path.name, "可读": False, "原因": "回执不是对象"})
            continue
        row = {k: raw.get(k) for k in _DISPATCH_PUBLIC_FIELDS if raw.get(k) not in (None, "")}
        targets = [_dispatch_target(item) for item in (raw.get("target_results") or [])
                   if isinstance(item, dict)]
        row["目标"] = targets
        row["回执"] = path.name
        row["写入时间"] = datetime.fromtimestamp(path.stat().st_mtime, BEIJING).isoformat()
        out.append(row)
    return out


@app.get("/public-api/考勤投递")
def public_attendance_dispatch():
    """考勤到底发出去没有——**不需要登录**。

    Owner 2026-07-28：「考勤我没有收到」。

    这个端点存在的理由和 `技能健康` 一模一样，是同一个死结的下一段：台账里考勤
    `rc=0`，看着是绿的，但绿只代表**进程正常退出**，不代表**消息发出去了**。
    真相在容器里的 `*.dispatch.json` 回执里，而 Coolify 的 `exec` 实测返回 404、
    `logs` 是空的、`/api/*` 在 Access 后面而 Owner 不登录——**没有任何人能拿到证据**。

    分得开的三件事（现在在台账上长得一模一样，都是 rc=0）：
      · `SENT`                          真发了
      · `NOTIFIER_CONFIG_MISSING`       通道没配好，一条没发
      · `NOT_SENT_NO_TARGET_SELECTED`   目标筛完是空的，一条没发

    公开边界：只出状态、时间、对象标签与通道名。**不出**考勤正文、模板文本、
    user_id——那是全员考勤数据和员工标识。
    """
    headers = {"Cache-Control": "no-store", "X-Robots-Tag": "noindex, nofollow"}
    now = datetime.now(BEIJING)
    if not ATTENDANCE_ARCHIVE_ROOT.is_dir():
        return JSONResponse({
            "生成时间": now.isoformat(), "可读": False,
            "原因": f"{ATTENDANCE_ARCHIVE_ROOT} 不存在——app 容器没挂考勤归档卷，"
                    f"或考勤从未产出过回执",
            "投递": [],
            "诚实边界": "读不到就说读不到，不拿空列表冒充『没有投递记录』。",
        }, headers=headers)
    receipts = _dispatch_receipts()
    latest = receipts[0] if receipts else None
    payload: dict[str, Any] = {}
    if latest:
        top_sent = latest.get("notification_status") == "SENT"
        targets = latest.get("目标") or []
        # 顶层状态和逐目标状态**可能对不上**。2026-07-28 首次真跑就撞上：顶层 SENT、
        # 目标行却判成失败（当时是端点读错了键）。键修好了，但这种不一致本身仍要
        # 报出来——它意味着「整体说成了、具体某个人没收到」，而那恰恰是 Owner 要问的。
        if targets and top_sent != all(t["成功"] for t in targets):
            payload["口径不一致"] = (
                f"顶层 notification_status={latest.get('notification_status')}，"
                f"但逐目标结果是 {[t['成功'] for t in targets]}——以逐目标为准，"
                f"整体说发了不等于每个人都收到了")
    return JSONResponse({
        "生成时间": now.isoformat(),
        "可读": True,
        "需要登录": False,
        "最近一次是否真的发出": (latest.get("notification_status") == "SENT") if latest else None,
        **payload,
        "为什么看这个": "台账 rc=0 只代表进程正常退出，不代表消息发出去了；"
                       "两者在台账上长得一模一样。",
        "投递": receipts,
    }, headers=headers)


@app.get("/public-api/项目成本")
def public_project_cost():
    """项目成本兼容 API；生产必须经过 Cloudflare Access。

    路由名保留 ``/public-api`` 只为兼容旧链接，不表示匿名授权。Origin guard
    独立校验 Access JWT，边缘策略误配时也不得返回客户名或财务金额。

    读不到就说读不到——不拿空列表冒充「没有项目」。
    """
    headers = {"Cache-Control": "no-store", "X-Robots-Tag": "noindex, nofollow"}
    if not RECENT_COST_PATH.exists():
        parent = RECENT_COST_PATH.parent
        reason = (f"{parent} 不存在——app 容器没挂 kmfa-logs 卷（部署配置问题）"
                  if not parent.exists()
                  else "刷新作业尚未产出：技能 project-cost-refresh 从未成功跑完一次")
        return JSONResponse(
            {"可读": False, "原因": reason, "项目": [],
             "诚实边界": "读不到就说读不到，不拿空列表冒充『没有项目』。"},
            headers=headers)
    try:
        payload = json.loads(RECENT_COST_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return JSONResponse(
            {"可读": False, "原因": f"产物无法解析：{type(exc).__name__}", "项目": []},
            headers=headers)
    try:
        _assert_current_project_cost_runtime(payload)
    except ValueError as exc:
        return JSONResponse(
            {
                "可读": False,
                "原因": f"项目成本运行态版本不兼容或不完整：{exc}",
                "项目": [],
            },
            status_code=503,
            headers=headers,
        )
    payload["可读"] = True
    payload["需要登录"] = True
    payload["产出时间"] = datetime.fromtimestamp(
        RECENT_COST_PATH.stat().st_mtime, BEIJING).isoformat()
    return JSONResponse(payload, headers=headers)


#: 项目成本页的外壳。刻意做成**服务端渲染的单页 HTML**：Owner 不登录、不看 SPA，
#: 出口必须是「打开就是数」。样式内联——CSP 拦外链，且这一页要能在任何网络下打开。
_COST_PAGE_SHELL = """<!doctype html><html lang="zh-CN"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow"><title>项目成本 · KMFA</title>
<style>
:root{--paper:#f6f7f9;--card:#fff;--ink:#10151c;--soft:#5c6672;--rule:#dfe3e8;
 --accent:#1d5c8f;--dim:#e8eef4;--bad:#9b2d2d;--badbg:#f7e8e8;
 --mono:ui-monospace,"SF Mono",Menlo,monospace;
 --sans:-apple-system,BlinkMacSystemFont,"PingFang SC","Hiragino Sans GB","Microsoft YaHei",sans-serif}
@media(prefers-color-scheme:dark){:root{--paper:#0c1015;--card:#141a21;--ink:#e6eaef;
 --soft:#8d98a5;--rule:#242c36;--accent:#6fa8d6;--dim:#17242f;--bad:#e08282;--badbg:#261616}}
*{box-sizing:border-box}body{margin:0;background:var(--paper);color:var(--ink);
 font-family:var(--sans);line-height:1.6}
.w{max-width:74rem;margin:0 auto;padding:2.2rem 1rem 4rem;display:flex;flex-direction:column;gap:1.5rem}
header{border-bottom:2px solid var(--ink);padding-bottom:.8rem}
.eb{font-family:var(--mono);font-size:.68rem;letter-spacing:.14em;color:var(--accent);text-transform:uppercase}
h1{font-size:clamp(1.4rem,4vw,2rem);font-weight:640;margin:.15rem 0 .3rem;letter-spacing:-.02em}
.sub{color:var(--soft);font-size:.86rem}
.strip{display:grid;grid-template-columns:repeat(auto-fit,minmax(9rem,1fr));gap:1px;
 background:var(--rule);border:1px solid var(--rule);border-radius:3px;overflow:hidden}
.st{background:var(--card);padding:.85rem 1rem}
.st b{display:block;font-family:var(--mono);font-size:1.45rem;font-weight:600;
 font-variant-numeric:tabular-nums;letter-spacing:-.02em}
.st span{font-size:.74rem;color:var(--soft)}
.note,.warn{border-left:3px solid var(--accent);background:var(--dim);
 padding:.75rem 1rem;border-radius:0 3px 3px 0;font-size:.86rem}
.warn{border-left-color:var(--bad);background:var(--badbg)}
.warn ul{margin:.4rem 0 0;padding-left:1.1rem}
.tw{overflow-x:auto;border:1px solid var(--rule);border-radius:3px}
table{border-collapse:collapse;width:100%;font-size:.8rem;background:var(--card);min-width:52rem}
th{text-align:left;font-weight:600;font-size:.66rem;letter-spacing:.05em;color:var(--soft);
 padding:.5rem .65rem;border-bottom:1px solid var(--rule);white-space:nowrap;
 text-transform:uppercase;position:sticky;top:0;background:var(--card)}
td{padding:.4rem .65rem;border-bottom:1px solid var(--rule)}
tr:last-child td{border-bottom:none}
td.n,th.n{font-family:var(--mono);font-variant-numeric:tabular-nums;text-align:right;white-space:nowrap}
td.k{font-family:var(--mono);font-size:.74rem;white-space:nowrap}
td.c{text-align:center;white-space:nowrap}td.b{font-weight:600}
code{font-family:var(--mono);font-size:.85em;background:var(--dim);padding:.05rem .28rem;border-radius:2px}
.bar{display:flex;flex-wrap:wrap;align-items:center;gap:.8rem;justify-content:space-between}
.dl{display:inline-block;background:var(--accent);color:#fff;text-decoration:none;
 padding:.5rem .95rem;border-radius:3px;font-size:.85rem;font-weight:600}
.dl:hover{filter:brightness(1.1)}
.hint{font-size:.78rem;color:var(--soft)}
th[data-s]{cursor:pointer;user-select:none}
th[data-s]:hover{color:var(--accent)}
th[data-s]::after{content:"";opacity:.35;margin-left:.28em;font-size:.9em}
th[data-s]:not([aria-sort])::after{content:"⇅"}
th[aria-sort="ascending"]::after{content:"↑";opacity:1}
th[aria-sort="descending"]::after{content:"↓";opacity:1}
td.neg{color:var(--bad);font-weight:600}
.grp{display:flex;gap:.55rem;align-items:center;flex-wrap:wrap}
.dl.alt{background:transparent;color:var(--accent);border:1px solid var(--accent);cursor:pointer;
 font-family:inherit}
.dl.alt:disabled{opacity:.55;cursor:default}
a.one{text-decoration:none;color:var(--accent);font-size:1rem}
a.one:hover{filter:brightness(1.25)}
#recalcmsg:empty{display:none}
#recalcmsg{background:var(--dim);border-left:3px solid var(--accent);padding:.6rem .9rem;
 border-radius:0 3px 3px 0}
footer{border-top:1px solid var(--rule);padding-top:.9rem;font-size:.76rem;color:var(--soft)}
</style></head><body><div class="w">{{BODY}}</div>
<script src="/static/project-cost.js" defer></script>
</body></html>"""

def _cost_num(value) -> Decimal | None:
    if value in (None, "", "None"):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _project_cost_review_counts(payload: dict) -> tuple[int, int]:
    """Read only the public-safe P1/P2 aggregate from a runtime payload."""

    control = payload.get("待确认") or {}
    try:
        return (
            int(control.get("P1开放复核数") or 0),
            int(control.get("P2已排除或提示数") or 0),
        )
    except (TypeError, ValueError):
        return 0, 0


def _assert_current_project_cost_runtime(payload: dict) -> None:
    """Reject legacy or incomplete project-cost payloads before interpretation."""

    if not isinstance(payload, dict):
        raise ValueError("runtime_payload_not_object")
    if payload.get("schema_version") != "kmfa.project_cost.current.v4":
        raise ValueError("runtime_schema_unsupported")
    snapshot_id = str(payload.get("快照ID") or "")
    if not snapshot_id:
        raise ValueError("runtime_snapshot_missing")
    status = str(payload.get("计算状态") or "")
    if status not in ("PASS", "PASS_WITH_OPEN_REVIEWS"):
        raise ValueError("runtime_calculation_not_publishable")
    projects = payload.get("项目")
    if not isinstance(projects, list) or any(
        not isinstance(project, dict) for project in projects
    ):
        raise ValueError("runtime_projects_invalid")
    if payload.get("项目数") != len(projects):
        raise ValueError("runtime_project_count_mismatch")
    for project in projects:
        margin_status = str(project.get("收入与毛利状态") or "")
        if not margin_status:
            raise ValueError("runtime_margin_status_missing")
        gross_profit = _cost_num(project.get("毛利"))
        closed_cost = _cost_num(project.get("项目成本"))
        gross_margin_text = project.get("毛利率")
        gross_margin_bps = project.get("毛利率基点")
        if margin_status != "READY":
            if (
                gross_profit is not None
                or closed_cost is not None
                or gross_margin_text not in (None, "")
                or gross_margin_bps is not None
            ):
                raise ValueError("runtime_blocked_margin_has_value")
            continue
        revenue = _cost_num(project.get("有效合同额"))
        cost = closed_cost
        incurred = _cost_num(project.get("项目已发生成本"))
        if (
            revenue is None
            or revenue <= 0
            or cost is None
            or (incurred is not None and cost < incurred)
            or gross_profit is None
            or isinstance(gross_margin_bps, bool)
            or not isinstance(gross_margin_bps, int)
        ):
            raise ValueError("runtime_margin_basis_invalid")
        if gross_profit != revenue - cost:
            raise ValueError("runtime_gross_profit_arithmetic")
        expected_bps = int(
            (
                gross_profit
                * Decimal(10_000)
                / revenue
            ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        )
        if gross_margin_bps != expected_bps:
            raise ValueError("runtime_gross_margin_arithmetic")
        if gross_margin_bps > 7000:
            raise ValueError("runtime_gross_margin_above_release_limit")
        expected_text = "%s%%" % format(
            Decimal(gross_margin_bps) / Decimal(100),
            ".2f",
        )
        if gross_margin_text != expected_text:
            raise ValueError("runtime_gross_margin_display_mismatch")
    review_control = payload.get("待确认")
    if not isinstance(review_control, dict) or review_control.get("状态") != status:
        raise ValueError("runtime_review_control_mismatch")
    source_binding = payload.get("封印来源")
    if not isinstance(source_binding, dict):
        raise ValueError("runtime_source_binding_missing")
    private_input_digest = str(
        source_binding.get("私有输入清单SHA256") or ""
    )
    if (
        source_binding.get("源码摘要算法")
        != "kmfa.project_cost.subject_tree.v1"
        or re.fullmatch(
            r"[0-9a-f]{64}",
            str(source_binding.get("源码SHA256") or ""),
        )
        is None
        or not isinstance(source_binding.get("源码文件数"), int)
        or isinstance(source_binding.get("源码文件数"), bool)
        or source_binding.get("源码文件数", 0) <= 0
        or re.fullmatch(
            r"[0-9a-f]{64}",
            private_input_digest,
        )
        is None
        or source_binding.get("输入清单类型")
        != "PRIVATE_MANIFEST_SHA256"
        or source_binding.get("输入清单SHA256") != private_input_digest
        or re.fullmatch(
            r"[0-9a-f]{64}",
            str(source_binding.get("选中来源绑定SHA256") or ""),
        )
        is None
    ):
        raise ValueError("runtime_source_binding_invalid")
    binding = payload.get("封印工作簿")
    if not isinstance(binding, dict):
        raise ValueError("sealed_workbook_binding_missing")
    filename = str(binding.get("文件名") or "")
    digest = str(binding.get("SHA256") or "").lower()
    size = binding.get("字节数")
    if (
        not filename
        or Path(filename).name != filename
        or "/" in filename
        or "\\" in filename
        or not re.fullmatch(r"[0-9a-f]{64}", digest)
        or not isinstance(size, int)
        or isinstance(size, bool)
        or size <= 0
        or binding.get("快照ID") != snapshot_id
    ):
        raise ValueError("sealed_workbook_binding_invalid")


def _sealed_project_cost_workbook(payload: dict) -> Path:
    """Resolve and hash-check the immutable workbook bound to this snapshot."""

    _assert_current_project_cost_runtime(payload)
    binding = payload.get("封印工作簿")
    assert isinstance(binding, dict)
    filename = str(binding.get("文件名") or "")
    digest = str(binding.get("SHA256") or "").lower()
    size = binding.get("字节数")
    assert isinstance(size, int) and not isinstance(size, bool)
    path = RECENT_COST_PATH.parent / filename
    if path.is_symlink() or not path.is_file() or path.stat().st_size != size:
        raise ValueError("sealed_workbook_unavailable")
    actual = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            actual.update(block)
    if actual.hexdigest() != digest:
        raise ValueError("sealed_workbook_hash_mismatch")
    return path


def _safe_spreadsheet_text(value) -> str:
    """Keep user-controlled text from becoming an Excel formula."""

    text = str(value or "")
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


PUBLIC_PROJECT_COST_PAGE_PATH = "/project-cost"
PUBLIC_PROJECT_COST_DOWNLOAD_PATH = "/project-cost/download"


@app.api_route(PUBLIC_PROJECT_COST_PAGE_PATH, methods=["GET", "HEAD"],
               response_class=HTMLResponse, include_in_schema=False)
@app.api_route("/项目成本", methods=["GET", "HEAD"], response_class=HTMLResponse,
               include_in_schema=False)
@app.api_route("/public-api/项目成本表", methods=["GET", "HEAD"],
               response_class=HTMLResponse, include_in_schema=False)
def public_project_cost_page(request: Request):
    """项目成本——公开报表打开就是数。

    为什么要有它（Owner 2026-07-29）：「我说了我只要我的项目成本！」「我没有看到
    你说的东西」「你不要放在本地，你推上网上去」。此前只有 `/public-api/项目成本`
    的 JSON——那是给机器读的，人打开看到的是一屏花括号。发文件也不行：卡片可能
    根本没露出来。所以出口必须是一个**打开就是数**的网页。

    三个地址均是面向访客的公开、只读报表入口；无登录、无重算能力：
      · `/project-cost` —— 新的短入口。
      · `/项目成本` 与 `/public-api/项目成本表` —— 保留既有书签和页面调用。

    公开页只暴露已通过项目成本发布门的快照，始终 `no-store`/`noindex`；
    JSON 兼容接口、重算和全部 `/api`、`/ops` 路由仍由 Cloudflare Access 与
    origin guard 双重保护。
    """
    # Keep legacy page links self-contained for existing bookmarks while all
    # three paths remain the same read-only public representation.
    download_path = (
        PUBLIC_PROJECT_COST_DOWNLOAD_PATH
        if request.url.path == PUBLIC_PROJECT_COST_PAGE_PATH
        else "/项目成本/下载"
    )
    headers = {
        "Cache-Control": "no-store",
        "X-Robots-Tag": "noindex, nofollow",
        "X-KMFA-Cost-Access": "public-read",
    }

    def page(body: str) -> HTMLResponse:
        return HTMLResponse(_COST_PAGE_SHELL.replace("{{BODY}}", body), headers=headers)

    if not RECENT_COST_PATH.exists():
        return page(
            '<div class="warn"><b>还没有数。</b>刷新作业 <code>project-cost-refresh</code>'
            "尚未成功跑完一次，或 app 容器没挂上共享卷。"
            "<br>读不到就说读不到——不拿空表冒充「没有项目」。</div>")
    try:
        payload = json.loads(RECENT_COST_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return page(f'<div class="warn"><b>产物无法解析</b>：{type(exc).__name__}</div>')
    try:
        _assert_current_project_cost_runtime(payload)
    except ValueError as exc:
        return HTMLResponse(
            _COST_PAGE_SHELL.replace(
                "{{BODY}}",
                '<div class="warn"><b>运行态版本不兼容或不完整。</b>'
                f"{html_escape(str(exc))}；等待 project-cost-refresh 成功发布新快照。</div>",
            ),
            status_code=503,
            headers=headers,
        )

    projects = [item for item in (payload.get("项目") or []) if isinstance(item, dict)]
    projects.sort(
        key=lambda project: (
            -(
                _cost_num(project.get("项目成本"))
                or _cost_num(project.get("项目已发生成本"))
                or 0
            ),
            str(project.get("合同编号") or ""),
        )
    )
    known_cost = [
        project for project in projects
        if _cost_num(project.get("项目已发生成本")) is not None
    ]
    nonzero_cost = [
        project for project in known_cost
        if _cost_num(project.get("项目已发生成本")) != 0
    ]
    closed_margin_count = sum(
        project.get("收入与毛利状态") == "READY"
        for project in projects
    )
    unclosed_cost_count = len(projects) - closed_margin_count
    total = sum(
        _cost_num(project.get("项目已发生成本")) or 0
        for project in known_cost
    )
    p1_open, p2_notices = _project_cost_review_counts(payload)
    ledger_coverage = payload.get("账簿覆盖") or {}
    try:
        stale_ledger_entities = int(
            ledger_coverage.get("早于报表截至月主体数") or 0
        )
    except (TypeError, ValueError):
        stale_ledger_entities = 0
    review_warning = (
        '<div class="warn"><b>存在开放复核项。</b>'
        f"当前有 {p1_open} 条 P1 观察未能唯一归属或未满足成本资格，"
        "它们均未进入任何项目的正式金额；这份结果只代表已唯一归属的合格事件。"
        f"另有 {p2_notices} 条已确定排除、别名修复或控制提示。</div>"
        if p1_open
        else ""
    )
    coverage_warning = (
        '<div class="warn"><b>账簿截止月早于报表截至日。</b>'
        f"有 {stale_ledger_entities} 个账簿主体尚未覆盖到"
        f"{html_escape(str(payload.get('截至日期') or '报表截至日'))} 所在月份。"
        "过账实际只代表各主体已选账簿截至月；之后只计入满足资格且未见过账冲突的应计。"
        "</div>"
        if stale_ledger_entities
        else ""
    )

    def money(value) -> str:
        number = _cost_num(value)
        return f"{number:,.2f}" if number is not None else "—"

    def cell(value, kind="n"):
        """数值单元格带原始 data-v，空值不冒充 0。"""
        number = _cost_num(value)
        shown = f"{number:,.2f}" if number is not None else "—"
        attr = f' data-v="{number}"' if number is not None else ' data-v=""'
        return f'<td class="{kind}"{attr}>{shown}</td>'

    def rate_cell(project):
        bps = project.get("毛利率基点")
        if isinstance(bps, int) and not isinstance(bps, bool):
            return (
                f'<td class="n b" data-v="{bps}">'
                f'{bps / 100:.2f}%</td>'
            )
        return '<td class="n" data-v="">待闭合</td>'

    def margin_status_text(project):
        if project.get("收入与毛利状态") == "READY":
            return "已闭合"
        coverage = str(project.get("项目成本覆盖") or "")
        ledger_period = str(project.get("账簿截至月份") or "")
        if coverage == "SOURCE_UNAVAILABLE":
            return "待成本闭合｜来源不可用"
        if "NO_QUALIFIED_EVENT" in coverage and ledger_period:
            return "待成本闭合｜账簿截至 %s｜无合格过账" % ledger_period
        if ledger_period:
            return "待成本闭合｜账簿截至 %s" % ledger_period
        return "待成本闭合"

    def download_cell(project):
        if project.get("收入与毛利状态") != "READY":
            return (
                '<td class="c" data-v=""><span class="hint" '
                'title="项目成本未闭合，不能生成正式项目财务分析表">待闭合</span></td>'
            )
        contract = quote(str(project.get("合同编号") or ""))
        return (
            '<td class="c" data-v=""><a class="one" '
            'title="只下载这一个已闭合合同" '
            f'href="{download_path}?合同={contract}">⬇</a></td>'
        )

    rows = "\n".join(
        "<tr>"
        f'<td class="k" data-v="{html_escape(str(p.get("合同编号") or ""))}">'
        f'{html_escape(str(p.get("合同编号") or ""))}</td>'
        f'<td data-v="{html_escape(str(p.get("项目名称") or p.get("甲方名称") or ""))}">'
        f'<b>{html_escape(str(p.get("项目名称") or "—"))}</b><br>'
        f'<span class="hint">{html_escape(str(p.get("甲方名称") or "—"))}</span></td>'
        f'<td class="c" data-v="{html_escape(str(p.get("施工状态") or ""))}">'
        f'{html_escape(str(p.get("施工状态") or "—"))}</td>'
        f'<td class="c" data-v="{html_escape(str(p.get("完工日期") or ""))}">'
        f'{html_escape(str(p.get("完工日期") or "—"))}</td>'
        + cell(p.get("有效合同额"))
        + cell(p.get("项目成本"), "n b")
        + cell(p.get("项目已发生成本"))
        + cell(p.get("毛利"), "n b")
        + rate_cell(p)
        + (
            f'<td class="c" data-v="{html_escape(margin_status_text(p))}">'
            f'{html_escape(margin_status_text(p))}</td>'
        )
        + download_cell(p)
        + "</tr>"
        for p in projects)

    footer = "这是公开、只读的已发布项目成本报表；重算与机器接口继续受保护。"

    body = f"""
<header>
  <div class="eb">KMFA · FORMAL PROJECT COST</div>
  <h1>项目成本</h1>
  <div class="sub">快照 {html_escape(str(payload.get("快照ID") or "—"))}
    　·　截至 {html_escape(str(payload.get("截至日期") or "—"))}
    　·　生成 {html_escape(str(payload.get("生成时间") or "—"))}</div>
</header>
<div class="strip">
  <div class="st"><b>{len(projects):,}</b><span>2026 全部项目</span></div>
  <div class="st"><b>{len(nonzero_cost)}</b><span>已有成本发生</span></div>
  <div class="st"><b>{closed_margin_count}</b><span>毛利口径已闭合</span></div>
  <div class="st"><b>{total:,.2f}</b><span>项目已发生成本合计</span></div>
  <div class="st"><b>{unclosed_cost_count}</b><span>成本尚未闭合</span></div>
</div>
<div class="note"><b>毛利口径：</b>毛利＝有效收入基数－已闭合项目成本；
  毛利率＝毛利÷有效收入基数。实际发生额仍是下限、人工或审批费用未闭合、
  或尚无受控完工预计时，毛利率显示“待闭合”，不会把低估成本算成高毛利。<br>
  <b>发布硬控制：</b>任何项目毛利率高于 70% 会阻断整批发布；
  该阈值只用于报错，绝不用于倒推、补差或压低毛利率。</div>
{coverage_warning}
{review_warning}
<div class="bar">
  <span class="grp">
    <a class="dl" href="{download_path}">⬇ 下载全部（Excel）</a>
  </span>
  <span class="hint">点表头按该列排序，再点一次反向　·　只有“已闭合”项目可单独下载</span>
</div>
<div class="tw"><table id="costtbl">
<thead><tr>
<th data-s="t">合同编号</th><th data-s="t">项目 / 甲方</th><th data-s="t">状态</th><th data-s="t">完工日</th>
<th class="n" data-s="n">有效收入基数</th><th class="n" data-s="n">项目成本（已闭合）</th>
<th class="n" data-s="n">已发生成本（下限）</th>
<th class="n" data-s="n">毛利</th><th class="n" data-s="n">毛利率</th>
<th data-s="t">口径状态</th>
<th>单独下载</th>
</tr></thead>
<tbody>{rows}</tbody></table></div>
<footer>{footer}
  收入、人工、审批费用和完工预计任一未闭合时，闭合项目成本、毛利与毛利率均保持空白；
  已发生成本只作为实际发生下限单独展示。</footer>"""
    return page(body)


#: Owner 那张《生产项目状态表》「信息表」的**列序原样**（30 列，2026-07-29 从
#: ~/Downloads/生产项目状态表.xlsx 逐列读出）。
#:
#: 为什么钉在这里：Owner 2026-07-29「项目成本单个项目下载下来的和我原来的格式
#: 根本不一样，你不要用乱七八糟的东西恶心我，这个东西很急，我和你说了无数遍」。
#: 上一版我把页签对齐了 `KMFA_项目成本_真实参考回放_8项目.xlsx`——**那是我自己
#: 生成的产物**，我却在注释里写成「对齐 Owner 手上那份」。拿自己的输出当基准，
#: 就是这句「说了无数遍」的由来。
#:
#: 这里只存**列名**：列名不是业务数据，可以进公开仓；真实甲方名与金额一个都不进。
OWNER_STATUS_COLUMNS = (
    "甲方名称", "省份", "合同号", "含税合同金额", "税率", "负责人", "项目类型",
    "开工时间", "完工时间", "实际工期", "施工状态", "结算时间", "开票时间", "回款时间",
    "完工后结算时间", "结算后开票时间", "开票后回款90%时间", "结算金额", "开票金额",
    "结算审计偏差率", "自有人工工时", "劳务人工工时", "生活住宿费", "交通费", "材料费",
    "其他费用", "项目成本表截止提供时间", "截止剩余时间", "是否提供项目成本表",
    "是否已计算提成",
)

#: 我算出来的东西**接在原表 30 列之后**，不插进去、不改名、不顶掉任何一列。
#: 插进去就等于改了 Owner 的表；顶掉就等于我认为我的口径比他的表更权威。
COMPUTED_COLUMNS = (
    "项目过账实际", "项目应计", "项目已发生成本", "正式材料成本",
    "正式租赁物流成本", "正式现场管理成本", "正式劳务承包成本",
    "主营成本已结转", "状态表已报直接成本", "支付系统已付观察",
    "账簿截至月份", "项目成本覆盖", "应计覆盖", "合同额口径", "收入与毛利状态",
)

#: 原表里这几列是 Excel 日期序列号（如 45999），导出要还成人看得懂的日期。
_DATE_COLUMNS = frozenset({
    "开工时间", "完工时间", "结算时间", "开票时间", "回款时间",
    "项目成本表截止提供时间",
})

#: 我的产物字段名 → 原表列名。只映**同一件事**，名字不同就在这里对上，
#: 不在这里的列一律留空——留空是「我不知道」，填 0 是「我说它是 0」，两码事。
_FIELD_TO_OWNER_COLUMN = {
    "合同编号": "合同号",
    "完工日期": "完工时间",
}


def _owner_row(project: dict) -> list:
    """按 Owner 列序摊一行。缺的留空，绝不臆造。"""
    reverse = {v: k for k, v in _FIELD_TO_OWNER_COLUMN.items()}
    row = []
    for column in OWNER_STATUS_COLUMNS:
        field = reverse.get(column, column)
        value = project.get(field, "")
        if value in (None, ""):
            row.append("")
            continue
        if column in _DATE_COLUMNS:
            text = str(value)
            row.append(text.split(" ")[0] if " " in text else text)
            continue
        number = _cost_num(value)
        row.append(number if number is not None else value)
    return row


def _computed_row(project: dict) -> list:
    buckets = project.get("报表归类") or {}

    def bucket_total(*keys):
        values = [_cost_num(buckets.get(key)) for key in keys]
        present = [value for value in values if value is not None]
        return sum(present) if present else ""

    calculated = {
        "正式材料成本": bucket_total("material", "fuel_power"),
        "正式租赁物流成本": bucket_total("rental", "logistics"),
        "正式现场管理成本": bucket_total(
            "own_labor", "travel", "lodging", "living",
            "road_parking", "vehicle", "other",
        ),
        "正式劳务承包成本": bucket_total("subcontract_labor"),
    }
    row = []
    for column in COMPUTED_COLUMNS:
        value = calculated.get(column, project.get(column, ""))
        if value in (None, ""):
            row.append("")
            continue
        number = _cost_num(value)
        row.append(number if number is not None else value)
    return row


@app.api_route(PUBLIC_PROJECT_COST_DOWNLOAD_PATH, methods=["GET", "HEAD"],
               include_in_schema=False)
@app.api_route("/项目成本/下载", methods=["GET", "HEAD"], include_in_schema=False)
@app.api_route("/public-api/项目成本表/下载", methods=["GET", "HEAD"],
               include_in_schema=False)
def public_project_cost_download(合同: str | None = None):
    """下载与当前运行态快照绑定的项目成本工作簿。

    全量下载直接返回由 Skill 生成、验证并封印的 8 页签工作簿，网站不再二次生成
    一份看似相同但无法证明同源的文件。指定合同时，才按真实竖版参考模板生成单项目
    《项目财务分析表》；未知值留空，不用 0 冒充。
    """
    if not RECENT_COST_PATH.exists():
        raise HTTPException(status_code=503, detail="项目成本产物还没生成（project-cost-refresh 未成功跑完）")
    try:
        payload = json.loads(RECENT_COST_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=503, detail=f"产物无法解析：{type(exc).__name__}") from exc
    try:
        _assert_current_project_cost_runtime(payload)
    except ValueError as exc:
        raise HTTPException(
            status_code=503,
            detail=f"项目成本运行态版本不兼容或不完整：{exc}",
        ) from exc
    p1_open, _ = _project_cost_review_counts(payload)

    projects = payload.get("项目") or []
    # 正式运行态已完成唯一合同身份解析；全部项目都输出。0.00 是经过完整
    # 选定期间检查后的明确零，空值才是来源不可用，两者不能再用过滤器混为一谈。
    rows = [project for project in projects if isinstance(project, dict)]

    tag = ""
    if 合同:
        wanted = str(合同).strip()
        rows = [p for p in rows if str(p.get("合同编号") or "").strip() == wanted]
        if not rows:
            raise HTTPException(status_code=404, detail=f"没有这个合同号的成本记录：{wanted}")
        tag = f"_{wanted}"

    import io as _io  # noqa: PLC0415

    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import (  # noqa: PLC0415
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.worksheet.page import PageMargins  # noqa: PLC0415

    # 单个合同 → **竖版《项目财务分析表》**，使用与 Skill 封印 PDF 相同的
    # B-family 行序和闭合项目财务分析成本口径。
    #
    # Owner 2026-07-30：「/Users/linzezhang/Downloads/KMFA_MetaData/销售绩效考核
    # 这里面的才是真实模版，你现在用的不知道是什么恶心东西」。
    # 那目录下是「竣工项目财务报表」PDF——单个项目的成本表是竖版分析表，
    # 而《生产项目状态表》的 30 列横表是**项目清单**的格式。两件事，我混了。
    # 我在这上面错了两版：先对齐自己生成的 xlsx，再把清单格式当成单项目格式。
    if 合同:
        from .project_statement import statement_header, statement_rows  # noqa: PLC0415

        project = rows[0]
        if str(project.get("收入与毛利状态") or "") != "READY":
            raise HTTPException(
                status_code=409,
                detail=(
                    "该项目成本尚未闭合；网页仅展示已发生成本下限，"
                    "禁止生成或下载正式项目财务分析表"
                ),
            )
        book = Workbook()
        ws = book.active
        ws.title = "项目财务分析表"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 31
        ws.column_dimensions["B"].width = 17
        ws.column_dimensions["C"].width = 48
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(
            left=Decimal("0.22"),
            right=Decimal("0.22"),
            top=Decimal("0.3"),
            bottom=Decimal("0.3"),
            header=Decimal("0.1"),
            footer=Decimal("0.1"),
        )

        title = ws.cell(row=1, column=1, value="项目财务分析表")
        title.font = Font(bold=True, size=16, color="17365D")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        ws.row_dimensions[1].height = 26

        line = 2
        for label, value in statement_header(project):
            ws.cell(row=line, column=1, value=label).font = Font(bold=True, size=9)
            ws.cell(
                row=line,
                column=2,
                value=_safe_spreadsheet_text(value),
            )
            ws.merge_cells(start_row=line, start_column=2, end_row=line, end_column=3)
            ws.row_dimensions[line].height = 17
            line += 1
        header_line = line
        ws.cell(row=line, column=1, value="项目")
        ws.cell(row=line, column=2, value="金额（元）")
        ws.cell(row=line, column=3, value="备注")
        for column in range(1, 4):
            current = ws.cell(row=line, column=column)
            current.font = Font(bold=True, color="FFFFFF", size=9)
            current.fill = PatternFill("solid", fgColor="17365D")
            current.alignment = Alignment(
                horizontal="center", vertical="center"
            )
        line += 1

        thin = Side(style="thin", color="7F8C8D")
        section_fill = PatternFill("solid", fgColor="9DC3E6")
        subtotal_fill = PatternFill("solid", fgColor="DDEBF7")
        total_fill = PatternFill("solid", fgColor="F4B183")
        policy_fill = PatternFill("solid", fgColor="FFF2CC")
        section_labels = {"二、资金运用及各项支出"}
        subtotal_labels = {
            "（一）原材料", "（二）租赁费", "（三）保险费",
            "（四）现场管理费", "（五）工资（承包费）支出", "（六）信息费",
            "（七）税金",
        }
        total_labels = {"一、合同额", "项目产值", "三 利润"}
        policy_labels = {
            "（八） 分摊的管理费用（合同的2%）",
        }
        for label, amount, note in statement_rows(project):
            ws.cell(row=line, column=1, value=label)
            cell = ws.cell(row=line, column=2)
            if amount is not None:
                cell.value = amount
                cell.number_format = "#,##0.00;[Red](#,##0.00);0.00"
            ws.cell(row=line, column=3, value=note or None)
            fill = None
            if label in total_labels:
                fill = total_fill
            elif label in section_labels:
                fill = section_fill
            elif label in subtotal_labels:
                fill = subtotal_fill
            elif label in policy_labels:
                fill = policy_fill
            for column in range(1, 4):
                current = ws.cell(row=line, column=column)
                current.border = Border(
                    left=thin, right=thin, top=thin, bottom=thin
                )
                current.font = Font(
                    bold=fill is not None,
                    size=8,
                    color="9C0006" if label in policy_labels else "000000",
                )
                current.alignment = Alignment(
                    horizontal="right" if column == 2 else "left",
                    vertical="center",
                    wrap_text=True,
                    indent=1 if column == 1 and fill is None else 0,
                )
                if fill is not None:
                    current.fill = fill
            ws.row_dimensions[line].height = 14.5
            line += 1

        line += 1
        ws.cell(row=line, column=1, value="项目经理：")
        ws.cell(row=line, column=3,
                value=f"日期：{datetime.now(BEIJING).strftime('%Y年%m月%d日')}")
        ws.print_title_rows = f"1:{header_line}"
        ws.print_area = f"A1:C{line}"

        note_ws = book.create_sheet("口径")
        note_ws.column_dimensions["A"].width = 22
        note_ws.column_dimensions["B"].width = 96
        for key, text in (
            ("生成时间", str(payload.get("生成时间") or "")),
            ("模版", "与 Skill 封印单项目 PDF 一致的 B-family 项目财务分析表"),
            ("空行", "表示「本系统没有这个数」，**不是 0**"),
            ("正式成本", "项目已发生成本＝项目过账实际＋合格应计；内部全程整数分"),
            ("收入与毛利状态", str(project.get("收入与毛利状态") or "")),
            (
                "已发生成本下限（非闭合项目成本）",
                str(project.get("项目已发生成本") or ""),
            ),
            ("闭合项目成本", str(project.get("项目成本") or "")),
            ("毛利率", str(project.get("毛利率") or "")),
            (
                "复核状态",
                f"{payload.get('计算状态') or ''}；P1 开放复核 {p1_open} 条，"
                "均未进入正式金额",
            ),
            ("（四）现场管理费", "仅承接正式事件分类；不能安全细分的金额在备注中保留"),
            ("（五）工资（承包费）支出", "正式劳务/分包/人工事件；不使用固定工时单价"),
            ("（八）分摊的管理费用", "保留模板原行；无合格政策时留空，禁止按合同额2%生成"),
            ("毛利", "有效收入基数减闭合项目成本；毛利率超过70%时整批禁止发布"),
            ("备注百分比", "占闭合项目成本比例"),
        ):
            note_ws.append([key, text])
        note_ws.cell(row=1, column=1).font = Font(bold=True)

        stream = _io.BytesIO()
        book.save(stream)
        stream.seek(0)
        from openpyxl import load_workbook  # noqa: PLC0415

        checked = load_workbook(stream, read_only=True, data_only=False)
        try:
            if any(
                cell.data_type == "f"
                or (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                )
                for sheet in checked.worksheets
                for row in sheet.iter_rows()
                for cell in row
            ):
                raise HTTPException(
                    status_code=500,
                    detail="单项目工作簿安全校验失败",
                )
        finally:
            checked.close()
        stream.seek(0)
        day = datetime.now(BEIJING).strftime("%Y%m%d")
        name = f"项目财务分析表_{合同}_{day}.xlsx"
        return Response(
            content=stream.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":
                    f"attachment; filename=KMFA_statement_{合同}_{day}.xlsx; "
                    f"filename*=UTF-8''{quote(name)}",
                "Cache-Control": "no-store",
                "X-Robots-Tag": "noindex, nofollow",
            },
        )

    try:
        sealed_workbook = _sealed_project_cost_workbook(payload)
    except (OSError, ValueError) as exc:
        raise HTTPException(
            status_code=503,
            detail=f"封印项目成本工作簿不可用：{exc}",
        ) from exc
    day = datetime.now(BEIJING).strftime("%Y%m%d")
    return FileResponse(
        sealed_workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"KMFA_项目成本报表_{day}.xlsx",
        headers={
            "Cache-Control": "no-store",
            "X-Robots-Tag": "noindex, nofollow",
            "X-KMFA-Snapshot-ID": str(payload.get("快照ID") or ""),
            "X-KMFA-Workbook-SHA256": str(
                (payload.get("封印工作簿") or {}).get("SHA256") or ""
            ),
        },
    )


#: 「重新计算」的请求标记。App 与 skills 是**两个容器**，App 里没有 run_skill.sh、
#: 也不该有——真让 App 去跑免 clone 下载并解析私有清单内全部文件的活，就是把 2026-07-28 那次
#: 「压测把线上打下线」原样重演一遍，只是换了个触发器。所以 App 只放一个标记。
#:
#: 放在 **app-state 卷**，不是日志卷。两个卷的读写方向是刻意反着的：
#:   kmfa-logs      skills 可写 / app **只读**
#:   kmfa-app-state app 可写   / skills **只读**（daily-backup 读它打包备份，注释写明「绝不写」）
#: 第一版把标记写进日志卷，线上直接 503「共享卷不可写」——那道只读边界是有意的，
#: 不该为了一个按钮把整个日志卷对 app 开成可写。
#:
#: skills 那边**删不掉**这个标记（它只读挂载），所以不能用「先删后跑」，
#: 改成比时间戳：标记里写请求时刻，skills 在自己可写的卷上记「上次处理到哪」，
#: 标记比记录新就跑。跑的过程中再点一次也不会被吞掉——时间戳又变新了。
COST_REFRESH_FLAG = APP_STATE_DIR / ".project_cost_refresh_requested"


@app.post("/项目成本/重算", include_in_schema=False)
@app.post("/public-api/项目成本表/重算", include_in_schema=False)
def public_project_cost_refresh():
    """请求重算项目成本——**不在这个容器里跑**，只放一个标记。

    Owner 2026-07-29：「不支持实时更新」。此前只有两种时机会重算：
    每次部署、以及每天 05:45 的排程。源数据变了要等到第二天，那不叫实时。

    返回里带上「上一次算完是什么时候」，好让人判断按下去之后有没有真的变——
    只回一句「已提交」而不给时间戳，跟没回一样。
    """
    headers = {"Cache-Control": "no-store", "X-Robots-Tag": "noindex, nofollow"}
    previous = None
    if RECENT_COST_PATH.exists():
        previous = datetime.fromtimestamp(
            RECENT_COST_PATH.stat().st_mtime, BEIJING).isoformat(timespec="seconds")
    try:
        COST_REFRESH_FLAG.parent.mkdir(parents=True, exist_ok=True)
        COST_REFRESH_FLAG.write_text(
            datetime.now(BEIJING).isoformat(timespec="seconds"), encoding="utf-8")
    except OSError as exc:
        # 报**这个**卷的名字。第一版把标记写进 kmfa-logs 才 503，改到 app-state 后
        # 这句话没跟着改，于是失败信息把人指向一个根本没参与的卷——
        # 「说错了地方的报错」比不报还费时间。
        return JSONResponse(
            {"已提交": False,
             "原因": (f"{COST_REFRESH_FLAG.parent} 写不进去（{type(exc).__name__}）"
                    "——app 容器没挂上 kmfa-app-state 卷（部署配置问题）"),
             "上次算完": previous},
            status_code=503, headers=headers)
    return JSONResponse(
        {"已提交": True,
         "说明": "skills 容器每分钟检查一次；重算会免 clone 按私有清单下载并解析全部选定文件，约 2–4 分钟。",
         "上次算完": previous,
         "怎么确认": "过几分钟刷新本页，看顶部「数据生成」时间有没有变。"},
        headers=headers)


@app.get("/public-api/技能健康")
def public_skill_health():
    """技能运行健康的**公开安全**摘要：只有「谁、什么时候、成没成」。

    为什么必须是公开的（2026-07-27 实测逼出来的）：
      Coolify 的 `exec` 返回 404、`logs` 返回空、`/api/排程健康` 在 Cloudflare Access 后面，
      而 Owner 明令不登录、不看页面——于是「技能到底跑没跑」**没有任何人能拿到证据**。
      判据只能落在 agent 自验上，那就必须有一个不需要凭据也能读到的产出物。

    公开边界：只出技能名（本就公开在仓库里）、时间戳、退出码、成功与否、次数。
    **不出**日志路径（会暴露目录结构）、不出投递开关（属运行策略）、不出任何业务数字。

    读不到就说读不到——不拿「没有坏消息」当好消息。
    """
    headers = {"Cache-Control": "no-store"}
    now = datetime.now(BEIJING)
    # Daily funds is intentionally outside the shared skills container and
    # ledger.  Its worker-to-app hand-off is the only admissible health source
    # for this one row; do not let a missing shared ledger turn that fact into
    # a fabricated "0 runs" result.
    daily_funds_row = _daily_funds_public_skill_health_row(now)
    if not SKILL_LEDGER_PATH.exists():
        return JSONResponse({
            "生成时间": now.isoformat(),
            "台账可读": False,
            "原因": "通用技能排程台账不存在——通用台账未验证；每日资金使用独立、无金额的轮询回执。",
            "技能": [daily_funds_row],
            "诚实边界": "读不到就说读不到，不猜、不拿「没有坏消息」当好消息。",
        }, headers=headers)

    by_skill: dict[str, list] = {}
    for line in SKILL_LEDGER_PATH.read_text(encoding="utf-8", errors="replace").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            r = json.loads(line)
        except json.JSONDecodeError:
            continue
        by_skill.setdefault(str(r.get("skill") or "?"), []).append(r)

    出 = []
    for skill in sorted(SCHEDULE_CONTRACT):
        if skill == "daily-funds":
            出.append(daily_funds_row)
            continue
        history = sorted(by_skill.get(skill, []), key=lambda r: str(r.get("ts")), reverse=True)
        # ✅/❌ 只由**排程跑**决定。压测跑（entrypoint 的全量压测）问的是另一个问题：
        # 「这个技能的机器还转不转」。混在一起，时间锚定的技能被拉到窗口外跑的合法失败
        # 就会顶掉当天真成功的排程结论——2026-07-28 实测：早上 10:36 考勤真发成功，
        # 晚上 19:44 压测重跑 REALTIME_REMINDER_INTEGRITY_FAILED，页面直接变红。
        # 但压测结果也不能藏：藏了就等于没做压测，回到「改一版等一天」。所以分两栏各报各的。
        # 老台账行没有 sweep 字段，一律按排程跑处理——改造不能把历史判没了。
        scheduled = [r for r in history if not r.get("sweep")]
        swept = [r for r in history if r.get("sweep")]
        last = scheduled[0] if scheduled else None
        距今小时 = None
        if last:
            try:
                距今小时 = round(
                    (now - datetime.fromisoformat(str(last["ts"]))).total_seconds() / 3600, 1)
            except (ValueError, KeyError, TypeError):
                距今小时 = None
        失败码 = _public_failure_code((last or {}).get("code")) if last and last.get("rc") else None
        # rc=0 底下也有截然不同的结局——例如 dws-data-auth 的「授权请求已发出」
        # 与「按设计没请求」，两者都是 rc=0，在页面上曾长得一模一样。
        # 「成功」不是终点，**成功里做了什么**才是；不报出来就等于又一种假绿。
        本次状态 = _public_failure_code((last or {}).get("code")) if last else None
        行 = {
            "技能": skill,
            "最近一次": (last or {}).get("ts"),
            "距今小时": 距今小时,
            "退出码": (last or {}).get("rc"),
            "成功": (last or {}).get("rc") == 0 if last else None,
            "运行次数": len(scheduled),
            # rc 只说「失败了」，失败码说「哪一种失败」——没有它就只能改一版等一天看会不会变绿。
            "失败码": 失败码,
            "本次状态": 本次状态,
        }
        if swept:
            最近压测 = swept[0]
            行["压测"] = {
                "最近一次": 最近压测.get("ts"),
                "退出码": 最近压测.get("rc"),
                "成功": 最近压测.get("rc") == 0,
                "次数": len(swept),
                "失败码": (_public_failure_code(最近压测.get("code"))
                          if 最近压测.get("rc") else None),
            }
        出.append(行)
    return JSONResponse({
        "生成时间": now.isoformat(),
        "台账可读": True,
        "技能": 出,
        "台账回传": _ledger_uplink_state(),
        "口径": "只报运行事实，不含任何业务数据；零运行次数即视为未跑通，不因日志新鲜而判健康。",
        "压测口径": (
            "「压测」是部署后把每个技能主动跑一遍，问的是「机器还转不转」；"
            "上面的成功/失败只由**排程跑**决定，问的是「今天这件事办成没有」。"
            "时间锚定的技能在压测里被拉到窗口外跑，合法拒绝会显示为压测失败，"
            "那不代表排程有问题。"
        ),
        "失败码口径": "白名单构造的机器状态令牌，形不合者一律不出；完整取证只落私有库，不上公开端点。",
    }, headers=headers)


@app.get("/api/排程健康")
def schedule_health():
    """排程执行健康——**不出任何业务数据**，只报「谁在什么时候跑了、成没成」。

    刻意做成这样：判断「排程是不是活着」不该需要登服务器，也不该需要谁口头汇报。
    """
    if not SKILL_LEDGER_PATH.exists():
        log_root = SKILL_LEDGER_PATH.parent
        if not log_root.exists():
            原因 = f"{log_root} 目录不存在——app 容器没挂 kmfa-logs 卷（部署配置问题，不是排程问题）"
        elif not any(log_root.iterdir()):
            原因 = f"{log_root} 已挂载但是空的——排程容器从未写入（查 skills 容器是否 Started）"
        else:
            原因 = f"{log_root} 有内容但没有 {SKILL_LEDGER_PATH.name}——排程从未完成过一次运行"
        return {
            "可读": False,
            "原因": 原因,
            "排程契约": SCHEDULE_CONTRACT,
            "逐项": [_daily_funds_schedule_row()],
            "每日资金": {
                **_daily_funds_status(),
                "业务流": _daily_funds_flow_state(),
            },
            "诚实边界": "读不到就说读不到，不猜、不拿「没有坏消息」当好消息。",
        }

    rows = []
    for line in SKILL_LEDGER_PATH.read_text(encoding="utf-8", errors="replace").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue

    now = datetime.now(BEIJING)
    # 台账是 append-only 全量历史（run_skill.sh 每次运行追加一行、日志按时间戳独立归档）。
    # 此前只取每技能最新一条是接口的缺陷，不是数据的缺陷——Owner 2026-07-21 点破后改为全量。
    by_skill: dict[str, list] = {}
    for r in rows:
        by_skill.setdefault(str(r.get("skill") or "?"), []).append(r)
    for v in by_skill.values():
        v.sort(key=lambda r: str(r.get("ts")), reverse=True)

    逐项 = []
    for skill, 约定 in sorted(SCHEDULE_CONTRACT.items()):
        if skill == "daily-funds":
            逐项.append(_daily_funds_schedule_row())
            continue
        # 跟 /public-api/技能健康 用**同一条**分法：结论只看排程跑，压测跑单独一栏。
        # 两个健康端点必须一致——一个修一个不修，会给出互相矛盾的信号，
        # 那比两个都错还糟：看到分歧的人只能猜哪个是真的。
        # 压测跑（entrypoint 部署后的全量压测）问「机器还转不转」，
        # 时间锚定的技能被拉到窗口外跑必然合法失败，那不是排程有问题。
        # 老台账行没有 sweep 字段，一律按排程跑处理。
        全部 = by_skill.get(skill, [])
        history = [r for r in 全部 if not r.get("sweep")]
        压测历史 = [r for r in 全部 if r.get("sweep")]
        last = history[0] if history else None
        距今小时 = None
        if last:
            try:
                距今小时 = round(
                    (now - datetime.fromisoformat(str(last["ts"]))).total_seconds() / 3600, 1)
            except (ValueError, KeyError, TypeError):
                距今小时 = None
        失败次数 = sum(1 for r in history if r.get("rc") != 0)
        连续失败 = 0
        for r in history:
            if r.get("rc") != 0:
                连续失败 += 1
            else:
                break
        逐项.append({
            "技能": skill,
            "业务模块": SKILL_MODULE.get(skill, "系统底座"),
            "约定时刻": 约定,
            "跑过": last is not None,
            "最近一次": (last or {}).get("ts"),
            "距今小时": 距今小时,
            "退出码": (last or {}).get("rc"),
            "成功": (last or {}).get("rc") == 0 if last else None,
            # 失败码：rc 只说「投递没成功」，而那对应十来种完全不同的原因。
            # 没有它，修一个失败技能就只能改一版等一天看会不会变绿——考勤为此拖了一个月。
            "失败码": (_public_failure_code((last or {}).get("code")) if last and last.get("rc") else None),
            "投递开关": (last or {}).get("delivery_enabled"),
            "次数": len(history),
            "失败次数": 失败次数,
            "成功率": (round(100 * (len(history) - 失败次数) / len(history)) if history else None),
            "连续失败": 连续失败,
            # 全量运行历史（最近在前，封顶 100 条防大账本撑爆响应；快照=当次独立日志文件）
            "历史": [{
                "ts": r.get("ts"), "rc": r.get("rc"), "成功": r.get("rc") == 0,
                "失败码": (_public_failure_code(r.get("code")) if r.get("rc") else None),
                "投递开关": r.get("delivery_enabled"), "快照": r.get("log"),
                # 结果摘要=当次日志最后一行有内容的输出（Owner：「skills 的结果呢」——
                # 不点快照也要能看到这次跑出了什么）。只取最近 8 条，读文件 IO 有界。
                "摘要": (_log_tail_line(r.get("log")) if i < 8 else None),
            } for i, r in enumerate(history[:100])],
            # 压测结果一条不藏——藏了就等于没做压测，回到「改一版等一天」。
            "压测": ({
                "最近一次": 压测历史[0].get("ts"),
                "退出码": 压测历史[0].get("rc"),
                "成功": 压测历史[0].get("rc") == 0,
                "失败码": (_public_failure_code(压测历史[0].get("code"))
                          if 压测历史[0].get("rc") else None),
                "次数": len(压测历史),
                "摘要": _log_tail_line(压测历史[0].get("log")),
            } if 压测历史 else None),
        })

    跑过的 = [x for x in 逐项 if x["跑过"]]
    失败的 = [x for x in 跑过的 if x["成功"] is False]
    空跑的 = [x for x in 跑过的 if str(x["投递开关"]) == "0"]
    # 近 24 小时战报：首页要一眼看到「昨晚到现在都跑了什么、成没成、干了什么」
    近24 = []
    for x in 逐项:
        for h in (x.get("历史") or []):
            try:
                if (now - datetime.fromisoformat(str(h["ts"]))).total_seconds() <= 86400:
                    近24.append({"技能": x["技能"], "业务模块": x["业务模块"], **h})
            except (ValueError, TypeError):
                continue
    近24.sort(key=lambda r: str(r.get("ts")), reverse=True)
    return {
        "可读": True,
        "近24小时": 近24[:30],
        "总执行次数": len(rows),
        "有记录的技能数": f"{len(跑过的)}/{len(SCHEDULE_CONTRACT)}",
        "失败数": len(失败的),
        "仍在空跑数": len(空跑的),
        "结论": (
            "从未执行过任何排程——排程链是断的" if not 跑过的
            else f"有 {len(失败的)} 个技能最近一次失败" if 失败的
            else f"有 {len(空跑的)} 个技能仍按空跑（投递开关=0），消息发不出去" if 空跑的
            else "最近一次执行全部成功且投递已开"
        ),
        "逐项": 逐项,
        "每日资金": {
            **_daily_funds_status(),
            "业务流": _daily_funds_flow_state(),
        },
        "诚实边界": "只报执行事实，不报业务内容；「没记录」一律显示为未跑过，不美化。",
    }


@app.get("/api/排程健康/快照")
def schedule_run_snapshot(log: str):
    """读取某一次运行的日志快照（run_skill.sh 每次运行写独立时间戳日志，即当时快照）。

    只许读排程日志根目录之内的文件——路径防穿越与 /assets 同款收紧；
    只回尾部 64KB：快照是给人复盘的，不是给人下载全量日志的。
    """
    log_root = SKILL_LEDGER_PATH.parent.resolve()
    target = Path(log).resolve()
    if not str(target).startswith(str(log_root) + "/") or not target.is_file():
        raise HTTPException(status_code=404, detail="快照不存在或不在排程日志目录内")
    data = target.read_bytes()
    tail = data[-65536:]
    return {
        "路径": str(target),
        "总字节": len(data),
        "截取": len(tail) < len(data),
        "内容": tail.decode("utf-8", errors="replace"),
    }
