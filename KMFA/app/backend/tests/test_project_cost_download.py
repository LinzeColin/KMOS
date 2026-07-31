# -*- coding: utf-8 -*-
"""Single-project downloads must preserve the website's closed-cost boundary."""

from __future__ import annotations

import io
import json

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import main
from app.main import app
from app.project_statement import statement_rows


client = TestClient(app)


def _runtime() -> dict:
    return {
        "schema_version": "kmfa.project_cost.current.v4",
        "快照ID": "kmfa-pc-2099-download",
        "计算状态": "PASS_WITH_OPEN_REVIEWS",
        "待确认": {
            "状态": "PASS_WITH_OPEN_REVIEWS",
            "P0阻断数": 0,
            "P1开放复核数": 1,
            "P2已排除或提示数": 0,
        },
        "项目数": 2,
        "封印来源": {
            "源码摘要算法": "kmfa.project_cost.subject_tree.v1",
            "源码SHA256": "a" * 64,
            "源码文件数": 1,
            "输入清单类型": "PRIVATE_MANIFEST_SHA256",
            "输入清单SHA256": "b" * 64,
            "私有输入清单SHA256": "b" * 64,
            "选中来源绑定SHA256": "c" * 64,
        },
        "封印工作簿": {
            "文件名": "sealed-download.xlsx",
            "SHA256": "d" * 64,
            "字节数": 1,
            "快照ID": "kmfa-pc-2099-download",
        },
        "生成时间": "2099-02-05T00:00:00+08:00",
        "项目": [
            {
                "合同编号": "SYNTHETIC-READY",
                "项目名称": "合成闭合项目",
                "甲方名称": "合成客户甲",
                "含税合同金额": "3000.00",
                "有效合同额": "2500.00",
                "项目已发生成本": "900.00",
                "项目成本": "1000.00",
                "毛利": "1500.00",
                "毛利率": "60.00%",
                "毛利率基点": 6000,
                "收入与毛利状态": "READY",
                "报表归类": {
                    "material": "100.00",
                    "fuel_power": "50.00",
                    "rental": "100.00",
                    "logistics": "50.00",
                    "travel": "25.00",
                    "lodging": "25.00",
                    "living": "25.00",
                    "road_parking": "25.00",
                    "vehicle": "25.00",
                    "other": "25.00",
                    "own_labor": "200.00",
                    "subcontract_labor": "200.00",
                    "information_fee": "50.00",
                    "tax": "100.00",
                },
                "项目成本覆盖": "FULL_SELECTED_GL_PERIOD;POSTING_PRESENT",
            },
            {
                "合同编号": "SYNTHETIC-BLOCKED",
                "项目名称": "合成未闭合项目",
                "甲方名称": "合成客户乙",
                "含税合同金额": "2000.00",
                "项目已发生成本": "700.00",
                "项目成本": None,
                "毛利": None,
                "毛利率": None,
                "毛利率基点": None,
                "收入与毛利状态": "BLOCKED_COST_COMPLETENESS",
                "报表归类": {"material": "700.00"},
                "项目成本覆盖": "FULL_SELECTED_GL_PERIOD;POSTING_PRESENT",
            },
        ],
    }


def _write_runtime(tmp_path, monkeypatch) -> dict:
    payload = _runtime()
    path = tmp_path / "recent_completed.json"
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    monkeypatch.setattr(main, "RECENT_COST_PATH", path)
    return payload


def _amounts_by_label(workbook_bytes: bytes) -> dict[str, object]:
    book = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    try:
        sheet = book["项目财务分析表"]
        return {
            str(row[0].value): row[1].value
            for row in sheet.iter_rows()
            if row[0].value
        }
    finally:
        book.close()


def test_blocked_project_cannot_download_a_formal_statement(
    tmp_path,
    monkeypatch,
) -> None:
    payload = _write_runtime(tmp_path, monkeypatch)
    blocked = payload["项目"][1]
    response = client.get(
        "/项目成本/下载",
        params={"合同": blocked["合同编号"]},
    )
    assert response.status_code == 409
    assert "成本尚未闭合" in response.json()["detail"]
    assert "正式项目财务分析表" in response.json()["detail"]

    try:
        statement_rows(blocked)
    except ValueError as exc:
        assert "未闭合" in str(exc)
    else:  # pragma: no cover - fail-closed boundary
        raise AssertionError("blocked project unexpectedly rendered")


def test_page_only_links_closed_project_downloads(tmp_path, monkeypatch) -> None:
    _write_runtime(tmp_path, monkeypatch)
    body = client.get("/项目成本").text
    assert "/项目成本/下载?合同=SYNTHETIC-READY" in body
    assert "/项目成本/下载?合同=SYNTHETIC-BLOCKED" not in body
    assert "项目成本未闭合，不能生成正式项目财务分析表" in body


def test_ready_download_uses_closed_cost_and_margin_not_incurred_lower_bound(
    tmp_path,
    monkeypatch,
) -> None:
    payload = _write_runtime(tmp_path, monkeypatch)
    ready = payload["项目"][0]
    response = client.get(
        "/项目成本/下载",
        params={"合同": ready["合同编号"]},
    )
    assert response.status_code == 200
    amounts = _amounts_by_label(response.content)
    assert amounts["项目产值"] == 2500
    assert amounts["二、资金运用及各项支出"] == 1000
    assert amounts["（六）信息费"] == 50
    assert amounts["（七）税金"] == 100
    assert amounts["三 利润"] == 1500
    assert amounts["二、资金运用及各项支出"] != 900

    book = load_workbook(io.BytesIO(response.content), data_only=True)
    try:
        notes = {
            row[0].value: row[1].value
            for row in book["口径"].iter_rows()
            if row[0].value
        }
        assert notes["已发生成本下限（非闭合项目成本）"] == "900.00"
        assert notes["闭合项目成本"] == "1000.00"
        assert notes["毛利率"] == "60.00%"
    finally:
        book.close()
