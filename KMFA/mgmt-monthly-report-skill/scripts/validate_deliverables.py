#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
validate_deliverables.py

Purpose:
  Validate that the generated management report Excel and board PDF conform to the v6 baseline contract.

Usage:
  python scripts/validate_deliverables.py \
    --period 202607 \
    --input-dir ./inputs \
    --excel "./outputs/经营管理分析报表 202607.xlsx" \
    --pdf "./outputs/董事会经营分析摘要 202607.pdf" \
    --config ./config/v6_spec.json \
    --report "./outputs/自动验收报告 202607.json" \
    --strict

Exit codes:
  0 = passed
  2 = failed
"""
from __future__ import annotations

import argparse
import fnmatch
import json
import re
import sys
import zipfile
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    print(json.dumps({"status":"failed","errors":[f"openpyxl unavailable: {exc}"]}, ensure_ascii=False, indent=2))
    sys.exit(2)

MOJIBAKE_PATTERNS = ["�", "ä¸", "å", "æ", "çš", "Ã", "Â", "Ð", "Ñ"]
FORMULA_ERROR_PATTERNS = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!"]

@dataclass
class CheckResult:
    name: str
    status: str
    detail: str = ""

class Validator:
    def __init__(self, period: str, input_dir: Path, excel: Path, pdf: Path, config_path: Path, strict: bool):
        self.period = period
        self.input_dir = input_dir
        self.excel = excel
        self.pdf = pdf
        self.config_path = config_path
        self.strict = strict
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.checks: List[CheckResult] = []
        self.config = json.loads(config_path.read_text(encoding="utf-8"))
        self.wb = None

    def add(self, name: str, ok: bool, detail: str = "", warning: bool = False):
        if ok:
            self.checks.append(CheckResult(name, "passed", detail))
            return
        if warning and not self.strict:
            self.warnings.append(f"{name}: {detail}")
            self.checks.append(CheckResult(name, "warning", detail))
        else:
            self.errors.append(f"{name}: {detail}")
            self.checks.append(CheckResult(name, "failed", detail))

    def run(self) -> Dict[str, Any]:
        self.check_period()
        self.check_file_names()
        self.check_inputs()
        self.load_excel()
        if self.wb is not None:
            self.check_sheet_order()
            self.check_tab_colors()
            self.check_forbidden_terms_and_errors()
            self.check_no_series_in_xml()
            self.check_charts()
            self.check_percent_formats()
            self.check_funds_sheet_formulas()
            self.check_display_risk()
        self.check_pdf()
        status = "passed" if not self.errors else "failed"
        return {
            "status": status,
            "period": self.period,
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "excel": str(self.excel),
            "pdf": str(self.pdf),
            "errors": self.errors,
            "warnings": self.warnings,
            "checks": [asdict(c) for c in self.checks],
        }

    def check_period(self):
        ok = bool(re.fullmatch(r"20\d{4}", self.period))
        self.add("period_format", ok, "period must be YYYYMM, e.g. 202607")

    def check_file_names(self):
        expected_excel = self.config["output_files"]["excel"].replace("{YYYYMM}", self.period)
        expected_pdf = self.config["output_files"]["pdf"].replace("{YYYYMM}", self.period)
        self.add("excel_file_exists", self.excel.exists(), f"missing {self.excel}")
        self.add("pdf_file_exists", self.pdf.exists(), f"missing {self.pdf}")
        self.add("excel_file_name", self.excel.name == expected_excel, f"expected {expected_excel}, got {self.excel.name}")
        self.add("pdf_file_name", self.pdf.name == expected_pdf, f"expected {expected_pdf}, got {self.pdf.name}")

    def check_inputs(self):
        if not self.input_dir.exists():
            self.add("input_dir", False, f"missing input dir {self.input_dir}")
            return
        files = [p.name for p in self.input_dir.iterdir() if p.is_file()]
        for group in self.config.get("input_business_groups", []):
            matches = []
            for pattern in group.get("patterns", []):
                matches.extend([f for f in files if fnmatch.fnmatch(f, pattern)])
            matches = sorted(set(matches))
            ok = bool(matches) or not group.get("required", True)
            self.add(f"input_group_{group['id']}", ok, f"{group['display_name']} matches={matches}")
            # If required sheet groups are specified, verify canonical sheets or accepted aliases in the first matched xlsx.
            if ok and group.get("required_sheet_groups") and matches:
                all_sheetnames = []
                read_errors = []
                non_xlsx = []
                for m in matches:
                    candidate = self.input_dir / m
                    if candidate.suffix.lower() != ".xlsx":
                        non_xlsx.append(candidate.name)
                        continue
                    try:
                        wb = load_workbook(candidate, read_only=True, data_only=False)
                        all_sheetnames.extend([f"{candidate.name}::{s}" for s in wb.sheetnames])
                    except Exception as exc:
                        read_errors.append(f"{candidate.name}: {exc}")
                if non_xlsx and not all_sheetnames:
                    self.add(f"input_group_{group['id']}_xlsx", False, f"non-xlsx files={non_xlsx}; convert before processing")
                    continue
                if read_errors and not all_sheetnames:
                    self.add(f"input_group_{group['id']}_read", False, f"cannot read any candidate: {read_errors}")
                    continue
                missing = []
                matched = {}
                pure_names = [x.split('::',1)[1] for x in all_sheetnames]
                for sg in group["required_sheet_groups"]:
                    aliases = sg.get("aliases", [sg.get("canonical", "")])
                    hit = []
                    for alias in aliases:
                        hit.extend([s for s in pure_names if fnmatch.fnmatch(s, alias)])
                    hit = sorted(set(hit))
                    if not hit:
                        missing.append(sg.get("canonical", str(aliases)))
                    else:
                        matched[sg.get("canonical", str(aliases))] = hit
                self.add(f"input_group_{group['id']}_required_sheet_groups", not missing, f"missing={missing}; matched={matched}; files={matches}")
            elif ok and group.get("required_sheets") and matches:
                candidate = self.input_dir / matches[0]
                try:
                    wb = load_workbook(candidate, read_only=True, data_only=False)
                    missing = [s for s in group["required_sheets"] if s not in wb.sheetnames]
                    self.add(f"input_group_{group['id']}_required_sheets", not missing, f"missing sheets={missing}")
                except Exception as exc:
                    self.add(f"input_group_{group['id']}_read", False, f"cannot read {candidate.name}: {exc}")

    def load_excel(self):
        if not self.excel.exists():
            return
        try:
            self.wb = load_workbook(self.excel, data_only=False)
            self.add("excel_open", True, "opened workbook")
        except Exception as exc:
            self.add("excel_open", False, str(exc))
            self.wb = None

    def check_sheet_order(self):
        expected = self.config["visible_sheet_order"]
        actual = [ws.title for ws in self.wb.worksheets if ws.sheet_state == "visible"]
        self.add("visible_sheet_order", actual == expected, f"expected={expected}; actual={actual}")

    def check_tab_colors(self):
        expected = self.config.get("tab_colors", {})
        for name, color in expected.items():
            if name not in self.wb.sheetnames:
                self.add(f"tab_color_{name}", False, "sheet missing")
                continue
            ws = self.wb[name]
            actual = None
            if ws.sheet_properties.tabColor is not None:
                actual = ws.sheet_properties.tabColor.rgb
            self.add(f"tab_color_{name}", actual == color, f"expected={color}, actual={actual}")

    def iter_visible_cells(self):
        for ws in self.wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            for row in ws.iter_rows():
                for cell in row:
                    yield ws, cell

    def check_forbidden_terms_and_errors(self):
        forbidden = self.config.get("forbidden_visible_terms", [])
        found_forbidden = []
        found_errors = []
        found_mojibake = []
        found_hashes = []
        for ws, cell in self.iter_visible_cells():
            v = cell.value
            if v is None:
                continue
            text = str(v)
            for term in forbidden:
                if term and term in text:
                    found_forbidden.append(f"{ws.title}!{cell.coordinate}:{term}")
            for err in FORMULA_ERROR_PATTERNS:
                if err in text:
                    found_errors.append(f"{ws.title}!{cell.coordinate}:{err}")
            for pat in MOJIBAKE_PATTERNS:
                if pat in text:
                    found_mojibake.append(f"{ws.title}!{cell.coordinate}:{pat}")
            if "#####" in text:
                found_hashes.append(f"{ws.title}!{cell.coordinate}")
        self.add("forbidden_terms", not found_forbidden, "; ".join(found_forbidden[:20]))
        self.add("formula_error_text", not found_errors, "; ".join(found_errors[:20]))
        self.add("mojibake_text", not found_mojibake, "; ".join(found_mojibake[:20]))
        self.add("literal_hash_marks", not found_hashes, "; ".join(found_hashes[:20]))

    def check_no_series_in_xml(self):
        if not self.excel.exists():
            return
        found = []
        try:
            with zipfile.ZipFile(self.excel) as z:
                for name in z.namelist():
                    if not name.endswith((".xml", ".rels")):
                        continue
                    data = z.read(name).decode("utf-8", errors="ignore")
                    if re.search(r"Series\s*1|Series1|Series\s*2|Series2", data):
                        found.append(name)
        except Exception as exc:
            self.add("xml_series_scan", False, f"cannot scan xml: {exc}", warning=True)
            return
        self.add("xml_no_series_default", not found, f"found default series in {found[:10]}")

    def chart_title(self, chart) -> str:
        try:
            if chart.title is None:
                return ""
            if chart.title.tx and chart.title.tx.rich and chart.title.tx.rich.p:
                parts=[]
                for p in chart.title.tx.rich.p:
                    for r in p.r:
                        if r.t:
                            parts.append(str(r.t))
                return "".join(parts)
        except Exception:
            pass
        try:
            return str(chart.title)
        except Exception:
            return ""

    def check_charts(self):
        req = self.config.get("chart_requirements", {})
        for sheet, titles in req.items():
            if sheet not in self.wb.sheetnames:
                self.add(f"charts_{sheet}", False, "sheet missing")
                continue
            ws = self.wb[sheet]
            actual_titles = [self.chart_title(c) for c in getattr(ws, "_charts", [])]
            missing = [t for t in titles if t not in actual_titles]
            self.add(f"charts_{sheet}", not missing, f"missing={missing}; actual={actual_titles}")

    def check_percent_formats(self):
        bad = []
        for ws, cell in self.iter_visible_cells():
            fmt = str(cell.number_format or "")
            if "%" in fmt and "0.00%" not in fmt:
                bad.append(f"{ws.title}!{cell.coordinate}:{fmt}")
            if isinstance(cell.value, str):
                # Flag literal percentages that are not exactly two decimals, ignore placeholders.
                for match in re.findall(r"(?<![\d.])(\d+(?:\.\d+)?)%", cell.value):
                    if "." not in match or len(match.split(".")[-1]) != 2:
                        bad.append(f"{ws.title}!{cell.coordinate}:{match}%")
        self.add("percentage_two_decimals", not bad, "; ".join(bad[:25]))

    def find_row_containing(self, ws, text: str) -> Optional[int]:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and text in str(cell.value):
                    return cell.row
        return None

    def check_funds_sheet_formulas(self):
        sheet = "04_2026年资金汇总"
        if sheet not in self.wb.sheetnames:
            self.add("funds_sheet_exists", False, f"missing {sheet}")
            return
        ws = self.wb[sheet]
        inflow_row = self.find_row_containing(ws, "资金总流入合计") or self.find_row_containing(ws, "资金总流入")
        outflow_row = self.find_row_containing(ws, "资金总流出合计") or self.find_row_containing(ws, "资金总流出")
        net_row = 30
        net_label_found = any(ws.cell(net_row, c).value and "累计资金净流量" in str(ws.cell(net_row, c).value) for c in range(1, min(ws.max_column, 8)+1))
        self.add("funds_inflow_row", inflow_row is not None, f"row={inflow_row}")
        self.add("funds_outflow_row", outflow_row is not None, f"row={outflow_row}")
        self.add("funds_row30_label", net_label_found, f"row30 values={[ws.cell(net_row,c).value for c in range(1,min(ws.max_column,8)+1)]}")
        def formulas_in_row(row: Optional[int], start_col: int = 3) -> Tuple[int, int, List[str]]:
            if row is None:
                return (0, 0, [])
            formula_count = 0
            value_count = 0
            bad_cells=[]
            for c in range(start_col, ws.max_column+1):
                v = ws.cell(row, c).value
                if v is None:
                    continue
                value_count += 1
                if isinstance(v, str) and v.startswith("="):
                    formula_count += 1
                else:
                    # Treat month labels/text as non-data, but flag numeric hardcoded values.
                    if isinstance(v, (int, float)):
                        bad_cells.append(ws.cell(row,c).coordinate)
            return (formula_count, value_count, bad_cells)
        for label, row in [("funds_inflow_formula", inflow_row), ("funds_outflow_formula", outflow_row), ("funds_cumulative_net_formula", net_row)]:
            formula_count, value_count, bad_cells = formulas_in_row(row)
            ok = formula_count > 0 and not bad_cells
            self.add(label, ok, f"formula_count={formula_count}, value_count={value_count}, hardcoded_numeric={bad_cells[:10]}")

    def check_display_risk(self):
        risky=[]
        for ws in self.wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            used_cols = set()
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        used_cols.add(cell.column_letter)
            for col_letter, dim in ws.column_dimensions.items():
                if col_letter in used_cols and dim.width is not None and dim.width < 8:
                    risky.append(f"{ws.title}!{col_letter}:width={dim.width}")
            # Text too long for non-wrapped cells.
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    if len(text) >= 18 and not cell.alignment.wrap_text:
                        risky.append(f"{ws.title}!{cell.coordinate}:not_wrapped_len={len(text)}")
                    if isinstance(cell.value, (int, float)):
                        width = ws.column_dimensions[cell.column_letter].width or 8
                        if abs(float(cell.value)) >= 1000000 and width < 11:
                            risky.append(f"{ws.title}!{cell.coordinate}:numeric_width={width}")
        self.add("display_risk_heuristic", not risky, "; ".join(risky[:30]), warning=True)

    def check_pdf(self):
        if not self.pdf.exists():
            return
        size = self.pdf.stat().st_size
        self.add("pdf_size", size > 30_000, f"size={size}")
        try:
            data = self.pdf.read_bytes()[:2048]
            self.add("pdf_header", data.startswith(b"%PDF"), "PDF header check")
        except Exception as exc:
            self.add("pdf_read", False, str(exc))


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--period", required=True)
    ap.add_argument("--input-dir", required=True)
    ap.add_argument("--excel", required=True)
    ap.add_argument("--pdf", required=True)
    ap.add_argument("--config", required=True)
    ap.add_argument("--report", required=True)
    ap.add_argument("--strict", action="store_true")
    args = ap.parse_args(argv)

    v = Validator(
        period=args.period,
        input_dir=Path(args.input_dir),
        excel=Path(args.excel),
        pdf=Path(args.pdf),
        config_path=Path(args.config),
        strict=args.strict,
    )
    result = v.run()
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["status"] == "passed" else 2

if __name__ == "__main__":
    raise SystemExit(main())
