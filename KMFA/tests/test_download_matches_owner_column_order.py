# -*- coding: utf-8 -*-
"""网站全量下载必须与 canonical Skill 的封印工作簿完全同源。

单项目下载仍使用竖版参考模板；全量下载不再由 App 根据 JSON 另造一份横表。
"""
from __future__ import annotations

import hashlib
import io
import importlib
import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

REPO = Path(__file__).resolve().parents[2]
BACKEND = REPO / "KMFA/app/backend"
SHEETS = (
    "01_项目成本表",
    "02_成本明细",
    "03_生命周期对照",
    "04_收入与现金",
    "05_来源与核销",
    "06_差异与待确认",
    "07_项目身份",
    "08_运行说明",
)
SNAPSHOT = "kmfa-pc-2099-synthetic"

PROJECTS = [
    {
        "合同编号": "KMX2099001-001",
        "项目名称": "合成项目甲",
        "甲方名称": "合成客户甲",
        "施工状态": "已完工",
        "含税合同金额": "128000",
        "项目过账实际": "3614.25",
        "项目应计": "18500",
        "项目已发生成本": "22114.25",
        "有效合同额": None,
        "毛利": None,
        "毛利率": None,
        "毛利率基点": None,
        "收入与毛利状态": "BLOCKED_COST_COMPLETENESS",
        "报表归类": {
            "material": "774.25",
            "travel": "1470",
            "lodging": "1370",
            "own_labor": "18500",
        },
        "项目成本覆盖": "GL_SELECTED_THROUGH_2099-02;POSTING_PRESENT",
    },
    {
        "合同编号": "KMX2099001-002",
        "项目名称": "合成项目乙",
        "甲方名称": "合成客户乙",
        "项目过账实际": "9000",
        "项目应计": "0",
        "项目已发生成本": "9000",
        "含税合同金额": "50000",
        "有效合同额": None,
        "毛利": None,
        "毛利率": None,
        "毛利率基点": None,
        "收入与毛利状态": "BLOCKED_COST_COMPLETENESS",
        "报表归类": {"other": "9000"},
        "项目成本覆盖": "GL_SELECTED_THROUGH_2099-02;POSTING_PRESENT",
    },
]


def _client(tmp_path: Path):
    workbook_path = tmp_path / "sealed-canonical.xlsx"
    book = Workbook()
    book.active.title = SHEETS[0]
    for title in SHEETS[1:]:
        book.create_sheet(title)
    for sheet in book.worksheets:
        sheet["A1"] = "snapshot_id"
        sheet["B1"] = SNAPSHOT
    book[SHEETS[0]]["A3"] = "合成封印工作簿"
    book.save(workbook_path)
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()

    payload = {
        "schema_version": "kmfa.project_cost.current.v4",
        "生成时间": "2099-02-05T00:00:00+00:00",
        "快照ID": SNAPSHOT,
        "截至日期": "2099-02-05",
        "计算状态": "PASS",
        "项目数": len(PROJECTS),
        "封印来源": {
            "源码摘要算法": "kmfa.project_cost.subject_tree.v1",
            "源码SHA256": "a" * 64,
            "源码文件数": 1,
            "输入清单类型": "PRIVATE_MANIFEST_SHA256",
            "输入清单SHA256": "b" * 64,
            "私有输入清单SHA256": "b" * 64,
            "选中来源绑定SHA256": "c" * 64,
        },
        "待确认": {
            "状态": "PASS",
            "P0阻断数": 0,
            "P1开放复核数": 0,
            "P2已排除或提示数": 0,
        },
        "项目": PROJECTS,
        "封印工作簿": {
            "文件名": workbook_path.name,
            "SHA256": digest,
            "字节数": workbook_path.stat().st_size,
            "快照ID": SNAPSHOT,
        },
    }
    artifact = tmp_path / "recent_completed.json"
    artifact.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    sys.path.insert(0, str(BACKEND))
    os.environ["KMFA_RECENT_COST"] = str(artifact)
    import app.main as module  # noqa: PLC0415

    importlib.reload(module)
    from fastapi.testclient import TestClient  # noqa: PLC0415

    return (
        TestClient(module.app, raise_server_exceptions=False),
        module,
        workbook_path,
        digest,
    )


def test_full_download_returns_the_exact_sealed_bytes(tmp_path):
    client, _, workbook_path, digest = _client(tmp_path)
    response = client.get("/项目成本/下载")
    assert response.status_code == 200
    assert response.content == workbook_path.read_bytes()
    assert response.headers["X-KMFA-Workbook-SHA256"] == digest
    assert response.headers["X-KMFA-Snapshot-ID"] == SNAPSHOT


def test_full_download_keeps_the_canonical_eight_sheet_contract(tmp_path):
    client, _, _, _ = _client(tmp_path)
    response = client.get("/项目成本/下载")
    book = load_workbook(io.BytesIO(response.content), read_only=True)
    assert tuple(book.sheetnames) == SHEETS
    assert book[SHEETS[0]]["A3"].value == "合成封印工作簿"
    book.close()


def test_tampered_sealed_workbook_is_rejected(tmp_path):
    client, _, workbook_path, _ = _client(tmp_path)
    changed = bytearray(workbook_path.read_bytes())
    changed[-1] ^= 1
    workbook_path.write_bytes(bytes(changed))
    response = client.get("/项目成本/下载")
    assert response.status_code == 503
    assert "hash_mismatch" in response.text


def test_single_contract_download_is_the_vertical_statement(tmp_path):
    client, _, _, _ = _client(tmp_path)
    response = client.get(
        "/项目成本/下载",
        params={"合同": "KMX2099001-001"},
    )
    assert response.status_code == 200
    book = load_workbook(io.BytesIO(response.content), read_only=True)
    assert book.sheetnames[0] == "项目财务分析表"
    assert "信息表" not in book.sheetnames
    book.close()


def test_owner_status_columns_remain_available_for_template_mapping(tmp_path):
    _, module, _, _ = _client(tmp_path)
    assert len(module.OWNER_STATUS_COLUMNS) == 30
    assert module.OWNER_STATUS_COLUMNS[:3] == ("甲方名称", "省份", "合同号")
