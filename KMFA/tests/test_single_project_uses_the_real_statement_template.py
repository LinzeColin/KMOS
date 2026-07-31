# -*- coding: utf-8 -*-
"""单个项目下载 = 竖版《项目财务分析表》。

行序来自 canonical Skill 封印 PDF 使用的 B 系模板；测试数据全部为合成值。
参考 PDF 只用于验证布局，不作为计算输入，也不进入公开仓。正式金额仅来自
canonical Skill 的闭合项目成本分类，模板中的 2% 管理费在缺少合格政策时必须留空。
"""
from __future__ import annotations

import importlib
import hashlib
import io
import json
import os
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
BACKEND = REPO / "KMFA/app/backend"

#: B 系模板表体行序的独立公开安全转录；不 import 主程序常量，避免自证。
PDF_ROWS_AS_READ = [
    "一、合同额", "项目产值", "二、资金运用及各项支出",
    "（一）原材料", "采购材料",
    "（二）租赁费", "其中:1.机械费", "（三）保险费",
    "（四）现场管理费", "1.自有人员工资", "2.差旅费", "3.招待费",
    "4.运输费", "5.办公费", "6.房租", "7.水电费", "8.备用金", "9.其他费用",
    "（五）工资（承包费）支出", "外协人员工资", "外协人员生活费",
    "临时用工费用", "（六）信息费", "（七）税金",
    "（八） 分摊的管理费用（合同的2%）", "已发生尚未支付费用", "三 利润",
]

#: 使用合成正式事件分类逐格核对；真实项目名与金额不进入公开仓。
SAMPLE = {
    "schema_version": "kmfa.project_cost.current.v4",
    "生成时间": "2026-07-30T15:00:00+08:00",
    "快照ID": "kmfa-pc-2099-statement",
    "截至日期": "2099-12-31",
    "计算状态": "PASS",
    "项目数": 2,
    "封印来源": {
        "源码摘要算法": "kmfa.project_cost.subject_tree.v1",
        "源码SHA256": "a" * 64,
        "源码文件数": 1,
        "输入清单类型": "PRIVATE_MANIFEST_SHA256",
        "输入清单SHA256": "b" * 64,
        "私有输入清单SHA256": "b" * 64,
        "选中来源绑定SHA256": "c" * 64,
    },
    "待确认": {
        "状态": "PASS",
        "P0阻断数": 0,
        "P1开放复核数": 0,
        "P2已排除或提示数": 0,
    },
    "项目": [{
        "合同编号": "KMX20991222-085", "项目名称": "合成竣工项目甲", "甲方名称": "合成客户甲",
        "施工状态": "已完工", "开工时间": "2099-12-23 00:00:00", "完工日期": "2099-12-30",
        "含税合同金额": "45000", "材料费": "494.10", "交通费": "542.81",
        "生活住宿费": "1300", "其他费用": "56", "自有人工工时": "41",
        "劳务人工工时": "17", "项目过账实际": "11402.81", "项目应计": "24098.95",
        "项目已发生成本": "35501.76", "项目成本": "35501.76",
        "有效合同额": "45000", "收入桥": "0.00",
        "毛利": "9498.24", "毛利率": "21.11%",
        "毛利率基点": 2111, "收入与毛利状态": "READY",
        "报表归类": {"material": "494.10", "travel": "542.81", "lodging": "1300",
                  "other": "56", "own_labor": "24098.85", "subcontract_labor": "9010"},
        "项目成本覆盖": "FULL_SELECTED_GL_PERIOD;POSTING_PRESENT",
    }, {
        "合同编号": "KMX20991222-086", "项目名称": "合成竣工项目乙", "甲方名称": "合成客户乙",
        "含税合同金额": "50000", "项目过账实际": "9000", "项目应计": "0",
        "项目已发生成本": "9000", "有效合同额": None, "毛利": None,
        "毛利率": None, "毛利率基点": None,
        "收入与毛利状态": "BLOCKED_COST_COMPLETENESS",
        "报表归类": {"other": "9000"},
        "项目成本覆盖": "FULL_SELECTED_GL_PERIOD;POSTING_PRESENT",
    }],
}


def _client(tmp_path: Path):
    from openpyxl import Workbook  # noqa: PLC0415

    payload = json.loads(json.dumps(SAMPLE, ensure_ascii=False))
    workbook_path = tmp_path / "sealed-statement-source.xlsx"
    book = Workbook()
    book.active.title = "01_项目成本表"
    for title in (
        "02_成本明细",
        "03_生命周期对照",
        "04_收入与现金",
        "05_来源与核销",
        "06_差异与待确认",
        "07_项目身份",
        "08_运行说明",
    ):
        book.create_sheet(title)
    book.save(workbook_path)
    payload["封印工作簿"] = {
        "文件名": workbook_path.name,
        "SHA256": hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
        "字节数": workbook_path.stat().st_size,
        "快照ID": payload["快照ID"],
    }
    artifact = tmp_path / "recent_completed.json"
    artifact.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    sys.path.insert(0, str(BACKEND))
    os.environ["KMFA_RECENT_COST"] = str(artifact)
    import app.main as m  # noqa: PLC0415

    importlib.reload(m)
    from fastapi.testclient import TestClient  # noqa: PLC0415

    return TestClient(m.app, raise_server_exceptions=False)


def _statement(tmp_path: Path, key: str = "KMX20991222-085"):
    from openpyxl import load_workbook  # noqa: PLC0415

    r = _client(tmp_path).get("/项目成本/下载", params={"合同": key})
    assert r.status_code == 200, r.text[:300]
    book = load_workbook(io.BytesIO(r.content))
    return book, r


def _body(ws):
    """表体 → {行标签: (金额, 备注)}，并保留出现顺序。"""
    order, data = [], {}
    for row in ws.iter_rows(values_only=True):
        label = row[0]
        if not isinstance(label, str) or label not in PDF_ROWS_AS_READ:
            continue
        order.append(label)
        data[label] = (row[1] if len(row) > 1 else None,
                       row[2] if len(row) > 2 else None)
    return order, data


def test_it_is_the_vertical_statement_not_a_flat_row(tmp_path):
    """**本文件的正主。** 单项目件必须是竖版分析表，不是横表一行。"""
    book, _ = _statement(tmp_path)
    assert book.sheetnames[0] == "项目财务分析表", f"页签不对：{book.sheetnames}"
    # 第二版那个 30 列横表是**清单**格式，不该出现在单项目件里
    assert "信息表" not in book.sheetnames, "单项目件又变成项目清单格式了"


def test_the_row_order_is_the_pdfs_row_order(tmp_path):
    """行序、层级、编号、括号写法一字不改。"""
    book, _ = _statement(tmp_path)
    order, _ = _body(book["项目财务分析表"])
    assert order == PDF_ROWS_AS_READ, (
        "行序与真 PDF 不一致。\n"
        f"  少了：{[r for r in PDF_ROWS_AS_READ if r not in order]}\n"
        f"  多了/错序：{[r for r in order if r not in PDF_ROWS_AS_READ]}")


def test_the_amounts_land_on_the_pdfs_own_lines(tmp_path):
    """我有的数必须落在模版**它自己那一行**上，不是挪到别处。"""
    book, _ = _statement(tmp_path)
    _, data = _body(book["项目财务分析表"])
    for label, expected in (
        ("一、合同额", 45000),
        ("项目产值", 45000),
        ("（一）原材料", 494.10),
        ("1.自有人员工资", 24098.85),
        ("2.差旅费", 1842.81),
        ("6.房租", 1300.00),
        ("9.其他费用", 56.00),
        ("（五）工资（承包费）支出", 9010.00),
        ("二、资金运用及各项支出", 35501.76),
        ("三 利润", 9498.24),
    ):
        got = data[label][0]
        assert got == expected, f"{label}：出的 {got}，应为 {expected}"
    assert data["（八） 分摊的管理费用（合同的2%）"][0] is None
    assert data["（七）税金"][0] is None


def test_formal_cost_share_and_policy_blanks_are_explicit(tmp_path):
    """成本占比以闭合成本为分母；无依据的政策费不伪装成 0。"""
    book, _ = _statement(tmp_path)
    _, data = _body(book["项目财务分析表"])
    assert "1.39%" in (data["（一）原材料"][1] or "")
    assert data["（八） 分摊的管理费用（合同的2%）"][0] is None
    assert "留空" in (data["（八） 分摊的管理费用（合同的2%）"][1] or "")
    assert data["三 利润"][0] == 9498.24
    assert "毛利" in (data["三 利润"][1] or "")


def test_free_text_and_share_coexist_in_the_note(tmp_path):
    """工时说明与正式成本占比必须同时保留。"""
    book, _ = _statement(tmp_path)
    _, data = _body(book["项目财务分析表"])
    note = data["（五）工资（承包费）支出"][1] or ""
    assert "17" in note and "25.38%" in note, f"文字与比例没并存：{note!r}"


def test_lines_i_have_no_data_for_stay_empty(tmp_path):
    """没有的行**留空**。留空是「我不知道」，填 0 是「我说它是 0」。"""
    book, _ = _statement(tmp_path)
    _, data = _body(book["项目财务分析表"])
    for label in (
        "（二）租赁费",
        "（三）保险费",
        "（六）信息费",
        "（七）税金",
        "4.运输费",
        "（八） 分摊的管理费用（合同的2%）",
    ):
        assert data[label][0] in (None, ""), \
            f"{label} 我并没有这个数，却填了 {data[label][0]!r}"


def test_money_never_silently_disappears(tmp_path):
    """正式分类里的 ``other`` 在模板里没有对应行——不能因此当它不存在。

    它已经算进成本合计了；如果既不进任何小计、又不写明，
    表里的分项就永远加不出合计，而看表的人不知道差额去哪了。
    """
    book, _ = _statement(tmp_path)
    _, data = _body(book["项目财务分析表"])
    site = data["（四）现场管理费"]
    assert site[0] == 25997.66, f"（四）小计没含其他费用：{site[0]}"
    assert "正式成本 56.00" in (site[1] or ""), f"未细分成本没在备注里交代：{site[1]!r}"


def test_the_second_level_subtotal_is_the_closed_project_cost(tmp_path):
    """B 系二级总额必须等于网站同一项目的闭合项目成本。"""
    book, _ = _statement(tmp_path)
    _, data = _body(book["项目财务分析表"])
    assert data["二、资金运用及各项支出"][0] == 35501.76


def test_the_header_block_copies_the_pdfs_fields(tmp_path):
    """表头区：项目名称／合同编号／开工时间／完工时间，措辞照抄。"""
    book, _ = _statement(tmp_path)
    ws = book["项目财务分析表"]
    text = "\n".join(
        "\t".join("" if c is None else str(c) for c in row)
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True))
    for must in ("项目财务分析表", "项目名称：", "合同编号", "开工时间", "完工时间",
                 "金额（元）", "备注", "合成竣工项目甲", "KMX20991222-085"):
        assert must in text, f"表头缺「{must}」：\n{text}"
    assert "2099/12/23" in text and "2099/12/30" in text, "日期没按模版的斜杠写法"


def test_there_is_a_signature_line_like_the_pdf(tmp_path):
    """PDF 末行是「项目经理: 日期：…」——交付件要能签字。"""
    book, _ = _statement(tmp_path)
    ws = book["项目财务分析表"]
    tail = "\n".join(
        "\t".join("" if c is None else str(c) for c in row)
        for row in ws.iter_rows(min_row=ws.max_row - 3, max_row=ws.max_row, values_only=True))
    assert "项目经理" in tail and "日期" in tail, f"没有签字行：\n{tail}"


def test_the_full_download_is_the_canonical_sealed_format(tmp_path):
    """不带合同号时直接返回 Skill 的 8 页签封印工作簿。"""
    from openpyxl import load_workbook  # noqa: PLC0415

    r = _client(tmp_path).get("/项目成本/下载")
    book = load_workbook(io.BytesIO(r.content))
    assert book.sheetnames[0] == "01_项目成本表"
    assert len(book.sheetnames) == 8


def test_the_caliber_sheet_explains_where_each_number_came_from(tmp_path):
    """银行/税务要能看懂。口径放第二个页签，不挤占数据表。"""
    book, _ = _statement(tmp_path)
    assert book.sheetnames[0] == "项目财务分析表"
    assert "口径" in book.sheetnames
    text = "\n".join(
        "\t".join("" if c is None else str(c) for c in row)
        for row in book["口径"].iter_rows(values_only=True))
    assert "不是 0" in text, "没有写明空行的含义"
    assert "闭合项目成本" in text


def test_the_filename_says_it_is_a_statement(tmp_path):
    _, r = _statement(tmp_path)
    disposition = r.headers.get("content-disposition", "")
    assert "KMX20991222-085" in disposition
    assert "statement" in disposition or "%E5%88%86%E6%9E%90%E8%A1%A8" in disposition, \
        f"文件名没体现这是分析表：{disposition}"
