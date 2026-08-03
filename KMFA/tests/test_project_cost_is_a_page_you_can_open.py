# -*- coding: utf-8 -*-
"""项目成本得是一页**打开就是数**的网页，不是 JSON、不是要下载的文件。

Owner 2026-07-29：「我说了我只要我的项目成本！」「我没有看到你说的东西」
「你不要放在本地，你推上网上去」。

在此之前出口只有两种，两种都没送到：
  · `/public-api/项目成本` 是 JSON——人打开看到的是一屏花括号；
  · 发 Excel 文件——卡片可能根本没在对话里露出来。

所以判据是「访客用浏览器打开 `/project-cost`，看到的是表格」。历史
`/项目成本*` 与 `/public-api/项目成本*` 继续受 Cloudflare Access 与 origin JWT
保护，公开页只读且不提供重算。
"""
from __future__ import annotations

import importlib
import hashlib
import json
import os
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]

SAMPLE = {
    "schema_version": "kmfa.project_cost.current.v4",
    "生成时间": "2026-07-29T09:00:00+08:00",
    "快照ID": "kmfa-pc-2099-synthetic",
    "截至日期": "2099-07-30",
    "计算状态": "PASS",
    "项目数": 4,
    "封印来源": {
        "源码摘要算法": "kmfa.project_cost.subject_tree.v1",
        "源码SHA256": "a" * 64,
        "源码文件数": 1,
        "输入清单类型": "PRIVATE_MANIFEST_SHA256",
        "输入清单SHA256": "b" * 64,
        "私有输入清单SHA256": "b" * 64,
        "选中来源绑定SHA256": "c" * 64,
    },
    "待确认": {
        "状态": "PASS",
        "P0阻断数": 0,
        "P1开放复核数": 0,
        "P2已排除或提示数": 0,
    },
    "正式成本口径": "项目已发生成本 = 项目过账实际 + 合格应计",
    "项目": [
        {"合同编号": "KMX2099001-001", "项目名称": "合成项目甲", "甲方名称": "合成客户甲",
         "施工状态": "已完工", "完工日期": "2099-06-30", "含税合同金额": "100000",
         "项目过账实际": "40000", "项目应计": "7000", "项目已发生成本": "47000",
         "项目成本": "47000",
         "有效合同额": "100000", "收入桥": "0.00",
         "毛利": "53000", "毛利率": "53.00%",
         "毛利率基点": 5300, "收入与毛利状态": "READY",
         "主营成本已结转": "39000", "状态表已报直接成本": "12000",
         "支付系统已付观察": "9000", "项目成本覆盖": "FULL_SELECTED_GL_PERIOD;POSTING_PRESENT",
         "账簿截至月份": "2099-06",
         "报表归类": {"material": "10000", "other": "30000", "subcontract_labor": "7000"}},
        {"合同编号": "KMX2099001-002", "项目名称": "合成项目乙", "甲方名称": "合成客户乙",
         "施工状态": "施工中", "含税合同金额": "50000",
         "项目过账实际": "-9000", "项目应计": "0", "项目已发生成本": "-9000",
         "项目成本": None,
         "有效合同额": None, "毛利": None, "毛利率": None,
         "毛利率基点": None, "收入与毛利状态": "BLOCKED_COST_COMPLETENESS",
         "项目成本覆盖": "FULL_SELECTED_GL_PERIOD;POSTING_PRESENT",
         "账簿截至月份": "2099-06",
         "报表归类": {"other": "-9000"}},
        {"合同编号": "KMX2099001-003", "项目名称": "合成项目丙", "甲方名称": "合成客户丙",
         "施工状态": "待入场", "含税合同金额": "80000",
         "项目过账实际": "0", "项目应计": "0", "项目已发生成本": "0",
         "项目成本": None,
         "有效合同额": None, "毛利": None, "毛利率": None,
         "毛利率基点": None, "收入与毛利状态": "BLOCKED_COST_COMPLETENESS",
         "项目成本覆盖": "FULL_SELECTED_GL_PERIOD;NO_QUALIFIED_EVENT",
         "账簿截至月份": "2099-06",
         "报表归类": {}},
        {"合同编号": "KMX2099001-004", "项目名称": "合成项目丁", "甲方名称": "合成客户丁",
         "项目已发生成本": None, "项目成本": None,
         "有效合同额": None, "毛利": None, "毛利率": None,
         "毛利率基点": None, "收入与毛利状态": "BLOCKED_COST_COMPLETENESS",
         "项目成本覆盖": "SOURCE_UNAVAILABLE",
         "报表归类": {}},
    ],
}


def _client(tmp_path: Path):
    from openpyxl import Workbook  # noqa: PLC0415

    payload = json.loads(json.dumps(SAMPLE, ensure_ascii=False))
    workbook_path = tmp_path / "sealed-canonical.xlsx"
    book = Workbook()
    book.active.title = "01_项目成本表"
    for title in (
        "02_成本明细",
        "03_生命周期对照",
        "04_收入与现金",
        "05_来源与核销",
        "06_差异与待确认",
        "07_项目身份",
        "08_运行说明",
    ):
        book.create_sheet(title)
    book["01_项目成本表"]["A1"] = "合成封印工作簿"
    book.save(workbook_path)
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    payload["封印工作簿"] = {
        "文件名": workbook_path.name,
        "SHA256": digest,
        "字节数": workbook_path.stat().st_size,
        "快照ID": payload["快照ID"],
    }
    artifact = tmp_path / "recent_completed.json"
    artifact.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    sys.path.insert(0, str(REPO / "KMFA/app/backend"))
    os.environ["KMFA_RECENT_COST"] = str(artifact)
    import app.main as m  # noqa: PLC0415

    importlib.reload(m)
    from fastapi.testclient import TestClient  # noqa: PLC0415

    return TestClient(m.app, raise_server_exceptions=False)


def test_it_returns_html_not_json(tmp_path):
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/html"), \
        "还是 JSON——人打开看到的是一屏花括号"
    assert "<table" in r.text


def test_legacy_public_api_name_is_still_access_protected(tmp_path):
    """兼容路径名不等于匿名授权；生产 origin guard 必须覆盖它。"""
    from app import main as m, private_access  # noqa: PLC0415

    _client(tmp_path)
    paths = {getattr(r, "path", "") for r in m.app.routes}
    assert "/public-api/项目成本表" in paths
    assert any(
        "/public-api/项目成本".startswith(prefix)
        for prefix in private_access.PRIVATE_PATH_ROOTS
    )


def test_the_numbers_are_actually_on_the_page(tmp_path):
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert "KMX2099001-001" in r.text
    assert "47,000.00" in r.text, "项目已发生成本没渲出来"
    assert "合成项目甲" in r.text


def test_a_negative_cost_project_is_not_filtered_out(tmp_path):
    """成本为负 = 金蝶红字冲销超过借方，那是最该被看见的一条。

    按「> 0」过滤会把它连同金额一起从页面和合计里抹掉——
    这条线整晚都在修的就是这种静默过滤。
    """
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert "KMX2099001-002" in r.text, "负成本项目被滤掉了"
    assert "-9,000.00" in r.text


def test_verified_zero_and_missing_are_both_shown_but_not_conflated(tmp_path):
    """0.00 是完整期间无事件；空值才是来源不可用。两者都要看见。"""
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert "KMX2099001-003" in r.text
    assert "KMX2099001-004" in r.text
    assert "账簿截至 2099-06｜无合格过账" in r.text
    assert "来源不可用" in r.text
    assert "0.00" in r.text


def test_all_canonical_projects_are_kept_on_the_page(tmp_path):
    r = _client(tmp_path).get("/public-api/项目成本表")
    for suffix in ("001", "002", "003", "004"):
        assert f"KMX2099001-{suffix}" in r.text


def test_it_says_so_when_there_is_no_artifact(tmp_path):
    """读不到就说读不到，不拿空表冒充「没有项目」。"""
    sys.path.insert(0, str(REPO / "KMFA/app/backend"))
    os.environ["KMFA_RECENT_COST"] = str(tmp_path / "nope.json")
    import app.main as m  # noqa: PLC0415

    importlib.reload(m)
    from fastapi.testclient import TestClient  # noqa: PLC0415

    r = TestClient(m.app, raise_server_exceptions=False).get("/public-api/项目成本表")
    assert r.status_code == 200
    assert "还没有数" in r.text
    assert "<table" not in r.text, "读不到却渲了一张空表"


def test_it_is_not_indexed(tmp_path):
    """真实客户名与合同额在这一页上，绝不能进搜索引擎。"""
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert "noindex" in r.headers.get("X-Robots-Tag", "")
    # 断包含不断相等：中间件会再追加 no-transform（防边缘层注入第三方脚本），
    # 写死相等会把「中间件在干活」判成回归。
    assert "no-store" in r.headers.get("Cache-Control", "")


def test_margin_column_is_explicit_and_incomplete_rows_stay_blocked(tmp_path):
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert ">毛利率<" in r.text
    assert ">毛利<" in r.text
    assert "项目成本（已闭合）" in r.text
    assert "已发生成本（下限）" in r.text
    assert "53.00%" in r.text
    assert "待成本闭合" in r.text


def test_unclosed_headline_counts_blocked_rows_not_only_missing_incurred_cost(tmp_path):
    """已有下限成本不等于成本闭合；顶部必须如实显示 1 READY、3 BLOCKED。"""
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert re.search(
        r'<div class="st"><b>1</b><span>毛利口径已闭合</span></div>',
        r.text,
    )
    assert re.search(
        r'<div class="st"><b>3</b><span>成本尚未闭合</span></div>',
        r.text,
    )


def test_requested_observation_and_posting_columns_are_removed(tmp_path):
    r = _client(tmp_path).get("/public-api/项目成本表")
    for removed in (
        "支付观察",
        "主营成本结转",
        "状态表观察",
        ">过账实际<",
        ">合格应计<",
    ):
        assert removed not in r.text


def test_react_dashboard_no_longer_fetches_or_renders_legacy_margin_upper_bound():
    source = (
        REPO / "KMFA/app/frontend/src/App.jsx"
    ).read_text(encoding="utf-8")
    assert "/api/项目毛利" not in source
    assert "毛利上限率" not in source
    assert "项目口径毛利（含在建）" not in source


def test_runtime_with_margin_above_seventy_percent_is_rejected(tmp_path):
    payload = json.loads(json.dumps(SAMPLE, ensure_ascii=False))
    first = payload["项目"][0]
    first.update(
        {
            "有效合同额": "100000",
            "项目已发生成本": "20000",
            "项目成本": "20000",
            "毛利": "80000",
            "毛利率": "80.00%",
            "毛利率基点": 8000,
            "收入与毛利状态": "READY",
        }
    )
    from openpyxl import Workbook  # noqa: PLC0415

    workbook_path = tmp_path / "sealed-over-limit.xlsx"
    workbook = Workbook()
    workbook.save(workbook_path)
    payload["封印工作簿"] = {
        "文件名": workbook_path.name,
        "SHA256": hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
        "字节数": workbook_path.stat().st_size,
        "快照ID": payload["快照ID"],
    }
    artifact = tmp_path / "over-limit.json"
    artifact.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    sys.path.insert(0, str(REPO / "KMFA/app/backend"))
    os.environ["KMFA_RECENT_COST"] = str(artifact)
    import app.main as m  # noqa: PLC0415

    importlib.reload(m)
    from fastapi.testclient import TestClient  # noqa: PLC0415

    response = TestClient(m.app, raise_server_exceptions=False).get(
        "/public-api/项目成本表"
    )
    assert response.status_code == 503
    assert "runtime_gross_margin_above_release_limit" in response.text


def _page_script(client) -> str:
    """跟着页面真正引的 <script src> 去取脚本——浏览器就是这么拿到它的。

    2026-07-29 教训：这两条断言原本写成 `in r.text`，于是
    「aria-sort」命中的是 CSS 规则、「空值永远沉底」命中的是脚本里的注释，
    而那整块脚本当时是内联的、被 CSP 拒绝执行，排序**从来没能用过**。
    在 HTML 文本里找字符串，证明不了浏览器会执行它。
    """
    html = client.get("/public-api/项目成本表").text
    srcs = re.findall(r"""<script[^>]*\bsrc\s*=\s*["']([^"']+)["']""", html, re.I)
    assert srcs, "页面没引任何脚本——排序是靠 JS 活的"
    return "\n".join(client.get(s).text for s in srcs)


def test_numeric_columns_sort_by_value_not_by_the_printed_string(tmp_path):
    """排序读 data-v 的数值。按显示的千分位字符串排，「1,000,000」会排在「9,000」前面。"""
    client = _client(tmp_path)
    r = client.get("/public-api/项目成本表")
    assert 'data-v="47000.0"' in r.text or 'data-v="47000"' in r.text
    assert "th[data-s]" in r.text, "表头没做成可点"
    js = _page_script(client)
    assert "parseFloat" in js, "数值列没按数值比，会退化成字符串序"
    assert "aria-sort" in js, "排序状态没写回表头"


def test_empty_cells_always_sink_to_the_bottom(tmp_path):
    """空值不管升序降序都沉底——把「没有数」排到有数的前面等于让缺失冒充最小值。"""
    js = _page_script(_client(tmp_path))
    assert "空值永远沉底" in js
    # 沉底靠的是「不参与升降序取反」：命中空值时直接返回定值
    assert "return 1;" in js and "return -1;" in js, "空值分支没有绕开升降序取反"


def test_there_is_a_download_button(tmp_path):
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert 'href="/项目成本/下载"' in r.text


def test_the_download_is_the_canonical_sealed_workbook(tmp_path):
    """全量下载必须返回 Skill 封印的 8 页签工作簿。"""
    import io

    import openpyxl

    r = _client(tmp_path).get("/项目成本/下载")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    book = openpyxl.load_workbook(io.BytesIO(r.content))
    assert book.sheetnames == [
        "01_项目成本表",
        "02_成本明细",
        "03_生命周期对照",
        "04_收入与现金",
        "05_来源与核销",
        "06_差异与待确认",
        "07_项目身份",
        "08_运行说明",
    ]
    assert book["01_项目成本表"]["A1"].value == "合成封印工作簿"


def test_the_download_filename_survives_chinese(tmp_path):
    """中文文件名要走 RFC 5987，否则浏览器存下来是一串下划线或乱码。"""
    r = _client(tmp_path).get("/项目成本/下载")
    disposition = r.headers.get("content-disposition", "")
    assert "filename*=utf-8''" in disposition.lower()
    assert "attachment" in disposition


def test_the_download_says_no_when_there_is_no_artifact(tmp_path):
    import importlib
    import os
    import sys

    sys.path.insert(0, str(REPO / "KMFA/app/backend"))
    os.environ["KMFA_RECENT_COST"] = str(tmp_path / "nope.json")
    import app.main as m  # noqa: PLC0415

    importlib.reload(m)
    from fastapi.testclient import TestClient  # noqa: PLC0415

    r = TestClient(m.app, raise_server_exceptions=False).get("/项目成本/下载")
    assert r.status_code == 503, "产物不在却给了一个文件——那文件里会是空的"


def test_the_homepage_link_survives_react_hydration():
    """链接必须写在 React 壳里，不能只写在静态壳里。

    2026-07-29 Owner：「和主页没有连接在一起」。此前链接只在 index.html 的静态壳，
    浏览器一加载 JS，React 接管就把整块换掉——「首页有链接」只在没有 JS 时成立。
    """
    shell = (REPO / "KMFA/app/frontend/src/PublicAppShell.jsx").read_text(encoding="utf-8")
    assert shell.count('href="/project-cost"') >= 2, "React 壳里没有项目成本入口"
    assert "data-shell-cost-entry" in shell, "入口没有可测锚点"

    built = list((REPO / "KMFA/app/frontend/dist/assets").glob("PublicAppShell-*.js"))
    assert built, "找不到构建产物"
    assert any("shell-cost-entry" in f.read_text(encoding="utf-8", errors="replace")
               for f in built), "改了源码但没重新构建，线上还是旧的"


def test_the_entry_is_in_the_component_the_root_actually_renders():
    """入口必须在 App.jsx——**根路径渲染的是它**，不是 PublicAppShell。

    2026-07-29 栽在这上面：入口写进了 PublicAppShell.jsx，而 main.jsx 里
    只有 `/workspace` 才加载那个组件，根路径加载的是 App.jsx。
    我当时只 grep 了 JS 包里有没有那个字符串，**没验它渲不渲得出来**——
    包里有 ≠ 用户看得见。Owner 连着两次说「首页依旧进不去项目成本」。
    """
    main_js = (REPO / "KMFA/app/frontend/src/main.jsx").read_text(encoding="utf-8")
    assert "loadPrivateOperationsApp()" in main_js and "isPublicWorkspace" in main_js, \
        "路由结构变了——重新确认根路径到底渲染哪个组件，别再照着旧假设放入口"

    app_jsx = (REPO / "KMFA/app/frontend/src/App.jsx").read_text(encoding="utf-8")
    assert 'href="/project-cost"' in app_jsx, "根路径渲染的组件里没有项目成本入口"
    assert 'href="/project-cost/download"' in app_jsx, "根路径渲染的组件里没有下载入口"

    built = list((REPO / "KMFA/app/frontend/dist/assets").glob("App-*.js"))
    assert built, "找不到 App 的构建产物"
    assert any("/project-cost" in f.read_text(encoding="utf-8", errors="replace") for f in built), \
        "改了源码但没重新构建，线上还是旧的"


def test_only_closed_rows_can_be_downloaded_as_formal_statements(tmp_path):
    """未闭合下限不能经下载按钮重新命名成正式项目成本。"""
    r = _client(tmp_path).get("/public-api/项目成本表")
    assert r.text.count('class="one"') == 1
    assert "/项目成本/下载?合同=KMX2099001-001" in r.text
    for suffix in ("002", "003", "004"):
        assert f"/项目成本/下载?合同=KMX2099001-{suffix}" not in r.text


def test_a_single_contract_download_is_the_vertical_statement(tmp_path):
    """单合同件是竖版《项目财务分析表》——照抄 Owner 的「竣工项目财务报表」模版。

    2026-07-30 改向：本条上一版断言它是 30 列横表的一行。那是搞混了两种格式——
    横表是**项目清单**，单个项目要的是竖版分析表。
    """
    import io

    import openpyxl

    r = _client(tmp_path).get("/项目成本/下载", params={"合同": "KMX2099001-001"})
    assert r.status_code == 200
    book = openpyxl.load_workbook(io.BytesIO(r.content))
    assert book.sheetnames[0] == "项目财务分析表", f"页签不对：{book.sheetnames}"
    labels = [row[0] for row in book["项目财务分析表"].iter_rows(values_only=True)]
    for must in (
        "一、合同额",
        "项目产值",
        "二、资金运用及各项支出",
        "（七）税金",
        "三 利润",
    ):
        assert must in labels, f"模版缺行「{must}」"
    assert "KMX2099001-001" in r.headers.get("content-disposition", "")


def test_an_unknown_contract_is_a_404_not_an_empty_workbook(tmp_path):
    """找不到就说找不到。回一份空表比报错更糟——拿到手的人会以为这个项目成本是 0。"""
    r = _client(tmp_path).get("/项目成本/下载", params={"合同": "KMX-不存在"})
    assert r.status_code == 404


def test_recompute_only_files_a_request_and_never_runs_the_job_here(tmp_path, monkeypatch):
    """「重新计算」只放标记，真正的活在 skills 容器里跑。

    让 App 容器去跑克隆私有库解析上千张表的活，就是把 2026-07-28 那次
    「压测把线上打下线」原样重演一遍，只是换了个触发器。
    """
    c = _client(tmp_path)
    import app.main as m  # noqa: PLC0415

    flag = tmp_path / ".refresh_requested"
    monkeypatch.setattr(m, "COST_REFRESH_FLAG", flag)
    r = c.post("/项目成本/重算")
    assert r.status_code == 200
    body = r.json()
    assert body["已提交"] is True
    assert body["怎么确认"], "只说「已提交」而不说怎么确认，跟没回一样"
    assert flag.exists(), "标记没落到共享卷上，skills 那边永远看不到"

    src = (REPO / "KMFA/app/backend/app/main.py").read_text(encoding="utf-8")
    tail = src.split("def public_project_cost_refresh")[1][:1400]
    assert "subprocess" not in tail, "重算端点里出现了子进程调用——重活不该在 App 容器里跑"


def test_the_flag_goes_where_the_app_can_actually_write():
    """标记必须落在 **app 可写**的卷上。

    2026-07-29 线上实测：第一版把它写进日志卷，按钮直接 503
    「共享卷不可写——app 容器没挂 kmfa-logs 卷」。两个卷的读写方向是刻意反着的：
      kmfa-logs       skills 可写 / app **只读**
      kmfa-app-state  app 可写   / skills **只读**（daily-backup 读它打包备份，「绝不写」）
    那道只读边界是有意的，不该为了一个按钮把整个日志卷对 app 开成可写。
    """
    src = (REPO / "KMFA/app/backend/app/main.py").read_text(encoding="utf-8")
    line = next(l for l in src.splitlines() if l.startswith("COST_REFRESH_FLAG"))
    assert "APP_STATE_DIR" in line, f"标记又放回 app 只读的卷上了：{line}"


def test_the_failure_message_names_the_volume_that_is_actually_involved(tmp_path, monkeypatch):
    """写失败时报的卷名必须是**真参与的那个**。

    标记从 kmfa-logs 挪到 kmfa-app-state 之后，这句报错没跟着改，
    于是失败信息把人指向一个根本没参与的卷。**指错地方的报错比不报还费时间**——
    照着它去查日志卷的挂载，会发现挂载完全正常，然后一无所获。
    """
    client = _client(tmp_path)
    from app import main as m  # noqa: PLC0415

    # 让写标记必然失败：把它指到一个「父目录是文件」的路径上
    blocker = tmp_path / "not_a_dir"
    blocker.write_text("x", encoding="utf-8")
    monkeypatch.setattr(m, "COST_REFRESH_FLAG", blocker / "flag")

    body = client.post("/项目成本/重算").json()
    assert body["已提交"] is False
    reason = body["原因"]
    assert "app-state" in reason, f"报错没提真正涉及的卷：{reason}"
    assert "kmfa-logs" not in reason, f"报错还在指向没参与的日志卷：{reason}"


def test_the_skills_side_actually_watches_for_the_flag():
    """标记要有人看。没人看的标记 = 按钮按下去什么也不会发生。"""
    cron = (REPO / "KMFA/deploy/skills-runtime/crontab.txt").read_text(encoding="utf-8")
    line = next((l for l in cron.splitlines()
                 if "refresh_requested" in l and not l.lstrip().startswith("#")), None)
    assert line, "crontab 里没有看标记的那一跳"
    assert line.startswith("* * * * *"), "不是每分钟看一次，按下去要等很久才有反应"
    assert "nice -n" in line, "重算没让出 CPU 优先级"
    # skills 只读挂载 app-state，**删不掉**请求标记，所以只能比时间戳。
    # 失败时可以删除自己在日志卷创建的 pending 副本，但不能删除 $R。
    assert 'rm -f "$R"' not in line, "skills 对 app-state 只读，删不掉请求标记"
    assert "-nt" in line, "没有比时间戳——那就分不出「这次点的」和「上次点的」"
    assert "/var/lib/kmfa/state/" in line and "/var/log/kmfa/" in line, \
        "两个卷都要用到：读 app 写的标记，写自己这边的处理记录"
    assert 'cp -p "$R" "$P"' in line, "运行前没有冻结本次请求时间戳"
    assert line.index('cp -p "$R" "$P"') < line.index("run_skill.sh project-cost-refresh"), \
        "必须先冻结请求时间戳，再开始长任务"
    assert line.index("run_skill.sh project-cost-refresh") < line.index('mv -f "$P" "$S"'), \
        "请求尚未成功就被标成已处理，失败后不会自动重试"


def test_project_cost_lock_contention_is_retryable_not_success():
    """flock 抢不到不能返回 0，否则 cron 会把未运行的请求标成成功。"""
    wrapper = (
        REPO / "KMFA/deploy/skills-runtime/run_skill.sh"
    ).read_text(encoding="utf-8")
    lock_guard = next(
        line for line in wrapper.splitlines()
        if "flock -n 9" in line and "上一轮仍在运行" in line
    )
    assert "exit 75" in lock_guard, "锁竞争必须返回 EX_TEMPFAIL，交给 cron 重试"
    assert "exit 0" not in lock_guard, "锁竞争没有产生新快照，不得冒充成功"


def test_project_cost_secrets_are_synthesised_for_cron():
    entrypoint = (
        REPO / "KMFA/deploy/skills-runtime/entrypoint.sh"
    ).read_text(encoding="utf-8")
    assert "KMFA_PRIVATE_DB_READ_TOKEN KMFA_PAYROLL_PASSWORD" in entrypoint
    assert "printf '%s=%q\\n'" in entrypoint, "secret 值没有 shell-safe 转义"


def test_healthcheck_requires_current_schema_binding_and_freshness():
    health = (
        REPO / "KMFA/deploy/skills-runtime/healthcheck.sh"
    ).read_text(encoding="utf-8")
    assert "36*3600" in health
    assert "kmfa.project_cost.current.v4" in health
    assert (
        'payload.get("计算状态") not in '
        '("PASS", "PASS_WITH_OPEN_REVIEWS")'
    ) in health
    assert "封印工作簿" in health
    assert 'row.get("skill") == "project-cost-refresh"' in health
