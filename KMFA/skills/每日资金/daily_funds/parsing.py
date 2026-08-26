"""Deterministic parsers for the two required daily-funds fact families.

There is deliberately no heuristic OCR in this module.  Only a bounded,
deterministic offline OCR fallback may open a supported image or scanned PDF;
unknown column shapes, MIME/magic mismatches and unsupported identifiers are
rejected with machine codes.  OCR results remain unsupported until real-sample
layout calibration proves the source profile, rather than silently inventing
financial facts.
"""

from __future__ import annotations

import csv
import io
import json
import re
import struct
import subprocess
import tempfile
from dataclasses import dataclass, replace
from decimal import Decimal, InvalidOperation, ROUND_FLOOR
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping

from .contracts import ContractError, parse_amount_to_fen
from .models import (
    CashflowObservation,
    PaymentRequestObservation,
    AccountSnapshot,
    ParsedFacts,
    ParserEvidence,
    SourceRef,
    Transaction,
)

ACCOUNT_FAMILY = "资金账户明细表"
TRANSACTION_FAMILIES = frozenset({"资金流水明细", "资金明细"})
# v13 adds the frozen task-pack's legacy ``.xls`` route without weakening the
# structured-source contract.  An old workbook opens only after its OLE
# container is intact, it exposes one ordinary worksheet, it has no macro
# streams or BIFF formula records, and the unchanged complete account or
# transaction schema passes.  This keeps cached formula values from becoming
# financial facts merely because xlrd can expose them as scalar cells.
#
# It retains v12's bounded deterministic OCR fallback and transaction identity
# requirements: bank_id is optional and a source-row fact identifier is used
# only when a source does not expose a transaction identifier.  It also
# requires an ambiguously titled structured ``资金明细`` document to open exactly
# one complete fact schema, matching the existing generic-image gate.
#
# It retains v5's narrow source-classification rule: a generic ``资金明细``
# image may be treated as an account snapshot only when its OCR table satisfies
# the account schema *and* cannot satisfy the transaction schema.  When both
# candidates fail at the same values-free OCR phase, v10 retains that bounded
# diagnosis for the protected capability receipt.  It never turns a failed
# candidate into a fact or relaxes either schema.  When a complete header is
# visually aligned but Tesseract splits its cells into separate lines, v10
# reassembles only those overlapping OCR lines and then applies the same exact
# aliases, confidence, row and fact rules.  Capability receipts are versioned,
# so a rule change cannot inherit a prior parser's production-support assertion.
PARSER_VERSION = "kmfa.daily_funds.parser.v13"
# This parser is deliberately separate from ``PARSER_VERSION``.  It can
# create a chart-only receipt from a narrow receipt/payment screenshot without
# weakening the two-fact account-balance publication contract.
# v5 keeps v4's bounded sparse-layout fallback and adds a source-family
# admission gate.  v7 permits the explicitly allowed ``资金明细`` document
# family to reach this *separate* chart-only parser even when it cannot form a
# formal account/transaction fact.  It still adds one *consensus-only*
# table-layout recovery pair.  It is reached only after both existing header
# passes are missing, and both recovery modes must independently satisfy the
# unchanged date, row, amount-confidence and footer-total rules with exactly
# the same result.  A single alternate OCR reading can therefore never create
# a chart point.  v8 adds a bounded headerless fixed-table recovery for the
# known source family only: it requires repeated same-day date cells, one
# visible ``合计`` footer, exactly two stable right-aligned money columns and exact
# independent OCR agreement.  v9 adds a deterministic grid-removal image
# normalization only after every original layout pass stops in a layout/OCR
# gate.  The normalized image must still pass two independent segmenters with
# exactly the same date and totals, then the unchanged row/footer checks.
# v10 adds a final contrast-preserving rendering repair after those layout
# paths stop at an allowed layout/OCR gate.  Fixed grayscale autocontrast,
# bounded Lanczos scaling and a fixed unsharp mask do not change OCR text,
# aliases, amount/date parsing, or the confidence threshold.  Both fixed
# segmenters still need to produce identical, footer-reconciled totals.
# Chart admission never produces a formal account-balance publication.
CASHFLOW_OBSERVATION_PARSER_VERSION = "kmfa.daily_funds.cashflow_observation.v11"
# This parser is deliberately a separate report family: the source image is a
# daily payment-request sheet, not a bank statement and not a completed cash
# receipt/payment flow.  It can expose one total only after three fixed OCR
# segmentations agree on the fixed fields required by its visual profile.  v2
# adds the approved horizontal message-summary profile.  It records the exact
# message day as its date basis because that compact profile has no visible
# document-date cell.
PAYMENT_REQUEST_OBSERVATION_PARSER_VERSION = "kmfa.daily_funds.payment_request_observation.v2"

_SHA256 = re.compile(r"^[0-9a-f]{64}$")
_OCCURRENCE_PATH = re.compile(
    r"^Private-KMDatabase/KMFA/daily_funds/raw/occurrences/"
    r"(?P<year>20\d{2})/(?P<month>0[1-9]|1[0-2])/"
    r"(?P<day>0[1-9]|[12]\d|3[01])/"
    r"(?P<message>[0-9a-f]{64})/(?P<index>0|[1-9]\d*)\.json$"
)
_MIME_TOKEN = re.compile(r"^[a-z0-9!#$&^_.+-]+/[a-z0-9!#$&^_.+-]+$")
_ZIP_MAGICS = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")
_TEXT_SUFFIXES = frozenset({".csv", ".txt"})
_WORKBOOK_SUFFIXES = frozenset({".xls", ".xlsx", ".xlsm"})
_ALLOWED_SUFFIXES = _TEXT_SUFFIXES | _WORKBOOK_SUFFIXES
_OCR_IMAGE_SUFFIXES = frozenset({".png", ".jpg", ".jpeg", ".bmp", ".webp"})
_OCR_PDF_SUFFIXES = frozenset({".pdf"})
_OCR_SUFFIXES = _OCR_IMAGE_SUFFIXES | _OCR_PDF_SUFFIXES
_CAPABILITY_SUFFIXES = _ALLOWED_SUFFIXES | frozenset({
    ".pdf", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp",
})
_CSV_MIME = frozenset({
    "text/csv",
    "application/csv",
    "text/plain",
    "application/vnd.ms-excel",
    "application/octet-stream",
})
_XLSX_MIME = frozenset({
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
})
_XLSM_MIME = frozenset({
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "application/octet-stream",
})
_XLS_MIME = frozenset({
    "application/vnd.ms-excel",
    "application/octet-stream",
})
_XLS_FORMULA_RECORD_IDS = frozenset({
    0x0006,  # FORMULA
    0x0221,  # ARRAY
    0x0236,  # TABLE
    0x04BC,  # SHRFMLA
})
_XLS_MACRO_COMPONENTS = frozenset({"macros", "vba", "_vba_project_cur"})
_XLS_BOUNDSHEET_RECORD = 0x0085
_XLS_FILEPASS_RECORD = 0x002F
_OCR_IMAGE_MIME = {
    ".png": frozenset({"image/png", "application/octet-stream"}),
    ".jpg": frozenset({"image/jpeg", "application/octet-stream"}),
    ".jpeg": frozenset({"image/jpeg", "application/octet-stream"}),
    ".bmp": frozenset({"image/bmp", "image/x-ms-bmp", "application/octet-stream"}),
    ".webp": frozenset({"image/webp", "application/octet-stream"}),
    ".pdf": frozenset({"application/pdf", "application/octet-stream"}),
}
_OCR_MAGIC_BY_SUFFIX = {
    ".png": "PNG",
    ".jpg": "JPEG",
    ".jpeg": "JPEG",
    ".bmp": "BMP",
    ".webp": "WEBP",
    ".pdf": "PDF",
}
_OCR_CANONICAL_SUFFIX_BY_MAGIC = {
    "PNG": ".png",
    "JPEG": ".jpg",
    "BMP": ".bmp",
    "WEBP": ".webp",
    "PDF": ".pdf",
}
OCR_MIN_CONFIDENCE_BPS = 9_800
OCR_MAX_ATTACHMENT_BYTES = 50 * 1024 * 1024
OCR_TIMEOUT_SECONDS = 90
OCR_LANGUAGE = "chi_sim+eng"
OCR_PRIMARY_PSM = 6
# PSM 11 is a bounded, deterministic sparse-text layout fallback.  It is not
# a general OCR search: formal parsing can reach it only after every required
# source-family candidate stopped at the exact missing-header gate under the
# primary table mode.
OCR_HEADER_FALLBACK_PSM = 11
# A bounded pair of alternative table segmenters used only for the cashflow
# observation's consensus recovery.  They are deliberately not a general
# retry set for formal facts, and cannot be reached after a visible row or
# footer failure.
OCR_CASHFLOW_CONSENSUS_PSMS = (4, 12)
_OCR_ALLOWED_PSMS = frozenset({OCR_PRIMARY_PSM, OCR_HEADER_FALLBACK_PSM, *OCR_CASHFLOW_CONSENSUS_PSMS})
# Grid screenshots are not a new source format.  These fixed, independent
# Tesseract modes are only used after the untouched image exhausted its
# existing layout-only recovery path.  Requiring both results avoids choosing
# whichever preprocessed reading produces a convenient number.
OCR_GRID_PREPROCESS_PSMS = (OCR_PRIMARY_PSM, OCR_HEADER_FALLBACK_PSM)
_OCR_GRID_THRESHOLD = 192
_OCR_GRID_LINE_COVERAGE_BPS = 7_000
_OCR_GRID_MAX_PIXELS = 20_000_000
_OCR_GRID_SCALE = 2
_OCR_ENHANCED_MAX_PIXELS = 20_000_000
_OCR_ENHANCED_SMALL_IMAGE_PIXELS = 2_200_000
_OCR_ENHANCED_MEDIUM_IMAGE_PIXELS = 5_000_000
_OCR_ENHANCED_SMALL_SCALE = 3
_OCR_ENHANCED_MEDIUM_SCALE = 2
_OCR_ENHANCED_SHARPEN_RADIUS = 2
_OCR_ENHANCED_SHARPEN_PERCENT = 175
_OCR_ENHANCED_SHARPEN_THRESHOLD = 2
_CASHFLOW_ORIGINAL_LAYOUT_RECOVERY_CODES = frozenset({
    "CASHFLOW_OBSERVATION_HEADER_MISSING",
    "CASHFLOW_OBSERVATION_LAYOUT_CONSENSUS_MISSING",
    "OCR_LOW_CONFIDENCE",
    "OCR_TSV_INVALID",
})
_CASHFLOW_IMAGE_LAYOUT_RECOVERY_CODES = frozenset({
    "CASHFLOW_OBSERVATION_HEADER_MISSING",
    "OCR_GRID_RULES_NOT_FOUND",
    "OCR_LOW_CONFIDENCE",
    "OCR_TSV_INVALID",
})


class ParseError(ContractError):
    pass


# A generic source label is intentionally never enough to classify a financial
# fact.  These sets describe only a uniform *parser phase* across the two
# candidate schemas; they contain no OCR text, field name, account, amount or
# source identifier.  Mixed phases remain the existing generic unresolved
# result so that a diagnostic cannot overstate what was observed.
_GENERIC_OCR_HEADER_PHASE_CODES = frozenset({
    "OCR_HEADER_MAPPING_MISSING",
    "OCR_HEADER_ROW_AMBIGUOUS",
    "OCR_COLUMN_MAPPING_AMBIGUOUS",
    "COLUMN_HEADER_DUPLICATE",
})
_GENERIC_OCR_ROW_PHASE_CODES = frozenset({
    "OCR_ROW_REQUIRED_CELL_MISSING",
    "SOURCE_ROWS_EMPTY",
})
_GENERIC_OCR_CONFIDENCE_PHASE_CODES = frozenset({"OCR_LOW_CONFIDENCE"})
_SOURCE_INTEGRITY_PARSE_CODES = frozenset({
    "SOURCE_LINEAGE_INVALID",
    "SOURCE_VERSION_MISMATCH",
    "SOURCE_PAYLOAD_HASH_MISMATCH",
})


@dataclass(frozen=True)
class OcrParsedAttachment:
    """A deterministic OCR result that has not yet crossed the runtime template gate.

    The layout fingerprint deliberately contains only canonical field names,
    geometry buckets and parser metadata.  It contains no OCR text, document
    identifier, amount or account value, so the runtime may retain it in its
    private journal as a calibration key without creating a second raw store.
    """

    facts: ParsedFacts
    layout_fingerprint: str


@dataclass(frozen=True)
class _OcrWord:
    text: str
    confidence_bps: int
    page: int
    block: int
    paragraph: int
    line: int
    left: int
    top: int
    width: int
    height: int


@dataclass(frozen=True)
class _OcrHeaderCell:
    label: str
    field: str | None
    left: int
    right: int
    top: int
    bottom: int
    confidence_bps: int


def normalize_header(value: object) -> str:
    return re.sub(r"[\s_\-（）()【】\[\]：:]+", "", str(value or "").strip()).lower()


ALIASES: Mapping[str, tuple[str, ...]] = {
    "business_date": ("业务日期", "数据日期", "截止日期", "日期", "统计日期"),
    "company": ("公司", "公司名称", "主体", "账户主体", "所属公司"),
    "bank": ("开户行", "银行", "银行名称", "银行机构"),
    "account": ("账号", "账户", "银行账号", "账户号"),
    "ending": ("期末可用余额", "可用余额", "账户余额", "余额", "期末余额"),
    "opening": ("期初可用余额", "期初余额", "上日余额", "前日余额"),
    "currency": ("币种", "货币", "currency"),
    "transaction_id": ("流水号", "交易流水号", "交易编号", "凭证号", "业务编号"),
    "occurred_at": ("发生时间", "交易时间", "记账时间", "时间"),
    "inflow": ("流入", "收入", "贷方发生额", "入账金额"),
    "outflow": ("流出", "支出", "转出", "借方发生额", "出账金额"),
    "category": ("收支类别", "类别", "费用类别"),
    "amount": ("金额", "发生额", "交易金额"),
    "direction": ("收支方向", "方向", "借贷方向", "交易方向"),
    "adjustment": ("调整", "调整金额", "其他调整"),
    "internal": ("内部调拨", "是否内部调拨", "调拨标记"),
    "transfer_id": ("调拨编号", "内部调拨编号", "关联流水号", "关联交易号"),
}


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _trim_trailing_blank(values: Iterable[object]) -> list[object]:
    materialized = list(values)
    while materialized and _is_blank(materialized[-1]):
        materialized.pop()
    return materialized


def _validated_headers(headers: Iterable[object]) -> list[str]:
    validated: list[str] = []
    seen: set[str] = set()
    for raw_header in headers:
        header = str(raw_header or "").strip()
        normalized = normalize_header(header)
        if not header or not normalized:
            raise ParseError("COLUMN_HEADER_EMPTY")
        if normalized in seen:
            raise ParseError("COLUMN_HEADER_DUPLICATE")
        seen.add(normalized)
        validated.append(header)
    if not validated:
        raise ParseError("COLUMN_MAPPING_EMPTY")
    return validated


def _column_map(headers: Iterable[object]) -> dict[str, str]:
    validated = _validated_headers(headers)
    normalized = {normalize_header(header): header for header in validated}
    mapped: dict[str, str] = {}
    for field, aliases in ALIASES.items():
        matches = [
            normalized[normalized_alias]
            for alias in aliases
            if (normalized_alias := normalize_header(alias)) in normalized
        ]
        if len(matches) > 1:
            raise ParseError("COLUMN_MAPPING_AMBIGUOUS_" + field.upper())
        if matches:
            mapped[field] = matches[0]
    return mapped


def _required(mapped: Mapping[str, str], fields: Iterable[str]) -> None:
    missing = [field for field in fields if field not in mapped]
    if missing:
        raise ParseError("COLUMN_MAPPING_MISSING_" + "_".join(sorted(missing)).upper())


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ParseError("BUSINESS_DATE_INVALID")


def _parse_datetime(value: object) -> datetime | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ParseError("TRANSACTION_TIME_INVALID")


def _required_text(row: Mapping[str, object], column: str, code: str) -> str:
    value = str(row.get(column) or "").strip()
    if not value:
        raise ParseError(code)
    return value


def _required_identifier(row: Mapping[str, object], column: str, missing_code: str, non_text_code: str) -> str:
    value = row.get(column)
    if _is_blank(value):
        raise ParseError(missing_code)
    # Account and transaction identifiers can legally begin with zero.  An
    # Excel numeric cell has already lost that information, so it is not safe
    # to stringify and guess it back.
    if not isinstance(value, str):
        raise ParseError(non_text_code)
    text = value.strip()
    if not text:
        raise ParseError(missing_code)
    return text


def _optional_identifier(row: Mapping[str, object], column: str, non_text_code: str) -> str | None:
    value = row.get(column)
    if _is_blank(value):
        return None
    if not isinstance(value, str):
        raise ParseError(non_text_code)
    text = value.strip()
    return text or None


def _amount_to_fen(value: object) -> int:
    # XLSX numeric cells are exposed by openpyxl as Python floats.  Convert
    # their canonical decimal spelling before the integer-fen contract; never
    # perform binary float arithmetic or display-format rounding.
    if isinstance(value, float):
        return parse_amount_to_fen(str(value))
    return parse_amount_to_fen(value)


def _bool(value: object) -> bool:
    text = normalize_header(value)
    if text in {"", "否", "no", "false", "0", "n"}:
        return False
    if text in {"是", "yes", "true", "1", "y", "内部", "调拨"}:
        return True
    raise ParseError("INTERNAL_TRANSFER_FLAG_INVALID")


def _date_from_filename(filename: str) -> date | None:
    match = re.search(r"(?<!\d)(20\d{2})[-_.年]?(\d{1,2})[-_.月]?(\d{1,2})(?:日)?(?!\d)", filename)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _declared_mime(mime: str | None) -> str | None:
    if mime is None or not str(mime).strip():
        return None
    normalized = str(mime).split(";", 1)[0].strip().lower()
    if not _MIME_TOKEN.fullmatch(normalized):
        raise ParseError("MIME_DECLARATION_INVALID")
    return normalized


def _capability_magic(payload: bytes) -> str:
    """Classify bytes for a values-free capability receipt, never parsing them.

    This deliberately stops at stable file signatures.  In particular, a
    recognised image or PDF signature is evidence that the attachment needs a
    reviewed parser, *not* permission to OCR it or treat it as financial data.
    """

    if not payload:
        return "EMPTY"
    if payload.startswith(_ZIP_MAGICS):
        return "ZIP"
    if payload.startswith(b"\x89PNG\r\n\x1a\n"):
        return "PNG"
    if payload.startswith(b"\xff\xd8\xff"):
        return "JPEG"
    if payload.startswith((b"GIF87a", b"GIF89a")):
        return "GIF"
    if payload.startswith(b"BM"):
        return "BMP"
    if payload.startswith(b"RIFF") and payload[8:12] == b"WEBP":
        return "WEBP"
    if payload.startswith(b"%PDF-"):
        return "PDF"
    if payload.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "OLE"
    sample = payload[:4096]
    if b"\x00" in sample:
        return "BINARY"
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            sample.decode(encoding)
            return "TEXT"
        except UnicodeDecodeError:
            continue
    return "BINARY"


def attachment_capability_metadata(*, filename: str, payload: bytes, mime: str | None = None) -> tuple[str, str | None, str]:
    """Return bounded, values-free metadata for every Git-readback attachment.

    The returned tuple can safely be persisted for a supported *or* rejected
    type.  It intentionally does not inspect document text, OCR images, or
    derive a financial family from a filename.
    """

    candidate_suffix = Path(filename).suffix.lower()
    suffix = candidate_suffix if candidate_suffix in _CAPABILITY_SUFFIXES else "UNKNOWN_SUFFIX"
    try:
        declared_mime = _declared_mime(mime)
    except ParseError:
        # The parse path records MIME_DECLARATION_INVALID as the failure code;
        # retaining an arbitrary malformed string would violate the redacted
        # capability-journal contract.
        declared_mime = None
    return suffix, declared_mime, _capability_magic(payload)


def inspect_attachment_format(*, filename: str, payload: bytes, mime: str | None = None) -> ParserEvidence:
    """Validate filename, declared MIME and byte magic before parser-open.

    This returns only values-free metadata.  It is deliberately not a global
    support claim: production support is recorded only after the private-Git
    readback byte subsequently opens and parses successfully.
    """

    suffix = Path(filename).suffix.lower()
    if suffix not in _ALLOWED_SUFFIXES:
        raise ParseError("UNSUPPORTED_ATTACHMENT")
    declared_mime = _declared_mime(mime)
    if suffix in _TEXT_SUFFIXES:
        if payload.startswith(_ZIP_MAGICS) or b"\x00" in payload:
            raise ParseError("FORMAT_MAGIC_MISMATCH")
        if declared_mime is not None and declared_mime not in _CSV_MIME:
            raise ParseError("MIME_SUFFIX_MISMATCH")
        return ParserEvidence("CSV", suffix, declared_mime, "TEXT", PARSER_VERSION)
    if suffix == ".xls":
        if not payload.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
            raise ParseError("FORMAT_MAGIC_MISMATCH")
        if declared_mime is not None and declared_mime not in _XLS_MIME:
            raise ParseError("MIME_SUFFIX_MISMATCH")
        return ParserEvidence("XLS", suffix, declared_mime, "OLE", PARSER_VERSION)
    if not payload.startswith(_ZIP_MAGICS):
        raise ParseError("FORMAT_MAGIC_MISMATCH")
    allowed_mime = _XLSX_MIME if suffix == ".xlsx" else _XLSM_MIME
    if declared_mime is not None and declared_mime not in allowed_mime:
        raise ParseError("MIME_SUFFIX_MISMATCH")
    return ParserEvidence("XLSX" if suffix == ".xlsx" else "XLSM", suffix, declared_mime, "ZIP", PARSER_VERSION)


def _csv_delimiter(text: str) -> str:
    for line in text.splitlines():
        if line.strip():
            counts = {delimiter: line.count(delimiter) for delimiter in (",", "\t", ";")}
            maximum = max(counts.values())
            if maximum == 0:
                return ","
            winners = [delimiter for delimiter, count in counts.items() if count == maximum]
            if len(winners) != 1:
                raise ParseError("CSV_DELIMITER_AMBIGUOUS")
            return winners[0]
    raise ParseError("COLUMN_MAPPING_EMPTY")


def _csv_rows(payload: bytes) -> list[dict[str, object]]:
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = payload.decode("gb18030")
        except UnicodeDecodeError as exc:
            raise ParseError("CORRUPT_ATTACHMENT") from exc
    delimiter = _csv_delimiter(text)
    try:
        reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter, strict=True)
        headers = _validated_headers(next(reader))
        rows: list[dict[str, object]] = []
        for raw_row in reader:
            if not raw_row or all(_is_blank(value) for value in raw_row):
                continue
            if len(raw_row) != len(headers):
                raise ParseError("CSV_ROW_WIDTH_INVALID")
            rows.append(dict(zip(headers, raw_row)))
        return rows
    except StopIteration as exc:
        raise ParseError("COLUMN_MAPPING_EMPTY") from exc
    except csv.Error as exc:
        raise ParseError("CORRUPT_ATTACHMENT") from exc


def _xlsx_rows(payload: bytes) -> list[dict[str, object]]:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:  # container has a locked dependency
        raise ParseError("XLSX_RUNTIME_DEPENDENCY_MISSING") from exc
    formula_book = value_book = None
    try:
        # Cached formula values must not be treated as an independently
        # verified financial source.  Open the same workbook in formula view
        # first, then only read values if its sole worksheet contains no formula.
        formula_book = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=False, keep_vba=False)
        # Until a target-group real sample freezes an explicit multi-sheet
        # template, selecting ``active`` would silently discard the remaining
        # worksheets.  A financial source must be complete or fail closed;
        # reject every multi-sheet workbook rather than guessing where facts
        # belong (including a hidden or auxiliary sheet).
        if len(formula_book.worksheets) != 1:
            raise ParseError("XLSX_WORKSHEET_AMBIGUOUS")
        formula_sheet = formula_book.worksheets[0]
        for formula_row in formula_sheet.iter_rows():
            if any(cell.data_type == "f" for cell in formula_row):
                raise ParseError("XLSX_FORMULA_UNSUPPORTED")
        value_book = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True, keep_vba=False)
        if len(value_book.worksheets) != 1:
            raise ParseError("XLSX_WORKSHEET_AMBIGUOUS")
        sheet = value_book.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        headers = _validated_headers(_trim_trailing_blank(next(iterator)))
        rows: list[dict[str, object]] = []
        for raw_row in iterator:
            values = list(raw_row)
            if any(not _is_blank(value) for value in values[len(headers):]):
                raise ParseError("XLSX_ROW_WIDTH_INVALID")
            values = values[:len(headers)]
            if not values or all(_is_blank(value) for value in values):
                continue
            rows.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers)})
        return rows
    except ParseError:
        raise
    except StopIteration as exc:
        raise ParseError("COLUMN_MAPPING_EMPTY") from exc
    except Exception as exc:  # openpyxl has many format-specific exceptions
        raise ParseError("CORRUPT_ATTACHMENT") from exc
    finally:
        if formula_book is not None:
            formula_book.close()
        if value_book is not None:
            value_book.close()


def _xls_biff_records(workbook_stream: bytes) -> tuple[tuple[int, bytes], ...]:
    """Read the bounded BIFF record framing required for a legacy XLS gate.

    xlrd intentionally returns cached formula values for historic workbooks.
    We therefore inspect the raw Workbook stream first, without evaluating any
    record, and reject formula/encryption records before asking xlrd for cells.
    """

    records: list[tuple[int, bytes]] = []
    offset = 0
    while offset < len(workbook_stream):
        if len(workbook_stream) - offset < 4:
            # Some valid CFB writers allocate the Workbook stream to a sector
            # boundary and leave only zero padding after the final BIFF EOF.
            # It is not a BIFF record and must not make a valid static table
            # look corrupt; any non-zero tail remains a hard rejection.
            if not any(workbook_stream[offset:]):
                break
            raise ParseError("CORRUPT_ATTACHMENT")
        record_id, payload_size = struct.unpack_from("<HH", workbook_stream, offset)
        if record_id == 0 and payload_size == 0 and not any(workbook_stream[offset:]):
            break
        start = offset + 4
        end = start + payload_size
        if end > len(workbook_stream):
            raise ParseError("CORRUPT_ATTACHMENT")
        records.append((record_id, workbook_stream[start:end]))
        offset = end
    if not records or records[0][0] not in {0x0009, 0x0809}:
        raise ParseError("CORRUPT_ATTACHMENT")
    return tuple(records)


def _xls_workbook_preflight(payload: bytes) -> None:
    """Prove the narrow legacy-XLS shape before values become table rows."""

    try:
        import olefile  # type: ignore[import-not-found]
    except ImportError as exc:  # container has a locked dependency
        raise ParseError("XLS_RUNTIME_DEPENDENCY_MISSING") from exc

    try:
        with olefile.OleFileIO(io.BytesIO(payload), raise_defects=olefile.DEFECT_INCORRECT) as compound:
            if getattr(compound, "parsing_issues", ()):
                raise ParseError("CORRUPT_ATTACHMENT")
            stream_paths = compound.listdir(streams=True, storages=False)
            if any(
                any(str(component).casefold() in _XLS_MACRO_COMPONENTS for component in path)
                for path in stream_paths
            ):
                raise ParseError("XLS_MACRO_UNSUPPORTED")
            workbook_paths = [
                path
                for path in stream_paths
                if len(path) == 1 and str(path[0]).casefold() in {"book", "workbook"}
            ]
            if len(workbook_paths) != 1:
                raise ParseError("CORRUPT_ATTACHMENT")
            workbook_stream = compound.openstream(workbook_paths[0]).read()
    except ParseError:
        raise
    except Exception as exc:  # OLE parser reports many malformed-container shapes
        raise ParseError("CORRUPT_ATTACHMENT") from exc

    records = _xls_biff_records(workbook_stream)
    sheet_types: list[int] = []
    for record_id, record_payload in records:
        if record_id == _XLS_FILEPASS_RECORD:
            raise ParseError("XLS_ENCRYPTED_UNSUPPORTED")
        if record_id in _XLS_FORMULA_RECORD_IDS:
            raise ParseError("XLS_FORMULA_UNSUPPORTED")
        if record_id == _XLS_BOUNDSHEET_RECORD:
            if len(record_payload) < 6:
                raise ParseError("CORRUPT_ATTACHMENT")
            sheet_types.append(record_payload[5])
    if len(sheet_types) != 1:
        raise ParseError("XLS_WORKSHEET_AMBIGUOUS")
    if sheet_types[0] != 0:
        # Excel 4 macro, chart and VBA module sheets cannot be interpreted as
        # a financial table by this parser.
        raise ParseError("XLS_WORKSHEET_TYPE_UNSUPPORTED")


def _xls_cell_value(cell: Any, *, book: Any, xlrd: Any) -> object:
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, book.datemode)
    return cell.value


def _xls_rows(payload: bytes) -> list[dict[str, object]]:
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as exc:  # container has a locked dependency
        raise ParseError("XLS_RUNTIME_DEPENDENCY_MISSING") from exc

    _xls_workbook_preflight(payload)
    book = None
    try:
        book = xlrd.open_workbook(file_contents=payload, on_demand=False, formatting_info=False)
        if book.nsheets != 1:
            raise ParseError("XLS_WORKSHEET_AMBIGUOUS")
        sheet = book.sheet_by_index(0)
        headers = _validated_headers(_trim_trailing_blank(
            _xls_cell_value(cell, book=book, xlrd=xlrd) for cell in sheet.row(0)
        ))
        rows: list[dict[str, object]] = []
        for row_index in range(1, sheet.nrows):
            values = [_xls_cell_value(cell, book=book, xlrd=xlrd) for cell in sheet.row(row_index)]
            if any(not _is_blank(value) for value in values[len(headers):]):
                raise ParseError("XLS_ROW_WIDTH_INVALID")
            values = values[:len(headers)]
            if not values or all(_is_blank(value) for value in values):
                continue
            rows.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers)})
        return rows
    except ParseError:
        raise
    except IndexError as exc:
        raise ParseError("COLUMN_MAPPING_EMPTY") from exc
    except Exception as exc:  # xlrd has many format-specific exceptions
        raise ParseError("CORRUPT_ATTACHMENT") from exc
    finally:
        if book is not None:
            release = getattr(book, "release_resources", None)
            if callable(release):
                release()


def _rows_from_bytes(filename: str, payload: bytes, mime: str | None) -> tuple[list[dict[str, object]], ParserEvidence]:
    evidence = inspect_attachment_format(filename=filename, payload=payload, mime=mime)
    if evidence.format == "CSV":
        return _csv_rows(payload), evidence
    if evidence.format == "XLS":
        return _xls_rows(payload), evidence
    return _xlsx_rows(payload), evidence


def _ocr_integer(value: object, *, code: str, minimum: int | None = None) -> int:
    try:
        parsed = int(str(value))
    except (TypeError, ValueError) as exc:
        raise ParseError(code) from exc
    if minimum is not None and parsed < minimum:
        raise ParseError(code)
    return parsed


def _ocr_confidence_bps(value: object) -> int:
    try:
        score = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ParseError("OCR_CONFIDENCE_INVALID") from exc
    if not score.is_finite() or score < 0 or score > 100:
        raise ParseError("OCR_CONFIDENCE_INVALID")
    return int((score * 100).to_integral_value(rounding=ROUND_FLOOR))


def _validate_ocr_min_confidence(value: object) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 1 or value > 10_000:
        raise ParseError("OCR_CONFIDENCE_THRESHOLD_INVALID")
    return value


def inspect_ocr_attachment_format(*, filename: str, payload: bytes, mime: str | None = None) -> ParserEvidence:
    """Validate one image or scanned-PDF before a deterministic OCR open.

    OCR is not a bypass around the normal attachment contract.  A recognised
    filename type, byte magic or declared MIME disagreement is a hard failure.
    An opaque, unrecognised client-side suffix may be normalised only from the
    verified byte signature, then remains subject to the same MIME and parser
    gates.  The result contains only format metadata and can therefore be used
    in the existing redacted capability journal.
    """

    if len(payload) > OCR_MAX_ATTACHMENT_BYTES:
        raise ParseError("OCR_ATTACHMENT_TOO_LARGE")
    candidate_suffix = Path(filename).suffix.lower()
    magic = _capability_magic(payload)
    if candidate_suffix in _OCR_SUFFIXES:
        suffix = candidate_suffix
        if magic != _OCR_MAGIC_BY_SUFFIX[suffix]:
            raise ParseError("FORMAT_MAGIC_MISMATCH")
    elif candidate_suffix in _CAPABILITY_SUFFIXES:
        # A recognised non-OCR type (for example a workbook or GIF) is a
        # concrete contradictory declaration, not an opaque transport suffix.
        raise ParseError("UNSUPPORTED_ATTACHMENT")
    else:
        suffix = _OCR_CANONICAL_SUFFIX_BY_MAGIC.get(magic)
        if suffix is None:
            raise ParseError("UNSUPPORTED_ATTACHMENT")
    declared_mime = _declared_mime(mime)
    if declared_mime is not None and declared_mime not in _OCR_IMAGE_MIME[suffix]:
        raise ParseError("MIME_SUFFIX_MISMATCH")
    return ParserEvidence(
        format=f"OCR_{magic}",
        suffix=suffix,
        declared_mime=declared_mime,
        magic=magic,
        parser_version=PARSER_VERSION,
    )


def is_ocr_attachment(filename: str, *, payload: bytes | None = None) -> bool:
    """Whether an attachment can enter the explicit image/scanned-PDF fallback.

    Known OCR suffixes retain their existing parser-time magic validation.  A
    non-capability suffix is admitted only when raw bytes establish one of the
    same bounded OCR formats; a known conflicting capability suffix cannot be
    converted into an OCR input by its payload.
    """

    suffix = Path(filename).suffix.lower()
    if suffix in _OCR_SUFFIXES:
        return True
    if payload is None or suffix in _CAPABILITY_SUFFIXES:
        return False
    return _capability_magic(payload) in _OCR_CANONICAL_SUFFIX_BY_MAGIC


def _run_ocr_command(
    command: list[str],
    *,
    runner: Callable[..., Any],
    failure_code: str,
) -> str:
    """Run an offline OCR utility without retaining its diagnostic stream."""

    try:
        completed = runner(
            command,
            capture_output=True,
            text=True,
            timeout=OCR_TIMEOUT_SECONDS,
            check=False,
        )
    except FileNotFoundError as exc:
        raise ParseError("OCR_RUNTIME_UNAVAILABLE") from exc
    except subprocess.TimeoutExpired as exc:
        raise ParseError("OCR_TIMEOUT") from exc
    except OSError as exc:
        raise ParseError("OCR_RUNTIME_UNAVAILABLE") from exc
    if getattr(completed, "returncode", 1) != 0:
        raise ParseError(failure_code)
    stdout = getattr(completed, "stdout", "")
    return stdout if isinstance(stdout, str) else ""


def _pdf_page_count(path: Path, *, runner: Callable[..., Any]) -> int:
    output = _run_ocr_command(["pdfinfo", str(path)], runner=runner, failure_code="OCR_PDF_METADATA_FAILED")
    pages: int | None = None
    for line in output.splitlines():
        key, separator, value = line.partition(":")
        if separator and key.strip().lower() == "pages":
            if pages is not None:
                raise ParseError("OCR_PDF_METADATA_INVALID")
            pages = _ocr_integer(value.strip(), code="OCR_PDF_METADATA_INVALID", minimum=1)
    if pages is None:
        raise ParseError("OCR_PDF_METADATA_INVALID")
    return pages


def _preprocess_ocr_grid_image(*, image: Path, root: Path) -> Path:
    """Remove only long table rules from one bounded OCR image.

    Historical daily-funds screenshots are compact ruled tables.  Tesseract
    can otherwise merge every cell into one logical line even though the text
    itself is legible.  This is deliberately an image-format repair, not a
    semantic OCR fallback: it thresholds pixels, erases only rows/columns
    whose dark coverage proves a table rule, doubles the scale, and keeps the
    transformed bytes inside the existing temporary directory.

    The caller still has to pass the unchanged header, confidence, row and
    footer-total gates twice under independent page-segmentation modes.
    """

    try:
        from PIL import Image
    except ImportError as exc:
        raise ParseError("OCR_PREPROCESS_RUNTIME_UNAVAILABLE") from exc
    try:
        with Image.open(image) as source:
            width, height = source.size
            if (
                width < 16
                or height < 16
                or width * height > _OCR_GRID_MAX_PIXELS
            ):
                raise ParseError("OCR_IMAGE_DIMENSIONS_UNSUPPORTED")
            source.load()
            binary = source.convert("L").point(
                lambda value: 0 if value < _OCR_GRID_THRESHOLD else 255,
                mode="L",
            )
    except ParseError:
        raise
    except (OSError, ValueError) as exc:
        raise ParseError("OCR_IMAGE_PREPROCESS_INVALID") from exc

    pixels = binary.load()
    minimum_row_dark = (width * _OCR_GRID_LINE_COVERAGE_BPS + 9_999) // 10_000
    minimum_column_dark = (height * _OCR_GRID_LINE_COVERAGE_BPS + 9_999) // 10_000
    rule_rows = tuple(
        y for y in range(height)
        if sum(1 for x in range(width) if pixels[x, y] == 0) >= minimum_row_dark
    )
    rule_columns = tuple(
        x for x in range(width)
        if sum(1 for y in range(height) if pixels[x, y] == 0) >= minimum_column_dark
    )
    if not rule_rows or not rule_columns:
        raise ParseError("OCR_GRID_RULES_NOT_FOUND")
    for y in rule_rows:
        for x in range(width):
            pixels[x, y] = 255
    for x in rule_columns:
        for y in range(height):
            pixels[x, y] = 255

    normalized = binary.resize(
        (width * _OCR_GRID_SCALE, height * _OCR_GRID_SCALE),
        resample=Image.Resampling.NEAREST,
    )
    output = root / "ocr-grid-normalized.png"
    try:
        normalized.save(output, format="PNG")
    except OSError as exc:
        raise ParseError("OCR_IMAGE_PREPROCESS_INVALID") from exc
    return output


def _preprocess_ocr_enhanced_image(*, image: Path, root: Path) -> Path:
    """Render one bounded source image for a final strict OCR attempt.

    This is a deterministic presentation repair for compact screenshots, not
    a semantic interpretation step.  It retains the source pixels in
    grayscale, applies only fixed autocontrast and sharpening, and bounds the
    output dimensions before Tesseract sees it.  The caller must still satisfy
    the unchanged header, confidence, date, row and footer-total gates twice.
    """

    try:
        from PIL import Image, ImageFilter, ImageOps
    except ImportError as exc:
        raise ParseError("OCR_PREPROCESS_RUNTIME_UNAVAILABLE") from exc
    try:
        with Image.open(image) as source:
            width, height = source.size
            pixels = width * height
            if width < 16 or height < 16 or pixels > _OCR_ENHANCED_MAX_PIXELS:
                raise ParseError("OCR_IMAGE_DIMENSIONS_UNSUPPORTED")
            source.load()
            if pixels <= _OCR_ENHANCED_SMALL_IMAGE_PIXELS:
                scale = _OCR_ENHANCED_SMALL_SCALE
            elif pixels <= _OCR_ENHANCED_MEDIUM_IMAGE_PIXELS:
                scale = _OCR_ENHANCED_MEDIUM_SCALE
            else:
                scale = 1
            rendered = ImageOps.autocontrast(source.convert("L"))
            if scale > 1:
                rendered = rendered.resize(
                    (width * scale, height * scale),
                    resample=Image.Resampling.LANCZOS,
                )
            rendered = rendered.filter(ImageFilter.UnsharpMask(
                radius=_OCR_ENHANCED_SHARPEN_RADIUS,
                percent=_OCR_ENHANCED_SHARPEN_PERCENT,
                threshold=_OCR_ENHANCED_SHARPEN_THRESHOLD,
            ))
            output = root / "ocr-enhanced.png"
            rendered.save(output, format="PNG")
    except ParseError:
        raise
    except (OSError, ValueError) as exc:
        raise ParseError("OCR_IMAGE_PREPROCESS_INVALID") from exc
    return output


def _preprocess_ocr_binarized_image(*, image: Path, root: Path) -> Path:
    """Apply one deterministic global threshold after the bounded enhancement.

    This is a last image-layout repair for screenshot contrast only.  The
    threshold is derived from the image histogram (Otsu's method), not chosen
    per attachment, and it cannot change the downstream header, confidence,
    date, row, or footer-reconciliation requirements.
    """

    try:
        from PIL import Image, ImageFilter, ImageOps
    except ImportError as exc:
        raise ParseError("OCR_PREPROCESS_RUNTIME_UNAVAILABLE") from exc
    try:
        with Image.open(image) as source:
            width, height = source.size
            pixels = width * height
            if width < 16 or height < 16 or pixels > _OCR_ENHANCED_MAX_PIXELS:
                raise ParseError("OCR_IMAGE_DIMENSIONS_UNSUPPORTED")
            source.load()
            if pixels <= _OCR_ENHANCED_SMALL_IMAGE_PIXELS:
                scale = _OCR_ENHANCED_SMALL_SCALE
            elif pixels <= _OCR_ENHANCED_MEDIUM_IMAGE_PIXELS:
                scale = _OCR_ENHANCED_MEDIUM_SCALE
            else:
                scale = 1
            rendered = ImageOps.autocontrast(source.convert("L"))
            if scale > 1:
                rendered = rendered.resize(
                    (width * scale, height * scale),
                    resample=Image.Resampling.LANCZOS,
                )
            rendered = rendered.filter(ImageFilter.UnsharpMask(
                radius=_OCR_ENHANCED_SHARPEN_RADIUS,
                percent=_OCR_ENHANCED_SHARPEN_PERCENT,
                threshold=_OCR_ENHANCED_SHARPEN_THRESHOLD,
            ))
            histogram = rendered.histogram()
            total = sum(histogram)
            weighted_total = sum(index * count for index, count in enumerate(histogram))
            weight_background = 0
            weighted_background = 0
            best_threshold = 0
            best_variance = -1.0
            for threshold, count in enumerate(histogram):
                weight_background += count
                if weight_background == 0:
                    continue
                weight_foreground = total - weight_background
                if weight_foreground == 0:
                    break
                weighted_background += threshold * count
                mean_background = weighted_background / weight_background
                mean_foreground = (weighted_total - weighted_background) / weight_foreground
                variance = weight_background * weight_foreground * (mean_background - mean_foreground) ** 2
                if variance > best_variance:
                    best_variance = variance
                    best_threshold = threshold
            if best_variance < 0:
                raise ParseError("OCR_IMAGE_PREPROCESS_INVALID")
            binary = rendered.point(lambda value: 0 if value <= best_threshold else 255, mode="L")
            output = root / "ocr-binarized.png"
            binary.save(output, format="PNG")
    except ParseError:
        raise
    except (OSError, ValueError) as exc:
        raise ParseError("OCR_IMAGE_PREPROCESS_INVALID") from exc
    return output


def _ocr_tsv(
    *,
    payload: bytes,
    evidence: ParserEvidence,
    runner: Callable[..., Any],
    psm: int = OCR_PRIMARY_PSM,
    preprocess_grid: bool = False,
    preprocess_enhanced: bool = False,
    preprocess_binarized: bool = False,
) -> str:
    """Generate in-memory Tesseract TSV for exactly one bounded document page.

    The input file and rendered PDF page live only inside a temporary directory.
    Neither OCR text nor utility stderr is written to a log, status file, Git
    repository or exception message.
    """

    if psm not in _OCR_ALLOWED_PSMS:
        raise ParseError("OCR_PSM_INVALID")
    if sum((preprocess_grid, preprocess_enhanced, preprocess_binarized)) > 1:
        raise ParseError("OCR_PREPROCESS_MODE_INVALID")
    with tempfile.TemporaryDirectory(prefix="daily-funds-ocr-") as temporary:
        root = Path(temporary)
        source = root / f"input{evidence.suffix}"
        source.write_bytes(payload)
        image = source
        if evidence.magic == "PDF":
            if _pdf_page_count(source, runner=runner) != 1:
                raise ParseError("OCR_PDF_PAGE_AMBIGUOUS")
            prefix = root / "render"
            _run_ocr_command(
                ["pdftoppm", "-png", "-r", "300", "-f", "1", "-l", "1", str(source), str(prefix)],
                runner=runner,
                failure_code="OCR_PDF_RENDER_FAILED",
            )
            rendered = tuple(sorted(root.glob("render-*.png")))
            if len(rendered) != 1 or rendered[0].is_symlink() or not rendered[0].is_file():
                raise ParseError("OCR_PDF_RENDER_FAILED")
            image = rendered[0]
        if preprocess_grid:
            image = _preprocess_ocr_grid_image(image=image, root=root)
        elif preprocess_enhanced:
            image = _preprocess_ocr_enhanced_image(image=image, root=root)
        elif preprocess_binarized:
            image = _preprocess_ocr_binarized_image(image=image, root=root)
        return _run_ocr_command(
            ["tesseract", str(image), "stdout", "-l", OCR_LANGUAGE, "--psm", str(psm), "tsv"],
            runner=runner,
            failure_code="OCR_ENGINE_FAILED",
        )


def _parse_tesseract_tsv(text: str) -> tuple[_OcrWord, ...]:
    expected = {
        "level", "page_num", "block_num", "par_num", "line_num", "word_num",
        "left", "top", "width", "height", "conf", "text",
    }
    try:
        reader = csv.DictReader(io.StringIO(text, newline=""), delimiter="\t", strict=True)
    except csv.Error as exc:
        raise ParseError("OCR_TSV_INVALID") from exc
    if reader.fieldnames is None or set(reader.fieldnames) != expected or len(reader.fieldnames) != len(expected):
        raise ParseError("OCR_TSV_INVALID")
    words: list[_OcrWord] = []
    try:
        for raw in reader:
            if raw.get("level") != "5":
                continue
            token = str(raw.get("text") or "").strip()
            if not token:
                continue
            if len(token) > 512:
                raise ParseError("OCR_TOKEN_INVALID")
            words.append(_OcrWord(
                text=token,
                confidence_bps=_ocr_confidence_bps(raw.get("conf")),
                page=_ocr_integer(raw.get("page_num"), code="OCR_TSV_INVALID", minimum=1),
                block=_ocr_integer(raw.get("block_num"), code="OCR_TSV_INVALID", minimum=0),
                paragraph=_ocr_integer(raw.get("par_num"), code="OCR_TSV_INVALID", minimum=0),
                line=_ocr_integer(raw.get("line_num"), code="OCR_TSV_INVALID", minimum=0),
                left=_ocr_integer(raw.get("left"), code="OCR_TSV_INVALID", minimum=0),
                top=_ocr_integer(raw.get("top"), code="OCR_TSV_INVALID", minimum=0),
                width=_ocr_integer(raw.get("width"), code="OCR_TSV_INVALID", minimum=1),
                height=_ocr_integer(raw.get("height"), code="OCR_TSV_INVALID", minimum=1),
            ))
    except csv.Error as exc:
        raise ParseError("OCR_TSV_INVALID") from exc
    if not words:
        raise ParseError("OCR_OUTPUT_EMPTY")
    return tuple(words)


def _ocr_lines(words: Iterable[_OcrWord]) -> tuple[tuple[_OcrWord, ...], ...]:
    grouped: dict[tuple[int, int, int, int], list[_OcrWord]] = {}
    for word in words:
        grouped.setdefault((word.page, word.block, word.paragraph, word.line), []).append(word)
    ordered = sorted(
        grouped.values(),
        key=lambda row: (
            min(word.page for word in row),
            min(word.top for word in row),
            min(word.left for word in row),
            min(word.block for word in row),
            min(word.paragraph for word in row),
            min(word.line for word in row),
        ),
    )
    return tuple(tuple(sorted(row, key=lambda word: (word.left, word.top, word.text))) for row in ordered)


def _ocr_visual_rows(lines: tuple[tuple[_OcrWord, ...], ...]) -> tuple[tuple[int, tuple[_OcrWord, ...]], ...]:
    """Reassemble one visually aligned row when Tesseract splits its columns.

    Tesseract can assign separate ``line_num`` values to cells that share one
    rendered table row.  Group only OCR lines whose vertical boxes overlap;
    adjacent rows remain separate, and callers still apply the same strict
    alias, confidence, and row-total gates after this layout-only repair.
    """

    grouped: list[tuple[int, tuple[_OcrWord, ...]]] = []
    current_index: int | None = None
    current_top: int | None = None
    current_bottom: int | None = None
    current_words: list[_OcrWord] = []

    def flush() -> None:
        nonlocal current_index, current_top, current_bottom, current_words
        if current_index is not None and current_words:
            grouped.append((current_index, tuple(sorted(current_words, key=lambda word: (word.left, word.top, word.text)))))
        current_index = None
        current_top = None
        current_bottom = None
        current_words = []

    for index, words in enumerate(lines):
        line_top = min(word.top for word in words)
        line_bottom = max(word.top + word.height for word in words)
        if current_index is None:
            current_index = index
            current_top = line_top
            current_bottom = line_bottom
            current_words.extend(words)
            continue
        assert current_top is not None and current_bottom is not None
        if line_top <= current_bottom and line_bottom >= current_top:
            current_top = min(current_top, line_top)
            current_bottom = max(current_bottom, line_bottom)
            current_words.extend(words)
            continue
        flush()
        current_index = index
        current_top = line_top
        current_bottom = line_bottom
        current_words.extend(words)
    flush()
    return tuple(grouped)


def _ocr_alias_candidates(words: tuple[_OcrWord, ...]) -> list[tuple[int, int, str]]:
    aliases: dict[str, set[str]] = {}
    for field, values in ALIASES.items():
        for alias in values:
            aliases.setdefault(normalize_header(alias), set()).add(field)
    candidates: list[tuple[int, int, str]] = []
    for start in range(len(words)):
        for end in range(start + 1, min(len(words), start + 3) + 1):
            fields = aliases.get(normalize_header("".join(word.text for word in words[start:end])))
            if fields is None:
                continue
            if len(fields) != 1:
                raise ParseError("OCR_COLUMN_MAPPING_AMBIGUOUS")
            candidates.append((start, end, next(iter(fields))))
    return candidates


def _ocr_header_cells(words: tuple[_OcrWord, ...]) -> tuple[_OcrHeaderCell, ...]:
    candidates = _ocr_alias_candidates(words)
    # Duplicate non-overlapping aliases are a source ambiguity, never a reason
    # to select whichever one Tesseract happened to emit first.
    for field in {field for _, _, field in candidates}:
        positions = [(start, end) for start, end, candidate_field in candidates if candidate_field == field]
        if any(end_a <= start_b or end_b <= start_a for index, (start_a, end_a) in enumerate(positions) for start_b, end_b in positions[index + 1:]):
            raise ParseError("OCR_COLUMN_MAPPING_AMBIGUOUS")
    selected: list[tuple[int, int, str]] = []
    occupied: set[int] = set()
    fields: set[str] = set()
    for start, end, field in sorted(candidates, key=lambda item: (-(item[1] - item[0]), item[0], item[2])):
        if field in fields or any(index in occupied for index in range(start, end)):
            continue
        selected.append((start, end, field))
        occupied.update(range(start, end))
        fields.add(field)
    selected_by_start = {start: (end, field) for start, end, field in selected}
    cells: list[_OcrHeaderCell] = []
    index = 0
    while index < len(words):
        end, field = selected_by_start.get(index, (index + 1, None))
        segment = words[index:end]
        cells.append(_OcrHeaderCell(
            label="".join(word.text for word in segment),
            field=field,
            left=min(word.left for word in segment),
            right=max(word.left + word.width for word in segment),
            top=min(word.top for word in segment),
            bottom=max(word.top + word.height for word in segment),
            confidence_bps=min(word.confidence_bps for word in segment),
        ))
        index = end
    if len({cell.label for cell in cells}) != len(cells):
        raise ParseError("COLUMN_HEADER_DUPLICATE")
    return tuple(sorted(cells, key=lambda cell: (cell.left, cell.right, cell.label)))


def _ocr_required_fields(family: str, fields: set[str]) -> bool:
    if family == ACCOUNT_FAMILY:
        return {"company", "bank", "account", "ending"} <= fields
    if family in TRANSACTION_FAMILIES:
        # ``bank_id`` is optional and the sealed transaction schema has no
        # required source transaction identifier.  A deterministic source-row
        # identity is assigned later only for deduplication; it is never a
        # fabricated business identifier or financial value.
        identity = {"company", "account"} <= fields
        flow = "inflow" in fields or "outflow" in fields
        amount_direction = {"amount", "direction"} <= fields
        return identity and (flow or amount_direction) and not (flow and amount_direction)
    raise ParseError("DOCUMENT_FAMILY_UNSUPPORTED")


def _select_ocr_header(
    lines: tuple[tuple[_OcrWord, ...], ...],
    *,
    family: str,
    min_confidence_bps: int,
) -> tuple[int, tuple[_OcrHeaderCell, ...]]:
    candidates: list[tuple[int, tuple[_OcrHeaderCell, ...], int]] = []
    for index, words in enumerate(lines):
        try:
            cells = _ocr_header_cells(words)
        except ParseError:
            continue
        fields = {cell.field for cell in cells if cell.field is not None}
        if not _ocr_required_fields(family, fields):
            continue
        candidates.append((index, cells, len(fields)))
    if not candidates:
        raise ParseError("OCR_HEADER_MAPPING_MISSING")
    candidates.sort(key=lambda item: item[2], reverse=True)
    if len(candidates) > 1 and candidates[0][2] == candidates[1][2]:
        raise ParseError("OCR_HEADER_ROW_AMBIGUOUS")
    index, cells, _ = candidates[0]
    if any(cell.field is not None and cell.confidence_bps < min_confidence_bps for cell in cells):
        raise ParseError("OCR_LOW_CONFIDENCE")
    return index, cells


def _ocr_cell_index(word: _OcrWord, cells: tuple[_OcrHeaderCell, ...]) -> int:
    center_twice = 2 * word.left + word.width
    for index in range(len(cells) - 1):
        # Compare both centers using integers.  ``center_twice`` is twice the
        # word centre; the midpoint between adjacent header centres is one
        # half of the sum below, so its equivalent scale is
        # ``2 * center_twice < header_center_sum``.  Keeping this integral
        # avoids a rounded boundary moving a money field into its neighbour.
        header_center_sum = cells[index].left + cells[index].right + cells[index + 1].left + cells[index + 1].right
        if 2 * center_twice < header_center_sum:
            return index
    return len(cells) - 1


def _ocr_rows(
    lines: tuple[tuple[_OcrWord, ...], ...],
    *,
    header_index: int,
    cells: tuple[_OcrHeaderCell, ...],
    family: str,
    min_confidence_bps: int,
) -> list[dict[str, object]]:
    header_page = lines[header_index][0].page
    header_bottom = max(cell.bottom for cell in cells)
    field_by_label = {cell.label: cell.field for cell in cells}
    rows: list[dict[str, object]] = []
    for words in lines[header_index + 1:]:
        if words[0].page != header_page or min(word.top for word in words) <= header_bottom:
            continue
        by_index: dict[int, list[_OcrWord]] = {index: [] for index in range(len(cells))}
        for word in words:
            by_index[_ocr_cell_index(word, cells)].append(word)
        row: dict[str, object] = {}
        confidence: dict[str, int | None] = {}
        for index, cell in enumerate(cells):
            values = sorted(by_index[index], key=lambda word: (word.left, word.top, word.text))
            row[cell.label] = "".join(word.text for word in values)
            confidence[cell.label] = min((word.confidence_bps for word in values), default=None)
        mapped = {field: str(row[label] or "") for label, field in field_by_label.items() if field is not None}
        if not any(value for value in mapped.values()):
            continue
        fields = {field for field, value in mapped.items() if value}
        if not _ocr_required_fields(family, fields):
            raise ParseError("OCR_ROW_REQUIRED_CELL_MISSING")
        for label, field in field_by_label.items():
            if field is not None and row[label] and (confidence[label] is None or confidence[label] < min_confidence_bps):
                raise ParseError("OCR_LOW_CONFIDENCE")
        rows.append(row)
    if not rows:
        raise ParseError("SOURCE_ROWS_EMPTY")
    return rows


def _ocr_layout_fingerprint(
    *,
    family: str,
    evidence: ParserEvidence,
    cells: tuple[_OcrHeaderCell, ...],
) -> str:
    width = max(cell.right for cell in cells)
    if width <= 0:
        raise ParseError("OCR_LAYOUT_INVALID")
    layout = {
        "family": family,
        "format": evidence.format,
        "magic": evidence.magic,
        "columns": [cell.field or "UNMAPPED" for cell in cells],
        "centers_bps": [((cell.left + cell.right) * 10_000) // (2 * width) for cell in cells],
    }
    return sha256(json.dumps(layout, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest()


def deterministic_ocr_runtime_ready(*, runner: Callable[..., Any] = subprocess.run) -> bool:
    """Return whether the isolated image has the two offline OCR tools.

    This uses no document input and intentionally discards all command output.
    It is suitable for the values-free runtime isolation audit, not as a claim
    that any source layout or financial value has been verified.
    """

    try:
        languages = _run_ocr_command(["tesseract", "--list-langs"], runner=runner, failure_code="OCR_RUNTIME_UNAVAILABLE")
        _run_ocr_command(["pdfinfo", "-v"], runner=runner, failure_code="OCR_RUNTIME_UNAVAILABLE")
        _run_ocr_command(["pdftoppm", "-v"], runner=runner, failure_code="OCR_RUNTIME_UNAVAILABLE")
    except ParseError:
        return False
    return "chi_sim" in {line.strip() for line in languages.splitlines()}


def _parse_ocr_table(
    *,
    family: str,
    filename: str,
    source: SourceRef,
    evidence: ParserEvidence,
    lines: tuple[tuple[_OcrWord, ...], ...],
    min_confidence_bps: int,
) -> OcrParsedAttachment:
    """Parse one already-OCRed table under one exact source-family schema."""

    active_lines = lines
    try:
        header_index, cells = _select_ocr_header(
            active_lines,
            family=family,
            min_confidence_bps=min_confidence_bps,
        )
    except ParseError as exc:
        # This is a layout-only repair for a known Tesseract behaviour: cells
        # that visually share a table row may be emitted as distinct OCR
        # lines.  It is deliberately available only when the original pass
        # found no complete header at all.  An ambiguous header, low
        # confidence header or row/fact error remains a hard failure rather
        # than a reason to try a different semantic interpretation.
        if str(exc).split(":", 1)[0] != "OCR_HEADER_MAPPING_MISSING":
            raise
        visual_lines = tuple(words for _, words in _ocr_visual_rows(lines))
        if visual_lines == lines:
            raise
        active_lines = visual_lines
        header_index, cells = _select_ocr_header(
            active_lines,
            family=family,
            min_confidence_bps=min_confidence_bps,
        )
    rows = _ocr_rows(
        active_lines,
        header_index=header_index,
        cells=cells,
        family=family,
        min_confidence_bps=min_confidence_bps,
    )
    facts = _facts_from_rows(family=family, filename=filename, source=source, rows=rows, parser_evidence=evidence)
    return OcrParsedAttachment(
        facts=facts,
        layout_fingerprint=_ocr_layout_fingerprint(family=family, evidence=evidence, cells=cells),
    )


def _select_ocr_cashflow_header(
    lines: tuple[tuple[_OcrWord, ...], ...],
    *,
    min_confidence_bps: int,
) -> tuple[int, tuple[_OcrHeaderCell, ...]]:
    """Select the one strict receipt/payment table header in an OCR image.

    The chart-only profile deliberately requires both money directions, a
    date, and a bank column.  That is narrow enough to avoid treating an
    arbitrary image containing a number as financial evidence, while still
    matching the existing daily ``资金明细`` screenshot layout.
    """

    required = {"business_date", "inflow", "outflow", "bank"}
    candidates: list[tuple[int, tuple[_OcrHeaderCell, ...], int]] = []
    for index, words in enumerate(lines):
        try:
            cells = _ocr_header_cells(words)
        except ParseError:
            continue
        fields = {cell.field for cell in cells if cell.field is not None}
        if not required <= fields:
            continue
        candidates.append((index, cells, len(fields)))
    # Receipt/payment screenshots can put each heading cell in a distinct
    # Tesseract line despite their boxes being vertically aligned.  This is a
    # layout reassembly only; no looser alias or amount rule is introduced.
    if not candidates:
        for index, words in _ocr_visual_rows(lines):
            try:
                cells = _ocr_header_cells(words)
            except ParseError:
                continue
            fields = {cell.field for cell in cells if cell.field is not None}
            if not required <= fields:
                continue
            candidates.append((index, cells, len(fields)))
    if not candidates:
        return _select_ocr_cashflow_template_header(
            lines,
            min_confidence_bps=min_confidence_bps,
        )
    candidates.sort(key=lambda item: item[2], reverse=True)
    if len(candidates) > 1 and candidates[0][2] == candidates[1][2]:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_AMBIGUOUS")
    index, cells, _ = candidates[0]
    if any(cell.field in required and cell.confidence_bps < min_confidence_bps for cell in cells):
        raise ParseError("OCR_LOW_CONFIDENCE")
    return index, cells


_CASHFLOW_TEMPLATE_MAX_ANCHOR_VERTICAL_SPAN = 160
_CASHFLOW_TEMPLATE_EDGE_TOLERANCE = 16
_CASHFLOW_TEMPLATE_MIN_EDGE_OBSERVATIONS = 2
_CASHFLOW_TEMPLATE_MIN_MONEY_COLUMN_GAP = 32
_CASHFLOW_HEADERLESS_MIN_DATE_OBSERVATIONS = 2


def _cashflow_template_anchor_pair(
    lines: tuple[tuple[_OcrWord, ...], ...],
) -> tuple[int, _OcrHeaderCell, _OcrHeaderCell]:
    """Return the single exact date and category anchor pair for the layout.

    The source's date and category captions are explicit table labels.  They
    are not inferred from row content, and the pair must occur once on one
    page in left-to-right order.  This leaves all source layouts with a
    duplicated or incomplete anchor fail-closed.
    """

    anchors: dict[str, list[tuple[int, int, _OcrHeaderCell]]] = {
        "business_date": [],
        "category": [],
    }
    for line_index, words in enumerate(lines):
        try:
            cells = _ocr_header_cells(words)
        except ParseError:
            continue
        page = words[0].page
        for cell in cells:
            if cell.field in anchors:
                anchors[cell.field].append((page, line_index, cell))

    candidates: list[tuple[int, _OcrHeaderCell, _OcrHeaderCell]] = []
    for date_page, date_index, date_cell in anchors["business_date"]:
        for category_page, category_index, category_cell in anchors["category"]:
            if date_page != category_page or date_cell.right >= category_cell.left:
                continue
            vertical_span = abs(
                (date_cell.top + date_cell.bottom) - (category_cell.top + category_cell.bottom)
            )
            if vertical_span > _CASHFLOW_TEMPLATE_MAX_ANCHOR_VERTICAL_SPAN:
                continue
            candidates.append((min(date_index, category_index), date_cell, category_cell))

    if not candidates:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")
    if len(candidates) != 1:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_AMBIGUOUS")
    return candidates[0]


def _cashflow_template_money_cells(
    lines: tuple[tuple[_OcrWord, ...], ...],
    *,
    header_index: int,
    date_cell: _OcrHeaderCell,
    category_cell: _OcrHeaderCell,
    min_confidence_bps: int,
) -> tuple[_OcrHeaderCell, ...]:
    """Derive two fixed money columns from repeated right-aligned values.

    This is a bounded layout calibration, not value inference: only
    parseable high-confidence numeric OCR words to the right of the category
    anchor take part, each retained column needs repeated observations, and
    the first two stable columns must be materially separated.  The resulting
    cells are then subject to the existing per-row and footer-total gates.
    """

    header_page = lines[header_index][0].page
    header_bottom = max(date_cell.bottom, category_cell.bottom)
    numeric_words: list[_OcrWord] = []
    for words in lines[header_index + 1:]:
        if words[0].page != header_page or min(word.top for word in words) <= header_bottom:
            continue
        for word in words:
            if word.left + word.width <= category_cell.right or word.confidence_bps < min_confidence_bps:
                continue
            try:
                _cashflow_observation_amount(word.text)
            except ParseError:
                continue
            numeric_words.append(word)

    clusters: list[list[_OcrWord]] = []
    for word in sorted(numeric_words, key=lambda item: (item.left + item.width, item.left, item.top)):
        right = word.left + word.width
        if not clusters or right - (clusters[-1][-1].left + clusters[-1][-1].width) > _CASHFLOW_TEMPLATE_EDGE_TOLERANCE:
            clusters.append([word])
            continue
        clusters[-1].append(word)

    right_edges = [
        sorted(word.left + word.width for word in cluster)[len(cluster) // 2]
        for cluster in clusters
        if len(cluster) >= _CASHFLOW_TEMPLATE_MIN_EDGE_OBSERVATIONS
    ]
    if len(right_edges) < 2:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")
    outflow_right, inflow_right = right_edges[:2]
    if inflow_right - outflow_right < _CASHFLOW_TEMPLATE_MIN_MONEY_COLUMN_GAP:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")
    detail_left = date_cell.right + 1
    detail_right = category_cell.left - 1
    if detail_left >= detail_right:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")

    tail_center = inflow_right + ((inflow_right - outflow_right) // 2)
    return (
        _OcrHeaderCell(
            label="cashflow_template_date",
            field="business_date",
            left=date_cell.left,
            right=date_cell.right,
            top=date_cell.top,
            bottom=date_cell.bottom,
            confidence_bps=date_cell.confidence_bps,
        ),
        _OcrHeaderCell(
            label="cashflow_template_details",
            field=None,
            left=detail_left,
            right=detail_right,
            top=category_cell.top,
            bottom=category_cell.bottom,
            confidence_bps=min_confidence_bps,
        ),
        _OcrHeaderCell(
            label="cashflow_template_category",
            field=None,
            left=category_cell.left,
            right=category_cell.right,
            top=category_cell.top,
            bottom=category_cell.bottom,
            confidence_bps=category_cell.confidence_bps,
        ),
        _OcrHeaderCell(
            label="cashflow_template_outflow",
            field="outflow",
            left=category_cell.right + 1,
            right=outflow_right,
            top=category_cell.top,
            bottom=category_cell.bottom,
            confidence_bps=min_confidence_bps,
        ),
        _OcrHeaderCell(
            label="cashflow_template_inflow",
            field="inflow",
            left=outflow_right + 1,
            right=inflow_right,
            top=category_cell.top,
            bottom=category_cell.bottom,
            confidence_bps=min_confidence_bps,
        ),
        _OcrHeaderCell(
            label="cashflow_template_tail",
            field=None,
            left=tail_center - 1,
            right=tail_center + 1,
            top=category_cell.top,
            bottom=category_cell.bottom,
            confidence_bps=min_confidence_bps,
        ),
    )


def _select_ocr_cashflow_template_header(
    lines: tuple[tuple[_OcrWord, ...], ...],
    *,
    min_confidence_bps: int,
) -> tuple[int, tuple[_OcrHeaderCell, ...]]:
    """Build the verified fixed-layout profile when header aliases are split."""

    header_index, date_cell, category_cell = _cashflow_template_anchor_pair(lines)
    if date_cell.confidence_bps < min_confidence_bps or category_cell.confidence_bps < min_confidence_bps:
        raise ParseError("OCR_LOW_CONFIDENCE")
    return (
        header_index,
        _cashflow_template_money_cells(
            lines,
            header_index=header_index,
            date_cell=date_cell,
            category_cell=category_cell,
            min_confidence_bps=min_confidence_bps,
        ),
    )


def _select_ocr_cashflow_headerless_template(
    lines: tuple[tuple[_OcrWord, ...], ...],
    *,
    received_at: datetime,
    min_confidence_bps: int,
) -> tuple[tuple[tuple[_OcrWord, ...], ...], tuple[_OcrHeaderCell, ...]]:
    """Recover one fixed receipt/payment table whose header OCR is unreadable.

    This is deliberately narrower than the visible-header path.  It has no
    semantic alias expansion: the source must expose at least two date cells
    in one stable column, exactly one visible ``合计`` footer, and exactly two
    stable right-aligned amount columns.  The normal row/footer reconciliation then
    reuses those derived column boundaries.  A title, a lone date, a summary
    card, or an arbitrary image with numbers therefore cannot create a chart
    point through this recovery.
    """

    visual_rows = _ocr_visual_rows(lines)
    date_candidates: list[tuple[int, _OcrWord, date]] = []
    for row_index, words in visual_rows:
        for word in words:
            try:
                parsed = _cashflow_observation_date(word.text, received_at=received_at)
            except ParseError:
                continue
            date_candidates.append((row_index, word, parsed))

    date_clusters: list[list[tuple[int, _OcrWord, date]]] = []
    for candidate in sorted(date_candidates, key=lambda item: (item[1].page, item[1].left + item[1].width, item[1].top)):
        row_index, word, _ = candidate
        right = word.left + word.width
        if not date_clusters:
            date_clusters.append([candidate])
            continue
        previous = date_clusters[-1][-1][1]
        previous_right = previous.left + previous.width
        if word.page != previous.page or right - previous_right > _CASHFLOW_TEMPLATE_EDGE_TOLERANCE:
            date_clusters.append([candidate])
        else:
            date_clusters[-1].append(candidate)

    stable_date_clusters = [
        cluster
        for cluster in date_clusters
        if len({row_index for row_index, _, _ in cluster}) >= _CASHFLOW_HEADERLESS_MIN_DATE_OBSERVATIONS
        and len({parsed for _, _, parsed in cluster}) == 1
    ]
    if not stable_date_clusters:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")
    if len(stable_date_clusters) != 1:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_AMBIGUOUS")

    date_cluster = stable_date_clusters[0]
    if any(word.confidence_bps < min_confidence_bps for _, word, _ in date_cluster):
        raise ParseError("OCR_LOW_CONFIDENCE")
    date_page = date_cluster[0][1].page
    date_rows = {row_index for row_index, _, _ in date_cluster}
    date_left = min(word.left for _, word, _ in date_cluster)
    date_right = max(word.left + word.width for _, word, _ in date_cluster)

    footer_candidates = [
        (row_index, words)
        for row_index, words in visual_rows
        if words[0].page == date_page and "合计" in normalize_header("".join(word.text for word in words))
    ]
    if not footer_candidates:
        raise ParseError("CASHFLOW_OBSERVATION_TOTAL_MISSING")
    if len(footer_candidates) != 1:
        raise ParseError("CASHFLOW_OBSERVATION_TOTAL_AMBIGUOUS")
    footer_index, footer_words = footer_candidates[0]
    if not date_rows or max(date_rows) >= footer_index:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")

    body_rows = tuple(
        words
        for row_index, words in visual_rows
        if words[0].page == date_page and min(date_rows) <= row_index <= footer_index
    )
    footer_amount_words: list[_OcrWord] = []
    for word in footer_words:
        if word.left + word.width <= date_right:
            continue
        try:
            _cashflow_observation_amount(word.text)
        except ParseError:
            continue
        footer_amount_words.append(word)
    if len(footer_amount_words) < 2:
        raise ParseError("CASHFLOW_OBSERVATION_TOTAL_MISSING")
    # Geometry alone cannot safely assign an outflow/inflow meaning when a
    # footer has a third money-like column (for example a running balance).
    # The headerless path is deliberately unavailable in that shape; a
    # readable header must establish field identity instead.
    if len(footer_amount_words) != 2:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_AMBIGUOUS")
    footer_amount_words.sort(key=lambda word: (word.left + word.width, word.left, word.top))
    outflow_footer, inflow_footer = footer_amount_words[-2:]
    outflow_footer_right = outflow_footer.left + outflow_footer.width
    inflow_footer_right = inflow_footer.left + inflow_footer.width
    if inflow_footer_right - outflow_footer_right < _CASHFLOW_TEMPLATE_MIN_MONEY_COLUMN_GAP:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")
    if outflow_footer.confidence_bps < min_confidence_bps or inflow_footer.confidence_bps < min_confidence_bps:
        raise ParseError("OCR_LOW_CONFIDENCE")

    def column_words(right_edge: int) -> tuple[_OcrWord, ...]:
        matched: list[_OcrWord] = []
        for words in body_rows:
            for word in words:
                if abs((word.left + word.width) - right_edge) > _CASHFLOW_TEMPLATE_EDGE_TOLERANCE:
                    continue
                try:
                    _cashflow_observation_amount(word.text)
                except ParseError:
                    continue
                matched.append(word)
        return tuple(matched)

    outflow_words = column_words(outflow_footer_right)
    inflow_words = column_words(inflow_footer_right)
    if not outflow_words or not inflow_words:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")
    if any(word.confidence_bps < min_confidence_bps for word in (*outflow_words, *inflow_words)):
        raise ParseError("OCR_LOW_CONFIDENCE")
    if len(outflow_words) + len(inflow_words) < _CASHFLOW_HEADERLESS_MIN_DATE_OBSERVATIONS + 2:
        raise ParseError("CASHFLOW_OBSERVATION_ROWS_EMPTY")

    outflow_left = min(word.left for word in outflow_words)
    outflow_right = max(word.left + word.width for word in outflow_words)
    inflow_left = min(word.left for word in inflow_words)
    inflow_right = max(word.left + word.width for word in inflow_words)
    if date_right + 1 >= outflow_left or outflow_right >= inflow_left:
        raise ParseError("CASHFLOW_OBSERVATION_HEADER_MISSING")
    width = max(word.left + word.width for words in body_rows for word in words)
    if width <= inflow_right:
        width = inflow_right

    cells: list[_OcrHeaderCell] = []
    if date_left > 0:
        cells.append(_OcrHeaderCell(
            label="cashflow_headerless_prefix",
            field=None,
            left=0,
            right=date_left - 1,
            top=0,
            bottom=0,
            confidence_bps=min_confidence_bps,
        ))
    cells.append(_OcrHeaderCell(
        label="cashflow_headerless_date",
        field="business_date",
        left=date_left,
        right=date_right,
        top=0,
        bottom=0,
        confidence_bps=min(word.confidence_bps for _, word, _ in date_cluster),
    ))
    cells.append(_OcrHeaderCell(
        label="cashflow_headerless_details",
        field=None,
        left=date_right + 1,
        right=outflow_left - 1,
        top=0,
        bottom=0,
        confidence_bps=min_confidence_bps,
    ))
    cells.append(_OcrHeaderCell(
        label="cashflow_headerless_outflow",
        field="outflow",
        left=outflow_left,
        right=outflow_right,
        top=0,
        bottom=0,
        confidence_bps=min(word.confidence_bps for word in outflow_words),
    ))
    if outflow_right + 1 <= inflow_left - 1:
        cells.append(_OcrHeaderCell(
            label="cashflow_headerless_gap",
            field=None,
            left=outflow_right + 1,
            right=inflow_left - 1,
            top=0,
            bottom=0,
            confidence_bps=min_confidence_bps,
        ))
    cells.append(_OcrHeaderCell(
        label="cashflow_headerless_inflow",
        field="inflow",
        left=inflow_left,
        right=inflow_right,
        top=0,
        bottom=0,
        confidence_bps=min(word.confidence_bps for word in inflow_words),
    ))
    if inflow_right < width:
        cells.append(_OcrHeaderCell(
            label="cashflow_headerless_tail",
            field=None,
            left=inflow_right + 1,
            right=width,
            top=0,
            bottom=0,
            confidence_bps=min_confidence_bps,
        ))
    return body_rows, tuple(cells)


def _cashflow_observation_amount(value: object) -> int:
    text = str(value or "").strip()
    if not text:
        return 0
    try:
        amount = parse_amount_to_fen(text)
    except ContractError as exc:
        raise ParseError("CASHFLOW_OBSERVATION_AMOUNT_INVALID") from exc
    if amount < 0:
        raise ParseError("CASHFLOW_OBSERVATION_AMOUNT_NEGATIVE")
    return amount


_CASHFLOW_SHORT_DATE = re.compile(r"^(?P<month>0?[1-9]|1[0-2])月(?P<day>0?[1-9]|[12]\d|3[01])日?$")


def _cashflow_observation_date(value: object, *, received_at: datetime) -> date:
    """Resolve a screenshot's month/day cell against its source-message date.

    A source image regularly renders ``08月07日`` rather than a full year.
    The received timestamp supplies only the calendar year boundary; the
    parser rejects dates more than a year away instead of silently guessing an
    arbitrary period.
    """

    text = str(value or "").strip().replace(" ", "")
    if not text:
        raise ParseError("CASHFLOW_OBSERVATION_DATE_MISSING")
    try:
        parsed = _parse_date(text)
    except ParseError:
        matched = _CASHFLOW_SHORT_DATE.fullmatch(text)
        if matched is None:
            raise ParseError("CASHFLOW_OBSERVATION_DATE_INVALID")
        reference = received_at.date()
        try:
            parsed = date(reference.year, int(matched.group("month")), int(matched.group("day")))
        except ValueError as exc:
            raise ParseError("CASHFLOW_OBSERVATION_DATE_INVALID") from exc
        # A December report received in January belongs to the immediately
        # preceding year.  No other year inference is allowed.
        if (parsed - reference).days > 31:
            try:
                parsed = date(parsed.year - 1, parsed.month, parsed.day)
            except ValueError as exc:
                raise ParseError("CASHFLOW_OBSERVATION_DATE_INVALID") from exc
    reference = received_at.date()
    if abs((reference - parsed).days) > 366:
        raise ParseError("CASHFLOW_OBSERVATION_DATE_OUT_OF_RANGE")
    return parsed


def _ocr_cashflow_observation_totals(
    lines: tuple[tuple[_OcrWord, ...], ...],
    *,
    header_index: int,
    cells: tuple[_OcrHeaderCell, ...],
    received_at: datetime,
    min_confidence_bps: int,
) -> tuple[date, int, int]:
    """Return a zero-leak, footer-reconciled daily receipt/payment total.

    Individual receipt/payment rows are read only long enough to recompute the
    visible ``合计`` footer.  No row description, bank, counterparty or amount
    text is returned or persisted by this function.
    """

    header_page = lines[header_index][0].page
    header_bottom = max(cell.bottom for cell in cells)
    body_rows = tuple(
        words
        for words in lines[header_index + 1:]
        if words[0].page == header_page and min(word.top for word in words) > header_bottom
    )
    return _ocr_cashflow_observation_totals_from_rows(
        body_rows,
        cells=cells,
        received_at=received_at,
        min_confidence_bps=min_confidence_bps,
    )


def _ocr_cashflow_observation_totals_from_rows(
    rows: Iterable[tuple[_OcrWord, ...]],
    *,
    cells: tuple[_OcrHeaderCell, ...],
    received_at: datetime,
    min_confidence_bps: int,
) -> tuple[date, int, int]:
    """Reconcile one already-bounded set of table rows against its footer.

    The normal header path derives its rows from the selected OCR header.  The
    narrow headerless recovery below derives an equally bounded row span from
    repeated date cells and a visible footer.  Both paths deliberately share
    this exact amount/date/footer gate so the fallback cannot weaken it.
    """

    field_by_label = {cell.label: cell.field for cell in cells}
    row_inflow = 0
    row_outflow = 0
    business_dates: set[date] = set()
    total: tuple[int, int] | None = None
    row_count = 0
    for words in rows:
        by_index: dict[int, list[_OcrWord]] = {index: [] for index in range(len(cells))}
        for word in words:
            by_index[_ocr_cell_index(word, cells)].append(word)
        row: dict[str, str] = {}
        confidence: dict[str, int | None] = {}
        for index, cell in enumerate(cells):
            values = sorted(by_index[index], key=lambda word: (word.left, word.top, word.text))
            row[cell.label] = "".join(word.text for word in values)
            confidence[cell.label] = min((word.confidence_bps for word in values), default=None)
        mapped = {field: row[label] for label, field in field_by_label.items() if field is not None}
        inflow_text = mapped.get("inflow", "")
        outflow_text = mapped.get("outflow", "")
        row_text = normalize_header("".join(row.values()))
        is_total = "合计" in row_text
        if not inflow_text and not outflow_text:
            # Wrapped descriptions have neither money direction.  They cannot
            # change a footer total and are intentionally not treated as rows.
            if is_total:
                raise ParseError("CASHFLOW_OBSERVATION_TOTAL_MISSING")
            continue
        for label, field in field_by_label.items():
            if field in {"business_date", "inflow", "outflow", "bank"} and row[label] and (
                confidence[label] is None or confidence[label] < min_confidence_bps
            ):
                raise ParseError("OCR_LOW_CONFIDENCE")
        inflow = _cashflow_observation_amount(inflow_text)
        outflow = _cashflow_observation_amount(outflow_text)
        if inflow == 0 and outflow == 0:
            raise ParseError("CASHFLOW_OBSERVATION_ZERO_ROW")
        if is_total:
            if total is not None:
                raise ParseError("CASHFLOW_OBSERVATION_TOTAL_AMBIGUOUS")
            total = (inflow, outflow)
            continue
        business_dates.add(_cashflow_observation_date(mapped.get("business_date"), received_at=received_at))
        row_inflow += inflow
        row_outflow += outflow
        row_count += 1
    if row_count == 0:
        raise ParseError("CASHFLOW_OBSERVATION_ROWS_EMPTY")
    if len(business_dates) != 1:
        raise ParseError("CASHFLOW_OBSERVATION_DATE_AMBIGUOUS")
    if total is None:
        raise ParseError("CASHFLOW_OBSERVATION_TOTAL_MISSING")
    if total != (row_inflow, row_outflow):
        raise ParseError("CASHFLOW_OBSERVATION_TOTAL_MISMATCH")
    return next(iter(business_dates)), row_inflow, row_outflow


def _cashflow_observation_from_ocr(
    *,
    payload: bytes,
    evidence: ParserEvidence,
    runner: Callable[..., Any],
    psm: int,
    received_at: datetime,
    min_confidence_bps: int,
    preprocess_grid: bool = False,
    preprocess_enhanced: bool = False,
    preprocess_binarized: bool = False,
) -> tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]:
    """Apply one fixed OCR segmentation mode to the unchanged cashflow gate.

    This helper intentionally returns only the final date/totals and the
    layout cells needed for a non-sensitive layout fingerprint.  Individual
    OCR rows never escape the parser, including for the consensus fallback.
    """

    words = _parse_tesseract_tsv(_ocr_tsv(
        payload=payload,
        evidence=evidence,
        runner=runner,
        psm=psm,
        preprocess_grid=preprocess_grid,
        preprocess_enhanced=preprocess_enhanced,
        preprocess_binarized=preprocess_binarized,
    ))
    lines = _ocr_lines(words)
    headerless = False
    try:
        header_index, cells = _select_ocr_cashflow_header(
            lines,
            min_confidence_bps=min_confidence_bps,
        )
    except ParseError as exc:
        # The source can retain a verified, footer-reconciled table while OCR
        # cannot read its header captions.  Only the exact missing-header
        # case reaches the fixed geometry recovery; ambiguity, low confidence
        # or any row/footer failure stays a hard failure.
        if str(exc).split(":", 1)[0] != "CASHFLOW_OBSERVATION_HEADER_MISSING":
            raise
        headerless = True
        body_rows, cells = _select_ocr_cashflow_headerless_template(
            lines,
            received_at=received_at,
            min_confidence_bps=min_confidence_bps,
        )
        business_date, inflow_fen, outflow_fen = _ocr_cashflow_observation_totals_from_rows(
            body_rows,
            cells=cells,
            received_at=received_at,
            min_confidence_bps=min_confidence_bps,
        )
    else:
        business_date, inflow_fen, outflow_fen = _ocr_cashflow_observation_totals(
            lines,
            header_index=header_index,
            cells=cells,
            received_at=received_at,
            min_confidence_bps=min_confidence_bps,
        )
    return business_date, inflow_fen, outflow_fen, cells, headerless


def _cashflow_observation_from_grid_preprocessed_ocr(
    *,
    payload: bytes,
    evidence: ParserEvidence,
    runner: Callable[..., Any],
    received_at: datetime,
    min_confidence_bps: int,
) -> tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]:
    """Require two strict OCR readings after deterministic grid removal.

    This route is intentionally unavailable until the original image exhausted
    the established layout path.  Each result still independently validates
    its source table and footer; agreement is only over the three derived
    chart values, never over raw OCR text.
    """

    results: list[tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]] = []
    for psm in OCR_GRID_PREPROCESS_PSMS:
        try:
            results.append(_cashflow_observation_from_ocr(
                payload=payload,
                evidence=evidence,
                runner=runner,
                psm=psm,
                received_at=received_at,
                min_confidence_bps=min_confidence_bps,
                preprocess_grid=True,
            ))
        except ParseError as exc:
            if str(exc).split(":", 1)[0] not in _CASHFLOW_IMAGE_LAYOUT_RECOVERY_CODES:
                raise
            raise ParseError("CASHFLOW_OBSERVATION_GRID_PREPROCESS_CONSENSUS_MISSING") from None
    first, second = results
    if first[:3] != second[:3]:
        raise ParseError("CASHFLOW_OBSERVATION_GRID_PREPROCESS_DISAGREEMENT")
    return first


def _cashflow_observation_from_enhanced_ocr(
    *,
    payload: bytes,
    evidence: ParserEvidence,
    runner: Callable[..., Any],
    received_at: datetime,
    min_confidence_bps: int,
) -> tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]:
    """Require two strict readings after contrast-preserving rendering.

    This final layout repair remains unavailable until the original path has
    already stopped at a layout/OCR gate.  It is not a fallback after a
    financial field, row, date or footer decision, and cannot select values
    from one favourable OCR output.
    """

    results: list[tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]] = []
    for psm in OCR_GRID_PREPROCESS_PSMS:
        try:
            results.append(_cashflow_observation_from_ocr(
                payload=payload,
                evidence=evidence,
                runner=runner,
                psm=psm,
                received_at=received_at,
                min_confidence_bps=min_confidence_bps,
                preprocess_enhanced=True,
            ))
        except ParseError as exc:
            if str(exc).split(":", 1)[0] not in _CASHFLOW_IMAGE_LAYOUT_RECOVERY_CODES:
                raise
            raise ParseError("CASHFLOW_OBSERVATION_ENHANCED_PREPROCESS_CONSENSUS_MISSING") from None
    first, second = results
    if first[:3] != second[:3]:
        raise ParseError("CASHFLOW_OBSERVATION_ENHANCED_PREPROCESS_DISAGREEMENT")
    return first


def _cashflow_observation_from_binarized_ocr(
    *,
    payload: bytes,
    evidence: ParserEvidence,
    runner: Callable[..., Any],
    received_at: datetime,
    min_confidence_bps: int,
) -> tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]:
    """Require matching strict readings after fixed Otsu binarization."""

    results: list[tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]] = []
    for psm in OCR_GRID_PREPROCESS_PSMS:
        try:
            results.append(_cashflow_observation_from_ocr(
                payload=payload,
                evidence=evidence,
                runner=runner,
                psm=psm,
                received_at=received_at,
                min_confidence_bps=min_confidence_bps,
                preprocess_binarized=True,
            ))
        except ParseError as exc:
            if str(exc).split(":", 1)[0] not in _CASHFLOW_IMAGE_LAYOUT_RECOVERY_CODES:
                raise
            raise ParseError("CASHFLOW_OBSERVATION_BINARIZED_PREPROCESS_CONSENSUS_MISSING") from None
    first, second = results
    if first[:3] != second[:3]:
        raise ParseError("CASHFLOW_OBSERVATION_BINARIZED_PREPROCESS_DISAGREEMENT")
    return first


def parse_cashflow_observation(
    *,
    family: str,
    filename: str,
    payload: bytes,
    source: SourceRef,
    received_at: datetime,
    mime: str | None = None,
    min_confidence_bps: int = OCR_MIN_CONFIDENCE_BPS,
    runner: Callable[..., Any] = subprocess.run,
) -> CashflowObservation:
    """Parse one footer-reconciled receipt/payment screenshot for the UI only.

    It requires an explicit flow document family and does not return account,
    counterparty, bank, or individual-row fields.  Calling code must keep the
    result out of formal account/transaction reconciliation and publication.
    """

    if family not in TRANSACTION_FAMILIES:
        raise ParseError("CASHFLOW_OBSERVATION_FAMILY_UNSUPPORTED")
    if not isinstance(received_at, datetime):
        raise ParseError("CASHFLOW_OBSERVATION_RECEIVED_AT_INVALID")
    if not payload:
        raise ParseError("CORRUPT_ATTACHMENT")
    _validate_source(source, payload)
    threshold = _validate_ocr_min_confidence(min_confidence_bps)
    evidence = replace(
        inspect_ocr_attachment_format(filename=filename, payload=payload, mime=mime),
        parser_version=CASHFLOW_OBSERVATION_PARSER_VERSION,
    )

    def require_headerless_consensus(
        candidate: tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool],
    ) -> tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]:
        """Require two independent layout passes for a headerless result."""

        consensus: list[tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]] = []
        for consensus_psm in OCR_CASHFLOW_CONSENSUS_PSMS:
            try:
                consensus.append(_cashflow_observation_from_ocr(
                    payload=payload,
                    evidence=evidence,
                    runner=runner,
                    psm=consensus_psm,
                    received_at=received_at,
                    min_confidence_bps=threshold,
                ))
            except ParseError as exc:
                if str(exc).split(":", 1)[0] not in _CASHFLOW_IMAGE_LAYOUT_RECOVERY_CODES:
                    raise
                raise ParseError("CASHFLOW_OBSERVATION_LAYOUT_CONSENSUS_MISSING") from None
        if any(result[:3] != candidate[:3] for result in consensus):
            raise ParseError("CASHFLOW_OBSERVATION_LAYOUT_CONSENSUS_DISAGREEMENT")
        return candidate

    def parse_original_layout() -> tuple[date, int, int, tuple[_OcrHeaderCell, ...]]:
        """Run the pre-v9 strict layout path without altering its semantics."""

        try:
            primary = _cashflow_observation_from_ocr(
                payload=payload,
                evidence=evidence,
                runner=runner,
                psm=OCR_PRIMARY_PSM,
                received_at=received_at,
                min_confidence_bps=threshold,
            )
        except ParseError as exc:
            # Some real screenshot tables have a complete visual header but
            # PSM 6 emits it as sparse text.  Re-run exactly once under PSM 11
            # only when the first pass reached the missing-header gate.  A
            # header ambiguity, low confidence result, or any later row/footer
            # error must never be retried under a different interpretation.
            if str(exc).split(":", 1)[0] != "CASHFLOW_OBSERVATION_HEADER_MISSING":
                raise
            try:
                fallback = _cashflow_observation_from_ocr(
                    payload=payload,
                    evidence=evidence,
                    runner=runner,
                    psm=OCR_HEADER_FALLBACK_PSM,
                    received_at=received_at,
                    min_confidence_bps=threshold,
                )
            except ParseError as fallback_exc:
                if str(fallback_exc).split(":", 1)[0] != "CASHFLOW_OBSERVATION_HEADER_MISSING":
                    raise
                # The two alternate page-segmentation modes are not ordinary
                # retries.  Both must independently reach the complete strict
                # result and agree exactly; this remains a layout-only recovery
                # rather than a way to choose whichever OCR output has a value.
                consensus: list[tuple[date, int, int, tuple[_OcrHeaderCell, ...], bool]] = []
                for consensus_psm in OCR_CASHFLOW_CONSENSUS_PSMS:
                    try:
                        consensus.append(_cashflow_observation_from_ocr(
                            payload=payload,
                            evidence=evidence,
                            runner=runner,
                            psm=consensus_psm,
                            received_at=received_at,
                            min_confidence_bps=threshold,
                        ))
                    except ParseError as consensus_exc:
                        if str(consensus_exc).split(":", 1)[0] not in _CASHFLOW_IMAGE_LAYOUT_RECOVERY_CODES:
                            raise
                        continue
                if len(consensus) != len(OCR_CASHFLOW_CONSENSUS_PSMS):
                    raise ParseError("CASHFLOW_OBSERVATION_LAYOUT_CONSENSUS_MISSING")
                first, second = consensus
                if first[:3] != second[:3]:
                    raise ParseError("CASHFLOW_OBSERVATION_LAYOUT_CONSENSUS_DISAGREEMENT")
                business_date, inflow_fen, outflow_fen, cells, _ = first
            else:
                if fallback[4]:
                    fallback = require_headerless_consensus(fallback)
                business_date, inflow_fen, outflow_fen, cells, _ = fallback
        else:
            if primary[4]:
                primary = require_headerless_consensus(primary)
            business_date, inflow_fen, outflow_fen, cells, _ = primary
        return business_date, inflow_fen, outflow_fen, cells

    try:
        business_date, inflow_fen, outflow_fen, cells = parse_original_layout()
    except ParseError as original_error:
        # Grid removal is purely a deterministic image-layout repair.  It is
        # not allowed after a date, amount, field identity, row, or footer
        # decision.  When it cannot reproduce a strict result, preserve the
        # original public-safe failure rather than exposing preprocessing
        # internals or weakening the original gate.
        original_code = str(original_error).split(":", 1)[0]
        if original_code == "CASHFLOW_OBSERVATION_LAYOUT_CONSENSUS_DISAGREEMENT":
            raise ParseError("CASHFLOW_OBSERVATION_LAYOUT_CONSENSUS_MISSING") from None
        if original_code not in _CASHFLOW_ORIGINAL_LAYOUT_RECOVERY_CODES:
            raise
        try:
            business_date, inflow_fen, outflow_fen, cells, _ = _cashflow_observation_from_grid_preprocessed_ocr(
                payload=payload,
                evidence=evidence,
                runner=runner,
                received_at=received_at,
                min_confidence_bps=threshold,
            )
        except ParseError as grid_error:
            if str(grid_error).split(":", 1)[0] != "CASHFLOW_OBSERVATION_GRID_PREPROCESS_CONSENSUS_MISSING":
                raise original_error from None
            try:
                business_date, inflow_fen, outflow_fen, cells, _ = _cashflow_observation_from_enhanced_ocr(
                    payload=payload,
                    evidence=evidence,
                    runner=runner,
                    received_at=received_at,
                    min_confidence_bps=threshold,
                )
            except ParseError as enhanced_error:
                if str(enhanced_error).split(":", 1)[0] != "CASHFLOW_OBSERVATION_ENHANCED_PREPROCESS_CONSENSUS_MISSING":
                    raise original_error from None
                try:
                    business_date, inflow_fen, outflow_fen, cells, _ = _cashflow_observation_from_binarized_ocr(
                        payload=payload,
                        evidence=evidence,
                        runner=runner,
                        received_at=received_at,
                        min_confidence_bps=threshold,
                    )
                except ParseError:
                    raise original_error from None
    return CashflowObservation(
        business_date=business_date,
        inflow_fen=inflow_fen,
        outflow_fen=outflow_fen,
        source=source,
        parser_evidence=evidence,
        layout_fingerprint=_ocr_layout_fingerprint(
            family="cashflow_observation",
            evidence=evidence,
            cells=cells,
        ),
    )


_PAYMENT_REQUEST_OCR_PSMS = (6, 11, 12)
_PAYMENT_REQUEST_SHEET_LAYOUT = "SHEET"
_PAYMENT_REQUEST_MESSAGE_STRIP_LAYOUT = "MESSAGE_STRIP"
_PAYMENT_REQUEST_SHEET_MIN_WIDTH = 640
_PAYMENT_REQUEST_SHEET_MIN_HEIGHT = 900
_PAYMENT_REQUEST_STRIP_MIN_WIDTH = 960
_PAYMENT_REQUEST_STRIP_MIN_HEIGHT = 200
_PAYMENT_REQUEST_STRIP_MIN_ASPECT_RATIO = 4.0
_PAYMENT_REQUEST_STRIP_MAX_ASPECT_RATIO = 6.0
_PAYMENT_REQUEST_MAX_PIXELS = 16_000_000
_PAYMENT_REQUEST_MAX_REPORT_LAG_DAYS = 7
_PAYMENT_REQUEST_SHEET_CROPS = {
    # These relative crops describe the frozen ``待付款请示明细表`` layout.
    # The parser never searches arbitrary image text for a number: every
    # accepted value must appear in the labelled grand-total cell.
    "title": (0.00, 0.00, 0.65, 0.035),
    "business_date": (0.65, 0.00, 0.95, 0.035),
    "grand_total_label": (0.60, 0.985, 0.76, 1.00),
    "grand_total": (0.74, 0.985, 0.87, 1.00),
}
_PAYMENT_REQUEST_MESSAGE_STRIP_CROPS = {
    # The exact-group message summary is a horizontal export strip.  Its
    # footer retains the fixed label/amount cells while the document title and
    # business-date cells are outside the captured area.
    "strip_grand_total_label": (0.58, 0.80, 0.78, 0.995),
    "strip_grand_total": (0.74, 0.80, 0.96, 0.995),
}
_PAYMENT_REQUEST_DATE = re.compile(
    r"(?P<year>20\d{2})\s*[-./]\s*(?P<month>\d{1,2})\s*[-./]\s*(?P<day>\d{1,2})"
)
_PAYMENT_REQUEST_AMOUNT = re.compile(r"\d+(?:\.\d{1,2})?")


def _payment_request_layout_and_crops(
    *,
    width: int,
    height: int,
) -> tuple[str, Mapping[str, tuple[float, float, float, float]]]:
    """Select one frozen payment-request layout from image geometry."""

    if width * height > _PAYMENT_REQUEST_MAX_PIXELS:
        raise ParseError("PAYMENT_REQUEST_IMAGE_DIMENSIONS_UNSUPPORTED")
    if width >= _PAYMENT_REQUEST_SHEET_MIN_WIDTH and height >= _PAYMENT_REQUEST_SHEET_MIN_HEIGHT:
        return _PAYMENT_REQUEST_SHEET_LAYOUT, _PAYMENT_REQUEST_SHEET_CROPS
    aspect_ratio = width / height if height else 0.0
    if (
        width >= _PAYMENT_REQUEST_STRIP_MIN_WIDTH
        and height >= _PAYMENT_REQUEST_STRIP_MIN_HEIGHT
        and _PAYMENT_REQUEST_STRIP_MIN_ASPECT_RATIO <= aspect_ratio <= _PAYMENT_REQUEST_STRIP_MAX_ASPECT_RATIO
    ):
        return _PAYMENT_REQUEST_MESSAGE_STRIP_LAYOUT, _PAYMENT_REQUEST_MESSAGE_STRIP_CROPS
    raise ParseError("PAYMENT_REQUEST_IMAGE_DIMENSIONS_UNSUPPORTED")


def _payment_request_crop_texts(
    *,
    payload: bytes,
    evidence: ParserEvidence,
    runner: Callable[..., Any],
) -> tuple[str, dict[str, tuple[str, ...]]]:
    """Read only the fixed visual cells required by one approved layout."""

    if evidence.magic not in {"PNG", "JPEG", "BMP", "WEBP", "PDF"}:
        raise ParseError("PAYMENT_REQUEST_IMAGE_UNSUPPORTED")
    try:
        from PIL import Image
    except ImportError as exc:
        raise ParseError("OCR_RUNTIME_UNAVAILABLE") from exc

    with tempfile.TemporaryDirectory(prefix="daily-funds-payment-request-") as temporary:
        root = Path(temporary)
        source = root / f"input{evidence.suffix}"
        source.write_bytes(payload)
        try:
            image_path = source
            if evidence.magic == "PDF":
                try:
                    page_count = _pdf_page_count(source, runner=runner)
                except ParseError as exc:
                    raise ParseError("PAYMENT_REQUEST_PDF_METADATA_INVALID") from exc
                if page_count != 1:
                    raise ParseError("PAYMENT_REQUEST_PDF_PAGE_AMBIGUOUS")
                prefix = root / "payment-request-render"
                _run_ocr_command(
                    ["pdftoppm", "-png", "-r", "300", "-f", "1", "-l", "1", str(source), str(prefix)],
                    runner=runner,
                    failure_code="PAYMENT_REQUEST_PDF_RENDER_FAILED",
                )
                rendered = tuple(sorted(root.glob("payment-request-render-*.png")))
                if len(rendered) != 1 or rendered[0].is_symlink() or not rendered[0].is_file():
                    raise ParseError("PAYMENT_REQUEST_PDF_RENDER_FAILED")
                image_path = rendered[0]
            with Image.open(image_path) as image:
                image.load()
                width, height = image.size
                layout, crops = _payment_request_layout_and_crops(width=width, height=height)
                rendered = image.convert("RGB")
                regions: dict[str, Path] = {}
                for name, (left, top, right, bottom) in crops.items():
                    bounds = (
                        int(width * left),
                        int(height * top),
                        int(width * right),
                        int(height * bottom),
                    )
                    crop = rendered.crop(bounds)
                    if crop.width < 8 or crop.height < 8:
                        raise ParseError("PAYMENT_REQUEST_IMAGE_DIMENSIONS_UNSUPPORTED")
                    scaled = crop.resize(
                        (crop.width * 4, crop.height * 4),
                        # Bicubic is the fixed rendering used by the
                        # validated source profile.  It preserves the thin
                        # Chinese glyph strokes in the title and footer cells
                        # better than a sharpening-oriented resampler.
                        resample=Image.Resampling.BICUBIC,
                    )
                    path = root / f"payment-{name}.png"
                    scaled.save(path, format="PNG")
                    regions[name] = path
        except ParseError:
            raise
        except (OSError, ValueError) as exc:
            raise ParseError("PAYMENT_REQUEST_IMAGE_INVALID") from exc

        output: dict[str, tuple[str, ...]] = {}
        for name, path in regions.items():
            language = "eng" if name.endswith("grand_total") else OCR_LANGUAGE
            values: list[str] = []
            for psm in _PAYMENT_REQUEST_OCR_PSMS:
                command = [
                    "tesseract", str(path), "stdout", "-l", language,
                    "--psm", str(psm),
                ]
                if name.endswith("grand_total"):
                    command.extend(("-c", "tessedit_char_whitelist=0123456789.,"))
                values.append(_run_ocr_command(
                    command,
                    runner=runner,
                    failure_code="PAYMENT_REQUEST_OCR_ENGINE_FAILED",
                ).strip())
            output[name] = tuple(values)
        return layout, output


def _payment_request_date(value: str) -> date:
    matches = tuple(_PAYMENT_REQUEST_DATE.finditer(value))
    if len(matches) != 1:
        raise ParseError("PAYMENT_REQUEST_DATE_INVALID")
    match = matches[0]
    try:
        return date(
            int(match.group("year")),
            int(match.group("month")),
            int(match.group("day")),
        )
    except ValueError as exc:
        raise ParseError("PAYMENT_REQUEST_DATE_INVALID") from exc


def _payment_request_amount(value: str) -> int:
    matches = tuple(_PAYMENT_REQUEST_AMOUNT.findall(value))
    if len(matches) != 1:
        raise ParseError("PAYMENT_REQUEST_TOTAL_INVALID")
    try:
        amount = parse_amount_to_fen(matches[0])
    except ContractError as exc:
        raise ParseError("PAYMENT_REQUEST_TOTAL_INVALID") from exc
    if amount <= 0:
        raise ParseError("PAYMENT_REQUEST_TOTAL_INVALID")
    return amount


def _payment_request_consensus(values: tuple[object, ...], *, code: str) -> object:
    if len(values) != len(_PAYMENT_REQUEST_OCR_PSMS) or any(value in (None, "") for value in values):
        raise ParseError(code)
    if len(set(values)) != 1:
        raise ParseError(code)
    return values[0]


def _payment_request_layout_fingerprint(
    *,
    layout: str,
    evidence: ParserEvidence,
    crops: Mapping[str, tuple[float, float, float, float]],
) -> str:
    return sha256(
        "\x1f".join((
            "payment_request_observation.v2",
            layout,
            evidence.format,
            evidence.magic,
            *(
                f"{name}:{left:.3f},{top:.3f},{right:.3f},{bottom:.3f}"
                for name, (left, top, right, bottom) in crops.items()
            ),
        )).encode("utf-8")
    ).hexdigest()


def parse_payment_request_observation(
    *,
    filename: str,
    payload: bytes,
    source: SourceRef,
    received_at: datetime,
    mime: str | None = None,
    runner: Callable[..., Any] = subprocess.run,
) -> PaymentRequestObservation | None:
    """Return one verified daily payment-request total or ``None`` for other images.

    A full sheet uses its title, document date, label and total.  The approved
    horizontal message summary has no title/date cells, so all three OCR reads
    must identify its fixed total label and amount; its date is the exact DWS
    message day.  Other exact-group images remain non-candidates.
    """

    if not isinstance(received_at, datetime):
        raise ParseError("PAYMENT_REQUEST_RECEIVED_AT_INVALID")
    if not payload:
        raise ParseError("CORRUPT_ATTACHMENT")
    _validate_source(source, payload)
    evidence = replace(
        inspect_ocr_attachment_format(filename=filename, payload=payload, mime=mime),
        parser_version=PAYMENT_REQUEST_OBSERVATION_PARSER_VERSION,
    )
    layout, regions = _payment_request_crop_texts(payload=payload, evidence=evidence, runner=runner)
    if layout == _PAYMENT_REQUEST_MESSAGE_STRIP_LAYOUT:
        grand_total_label_votes = sum(
            "总" in value and "合计" in value
            for value in regions["strip_grand_total_label"]
        )
        if grand_total_label_votes == 0:
            return None
        if grand_total_label_votes != len(_PAYMENT_REQUEST_OCR_PSMS):
            raise ParseError("PAYMENT_REQUEST_GRAND_TOTAL_LABEL_MISSING")
        request_total_fen = _payment_request_consensus(
            tuple(_payment_request_amount(value) for value in regions["strip_grand_total"]),
            code="PAYMENT_REQUEST_TOTAL_CONSENSUS_MISSING",
        )
        assert isinstance(request_total_fen, int)
        from zoneinfo import ZoneInfo

        return PaymentRequestObservation(
            business_date=received_at.astimezone(ZoneInfo("Asia/Shanghai")).date(),
            date_basis="MESSAGE_DAY",
            request_total_fen=request_total_fen,
            source=source,
            parser_evidence=evidence,
            layout_fingerprint=_payment_request_layout_fingerprint(
                layout=layout,
                evidence=evidence,
                crops=_PAYMENT_REQUEST_MESSAGE_STRIP_CROPS,
            ),
        )

    title_votes = sum("付款" in value for value in regions["title"])
    if title_votes == 0:
        return None
    if title_votes != len(_PAYMENT_REQUEST_OCR_PSMS):
        raise ParseError("PAYMENT_REQUEST_TITLE_AMBIGUOUS")

    business_date = _payment_request_consensus(
        tuple(_payment_request_date(value) for value in regions["business_date"]),
        code="PAYMENT_REQUEST_DATE_CONSENSUS_MISSING",
    )
    assert isinstance(business_date, date)
    grand_total_label_votes = sum(
        "总" in value and "合计" in value
        for value in regions["grand_total_label"]
    )
    if grand_total_label_votes < 2:
        raise ParseError("PAYMENT_REQUEST_GRAND_TOTAL_LABEL_MISSING")
    request_total_fen = _payment_request_consensus(
        tuple(_payment_request_amount(value) for value in regions["grand_total"]),
        code="PAYMENT_REQUEST_TOTAL_CONSENSUS_MISSING",
    )
    assert isinstance(request_total_fen, int)

    from zoneinfo import ZoneInfo

    received_day = received_at.astimezone(ZoneInfo("Asia/Shanghai")).date()
    lag_days = (received_day - business_date).days
    if lag_days < 0 or lag_days > _PAYMENT_REQUEST_MAX_REPORT_LAG_DAYS:
        raise ParseError("PAYMENT_REQUEST_DATE_OUT_OF_RANGE")
    return PaymentRequestObservation(
        business_date=business_date,
        date_basis="DOCUMENT_DAY",
        request_total_fen=request_total_fen,
        source=source,
        parser_evidence=evidence,
        layout_fingerprint=_payment_request_layout_fingerprint(
            layout=layout,
            evidence=evidence,
            crops=_PAYMENT_REQUEST_SHEET_CROPS,
        ),
    )


def _generic_ocr_unresolved_code(failures: Iterable[ParseError]) -> str:
    """Return a conservative values-free reason for zero generic candidates.

    The two strict schemas are still both required to fail.  A more specific
    code is returned only if every failed candidate stopped in the same broad
    OCR phase; mixed or unknown failures retain the original fail-closed code.
    """

    codes = {str(failure).split(":", 1)[0] for failure in failures}
    if codes and codes <= _GENERIC_OCR_HEADER_PHASE_CODES:
        return "OCR_GENERIC_HEADER_SCHEMA_MISSING"
    if codes and codes <= _GENERIC_OCR_ROW_PHASE_CODES:
        return "OCR_GENERIC_ROW_SCHEMA_MISSING"
    if codes and codes <= _GENERIC_OCR_CONFIDENCE_PHASE_CODES:
        return "OCR_GENERIC_CONFIDENCE_BLOCKED"
    return "OCR_GENERIC_FAMILY_UNRESOLVED"


def parse_ocr_attachment(
    *,
    family: str,
    filename: str,
    payload: bytes,
    source: SourceRef,
    mime: str | None = None,
    min_confidence_bps: int = OCR_MIN_CONFIDENCE_BPS,
    runner: Callable[..., Any] = subprocess.run,
) -> OcrParsedAttachment:
    """Open one image/scanned-PDF through a deterministic offline OCR table path.

    This function proves only a strict, in-memory parser-open.  The runtime
    adds the separate two-business-day layout calibration gate before it can
    treat the resulting facts as a supported source for reconciliation.
    """

    if family not in TRANSACTION_FAMILIES | {ACCOUNT_FAMILY}:
        raise ParseError("DOCUMENT_FAMILY_UNSUPPORTED")
    if not payload:
        raise ParseError("CORRUPT_ATTACHMENT")
    _validate_source(source, payload)
    threshold = _validate_ocr_min_confidence(min_confidence_bps)
    evidence = inspect_ocr_attachment_format(filename=filename, payload=payload, mime=mime)
    words = _parse_tesseract_tsv(_ocr_tsv(payload=payload, evidence=evidence, runner=runner))
    lines = _ocr_lines(words)
    # ``资金明细`` is an allowed transaction family, but it is also the one
    # generic source label used by historical image messages.  Do not let that
    # label force an account-looking table through the transaction schema.  We
    # accept the image only when the already-produced OCR table validates under
    # exactly one of the two complete fact schemas; zero or two matches remain
    # fail-closed.  OCR itself executes once, so this adds no second raw read
    # or a heuristic retry path.
    candidate_families = (ACCOUNT_FAMILY, "资金明细") if family == "资金明细" else (family,)
    def parse_candidates(active_lines: tuple[tuple[_OcrWord, ...], ...]) -> tuple[list[OcrParsedAttachment], list[ParseError]]:
        candidates: list[OcrParsedAttachment] = []
        failures: list[ParseError] = []
        for candidate_family in candidate_families:
            try:
                candidates.append(_parse_ocr_table(
                    family=candidate_family,
                    filename=filename,
                    source=source,
                    evidence=evidence,
                    lines=active_lines,
                    min_confidence_bps=threshold,
                ))
            except ParseError as exc:
                failures.append(exc)
        return candidates, failures

    candidates, failures = parse_candidates(lines)
    # Some real table screenshots emit a sparse, non-tabular OCR layout under
    # PSM 6 even though their standard headers and rows are visible under the
    # deterministic sparse-text mode.  This fallback is intentionally narrow:
    # it runs once, only after *every* candidate reached the exact header-missing
    # gate.  Any ambiguity, low confidence, invalid row, amount or source
    # integrity condition remains a hard failure without a second OCR attempt.
    if candidates == [] and failures and all(
        str(failure).split(":", 1)[0] == "OCR_HEADER_MAPPING_MISSING"
        for failure in failures
    ):
        fallback_words = _parse_tesseract_tsv(_ocr_tsv(
            payload=payload,
            evidence=evidence,
            runner=runner,
            psm=OCR_HEADER_FALLBACK_PSM,
        ))
        candidates, failures = parse_candidates(_ocr_lines(fallback_words))

    if len(candidates) == 1:
        return candidates[0]
    if family == "资金明细":
        if len(candidates) > 1:
            raise ParseError("OCR_GENERIC_FAMILY_AMBIGUOUS")
        raise ParseError(_generic_ocr_unresolved_code(failures))
    # A non-generic source family has exactly one candidate, so preserving the
    # original strict parser code is both more precise and backward compatible.
    raise failures[0]


def _validate_source(source: SourceRef, payload: bytes) -> None:
    if not all(isinstance(value, str) and _SHA256.fullmatch(value) for value in (
        source.attachment_sha256,
        source.message_id_hash,
        source.source_version,
    )):
        raise ParseError("SOURCE_LINEAGE_INVALID")
    if source.source_version != source.attachment_sha256:
        raise ParseError("SOURCE_VERSION_MISMATCH")
    if sha256(payload).hexdigest() != source.attachment_sha256:
        raise ParseError("SOURCE_PAYLOAD_HASH_MISMATCH")
    match = _OCCURRENCE_PATH.fullmatch(source.occurrence_path)
    if match is None or match.group("message") != source.message_id_hash:
        raise ParseError("SOURCE_LINEAGE_INVALID")
    try:
        date(int(match.group("year")), int(match.group("month")), int(match.group("day")))
    except ValueError as exc:
        raise ParseError("SOURCE_LINEAGE_INVALID") from exc


def _business_date(rows: list[dict[str, object]], mapped: Mapping[str, str], filename: str) -> date:
    filename_date = _date_from_filename(filename)
    if "business_date" in mapped:
        values = {_parse_date(row.get(mapped["business_date"])) for row in rows}
        if len(values) != 1:
            raise ParseError("BUSINESS_DATE_AMBIGUOUS")
        business_date = next(iter(values))
        if filename_date is not None and filename_date != business_date:
            raise ParseError("BUSINESS_DATE_FILENAME_MISMATCH")
        return business_date
    if filename_date is None:
        raise ParseError("BUSINESS_DATE_MISSING")
    return filename_date


def _facts_from_rows(
    *,
    family: str,
    filename: str,
    source: SourceRef,
    rows: list[dict[str, object]],
    parser_evidence: ParserEvidence,
) -> ParsedFacts:
    """Build one fact family from rows already opened by a strict parser."""

    if family not in TRANSACTION_FAMILIES | {ACCOUNT_FAMILY}:
        raise ParseError("DOCUMENT_FAMILY_UNSUPPORTED")
    if not rows:
        raise ParseError("SOURCE_ROWS_EMPTY")
    mapped = _column_map(rows[0].keys())
    business_date = _business_date(rows, mapped, filename)
    if family == ACCOUNT_FAMILY:
        _required(mapped, ("company", "bank", "account", "ending"))
        accounts: list[AccountSnapshot] = []
        seen_accounts: set[tuple[date, str, str, str]] = set()
        for row in rows:
            company = _required_text(row, mapped["company"], "ACCOUNT_COMPANY_MISSING")
            bank = _required_text(row, mapped["bank"], "ACCOUNT_BANK_MISSING")
            account = _required_identifier(
                row,
                mapped["account"],
                "ACCOUNT_NUMBER_MISSING",
                "ACCOUNT_NUMBER_NON_TEXT",
            )
            key = (business_date, company, bank, account)
            if key in seen_accounts:
                raise ParseError("ACCOUNT_SNAPSHOT_DUPLICATE")
            seen_accounts.add(key)
            ending = _amount_to_fen(row.get(mapped["ending"]))
            opening = _amount_to_fen(row.get(mapped["opening"])) if mapped.get("opening") and not _is_blank(row.get(mapped["opening"])) else None
            currency = str(row.get(mapped.get("currency", "")) or "CNY").strip().upper() or "CNY"
            if currency != "CNY":
                raise ParseError("CURRENCY_UNSUPPORTED")
            accounts.append(AccountSnapshot(business_date, company, bank, account, ending, opening, currency, source))
        return ParsedFacts(business_date, family, tuple(accounts), tuple(), source.source_version, parser_evidence)

    # The sealed transaction schema requires a company/account alias but makes
    # bank_id optional and does not require a source transaction identifier.
    # Retain a provided identifier when present; otherwise the stable row
    # position within this immutable attachment is a fact-local identity used
    # exclusively for deduplication.
    _required(mapped, ("company", "account"))
    has_flow_columns = "inflow" in mapped or "outflow" in mapped
    has_amount_direction = {"amount", "direction"} <= set(mapped)
    # These are two alternative financial encodings.  Without a frozen real
    # template proving their relationship, preferring one and ignoring the
    # other could silently publish a different amount than the attachment
    # states.  Treat coexistence as ambiguity rather than an implicit choice.
    if has_flow_columns and has_amount_direction:
        raise ParseError("TRANSACTION_AMOUNT_MAPPING_AMBIGUOUS")
    if not has_flow_columns and not has_amount_direction:
        raise ParseError("TRANSACTION_AMOUNT_MAPPING_MISSING")
    transactions: list[Transaction] = []
    seen_transactions: set[tuple[date, str, str | None, str, str]] = set()
    for row_index, row in enumerate(rows, start=1):
        company = _required_text(row, mapped["company"], "TRANSACTION_COMPANY_MISSING")
        bank = (
            _required_text(row, mapped["bank"], "TRANSACTION_BANK_MISSING")
            if "bank" in mapped and not _is_blank(row.get(mapped["bank"]))
            else None
        )
        account = _required_identifier(
            row,
            mapped["account"],
            "TRANSACTION_ACCOUNT_MISSING",
            "TRANSACTION_ACCOUNT_NON_TEXT",
        )
        transaction_id = (
            _required_identifier(
                row,
                mapped["transaction_id"],
                "TRANSACTION_ID_MISSING",
                "TRANSACTION_ID_NON_TEXT",
            )
            if "transaction_id" in mapped and not _is_blank(row.get(mapped["transaction_id"]))
            else f"source-row-{row_index}"
        )
        key = (business_date, company, bank, account, transaction_id)
        if key in seen_transactions:
            raise ParseError("TRANSACTION_DUPLICATE")
        seen_transactions.add(key)
        if has_flow_columns:
            inflow = _amount_to_fen(row.get(mapped["inflow"])) if "inflow" in mapped and not _is_blank(row.get(mapped["inflow"])) else 0
            outflow = _amount_to_fen(row.get(mapped["outflow"])) if "outflow" in mapped and not _is_blank(row.get(mapped["outflow"])) else 0
        else:
            amount = _amount_to_fen(row.get(mapped["amount"]))
            direction = normalize_header(row.get(mapped["direction"]))
            if direction in {"收入", "流入", "贷", "贷方", "in", "inflow"}:
                inflow, outflow = amount, 0
            elif direction in {"支出", "流出", "借", "借方", "out", "outflow"}:
                inflow, outflow = 0, amount
            else:
                raise ParseError("TRANSACTION_DIRECTION_INVALID")
        if inflow < 0 or outflow < 0 or (inflow and outflow):
            raise ParseError("TRANSACTION_FLOW_INVALID")
        adjustment = _amount_to_fen(row.get(mapped["adjustment"])) if mapped.get("adjustment") and not _is_blank(row.get(mapped["adjustment"])) else 0
        internal = _bool(row.get(mapped["internal"])) if "internal" in mapped else False
        transfer_id = _optional_identifier(row, mapped["transfer_id"], "TRANSFER_ID_NON_TEXT") if "transfer_id" in mapped else None
        if internal and not transfer_id:
            raise ParseError("INTERNAL_TRANSFER_ID_MISSING")
        transactions.append(Transaction(
            business_date,
            company,
            bank,
            account,
            transaction_id,
            _parse_datetime(row.get(mapped["occurred_at"])) if "occurred_at" in mapped else None,
            inflow,
            outflow,
            adjustment,
            internal,
            transfer_id,
            source,
        ))
    return ParsedFacts(business_date, family, tuple(), tuple(transactions), source.source_version, parser_evidence)


def parse_attachment(
    *,
    family: str,
    filename: str,
    payload: bytes,
    source: SourceRef,
    mime: str | None = None,
) -> ParsedFacts:
    """Open and parse one structured source file into exactly one fact family.

    Images and scanned PDFs deliberately use :func:`parse_ocr_attachment`
    instead.  Keeping that route explicit prevents an unsupported binary from
    silently becoming a financial fact merely because its filename resembles a
    supported spreadsheet.
    """

    if family not in TRANSACTION_FAMILIES | {ACCOUNT_FAMILY}:
        raise ParseError("DOCUMENT_FAMILY_UNSUPPORTED")
    if not payload:
        raise ParseError("CORRUPT_ATTACHMENT")
    _validate_source(source, payload)
    rows, parser_evidence = _rows_from_bytes(filename, payload, mime)
    return _facts_from_rows(
        family=family,
        filename=filename,
        source=source,
        rows=rows,
        parser_evidence=parser_evidence,
    )


def parse_generic_structured_attachment(
    *,
    filename: str,
    payload: bytes,
    source: SourceRef,
    mime: str | None = None,
) -> ParsedFacts:
    """Resolve a ``资金明细`` spreadsheet through both frozen fact schemas.

    ``资金明细`` is an allowed source label, but it is not itself proof that
    a spreadsheet contains a transaction fact rather than an account snapshot.
    A structured file is therefore accepted only when exactly one complete
    schema opens.  This mirrors the existing generic-image gate without
    inventing a filename-based family or silently preferring one valid schema
    over another.

    Images and scanned PDFs remain on :func:`parse_ocr_attachment`, because
    that path retains the layout fingerprint required for the separate OCR
    calibration gate.
    """

    candidates: list[ParsedFacts] = []
    for candidate_family in (ACCOUNT_FAMILY, "资金明细"):
        try:
            candidates.append(parse_attachment(
                family=candidate_family,
                filename=filename,
                payload=payload,
                source=source,
                mime=mime,
            ))
        except ParseError as exc:
            # A generic label may be unresolved, but it must never downgrade a
            # broken raw-readback lineage into an ordinary schema miss.  The
            # caller needs the original integrity error to fail the batch.
            if str(exc).split(":", 1)[0] in _SOURCE_INTEGRITY_PARSE_CODES:
                raise
            continue
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        raise ParseError("GENERIC_SOURCE_SCHEMA_AMBIGUOUS")
    raise ParseError("GENERIC_SOURCE_SCHEMA_UNRESOLVED")
