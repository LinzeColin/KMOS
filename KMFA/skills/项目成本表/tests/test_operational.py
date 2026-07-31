from __future__ import annotations

import json
import zipfile
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

import project_cost_table.operational as operational
from project_cost_table.operational import (
    ProjectCostError,
    _attendance_assignments,
    _formal_cost_categories,
    _synthetic_ledger_book,
    _statement_buckets,
    _statement_rows,
    generate_outputs,
    governed_financial_analysis_revenue,
    governed_gross_margin,
    governed_contract_revenue,
    project_financial_analysis_components,
    labor_posted_component_reconciliation,
    labor_posted_reconciliation,
    ledger_book_metadata,
    parse_dws_approvals,
    parse_ledger_books,
    parse_ocr_paid_project_costs,
    parse_ocr_shared_information_fees,
    parse_status,
    project_margin_cost_completeness_blockers,
    project_level_residual_labor_allocate,
    qualify_cost_accruals,
    runtime_projection,
    sha256_bytes,
    split_project_work_units_by_entity,
    stable_json,
    subject_source_binding,
    verify_output,
    write_runtime_projection,
)

SUBJECT_BINDING = subject_source_binding()
EMPTY_SOURCE_BINDING_DIGEST = sha256_bytes(stable_json([]))


def _snapshot() -> dict:
    return {
        "schema_version": "kmfa.project_cost.snapshot.v2",
        "snapshot_id": "kmfa-pc-2099-synthetic",
        "generated_at": "2099-02-05T00:00:00+00:00",
        "skill_version": "0.0.6",
        "core_version": "0.2.0",
        "year": 2099,
        "as_of": "2099-02-05",
        "currency": "CNY",
        "subject_binding": SUBJECT_BINDING,
        "private_input_manifest_sha256": "b" * 64,
        "input_manifest_binding": {
            "kind": "PRIVATE_MANIFEST_SHA256",
            "digest": "b" * 64,
        },
        "selected_source_binding_digest": EMPTY_SOURCE_BINDING_DIGEST,
        "sources": [],
        "project_count": 2,
        "projects": [
            {
                "canonical_contract_id": "KMX20990101-001",
                "contract_base": "KMX20990101-001",
                "project_name": "合成项目甲",
                "customer": "合成客户甲",
                "contractor": "合成企业甲",
                "contract_amount_cents": 100_000,
                "job_posted_actual_cents": 12_345,
                "cost_accrued_cents": 7_655,
                "job_cost_incurred_cents": 20_000,
                "gross_margin_cost_basis_cents": 20_000,
                "effective_revenue_cents": 50_000,
                "gross_margin_status": "READY",
                "gl_recognized_cogs_cents": 8_000,
                "business_reported_direct_cost_cents": 5_000,
                "payment_system_paid_observed_cents": None,
                "status_business_components_cents": {},
                "own_work_units": 2,
                "external_work_units": None,
                "job_cost_coverage": "FULL_SELECTED_GL_PERIOD;POSTING_PRESENT",
                "accrual_coverage": "QUALIFIED_ACCRUAL_PRESENT",
            },
            {
                "canonical_contract_id": "KMX20990102-002",
                "contract_base": "KMX20990102-002",
                "project_name": "合成项目乙",
                "customer": "合成客户乙",
                "contractor": "合成企业甲",
                "contract_amount_cents": 200_000,
                "job_posted_actual_cents": 0,
                "cost_accrued_cents": 0,
                "job_cost_incurred_cents": 0,
                "gross_margin_cost_basis_cents": None,
                "effective_revenue_cents": None,
                "gross_margin_status": "BLOCKED_COST_COMPLETENESS",
                "gl_recognized_cogs_cents": 0,
                "business_reported_direct_cost_cents": None,
                "payment_system_paid_observed_cents": None,
                "status_business_components_cents": {},
                "job_cost_coverage": "FULL_SELECTED_GL_PERIOD;NO_QUALIFIED_EVENT",
                "accrual_coverage": "NO_QUALIFIED_ACCRUAL_EVENT",
            },
        ],
        "events": [
            {
                "project": "KMX20990101-001",
                "plane": "JOB_POSTED_ACTUAL",
                "category": "材料",
                "amount_cents": 12_345,
            },
            {
                "project": "KMX20990101-001",
                "plane": "COST_ACCRUED",
                "category": "自有人工-工资应计",
                "amount_cents": 7_655,
            },
        ],
        "reviews": [],
        "coverage": {
            "labor_wage_component_control_cents": 0,
            "labor_allocated_accrual_cents": 0,
            "labor_unallocated_cents": 0,
            "ledger_selected_book_count": 2,
            "qualified_accrual_event_count": 1,
            "labor_wage_component_event_count": 1,
        },
        "diagnostics": {"labor": {"already_posted_cents": 0}},
        "formula_contract": {
            "JOB_COST_INCURRED": "JOB_POSTED_ACTUAL + COST_ACCRUED",
            "planes_not_combined": [
                "GL_RECOGNIZED_COGS",
                "BUSINESS_REPORTED_DIRECT_COST",
                "PAYMENT_SYSTEM_PAID_OBSERVED",
            ],
            "automatic_fixed_labor_rate": False,
            "labor_allocation": (
                "auditable wage component × approved payroll days; largest remainder"
            ),
            "labor_unallocated_preserved": True,
            "automatic_management_fee_percent": False,
            "historical_reference_in_calculate": False,
        },
    }


def test_runtime_projection_keeps_formal_cost_and_observation_planes_separate():
    projection = runtime_projection(_snapshot())
    assert projection["schema_version"] == "kmfa.project_cost.current.v4"
    assert projection["项目数"] == 2
    first, second = projection["项目"]
    assert first["项目过账实际"] == "123.45"
    assert first["项目应计"] == "76.55"
    assert first["项目已发生成本"] == "200.00"
    assert first["项目成本"] == "200.00"
    assert first["主营成本已结转"] == "80.00"
    assert first["毛利"] == "300.00"
    assert first["毛利率"] == "60.00%"
    assert first["毛利率基点"] == 6000
    assert second["项目已发生成本"] == "0.00"
    assert second["项目成本"] is None
    assert "固定人工单价" in projection["禁止"]
    assert "自动合同额2%管理费" in projection["禁止"]
    assert "参考报表回填" in projection["禁止"]
    assert projection["计算状态"] == "PASS"
    assert projection["待确认"]["P1开放复核数"] == 0


def test_statement_template_uses_current_b_family_and_conserves_cents():
    snapshot = _snapshot()
    categories = _formal_cost_categories(snapshot)["KMX20990101-001"]
    buckets = _statement_buckets(categories)
    rows = {
        label: (amount, note)
        for label, amount, note in _statement_rows(snapshot["projects"][0], buckets)
    }
    assert rows["（一）原材料"][0] == 12_345
    assert rows["1.自有人员工资"][0] == 7_655
    assert rows["二、资金运用及各项支出"][0] == 20_000
    assert rows["（八） 分摊的管理费用（合同的2%）"][0] is None
    assert rows["三 利润"][0] == 30_000


def test_runtime_projection_rejects_a_one_cent_category_drift():
    snapshot = _snapshot()
    snapshot["events"][1]["amount_cents"] = 7_654
    with pytest.raises(ProjectCostError) as caught:
        runtime_projection(snapshot)
    assert caught.value.code == "RUNTIME_CATEGORY_CONSERVATION"


def test_runtime_projection_blocks_open_p1_instead_of_publishing_a_lower_bound():
    snapshot = _snapshot()
    snapshot["reviews"] = [
        {
            "severity": "P1",
            "type": "SYNTHETIC_IDENTITY_UNRESOLVED",
            "amount_cents": 999_999,
        },
        {
            "severity": "P2",
            "type": "SYNTHETIC_EXCLUDED",
            "amount_cents": 888_888,
        },
    ]
    with pytest.raises(ProjectCostError) as caught:
        runtime_projection(snapshot)
    assert caught.value.code == "P1_REVIEW_OPEN"


def test_margin_above_seventy_percent_blocks_instead_of_being_clamped():
    with pytest.raises(ProjectCostError) as caught:
        governed_gross_margin(
            revenue_cents=100_000,
            cost_cents=20_000,
            basis_status="READY",
        )
    assert caught.value.code == "GROSS_MARGIN_SANITY_GATE"


def test_incomplete_margin_basis_returns_null_without_backsolving_cost():
    result = governed_gross_margin(
        revenue_cents=100_000,
        cost_cents=1,
        basis_status="BLOCKED_COST_COMPLETENESS",
    )
    assert result == {
        "gross_profit_cents": None,
        "gross_margin_bps": None,
        "status": "BLOCKED_COST_COMPLETENESS",
    }


def test_project_scoped_p1_blocks_margin_cost_completeness():
    blockers = project_margin_cost_completeness_blockers(
        [
            {
                "severity": "P1",
                "type": "LABOR_PROJECT_LEVEL_PERIOD_SOURCE_MISSING",
                "project": "KMX20990101-001",
            },
            {
                "severity": "P2",
                "type": "LABOR_PROJECT_LEVEL_MULTI_ENTITY_SOURCE_SPLIT",
                "project": "KMX20990101-001",
            },
            {
                "severity": "P1",
                "type": "SOURCE_CANDIDATE_REJECTED",
            },
        ]
    )
    assert blockers == {
        "KMX20990101-001": (
            "LABOR_PROJECT_LEVEL_PERIOD_SOURCE_MISSING",
        )
    }


def test_closed_cost_basis_cannot_be_below_incurred_cost():
    snapshot = _snapshot()
    snapshot["projects"][0]["gross_margin_cost_basis_cents"] = 19_999
    with pytest.raises(ProjectCostError) as caught:
        runtime_projection(snapshot)
    assert caught.value.code == "GROSS_MARGIN_COST_BELOW_INCURRED"


def test_completed_settlement_is_distinct_from_invoice_and_original_contract():
    project = {
        "contract_amount_cents": 91_000,
        "construction_status_master": "已完工",
    }
    status = {
        "construction_status": "已完工",
        "status_contract_amount_cents": 91_000,
        "settlement_amount_cents": 157_300,
        "invoice_amount_cents": 145_200,
    }
    result = governed_contract_revenue(project, status)
    assert result == {
        "effective_revenue_cents": 157_300,
        "status": "READY_SETTLEMENT_REGISTER",
        "basis": "GOVERNED_SETTLEMENT_REGISTER",
    }


def test_financial_analysis_revenue_uses_approved_project_output_not_settlement():
    result = governed_financial_analysis_revenue(
        {
            "contract_amount_cents": 91_000,
            "construction_status_master": "已完工",
        },
        {
            "construction_status": "已完工",
            "status_contract_amount_cents": 91_000,
            "settlement_amount_cents": 157_300,
            "invoice_amount_cents": 145_200,
        },
    )
    assert result == {
        "effective_revenue_cents": 145_200,
        "status": "READY_APPROVED_INVOICED_PROJECT_OUTPUT",
        "basis": "APPROVED_INVOICED_PROJECT_OUTPUT",
        "original_contract_cents": 91_000,
        "settlement_cents": 157_300,
        "revenue_bridge_cents": 54_200,
    }


def test_financial_analysis_policy_is_exact_and_credits_direct_tax_once():
    result = project_financial_analysis_components(
        revenue_cents=11_300,
        invoice_events=[
            {
                "event_id": "invoice-1",
                "invoice_gross_cents": 11_300,
                "amount_cents": 1_300,
            }
        ],
        direct_project_tax_in_incurred_cents=500,
    )
    assert result["status"] == "READY"
    assert result["management_allocation_cents"] == 226
    assert result["tax_components_cents"] == {
        "output_vat": 1_300,
        "surcharge": 156,
        "income_tax": 226,
        "stamp_tax": 3,
    }
    assert result["tax_provision_cents"] == 1_685
    assert result["direct_project_tax_credit_cents"] == 500
    assert result["incremental_tax_provision_cents"] == 1_185
    assert result["analysis_increment_cents"] == 1_411


def test_financial_analysis_tax_blocks_invoice_control_drift_and_overcredit():
    drift = project_financial_analysis_components(
        revenue_cents=11_301,
        invoice_events=[
            {
                "event_id": "invoice-1",
                "invoice_gross_cents": 11_300,
                "amount_cents": 1_300,
            }
        ],
        direct_project_tax_in_incurred_cents=0,
    )
    assert drift["status"] == "BLOCKED_INVOICE_REVENUE_RECONCILIATION"
    overcredit = project_financial_analysis_components(
        revenue_cents=11_300,
        invoice_events=[
            {
                "event_id": "invoice-1",
                "invoice_gross_cents": 11_300,
                "amount_cents": 1_300,
            }
        ],
        direct_project_tax_in_incurred_cents=1_686,
    )
    assert overcredit["status"] == "BLOCKED_DIRECT_TAX_EXCEEDS_PROVISION"


def test_revenue_basis_blocks_when_master_and_status_contract_conflict():
    result = governed_contract_revenue(
        {
            "contract_amount_cents": 100_000,
            "construction_status_master": "已完工",
        },
        {
            "construction_status": "已完工",
            "status_contract_amount_cents": 110_000,
            "settlement_amount_cents": 120_000,
        },
    )
    assert result["effective_revenue_cents"] is None
    assert result["status"] == "BLOCKED_CONTRACT_REGISTER_CONFLICT"


def test_reporting_cohort_includes_target_year_and_operational_carryovers_only():
    projects = [
        {
            "contract_base": "KMX20990101-001",
            "year": 2099,
            "created_date": "2099-01-01",
            "construction_status_master": "",
        },
        {
            "contract_base": "KMX20981201-002",
            "year": 2098,
            "created_date": "2098-12-01",
            "construction_status_master": "",
        },
        {
            "contract_base": "KMX20981202-003",
            "year": 2098,
            "created_date": "2098-12-02",
            "construction_status_master": "",
        },
        {
            "contract_base": "KMX20981203-004",
            "year": 2098,
            "created_date": "2098-12-03",
            "construction_status_master": "",
        },
    ]
    status = {
        "KMX20981201-002": {
            "construction_status": "已完工",
            "completion_date": "2099-02-05",
        },
        "KMX20981202-003": {
            "construction_status": "施工中",
            "start_date": "2098-12-20",
        },
        "KMX20981203-004": {
            "construction_status": "已完工",
            "completion_date": "2098-12-28",
            "invoice_date": "2099-01-10",
            "cash_in_date": "2099-01-20",
        },
    }
    cohort = operational.reporting_project_cohort(
        projects,
        status,
        year=2099,
    )
    assert [row["contract_base"] for row in cohort] == [
        "KMX20981201-002",
        "KMX20981202-003",
        "KMX20990101-001",
    ]


def test_status_parser_preserves_project_cost_report_control_fields(
    tmp_path: Path,
):
    master_path = tmp_path / "红圈主合同.xlsx"
    status_path = tmp_path / "生产项目状态表.xlsx"
    operational._synthetic_master(master_path)
    operational._synthetic_status(status_path)
    projects, _metadata = operational.parse_master(master_path)
    rows, reviews, diagnostics = parse_status(status_path, projects, 2099)
    row = rows["KMX20990101-001"]
    assert row["project_cost_report_provided"] is True
    assert row["project_cost_report_deadline"] == "2099-03-01"
    assert row["commission_calculated"] is True
    assert row["status_contract_amount_cents"] == 100_000
    assert diagnostics["provided_project_cost_report_rows"] == 1
    assert diagnostics["provided_project_cost_report_with_direct_components"] == 1
    assert not [review for review in reviews if review["severity"] == "P1"]


def test_runtime_projection_blocks_any_p0_review():
    snapshot = _snapshot()
    snapshot["reviews"] = [
        {"severity": "P0", "type": "SYNTHETIC_FORMAL_SOURCE_UNAVAILABLE"}
    ]
    with pytest.raises(ProjectCostError) as caught:
        runtime_projection(snapshot)
    assert caught.value.code == "P0_REVIEW_OPEN"


def test_ledger_keeps_two_legal_identical_rows_in_one_source_view():
    payload = _synthetic_ledger_book(
        entity="合成企业甲",
        contract="KMX20990101-001",
        customer="合成客户甲",
        amount=Decimal("100.00"),
        account="5001001-生产成本_原材料",
        include_research_column=True,
        same_sheet_occurrences=2,
    )
    metadata = ledger_book_metadata("合成企业甲-明细账.xlsx", payload)
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "year": 2099,
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
        }
    ]
    events, reviews, diagnostics = parse_ledger_books(
        [{"metadata": metadata, "payload": payload}],
        projects,
        2099,
        "2099-02-28",
        {},
    )
    actual = [
        event for event in events if event["plane"] == "JOB_POSTED_ACTUAL"
    ]
    assert len(actual) == 2
    assert sum(event["amount_cents"] for event in actual) == 20_000
    assert diagnostics["semantic_duplicate_rows"] == 0
    assert not [row for row in reviews if row["severity"] == "P0"]


def test_ledger_short_and_full_entity_names_share_one_coverage_key():
    payloads = [
        _synthetic_ledger_book(
            entity=entity,
            contract="KMX20990101-001",
            customer="合成客户甲",
            amount=amount,
            account="5001001-生产成本_原材料",
            include_research_column=True,
        )
        for entity, amount in (
            ("合成企业甲", Decimal("100.00")),
            ("区域合成企业甲有限公司", Decimal("200.00")),
        )
    ]
    books = [
        {
            "metadata": ledger_book_metadata(
                "%s-明细账.xlsx" % entity,
                payload,
            ),
            "payload": payload,
        }
        for entity, payload in zip(
            ("合成企业甲", "区域合成企业甲有限公司"),
            payloads,
        )
    ]
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "year": 2099,
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
        }
    ]
    events, _reviews, diagnostics = parse_ledger_books(
        books,
        projects,
        2099,
        "2099-02-28",
        {},
    )
    actual = [
        event for event in events if event["plane"] == "JOB_POSTED_ACTUAL"
    ]
    assert len(actual) == 2
    assert {event["entity"] for event in actual} == {"合成企业甲"}
    assert {event["source_entity"] for event in actual} == {
        "合成企业甲",
        "区域合成企业甲有限公司",
    }
    assert list(diagnostics["period_ends_by_entity"]) == ["合成企业甲"]


def test_other_contract_review_is_not_multiplied_by_export_views():
    payloads = [
        _synthetic_ledger_book(
            entity="合成企业甲",
            contract="KMX20980101-999",
            customer="合成客户甲",
            amount=Decimal("100.00"),
            account="5001001-生产成本_原材料",
            include_research_column=include_research,
        )
        for include_research in (False, True)
    ]
    books = [
        {
            "metadata": ledger_book_metadata(
                "合成企业甲-明细账-%d.xlsx" % index,
                payload,
            ),
            "payload": payload,
        }
        for index, payload in enumerate(payloads, 1)
    ]
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "year": 2099,
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
        }
    ]
    events, reviews, diagnostics = parse_ledger_books(
        books,
        projects,
        2099,
        "2099-02-28",
        {},
    )
    excluded = [
        row
        for row in reviews
        if row["type"] == "LEDGER_OTHER_CONTRACT_EXCLUDED"
    ]
    assert events == []
    assert len(excluded) == 1
    assert excluded[0]["amount_cents"] == 10_000
    assert diagnostics["semantic_duplicate_rows"] == 1


def test_unallocated_5001_pool_is_preserved_and_blocks_publication():
    payloads = [
        _synthetic_ledger_book(
            entity="合成企业甲",
            contract="",
            customer="",
            amount=Decimal("100.00"),
            account="5001007-生产成本_劳务",
            include_research_column=include_research,
        )
        for include_research in (False, True)
    ]
    books = [
        {
            "metadata": ledger_book_metadata(
                "合成企业甲-未分配-%d.xlsx" % index,
                payload,
            ),
            "payload": payload,
        }
        for index, payload in enumerate(payloads, 1)
    ]
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "year": 2099,
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
        }
    ]
    events, reviews, diagnostics = parse_ledger_books(
        books,
        projects,
        2099,
        "2099-02-28",
        {},
    )
    assert len(events) == 1
    assert events[0]["plane"] == "UNALLOCATED_LEDGER_COST_POOL"
    assert events[0]["project"] is None
    assert diagnostics["unallocated_cost_rows"] == 1
    assert diagnostics["unallocated_cost_net_cents"] == 10_000
    assert diagnostics["unallocated_cost_absolute_cents"] == 10_000
    assert diagnostics["semantic_duplicate_rows"] == 1
    open_pool = [
        row
        for row in reviews
        if row["type"] == "UNALLOCATED_LEDGER_COST_POOL_OPEN"
    ]
    assert len(open_pool) == 1
    assert open_pool[0]["severity"] == "P1"


def test_ledger_exact_project_name_in_summary_recovers_project_identity():
    payload = _synthetic_ledger_book(
        entity="合成企业甲",
        contract="KMX999_不分项目",
        customer="",
        amount=Decimal("100.00"),
        account="5001003-生产成本_工资",
        include_research_column=True,
    )
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目成本",
            "year": 2099,
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
        }
    ]
    metadata = ledger_book_metadata("合成企业甲-明细账.xlsx", payload)
    events, reviews, diagnostics = parse_ledger_books(
        [{"metadata": metadata, "payload": payload}],
        projects,
        2099,
        "2099-02-28",
        {},
    )
    actual = [
        event for event in events if event["plane"] == "JOB_POSTED_ACTUAL"
    ]
    assert len(actual) == 1
    assert actual[0]["project"] == "KMX20990101-001"
    assert actual[0]["identity_reason"] == "LEDGER_NARRATIVE_EXACT_PROJECT_NAME"
    assert diagnostics["unallocated_cost_rows"] == 0
    assert not [
        row
        for row in reviews
        if row["type"] == "UNALLOCATED_LEDGER_COST_POOL_OPEN"
    ]


def test_payment_observation_never_creates_cost_without_occurrence_evidence():
    posted = [
        {
            "event_id": "posted-material",
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-01",
        }
    ]
    paid = [
        {
            "event_id": "paid-labor",
            "project": "KMX20990101-001",
            "category": "劳务/人工",
            "amount_cents": 10_000,
            "posting_date": "2099-02-05",
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(posted, [], paid)
    assert accruals == []
    assert reviews[0]["severity"] == "P2"
    assert (
        reviews[0]["type"]
        == "PAID_COST_OBSERVATION_EXCLUDED_FROM_ACCRUAL"
    )
    assert reviews[0]["posting_candidate_count"] == 0
    assert diagnostics["posting_link_required_count"] == 0


def test_paid_reimbursement_occurrence_creates_one_accrual_when_unrepresented():
    paid = [
        {
            "event_id": "paid-reimbursement",
            "project": "KMX20990101-001",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-05",
            "cost_occurrence_evidenced": True,
            "cost_occurrence_basis": "FINANCE_REGISTER_EMPLOYEE_REIMBURSEMENT",
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals([], [], paid)
    assert len(accruals) == 1
    assert accruals[0]["amount_cents"] == 10_000
    assert accruals[0]["plane"] == "COST_ACCRUED"
    assert [
        row
        for row in reviews
        if row["type"] == "PAID_COST_OCCURRENCE_PROMOTED_TO_ACCRUAL"
    ]
    assert diagnostics["paid_occurrence_accrual_count"] == 1


def test_paid_occurrence_does_not_duplicate_verified_approval():
    approved = [
        {
            "event_id": "approved-reimbursement",
            "approval_id": "APP-1",
            "project": "KMX20990101-001",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-05",
            "approval_authority_verified": True,
        }
    ]
    paid = [
        {
            "event_id": "paid-reimbursement",
            "project": "KMX20990101-001",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-06",
            "cost_occurrence_evidenced": True,
            "cost_occurrence_basis": "FINANCE_REGISTER_EMPLOYEE_REIMBURSEMENT",
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        [],
        approved,
        paid,
    )
    assert len(accruals) == 1
    assert accruals[0]["evidence_event_ids"] == ["approved-reimbursement"]
    assert [
        row
        for row in reviews
        if row["type"] == "PAID_COST_OCCURRENCE_ALREADY_REPRESENTED"
    ]
    assert diagnostics["paid_occurrence_existing_representation_count"] == 1


def test_payment_with_partial_posting_remains_observation_only():
    posted = [
        {
            "event_id": "posted-material-part",
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "category": "材料",
            "amount_cents": 4_000,
            "posting_date": "2099-02-01",
        }
    ]
    paid = [
        {
            "event_id": "approved-material-total",
            "project": "KMX20990101-001",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-05",
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        posted,
        [],
        paid,
    )
    assert accruals == []
    assert (
        reviews[0]["type"]
        == "PAID_COST_OBSERVATION_EXCLUDED_FROM_ACCRUAL"
    )
    assert reviews[0]["severity"] == "P2"
    assert reviews[0]["posting_candidate_count"] == 1
    assert diagnostics["posting_link_required_count"] == 0


def test_exact_payment_and_posting_still_do_not_change_cost_formula():
    posted = [
        {
            "event_id": "posted-rental",
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "category": "设备租赁",
            "amount_cents": 386_500,
            "posting_date": "2099-02-05",
        }
    ]
    paid = [
        {
            "event_id": "paid-rental",
            "project": "KMX20990101-001",
            "category": "设备租赁",
            "amount_cents": 386_500,
            "posting_date": "2099-02-05",
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(posted, [], paid)
    assert accruals == []
    assert reviews[0]["severity"] == "P2"
    assert (
        reviews[0]["type"]
        == "PAID_COST_OBSERVATION_EXCLUDED_FROM_ACCRUAL"
    )
    assert reviews[0]["posting_candidate_count"] == 1
    assert diagnostics["posting_link_required_count"] == 0


def test_approved_cost_is_not_double_accrued_over_unallocated_5001_candidate():
    ledger = [
        {
            "event_id": "unallocated-material",
            "project": None,
            "plane": "UNALLOCATED_LEDGER_COST_POOL",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-08",
        }
    ]
    approved = [
        {
            "event_id": "approved-material",
            "project": "KMX20990101-001",
            "plane": "DWS_APPROVED_COST",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-05",
            "approval_authority_verified": True,
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        ledger,
        approved,
        [],
    )
    assert accruals == []
    assert diagnostics["posting_link_required_count"] == 1
    assert [
        row
        for row in reviews
        if row["type"] == "APPROVED_COST_UNALLOCATED_POSTING_LINK_REQUIRED"
    ]


def test_unrelated_unallocated_same_amount_does_not_hide_approved_cost():
    ledger = [
        {
            "event_id": "unallocated-lodging",
            "project": None,
            "plane": "UNALLOCATED_LEDGER_COST_POOL",
            "category": "其他直接成本",
            "amount_cents": 100_000,
            "posting_date": "2099-02-28",
            "summary": "报销外协施工人员住宿费用",
        }
    ]
    approved = [
        {
            "event_id": "approved-customer-gift",
            "approval_id": "APPROVED-CUSTOMER-GIFT",
            "project": "KMX20990101-001",
            "category": "其他直接成本",
            "amount_cents": 100_000,
            "posting_date": "2099-01-28",
            "summary": "项目合同签订后为客户购买礼品",
            "approval_authority_verified": True,
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        ledger,
        approved,
        [],
    )
    assert len(accruals) == 1
    assert accruals[0]["amount_cents"] == 100_000
    assert diagnostics["approved_unallocated_posting_link_count"] == 0
    assert not [
        row
        for row in reviews
        if row["type"] == "APPROVED_COST_UNALLOCATED_POSTING_LINK_REQUIRED"
    ]


def test_dws_reaction_cannot_create_a_formal_accrual_without_authority():
    observed_reaction = [
        {
            "event_id": "dws-unverified",
            "project": "KMX20990101-001",
            "category": "材料",
            "amount_cents": 311_000,
            "posting_date": "2099-02-05",
            "approval_authority_verified": False,
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        [],
        observed_reaction,
        [],
    )
    assert accruals == []
    assert reviews[0]["severity"] == "P1"
    assert reviews[0]["type"] == "DWS_APPROVER_AUTHORITY_UNVERIFIED_EXCLUDED"
    assert diagnostics["dws_reaction_formal_amount_use"] is False


def test_explicit_approved_cost_accrues_even_when_payment_status_is_unpaid():
    approved = [
        {
            "event_id": "dws-approved",
            "approval_id": "SYN-APPROVED-1",
            "project": "KMX20990101-001",
            "category": "设备租赁",
            "amount_cents": 240_000,
            "posting_date": "2099-02-05",
            "approval_authority_verified": True,
            "payment_status": "未支付",
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals([], approved, [])
    assert reviews == []
    assert len(accruals) == 1
    assert accruals[0]["plane"] == "COST_ACCRUED"
    assert accruals[0]["amount_cents"] == 240_000
    assert diagnostics["dws_approved_cost_formal_count"] == 1


def test_paid_observation_plus_independent_approval_creates_exactly_one_accrual():
    approved = [
        {
            "event_id": "approved-tax",
            "approval_id": "SYN-APPROVED-TAX",
            "project": "KMX20990101-001",
            "category": "项目税费",
            "amount_cents": 12_300,
            "posting_date": "2099-02-05",
            "approval_authority_verified": True,
        }
    ]
    paid = [
        {
            "event_id": "paid-tax",
            "project": "KMX20990101-001",
            "category": "项目税费",
            "amount_cents": 12_300,
            "posting_date": "2099-02-06",
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        [],
        approved,
        paid,
    )
    assert len(accruals) == 1
    assert accruals[0]["amount_cents"] == 12_300
    assert sum(row["amount_cents"] for row in accruals) == 12_300
    assert [
        row
        for row in reviews
        if row["type"] == "PAID_COST_OBSERVATION_EXCLUDED_FROM_ACCRUAL"
    ]
    assert diagnostics["dws_reaction_paid_observation_link_count"] == 1


def test_approved_cost_posting_reconciliation_is_one_to_one():
    posted = [
        {
            "event_id": "posted-%d" % index,
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "category": "交通/差旅",
            "amount_cents": 41_650,
            "posting_date": "2099-02-05",
            "summary": "同一批差旅报销",
        }
        for index in range(3)
    ]
    approved = [
        {
            "event_id": "approved-%d" % index,
            "approval_id": "APPROVED-%d" % index,
            "project": "KMX20990101-001",
            "category": "交通/差旅",
            "amount_cents": 41_650,
            "posting_date": "2099-02-05",
            "summary": "同一批差旅报销",
            "approval_authority_verified": True,
        }
        for index in range(4)
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        posted,
        approved,
        [],
    )
    assert len(accruals) == 1
    assert accruals[0]["amount_cents"] == 41_650
    assert diagnostics["approved_posting_exact_match_count"] == 3
    assert len(
        [
            row
            for row in reviews
            if row["type"]
            == "APPROVED_COST_POSTING_MATCHED_ONE_TO_ONE"
        ]
    ) == 3


def test_nearby_same_category_different_amount_no_longer_blocks_accrual():
    posted = [
        {
            "event_id": "posted-partial",
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "category": "材料",
            "amount_cents": 4_000,
            "posting_date": "2099-02-01",
            "summary": "部分材料",
        }
    ]
    approved = [
        {
            "event_id": "approved-total",
            "approval_id": "APPROVED-TOTAL",
            "project": "KMX20990101-001",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-05",
            "summary": "另一笔材料采购",
            "approval_authority_verified": True,
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        posted,
        approved,
        [],
    )
    assert len(accruals) == 1
    assert accruals[0]["amount_cents"] == 10_000
    assert not [
        row
        for row in reviews
        if row["type"] == "APPROVED_COST_POSTING_LINK_REQUIRED"
    ]
    assert diagnostics["posting_link_required_count"] == 0


def test_semantic_gross_net_posting_match_suppresses_one_duplicate():
    posted = [
        {
            "event_id": "posted-net",
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "category": "交通/差旅",
            "amount_cents": 10_000,
            "posting_date": "2099-02-07",
            "summary": "设备维修项目差旅住宿费用入账",
        }
    ]
    approved = [
        {
            "event_id": "approved-gross",
            "approval_id": "APPROVED-GROSS",
            "project": "KMX20990101-001",
            "category": "交通/差旅",
            "amount_cents": 11_300,
            "posting_date": "2099-02-05",
            "summary": "设备维修项目差旅住宿费用报销",
            "approval_authority_verified": True,
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        posted,
        approved,
        [],
    )
    assert accruals == []
    assert diagnostics["approved_posting_fuzzy_match_count"] == 1
    matched = [
        row
        for row in reviews
        if row["type"] == "APPROVED_COST_POSTING_MATCHED_ONE_TO_ONE"
    ]
    assert matched[0]["match_kind"] == "SEMANTIC_GROSS_NET_13_PERCENT"


def test_ambiguous_fuzzy_posting_match_remains_blocked():
    posted = [
        {
            "event_id": "posted-net-%d" % index,
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "category": "交通/差旅",
            "amount_cents": 10_000,
            "posting_date": "2099-02-07",
            "summary": "设备维修项目差旅住宿费用入账",
        }
        for index in range(2)
    ]
    approved = [
        {
            "event_id": "approved-gross",
            "approval_id": "APPROVED-GROSS",
            "project": "KMX20990101-001",
            "category": "交通/差旅",
            "amount_cents": 11_300,
            "posting_date": "2099-02-05",
            "summary": "设备维修项目差旅住宿费用报销",
            "approval_authority_verified": True,
        }
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        posted,
        approved,
        [],
    )
    assert accruals == []
    assert diagnostics["approved_posting_ambiguous_count"] == 1
    assert reviews[0]["type"] == "APPROVED_COST_POSTING_MATCH_AMBIGUOUS"
    assert reviews[0]["severity"] == "P1"


def test_unallocated_posting_candidate_is_consumed_one_to_one():
    ledger = [
        {
            "event_id": "unallocated-material",
            "project": None,
            "plane": "UNALLOCATED_LEDGER_COST_POOL",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-08",
        }
    ]
    approved = [
        {
            "event_id": "approved-material-%d" % index,
            "approval_id": "APPROVED-MATERIAL-%d" % index,
            "project": "KMX20990101-001",
            "category": "材料",
            "amount_cents": 10_000,
            "posting_date": "2099-02-05",
            "approval_authority_verified": True,
        }
        for index in range(2)
    ]
    accruals, reviews, diagnostics = qualify_cost_accruals(
        ledger,
        approved,
        [],
    )
    assert len(accruals) == 1
    assert diagnostics["approved_unallocated_posting_link_count"] == 1
    assert len(
        [
            row
            for row in reviews
            if row["type"]
            == "APPROVED_COST_UNALLOCATED_POSTING_LINK_REQUIRED"
        ]
    ) == 1


def test_approved_cost_detail_uses_exact_project_and_excludes_deposit(
    tmp_path: Path,
):
    path = tmp_path / "项目成本统计_截至20990205.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "资金支出明细"
    sheet.append(
        [
            "费用明细编号",
            "创建时间",
            "审批状态",
            "费用项目",
            "费用说明",
            "金额",
            "业务类型",
            "成本清单明细",
            "关联主合同",
            "关联主合同(费用报销)",
            "任务单",
            "申请编号(费用报销)",
        ]
    )
    sheet.append(
        [
            "DETAIL-1",
            "2099-02-05 09:00",
            "已通过",
            "采购费",
            "项目材料",
            Decimal("123.45"),
            "项目报销",
            "",
            "",
            "",
            "KMX20990101-001--Z",
            "PARENT-1",
        ]
    )
    sheet.append(
        [
            "DETAIL-2",
            "2099-02-05 10:00",
            "已通过",
            "押金",
            "设备押金",
            Decimal("50.00"),
            "项目付款",
            "",
            "",
            "",
            "KMX20990101-001--Z",
            "PARENT-2",
        ]
    )
    sheet.append(
        [
            "DETAIL-3",
            "2099-02-05 11:00",
            "已通过",
            "税费",
            "项目印花税",
            Decimal("12.30"),
            "项目报销",
            "",
            "",
            "",
            "KMX20990101-001--Z",
            "PARENT-3",
        ]
    )
    workbook.save(path)
    workbook.close()
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    events, reviews, diagnostics = operational.parse_approved_cost_detail(
        path,
        projects,
        {},
        2099,
        "2099-02-28",
    )
    assert [(row["category"], row["amount_cents"]) for row in events] == [
        ("材料", 12_345),
        ("项目税费", 1_230),
    ]
    assert all(row["approval_authority_verified"] for row in events)
    assert diagnostics["mapped_rows"] == 2
    assert diagnostics["non_cost_rows"] == 1
    assert [
        row
        for row in reviews
        if row["type"] == "APPROVED_COST_DETAIL_NON_COST_EXCLUDED"
    ]


def test_approved_cost_summary_is_suppressed_on_exact_parent_reconciliation():
    dws = [
        {
            "event_id": "summary",
            "approval_id": "PARENT-1",
            "project": "KMX20990101-001",
            "amount_cents": 30_000,
            "approval_authority_verified": True,
        }
    ]
    detail = [
        {
            "event_id": "detail-1",
            "parent_approval_id": "PARENT-1",
            "project": "KMX20990101-001",
            "amount_cents": 10_000,
        },
        {
            "event_id": "detail-2",
            "parent_approval_id": "PARENT-1",
            "project": "KMX20990101-001",
            "amount_cents": 20_000,
        },
    ]
    combined, reviews, diagnostics = (
        operational.merge_approved_cost_sources(dws, detail)
    )
    assert [row["event_id"] for row in combined] == [
        "detail-1",
        "detail-2",
    ]
    assert diagnostics["exact_parent_match_count"] == 1
    assert diagnostics["conflicting_parent_count"] == 0
    assert reviews[0]["type"] == "APPROVED_COST_PARENT_EXACT_DUPLICATE"


def test_project_invoice_output_vat_is_observation_only_and_keeps_red_invoice(
    tmp_path: Path,
):
    path = tmp_path / "项目开票_导出文件_20990205.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "项目开票"
    sheet.append(
        [
            "发票号码",
            "开票状态",
            "开票日期",
            "本次开票含税金额(元)",
            "税率(%)",
            "开票单位",
            "审批状态",
            "合同编号(合同名称)",
        ]
    )
    rows = [
        [
            "INV-13",
            "已开票",
            "2099-02-01",
            Decimal("113.00"),
            "13%",
            "合成企业甲",
            "已通过",
            "KMX20990101-001",
        ],
        # An exact duplicate export row is one business fact.
        [
            "INV-13",
            "已开票",
            "2099-02-01",
            Decimal("113.00"),
            "13%",
            "合成企业甲",
            "已通过",
            "KMX20990101-001",
        ],
        [
            "INV-RED-6",
            "已开票",
            "2099-02-02",
            Decimal("-106.00"),
            "6%",
            "合成企业甲",
            "已通过",
            "KMX20990101-001",
        ],
        [
            "INV-ZERO",
            "已开票",
            "2099-02-03",
            Decimal("100.00"),
            "0%",
            "合成企业甲",
            "已通过",
            "KMX20990101-001",
        ],
        [
            "INV-NOT-ISSUED",
            "未开票",
            "2099-02-04",
            Decimal("113.00"),
            "13%",
            "合成企业甲",
            "已通过",
            "KMX20990101-001",
        ],
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    events, reviews, diagnostics = operational.parse_project_invoice_tax(
        path,
        projects,
        2099,
        "2099-02-28",
    )
    assert [row["type"] for row in reviews] == [
        "PROJECT_INVOICE_OUTPUT_VAT_EXCLUDED_FROM_COST"
    ]
    assert [
        (row["category"], row["amount_cents"])
        for row in events
    ] == [
        ("项目开票销项税额（观察）", 1_300),
        ("项目开票销项税额（观察）", -600),
    ]
    assert all(
        row["plane"] == "PROJECT_INVOICE_OUTPUT_VAT_OBSERVED"
        for row in events
    )
    assert diagnostics["mapped_rows"] == 2
    assert diagnostics["zero_rate_rows"] == 1
    assert diagnostics["formal_amount_use"] is False
    assert diagnostics["observation_amount_use"] is True
    assert diagnostics["output_vat_in_project_cost"] is False
    assert diagnostics["company_tax_allocation_used"] is False


def test_project_invoice_conflict_excludes_the_complete_invoice_project_group(
    tmp_path: Path,
):
    path = tmp_path / "项目开票_导出文件_冲突.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "发票号码",
            "开票状态",
            "开票日期",
            "本次开票含税金额(元)",
            "税率(%)",
            "开票单位",
            "审批状态",
            "合同编号(合同名称)",
        ]
    )
    for amount in (Decimal("113.00"), Decimal("226.00")):
        sheet.append(
            [
                "INV-CONFLICT",
                "已开票",
                "2099-02-01",
                amount,
                "13%",
                "合成企业甲",
                "已通过",
                "KMX20990101-001",
            ]
        )
    workbook.save(path)
    workbook.close()
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    events, reviews, diagnostics = operational.parse_project_invoice_tax(
        path,
        projects,
        2099,
        "2099-02-28",
    )
    assert events == []
    assert diagnostics["conflicting_invoice_project_count"] == 1
    assert reviews[0]["type"] == "PROJECT_INVOICE_ALLOCATION_CONFLICT"


def test_project_invoice_number_is_recovered_from_unambiguous_attachment_name(
    tmp_path: Path,
):
    path = tmp_path / "项目开票_导出文件_附件票号.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "发票号码",
            "开票状态",
            "开票日期",
            "本次开票含税金额(元)",
            "税率(%)",
            "开票单位",
            "审批状态",
            "合同编号(合同名称)",
            "发票回填",
            "发票回填",
        ]
    )
    sheet.append(
        [
            "待开票",
            "已开票",
            "2099-02-01",
            Decimal("113.00"),
            "13%",
            "合成企业甲",
            "已通过",
            "KMX20990101-001",
            "dzfp_12345678901234567890_合成.pdf",
            "https://example.invalid/download",
        ]
    )
    workbook.save(path)
    workbook.close()
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]

    events, reviews, diagnostics = operational.parse_project_invoice_tax(
        path,
        projects,
        2099,
        "2099-02-28",
    )

    assert [row["type"] for row in reviews] == [
        "PROJECT_INVOICE_OUTPUT_VAT_EXCLUDED_FROM_COST"
    ]
    assert len(events) == 1
    assert events[0]["amount_cents"] == 1_300
    assert events[0]["invoice_number_hash"] == sha256_bytes(
        b"12345678901234567890"
    )[:16]
    assert diagnostics["recovered_invoice_number_rows"] == 1
    assert diagnostics["conflicting_attachment_number_rows"] == 0


def test_project_invoice_attachment_conflict_fails_closed_and_is_scoped(
    tmp_path: Path,
):
    path = tmp_path / "项目开票_导出文件_附件票号冲突.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "发票号码",
            "开票状态",
            "开票日期",
            "本次开票含税金额(元)",
            "税率(%)",
            "开票单位",
            "审批状态",
            "合同编号(合同名称)",
            "发票回填",
            "发票回填",
        ]
    )
    sheet.append(
        [
            "待开票",
            "已开票",
            "2099-02-01",
            Decimal("113.00"),
            "13%",
            "合成企业甲",
            "已通过",
            "KMX20990101-001",
            "dzfp_12345678901234567890_合成.pdf",
            "dzfp_09876543210987654321_冲突.pdf",
        ]
    )
    workbook.save(path)
    workbook.close()
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]

    events, reviews, diagnostics = operational.parse_project_invoice_tax(
        path,
        projects,
        2099,
        "2099-02-28",
    )

    assert events == []
    assert reviews[0]["type"] == "PROJECT_INVOICE_ATTACHMENT_NUMBER_CONFLICT"
    assert reviews[0]["project"] == "KMX20990101-001"
    assert diagnostics["conflicting_attachment_number_rows"] == 1


def test_ocr_project_tax_register_requires_customer_gross_and_component_total(
    tmp_path: Path,
):
    path = tmp_path / "tax-register.jsonl"
    path.write_text(
        json.dumps(
            {
                "file": "synthetic-tax-register.png",
                "text": "\n".join(
                    (
                        "缴纳税款登记表-跨区域涉税事项报告预（2099年2月）",
                        "开票金额（含税）",
                        "预缴税金",
                        "合计",
                        "2099/2/5",
                        "合成客户甲有限公司",
                        "合成工程甲",
                        "113.00",
                        "100.00",
                        "13.00",
                        "2.00",
                        "0.10",
                        "0.06",
                        "0.04",
                        "0.20",
                        "2.40",
                    )
                ),
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    projects = [
        {
            "contract_base": "KMX20990101-001",
            "customer": "合成客户甲有限公司",
        },
        {
            "contract_base": "KMX20990102-002",
            "customer": "合成客户乙有限公司",
        },
    ]
    invoices = [
        {
            "event_id": "invoice-a",
            "project": "KMX20990101-001",
            "invoice_gross_cents": 11_300,
            "posting_date": "2099-02-05",
        },
        {
            "event_id": "invoice-b",
            "project": "KMX20990102-002",
            "invoice_gross_cents": 11_300,
            "posting_date": "2099-02-05",
        },
    ]

    events, reviews, sources, diagnostics = (
        operational.parse_ocr_project_tax_registers(
            path,
            (tmp_path,),
            projects,
            invoices,
            "2099-02-28",
        )
    )

    assert len(events) == 1
    assert events[0]["project"] == "KMX20990101-001"
    assert events[0]["amount_cents"] == 240
    assert events[0]["tax_component_count"] == 5
    assert events[0]["approval_authority_verified"] is True
    assert events[0]["plane"] == "PROJECT_TAX_REGISTER_EVIDENCE"
    assert [row["type"] for row in reviews] == [
        "OCR_PROJECT_TAX_REGISTER_EVIDENCE_ACCEPTED"
    ]
    assert sources[0]["source_slot"] == "ocr_project_tax_register"
    assert diagnostics["qualified_event_count"] == 1
    assert diagnostics["output_vat_substituted"] is False

    posted = [
        {
            "event_id": "posted-tax",
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "category": "项目税费",
            "amount_cents": 240,
            "posting_date": "2099-02-05",
            "summary": "项目预缴税金",
        }
    ]
    accruals, reconciliation_reviews, reconciliation = (
        qualify_cost_accruals(posted, events, [])
    )
    assert accruals == []
    assert reconciliation["approved_posting_exact_match_count"] == 1
    assert reconciliation_reviews[0]["type"] == (
        "APPROVED_COST_POSTING_MATCHED_ONE_TO_ONE"
    )


def test_ocr_project_tax_register_parses_dotted_thousands_total(
    tmp_path: Path,
):
    assert operational._tax_register_money_cents("2.179.82") == 217_982
    path = tmp_path / "tax-register.jsonl"
    path.write_text(
        json.dumps(
            {
                "file": "dotted-total.png",
                "text": "\n".join(
                    (
                        "缴纳税款登记表-跨区域涉税事项报告预",
                        "开票金额（含税）",
                        "预缴税金",
                        "合计",
                        "2099/2/6",
                        "合成客户甲有限公司",
                        "合成工程甲",
                        "99,000.00",
                        "90,825.69",
                        "8,174.31",
                        "1,816.51",
                        "90.83",
                        "54.50",
                        "36.33",
                        "181.65",
                        "2.179.82",
                    )
                ),
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    events, reviews, _sources, diagnostics = (
        operational.parse_ocr_project_tax_registers(
            path,
            (tmp_path,),
            [
                {
                    "contract_base": "KMX20990101-001",
                    "customer": "合成客户甲有限公司",
                }
            ],
            [
                {
                    "event_id": "invoice-a",
                    "project": "KMX20990101-001",
                    "invoice_gross_cents": 9_900_000,
                    "posting_date": "2099-02-06",
                }
            ],
            "2099-02-28",
        )
    )
    assert [event["amount_cents"] for event in events] == [217_982]
    assert reviews[-1]["type"] == (
        "OCR_PROJECT_TAX_REGISTER_EVIDENCE_ACCEPTED"
    )
    assert diagnostics["qualified_event_count"] == 1


def test_ocr_project_tax_register_conflict_excludes_business_identity(
    tmp_path: Path,
):
    path = tmp_path / "tax-register.jsonl"
    records = []
    for filename, components, total in (
        ("conflict-a.png", ("2.00", "0.10", "0.06", "0.04", "0.20"), "2.40"),
        ("conflict-b.png", ("2.00", "0.10", "0.06", "0.04", "0.30"), "2.50"),
    ):
        records.append(
            json.dumps(
                {
                    "file": filename,
                    "text": "\n".join(
                        (
                            "缴纳税款登记表-跨区域涉税事项报告预",
                            "开票金额（含税）",
                            "预缴税金",
                            "合计",
                            "2099/2/5",
                            "合成客户甲有限公司",
                            "合成工程甲",
                            "113.00",
                            "100.00",
                            "13.00",
                            *components,
                            total,
                        )
                    ),
                },
                ensure_ascii=False,
            )
        )
    path.write_text("\n".join(records) + "\n", encoding="utf-8")

    events, reviews, _sources, diagnostics = (
        operational.parse_ocr_project_tax_registers(
            path,
            (tmp_path,),
            [
                {
                    "contract_base": "KMX20990101-001",
                    "customer": "合成客户甲有限公司",
                }
            ],
            [
                {
                    "event_id": "invoice-a",
                    "project": "KMX20990101-001",
                    "invoice_gross_cents": 11_300,
                    "posting_date": "2099-02-05",
                }
            ],
            "2099-02-28",
        )
    )
    assert events == []
    assert reviews[0]["type"] == (
        "OCR_PROJECT_TAX_REGISTER_AMOUNT_CONFLICT"
    )
    assert diagnostics["business_amount_conflict_count"] == 1


def test_funding_plan_uses_description_not_customer_name_for_cost_category(
    tmp_path: Path,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "项目资金计划"
    sheet.append(
        [
            "申请编号",
            "关联主合同",
            "累计报销金额",
            "审批状态",
            "报销说明",
            "收款单位",
            "收款账户",
            "支付状态",
        ]
    )
    sheet.append(
        [
            "APP209902050001",
            "KMX20990101-001 某材料科技项目",
            Decimal("123.45"),
            "已通过",
            "坐轮渡回公司",
            "",
            "",
            "未支付",
        ]
    )
    payload = tmp_path / "资金计划.xlsx"
    workbook.save(payload)
    workbook.close()
    archive_path = tmp_path / "DWS_Outputs.zip"
    with zipfile.ZipFile(
        archive_path,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as archive:
        archive.write(
            payload,
            "DWS_Outputs/生产管理群/files/02/项目资金计划.xlsx",
        )
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "某材料科技项目",
            "customer": "某材料科技公司",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    events, reviews, _sources, diagnostics = parse_dws_approvals(
        (archive_path,),
        (tmp_path,),
        projects,
        {},
        2099,
        "2099-02-28",
    )
    assert reviews == []
    assert len(events) == 1
    assert events[0]["plane"] == "DWS_APPROVED_COST"
    assert events[0]["category"] == "交通/差旅"
    assert events[0]["amount_cents"] == 12_345
    assert diagnostics["funding_plan_mapped_row_count"] == 1


def _write_funding_snapshot(
    path: Path,
    *,
    approval_state: str,
    amount: str = "123.45",
    payment_status: str = "未支付",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "项目资金支出"
    sheet.append(
        [
            "申请编号",
            "关联主合同",
            "累计报销金额",
            "审批状态",
            "报销说明",
            "收款单位",
            "收款账户",
            "支付状态",
        ]
    )
    sheet.append(
        [
            "APP209902010001",
            "KMX20990101-001 合成项目甲",
            Decimal(amount),
            approval_state,
            "项目交通费",
            "",
            "",
            payment_status,
        ]
    )
    workbook.save(path)
    workbook.close()


def _funding_projects() -> list:
    return [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]


def test_standalone_funding_plan_uses_latest_explicit_approval_state(
    tmp_path: Path,
):
    pending = tmp_path / "项目资金支出-2099.02.05.xlsx"
    approved = tmp_path / "项目资金支出-2099.02.06.xlsx"
    _write_funding_snapshot(pending, approval_state="审批中")
    _write_funding_snapshot(approved, approval_state="已通过")

    events, reviews, sources, diagnostics = parse_dws_approvals(
        (),
        (tmp_path,),
        _funding_projects(),
        {},
        2099,
        "2099-02-28",
        (pending, approved),
    )

    assert reviews == []
    assert len(events) == 1
    assert events[0]["amount_cents"] == 12_345
    assert events[0]["approval_state_source"].startswith(
        "latest_snapshot:"
    )
    assert len([source for source in sources if source["selected"]]) == 2
    assert diagnostics["standalone_funding_plan_count"] == 2
    assert diagnostics["funding_plan_mapped_row_count"] == 1


def test_funding_snapshot_date_does_not_cross_directory_boundary():
    assert operational._funding_snapshot_label_date(
        "2026-05/2026.5.15付款计划.xlsx",
        "2026-05-14",
    ) == "2026-05-15"


def test_standalone_funding_plan_latest_revocation_excludes_prior_approval(
    tmp_path: Path,
):
    approved = tmp_path / "项目资金支出-2099.02.05.xlsx"
    revoked = tmp_path / "项目资金支出-2099.02.06.xlsx"
    _write_funding_snapshot(approved, approval_state="已通过")
    _write_funding_snapshot(revoked, approval_state="已撤回")

    events, reviews, _sources, diagnostics = parse_dws_approvals(
        (),
        (tmp_path,),
        _funding_projects(),
        {},
        2099,
        "2099-02-28",
        (approved, revoked),
    )

    assert reviews == []
    assert events == []
    state_resolution = diagnostics["funding_plan_state_resolution"]
    assert state_resolution["latest_nonapproved_application_count"] == 1


def test_standalone_funding_plan_latest_snapshot_conflict_fails_closed(
    tmp_path: Path,
):
    left = tmp_path / "项目资金支出-A-2099.02.07.xlsx"
    right = tmp_path / "项目资金支出-B-2099.02.07.xlsx"
    _write_funding_snapshot(left, approval_state="已通过")
    _write_funding_snapshot(
        right,
        approval_state="已通过",
        amount="223.45",
    )

    events, reviews, _sources, diagnostics = parse_dws_approvals(
        (),
        (tmp_path,),
        _funding_projects(),
        {},
        2099,
        "2099-02-28",
        (left, right),
    )

    assert events == []
    assert reviews[0]["type"] == (
        "APPROVED_COST_LATEST_SNAPSHOT_CONFLICT"
    )
    assert (
        diagnostics["funding_plan_state_resolution"][
            "latest_snapshot_conflict_count"
        ]
        == 1
    )


def test_standalone_funding_plan_explicit_zero_needs_no_payment(
    tmp_path: Path,
):
    zero = tmp_path / "项目资金支出-2099.02.07.xlsx"
    _write_funding_snapshot(
        zero,
        approval_state="已通过",
        amount="0",
        payment_status="无需支付",
    )

    events, reviews, _sources, diagnostics = parse_dws_approvals(
        (),
        (tmp_path,),
        _funding_projects(),
        {},
        2099,
        "2099-02-28",
        (zero,),
    )

    assert events == []
    assert reviews[0]["severity"] == "P2"
    assert reviews[0]["type"] == "DWS_APPROVED_ZERO_COST_EXCLUDED"
    assert diagnostics["funding_plan_mapped_row_count"] == 0


def test_official_project_attendance_maps_exact_employee_project_days(
    tmp_path: Path,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2月项目差旅费"
    sheet.append(["2099年02月份各项目差旅食宿费"])
    sheet.append(
        [
            "序号",
            "任务单号",
            "项目名称",
            "姓名",
            "费用类别",
            "1",
            "2",
            "3",
        ]
    )
    sheet.append(
        [
            1,
            "KMX20990101-001",
            "合成项目甲",
            "合成人员甲",
            "餐费",
            35,
            0,
            35,
        ]
    )
    sheet.append([None, None, None, None, "住宿", 0, 35, 0])
    attendance_path = tmp_path / "生产部考勤表2月份-核对后.xlsx"
    workbook.save(attendance_path)
    workbook.close()
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    assignments, sources, diagnostics, reviews = _attendance_assignments(
        (tmp_path,),
        "2099-02",
        projects,
        {},
        (tmp_path,),
    )
    assert reviews == []
    assert assignments["合成人员甲"]["KMX20990101-001"] == {
        "20990201",
        "20990202",
        "20990203",
    }
    assert len(sources) == 1
    assert (
        diagnostics["official_project_attendance"][
            "mapped_employee_project_days"
        ]
        == 3
    )


def test_current_contract_header_and_short_cost_categories_map_project_days(
    tmp_path: Path,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "6月项目差旅费"
    sheet.append(["2099年06月份各项目自有工人工时（截止至6.30日）"])
    sheet.append(
        [
            "序号",
            "项目名称",
            "合同号",
            "姓名",
            "费用类别",
            "1",
            "2",
            "3",
        ]
    )
    sheet.append(
        [
            1,
            "合成项目甲",
            "KMX20990101-001",
            "合成人员甲",
            "生",
            35,
            35,
            0,
        ]
    )
    sheet.append([None, None, None, None, "住", 0, 35, 35])
    attendance_path = tmp_path / "生产部考勤表6月份-核对后.xlsx"
    workbook.save(attendance_path)
    workbook.close()
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    assignments, _sources, diagnostics, reviews = (
        _attendance_assignments(
            (tmp_path,),
            "2099-06",
            projects,
            {},
            (tmp_path,),
        )
    )
    assert reviews == []
    assert assignments["合成人员甲"]["KMX20990101-001"] == {
        "20990601",
        "20990602",
        "20990603",
    }
    assert (
        diagnostics["official_project_attendance"][
            "mapped_employee_project_days"
        ]
        == 3
    )


def test_hash_bound_attendance_alias_is_selected_by_workbook_period(
    tmp_path: Path,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "6月项目差旅费"
    sheet.append(["2099年06月份各项目自有工人工时（截止至6.30日）"])
    sheet.append(
        [
            "序号",
            "项目名称",
            "合同号",
            "姓名",
            "费用类别",
            "1",
            "2",
        ]
    )
    sheet.append(
        [
            1,
            "合成项目甲",
            "KMX20990101-001",
            "合成人员甲",
            "生",
            35,
            35,
        ]
    )
    attendance_path = tmp_path / "209906.xlsx"
    workbook.save(attendance_path)
    workbook.close()
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    assignments, sources, diagnostics, reviews = _attendance_assignments(
        (tmp_path,),
        "2099-06",
        projects,
        {},
        (tmp_path,),
    )
    assert reviews == []
    assert assignments["合成人员甲"]["KMX20990101-001"] == {
        "20990601",
        "20990602",
    }
    assert len(sources) == 1
    assert (
        diagnostics["official_project_attendance"][
            "mapped_employee_project_days"
        ]
        == 2
    )


def _write_synthetic_payroll(
    path: Path,
    *,
    department: str = "综合部",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "分部门"
    sheet.append(
        [
            "序号",
            "公司",
            "姓名",
            "部门",
            "应计工资小计",
            "实出勤天-厂外",
            "实出勤天-厂内",
            "实际部门",
        ]
    )
    sheet.append(
        [
            1,
            "合成区域企业甲",
            "合成人员甲",
            department,
            Decimal("300.00"),
            3,
            0,
            department,
        ]
    )
    workbook.save(path)
    workbook.close()


def _write_synthetic_employer_burden(
    path: Path,
    *,
    row_period: int = 9902,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "合成区域企业甲209902"
    sheet.append(
        [
            "序号",
            "部门",
            "月份",
            "姓名",
            "单位应缴",
            None,
            "个人应缴",
        ]
    )
    sheet.append([None, None, None, None, "养老", "合计", "合计"])
    sheet.append(
        [
            1,
            "综合部",
            row_period,
            "合成人员甲",
            Decimal("40.00"),
            Decimal("60.00"),
            Decimal("10.00"),
        ]
    )
    workbook.save(path)
    workbook.close()


def _write_synthetic_project_attendance(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2月项目差旅费"
    sheet.append(["2099年02月份各项目差旅食宿费"])
    sheet.append(
        [
            "序号",
            "任务单号",
            "项目名称",
            "姓名",
            "费用类别",
            "1",
            "2",
            "3",
        ]
    )
    sheet.append(
        [
            1,
            "KMX20990101-001",
            "合成项目甲",
            "合成人员甲",
            "餐费",
            35,
            35,
            35,
        ]
    )
    workbook.save(path)
    workbook.close()


def _synthetic_labor_inputs(tmp_path: Path) -> tuple[Path, Path, Path, list[dict]]:
    payroll = tmp_path / "209902工资.xlsx"
    burden = tmp_path / "209902单位社保医保.xlsx"
    attendance = tmp_path / "生产部考勤表2月份-核对后.xlsx"
    _write_synthetic_payroll(payroll)
    _write_synthetic_employer_burden(burden)
    _write_synthetic_project_attendance(attendance)
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    return payroll, burden, attendance, projects


def test_non_production_employee_with_approved_project_days_gets_full_labor_cost(
    tmp_path: Path,
):
    payroll, burden, _attendance, projects = _synthetic_labor_inputs(tmp_path)
    events, reviews, _sources, diagnostics = (
        operational.parse_payroll_and_attendance(
            (payroll,),
            (tmp_path,),
            (burden,),
            projects,
            {},
            (),
            (tmp_path,),
            year=2099,
            as_of="2099-02-28",
        )
    )
    assert not [row for row in reviews if row["severity"] in ("P0", "P1")]
    by_category = {
        event["category"]: event["amount_cents"] for event in events
    }
    assert by_category == {
        "自有人工-工资应计": 30_000,
        "自有人工-雇主社保医保应计": 6_000,
    }
    assert diagnostics["fully_loaded_labor_control_cents"] == 36_000
    assert diagnostics["allocated_accrual_cents"] == 36_000
    assert diagnostics["unallocated_cents"] == 0
    assert diagnostics["conservation_delta_cents"] == 0


def test_missing_employer_burden_blocks_instead_of_guessing(
    tmp_path: Path,
):
    payroll, _burden, _attendance, projects = _synthetic_labor_inputs(
        tmp_path
    )
    events, reviews, _sources, diagnostics = (
        operational.parse_payroll_and_attendance(
            (payroll,),
            (tmp_path,),
            (),
            projects,
            {},
            (),
            (tmp_path,),
            year=2099,
            as_of="2099-02-28",
        )
    )
    assert [event["category"] for event in events] == [
        "自有人工-工资应计"
    ]
    assert any(
        row["type"] == "LABOR_EMPLOYER_BURDEN_MISSING"
        and row["severity"] == "P1"
        for row in reviews
    )
    assert diagnostics["employer_burden_control_cents"] == 0
    assert diagnostics["personal_deductions_used"] is False


def test_employer_burden_row_month_override_requires_file_and_sheet_agreement(
    tmp_path: Path,
):
    burden = tmp_path / "209902单位社保医保.xlsx"
    _write_synthetic_employer_burden(burden, row_period=9901)
    records, _sources, reviews, diagnostics = (
        operational.parse_employer_burden_workbooks(
            (burden,),
            (tmp_path,),
            year=2099,
            as_of="2099-02-28",
        )
    )
    assert len(records["2099-02"]) == 1
    assert diagnostics["employer_burden_control_cents"] == 6_000
    assert any(
        row["type"] == "EMPLOYER_BURDEN_ROW_PERIOD_OVERRIDDEN"
        and row["severity"] == "P2"
        for row in reviews
    )
    assert not [row for row in reviews if row["severity"] in ("P0", "P1")]


def test_ledger_excludes_postings_after_as_of():
    payload = _synthetic_ledger_book(
        entity="合成企业甲",
        contract="KMX20990101-001",
        customer="合成客户甲",
        amount=Decimal("100.00"),
        account="5001001-生产成本_原材料",
        include_research_column=True,
    )
    metadata = ledger_book_metadata("合成企业甲-明细账.xlsx", payload)
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "year": 2099,
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
        }
    ]
    events, reviews, _diagnostics = parse_ledger_books(
        [{"metadata": metadata, "payload": payload}],
        projects,
        2099,
        "2099-01-15",
        {},
    )
    assert events == []
    future = [
        row
        for row in reviews
        if row["type"] == "LEDGER_POSTING_AFTER_AS_OF_EXCLUDED"
    ]
    assert len(future) == 1
    assert future[0]["amount_cents"] == 10_000


def test_carryover_project_keeps_inception_to_date_prior_year_cost():
    payload = _synthetic_ledger_book(
        entity="合成企业甲",
        contract="KMX20981201-001",
        customer="合成客户甲",
        amount=Decimal("100.00"),
        account="5001001-生产成本_原材料",
        include_research_column=True,
        posting_date="2098-12-15",
    )
    metadata = ledger_book_metadata("合成企业甲-明细账.xlsx", payload)
    events, reviews, diagnostics = parse_ledger_books(
        [{"metadata": metadata, "payload": payload}],
        [
            {
                "canonical_contract_id": "KMX20981201-001",
                "contract_base": "KMX20981201-001",
                "year": 2098,
                "customer": "合成客户甲",
                "contractor": "合成企业甲",
                "created_date": "2098-12-01",
            }
        ],
        2099,
        "2099-02-28",
        {
            "KMX20981201-001": {
                "start_date": "2098-12-10",
                "completion_date": "2099-01-20",
            }
        },
    )
    assert not [row for row in reviews if row["severity"] in ("P0", "P1")]
    assert len(events) == 1
    assert events[0]["amount_cents"] == 10_000
    assert diagnostics["cohort_scan_start"] == "2098-10-17"


def test_ocr_preserves_identical_legal_payments_on_distinct_pages(
    tmp_path: Path,
):
    text = "合成客户甲 合成项目甲 材料款\n2月5日\n项目成本\n100.00"
    path = tmp_path / "ocr.jsonl"
    path.write_text(
        "\n".join(
            json.dumps({"file": filename, "text": text}, ensure_ascii=False)
            for filename in ("page-1.png", "page-2.png", "page-1.png")
        )
        + "\n",
        encoding="utf-8",
    )
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    events, reviews, _sources, _diagnostics = parse_ocr_paid_project_costs(
        path,
        (tmp_path,),
        projects,
        {},
        2099,
        "2099-02-28",
    )
    assert len(events) == 2
    assert sum(event["amount_cents"] for event in events) == 20_000
    duplicates = [
        row
        for row in reviews
        if row["type"]
        == "OCR_SAME_PHYSICAL_OCCURRENCE_DUPLICATE_EXCLUDED"
    ]
    assert len(duplicates) == 1


def test_ocr_marks_reimbursement_and_recovers_explicit_wage_payment_history(
    tmp_path: Path,
):
    text = "\n".join(
        (
            "2月5日",
            (
                "生产部用款合成项目甲测试外协人员工资（合同金额1万，"
                "1月10日已经支付200元，1月20日支"
            ),
            "2月5日",
            "项目成本",
            "100.00",
            "合成银行",
            "付0.03万，本次支付100元）",
        )
    )
    path = tmp_path / "ocr.jsonl"
    path.write_text(
        json.dumps(
            {"file": "finance-register.png", "text": text},
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲测试",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    events, reviews, _sources, diagnostics = parse_ocr_paid_project_costs(
        path,
        (tmp_path,),
        projects,
        {},
        2099,
        "2099-02-28",
    )
    assert reviews == []
    assert sorted(event["amount_cents"] for event in events) == [
        10_000,
        20_000,
        30_000,
    ]
    assert all(event["cost_occurrence_evidenced"] for event in events)
    assert diagnostics["embedded_wage_history_count"] == 2
    assert diagnostics["occurrence_evidenced_count"] == 3


def test_ocr_page_anchor_resolves_same_customer_rows_without_name_hardcoding(
    tmp_path: Path,
):
    text = "\n".join(
        (
            "5月22日",
            "合成客户甲 5号煅烧炉现场车削 外协人员工资",
            "项目成本",
            "7480.00",
            "合成客户甲 焊丝",
            "项目成本",
            "1235.00",
        )
    )
    path = tmp_path / "ocr.jsonl"
    path.write_text(
        json.dumps(
            {"file": "finance-register.png", "text": text},
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成客户甲7号锅炉省煤器",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        },
        {
            "canonical_contract_id": "KMX20990501-002",
            "contract_base": "KMX20990501-002",
            "project_name": "合成客户甲5号煅烧炉现场车削",
            "customer": "合成客户甲",
            "contractor": "合成企业甲",
            "created_date": "2099-05-01",
            "year": 2099,
        },
    ]
    status = {
        "KMX20990101-001": {
            "start_date": "2099-03-01",
            "completion_date": "2099-06-30",
            "project_type": "自有人员",
        },
        "KMX20990501-002": {
            "start_date": "2099-05-11",
            "completion_date": "2099-05-21",
            "project_type": "劳务外协",
            "external_work_units": 34,
        },
    }
    events, reviews, _sources, diagnostics = parse_ocr_paid_project_costs(
        path,
        (tmp_path,),
        projects,
        status,
        2099,
        "2099-05-31",
    )
    assert [(row["amount_cents"], row["project"]) for row in events] == [
        (748_000, "KMX20990501-002"),
        (123_500, "KMX20990501-002"),
    ]
    assert not [
        row
        for row in reviews
        if row["type"] == "OCR_PAID_OUTSIDE_FORMULA_EXCLUDED"
    ]
    assert diagnostics["page_anchor_inherited_count"] == 1


def test_shared_information_fee_uses_governed_invoice_weight_and_conserves_cents(
    tmp_path: Path,
):
    path = tmp_path / "ocr.jsonl"
    path.write_text(
        json.dumps(
            {
                "file": "transfer.png",
                "text": "\n".join(
                    (
                        "交易详情",
                        "-¥100.00",
                        "交易时间",
                        "2099-03-20 16:20:06",
                        "附言",
                        "合成客户甲信息费",
                    )
                ),
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲有限公司",
            "created_date": "2099-01-01",
            "year": 2099,
        },
        {
            "canonical_contract_id": "KMX20990102-002",
            "contract_base": "KMX20990102-002",
            "project_name": "合成项目乙",
            "customer": "合成客户甲有限公司",
            "created_date": "2099-01-02",
            "year": 2099,
        },
    ]
    status = {
        "KMX20990101-001": {"invoice_amount_cents": 30_000},
        "KMX20990102-002": {"invoice_amount_cents": 10_000},
    }
    events, reviews, _sources, diagnostics = (
        parse_ocr_shared_information_fees(
            path,
            (tmp_path,),
            projects,
            status,
            2099,
            "2099-03-31",
        )
    )
    assert reviews == []
    assert [
        (row["project"], row["amount_cents"])
        for row in events
    ] == [
        ("KMX20990101-001", 7_500),
        ("KMX20990102-002", 2_500),
    ]
    assert sum(row["amount_cents"] for row in events) == 10_000
    assert all(
        row["cost_occurrence_evidenced"] is True
        and row["category"] == "信息费"
        for row in events
    )
    assert diagnostics["allocated_transaction_count"] == 1


def test_one_cent_labor_posting_cannot_suppress_full_payroll_allocation():
    matched, residual = labor_posted_reconciliation(10_000, 1)
    assert matched == 1
    assert residual == 9_999
    assert matched + residual == 10_000


def test_project_level_residual_labor_uses_actual_control_quotient():
    result = project_level_residual_labor_allocate(
        wage_pool_cents=100_000,
        employer_burden_pool_cents=10_000,
        approved_unallocated_days=Decimal("100"),
        project_work_units={
            "project-a": Decimal("10"),
            "project-b": Decimal("20"),
        },
    )

    assert result["wage_by_project"] == {
        "project-a": 10_000,
        "project-b": 20_000,
    }
    assert result["employer_burden_by_project"] == {
        "project-a": 1_000,
        "project-b": 2_000,
    }
    assert result["wage_unallocated_cents"] == 70_000
    assert result["employer_burden_unallocated_cents"] == 7_000
    assert result["remaining_unallocated_days"] == "70"
    assert result["wage_control_delta_cents"] == 0
    assert result["employer_burden_control_delta_cents"] == 0


def test_project_level_residual_labor_rejects_time_overflow():
    with pytest.raises(ProjectCostError) as error:
        project_level_residual_labor_allocate(
            wage_pool_cents=100_000,
            employer_burden_pool_cents=10_000,
            approved_unallocated_days=Decimal("10"),
            project_work_units={"project-a": Decimal("11")},
        )

    assert error.value.code == "LABOR_PROJECT_LEVEL_TIME_OVERFLOW"


def test_project_work_units_split_across_observed_employment_entities():
    result = split_project_work_units_by_entity(
        Decimal("19"),
        {
            "entity-a": Decimal("1"),
            "entity-b": Decimal("1"),
        },
    )

    assert result == {
        "entity-a": Decimal("9.5"),
        "entity-b": Decimal("9.5"),
    }
    assert sum(result.values(), Decimal(0)) == Decimal("19")


def test_project_work_units_entity_split_conserves_repeating_ratio():
    result = split_project_work_units_by_entity(
        Decimal("20"),
        {
            "entity-a": Decimal("1"),
            "entity-b": Decimal("2"),
        },
    )

    assert set(result) == {"entity-a", "entity-b"}
    assert all(value > 0 for value in result.values())
    assert sum(result.values(), Decimal(0)) == Decimal("20")


def test_payroll_parser_allocates_actual_residual_pool_by_status_work_units(
    tmp_path: Path,
):
    payroll = tmp_path / "209902工资.xlsx"
    burden = tmp_path / "209902单位社保医保.xlsx"
    attendance_root = tmp_path / "attendance"
    attendance_root.mkdir()
    _write_synthetic_payroll(payroll, department="生产部")
    _write_synthetic_employer_burden(burden)
    projects = [
        {
            "canonical_contract_id": "KMX20990101-001",
            "contract_base": "KMX20990101-001",
            "project_name": "合成项目甲",
            "customer": "合成客户甲",
            "contractor": "合成区域企业甲",
            "created_date": "2099-01-01",
            "year": 2099,
        }
    ]
    status_map = {
        "KMX20990101-001": {
            "own_work_units": Decimal("1"),
            "completion_date": "2099-02-20",
        }
    }

    events, reviews, _sources, diagnostics = (
        operational.parse_payroll_and_attendance(
            (payroll,),
            (attendance_root,),
            (burden,),
            projects,
            status_map,
            (),
            (tmp_path,),
            year=2099,
            as_of="2099-02-28",
        )
    )

    assert {
        event["category"]: event["amount_cents"]
        for event in events
    } == {
        "自有人工-工资项目级分配": 10_000,
        "自有人工-单位负担项目级分配": 2_000,
    }
    assert not [
        row
        for row in reviews
        if row["severity"] in ("P0", "P1")
    ]
    assert diagnostics["fully_loaded_labor_control_cents"] == 36_000
    assert diagnostics["allocated_accrual_cents"] == 12_000
    assert diagnostics["unallocated_cents"] == 24_000
    assert diagnostics["conservation_delta_cents"] == 0
    assert diagnostics["fixed_daily_rate_used"] is False


def test_explicit_gl_wage_and_social_allocation_prevents_double_count():
    result = labor_posted_component_reconciliation(
        allocated_wage_cents=30_000,
        allocated_burden_cents=6_000,
        direct_wage_posted_cents=10_000,
        combined_wage_burden_posted_cents=26_000,
    )
    assert result == {
        "direct_wage_matched_cents": 10_000,
        "combined_matched_cents": 26_000,
        "matched_cents": 36_000,
        "wage_accrual_cents": 0,
        "employer_burden_accrual_cents": 0,
        "direct_wage_posted_excess_cents": 0,
        "combined_posted_excess_cents": 0,
    }


def test_combined_gl_labor_split_conserves_every_cent():
    result = labor_posted_component_reconciliation(
        allocated_wage_cents=10_000,
        allocated_burden_cents=2_000,
        direct_wage_posted_cents=1,
        combined_wage_burden_posted_cents=1,
    )
    assert result["matched_cents"] == 2
    assert (
        result["wage_accrual_cents"]
        + result["employer_burden_accrual_cents"]
        + result["matched_cents"]
        == 12_000
    )


def test_payroll_parser_matches_explicit_5001006_labor_by_legal_entity(
    tmp_path: Path,
):
    payroll, burden, _attendance, projects = _synthetic_labor_inputs(
        tmp_path
    )
    ledger_events = (
        {
            "event_id": "posted-wage",
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "account_code": "5001003",
            "amount_cents": 10_000,
            "posting_date": "2099-02-28",
            "entity": "合成区域企业甲有限公司",
            "summary": "项目工资",
        },
        {
            "event_id": "posted-wage-burden",
            "project": "KMX20990101-001",
            "plane": "JOB_POSTED_ACTUAL",
            "account_code": "5001006",
            "amount_cents": 26_000,
            "posting_date": "2099-02-28",
            "entity": "合成区域企业甲有限公司",
            "summary": "项目分摊工资和社保费用",
        },
    )
    events, reviews, _sources, diagnostics = (
        operational.parse_payroll_and_attendance(
            (payroll,),
            (tmp_path,),
            (burden,),
            projects,
            {},
            ledger_events,
            (tmp_path,),
            year=2099,
            as_of="2099-02-28",
        )
    )
    assert events == []
    assert not [
        row
        for row in reviews
        if row["severity"] in ("P0", "P1")
    ]
    assert diagnostics["fully_loaded_labor_control_cents"] == 36_000
    assert diagnostics["already_posted_cents"] == 36_000
    assert diagnostics["allocated_accrual_cents"] == 0
    assert diagnostics["conservation_delta_cents"] == 0


def test_runtime_projection_is_an_atomic_private_file(tmp_path: Path):
    destination = tmp_path / "runtime" / "recent_completed.json"
    binding = {
        "filename": "current_project_cost_synthetic_deadbeef.xlsx",
        "sha256": "a" * 64,
        "size_bytes": 123,
    }
    projection = write_runtime_projection(
        destination,
        _snapshot(),
        binding,
    )
    assert json.loads(destination.read_text(encoding="utf-8")) == projection
    assert projection["封印工作簿"]["SHA256"] == "a" * 64
    assert projection["封印工作簿"]["快照ID"] == _snapshot()["snapshot_id"]
    assert projection["封印来源"]["输入清单类型"] == "PRIVATE_MANIFEST_SHA256"
    assert projection["封印来源"]["输入清单SHA256"] == "b" * 64
    assert destination.stat().st_mode & 0o777 == 0o600
    assert not list(destination.parent.glob(".*.tmp-*"))


def test_verify_output_rejects_source_or_private_manifest_drift(
    tmp_path: Path,
    monkeypatch,
):
    derived_snapshot = _snapshot()
    derived_snapshot["private_input_manifest_sha256"] = None
    derived_snapshot["input_manifest_binding"] = {
        "kind": "SELECTED_SOURCE_DERIVED_SHA256",
        "digest": EMPTY_SOURCE_BINDING_DIGEST,
    }
    derived_output = tmp_path / "derived-sealed"
    generate_outputs(derived_output, derived_snapshot)
    derived_verification = verify_output(derived_output)
    assert derived_verification["input_manifest_binding"] == {
        "kind": "SELECTED_SOURCE_DERIVED_SHA256",
        "digest": EMPTY_SOURCE_BINDING_DIGEST,
    }
    derived_runtime = runtime_projection(derived_snapshot)
    assert derived_runtime["封印来源"]["输入清单SHA256"] == (
        EMPTY_SOURCE_BINDING_DIGEST
    )

    output = tmp_path / "sealed"
    generate_outputs(output, _snapshot())
    assert verify_output(
        output,
        expected_private_input_manifest_sha256="b" * 64,
    )["status"] == "PASS"

    with pytest.raises(ProjectCostError) as private_drift:
        verify_output(
            output,
            expected_private_input_manifest_sha256="d" * 64,
        )
    assert private_drift.value.code == "PRIVATE_MANIFEST_BINDING_MISMATCH"

    changed = dict(SUBJECT_BINDING)
    changed["digest"] = "e" * 64
    monkeypatch.setattr(operational, "subject_source_binding", lambda: changed)
    with pytest.raises(ProjectCostError) as source_drift:
        verify_output(output)
    assert source_drift.value.code == "SUBJECT_BINDING_MISMATCH"
