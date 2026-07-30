#!/usr/bin/env python3
"""Deterministic KMFA project-cost engine.

The module is intentionally source-only and Python 3.9 compatible.  It reads
private inputs in place, writes only to a caller-selected output directory, and
keeps every accounting observation plane separate.
"""

from __future__ import annotations

import csv
import gzip
import hashlib
import io
import json
import os
import re
import shutil
import stat
import tempfile
import uuid
import warnings
import zipfile
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path, PurePosixPath
from typing import Any, Dict, Iterable, Iterator, List, Mapping, Optional, Sequence, Set, Tuple


CORE_VERSION = "0.2.0"
SKILL_VERSION = "0.0.6"
SUBJECT_DIGEST_RECIPE = "kmfa.project_cost.subject_tree.v1"
CENT = Decimal("0.01")
MAX_GROSS_MARGIN_BPS = 7000
MAX_ARCHIVE_MEMBERS = 20000
MAX_ARCHIVE_UNCOMPRESSED = 1024 * 1024 * 1024
MAX_ARCHIVE_MEMBER = 256 * 1024 * 1024
MAX_COMPRESSION_RATIO = 250
CONTROL_SUMMARIES = ("期初余额", "本期合计", "本年累计", "结转本期损益")
PLACEHOLDER_CONTRACTS = ("KMX999", "KMX9999", "不分项目")
FORMULA_SHEETS = (
    "01_项目成本表",
    "02_成本明细",
    "03_生命周期对照",
    "04_收入与现金",
    "05_来源与核销",
    "06_差异与待确认",
    "07_项目身份",
    "08_运行说明",
)
STATEMENT_TEMPLATE_A = (
    ("一、合同额", "contract"),
    ("二、资金运用及各项支出", "sec2"),
    ("（一）原材料", "l2_material"),
    ("其中:1.主要材料", "d_material"),
    ("2.辅助材料", "blank"),
    ("2.1气体", "d_fuel_power"),
    ("2.2焊材", "blank"),
    ("2.3漆料", "blank"),
    ("2.4低值易损耗材", "blank"),
    ("3 外协 加工费", "blank"),
    ("（二）租赁费", "l2_rental"),
    ("其中:1.吊车租赁费", "blank"),
    ("2.脚手架租赁费", "blank"),
    ("3.物流运输费", "d_logistics"),
    ("（三）保险费", "blank"),
    ("（四）现场管理费", "l2_site"),
    ("1.管理人员工资", "d_own_labor"),
    ("2.差旅费", "d_travel"),
    ("2.1车票", "d_ticket"),
    ("2.2住宿", "d_lodging"),
    ("3.业务费用", "blank"),
    ("3.1招待费", "blank"),
    ("4.生活费用", "d_living"),
    ("4.1生活用品", "blank"),
    ("4.2生活费", "blank"),
    ("5.工程车辆使用费", "d_vehicle"),
    ("5.1加油费及保养", "d_vehicle_fuel"),
    ("5.2过路、停车费", "d_road_parking"),
    ("5.3维修费", "blank"),
    ("6.办公费", "blank"),
    ("7.安全防护费", "blank"),
    ("8.房租", "blank"),
    ("9.临电", "blank"),
    ("10.体检及工伤支出等", "blank"),
    ("11.罚款", "blank"),
    ("12.挂靠管理费", "blank"),
    ("（五）工资（承包费）支出", "l2_subcontract_labor"),
    ("（六）信息费", "blank"),
    ("三 1.1分摊的管理费用（合同的2%）", "allocation"),
    ("1.2占用的资金利息", "interest"),
    ("合计支出", "total"),
    ("（七）毛利", "profit"),
)
MONEY_FORMAT = '#,##0.00;[Red](#,##0.00);-'
CONTRACT_TOKEN_RE = re.compile(r"KMX[0-9A-Z]+(?:-[0-9A-Z]+)+", re.IGNORECASE)
DATE_DIR_RE = re.compile(r"(20\d{2})[-_.]?(\d{2})[-_.]?(\d{2})")
PERIOD_RE = re.compile(r"(20\d{2})年(?:第)?(\d{1,2})期")
PROHIBITED_TEXT_PREFIXES = ("=", "+", "-", "@")
DIRECT_LABOR_DEPARTMENT_TOKENS = ("生产", "项目施工")
NON_SITE_ADMIN_SEGMENTS = {
    "中国",
    "内蒙古",
    "新疆",
    "广西",
    "宁夏",
    "西藏",
    "北京",
    "天津",
    "上海",
    "重庆",
    "湖北",
    "湖南",
    "江西",
    "福建",
    "山东",
    "江苏",
    "浙江",
    "安徽",
    "河南",
    "河北",
    "广东",
    "贵州",
    "青海",
    "四川",
    "云南",
    "陕西",
    "山西",
    "甘肃",
    "辽宁",
    "吉林",
    "黑龙江",
    "海南",
}


class ProjectCostError(RuntimeError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__("%s: %s" % (code, message))
        self.code = code
        self.message = message


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def stable_json(value: Any) -> bytes:
    return (json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n").encode("utf-8")


def pretty_json(value: Any) -> bytes:
    return (json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def subject_source_binding(skill_root: Optional[Path] = None) -> Dict[str, Any]:
    """Hash the complete source-only Skill tree with a reproducible recipe."""

    root = (
        Path(skill_root).resolve()
        if skill_root is not None
        else Path(__file__).resolve().parents[2]
    )
    if root.is_symlink() or not root.is_dir():
        raise ProjectCostError(
            "SUBJECT_ROOT_INVALID",
            "Skill subject root must be a regular directory",
        )
    files: List[Dict[str, Any]] = []
    for path in sorted(
        (candidate for candidate in root.rglob("*") if candidate.is_file()),
        key=lambda candidate: candidate.relative_to(root).as_posix(),
    ):
        relative = path.relative_to(root).as_posix()
        parts = PurePosixPath(relative).parts
        if (
            path.is_symlink()
            or "__pycache__" in parts
            or any(part.startswith(".") for part in parts)
            or relative.endswith((".pyc", ".pyo"))
        ):
            continue
        files.append(
            {
                "path": relative,
                "sha256": sha256_file(path),
                "size_bytes": path.stat().st_size,
            }
        )
    if not files:
        raise ProjectCostError(
            "SUBJECT_TREE_EMPTY",
            "Skill subject tree contains no hashable files",
        )
    digest_payload = {
        "recipe": SUBJECT_DIGEST_RECIPE,
        "files": files,
    }
    return {
        "recipe": SUBJECT_DIGEST_RECIPE,
        "digest": sha256_bytes(stable_json(digest_payload)),
        "file_count": len(files),
        "files": files,
    }


def cents(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        amount = Decimal(str(value).replace(",", "").replace("￥", "").replace("¥", "").strip())
    except (InvalidOperation, ValueError):
        return None
    if not amount.is_finite():
        return None
    return int((amount * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def yuan_from_cents(value: Optional[int]) -> Optional[Decimal]:
    if value is None:
        return None
    return (Decimal(value) / 100).quantize(CENT)


def iso_date(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3))).isoformat()
        except ValueError:
            return None
    return None


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("（", "(").replace("）", ")").replace("－", "-").replace("—", "-")
    return re.sub(r"[\s_()·#]+", "", text).upper()


def spreadsheet_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    clean = value.replace("\x00", "").replace("\r", " ").replace("\n", " ").strip()
    if clean.startswith(PROHIBITED_TEXT_PREFIXES):
        return "'" + clean
    return clean


def contract_base(value: Any) -> str:
    text = "" if value is None else str(value).upper()
    text = text.replace("－", "-").replace("—", "-").replace(" ", "")
    match = CONTRACT_TOKEN_RE.search(text)
    if not match:
        return ""
    token = match.group(0)
    changed = True
    while changed:
        changed = False
        for pattern in (r"-XF--Z\d*$", r"-XF-Z\d*$", r"--Z\d*$", r"-Z\d*$", r"-XF$"):
            reduced = re.sub(pattern, "", token)
            if reduced != token:
                token = reduced
                changed = True
                break
    return token


def contract_year(value: Any) -> Optional[int]:
    base = contract_base(value)
    match = re.match(r"KMX(20\d{2})", base)
    return int(match.group(1)) if match else None


def contract_tokens(value: Any) -> Tuple[str, ...]:
    text = "" if value is None else str(value).upper().replace("－", "-").replace("—", "-")
    return tuple(dict.fromkeys(match.group(0) for match in CONTRACT_TOKEN_RE.finditer(text)))


def edit_distance(left: str, right: str) -> int:
    if left == right:
        return 0
    if not left:
        return len(right)
    if not right:
        return len(left)
    previous = list(range(len(right) + 1))
    for index, lchar in enumerate(left, 1):
        current = [index]
        for jndex, rchar in enumerate(right, 1):
            current.append(
                min(
                    current[-1] + 1,
                    previous[jndex] + 1,
                    previous[jndex - 1] + (0 if lchar == rchar else 1),
                )
            )
        previous = current
    return previous[-1]


def _safe_archive_name(name: str) -> None:
    if not name or "\x00" in name or "\\" in name:
        raise ProjectCostError("ARCHIVE_PATH_UNSAFE", "archive member has an unsafe name")
    pure = PurePosixPath(name)
    if pure.is_absolute() or any(part in ("", ".", "..") for part in pure.parts):
        raise ProjectCostError("ARCHIVE_PATH_UNSAFE", "archive member escapes its container")


def audit_archive(archive: zipfile.ZipFile) -> None:
    members = archive.infolist()
    if len(members) > MAX_ARCHIVE_MEMBERS:
        raise ProjectCostError("ARCHIVE_MEMBER_LIMIT", "archive has too many members")
    total = 0
    for member in members:
        _safe_archive_name(member.filename)
        mode = (member.external_attr >> 16) & 0xFFFF
        if stat.S_ISLNK(mode):
            raise ProjectCostError("ARCHIVE_SYMLINK", "archive contains a symbolic link")
        if member.file_size > MAX_ARCHIVE_MEMBER:
            raise ProjectCostError("ARCHIVE_MEMBER_SIZE", "archive member exceeds the size limit")
        total += member.file_size
        if total > MAX_ARCHIVE_UNCOMPRESSED:
            raise ProjectCostError("ARCHIVE_TOTAL_SIZE", "archive exceeds the uncompressed size limit")
        if member.file_size and member.compress_size == 0:
            raise ProjectCostError("ARCHIVE_RATIO", "archive member has an invalid compression ratio")
        if member.compress_size and member.file_size / member.compress_size > MAX_COMPRESSION_RATIO:
            raise ProjectCostError("ARCHIVE_RATIO", "archive member compression ratio is unsafe")


def sanitized_ooxml_bytes(payload: bytes) -> bytes:
    source = zipfile.ZipFile(io.BytesIO(payload))
    audit_archive(source)
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target:
        for info in source.infolist():
            data = source.read(info)
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                data = re.sub(
                    br"<dataValidations\b[^>]*>.*?</dataValidations>",
                    b"",
                    data,
                    flags=re.DOTALL,
                )
                data = re.sub(br"<extLst\b[^>]*>.*?</extLst>", b"", data, flags=re.DOTALL)
            target.writestr(info, data)
    return output.getvalue()


def open_xlsx_payload(payload: bytes, *, sanitize: bool = True):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ProjectCostError("DEPENDENCY_MISSING", "openpyxl is required") from exc
    prepared = sanitized_ooxml_bytes(payload) if sanitize else payload
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(io.BytesIO(prepared), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        try:
            sheet.reset_dimensions()
        except Exception:
            pass
    return workbook


def locate_header(sheet: Any, required: Sequence[str], max_rows: int = 10) -> Tuple[int, Dict[str, int]]:
    for row_number in range(1, min(max_rows, sheet.max_row or max_rows) + 1):
        try:
            row = next(
                sheet.iter_rows(
                    min_row=row_number,
                    max_row=row_number,
                    values_only=True,
                )
            )
        except StopIteration:
            break
        mapping: Dict[str, int] = {}
        for index, value in enumerate(row):
            key = str(value).strip() if value not in (None, "") else ""
            if key and key not in mapping:
                mapping[key] = index
        if all(name in mapping for name in required):
            return row_number, mapping
    raise ProjectCostError("XLSX_HEADER_NOT_FOUND", "required worksheet headers were not found")


def read_path_bytes(path: Path) -> bytes:
    metadata = Path(path).lstat()
    if not stat.S_ISREG(metadata.st_mode) or stat.S_ISLNK(metadata.st_mode):
        raise ProjectCostError("SOURCE_PATH_UNSAFE", "source must be a regular non-symlink file")
    return Path(path).read_bytes()


def iter_source_files(roots: Sequence[Path]) -> Iterator[Path]:
    seen: Set[Tuple[int, int]] = set()
    for root in roots:
        root_path = Path(root)
        if root_path.is_symlink() or not root_path.is_dir():
            raise ProjectCostError("DATA_ROOT_INVALID", "data root is unavailable or is a symlink: %s" % root_path)
        for current, directories, filenames in os.walk(str(root_path), followlinks=False):
            directories[:] = sorted(
                name
                for name in directories
                if name not in (".git", "__pycache__", "__MACOSX")
                and not name.startswith("._")
                and not (Path(current) / name).is_symlink()
            )
            for filename in sorted(filenames):
                if filename == ".DS_Store" or filename.startswith("._"):
                    continue
                path = Path(current) / filename
                try:
                    metadata = path.lstat()
                except OSError:
                    continue
                if not stat.S_ISREG(metadata.st_mode) or stat.S_ISLNK(metadata.st_mode):
                    continue
                identity = (metadata.st_dev, metadata.st_ino)
                if identity in seen:
                    continue
                seen.add(identity)
                yield path


def relative_to_any(path: Path, roots: Sequence[Path]) -> str:
    resolved = path.resolve()
    for index, root in enumerate(roots):
        try:
            relative = resolved.relative_to(Path(root).resolve())
            return "root%d/%s" % (index + 1, relative.as_posix())
        except ValueError:
            continue
    return path.name


def _date_score(path: Path) -> Tuple[int, int, int, int]:
    scores: List[Tuple[int, int, int]] = []
    for part in path.parts:
        match = DATE_DIR_RE.search(part)
        if match:
            scores.append(tuple(int(match.group(index)) for index in (1, 2, 3)))
    if scores:
        year, month, day = max(scores)
        return year, month, day, path.stat().st_mtime_ns
    stamp = datetime.fromtimestamp(path.stat().st_mtime)
    return stamp.year, stamp.month, stamp.day, path.stat().st_mtime_ns


def source_record(path: Path, roots: Sequence[Path], *, selected: bool, reason: str) -> Dict[str, Any]:
    return {
        "source_id": "src_" + sha256_bytes(relative_to_any(path, roots).encode("utf-8"))[:24],
        "relative_path": relative_to_any(path, roots),
        "absolute_path": str(path.resolve()),
        "sha256": sha256_file(path),
        "size_bytes": path.stat().st_size,
        "selected": selected,
        "selection_reason": reason,
    }


def discover_candidates(files: Sequence[Path]) -> Dict[str, List[Path]]:
    result: Dict[str, List[Path]] = {
        "master": [],
        "status": [],
        "payment": [],
        "approved_cost_detail": [],
        "project_invoice": [],
        "ledger_zip": [],
        "dingtalk_zip": [],
        "dws_zip": [],
        "reference_pdf": [],
    }
    for path in files:
        name = path.name
        lower = name.lower()
        if lower.endswith(".xlsx") and "红圈主合同" in name:
            result["master"].append(path)
        elif lower.endswith(".xlsx") and name == "生产项目状态表.xlsx":
            result["status"].append(path)
        elif lower.endswith(".xlsx") and "付款审批" in name:
            result["payment"].append(path)
        elif lower.endswith(".xlsx") and "项目成本统计" in name:
            result["approved_cost_detail"].append(path)
        elif lower.endswith(".xlsx") and "项目开票_导出文件" in name:
            result["project_invoice"].append(path)
        elif lower.endswith(".zip") and "金蝶" in name and ("账务" in name or "账套" in name):
            result["ledger_zip"].append(path)
        elif lower.endswith(".zip") and name == "DWS_Outputs.zip":
            result["dws_zip"].append(path)
        elif lower.endswith(".zip") and "钉钉" in name:
            result["dingtalk_zip"].append(path)
        elif lower.endswith(".pdf"):
            result["reference_pdf"].append(path)
    return result


def _row_dict(row: Sequence[Any], headers: Mapping[str, int]) -> Dict[str, Any]:
    return {key: row[index] if index < len(row) else None for key, index in headers.items()}


def parse_master(path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    payload = read_path_bytes(path)
    workbook = open_xlsx_payload(payload)
    best_sheet = None
    header_row = None
    headers = None
    for sheet in workbook.worksheets:
        try:
            row_number, mapping = locate_header(sheet, ("合同名称", "合同编号", "甲方", "含税合同额(元)"))
        except ProjectCostError:
            continue
        best_sheet, header_row, headers = sheet, row_number, mapping
        break
    if best_sheet is None or header_row is None or headers is None:
        workbook.close()
        raise ProjectCostError("MASTER_SCHEMA", "red-circle master schema was not recognized")
    projects: List[Dict[str, Any]] = []
    max_created = ""
    for row_number, row in enumerate(
        best_sheet.iter_rows(min_row=header_row + 1, values_only=True),
        header_row + 1,
    ):
        values = _row_dict(row, headers)
        raw_contract = str(values.get("合同编号") or "").strip()
        base = contract_base(raw_contract)
        if not base or any(marker in base for marker in PLACEHOLDER_CONTRACTS):
            continue
        created = iso_date(values.get("创建时间"))
        if created and created > max_created:
            max_created = created
        projects.append(
            {
                "canonical_contract_id": raw_contract,
                "contract_base": base,
                "year": contract_year(raw_contract),
                "project_name": str(values.get("合同名称") or "").strip(),
                "customer": str(values.get("甲方") or "").strip(),
                "contractor": str(values.get("乙方") or "").strip(),
                "construction_status_master": str(values.get("施工状态") or "").strip(),
                "created_date": created,
                "completion_date_master": iso_date(values.get("完工日期（产值上报）")),
                "acceptance_date": iso_date(values.get("验收日期")),
                "contract_amount_cents": cents(values.get("含税合同额(元)")),
                "tax_rate_source": str(values.get("税率(%)") or "").strip(),
                "settlement_amount_cents": cents(values.get("结算金额(元)")),
                "settlement_receivable_cents": cents(values.get("结算应收款(元)")),
                "invoiced_cents": cents(values.get("累计开票")),
                "invoice_receivable_cents": cents(values.get("开票应收款(元)")),
                "cash_in_cents": cents(values.get("项目收款金额")),
                "deposit_paid_cents": cents(values.get("保证金支付金额")),
                "deposit_returned_cents": cents(values.get("保证金退还金额")),
                "source_row": row_number,
            }
        )
    workbook.close()
    if not projects:
        raise ProjectCostError("MASTER_EMPTY", "red-circle master contains no canonical contracts")
    bases = [project["contract_base"] for project in projects]
    duplicate_base_count = len(bases) - len(set(bases))
    return projects, {
        "max_created_date": max_created,
        "project_count": len(projects),
        "historical_normalized_duplicate_count": duplicate_base_count,
    }


def select_master(
    candidates: Sequence[Path],
    roots: Sequence[Path],
) -> Tuple[Path, List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    if not candidates:
        raise ProjectCostError("MASTER_MISSING", "no red-circle master workbook was found")
    parsed: List[Tuple[Tuple[str, int, Tuple[int, int, int, int]], Path, List[Dict[str, Any]], Dict[str, Any]]] = []
    errors: List[Dict[str, Any]] = []
    seen_hashes: Set[str] = set()
    for path in candidates:
        digest = sha256_file(path)
        if digest in seen_hashes:
            continue
        seen_hashes.add(digest)
        try:
            projects, metadata = parse_master(path)
            score = (metadata["max_created_date"], metadata["project_count"], _date_score(path))
            parsed.append((score, path, projects, metadata))
        except Exception as exc:
            errors.append({"source": relative_to_any(path, roots), "error": str(exc)})
    if not parsed:
        raise ProjectCostError("MASTER_UNREADABLE", "all red-circle master candidates failed schema validation")
    parsed.sort(key=lambda item: item[0], reverse=True)
    selected = parsed[0]
    sources = []
    for _, path, _, metadata in parsed:
        sources.append(
            dict(
                source_record(
                    path,
                    roots,
                    selected=path == selected[1],
                    reason="latest in-workbook creation date, then project count and dated snapshot",
                ),
                logical_metadata=metadata,
                source_slot="project_master",
            )
        )
    return selected[1], selected[2], sources, errors


def _project_indexes(projects: Sequence[Mapping[str, Any]], year: int) -> Dict[str, Any]:
    # Callers pass a reporting cohort.  Do not re-filter it by the contract-id
    # year here: a prior-year contract can be an in-scope carryover project
    # when it starts, remains active, or completes in the reporting year.
    # The ``year`` argument remains part of the parser contract and is used by
    # callers for posting/date cut-offs and unresolved-row diagnostics.
    indexed_projects = list(projects)
    by_base = {
        str(project["contract_base"]): project
        for project in indexed_projects
    }
    by_customer: Dict[str, List[Mapping[str, Any]]] = defaultdict(list)
    by_contractor: Dict[str, List[Mapping[str, Any]]] = defaultdict(list)
    for project in indexed_projects:
        by_customer[normalize_text(project.get("customer"))].append(project)
        contractor = normalize_text(project.get("contractor"))
        if contractor:
            by_contractor[contractor].append(project)
    return {
        "projects": indexed_projects,
        "by_base": by_base,
        "by_customer": by_customer,
        "by_contractor": by_contractor,
    }


def reporting_project_cohort(
    projects: Sequence[Mapping[str, Any]],
    status_map: Mapping[str, Mapping[str, Any]],
    *,
    year: int,
) -> List[Mapping[str, Any]]:
    """Return target-year contracts plus evidenced carryover projects.

    Invoice and cash dates alone do not move a completed prior-year project
    into a new project-cost cohort.  A carryover needs a target-year start or
    completion date, a target-year master completion date, or a current
    non-completed production status.
    """

    selected: List[Mapping[str, Any]] = []
    for project in projects:
        contract_year_value = project.get("year")
        if not isinstance(contract_year_value, int):
            continue
        status = status_map.get(str(project["contract_base"]), {})
        construction_status = normalize_text(
            status.get("construction_status")
            or project.get("construction_status_master")
        )
        target_prefix = "%04d-" % year
        starts_in_year = str(status.get("start_date") or "").startswith(
            target_prefix
        )
        completes_in_year = any(
            str(value or "").startswith(target_prefix)
            for value in (
                status.get("completion_date"),
                project.get("completion_date_master"),
            )
        )
        active_carryover = (
            contract_year_value == year - 1
            and bool(status)
            and construction_status
            not in ("已完工", "已终止", "已取消", "作废")
        )
        if (
            contract_year_value == year
            or starts_in_year
            or completes_in_year
            or active_carryover
        ):
            selected.append(project)
    return sorted(
        selected,
        key=lambda project: (
            project.get("created_date") or "",
            str(project["contract_base"]),
        ),
    )


def _entity_ledger_period_end(
    entity_name: Any,
    period_ends_by_entity: Mapping[str, str],
) -> Optional[str]:
    """Resolve a governed entity label to exactly one selected GL entity."""

    entity_name_normalized = normalize_text(entity_name)
    if not entity_name_normalized:
        return None
    exact = period_ends_by_entity.get(entity_name_normalized)
    if exact:
        return str(exact)
    matches = [
        str(period_end)
        for entity, period_end in period_ends_by_entity.items()
        if len(entity) >= 4
        and (
            entity in entity_name_normalized
            or entity_name_normalized in entity
        )
    ]
    return matches[0] if len(matches) == 1 else None


def _project_ledger_period_end(
    project: Mapping[str, Any],
    period_ends_by_entity: Mapping[str, str],
) -> Optional[str]:
    return _entity_ledger_period_end(
        project.get("contractor"),
        period_ends_by_entity,
    )


def resolve_identity(
    text: str,
    customer: str,
    indexes: Mapping[str, Any],
    *,
    allow_text_only: bool,
) -> Tuple[Optional[Mapping[str, Any]], str, Tuple[str, ...]]:
    by_base = indexes["by_base"]
    raw_tokens = contract_tokens(text)
    for token in raw_tokens:
        base = contract_base(token)
        if base in by_base:
            reason = "EXACT_CONTRACT" if token == by_base[base]["canonical_contract_id"] else "CONTROLLED_SUFFIX_ALIAS"
            return by_base[base], reason, raw_tokens
    customer_key = normalize_text(customer)
    candidates = list(indexes["by_customer"].get(customer_key, ())) if customer_key else []
    if len(candidates) == 1 and raw_tokens:
        candidate = candidates[0]
        distances = [edit_distance(contract_base(token), str(candidate["contract_base"])) for token in raw_tokens]
        if min(distances) <= 2:
            return candidate, "CUSTOMER_UNIQUE_CONTRACT_TYPO_REPAIR", raw_tokens
        source_sequences = {contract_base(token).rsplit("-", 1)[-1] for token in raw_tokens}
        target_sequence = str(candidate["contract_base"]).rsplit("-", 1)[-1]
        if target_sequence in source_sequences:
            return candidate, "CUSTOMER_UNIQUE_SEQUENCE_ALIAS", raw_tokens
    if allow_text_only:
        normalized = normalize_text(text)
        text_hits: Dict[str, Mapping[str, Any]] = {}
        for key, items in indexes["by_customer"].items():
            if key and len(key) >= 4 and key in normalized and len(items) == 1:
                text_hits[str(items[0]["contract_base"])] = items[0]
        if len(text_hits) == 1:
            return next(iter(text_hits.values())), "UNIQUE_CUSTOMER_TEXT", raw_tokens
        # Contractor-only narrative matching is intentionally disabled.
        # Contractor names also occur in internal reimbursements and related-
        # party transactions, so they are not sufficient project identity
        # evidence. Earlier drafts special-cased real tenant company names
        # here; that made the rule both unsafe to publish and non-portable.
    return None, "UNRESOLVED", raw_tokens


def _customer_name(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return re.sub(r"^[A-Z0-9.]+_", "", text, flags=re.IGNORECASE).strip()


def _as_date(value: Any) -> Optional[date]:
    parsed = iso_date(value)
    if not parsed:
        return None
    try:
        return date.fromisoformat(parsed)
    except ValueError:
        return None


def _summary_service_date(summary: str, posting_date: Optional[str]) -> Optional[date]:
    """Extract a conservative service date from a voucher summary.

    Full dates win.  Month/day fragments are accepted only when they can be
    anchored to the posting year without landing materially after posting.
    """

    posting = _as_date(posting_date)
    full = re.search(r"(20\d{2})[./年-](\d{1,2})[./月-](\d{1,2})", summary)
    if full:
        try:
            return date(int(full.group(1)), int(full.group(2)), int(full.group(3)))
        except ValueError:
            pass
    if posting is None:
        return None
    fragments = re.findall(r"(?<!\d)(\d{1,2})[./月](\d{1,2})(?:日|号)?", summary)
    candidates: List[date] = []
    for month, day_value in fragments:
        try:
            candidate = date(posting.year, int(month), int(day_value))
        except ValueError:
            continue
        if candidate > posting + timedelta(days=7):
            try:
                candidate = date(posting.year - 1, int(month), int(day_value))
            except ValueError:
                continue
        if candidate <= posting + timedelta(days=7):
            candidates.append(candidate)
    return max(candidates) if candidates else None


def _project_window(
    project: Mapping[str, Any],
    status_map: Mapping[str, Mapping[str, Any]],
) -> Tuple[Optional[date], Optional[date], Optional[date]]:
    base = str(project["contract_base"])
    status = status_map.get(base, {})
    created = _as_date(project.get("created_date"))
    start = _as_date(status.get("start_date")) or created
    end = (
        _as_date(status.get("completion_date"))
        or _as_date(project.get("completion_date_master"))
    )
    return created, start, end


def resolve_ledger_identity(
    *,
    text: str,
    customer: str,
    sales_contract: str,
    posting_date: Optional[str],
    summary: str,
    indexes: Mapping[str, Any],
    status_map: Mapping[str, Mapping[str, Any]],
) -> Tuple[Optional[Mapping[str, Any]], str, Tuple[str, ...]]:
    """Resolve ledger identity without allowing customer-only historical bleed.

    An explicit non-placeholder contract is authoritative: it must resolve
    exactly or through the narrow typo rules in ``resolve_identity``.  Customer
    plus time-window recovery is considered only when the sales-contract
    dimension is empty or explicitly unallocated.
    """

    project, reason, tokens = resolve_identity(
        text,
        _customer_name(customer),
        indexes,
        allow_text_only=False,
    )
    if project is not None:
        return project, reason, tokens

    explicit_tokens = contract_tokens(sales_contract)
    meaningful_tokens = tuple(
        token
        for token in explicit_tokens
        if not any(marker in token.upper() for marker in PLACEHOLDER_CONTRACTS)
    )
    if meaningful_tokens:
        return None, "EXPLICIT_OTHER_CONTRACT", tokens

    customer_key = normalize_text(_customer_name(customer))
    candidates = list(indexes["by_customer"].get(customer_key, ())) if customer_key else []
    if not candidates:
        return None, "UNRESOLVED", tokens

    observed = _summary_service_date(summary, posting_date) or _as_date(posting_date)
    if observed is None:
        return None, "CUSTOMER_DATE_MISSING", tokens
    preaward = any(
        marker in summary
        for marker in ("投标", "标书", "勘察", "踏勘", "技术沟通", "现场沟通")
    )
    eligible: List[Mapping[str, Any]] = []
    for candidate in candidates:
        created, start, end = _project_window(candidate, status_map)
        if created is not None and observed < created:
            allowed_lead = 45 if preaward else 14
            if created - observed > timedelta(days=allowed_lead):
                continue
        grace_days = 90 if len(candidates) == 1 else 45
        if end is not None and observed > end + timedelta(days=grace_days):
            continue
        eligible.append(candidate)
    if len(eligible) == 1:
        resolved_reason = (
            (
                "PRE_AWARD_UNIQUE_NEXT_CONTRACT"
                if preaward
                else "PRECONTRACT_UNIQUE_CUSTOMER"
            )
            if _as_date(eligible[0].get("created_date"))
            and observed < _as_date(eligible[0].get("created_date"))
            else "UNIQUE_CUSTOMER_ACTIVE_WINDOW"
        )
        return eligible[0], resolved_reason, tokens
    if len(eligible) > 1:
        return None, "CUSTOMER_ACTIVE_WINDOW_AMBIGUOUS", tokens
    return None, "CUSTOMER_OUTSIDE_PROJECT_WINDOW", tokens


def parse_status(
    path: Path,
    projects: Sequence[Mapping[str, Any]],
    year: int,
) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    payload = read_path_bytes(path)
    workbook = open_xlsx_payload(payload)
    indexes = _project_indexes(projects, year)
    selected_sheet = None
    header_row = None
    headers = None
    for sheet in workbook.worksheets:
        try:
            row_number, mapping = locate_header(sheet, ("甲方名称", "合同号", "施工状态"))
        except ProjectCostError:
            continue
        selected_sheet, header_row, headers = sheet, row_number, mapping
        break
    if selected_sheet is None or header_row is None or headers is None:
        workbook.close()
        raise ProjectCostError("STATUS_SCHEMA", "production status schema was not recognized")
    result: Dict[str, Dict[str, Any]] = {}
    reviews: List[Dict[str, Any]] = []
    max_observed_date = ""
    provided_report_rows = 0
    provided_report_with_direct_components = 0
    for row_number, row in enumerate(
        selected_sheet.iter_rows(min_row=header_row + 1, values_only=True),
        header_row + 1,
    ):
        values = _row_dict(row, headers)
        raw_contract = str(values.get("合同号") or "").strip()
        customer = str(values.get("甲方名称") or "").strip()
        project, reason, tokens = resolve_identity(
            raw_contract + " " + customer,
            customer,
            indexes,
            allow_text_only=False,
        )
        if project is None:
            if contract_year(raw_contract) == year:
                reviews.append(
                    {
                        "severity": "P1",
                        "type": "STATUS_IDENTITY_UNRESOLVED",
                        "source_row": row_number,
                        "raw_contract": raw_contract,
                        "customer": customer,
                        "action": "人工确认状态表合同与红圈主合同的唯一映射",
                    }
                )
            continue
        direct_components = {
            "生活住宿费": cents(values.get("生活住宿费")),
            "交通费": cents(values.get("交通费")),
            "材料费": cents(values.get("材料费")),
            "其他费用": cents(values.get("其他费用")),
        }
        present = [amount for amount in direct_components.values() if amount is not None]
        observed_dates = [
            iso_date(values.get(name))
            for name in ("开工时间", "完工时间", "结算时间", "开票时间", "回款时间")
        ]
        for observed in observed_dates:
            if observed and observed > max_observed_date:
                max_observed_date = observed
        base = str(project["contract_base"])
        provided_text = normalize_text(values.get("是否提供项目成本表"))
        project_cost_report_provided = provided_text == "已提供"
        if project_cost_report_provided:
            provided_report_rows += 1
            if present:
                provided_report_with_direct_components += 1
        result[base] = {
            "source_contract": raw_contract,
            "identity_reason": reason,
            "construction_status": str(values.get("施工状态") or "").strip(),
            "project_type": str(values.get("项目类型") or "").strip(),
            "owner": str(values.get("负责人") or "").strip(),
            "start_date": iso_date(values.get("开工时间")),
            "completion_date": iso_date(values.get("完工时间")),
            "settlement_date": iso_date(values.get("结算时间")),
            "invoice_date": iso_date(values.get("开票时间")),
            "cash_in_date": iso_date(values.get("回款时间")),
            "actual_duration_days": values.get("实际工期"),
            "status_contract_amount_cents": cents(values.get("含税合同金额")),
            "status_tax_rate_source": str(values.get("税率") or "").strip(),
            "settlement_amount_cents": cents(values.get("结算金额")),
            "invoice_amount_cents": cents(values.get("开票金额")),
            "own_work_units": values.get("自有人工工时"),
            "external_work_units": values.get("劳务人工工时"),
            "business_components_cents": direct_components,
            "business_reported_direct_cost_cents": sum(present) if present else None,
            "project_cost_report_deadline": iso_date(
                values.get("项目成本表截止提供时间")
            ),
            "project_cost_report_provided": project_cost_report_provided,
            "project_cost_report_provided_source": str(
                values.get("是否提供项目成本表") or ""
            ).strip(),
            "commission_calculated": normalize_text(
                values.get("是否已计算提成")
            )
            == "是",
            "source_row": row_number,
        }
        if reason not in ("EXACT_CONTRACT", "CONTROLLED_SUFFIX_ALIAS"):
            reviews.append(
                {
                    "severity": "P2",
                    "type": "STATUS_IDENTITY_REPAIRED",
                    "project": base,
                    "source_row": row_number,
                    "raw_contract": raw_contract,
                    "reason": reason,
                    "action": "已基于唯一客户/序号证据保留别名；复核后可冻结",
                }
            )
    workbook.close()
    return result, reviews, {
        "mapped_rows": len(result),
        "max_observed_date": max_observed_date,
        "provided_project_cost_report_rows": provided_report_rows,
        "provided_project_cost_report_with_direct_components": (
            provided_report_with_direct_components
        ),
    }


def governed_contract_revenue(
    project: Mapping[str, Any],
    status: Optional[Mapping[str, Any]],
) -> Dict[str, Any]:
    """Resolve the contract-margin revenue plane without mixing lifecycle facts.

    Original contract value, settlement and billed amount are three different
    facts.  A completed project can use a positive governed settlement-register
    amount as its contract-margin revenue basis.  Invoice amount remains a
    separate lifecycle observation and is never substituted for settlement.
    Projects without a closed settlement/change basis remain blocked.
    """

    status_row = status or {}
    master_contract = project.get("contract_amount_cents")
    status_contract = status_row.get("status_contract_amount_cents")
    settlement = status_row.get("settlement_amount_cents")
    construction_status = normalize_text(
        status_row.get("construction_status")
        or project.get("construction_status_master")
    )

    positive_master = (
        master_contract
        if isinstance(master_contract, int)
        and not isinstance(master_contract, bool)
        and master_contract > 0
        else None
    )
    positive_status_contract = (
        status_contract
        if isinstance(status_contract, int)
        and not isinstance(status_contract, bool)
        and status_contract > 0
        else None
    )
    positive_settlement = (
        settlement
        if isinstance(settlement, int)
        and not isinstance(settlement, bool)
        and settlement > 0
        else None
    )
    if (
        positive_master is not None
        and positive_status_contract is not None
        and positive_master != positive_status_contract
    ):
        return {
            "effective_revenue_cents": None,
            "status": "BLOCKED_CONTRACT_REGISTER_CONFLICT",
            "basis": None,
        }
    if construction_status == "已完工" and positive_settlement is not None:
        return {
            "effective_revenue_cents": positive_settlement,
            "status": "READY_SETTLEMENT_REGISTER",
            "basis": "GOVERNED_SETTLEMENT_REGISTER",
        }
    return {
        "effective_revenue_cents": None,
        "status": "BLOCKED_CONTRACT_CHANGE_COMPLETENESS",
        "basis": None,
    }


def select_latest_tabular(
    candidates: Sequence[Path],
    roots: Sequence[Path],
    parser: Any,
    *,
    source_slot: str,
    parser_args: Sequence[Any],
) -> Tuple[Optional[Path], Any, List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    if not candidates:
        return None, None, [], [], {}
    parsed: List[Tuple[Tuple[Any, ...], Path, Any, List[Dict[str, Any]], Dict[str, Any]]] = []
    errors: List[Dict[str, Any]] = []
    seen_hashes: Set[str] = set()
    for path in candidates:
        digest = sha256_file(path)
        if digest in seen_hashes:
            continue
        seen_hashes.add(digest)
        try:
            result, reviews, metadata = parser(path, *parser_args)
            score = (
                metadata.get("max_observed_date", ""),
                metadata.get("mapped_rows", 0),
                _date_score(path),
            )
            parsed.append((score, path, result, reviews, metadata))
        except Exception as exc:
            errors.append({"source": relative_to_any(path, roots), "error": str(exc)})
    if not parsed:
        return None, None, [], errors, {}
    parsed.sort(key=lambda item: item[0], reverse=True)
    selected = parsed[0]
    sources: List[Dict[str, Any]] = []
    for _, path, _, _, metadata in parsed:
        sources.append(
            dict(
                source_record(
                    path,
                    roots,
                    selected=path == selected[1],
                    reason="latest content date, then mapped row count and dated snapshot",
                ),
                source_slot=source_slot,
                logical_metadata=metadata,
            )
        )
    error_reviews = [
        {
            "severity": "P1",
            "type": "SOURCE_CANDIDATE_REJECTED",
            "source": error.get("source"),
            "detail": error.get("error"),
            "action": "已跳过失败候选；如其应为权威来源需修复后以新快照重跑",
        }
        for error in errors
    ]
    return selected[1], selected[2], sources, selected[3] + error_reviews, selected[4]


def iter_nested_archive_files(
    payload: bytes,
    *,
    container: str,
    depth: int = 0,
    max_depth: int = 2,
) -> Iterator[Tuple[str, bytes]]:
    if depth > max_depth:
        return
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        audit_archive(archive)
        for info in archive.infolist():
            if info.is_dir():
                continue
            name = "%s!/%s" % (container, info.filename)
            lower = info.filename.lower()
            member = archive.read(info)
            if lower.endswith(".xlsx"):
                yield name, member
            elif lower.endswith(".zip") and depth < max_depth:
                try:
                    for nested in iter_nested_archive_files(
                        member,
                        container=name,
                        depth=depth + 1,
                        max_depth=max_depth,
                    ):
                        yield nested
                except (zipfile.BadZipFile, ProjectCostError):
                    continue


def ledger_book_metadata(name: str, payload: bytes) -> Dict[str, Any]:
    workbook = open_xlsx_payload(payload)
    if not workbook.worksheets:
        workbook.close()
        raise ProjectCostError("LEDGER_EMPTY", "ledger workbook has no worksheets")
    sheet = workbook.worksheets[0]
    first_rows: List[Tuple[Any, ...]] = []
    for row in sheet.iter_rows(min_row=1, max_row=min(5, sheet.max_row or 5), values_only=True):
        first_rows.append(tuple(row))
    row2 = first_rows[1] if len(first_rows) > 1 else ()
    company = ""
    for row in first_rows:
        for value in row:
            text = str(value).strip() if value not in (None, "") else ""
            match = re.search(r"公司名称[：:]?\s*(.+)", text)
            if match:
                company = match.group(1).strip()
                break
        if company:
            break
    if not company:
        company = next((str(value).strip() for value in row2 if value not in (None, "")), "")
    if company.startswith("公司名称："):
        company = company.split("：", 1)[1].strip()
    if not company or company in ("明细账", "UNKNOWN_ENTITY"):
        member_hint = PurePosixPath(name.rsplit("!/", 1)[-1]).stem
        generic_hint = re.sub(
            r"^明细账[-_ ]*",
            "",
            member_hint,
        )
        generic_hint = re.sub(
            r"[-_ ]*(?:明细账)?[-_ ]*20\d{2}.*$",
            "",
            generic_hint,
        ).strip("-_ ")
        if generic_hint:
            company = generic_hint
    period_text = " | ".join(
        str(value)
        for row in first_rows
        for value in row
        if value not in (None, "")
    ) + " | " + name
    periods = [(int(year), int(month)) for year, month in PERIOD_RE.findall(period_text)]
    if not periods:
        compact_periods = re.findall(r"(20\d{2})(\d{2})", period_text)
        periods = [(int(year), int(month)) for year, month in compact_periods if 1 <= int(month) <= 12]
    if periods:
        start = min(periods)
        end = max(periods)
    else:
        start = (0, 0)
        end = (0, 0)
    metadata = {
        "container_member": name,
        "company": company or "UNKNOWN_ENTITY",
        "period_start": "%04d-%02d" % start,
        "period_end": "%04d-%02d" % end,
        "sheet_count": len(workbook.sheetnames),
        "payload_sha256": sha256_bytes(payload),
        "size_bytes": len(payload),
    }
    workbook.close()
    return metadata


def collect_ledger_books(
    ledger_archives: Sequence[Path],
    roots: Sequence[Path],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    payloads: List[Dict[str, Any]] = []
    sources: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen_archive_hashes: Set[str] = set()
    seen_book_hashes: Set[str] = set()
    for archive_path in ledger_archives:
        archive_hash = sha256_file(archive_path)
        if archive_hash in seen_archive_hashes:
            sources.append(
                dict(
                    source_record(
                        archive_path,
                        roots,
                        selected=False,
                        reason="duplicate container SHA-256",
                    ),
                    source_slot="general_ledger_container",
                )
            )
            continue
        seen_archive_hashes.add(archive_hash)
        try:
            container_payload = read_path_bytes(archive_path)
            candidate_count = 0
            for member_name, member_payload in iter_nested_archive_files(
                container_payload,
                container=relative_to_any(archive_path, roots),
            ):
                # The outer ZIP itself is often named “...明细账.zip” and
                # contains both detailed ledgers and voucher-list projections.
                # Testing the full nested path therefore accidentally admitted
                # every `凭证列表.xlsx` and counted the same posting twice.
                # Only the final workbook name decides whether it is a ledger.
                logical_member_name = PurePosixPath(
                    member_name.rsplit("!/", 1)[-1]
                ).name
                if "明细账" not in logical_member_name:
                    continue
                digest = sha256_bytes(member_payload)
                if digest in seen_book_hashes:
                    continue
                seen_book_hashes.add(digest)
                try:
                    metadata = ledger_book_metadata(member_name, member_payload)
                except Exception as exc:
                    errors.append({"source": member_name, "error": str(exc)})
                    continue
                candidate_count += 1
                payloads.append({"metadata": metadata, "payload": member_payload})
            sources.append(
                dict(
                    source_record(
                        archive_path,
                        roots,
                        selected=candidate_count > 0,
                        reason="safe archive with distinct ledger workbooks",
                    ),
                    source_slot="general_ledger_container",
                    logical_metadata={"candidate_ledger_books": candidate_count},
                )
            )
        except Exception as exc:
            errors.append({"source": relative_to_any(archive_path, roots), "error": str(exc)})
            sources.append(
                dict(
                    source_record(
                        archive_path,
                        roots,
                        selected=False,
                        reason="archive rejected or unreadable",
                    ),
                    source_slot="general_ledger_container",
                )
            )
    if not payloads:
        return [], sources, errors
    # Distinct workbooks are all read.  Overlapping cumulative and monthly
    # exports are de-duplicated at voucher-row level later; filename period
    # containment is not evidence that the narrower book has no unique rows.
    selected_books = sorted(
        payloads,
        key=lambda item: (
            item["metadata"]["company"],
            item["metadata"]["period_start"],
            item["metadata"]["period_end"],
            item["metadata"]["payload_sha256"],
        ),
    )
    for item in selected_books:
        metadata = item["metadata"]
        metadata["selected"] = True
        metadata["selection_reason"] = "distinct ledger workbook; semantic row de-duplication applied"
        sources.append(
            {
                "source_id": "src_" + metadata["payload_sha256"][:24],
                "relative_path": metadata["container_member"],
                "absolute_path": None,
                "sha256": metadata["payload_sha256"],
                "size_bytes": metadata["size_bytes"],
                "selected": True,
                "selection_reason": metadata["selection_reason"],
                "source_slot": "general_ledger_book",
                "logical_metadata": {
                    "company": metadata["company"],
                    "period_start": metadata["period_start"],
                    "period_end": metadata["period_end"],
                    "sheet_count": metadata["sheet_count"],
                },
            }
        )
    return selected_books, sources, errors


def account_category(account: str) -> str:
    code = account.split("-", 1)[0].replace("_", "").strip()
    mappings = (
        ("500100401", "住宿"),
        ("500100402", "生活补助"),
        ("500100403", "交通/差旅"),
        ("500100404", "过路停车"),
        ("500100405", "车辆油费"),
        ("500100408", "物流运杂"),
        ("500100410", "外协"),
        ("500100411", "设备租赁"),
        ("500100499", "其他直接成本"),
        ("5001001", "材料"),
        ("5001002", "燃料及动力"),
        ("5001003", "自有人工过账"),
        ("5001006", "已过账制造费用分配"),
        ("5001007", "劳务/分包"),
    )
    for prefix, category in mappings:
        if code.startswith(prefix):
            return category
    if code.startswith("5001"):
        return "其他5001"
    if code.startswith("6401"):
        return "已结转主营业务成本"
    return "未分类"


def _safe_row_value(row: Sequence[Any], headers: Mapping[str, int], names: Sequence[str]) -> Any:
    for name in names:
        if name in headers:
            index = headers[name]
            return row[index] if index < len(row) else None
    return None


def parse_ledger_books(
    books: Sequence[Mapping[str, Any]],
    projects: Sequence[Mapping[str, Any]],
    year: int,
    as_of: str,
    status_map: Mapping[str, Mapping[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    indexes = _project_indexes(projects, year)
    project_window_starts = [
        candidate
        for project in projects
        for candidate in _project_window(project, status_map)[:2]
        if candidate is not None
    ]
    cohort_scan_start = (
        min(project_window_starts) - timedelta(days=45)
        if project_window_starts
        else date(year, 1, 1)
    )
    events: List[Dict[str, Any]] = []
    reviews: List[Dict[str, Any]] = []
    event_rep_counts: Dict[
        Tuple[Any, ...],
        Dict[Tuple[str, str], int],
    ] = defaultdict(lambda: defaultdict(int))
    review_rep_counts: Dict[
        Tuple[Any, ...],
        Dict[Tuple[str, str], int],
    ] = defaultdict(lambda: defaultdict(int))
    unallocated_rep_counts: Dict[
        Tuple[Any, ...],
        Dict[Tuple[str, str], int],
    ] = defaultdict(lambda: defaultdict(int))
    raw_entity_keys = {
        normalize_text(book["metadata"].get("company"))
        for book in books
        if normalize_text(book["metadata"].get("company"))
    }
    entity_aliases: Dict[str, str] = {}
    for entity_key in raw_entity_keys:
        contained = [
            candidate
            for candidate in raw_entity_keys
            if len(candidate) >= 4 and candidate in entity_key
        ]
        entity_aliases[entity_key] = min(
            contained or [entity_key],
            key=lambda candidate: (len(candidate), candidate),
        )
    period_ends_by_entity: Dict[str, str] = {}
    selected_period_end = ""
    minimum_period_end = ""
    scanned_sheets = 0
    scanned_rows = 0
    duplicate_rows = 0
    unresolved_current_customer_rows = 0
    unresolved_current_customer_cents = 0
    unallocated_cost_rows = 0
    unallocated_cost_net_cents = 0
    unallocated_cost_absolute_cents = 0
    unallocated_cost_buckets: Dict[
        Tuple[str, str, str, str],
        Dict[str, int],
    ] = defaultdict(lambda: {"row_count": 0, "net_cents": 0, "absolute_cents": 0})
    for book in books:
        metadata = book["metadata"]
        selected_period_end = max(selected_period_end, metadata["period_end"])
        if metadata["period_end"]:
            minimum_period_end = (
                min(minimum_period_end, metadata["period_end"])
                if minimum_period_end
                else metadata["period_end"]
            )
        workbook = open_xlsx_payload(book["payload"])
        entity = metadata["company"]
        raw_entity_key = normalize_text(entity)
        entity_key = entity_aliases.get(raw_entity_key, raw_entity_key)
        # One legal entity is exported under both a short display name and a
        # full company name in the supplied Kingdee bundles.  Event identity
        # already uses ``entity_key`` to collapse those representations, so
        # period coverage must use the same key.  Keeping the raw label here
        # made a project with otherwise exact ledger rows look as if it had
        # two entities and incorrectly raised PROJECT_COST_SOURCE_UNAVAILABLE.
        period_ends_by_entity[entity_key] = max(
            period_ends_by_entity.get(entity_key, ""),
            str(metadata["period_end"]),
        )
        for sheet in workbook.worksheets:
            title = sheet.title
            many_sheets = len(workbook.sheetnames) > 3
            if many_sheets and not title.startswith(("5001", "6401")):
                continue
            try:
                header_row, headers = locate_header(sheet, ("科目", "摘要"))
            except ProjectCostError:
                continue
            scanned_sheets += 1
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True),
                header_row + 1,
            ):
                scanned_rows += 1
                account = str(_safe_row_value(row, headers, ("科目",)) or "").strip()
                if not account.startswith(("5001", "6401")):
                    continue
                summary = str(_safe_row_value(row, headers, ("摘要",)) or "").strip()
                if any(marker in summary for marker in CONTROL_SUMMARIES):
                    continue
                voucher = str(_safe_row_value(row, headers, ("凭证字号", "凭证号")) or "").strip()
                if not voucher:
                    continue
                debit_minor = cents(_safe_row_value(row, headers, ("借方", "借方金额"))) or 0
                credit_minor = cents(_safe_row_value(row, headers, ("贷方", "贷方金额"))) or 0
                net_minor = debit_minor - credit_minor
                if net_minor == 0:
                    continue
                customer = str(_safe_row_value(row, headers, ("客户", "甲方", "客户名称")) or "")
                sales_contract = str(_safe_row_value(row, headers, ("销售合同号",)) or "")
                supplier = str(
                    _safe_row_value(row, headers, ("供应商", "供应商名称")) or ""
                )
                posting_date = iso_date(_safe_row_value(row, headers, ("日期", "过账日期")))
                posting_date_value = _as_date(posting_date)
                if (
                    posting_date_value is not None
                    and posting_date_value < cohort_scan_start
                ):
                    continue
                if posting_date and posting_date > as_of:
                    reviews.append(
                        {
                            "severity": "P1",
                            "type": "LEDGER_POSTING_AFTER_AS_OF_EXCLUDED",
                            "entity": entity,
                            "sheet": title,
                            "row": row_number,
                            "voucher": voucher,
                            "posting_date": posting_date,
                            "as_of": as_of,
                            "amount_cents": net_minor,
                            "action": "过账日晚于报表截至日；已从正式金额排除，待下一截至日重算",
                        }
                    )
                    continue
                if account.startswith("5001") and any(
                    marker in summary for marker in ("结转生产成本", "结转成本")
                ):
                    # Transfer to 6401 clears the WIP account but does not undo
                    # the project cost incurred.  The matching 6401 debit is
                    # reported independently as recognized COGS.
                    continue
                text = title + " | " + " | ".join("" if value is None else str(value) for value in row)
                project, reason, tokens = resolve_ledger_identity(
                    text=text,
                    customer=customer,
                    sales_contract=sales_contract,
                    posting_date=posting_date,
                    summary=summary,
                    indexes=indexes,
                    status_map=status_map,
                )
                if project is None and account.startswith("5001"):
                    narrative_project, narrative_reason = (
                        resolve_narrative_identity(
                            summary,
                            posting_date,
                            indexes,
                            status_map,
                        )
                    )
                    if narrative_project is not None:
                        project = narrative_project
                        reason = "LEDGER_%s" % narrative_reason
                account_code = account.split("-", 1)[0].split("_", 1)[0].strip()
                representation = (
                    str(metadata["payload_sha256"]),
                    str(title),
                )
                if project is None:
                    # Unassigned 5001 rows were previously dropped silently.
                    # They are a cost-completeness risk even when no current
                    # customer token is present.  Preserve a deduplicated pool
                    # and block formal margin publication until governed
                    # project/WBS evidence assigns or excludes it.
                    unallocated_reasons = {
                        "UNRESOLVED",
                        "CUSTOMER_DATE_MISSING",
                        "CUSTOMER_ACTIVE_WINDOW_AMBIGUOUS",
                    }
                    if account.startswith("5001") and reason in unallocated_reasons:
                        pool_key = (
                            entity_key,
                            posting_date,
                            voucher,
                            normalize_text(account_code),
                            normalize_text(customer),
                            normalize_text(sales_contract),
                            net_minor,
                            normalize_text(summary),
                        )
                        other_max = max(
                            (
                                count
                                for source_representation, count
                                in unallocated_rep_counts[pool_key].items()
                                if source_representation != representation
                            ),
                            default=0,
                        )
                        unallocated_rep_counts[pool_key][representation] += 1
                        occurrence = unallocated_rep_counts[pool_key][
                            representation
                        ]
                        if occurrence <= other_max:
                            duplicate_rows += 1
                            continue
                        unallocated_cost_rows += 1
                        unallocated_cost_net_cents += net_minor
                        unallocated_cost_absolute_cents += abs(net_minor)
                        bucket_key = (
                            entity_key,
                            (posting_date or str(metadata["period_end"]))[:7],
                            account_category(account),
                            reason,
                        )
                        bucket = unallocated_cost_buckets[bucket_key]
                        bucket["row_count"] += 1
                        bucket["net_cents"] += net_minor
                        bucket["absolute_cents"] += abs(net_minor)
                        events.append(
                            {
                                "event_id": "unallocated_"
                                + sha256_bytes(
                                    stable_json([list(pool_key), occurrence])
                                )[:24],
                                "project": None,
                                "plane": "UNALLOCATED_LEDGER_COST_POOL",
                                "category": account_category(account),
                                "amount_cents": net_minor,
                                "posting_date": posting_date,
                                "account_code": account_code,
                                "voucher": voucher,
                                "summary": summary,
                                "supplier": supplier,
                                "sales_contract": sales_contract,
                                "entity": entity,
                                "source_id": "src_"
                                + metadata["payload_sha256"][:24],
                                "source_member": metadata["container_member"],
                                "sheet": title,
                                "row": row_number,
                                "identity_reason": reason,
                                "link_text": normalize_text(text),
                            }
                        )
                    customer_candidates = indexes["by_customer"].get(
                        normalize_text(_customer_name(customer)),
                        (),
                    )
                    if reason == "EXPLICIT_OTHER_CONTRACT":
                        review_key = (
                            "EXPLICIT_OTHER_CONTRACT",
                            entity_key,
                            posting_date,
                            voucher,
                            normalize_text(account_code),
                            normalize_text(customer),
                            normalize_text(sales_contract),
                            net_minor,
                            normalize_text(summary),
                        )
                        other_max = max(
                            (
                                count
                                for source_representation, count
                                in review_rep_counts[review_key].items()
                                if source_representation != representation
                            ),
                            default=0,
                        )
                        review_rep_counts[review_key][representation] += 1
                        occurrence = review_rep_counts[review_key][representation]
                        if occurrence <= other_max:
                            duplicate_rows += 1
                            continue
                        reviews.append(
                            {
                                "severity": "P2",
                                "type": "LEDGER_OTHER_CONTRACT_EXCLUDED",
                                "source": metadata["container_member"],
                                "sheet": title,
                                "row": row_number,
                                "raw_contracts": list(tokens),
                                "posting_date": posting_date,
                                "amount_cents": net_minor,
                                "reason": reason,
                                "action": "明示为其他合同；从本报告项目组合成本公式中排除",
                            }
                        )
                    elif customer_candidates:
                        unresolved_current_customer_rows += 1
                        unresolved_current_customer_cents += net_minor
                        outside_window = reason == "CUSTOMER_OUTSIDE_PROJECT_WINDOW"
                        reviews.append(
                            {
                                "severity": "P2" if outside_window else "P1",
                                "type": (
                                    "LEDGER_OUTSIDE_PROJECT_WINDOW_EXCLUDED"
                                    if outside_window
                                    else "LEDGER_IDENTITY_UNRESOLVED"
                                ),
                                "source": metadata["container_member"],
                                "sheet": title,
                                "row": row_number,
                                "raw_contracts": list(tokens),
                                "customer": customer,
                                "sales_contract": sales_contract,
                                "posting_date": posting_date,
                                "amount_cents": net_minor,
                                "reason": reason,
                                "candidate_projects": [
                                    str(candidate["contract_base"])
                                    for candidate in customer_candidates
                                ],
                                "action": (
                                    "客户相同但凭证服务日不在本报告候选项目窗口；"
                                    "保留审计并从本报告项目公式排除"
                                    if outside_window
                                    else "保留在未分配池；确认合同或项目窗口后再归属"
                                ),
                            }
                        )
                    continue
                if account.startswith("5001"):
                    plane = "JOB_POSTED_ACTUAL"
                elif account.startswith("6401"):
                    plane = "GL_RECOGNIZED_COGS"
                else:
                    continue
                project_base = str(project["contract_base"])
                # The same business row is exported twice in some Kingdee
                # workbooks: once on the plain subject sheet and once on an
                # auxiliary-account-expanded sheet.  The row-level `科目`
                # value therefore appears both as `500100403-...` and
                # `500100403_06.037...-...`.  The auxiliary suffix is a view,
                # not a second posting, so it must not participate in the
                # business-event identity.
                key = (
                    entity_key,
                    posting_date,
                    voucher,
                    normalize_text(account_code),
                    normalize_text(customer),
                    normalize_text(sales_contract),
                    project_base,
                    net_minor,
                    normalize_text(summary),
                )
                other_max = max(
                    (
                        count
                        for source_representation, count in event_rep_counts[key].items()
                        if source_representation != representation
                    ),
                    default=0,
                )
                event_rep_counts[key][representation] += 1
                occurrence = event_rep_counts[key][representation]
                if occurrence <= other_max:
                    duplicate_rows += 1
                    continue
                event_id = "evt_" + sha256_bytes(
                    stable_json([list(key), occurrence])
                )[:24]
                events.append(
                    {
                        "event_id": event_id,
                        "project": project_base,
                        "plane": plane,
                        "category": account_category(account),
                        "amount_cents": net_minor,
                        "posting_date": posting_date,
                        "account_code": account_code,
                        "voucher": voucher,
                        "summary": summary,
                        "supplier": supplier,
                        "entity": entity_key,
                        "source_entity": entity,
                        "source_id": "src_" + metadata["payload_sha256"][:24],
                        "source_member": metadata["container_member"],
                        "sheet": title,
                        "row": row_number,
                        "identity_reason": reason,
                    }
                )
                if reason not in ("EXACT_CONTRACT", "CONTROLLED_SUFFIX_ALIAS"):
                    reviews.append(
                        {
                            "severity": "P2",
                            "type": "LEDGER_IDENTITY_REPAIRED",
                            "project": project_base,
                            "source": metadata["container_member"],
                            "sheet": title,
                            "row": row_number,
                            "raw_contracts": list(tokens),
                            "reason": reason,
                            "action": "已保留证据化别名；财务复核后可冻结",
                        }
                    )
        workbook.close()
    if unallocated_cost_rows:
        reviews.append(
            {
                "severity": "P1",
                "type": "UNALLOCATED_LEDGER_COST_POOL_OPEN",
                "row_count": unallocated_cost_rows,
                "net_cents": unallocated_cost_net_cents,
                "absolute_cents": unallocated_cost_absolute_cents,
                "action": (
                    "存在未归属的5001生产成本；必须以合同/项目/WBS证据逐笔归属"
                    "或证明排除，禁止静默遗漏、按合同额比例分摊或用于反推毛利"
                ),
            }
        )
    reviews = _dedupe_review_rows(reviews)
    return events, reviews, {
        "selected_book_count": len(books),
        "selected_period_end": selected_period_end,
        "minimum_period_end": minimum_period_end,
        "cohort_scan_start": cohort_scan_start.isoformat(),
        "period_ends_by_entity": dict(sorted(period_ends_by_entity.items())),
        "scanned_sheets": scanned_sheets,
        "scanned_rows": scanned_rows,
        "event_count": len(events),
        "semantic_duplicate_rows": duplicate_rows,
        "unresolved_current_customer_rows": unresolved_current_customer_rows,
        "unresolved_current_customer_cents": unresolved_current_customer_cents,
        "unallocated_cost_rows": unallocated_cost_rows,
        "unallocated_cost_net_cents": unallocated_cost_net_cents,
        "unallocated_cost_absolute_cents": unallocated_cost_absolute_cents,
        "unallocated_cost_buckets": [
            {
                "entity": entity,
                "period": period,
                "category": category,
                "reason": reason,
                **amounts,
            }
            for (entity, period, category, reason), amounts in sorted(
                unallocated_cost_buckets.items()
            )
        ],
    }


def _dedupe_review_rows(rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    seen: Set[str] = set()
    for row in rows:
        digest = sha256_bytes(stable_json(dict(row)))
        if digest in seen:
            continue
        seen.add(digest)
        result.append(dict(row))
    return result


def review_summary(reviews: Sequence[Mapping[str, Any]]) -> Dict[str, Any]:
    """Return a public-safe review control without leaking row details.

    P0 blocks verification and runtime publication. P1 observations remain
    outside every formal formula until identity and cost qualification are
    unique, so a run containing them is explicitly provisional. P2 records
    deterministic exclusions, repaired aliases, and non-blocking controls.
    """

    by_severity = {"P0": 0, "P1": 0, "P2": 0}
    by_type: Dict[str, int] = defaultdict(int)
    for review in reviews:
        severity = str(review.get("severity") or "").upper()
        if severity not in by_severity:
            raise ProjectCostError(
                "REVIEW_SEVERITY_INVALID",
                "every review row must use severity P0, P1, or P2",
            )
        review_type = str(review.get("type") or "").strip()
        if not review_type:
            raise ProjectCostError(
                "REVIEW_TYPE_INVALID",
                "every review row must declare a review type",
            )
        by_severity[severity] += 1
        by_type[review_type] += 1
    status = (
        "FAIL"
        if by_severity["P0"]
        else (
            "PASS_WITH_OPEN_REVIEWS"
            if by_severity["P1"]
            else "PASS"
        )
    )
    return {
        "status": status,
        "total_count": sum(by_severity.values()),
        "by_severity": by_severity,
        "by_type": dict(sorted(by_type.items())),
    }


def parse_payment(
    path: Path,
    projects: Sequence[Mapping[str, Any]],
    year: int,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    payload = read_path_bytes(path)
    workbook = open_xlsx_payload(payload)
    indexes = _project_indexes(projects, year)
    selected_sheet = None
    header_row = None
    headers = None
    for sheet in workbook.worksheets:
        try:
            row_number, mapping = locate_header(
                sheet,
                ("付款编号", "审批状态", "付款内容"),
            )
        except ProjectCostError:
            continue
        selected_sheet, header_row, headers = sheet, row_number, mapping
        break
    if selected_sheet is None or header_row is None or headers is None:
        workbook.close()
        raise ProjectCostError("PAYMENT_SCHEMA", "payment approval schema was not recognized")
    events: List[Dict[str, Any]] = []
    reviews: List[Dict[str, Any]] = []
    max_observed_date = ""
    eligible_rows = 0
    for row_number, row in enumerate(
        selected_sheet.iter_rows(min_row=header_row + 1, values_only=True),
        header_row + 1,
    ):
        values = _row_dict(row, headers)
        approval = str(values.get("审批状态") or "").strip()
        payment_status = str(
            values.get("支付状态(系统)") or values.get("支付状态") or ""
        ).strip()
        actual_minor = cents(values.get("实际支付金额"))
        payment_date = iso_date(values.get("支付日期"))
        application_date = iso_date(values.get("申请日期"))
        observed_date = payment_date or application_date
        if observed_date and observed_date > max_observed_date:
            max_observed_date = observed_date
        if approval != "已通过" or payment_status != "全部支付" or not actual_minor or actual_minor <= 0:
            continue
        if observed_date and not observed_date.startswith("%04d-" % year):
            continue
        eligible_rows += 1
        description = " | ".join(
            str(values.get(name) or "")
            for name in ("付款内容", "备注", "收款账户")
        )
        project, reason, tokens = resolve_identity(
            description,
            "",
            indexes,
            allow_text_only=True,
        )
        if project is None:
            normalized_description = normalize_text(description)
            has_customer_hint = any(
                normalize_text(candidate.get("customer")) in normalized_description
                for candidate in indexes["projects"]
                if len(normalize_text(candidate.get("customer"))) >= 4
            )
            has_year_contract = any(contract_year(token) == year for token in tokens)
            if has_customer_hint or has_year_contract:
                reviews.append(
                    {
                        "severity": "P2",
                        "type": "PAYMENT_PROJECT_UNRESOLVED",
                        "source_row": row_number,
                        "payment_id": str(values.get("付款编号") or ""),
                        "amount_cents": actual_minor,
                        "action": "保留为未分配支付观察；不得自动分摊",
                    }
                )
            continue
        project_base = str(project["contract_base"])
        event_key = (
            str(values.get("付款编号") or ""),
            project_base,
            actual_minor,
            observed_date,
        )
        events.append(
            {
                "event_id": "pay_" + sha256_bytes(stable_json(list(event_key)))[:24],
                "project": project_base,
                "plane": "PAYMENT_SYSTEM_PAID_OBSERVED",
                "category": "支付系统已付",
                "amount_cents": actual_minor,
                "posting_date": observed_date,
                "payment_id": str(values.get("付款编号") or ""),
                "summary": str(values.get("付款内容") or ""),
                "source_row": row_number,
                "identity_reason": reason,
            }
        )
    workbook.close()
    return events, _dedupe_review_rows(reviews), {
        "mapped_rows": len(events),
        "eligible_paid_rows": eligible_rows,
        "max_observed_date": max_observed_date,
    }


def parse_approved_cost_detail(
    path: Path,
    projects: Sequence[Mapping[str, Any]],
    status_map: Mapping[str, Mapping[str, Any]],
    year: int,
    as_of: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Read approved line-level project expenses from the business export.

    This source is an occurrence/approval register, not a cash observation.
    Rows are accepted only when the approval state is explicit and the
    contract resolves uniquely to the reporting cohort.  Deposits, advances,
    loans, fines and their reversals remain outside project direct cost.
    """

    payload = read_path_bytes(path)
    workbook = open_xlsx_payload(payload)
    indexes = _project_indexes(projects, year)
    selected_sheet = None
    header_row = None
    headers = None
    required = (
        "费用明细编号",
        "创建时间",
        "审批状态",
        "费用项目",
        "费用说明",
        "金额",
    )
    for sheet in workbook.worksheets:
        try:
            row_number, mapping = locate_header(sheet, required)
        except ProjectCostError:
            continue
        selected_sheet, header_row, headers = sheet, row_number, mapping
        break
    if selected_sheet is None or header_row is None or headers is None:
        workbook.close()
        raise ProjectCostError(
            "APPROVED_COST_DETAIL_SCHEMA",
            "approved project-cost detail schema was not recognized",
        )

    candidates_by_id: Dict[str, Dict[str, Any]] = {}
    conflicting_ids: Set[str] = set()
    reviews: List[Dict[str, Any]] = []
    max_observed_date = ""
    approved_rows = 0
    mapped_rows = 0
    non_cost_rows = 0
    outside_cohort_rows = 0
    unresolved_current_rows = 0
    cutoff = date.fromisoformat(as_of)
    project_bases = {
        str(project["contract_base"]): project for project in projects
    }
    non_cost_markers = (
        "保证金",
        "押金",
        "定金",
        "借款",
        "借支",
        "贷款",
        "回款",
        "收票",
        "罚款",
        "预支",
        "备用金",
    )
    for row_number, row in enumerate(
        selected_sheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ),
        header_row + 1,
    ):
        values = _row_dict(row, headers)
        if normalize_text(values.get("审批状态")) != "已通过":
            continue
        approved_rows += 1
        observed = iso_date(values.get("创建时间"))
        if not observed:
            continue
        observed_day = _as_date(observed)
        if observed_day is None or observed_day > cutoff:
            continue
        if observed > max_observed_date:
            max_observed_date = observed
        amount_minor = cents(values.get("金额"))
        if amount_minor in (None, 0):
            continue
        detail_id = normalize_text(values.get("费用明细编号"))
        if not detail_id:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "APPROVED_COST_DETAIL_ID_MISSING",
                    "source_row": row_number,
                    "action": "已通过成本缺少稳定明细编号；不得进入正式应计",
                }
            )
            continue
        raw_contract_values = tuple(
            str(values.get(name) or "").strip()
            for name in (
                "任务单",
                "关联主合同",
                "关联主合同(费用报销)",
            )
        )
        description_values = tuple(
            str(values.get(name) or "").strip()
            for name in (
                "费用项目",
                "费用说明",
                "成本清单明细",
                "业务类型",
            )
        )
        identity_text = " | ".join(
            value
            for value in raw_contract_values + description_values
            if value
        )
        project, identity_reason, tokens = resolve_identity(
            identity_text,
            "",
            indexes,
            allow_text_only=True,
        )
        if project is None:
            project, identity_reason = resolve_narrative_identity(
                identity_text,
                observed,
                indexes,
                status_map,
            )
        if project is None:
            has_current_contract = any(
                contract_year(token) == year for token in tokens
            )
            if has_current_contract:
                unresolved_current_rows += 1
                reviews.append(
                    {
                        "severity": "P1",
                        "type": "APPROVED_COST_DETAIL_PROJECT_UNRESOLVED",
                        "source_row": row_number,
                        "amount_cents": amount_minor,
                        "action": "当年已通过成本未能唯一归属项目；确认任务单后重跑",
                    }
                )
            else:
                outside_cohort_rows += 1
            continue
        project_base = str(project["contract_base"])
        project_record = project_bases[project_base]
        project_created = _as_date(project_record.get("created_date"))
        if (
            project_created is not None
            and observed_day < project_created - timedelta(days=90)
            and identity_reason
            not in ("EXACT_CONTRACT", "CONTROLLED_SUFFIX_ALIAS")
        ):
            reviews.append(
                {
                    "severity": "P1",
                    "type": "APPROVED_COST_DETAIL_BEFORE_PROJECT_WINDOW",
                    "project": project_base,
                    "source_row": row_number,
                    "amount_cents": amount_minor,
                    "action": "成本日期早于项目建立窗口；不得仅凭相似文本强行归属",
                }
            )
            continue
        cost_text = " | ".join(description_values)
        exclusion = next(
            (marker for marker in non_cost_markers if marker in cost_text),
            None,
        )
        if exclusion:
            non_cost_rows += 1
            reviews.append(
                {
                    "severity": "P2",
                    "type": "APPROVED_COST_DETAIL_NON_COST_EXCLUDED",
                    "project": project_base,
                    "source_row": row_number,
                    "amount_cents": amount_minor,
                    "reason": exclusion,
                    "action": "押金、借支等资金往来及罚款不进入项目直接成本",
                }
            )
            continue
        parent_approval_id = normalize_text(
            values.get("申请编号(费用报销)")
        )
        event = {
            "event_id": "wps_cost_"
            + sha256_bytes(
                stable_json(
                    [
                        detail_id,
                        project_base,
                        amount_minor,
                        observed,
                    ]
                )
            )[:24],
            "project": project_base,
            "plane": "WPS_APPROVED_COST_DETAIL",
            "category": _approved_cost_detail_category(
                str(values.get("费用项目") or ""),
                str(values.get("费用说明") or ""),
                str(values.get("成本清单明细") or ""),
            ),
            "amount_cents": amount_minor,
            "posting_date": observed,
            "summary": str(values.get("费用说明") or "").strip(),
            "source_member": path.name,
            "row": row_number,
            "identity_reason": identity_reason,
            "approval_id": detail_id,
            "parent_approval_id": parent_approval_id or None,
            "approval_authority_verified": True,
            "approval_state_source": "审批状态=已通过",
            "approval_source_kind": "WPS_APPROVED_COST_DETAIL",
        }
        existing = candidates_by_id.get(detail_id)
        if existing is None:
            candidates_by_id[detail_id] = event
            mapped_rows += 1
        elif (
            existing["project"],
            existing["category"],
            existing["amount_cents"],
            existing["posting_date"],
        ) != (
            event["project"],
            event["category"],
            event["amount_cents"],
            event["posting_date"],
        ):
            conflicting_ids.add(detail_id)
    workbook.close()
    for detail_id in sorted(conflicting_ids):
        candidates_by_id.pop(detail_id, None)
        reviews.append(
            {
                "severity": "P1",
                "type": "APPROVED_COST_DETAIL_ID_CONFLICT",
                "detail_id_hash": sha256_bytes(
                    detail_id.encode("utf-8")
                )[:16],
                "action": "同一费用明细编号在项目、金额或日期上冲突；整组排除",
            }
        )
    return (
        [
            candidates_by_id[key]
            for key in sorted(candidates_by_id)
        ],
        _dedupe_review_rows(reviews),
        {
            "approved_rows": approved_rows,
            "mapped_rows": len(candidates_by_id),
            "non_cost_rows": non_cost_rows,
            "outside_cohort_rows": outside_cohort_rows,
            "unresolved_current_rows": unresolved_current_rows,
            "conflicting_detail_id_count": len(conflicting_ids),
            "max_observed_date": max_observed_date,
            "formal_amount_use": bool(candidates_by_id),
        },
    )


def _tax_rate_fraction(value: Any) -> Optional[Decimal]:
    """Normalize an invoice tax rate to a fraction without float arithmetic."""

    if value in (None, "") or isinstance(value, bool):
        return None
    text = str(value).strip().replace("％", "%")
    percent = text.endswith("%")
    if percent:
        text = text[:-1].strip()
    try:
        rate = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if not rate.is_finite():
        return None
    if percent or rate > 1:
        rate /= Decimal(100)
    if rate < 0 or rate > 1:
        return None
    return rate


def _tax_from_gross_cents(gross_cents: int, rate: Decimal) -> int:
    """Return per-line output VAT from a tax-inclusive integer-cent amount."""

    if isinstance(gross_cents, bool) or not isinstance(gross_cents, int):
        raise ProjectCostError(
            "PROJECT_INVOICE_AMOUNT_INVALID",
            "project invoice gross amount must be integer cents",
        )
    if rate < 0 or rate > 1:
        raise ProjectCostError(
            "PROJECT_INVOICE_TAX_RATE_INVALID",
            "project invoice tax rate must be between zero and one",
        )
    return int(
        (
            Decimal(gross_cents)
            * rate
            / (Decimal(1) + rate)
        ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    )


def parse_project_invoice_tax(
    path: Path,
    projects: Sequence[Mapping[str, Any]],
    year: int,
    as_of: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Create project-direct output-VAT facts from approved issued invoices.

    The export is authoritative only for the project-specific invoice line.
    It is not used to spread a company tax return across projects.  Positive
    and red-letter negative invoice allocations are both retained.  Tax is
    rounded once per exported invoice/project line from the tax-inclusive
    amount, using integer cents and ``ROUND_HALF_UP``.
    """

    workbook = open_xlsx_payload(read_path_bytes(path))
    indexes = _project_indexes(projects, year)
    selected_sheet = None
    header_row = None
    headers = None
    required = (
        "发票号码",
        "开票状态",
        "开票日期",
        "本次开票含税金额(元)",
        "税率(%)",
        "开票单位",
        "审批状态",
        "合同编号(合同名称)",
    )
    for sheet in workbook.worksheets:
        try:
            row_number, mapping = locate_header(sheet, required)
        except ProjectCostError:
            continue
        selected_sheet, header_row, headers = sheet, row_number, mapping
        break
    if selected_sheet is None or header_row is None or headers is None:
        workbook.close()
        raise ProjectCostError(
            "PROJECT_INVOICE_SCHEMA",
            "project invoice export schema was not recognized",
        )

    cutoff = date.fromisoformat(as_of)
    candidates: Dict[Tuple[str, str], Dict[str, Any]] = {}
    conflicts: Set[Tuple[str, str]] = set()
    reviews: List[Dict[str, Any]] = []
    approved_issued_rows = 0
    mapped_rows = 0
    zero_rate_rows = 0
    outside_cohort_rows = 0
    unresolved_current_rows = 0
    max_observed_date = ""
    for row_number, row in enumerate(
        selected_sheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ),
        header_row + 1,
    ):
        values = _row_dict(row, headers)
        if (
            normalize_text(values.get("审批状态")) != "已通过"
            or normalize_text(values.get("开票状态")) != "已开票"
        ):
            continue
        approved_issued_rows += 1
        invoice_date = iso_date(values.get("开票日期"))
        invoice_day = _as_date(invoice_date)
        if invoice_day is None or invoice_day > cutoff:
            continue
        if invoice_date > max_observed_date:
            max_observed_date = invoice_date
        invoice_number = normalize_text(values.get("发票号码"))
        if not invoice_number or invoice_number == "待开票":
            reviews.append(
                {
                    "severity": "P1",
                    "type": "PROJECT_INVOICE_NUMBER_MISSING",
                    "source_row": row_number,
                    "action": "已开票记录缺少真实发票号码；不得形成项目税额",
                }
            )
            continue
        gross_cents = cents(values.get("本次开票含税金额(元)"))
        if gross_cents in (None, 0):
            reviews.append(
                {
                    "severity": "P1",
                    "type": "PROJECT_INVOICE_AMOUNT_INVALID",
                    "source_row": row_number,
                    "invoice_number_hash": sha256_bytes(
                        invoice_number.encode("utf-8")
                    )[:16],
                    "action": "已开票记录含税金额为空或为零；不得形成项目税额",
                }
            )
            continue
        rate = _tax_rate_fraction(values.get("税率(%)"))
        if rate is None:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "PROJECT_INVOICE_TAX_RATE_INVALID",
                    "source_row": row_number,
                    "invoice_number_hash": sha256_bytes(
                        invoice_number.encode("utf-8")
                    )[:16],
                    "action": "税率无法确定；不得猜测项目税额",
                }
            )
            continue
        raw_contract = str(
            values.get("合同编号(合同名称)") or ""
        ).strip()
        project, reason, tokens = resolve_identity(
            raw_contract,
            "",
            indexes,
            allow_text_only=False,
        )
        if project is None:
            if any(contract_year(token) == year for token in tokens):
                unresolved_current_rows += 1
                reviews.append(
                    {
                        "severity": "P1",
                        "type": "PROJECT_INVOICE_PROJECT_UNRESOLVED",
                        "source_row": row_number,
                        "invoice_number_hash": sha256_bytes(
                            invoice_number.encode("utf-8")
                        )[:16],
                        "action": "当年开票未能唯一归属报告项目；确认合同编号后重跑",
                    }
                )
            else:
                outside_cohort_rows += 1
            continue
        if reason not in ("EXACT_CONTRACT", "CONTROLLED_SUFFIX_ALIAS"):
            reviews.append(
                {
                    "severity": "P1",
                    "type": "PROJECT_INVOICE_IDENTITY_NOT_EXACT",
                    "source_row": row_number,
                    "invoice_number_hash": sha256_bytes(
                        invoice_number.encode("utf-8")
                    )[:16],
                    "action": "项目税额只接受发票导出中的精确合同号或受控后缀",
                }
            )
            continue
        project_base = str(project["contract_base"])
        tax_cents = _tax_from_gross_cents(gross_cents, rate)
        if tax_cents == 0:
            zero_rate_rows += 1
            continue
        key = (invoice_number, project_base)
        event = {
            "event_id": "invoice_vat_"
            + sha256_bytes(
                stable_json(
                    [
                        invoice_number,
                        project_base,
                        gross_cents,
                        format(rate, "f"),
                        invoice_date,
                    ]
                )
            )[:24],
            "project": project_base,
            "plane": "COST_ACCRUED",
            "category": "项目税费-销项税额",
            "amount_cents": tax_cents,
            "posting_date": invoice_date,
            "summary": "已审批且已开票的项目销项增值税",
            "source_member": path.name,
            "row": row_number,
            "identity_reason": reason,
            "invoice_number_hash": sha256_bytes(
                invoice_number.encode("utf-8")
            )[:16],
            "invoice_gross_cents": gross_cents,
            "tax_rate_fraction": format(rate, "f"),
            "tax_policy": "OUTPUT_VAT_FROM_TAX_INCLUSIVE_PROJECT_INVOICE",
        }
        existing = candidates.get(key)
        if existing is None:
            candidates[key] = event
            mapped_rows += 1
        elif (
            existing["amount_cents"],
            existing["invoice_gross_cents"],
            existing["tax_rate_fraction"],
            existing["posting_date"],
        ) != (
            event["amount_cents"],
            event["invoice_gross_cents"],
            event["tax_rate_fraction"],
            event["posting_date"],
        ):
            conflicts.add(key)
    workbook.close()
    for invoice_number, project_base in sorted(conflicts):
        candidates.pop((invoice_number, project_base), None)
        reviews.append(
            {
                "severity": "P1",
                "type": "PROJECT_INVOICE_ALLOCATION_CONFLICT",
                "project": project_base,
                "invoice_number_hash": sha256_bytes(
                    invoice_number.encode("utf-8")
                )[:16],
                "action": "同一发票与项目的金额、税率或日期冲突；整组排除",
            }
        )
    events = [
        candidates[key]
        for key in sorted(candidates)
    ]
    return events, _dedupe_review_rows(reviews), {
        "approved_issued_rows": approved_issued_rows,
        "mapped_rows": len(events),
        "zero_rate_rows": zero_rate_rows,
        "outside_cohort_rows": outside_cohort_rows,
        "unresolved_current_rows": unresolved_current_rows,
        "conflicting_invoice_project_count": len(conflicts),
        "max_observed_date": max_observed_date,
        "formal_amount_use": bool(events),
        "calculation": (
            "ROUND_HALF_UP(gross_invoice_cents * rate / (1 + rate)) "
            "per invoice/project row"
        ),
        "company_tax_allocation_used": False,
    }


def inspect_dingtalk_archives(
    candidates: Sequence[Path],
    roots: Sequence[Path],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    sources: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    if not candidates:
        return sources, errors
    distinct: List[Tuple[Tuple[int, int, int, int], Path, Dict[str, Any]]] = []
    seen_hashes: Set[str] = set()
    for path in candidates:
        digest = sha256_file(path)
        if digest in seen_hashes:
            continue
        seen_hashes.add(digest)
        try:
            with zipfile.ZipFile(path) as archive:
                audit_archive(archive)
                images = [
                    member
                    for member in archive.infolist()
                    if not member.is_dir()
                    and member.filename.lower().endswith((".png", ".jpg", ".jpeg"))
                ]
                distinct_crc = len({(member.CRC, member.file_size) for member in images})
            metadata = {
                "image_members": len(images),
                "distinct_image_crc_size": distinct_crc,
                "formal_amount_use": False,
                "role": "OCR_CANDIDATE_ONLY",
            }
            distinct.append((_date_score(path), path, metadata))
        except Exception as exc:
            errors.append({"source": relative_to_any(path, roots), "error": str(exc)})
    distinct.sort(key=lambda item: item[0], reverse=True)
    selected_path = distinct[0][1] if distinct else None
    for _, path, metadata in distinct:
        sources.append(
            dict(
                source_record(
                    path,
                    roots,
                    selected=path == selected_path,
                    reason="latest safe screenshot container; OCR remains candidate-only",
                ),
                source_slot="dingtalk_screenshot_candidates",
                logical_metadata=metadata,
            )
        )
    return sources, errors


def _message_date(value: Any) -> Optional[str]:
    if isinstance(value, (int, float)):
        try:
            stamp = float(value) / (1000 if float(value) > 10_000_000_000 else 1)
            return datetime.fromtimestamp(stamp).date().isoformat()
        except (OSError, OverflowError, ValueError):
            return None
    return iso_date(value)


def _amounts_followed_by_yuan(text: str) -> List[int]:
    result: List[int] = []
    for raw, unit in re.findall(r"(?<![\d.])(\d[\d,， ]*(?:\.\d{1,2})?)\s*(万)?元", text):
        minor = cents(raw.replace("，", ",").replace(" ", ""))
        if minor is None:
            continue
        if unit:
            minor *= 10_000
        if minor > 0:
            result.append(minor)
    return list(dict.fromkeys(result))


def _longest_common_substring_length(left: str, right: str) -> int:
    if not left or not right:
        return 0
    previous = [0] * (len(right) + 1)
    best = 0
    for lchar in left:
        current = [0]
        for index, rchar in enumerate(right, 1):
            value = previous[index - 1] + 1 if lchar == rchar else 0
            current.append(value)
            if value > best:
                best = value
        previous = current
    return best


def _narrative_category(text: str) -> str:
    mappings = (
        (("工资", "劳务", "人工"), "劳务/人工"),
        (("加工费", "外协"), "外协"),
        (
            (
                "项目税",
                "税费",
                "增值税",
                "附加税",
                "印花税",
                "税款",
            ),
            "项目税费",
        ),
        (("保险", "安责险"), "项目保险"),
        (("设备租赁", "吊车租赁", "机械租赁"), "设备租赁"),
        (("物流", "快递", "寄件", "运费", "货运", "发货"), "物流运杂"),
        (("住宿", "房租", "酒店", "宾馆"), "住宿"),
        (("过路", "停车", "高速费"), "过路停车"),
        (("加油", "油费", "车辆保养"), "车辆油费"),
        (("电费", "水费", "临电", "燃气费"), "燃料及动力"),
        (
            (
                "交通",
                "车费",
                "火车",
                "高铁",
                "飞机",
                "轮渡",
                "出差",
                "返程",
                "回途",
                "回公司",
            ),
            "交通/差旅",
        ),
        (("材料", "采购", "螺栓", "焊条"), "材料"),
        (("生活费", "餐费", "晚餐"), "生活补助"),
        (("信息费",), "信息费"),
    )
    for markers, category in mappings:
        if any(marker in text for marker in markers):
            return category
    return "其他直接成本"


def _approved_cost_detail_category(
    cost_item: str,
    description: str,
    cost_detail: str,
) -> str:
    """Classify a line-level approved project expense from its cost fields."""

    item = normalize_text(cost_item)
    explicit = (
        (("交通费", "差旅费"), "交通/差旅"),
        (("采购费", "材料费"), "材料"),
        (("外协人员工资", "劳务费"), "劳务/人工"),
        (("物流运输费", "运费"), "物流运杂"),
        (("车辆加油费", "车辆维修保养费"), "车辆油费"),
        (("生活费", "生活用品费"), "生活补助"),
        (("住宿费", "房租"), "住宿"),
        (("吊车租赁费", "脚手架租赁费", "设备租赁费"), "设备租赁"),
        (("员工保险费", "项目保险费"), "项目保险"),
        (("外发加工费", "工程外包费"), "外协"),
        (("税费", "项目税费"), "项目税费"),
    )
    for markers, category in explicit:
        if any(marker in item for marker in markers):
            return category
    return _narrative_category(
        " | ".join(
            value
            for value in (cost_item, description, cost_detail)
            if value
        )
    )


def _funding_plan_category(description: str, counterparty: str) -> str:
    """Classify an approved cost without letting project names contaminate it.

    The linked-contract field is identity evidence, not cost-nature evidence.
    In the production export, some customer names contain words such as
    ``材料``.  Mixing that field into the classifier mislabeled a ferry/travel
    reimbursement as material.  Cost nature is therefore derived only from
    the reimbursement description and payee/account text.
    """

    return _narrative_category(
        " | ".join(
            value.strip()
            for value in (description, counterparty)
            if value and value.strip()
        )
    )


def resolve_narrative_identity(
    text: str,
    observed_date: Optional[str],
    indexes: Mapping[str, Any],
    status_map: Mapping[str, Mapping[str, Any]],
) -> Tuple[Optional[Mapping[str, Any]], str]:
    direct, reason, _ = resolve_identity(
        text,
        "",
        indexes,
        allow_text_only=True,
    )
    if direct is not None:
        return direct, reason
    normalized = normalize_text(text)
    exact_name_candidates: List[Tuple[int, Mapping[str, Any]]] = []
    observed = _as_date(observed_date)
    for project in indexes["projects"]:
        project_name = normalize_text(project.get("project_name"))
        if len(project_name) < 6 or project_name not in normalized:
            continue
        _, start, end = _project_window(project, status_map)
        if observed is not None:
            if start and observed < start - timedelta(days=45):
                continue
            if end and observed > end + timedelta(days=90):
                continue
        exact_name_candidates.append((len(project_name), project))
    if exact_name_candidates:
        best_length = max(length for length, _ in exact_name_candidates)
        best = [
            project
            for length, project in exact_name_candidates
            if length == best_length
        ]
        if len(best) == 1:
            return best[0], "NARRATIVE_EXACT_PROJECT_NAME"
    by_customer: Dict[str, List[Mapping[str, Any]]] = defaultdict(list)
    customer_scores: Dict[str, int] = {}
    for project in indexes["projects"]:
        customer = str(project.get("customer") or "")
        key = normalize_text(customer)
        if not key:
            continue
        by_customer[key].append(project)
        core = key
        for generic in (
            "有限责任公司",
            "股份有限公司",
            "有限公司",
            "分公司",
            "集团",
            "科技",
            "实业",
            "材料",
            "化工",
            "水泥",
            "公司",
        ):
            core = core.replace(generic, "")
        score = _longest_common_substring_length(normalized, core)
        meaningful_three = any(
            core[index : index + 3] in normalized
            and core[index : index + 3]
            not in ("内蒙古", "湖北省", "武汉市", "福建省", "青海省", "江西省", "贵州省")
            for index in range(max(0, len(core) - 2))
        )
        if score >= 4 or (score == 3 and meaningful_three):
            customer_scores[key] = max(customer_scores.get(key, 0), score)
    if not customer_scores:
        return None, "NARRATIVE_CUSTOMER_UNRESOLVED"
    best_score = max(customer_scores.values())
    best_customers = [key for key, score in customer_scores.items() if score == best_score]
    if len(best_customers) != 1:
        return None, "NARRATIVE_CUSTOMER_AMBIGUOUS"
    candidates = by_customer[best_customers[0]]
    preaward = any(
        marker in text
        for marker in ("投标", "标书", "勘察", "踏勘", "技术沟通", "现场沟通")
    )
    if observed is not None:
        eligible: List[Mapping[str, Any]] = []
        for candidate in candidates:
            created, _, end = _project_window(candidate, status_map)
            if created and observed < created:
                if not preaward or created - observed > timedelta(days=45):
                    continue
            grace_days = 90 if len(candidates) == 1 else 45
            if end and observed > end + timedelta(days=grace_days):
                continue
            eligible.append(candidate)
        candidates = eligible
    if len(candidates) > 1:
        name_scores: List[Tuple[int, Mapping[str, Any]]] = []
        for candidate in candidates:
            project_name = normalize_text(candidate.get("project_name"))
            customer_name = normalize_text(candidate.get("customer"))
            residual = project_name.replace(customer_name, "")
            score = _longest_common_substring_length(normalized, residual)
            name_scores.append((score, candidate))
        name_scores.sort(key=lambda item: item[0], reverse=True)
        if (
            name_scores
            and name_scores[0][0] >= 5
            and (len(name_scores) == 1 or name_scores[0][0] > name_scores[1][0])
        ):
            return name_scores[0][1], "NARRATIVE_UNIQUE_PROJECT_NAME"
    if len(candidates) == 1:
        return candidates[0], "NARRATIVE_UNIQUE_CUSTOMER_WINDOW"
    return None, "NARRATIVE_PROJECT_AMBIGUOUS"


def parse_dws_approvals(
    candidates: Sequence[Path],
    roots: Sequence[Path],
    projects: Sequence[Mapping[str, Any]],
    status_map: Mapping[str, Mapping[str, Any]],
    year: int,
    as_of: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    indexes = _project_indexes(projects, year)
    events: List[Dict[str, Any]] = []
    funding_plan_events: Dict[str, Dict[str, Any]] = {}
    funding_plan_conflicts: Set[str] = set()
    reviews: List[Dict[str, Any]] = []
    sources: List[Dict[str, Any]] = []
    seen_archives: Set[str] = set()
    seen_messages: Set[str] = set()
    message_count = 0
    approved_count = 0
    funding_plan_member_count = 0
    funding_plan_approved_row_count = 0
    funding_plan_mapped_row_count = 0
    for path in candidates:
        digest = sha256_file(path)
        if digest in seen_archives:
            continue
        seen_archives.add(digest)
        selected_members = 0
        funding_members: List[Any] = []
        try:
            with zipfile.ZipFile(path) as archive:
                audit_archive(archive)
                members = [
                    info
                    for info in archive.infolist()
                    if not info.is_dir()
                    and info.filename.endswith("/付款请示群/chat_records/raw_messages.jsonl")
                ]
                for member in members:
                    selected_members += 1
                    for line_number, line in enumerate(
                        archive.read(member).decode("utf-8-sig", "replace").splitlines(),
                        1,
                    ):
                        try:
                            record = json.loads(line)
                        except json.JSONDecodeError:
                            continue
                        message_count += 1
                        message_id = str(record.get("openMessageId") or "")
                        if message_id and message_id in seen_messages:
                            continue
                        if message_id:
                            seen_messages.add(message_id)
                        reactions = record.get("emotionReplyList") or []
                        reaction_observed = any(
                            str(reaction.get("emoji") or "").upper() in ("OK", "同意", "赞")
                            and bool(reaction.get("replyUsers"))
                            for reaction in reactions
                            if isinstance(reaction, dict)
                        )
                        if not reaction_observed:
                            continue
                        approved_count += 1
                        content = str(record.get("content") or "")
                        observed = _message_date(record.get("createTime"))
                        if not observed or not observed.startswith("%04d-" % year) or observed > as_of:
                            continue
                        amounts = _amounts_followed_by_yuan(content)
                        exclusion = next(
                            (
                                marker
                                for marker in ("保证金", "押金", "借款", "贷款", "回款", "收票")
                                if marker in content
                            ),
                            None,
                        )
                        if exclusion or len(amounts) != 1:
                            if amounts:
                                reviews.append(
                                    {
                                        "severity": "P2",
                                        "type": "DWS_APPROVAL_NOT_COST_QUALIFIED",
                                        "date": observed,
                                        "amounts_cents": amounts,
                                        "reason": exclusion or "MULTIPLE_AMOUNTS",
                                        "source_member": member.filename,
                                        "line": line_number,
                                        "action": "保留批准观察；不进入项目成本公式",
                                    }
                                )
                            continue
                        project, identity_reason = resolve_narrative_identity(
                            content,
                            observed,
                            indexes,
                            status_map,
                        )
                        if project is None:
                            reviews.append(
                                {
                                    "severity": "P2",
                                    "type": "DWS_APPROVAL_OUTSIDE_FORMULA_EXCLUDED",
                                    "date": observed,
                                    "amount_cents": amounts[0],
                                    "identity_reason": identity_reason,
                                    "source_member": member.filename,
                                    "line": line_number,
                                    "action": (
                                        "保留批准观察；只有项目唯一且满足应计资格的记录"
                                        "才进入公式，本行确定留在观察面"
                                    ),
                                }
                            )
                            continue
                        project_base = str(project["contract_base"])
                        key = (message_id or member.filename, line_number, project_base, amounts[0])
                        events.append(
                            {
                                "event_id": "dws_" + sha256_bytes(stable_json(list(key)))[:24],
                                "project": project_base,
                                "plane": "DWS_APPROVAL_REACTION_OBSERVED",
                                "category": _narrative_category(content),
                                "amount_cents": amounts[0],
                                "posting_date": observed,
                                "summary": content,
                                "source_id": "src_" + digest[:24],
                                "source_member": member.filename,
                                "row": line_number,
                                "identity_reason": identity_reason,
                                "approval_authority_verified": False,
                            }
                        )
                funding_members = [
                    info
                    for info in archive.infolist()
                    if not info.is_dir()
                    and info.filename.lower().endswith(".xlsx")
                    and "项目资金计划" in PurePosixPath(info.filename).name
                ]
                for member in funding_members:
                    funding_plan_member_count += 1
                    try:
                        workbook = open_xlsx_payload(archive.read(member))
                        selected_sheet = None
                        header_row = None
                        headers = None
                        for sheet in workbook.worksheets:
                            try:
                                row_number, mapping = locate_header(
                                    sheet,
                                    (
                                        "申请编号",
                                        "关联主合同",
                                        "累计报销金额",
                                        "审批状态",
                                        "报销说明",
                                    ),
                                )
                            except ProjectCostError:
                                continue
                            selected_sheet = sheet
                            header_row = row_number
                            headers = mapping
                            break
                        if (
                            selected_sheet is None
                            or header_row is None
                            or headers is None
                        ):
                            raise ProjectCostError(
                                "DWS_FUNDING_PLAN_SCHEMA",
                                "project funding plan schema was not recognized",
                            )
                        for row_number, row in enumerate(
                            selected_sheet.iter_rows(
                                min_row=header_row + 1,
                                values_only=True,
                            ),
                            header_row + 1,
                        ):
                            values = _row_dict(row, headers)
                            approval_id = normalize_text(
                                values.get("申请编号")
                            )
                            if (
                                not approval_id
                                or normalize_text(values.get("审批状态"))
                                != "已通过"
                            ):
                                continue
                            funding_plan_approved_row_count += 1
                            amount_minor = cents(
                                values.get("累计报销金额")
                            )
                            if amount_minor is None or amount_minor <= 0:
                                continue
                            date_match = re.search(
                                r"(20\d{2})(\d{2})(\d{2})",
                                approval_id,
                            )
                            if not date_match:
                                reviews.append(
                                    {
                                        "severity": "P1",
                                        "type": "DWS_APPROVED_COST_DATE_UNRESOLVED",
                                        "approval_id_hash": sha256_bytes(
                                            approval_id.encode("utf-8")
                                        )[:16],
                                        "action": "审批日期不能从稳定申请编号解析；不得计入正式应计",
                                    }
                                )
                                continue
                            observed = "%s-%s-%s" % date_match.groups()
                            try:
                                date.fromisoformat(observed)
                            except ValueError:
                                continue
                            if (
                                not observed.startswith("%04d-" % year)
                                or observed > as_of
                            ):
                                continue
                            related_contract = str(
                                values.get("关联主合同") or ""
                            ).strip()
                            description = str(
                                values.get("报销说明") or ""
                            ).strip()
                            counterparty = " ".join(
                                str(values.get(name) or "")
                                for name in ("收款单位", "收款账户")
                            )
                            narrative = " | ".join(
                                value
                                for value in (
                                    related_contract,
                                    description,
                                    counterparty,
                                )
                                if value
                            )
                            exclusion = next(
                                (
                                    marker
                                    for marker in (
                                        "保证金",
                                        "押金",
                                        "定金",
                                        "借款",
                                        "贷款",
                                        "回款",
                                        "收票",
                                        "罚款",
                                    )
                                    if marker in narrative
                                ),
                                None,
                            )
                            if exclusion:
                                reviews.append(
                                    {
                                        "severity": "P2",
                                        "type": "DWS_APPROVED_NON_COST_EXCLUDED",
                                        "approval_id_hash": sha256_bytes(
                                            approval_id.encode("utf-8")
                                        )[:16],
                                        "amount_cents": amount_minor,
                                        "reason": exclusion,
                                        "action": "审批事实保留；非成本或尚未发生项目不进入成本公式",
                                    }
                                )
                                continue
                            project, identity_reason = (
                                resolve_narrative_identity(
                                    narrative,
                                    observed,
                                    indexes,
                                    status_map,
                                )
                            )
                            if project is None:
                                if related_contract:
                                    reviews.append(
                                        {
                                            "severity": "P2",
                                            "type": "DWS_APPROVED_COST_PROJECT_UNRESOLVED",
                                            "approval_id_hash": sha256_bytes(
                                                approval_id.encode("utf-8")
                                            )[:16],
                                            "amount_cents": amount_minor,
                                            "identity_reason": identity_reason,
                                            "action": "保留未分配审批成本；禁止关键词强行归项目",
                                        }
                                    )
                                continue
                            project_base = str(project["contract_base"])
                            candidate = {
                                "event_id": "dws_funding_"
                                + sha256_bytes(
                                    stable_json(
                                        [
                                            approval_id,
                                            project_base,
                                            amount_minor,
                                        ]
                                    )
                                )[:24],
                                "project": project_base,
                                "plane": "DWS_APPROVED_COST",
                                "category": _funding_plan_category(
                                    description,
                                    counterparty,
                                ),
                                "amount_cents": amount_minor,
                                "posting_date": observed,
                                "summary": description,
                                "counterparty": counterparty,
                                "source_id": "src_" + digest[:24],
                                "source_member": member.filename,
                                "row": row_number,
                                "identity_reason": identity_reason,
                                "approval_id": approval_id,
                                "approval_authority_verified": True,
                                "approval_state_source": "审批状态=已通过",
                                "payment_status": str(
                                    values.get("支付状态") or ""
                                ).strip(),
                            }
                            existing = funding_plan_events.get(approval_id)
                            if existing is None:
                                funding_plan_events[approval_id] = candidate
                                funding_plan_mapped_row_count += 1
                            elif (
                                existing["project"],
                                existing["category"],
                                existing["amount_cents"],
                                existing["posting_date"],
                            ) != (
                                candidate["project"],
                                candidate["category"],
                                candidate["amount_cents"],
                                candidate["posting_date"],
                            ):
                                funding_plan_conflicts.add(approval_id)
                        workbook.close()
                    except Exception as exc:
                        reviews.append(
                            {
                                "severity": "P1",
                                "type": "DWS_FUNDING_PLAN_REJECTED",
                                "source_member_hash": sha256_bytes(
                                    member.filename.encode("utf-8")
                                )[:16],
                                "detail": "%s: %s"
                                % (type(exc).__name__, str(exc)),
                                "action": "资金计划底表未读通；整批正式发布保持阻断",
                            }
                        )
        except Exception as exc:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "DWS_SOURCE_REJECTED",
                    "source": relative_to_any(path, roots),
                    "detail": str(exc),
                    "action": "修复 DWS 压缩包后重跑",
                }
            )
        sources.append(
            dict(
                source_record(
                    path,
                    roots,
                    selected=selected_members > 0 or bool(funding_members),
                    reason=(
                        "payment-request observations plus explicit approved-cost "
                        "rows from the production funding plan"
                    ),
                ),
                source_slot="dws_payment_approvals",
                logical_metadata={
                    "selected_raw_message_members": selected_members,
                    "selected_funding_plan_members": len(funding_members),
                    "message_count": message_count,
                    "approval_like_reaction_count": approved_count,
                },
            )
        )
    for approval_id in sorted(funding_plan_conflicts):
        funding_plan_events.pop(approval_id, None)
        reviews.append(
            {
                "severity": "P1",
                "type": "DWS_APPROVED_COST_EXPORT_CONFLICT",
                "approval_id_hash": sha256_bytes(
                    approval_id.encode("utf-8")
                )[:16],
                "action": "同一申请编号在导出中金额或项目冲突；禁止选边",
            }
        )
    events.extend(
        funding_plan_events[key]
        for key in sorted(funding_plan_events)
    )
    return events, _dedupe_review_rows(reviews), sources, {
        "message_count": message_count,
        "approval_like_reaction_count": approved_count,
        "funding_plan_member_count": funding_plan_member_count,
        "funding_plan_approved_row_count": funding_plan_approved_row_count,
        "funding_plan_mapped_row_count": len(funding_plan_events),
        "observed_event_count": len(events),
        "formal_amount_use": bool(funding_plan_events),
    }


def _plain_money_cents(value: str) -> Optional[int]:
    clean = value.strip().replace("¥", "").replace("￥", "").replace(",", "").replace("，", "")
    clean = clean.replace(" ", "")
    if not re.fullmatch(r"-?\d+(?:\.\d{1,2})?", clean):
        return None
    return cents(clean)


def parse_ocr_paid_project_costs(
    path: Optional[Path],
    roots: Sequence[Path],
    projects: Sequence[Mapping[str, Any]],
    status_map: Mapping[str, Mapping[str, Any]],
    year: int,
    as_of: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    if path is None:
        return [], [], [], {"provided": False, "qualified_event_count": 0}
    source_path = Path(path)
    if not source_path.is_file() or source_path.is_symlink():
        raise ProjectCostError("OCR_JSONL_INVALID", "OCR JSONL must be a regular file")
    indexes = _project_indexes(projects, year)
    events: List[Dict[str, Any]] = []
    reviews: List[Dict[str, Any]] = []
    seen_occurrences: Set[Tuple[Any, ...]] = set()
    page_count = 0
    candidate_count = 0
    for line_number, raw_line in enumerate(source_path.read_text(encoding="utf-8-sig").splitlines(), 1):
        try:
            record = json.loads(raw_line)
        except json.JSONDecodeError:
            continue
        page_count += 1
        text = str(record.get("text") or "")
        filename = str(record.get("file") or "")
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        page_date_match = re.search(r"(\d{1,2})月(\d{1,2})日", text)
        if not page_date_match:
            continue
        try:
            observed = date(
                year,
                int(page_date_match.group(1)),
                int(page_date_match.group(2)),
            ).isoformat()
        except ValueError:
            continue
        if observed > as_of:
            continue
        for index, line in enumerate(lines):
            if line != "项目成本":
                continue
            candidate_count += 1
            amount_minor = None
            amount_index = None
            for offset in range(1, 4):
                if index + offset >= len(lines):
                    break
                candidate = _plain_money_cents(lines[index + offset])
                if candidate is not None and candidate > 0:
                    amount_minor = candidate
                    amount_index = index + offset
                    break
            if amount_minor is None:
                continue
            description_parts: List[str] = []
            skipped_date = False
            for prior in range(index - 1, max(-1, index - 4), -1):
                value = lines[prior]
                if re.fullmatch(r"\d{1,2}月\d{1,2}日", value):
                    if description_parts or skipped_date:
                        break
                    skipped_date = True
                    continue
                if _plain_money_cents(value) is not None:
                    break
                if value in (
                    "付日常费用",
                    "转出他行或提现",
                    "他行转入或存现",
                    "营业回款",
                    "投标保证金回款",
                ):
                    break
                description_parts.insert(0, value)
            description = " ".join(description_parts).strip()
            if not description:
                continue
            exclusion = next(
                (
                    marker
                    for marker in ("保证金", "押金", "借支", "预支", "贷款", "还借款")
                    if marker in description
                ),
                None,
            )
            project, identity_reason = resolve_narrative_identity(
                description,
                observed,
                indexes,
                status_map,
            )
            if exclusion:
                reviews.append(
                    {
                        "severity": "P2",
                        "type": "OCR_NON_COST_EXCLUDED",
                        "date": observed,
                        "amount_cents": amount_minor,
                        "reason": exclusion,
                        "file": filename,
                        "jsonl_line": line_number,
                        "action": "明确为非成本支付；保留观察并从项目成本公式中排除",
                    }
                )
                continue
            if project is None:
                reviews.append(
                    {
                        "severity": "P2",
                        "type": "OCR_PAID_OUTSIDE_FORMULA_EXCLUDED",
                        "date": observed,
                        "amount_cents": amount_minor,
                        "reason": identity_reason,
                        "file": filename,
                        "jsonl_line": line_number,
                        "action": (
                            "保留支付观察；只有项目唯一且满足应计资格的记录"
                            "才进入公式，本行确定留在观察面"
                        ),
                    }
                )
                continue
            project_base = str(project["contract_base"])
            occurrence_key = (
                filename or "JSONL_LINE_%d" % line_number,
                sha256_bytes(text.encode("utf-8")),
                index,
                amount_index,
            )
            if occurrence_key in seen_occurrences:
                reviews.append(
                    {
                        "severity": "P2",
                        "type": "OCR_SAME_PHYSICAL_OCCURRENCE_DUPLICATE_EXCLUDED",
                        "file": filename,
                        "jsonl_line": line_number,
                        "amount_cents": amount_minor,
                        "action": "同一文件、同一页内容与同一金额位置重复导出；仅排除该物理重复",
                    }
                )
                continue
            seen_occurrences.add(occurrence_key)
            event_key = (
                occurrence_key,
                observed,
                project_base,
                amount_minor,
                normalize_text(description),
            )
            events.append(
                {
                    "event_id": "ocr_" + sha256_bytes(stable_json(event_key))[:24],
                    "project": project_base,
                    "plane": "OCR_PAID_PROJECT_COST_OBSERVED",
                    "category": _narrative_category(description),
                    "amount_cents": amount_minor,
                    "posting_date": observed,
                    "summary": description,
                    "source_id": "src_" + sha256_file(source_path)[:24],
                    "source_member": filename,
                    "row": line_number,
                    "identity_reason": identity_reason,
                    "ocr_amount_line": amount_index,
                }
            )
    sources = [
        dict(
            source_record(
                source_path,
                tuple(roots) + (source_path.parent,),
                selected=True,
                reason="caller-supplied OCR JSONL; qualified project-cost payment rows only",
            ),
            source_slot="dingtalk_ocr_paid_project_cost",
            logical_metadata={
                "page_count": page_count,
                "candidate_project_cost_rows": candidate_count,
                "qualified_event_count": len(events),
                "machine_derived": True,
            },
        )
    ]
    return events, _dedupe_review_rows(reviews), sources, {
        "provided": True,
        "page_count": page_count,
        "candidate_project_cost_rows": candidate_count,
        "qualified_event_count": len(events),
    }


def qualify_cost_accruals(
    ledger_events: Sequence[Mapping[str, Any]],
    approved_events: Sequence[Mapping[str, Any]],
    paid_events: Sequence[Mapping[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Promote approved, not-yet-posted facts into the accrual plane.

    Reconciliation is one-to-one.  An exact or strongly evidenced gross/net
    representation may suppress one approved line against one posted line;
    a nearby row of the same broad category is never enough on its own.
    """

    posted = [
        event
        for event in ledger_events
        if event.get("plane") == "JOB_POSTED_ACTUAL"
    ]
    unallocated_posted = [
        event
        for event in ledger_events
        if event.get("plane") == "UNALLOCATED_LEDGER_COST_POOL"
    ]
    accruals: List[Dict[str, Any]] = []
    reviews: List[Dict[str, Any]] = []
    posting_link_required = 0
    corroborated = 0
    approved_formal = 0
    exact_posting_matches = 0
    fuzzy_posting_matches = 0
    ambiguous_posting_matches = 0
    unallocated_posting_links = 0

    def close_dates(left: Any, right: Any, days: int) -> bool:
        ldate, rdate = _as_date(left), _as_date(right)
        return bool(ldate and rdate and abs((ldate - rdate).days) <= days)

    def category_family(event: Mapping[str, Any]) -> str:
        category = str(event.get("category") or "")
        if category in ("材料",):
            return "MATERIAL"
        if category in ("外协", "劳务/分包"):
            return "SUBCONTRACT"
        if category in ("自有人工过账", "自有人工-工资应计"):
            return "OWN_LABOR"
        if category in ("自有人工-雇主社保医保应计",):
            return "OWN_LABOR_BURDEN"
        if category in ("设备租赁",):
            return "RENTAL"
        if category in ("交通/差旅",):
            return "TRAVEL"
        if category in ("住宿",):
            return "LODGING"
        if category in ("生活补助",):
            return "LIVING"
        if category in ("物流运杂",):
            return "LOGISTICS"
        if category in ("过路停车",):
            return "ROAD_PARKING"
        if category in ("车辆油费",):
            return "VEHICLE"
        if category in ("燃料及动力",):
            return "UTILITIES"
        if category in ("项目税费",):
            return "PROJECT_TAX"
        if category in ("项目税费-销项税额",):
            return "PROJECT_OUTPUT_VAT"
        return normalize_text(category)

    def possible_postings(
        observation: Mapping[str, Any],
        days: int,
    ) -> List[Mapping[str, Any]]:
        family = category_family(observation)
        return [
            row
            for row in posted
            if str(row.get("project")) == str(observation.get("project"))
            and close_dates(
                row.get("posting_date"),
                observation.get("posting_date"),
                days,
            )
            and category_family(row) == family
        ]

    def date_distance(left: Any, right: Any) -> int:
        ldate, rdate = _as_date(left), _as_date(right)
        if ldate is None or rdate is None:
            return 10**9
        return abs((ldate - rdate).days)

    def event_text(event: Mapping[str, Any]) -> str:
        return normalize_text(
            " | ".join(
                str(event.get(key) or "")
                for key in ("summary", "description", "counterparty")
            )
        )

    def semantic_score(
        left: Mapping[str, Any],
        right: Mapping[str, Any],
    ) -> int:
        left_text, right_text = event_text(left), event_text(right)
        shorter = min(len(left_text), len(right_text))
        if shorter == 0:
            return 0
        common = _longest_common_substring_length(left_text, right_text)
        return int(Decimal(common) * Decimal(1000) / Decimal(shorter))

    def posting_match(
        approved: Mapping[str, Any],
        posting: Mapping[str, Any],
    ) -> Optional[Tuple[int, int, int, str]]:
        if (
            str(posting.get("project")) != str(approved.get("project"))
            or category_family(posting) != category_family(approved)
            or not close_dates(
                posting.get("posting_date"),
                approved.get("posting_date"),
                45,
            )
        ):
            return None
        approved_amount = int(approved.get("amount_cents") or 0)
        posted_amount = int(posting.get("amount_cents") or 0)
        if (
            approved_amount == 0
            or posted_amount == 0
            or (approved_amount > 0) != (posted_amount > 0)
        ):
            return None
        distance = date_distance(
            posting.get("posting_date"),
            approved.get("posting_date"),
        )
        similarity = semantic_score(approved, posting)
        if approved_amount == posted_amount:
            return 0, distance, -similarity, "EXACT_AMOUNT"

        # Gross reimbursement values and VAT-exclusive GL values can be two
        # representations of one occurrence.  This weaker relationship is
        # accepted only with strong narrative overlap and a narrow,
        # enumerated amount transformation.
        if similarity < 550:
            return None
        approved_abs, posted_abs = abs(approved_amount), abs(posted_amount)
        if len(event_text(approved)) < 8 or len(event_text(posting)) < 8:
            return None
        near_tolerance = max(300, int(
            (
                Decimal(approved_abs) * Decimal("0.01")
            ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        ))
        if abs(approved_abs - posted_abs) <= near_tolerance:
            return 1, distance, -similarity, "SEMANTIC_NEAR_AMOUNT"
        if approved_abs >= posted_abs:
            for percentage in (1, 3, 6, 9, 13):
                rate = Decimal(percentage) / Decimal(100)
                expected_net = int(
                    (
                        Decimal(approved_abs)
                        / (Decimal(1) + rate)
                    ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                )
                if abs(expected_net - posted_abs) <= 2:
                    return (
                        2,
                        distance,
                        -similarity,
                        "SEMANTIC_GROSS_NET_%d_PERCENT" % percentage,
                    )
        return None

    # Build a deterministic one-to-one bipartite match.  Repeated exact
    # amounts are safe because each physical posting is consumed once.
    posting_edges: List[
        Tuple[int, int, int, str, str, int, int, str]
    ] = []
    fuzzy_by_approved: Dict[
        int,
        List[Tuple[int, int, int, int]],
    ] = defaultdict(list)
    exact_candidate_approved: Set[int] = set()
    for approved_index, approved in enumerate(approved_events):
        if not approved.get("approval_authority_verified"):
            continue
        for posted_index, posting in enumerate(posted):
            candidate = posting_match(approved, posting)
            if candidate is None:
                continue
            rank, distance, negative_similarity, match_kind = candidate
            posting_edges.append(
                (
                    rank,
                    distance,
                    negative_similarity,
                    str(approved.get("event_id") or approved_index),
                    str(posting.get("event_id") or posted_index),
                    approved_index,
                    posted_index,
                    match_kind,
                )
            )
            if rank == 0:
                exact_candidate_approved.add(approved_index)
            else:
                fuzzy_by_approved[approved_index].append(
                    (
                        rank,
                        distance,
                        negative_similarity,
                        posted_index,
                    )
                )
    ambiguous_approved: Set[int] = set()
    for approved_index, candidates in fuzzy_by_approved.items():
        if approved_index in exact_candidate_approved:
            continue
        ordered = sorted(candidates)
        best_core = ordered[0][:3]
        if sum(candidate[:3] == best_core for candidate in ordered) > 1:
            ambiguous_approved.add(approved_index)
    approved_to_posted: Dict[int, Tuple[int, str]] = {}
    used_posted: Set[int] = set()
    for edge in sorted(posting_edges):
        (
            _rank,
            _distance,
            _negative_similarity,
            _approved_id,
            _posted_id,
            approved_index,
            posted_index,
            match_kind,
        ) = edge
        if (
            approved_index in ambiguous_approved
            or approved_index in approved_to_posted
            or posted_index in used_posted
        ):
            continue
        approved_to_posted[approved_index] = (posted_index, match_kind)
        used_posted.add(posted_index)

    # Exact unallocated 5001 candidates are also consumed one-to-one.  Their
    # lack of project identity remains a P1 link blocker; they are never
    # silently assigned to a project.
    unallocated_edges: List[Tuple[int, str, str, int, int]] = []
    for approved_index, approved in enumerate(approved_events):
        if (
            not approved.get("approval_authority_verified")
            or approved_index in approved_to_posted
            or approved_index in ambiguous_approved
        ):
            continue
        family = category_family(approved)
        amount = int(approved.get("amount_cents") or 0)
        for unallocated_index, posting in enumerate(unallocated_posted):
            if (
                int(posting.get("amount_cents") or 0) == amount
                and category_family(posting) == family
                and close_dates(
                    posting.get("posting_date"),
                    approved.get("posting_date"),
                    45,
                )
            ):
                unallocated_edges.append(
                    (
                        date_distance(
                            posting.get("posting_date"),
                            approved.get("posting_date"),
                        ),
                        str(approved.get("event_id") or approved_index),
                        str(posting.get("event_id") or unallocated_index),
                        approved_index,
                        unallocated_index,
                    )
                )
    approved_to_unallocated: Dict[int, int] = {}
    used_unallocated: Set[int] = set()
    for _, _, _, approved_index, unallocated_index in sorted(
        unallocated_edges
    ):
        if (
            approved_index in approved_to_unallocated
            or unallocated_index in used_unallocated
        ):
            continue
        approved_to_unallocated[approved_index] = unallocated_index
        used_unallocated.add(unallocated_index)

    for paid in paid_events:
        posting_candidates = possible_postings(paid, 10)
        matching_approved = next(
            (
                row
                for row in approved_events
                if str(row.get("project")) == str(paid.get("project"))
                and int(row.get("amount_cents") or 0) == int(paid.get("amount_cents") or 0)
                and category_family(row) == category_family(paid)
                and close_dates(row.get("posting_date"), paid.get("posting_date"), 20)
            ),
            None,
        )
        if matching_approved is not None:
            corroborated += 1
        # A bank/payment observation proves cash movement, not cost
        # occurrence.  Earlier revisions promoted an OCR payment to accrued
        # cost whenever no nearby posting was found.  That reverses the
        # evidence direction and can manufacture project cost from payment
        # alone.  Keep the observation and any corroborating links outside the
        # incurred-cost formula; an independently approved event is evaluated
        # in the next loop on its own authority.
        reviews.append(
            {
                "severity": "P2",
                "type": "PAID_COST_OBSERVATION_EXCLUDED_FROM_ACCRUAL",
                "project": paid.get("project"),
                "observation_event_id": paid.get("event_id"),
                "posting_candidate_count": len(posting_candidates),
                "approved_event_linked": matching_approved is not None,
                "amount_cents": paid.get("amount_cents"),
                "action": (
                    "支付只证明现金阶段，不单独证明成本发生；保留观察和链接，"
                    "仅由独立过账或已批准发生事实进入成本公式"
                ),
            }
        )
    for approved_index, approved in enumerate(approved_events):
        posting_candidates = possible_postings(approved, 45)
        if not approved.get("approval_authority_verified"):
            posting_link_required += bool(posting_candidates)
            reviews.append(
                {
                    "severity": "P1",
                    "type": "DWS_APPROVER_AUTHORITY_UNVERIFIED_EXCLUDED",
                    "project": approved.get("project"),
                    "observation_event_id": approved.get("event_id"),
                    "posting_candidate_count": len(posting_candidates),
                    "amount_cents": int(approved["amount_cents"]),
                    "action": (
                        "DWS 反应人权限未建模；该记录只能作为观察，不能独立形成正式应计"
                    ),
                }
            )
            continue
        if approved_index in ambiguous_approved:
            posting_link_required += 1
            ambiguous_posting_matches += 1
            reviews.append(
                {
                    "severity": "P1",
                    "type": "APPROVED_COST_POSTING_MATCH_AMBIGUOUS",
                    "project": approved.get("project"),
                    "observation_event_id": approved.get("event_id"),
                    "candidate_count": len(
                        fuzzy_by_approved.get(approved_index, ())
                    ),
                    "amount_cents": int(approved["amount_cents"]),
                    "action": (
                        "已通过成本存在多个同等强度的毛额/净额过账候选；"
                        "确认申请编号到凭证链接前不重复应计"
                    ),
                }
            )
            continue
        if approved_index in approved_to_posted:
            posted_index, match_kind = approved_to_posted[approved_index]
            if match_kind == "EXACT_AMOUNT":
                exact_posting_matches += 1
            else:
                fuzzy_posting_matches += 1
            reviews.append(
                {
                    "severity": "P2",
                    "type": "APPROVED_COST_POSTING_MATCHED_ONE_TO_ONE",
                    "project": approved.get("project"),
                    "observation_event_id": approved.get("event_id"),
                    "posting_event_id": posted[posted_index].get("event_id"),
                    "match_kind": match_kind,
                    "amount_cents": int(approved["amount_cents"]),
                    "posting_amount_cents": int(
                        posted[posted_index].get("amount_cents") or 0
                    ),
                    "action": "已通过成本与项目过账逐笔匹配；仅保留总账表示一次",
                }
            )
            continue
        if approved_index in approved_to_unallocated:
            posting_link_required += 1
            unallocated_posting_links += 1
            unallocated_index = approved_to_unallocated[approved_index]
            reviews.append(
                {
                    "severity": "P1",
                    "type": "APPROVED_COST_UNALLOCATED_POSTING_LINK_REQUIRED",
                    "project": approved.get("project"),
                    "observation_event_id": approved.get("event_id"),
                    "candidate_count": 1,
                    "posting_event_id": unallocated_posted[
                        unallocated_index
                    ].get("event_id"),
                    "amount_cents": int(approved["amount_cents"]),
                    "action": (
                        "已通过成本存在同额、同类别、近日期的未分配5001过账；"
                        "取得申请编号到凭证/项目辅助核算的稳定链接前不得重复应计，"
                        "也不得直接把该总账行强行归项目"
                    ),
                }
            )
            continue
        key = [
            approved.get("project"),
            approved.get("approval_id"),
            approved.get("amount_cents"),
            approved.get("posting_date"),
        ]
        accruals.append(
            {
                "event_id": "accr_approved_"
                + sha256_bytes(stable_json(key))[:24],
                "project": approved["project"],
                "plane": "COST_ACCRUED",
                "category": approved.get("category"),
                "amount_cents": int(approved["amount_cents"]),
                "posting_date": approved.get("posting_date"),
                "summary": "业务系统审批已通过、尚未见项目成本过账",
                "source_id": approved.get("source_id"),
                "source_member": approved.get("source_member"),
                "identity_reason": approved.get("identity_reason"),
                "evidence_event_ids": [approved.get("event_id")],
                "approval_id": approved.get("approval_id"),
            }
        )
        approved_formal += 1
    return accruals, _dedupe_review_rows(reviews), {
        "qualified_accrual_count": len(accruals),
        "posting_link_required_count": posting_link_required,
        "dws_reaction_paid_observation_link_count": corroborated,
        "dws_approved_cost_formal_count": approved_formal,
        "dws_reaction_formal_amount_use": False,
        "approved_posting_exact_match_count": exact_posting_matches,
        "approved_posting_fuzzy_match_count": fuzzy_posting_matches,
        "approved_posting_ambiguous_count": ambiguous_posting_matches,
        "approved_unallocated_posting_link_count": (
            unallocated_posting_links
        ),
        "one_to_one_posting_reconciliation": True,
    }


def merge_approved_cost_sources(
    dws_events: Sequence[Mapping[str, Any]],
    detail_events: Sequence[Mapping[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Prefer line detail when it exactly reconciles to a DWS parent approval.

    The funding-plan workbook is one row per reimbursement application while
    the cost-detail workbook is one row per expense line.  An application may
    therefore appear in both sources.  Exact parent amount/project agreement
    keeps the richer detail and suppresses the summary.  Any disagreement
    excludes both representations instead of choosing a convenient number.
    """

    detail_by_parent: Dict[str, List[Mapping[str, Any]]] = defaultdict(list)
    for event in detail_events:
        parent = normalize_text(event.get("parent_approval_id"))
        if parent:
            detail_by_parent[parent].append(event)
    dws_by_approval: Dict[str, Mapping[str, Any]] = {
        normalize_text(event.get("approval_id")): event
        for event in dws_events
        if event.get("approval_authority_verified")
        and normalize_text(event.get("approval_id"))
    }
    suppressed_dws: Set[str] = set()
    conflicting_parents: Set[str] = set()
    reviews: List[Dict[str, Any]] = []
    exact_parent_matches = 0
    for parent in sorted(set(detail_by_parent) & set(dws_by_approval)):
        detail_rows = detail_by_parent[parent]
        dws = dws_by_approval[parent]
        detail_projects = {
            str(event.get("project")) for event in detail_rows
        }
        detail_amount = sum(
            int(event.get("amount_cents") or 0)
            for event in detail_rows
        )
        if (
            detail_projects == {str(dws.get("project"))}
            and detail_amount == int(dws.get("amount_cents") or 0)
        ):
            exact_parent_matches += 1
            suppressed_dws.add(parent)
            reviews.append(
                {
                    "severity": "P2",
                    "type": "APPROVED_COST_PARENT_EXACT_DUPLICATE",
                    "parent_id_hash": sha256_bytes(
                        parent.encode("utf-8")
                    )[:16],
                    "detail_row_count": len(detail_rows),
                    "amount_cents": detail_amount,
                    "action": "资金计划汇总与费用明细逐行合计一致；仅保留明细表示一次",
                }
            )
        else:
            conflicting_parents.add(parent)
            reviews.append(
                {
                    "severity": "P1",
                    "type": "APPROVED_COST_PARENT_RECONCILIATION_CONFLICT",
                    "parent_id_hash": sha256_bytes(
                        parent.encode("utf-8")
                    )[:16],
                    "detail_row_count": len(detail_rows),
                    "detail_amount_cents": detail_amount,
                    "summary_amount_cents": int(
                        dws.get("amount_cents") or 0
                    ),
                    "action": "资金计划汇总与费用明细不一致；两种表示均不进入正式应计",
                }
            )
    result: List[Dict[str, Any]] = []
    for event in dws_events:
        approval_id = normalize_text(event.get("approval_id"))
        if approval_id in suppressed_dws or approval_id in conflicting_parents:
            continue
        result.append(dict(event))
    for event in detail_events:
        parent = normalize_text(event.get("parent_approval_id"))
        if parent in conflicting_parents:
            continue
        result.append(dict(event))
    return result, _dedupe_review_rows(reviews), {
        "dws_event_count": len(dws_events),
        "detail_event_count": len(detail_events),
        "exact_parent_match_count": exact_parent_matches,
        "conflicting_parent_count": len(conflicting_parents),
        "combined_event_count": len(result),
    }


def _decimal_units(value: Any) -> Decimal:
    if value in (None, "") or isinstance(value, bool):
        return Decimal(0)
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(0)
    return result if result > 0 else Decimal(0)


def largest_remainder_allocate(
    total_cents: int,
    weights: Mapping[str, Decimal],
) -> Dict[str, int]:
    """Allocate integer cents deterministically with exact conservation."""

    ordered = sorted(
        (str(key), _decimal_units(value))
        for key, value in weights.items()
        if _decimal_units(value) > 0
    )
    control = sum((value for _, value in ordered), Decimal(0))
    if total_cents < 0 or control <= 0:
        raise ProjectCostError(
            "LABOR_ALLOCATION_INPUT",
            "labor allocation requires non-negative cents and positive weights",
        )
    floors: Dict[str, int] = {}
    fractions: List[Tuple[Decimal, str]] = []
    for key, weight in ordered:
        exact = Decimal(total_cents) * weight / control
        floor_value = int(exact)
        floors[key] = floor_value
        fractions.append((exact - Decimal(floor_value), key))
    remaining = total_cents - sum(floors.values())
    for _, key in sorted(fractions, key=lambda item: (-item[0], item[1]))[:remaining]:
        floors[key] += 1
    if sum(floors.values()) != total_cents:
        raise ProjectCostError("LABOR_ALLOCATION_DRIFT", "largest-remainder allocation did not conserve cents")
    return floors


def labor_posted_reconciliation(
    allocated_wage_component_cents: int,
    posted_wage_component_cents: int,
) -> Tuple[int, int]:
    """Return ``(matched, accrual_residual)`` for one project-period.

    Matching by cents prevents a one-cent 5001003 posting from suppressing an
    entire payroll allocation. Negative posted controls are reviewed by the
    caller and are not used to reduce the accrual.
    """

    allocated = int(allocated_wage_component_cents)
    posted = max(0, int(posted_wage_component_cents))
    if allocated < 0:
        raise ProjectCostError(
            "LABOR_ALLOCATION_NEGATIVE",
            "allocated wage component must not be negative",
        )
    matched = min(allocated, posted)
    return matched, allocated - matched


def labor_posted_component_reconciliation(
    allocated_wage_cents: int,
    allocated_burden_cents: int,
    direct_wage_posted_cents: int,
    combined_wage_burden_posted_cents: int,
) -> Dict[str, int]:
    """Reconcile payroll components to project-period GL labor postings.

    ``5001003`` is matched only to the wage component.  A ``5001006`` row is
    eligible only when its summary explicitly identifies a wage/social
    allocation; because that row is combined, its matched amount is split
    deterministically across the remaining wage and employer-burden
    components.  The split is a reconciliation device, not a new accounting
    fact, and total cents remain exact.
    """

    wage = int(allocated_wage_cents)
    burden = int(allocated_burden_cents)
    direct = int(direct_wage_posted_cents)
    combined = int(combined_wage_burden_posted_cents)
    if min(wage, burden, direct, combined) < 0:
        raise ProjectCostError(
            "LABOR_POSTED_COMPONENT_NEGATIVE",
            "labor component reconciliation requires non-negative cents",
        )
    direct_match = min(wage, direct)
    wage_after_direct = wage - direct_match
    remaining_components = {
        "wage": Decimal(wage_after_direct),
        "burden": Decimal(burden),
    }
    combined_capacity = wage_after_direct + burden
    combined_match = min(combined_capacity, combined)
    combined_split = (
        largest_remainder_allocate(
            combined_match,
            remaining_components,
        )
        if combined_match
        else {"wage": 0, "burden": 0}
    )
    wage_accrual = wage_after_direct - combined_split.get("wage", 0)
    burden_accrual = burden - combined_split.get("burden", 0)
    return {
        "direct_wage_matched_cents": direct_match,
        "combined_matched_cents": combined_match,
        "matched_cents": direct_match + combined_match,
        "wage_accrual_cents": wage_accrual,
        "employer_burden_accrual_cents": burden_accrual,
        "direct_wage_posted_excess_cents": direct - direct_match,
        "combined_posted_excess_cents": combined - combined_match,
    }


def _open_payroll_workbook(path: Path, password_env: Optional[str]):
    payload = read_path_bytes(path)
    if zipfile.is_zipfile(io.BytesIO(payload)):
        return open_xlsx_payload(payload, sanitize=False)
    if not password_env:
        raise ProjectCostError(
            "PAYROLL_PASSWORD_REQUIRED",
            "encrypted payroll requires --payroll-password-env; the secret value is never logged",
        )
    password = os.environ.get(password_env)
    if not password:
        raise ProjectCostError(
            "PAYROLL_PASSWORD_UNAVAILABLE",
            "the named payroll password environment variable is not set",
        )
    try:
        import msoffcrypto  # type: ignore
    except ImportError as exc:
        raise ProjectCostError(
            "DEPENDENCY_MISSING",
            "msoffcrypto-tool is required only for encrypted payroll workbooks",
        ) from exc
    try:
        office = msoffcrypto.OfficeFile(io.BytesIO(payload))
        office.load_key(password=password, verify_password=True)
        output = io.BytesIO()
        office.decrypt(output)
        return open_xlsx_payload(output.getvalue(), sanitize=False)
    except Exception as exc:
        raise ProjectCostError("PAYROLL_DECRYPT_FAILED", "payroll workbook could not be decrypted") from exc


def _payroll_period(path: Path, workbook: Any) -> str:
    # Final payroll exports in the supplied source library use Chinese
    # two-digit-year names such as `26-05月份工资表...` and
    # `26-06月份工资表...-26.7.30.xlsx`.  Prefer the token explicitly followed
    # by 月/月份 so the later preparation date (`26.7.30`) cannot be mistaken
    # for the payroll period.
    short_candidates = re.findall(
        r"(?<!\d)(\d{2})[._年-](\d{1,2})月份?",
        path.name,
    )
    for year_text, month_text in short_candidates:
        month = int(month_text)
        if 1 <= month <= 12:
            return "20%s-%02d" % (year_text, month)
    candidates = re.findall(r"(20\d{2})[._-]?(\d{2})", path.name)
    for year_text, month_text in reversed(candidates):
        month = int(month_text)
        if 1 <= month <= 12:
            return "%s-%02d" % (year_text, month)
    for title in workbook.sheetnames:
        match = re.search(r"(?:(20\d{2})年?)?(\d{1,2})月", title)
        if match and match.group(1):
            return "%s-%02d" % (match.group(1), int(match.group(2)))
    raise ProjectCostError(
        "PAYROLL_PERIOD_UNKNOWN",
        "payroll period must be encoded as YYYYMM or YY-MM月份 in the filename",
    )


def _period_end(period: str) -> date:
    year, month = (int(value) for value in period.split("-", 1))
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def _payroll_header(sheet: Any) -> Tuple[int, Dict[str, int]]:
    for row_number in range(1, min(sheet.max_row or 20, 20) + 1):
        try:
            row = next(
                sheet.iter_rows(
                    min_row=row_number,
                    max_row=row_number,
                    values_only=True,
                )
            )
        except StopIteration:
            break
        mapping: Dict[str, int] = {}
        for index, value in enumerate(row):
            key = re.sub(r"\s+", "", str(value or ""))
            if key and key not in mapping:
                mapping[key] = index
        if "姓名" in mapping and "应计工资小计" in mapping and "部门" in mapping:
            return row_number, mapping
    raise ProjectCostError(
        "PAYROLL_HEADER_NOT_FOUND",
        "payroll sheet must expose 姓名/部门/应计工资小计",
    )


def _employment_entity_key(value: Any) -> str:
    """Return a narrow legal-employer key shared by payroll and burden files."""

    text = normalize_text(value)
    if not text:
        return ""
    # Some finalized social-insurance sheets append the payroll month to a
    # legal-employer label in the sheet title.  Remove only a terminal period
    # token; do not keep tenant-specific company aliases in public source.
    text = re.sub(
        r"[（(]?(?:20)?\d{4,6}(?:-\d+)?[）)]?$",
        "",
        text,
    )
    return _company_core(text)


def _compatible_entity_key(
    requested: str,
    candidates: Iterable[str],
) -> Optional[str]:
    """Resolve one legal-employer key by exact or unique containment.

    Final payroll, burden and ledger exports do not always use the same legal
    suffix.  A narrow containment match handles ``short name`` versus
    ``full legal name`` while refusing ambiguous or two-character brand-only
    matches.
    """

    normalized = _employment_entity_key(requested)
    available = sorted(
        {
            _employment_entity_key(candidate)
            for candidate in candidates
            if _employment_entity_key(candidate)
        }
    )
    if not normalized:
        return None
    if normalized in available:
        return normalized
    minimum_length = 2 if len(available) == 1 else 4
    matches = [
        candidate
        for candidate in available
        if min(len(normalized), len(candidate)) >= minimum_length
        and (
            normalized in candidate
            or candidate in normalized
        )
    ]
    return matches[0] if len(matches) == 1 else None


def _compact_payroll_period(value: Any) -> Optional[str]:
    text = normalize_text(value)
    match = re.fullmatch(r"(?:(20)?(\d{2}))[-./年]?(\d{1,2})", text)
    if not match:
        return None
    year = int(match.group(2))
    month = int(match.group(3))
    if not 1 <= month <= 12:
        return None
    return "20%02d-%02d" % (year, month)


def _sheet_declared_period(sheet: Any) -> Optional[str]:
    values: List[str] = [str(sheet.title or "")]
    for row in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row or 2, 2),
        values_only=True,
    ):
        values.extend(str(value or "") for value in row)
    text = " ".join(values)
    candidates = re.findall(r"(20\d{2})[-./年]?(\d{1,2})", text)
    normalized = {
        "%s-%02d" % (year, int(month))
        for year, month in candidates
        if 1 <= int(month) <= 12
    }
    return next(iter(normalized)) if len(normalized) == 1 else None


def _employer_burden_header(
    sheet: Any,
) -> Optional[Tuple[int, Dict[str, int]]]:
    """Locate a two-row 社保/医保 header and its employer-total column."""

    for row_number in range(1, min(sheet.max_row or 8, 8) + 1):
        try:
            parent = list(
                next(
                    sheet.iter_rows(
                        min_row=row_number,
                        max_row=row_number,
                        values_only=True,
                    )
                )
            )
            child = list(
                next(
                    sheet.iter_rows(
                        min_row=row_number + 1,
                        max_row=row_number + 1,
                        values_only=True,
                    )
                )
            )
        except StopIteration:
            continue
        compact = [
            re.sub(r"\s+", "", str(value or ""))
            for value in parent
        ]
        if not all(
            name in compact
            for name in ("序号", "部门", "月份", "姓名", "单位应缴")
        ):
            continue
        employer_start = compact.index("单位应缴")
        try:
            personal_start = compact.index("个人应缴")
        except ValueError:
            continue
        child_compact = [
            re.sub(r"\s+", "", str(value or ""))
            for value in child
        ]
        totals = [
            index
            for index in range(employer_start, min(personal_start, len(child_compact)))
            if child_compact[index] == "合计"
        ]
        if len(totals) != 1:
            continue
        return row_number, {
            "department": compact.index("部门"),
            "employee": compact.index("姓名"),
            "period": compact.index("月份"),
            "employer_total": totals[0],
        }
    return None


def parse_employer_burden_workbooks(
    workbooks: Sequence[Path],
    roots: Sequence[Path],
    *,
    year: int,
    as_of: str,
    password_env: Optional[str] = None,
) -> Tuple[
    Dict[str, Dict[Tuple[str, str], Dict[str, Any]]],
    List[Dict[str, Any]],
    List[Dict[str, Any]],
    Dict[str, Any],
]:
    """Parse employer-paid social/medical components without using deductions."""

    cutoff = date.fromisoformat(as_of)
    records_by_period: Dict[
        str,
        Dict[Tuple[str, str], Dict[str, Any]],
    ] = {}
    sources: List[Dict[str, Any]] = []
    reviews: List[Dict[str, Any]] = []
    period_meta: List[Dict[str, Any]] = []
    seen_periods: Set[str] = set()
    for raw_path in workbooks:
        path = Path(raw_path)
        if path.is_symlink() or not path.is_file():
            raise ProjectCostError(
                "EMPLOYER_BURDEN_SOURCE_INVALID",
                "employer burden source is unavailable or is a symlink",
            )
        workbook = _open_payroll_workbook(path, password_env)
        try:
            period = _payroll_period(path, workbook)
            if int(period[:4]) != year or _period_end(period) > cutoff:
                continue
            if period in seen_periods:
                raise ProjectCostError(
                    "EMPLOYER_BURDEN_PERIOD_DUPLICATE",
                    "provide exactly one finalized employer-burden workbook per period",
                )
            seen_periods.add(period)
            parsed: Dict[Tuple[str, str], Dict[str, Any]] = {}
            duplicate_keys: Set[Tuple[str, str]] = set()
            row_period_overrides = 0
            rejected_sheets = 0
            adjustment_block_candidate_rows = 0
            for sheet in workbook.worksheets:
                located = _employer_burden_header(sheet)
                if located is None:
                    continue
                header_row, indexes = located
                sheet_period = _sheet_declared_period(sheet)
                if sheet_period and sheet_period != period:
                    rejected_sheets += 1
                    reviews.append(
                        {
                            "severity": "P1",
                            "type": "EMPLOYER_BURDEN_SHEET_PERIOD_CONFLICT",
                            "period": period,
                            "source_sheet": sheet.title,
                            "declared_period": sheet_period,
                            "action": "文件月份与工作表月份冲突；该表不进入雇主负担成本",
                        }
                    )
                    continue
                entity_hint = ""
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(header_row, 3),
                    values_only=True,
                ):
                    entity_hint = next(
                        (
                            str(value).strip()
                            for value in row
                            if value not in (None, "")
                            and "有限公司" in str(value)
                        ),
                        entity_hint,
                    )
                entity = _employment_entity_key(entity_hint or sheet.title)
                if not entity:
                    reviews.append(
                        {
                            "severity": "P1",
                            "type": "EMPLOYER_BURDEN_ENTITY_UNRESOLVED",
                            "period": period,
                            "source_sheet": sheet.title,
                            "action": "单位负担表无法解析雇佣主体；该表不进入人工成本",
                        }
                    )
                    continue
                for row in sheet.iter_rows(
                    min_row=header_row + 2,
                    values_only=True,
                ):
                    employee = normalize_text(
                        row[indexes["employee"]]
                        if indexes["employee"] < len(row)
                        else None
                    )
                    amount = cents(
                        row[indexes["employer_total"]]
                        if indexes["employer_total"] < len(row)
                        else None
                    )
                    if not employee or amount is None or amount < 0:
                        continue
                    row_period = _compact_payroll_period(
                        row[indexes["period"]]
                        if indexes["period"] < len(row)
                        else None
                    )
                    # The same worksheet may contain a second, differently
                    # shaped contribution-base adjustment table below the
                    # finalized monthly rows.  It reuses the words
                    # ``单位应缴`` but has no payroll-month column at the
                    # primary index.  Treating that block as monthly employee
                    # burden shifts columns and invents duplicate employees.
                    if row_period is None:
                        if employee not in ("合计", "总计"):
                            adjustment_block_candidate_rows += 1
                        continue
                    if row_period and row_period != period:
                        if sheet_period == period:
                            row_period_overrides += 1
                        else:
                            reviews.append(
                                {
                                    "severity": "P1",
                                    "type": "EMPLOYER_BURDEN_ROW_PERIOD_CONFLICT",
                                    "period": period,
                                    "source_sheet": sheet.title,
                                    "action": "单位负担行月份与文件月份冲突且无一致表头证据；该行已排除",
                                }
                            )
                            continue
                    key = (entity, employee)
                    if key in parsed:
                        duplicate_keys.add(key)
                        continue
                    parsed[key] = {
                        "amount_cents": amount,
                        "department": normalize_text(
                            row[indexes["department"]]
                            if indexes["department"] < len(row)
                            else None
                        ),
                        "source_sheet": sheet.title,
                    }
            for key in duplicate_keys:
                parsed.pop(key, None)
            if duplicate_keys:
                reviews.append(
                    {
                        "severity": "P1",
                        "type": "EMPLOYER_BURDEN_EMPLOYEE_DUPLICATE",
                        "period": period,
                        "duplicate_employee_count": len(duplicate_keys),
                        "action": "同主体同员工单位负担重复；重复人员保持未分配且不输出身份",
                    }
                )
            if row_period_overrides:
                reviews.append(
                    {
                        "severity": "P2",
                        "type": "EMPLOYER_BURDEN_ROW_PERIOD_OVERRIDDEN",
                        "period": period,
                        "row_count": row_period_overrides,
                        "action": "文件名与工作表均明确同一月份，行内复制月份未采用并保留审计记录",
                    }
                )
            if adjustment_block_candidate_rows:
                reviews.append(
                    {
                        "severity": "P1",
                        "type": (
                            "EMPLOYER_BURDEN_ADJUSTMENT_BLOCK_REQUIRES_PERIOD_ALLOCATION"
                        ),
                        "period": period,
                        "candidate_row_count": (
                            adjustment_block_candidate_rows
                        ),
                        "action": (
                            "同一工作表含无月度字段的缴费基数调整块；"
                            "需取得调整所属月份后才能计入项目人工成本"
                        ),
                    }
                )
            if not parsed:
                reviews.append(
                    {
                        "severity": "P1",
                        "type": "EMPLOYER_BURDEN_PERIOD_EMPTY",
                        "period": period,
                        "action": "单位承担社保/医保来源没有可用人员行；该期间人工成本保持阻断",
                    }
                )
            records_by_period[period] = parsed
            source = dict(
                source_record(
                    path,
                    tuple(roots) + (path.parent,),
                    selected=True,
                    reason="caller-supplied finalized employer social/medical burden workbook",
                ),
                source_slot="payroll_and_time.employer_burden",
                logical_metadata={
                    "payroll_period": period,
                    "employee_row_count": len(parsed),
                    "contains_personal_data": True,
                    "personal_data_not_copied_to_output": True,
                },
            )
            sources.append(source)
            period_meta.append(
                {
                    "period": period,
                    "employee_row_count": len(parsed),
                    "row_period_override_count": row_period_overrides,
                    "adjustment_block_candidate_row_count": (
                        adjustment_block_candidate_rows
                    ),
                    "rejected_sheet_count": rejected_sheets,
                    "employer_burden_control_cents": sum(
                        int(record["amount_cents"])
                        for record in parsed.values()
                    ),
                }
            )
        finally:
            workbook.close()
    return records_by_period, sources, _dedupe_review_rows(reviews), {
        "provided": bool(workbooks),
        "periods": period_meta,
        "selected_period_count": len(records_by_period),
        "employer_burden_control_cents": sum(
            int(record["amount_cents"])
            for records in records_by_period.values()
            for record in records.values()
        ),
        "personal_deductions_used": False,
    }


def _attendance_file_rank(path: Path) -> Tuple[int, int, str]:
    name = path.name.lower()
    rank = 4 if "final" in name else 3 if "evening" in name else 2 if "morning" in name else 1
    return rank, path.stat().st_mtime_ns, path.name


def _selected_attendance_files(
    attendance_roots: Sequence[Path],
    period: str,
) -> Dict[str, Path]:
    selected: Dict[str, Path] = {}
    compact_period = period.replace("-", "")
    for root in attendance_roots:
        root_path = Path(root)
        if root_path.is_symlink() or not root_path.is_dir():
            raise ProjectCostError(
                "ATTENDANCE_ROOT_INVALID",
                "attendance root is unavailable or is a symlink",
            )
        for current, directories, filenames in os.walk(str(root_path), followlinks=False):
            directories[:] = sorted(
                name
                for name in directories
                if name not in (".git", "__pycache__", "__MACOSX")
                and not (Path(current) / name).is_symlink()
            )
            for filename in sorted(filenames):
                if not filename.endswith(".raw.jsonl.gz"):
                    continue
                match = re.search(r"(20\d{6})", filename)
                if not match or not match.group(1).startswith(compact_period):
                    continue
                path = Path(current) / filename
                current_choice = selected.get(match.group(1))
                if current_choice is None or _attendance_file_rank(path) > _attendance_file_rank(current_choice):
                    selected[match.group(1)] = path
    return selected


def _official_attendance_sheet_period(sheet: Any) -> Optional[str]:
    for row in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row or 3, 3),
        values_only=True,
    ):
        text = " ".join(str(value or "") for value in row)
        match = re.search(
            r"(20\d{2})年0?(\d{1,2})月份各项目(?:差旅|自有工人工时)",
            text,
        )
        if match and 1 <= int(match.group(2)) <= 12:
            return "%s-%02d" % (match.group(1), int(match.group(2)))
    return None


def _official_attendance_header(
    sheet: Any,
) -> Tuple[int, Dict[str, int], Dict[int, int]]:
    for row_number in range(1, min(sheet.max_row or 8, 8) + 1):
        try:
            row = next(
                sheet.iter_rows(
                    min_row=row_number,
                    max_row=row_number,
                    values_only=True,
                )
            )
        except StopIteration:
            break
        mapping: Dict[str, int] = {}
        days: Dict[int, int] = {}
        for index, value in enumerate(row):
            key = re.sub(r"\s+", "", str(value or ""))
            if key and key not in mapping:
                mapping[key] = index
            try:
                day_number = int(str(value).strip())
            except (TypeError, ValueError):
                continue
            if 1 <= day_number <= 31:
                days[index] = day_number
        if (
            ("任务单号" in mapping or "合同号" in mapping)
            and "姓名" in mapping
            and "费用类别" in mapping
            and days
        ):
            mapping["contract"] = (
                mapping["任务单号"]
                if "任务单号" in mapping
                else mapping["合同号"]
            )
            return row_number, mapping, days
    raise ProjectCostError(
        "OFFICIAL_ATTENDANCE_HEADER_NOT_FOUND",
        "official project attendance sheet must expose 任务单号或合同号/姓名/费用类别 and day columns",
    )


def _official_attendance_assignments(
    attendance_roots: Sequence[Path],
    period: str,
    projects: Sequence[Mapping[str, Any]],
    roots: Sequence[Path],
) -> Tuple[
    Dict[str, Dict[str, Set[str]]],
    List[Dict[str, Any]],
    Dict[str, Any],
    List[Dict[str, Any]],
]:
    """Read payroll-reviewed WPS project-day sheets as the exact time source."""

    candidates: List[Path] = []
    compact_period = period.replace("-", "")
    for root in attendance_roots:
        root_path = Path(root)
        if root_path.is_symlink() or not root_path.is_dir():
            raise ProjectCostError(
                "ATTENDANCE_ROOT_INVALID",
                "attendance root is unavailable or is a symlink",
            )
        for current, directories, filenames in os.walk(
            str(root_path),
            followlinks=False,
        ):
            directories[:] = sorted(
                name
                for name in directories
                if name not in (".git", "__pycache__", "__MACOSX")
                and not (Path(current) / name).is_symlink()
            )
            for filename in sorted(filenames):
                lower_filename = filename.lower()
                governed_alias = bool(
                    re.fullmatch(
                        re.escape(compact_period)
                        + r"(?:[-_][^.]+)?\.xlsx",
                        lower_filename,
                    )
                )
                if (
                    filename.startswith("._")
                    or not lower_filename.endswith(".xlsx")
                    or (
                        "生产部考勤表" not in filename
                        and not governed_alias
                    )
                ):
                    continue
                path = Path(current) / filename
                metadata = path.lstat()
                if stat.S_ISREG(metadata.st_mode) and not stat.S_ISLNK(
                    metadata.st_mode
                ):
                    candidates.append(path)
    assignments_by_source: List[
        Tuple[Path, Dict[str, Dict[str, Set[str]]], int, int]
    ] = []
    reviews: List[Dict[str, Any]] = []
    seen_digests: Set[str] = set()
    for path in sorted(candidates):
        digest = sha256_file(path)
        if digest in seen_digests:
            continue
        seen_digests.add(digest)
        workbook = None
        try:
            workbook = open_xlsx_payload(read_path_bytes(path))
            matched_sheet_count = 0
            source_assignments: Dict[
                str,
                Dict[str, Set[str]],
            ] = defaultdict(lambda: defaultdict(set))
            current_project_days = 0
            for sheet in workbook.worksheets:
                if _official_attendance_sheet_period(sheet) != period:
                    continue
                matched_sheet_count += 1
                header_row, headers, day_columns = _official_attendance_header(
                    sheet
                )
                contract_index = headers["contract"]
                employee_index = headers["姓名"]
                category_index = headers["费用类别"]
                current_project: Optional[str] = None
                current_employee = ""
                project_bases = {
                    str(project["contract_base"]) for project in projects
                }
                for row in sheet.iter_rows(
                    min_row=header_row + 1,
                    values_only=True,
                ):
                    raw_contract = (
                        row[contract_index]
                        if contract_index < len(row)
                        else None
                    )
                    if raw_contract not in (None, ""):
                        current_employee = ""
                        candidate_base = contract_base(raw_contract)
                        current_project = (
                            candidate_base
                            if candidate_base in project_bases
                            else None
                        )
                    raw_employee = (
                        row[employee_index]
                        if employee_index < len(row)
                        else None
                    )
                    if raw_employee not in (None, ""):
                        current_employee = normalize_text(raw_employee)
                    category = normalize_text(
                        row[category_index]
                        if category_index < len(row)
                        else None
                    )
                    if (
                        current_project is None
                        or not current_employee
                        or category
                        not in ("餐费", "餐补", "生活费", "住宿", "生", "住")
                    ):
                        continue
                    for column_index, day_number in day_columns.items():
                        value = (
                            row[column_index]
                            if column_index < len(row)
                            else None
                        )
                        if _decimal_units(value) <= 0:
                            continue
                        try:
                            work_date = date.fromisoformat(
                                "%s-%02d" % (period, day_number)
                            )
                        except ValueError:
                            continue
                        compact_day = work_date.strftime("%Y%m%d")
                        before = len(
                            source_assignments[current_employee][
                                current_project
                            ]
                        )
                        source_assignments[current_employee][
                            current_project
                        ].add(compact_day)
                        if (
                            len(
                                source_assignments[current_employee][
                                    current_project
                                ]
                            )
                            > before
                        ):
                            current_project_days += 1
            if matched_sheet_count:
                assignments_by_source.append(
                    (
                        path,
                        source_assignments,
                        matched_sheet_count,
                        current_project_days,
                    )
                )
        except Exception as exc:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "OFFICIAL_ATTENDANCE_SOURCE_REJECTED",
                    "source": relative_to_any(
                        path,
                        tuple(roots) + tuple(attendance_roots),
                    ),
                    "detail": "%s: %s"
                    % (type(exc).__name__, str(exc)),
                    "action": "工资复核项目日底表未读通；该期间人工分配保持阻断",
                }
            )
        finally:
            if workbook is not None:
                workbook.close()
    if len(assignments_by_source) > 1:
        reviews.append(
            {
                "severity": "P1",
                "type": "OFFICIAL_ATTENDANCE_PERIOD_DUPLICATE",
                "period": period,
                "source_count": len(assignments_by_source),
                "action": "同期间存在多份不同字节的核定生产考勤；需明确唯一终稿后重跑",
            }
        )
        return {}, [], {
            "period": period,
            "selected_workbook_count": 0,
            "matched_sheet_count": 0,
            "mapped_employee_project_days": 0,
        }, reviews
    if not assignments_by_source:
        return {}, [], {
            "period": period,
            "selected_workbook_count": 0,
            "matched_sheet_count": 0,
            "mapped_employee_project_days": 0,
        }, reviews
    path, assignments, sheet_count, project_days = assignments_by_source[0]
    source = dict(
        source_record(
            path,
            tuple(roots) + tuple(attendance_roots),
            selected=True,
            reason="unique payroll-reviewed production project attendance workbook for the period",
        ),
        source_slot="payroll_and_time.official_project_attendance",
        logical_metadata={
            "payroll_period": period,
            "matched_sheet_count": sheet_count,
            "contains_personal_data": True,
            "personal_data_not_copied_to_output": True,
        },
    )
    return assignments, [source], {
        "period": period,
        "selected_workbook_count": 1,
        "matched_sheet_count": sheet_count,
        "mapped_employee_project_days": project_days,
    }, reviews


def _company_core(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(
        r"(有限责任公司|股份有限公司|集团有限公司|集团公司|有限公司|公司)$",
        "",
        text,
    )
    return text


def _site_segments(location: str) -> Set[str]:
    parts = re.split(
        r"壮族自治区|回族自治区|维吾尔自治区|特别行政区|自治区|自治州|"
        r"省|市|区|县|旗|镇|街道",
        normalize_text(location),
    )
    return {
        part
        for part in parts
        if 2 <= len(part) <= 8
        and re.fullmatch(r"[\u4e00-\u9fff]+", part)
        and part not in NON_SITE_ADMIN_SEGMENTS
    }


def _resolve_attendance_project(
    location: str,
    work_date: date,
    projects: Sequence[Mapping[str, Any]],
    status_map: Mapping[str, Mapping[str, Any]],
) -> Tuple[Optional[str], str]:
    segments = _site_segments(location)
    normalized_location = normalize_text(location)
    scored: List[Tuple[int, str]] = []
    for project in projects:
        base = str(project["contract_base"])
        status = status_map.get(base, {})
        construction = normalize_text(
            status.get("construction_status")
            or project.get("construction_status_master")
        )
        if any(token in construction for token in ("待入场", "未开工", "已取消", "作废")):
            continue
        start = _as_date(status.get("start_date") or project.get("created_date"))
        if start and work_date < start:
            continue
        project_text = normalize_text(
            "%s %s" % (project.get("project_name") or "", project.get("customer") or "")
        )
        segment_matches = sorted(segment for segment in segments if segment in project_text)
        score = sum(3 + min(len(segment), 4) for segment in segment_matches)
        customer_core = _company_core(project.get("customer"))
        common = (
            _longest_common_substring_length(customer_core, normalized_location)
            if len(customer_core) >= 4
            else 0
        )
        if common >= 4:
            score += 8 + common
        if score:
            scored.append((score, base))
    if not scored:
        return None, "NO_UNIQUE_PROJECT_SITE"
    best = max(score for score, _ in scored)
    winners = sorted(base for score, base in scored if score == best)
    if len(winners) == 1:
        return winners[0], "UNIQUE_PROJECT_SITE_AND_ACTIVE_WINDOW"
    starts: List[Tuple[date, str]] = []
    by_base = {str(project["contract_base"]): project for project in projects}
    for base in winners:
        project = by_base[base]
        status = status_map.get(base, {})
        start = _as_date(status.get("start_date") or project.get("created_date"))
        if start:
            starts.append((start, base))
    if starts:
        latest = max(start for start, _ in starts)
        latest_winners = sorted(base for start, base in starts if start == latest)
        if len(latest_winners) == 1:
            return latest_winners[0], "SAME_SITE_LATEST_EFFECTIVE_PROJECT"
    return None, "AMBIGUOUS_PROJECT_SITE"


def _attendance_assignments(
    attendance_roots: Sequence[Path],
    period: str,
    projects: Sequence[Mapping[str, Any]],
    status_map: Mapping[str, Mapping[str, Any]],
    roots: Sequence[Path],
) -> Tuple[Dict[str, Dict[str, Set[str]]], List[Dict[str, Any]], Dict[str, Any], List[Dict[str, Any]]]:
    (
        official_assignments,
        official_sources,
        official_meta,
        official_reviews,
    ) = _official_attendance_assignments(
        attendance_roots,
        period,
        projects,
        roots,
    )
    selected = _selected_attendance_files(attendance_roots, period)
    assignments: Dict[str, Dict[str, Set[str]]] = defaultdict(
        lambda: defaultdict(set)
    )
    for employee, project_rows in official_assignments.items():
        for project, days in project_rows.items():
            assignments[employee][project].update(days)
    sources: List[Dict[str, Any]] = list(official_sources)
    reviews: List[Dict[str, Any]] = list(official_reviews)
    employee_days_with_location = 0
    mapped_employee_days = 0
    ambiguous_employee_days = 0
    official_location_conflicts = 0
    for compact_day, path in sorted(selected.items()):
        try:
            work_date = datetime.strptime(compact_day, "%Y%m%d").date()
            metadata = path.lstat()
            if not stat.S_ISREG(metadata.st_mode) or stat.S_ISLNK(metadata.st_mode):
                raise ProjectCostError("ATTENDANCE_SOURCE_UNSAFE", "attendance input is not a regular file")
            with gzip.open(path, "rt", encoding="utf-8") as handle:
                for line in handle:
                    record = json.loads(line)
                    if record.get("type") != "employee_attendance":
                        continue
                    employee = normalize_text((record.get("member") or {}).get("name"))
                    payload = (
                        (((record.get("record") or {}).get("final") or {}).get("payload") or {})
                        .get("result")
                        or {}
                    )
                    locations = {
                        normalize_text(item.get("locationText"))
                        for item in payload.get("recordList") or []
                        if normalize_text(item.get("locationText"))
                    }
                    if not employee or not locations:
                        continue
                    employee_days_with_location += 1
                    resolved = [
                        _resolve_attendance_project(location, work_date, projects, status_map)
                        for location in locations
                    ]
                    resolutions = {project for project, _ in resolved}
                    resolutions.discard(None)
                    if len(resolutions) == 1:
                        project = next(iter(resolutions))
                        official_day_projects = {
                            candidate_project
                            for candidate_project, days in assignments.get(
                                employee,
                                {},
                            ).items()
                            if compact_day in days
                        }
                        if official_day_projects:
                            if project not in official_day_projects:
                                official_location_conflicts += 1
                        else:
                            assignments[employee][project].add(compact_day)
                            mapped_employee_days += 1
                    elif len(resolutions) > 1 or any(
                        reason == "AMBIGUOUS_PROJECT_SITE" for _, reason in resolved
                    ):
                        ambiguous_employee_days += 1
            sources.append(
                dict(
                    source_record(
                        path,
                        tuple(roots) + tuple(attendance_roots),
                        selected=True,
                        reason="latest official DingTalk attendance capture for the work date",
                    ),
                    source_slot="payroll_and_time.attendance",
                    logical_metadata={
                        "payroll_period": period,
                        "work_date": work_date.isoformat(),
                        "contains_personal_data": True,
                        "personal_data_not_copied_to_output": True,
                    },
                )
            )
        except Exception as exc:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "ATTENDANCE_SOURCE_REJECTED",
                    "source": relative_to_any(path, tuple(roots) + tuple(attendance_roots)),
                    "detail": "%s: %s" % (type(exc).__name__, str(exc)),
                    "action": "该日不进入人工分摊；其余日期继续计算",
                }
            )
    if official_location_conflicts:
        reviews.append(
            {
                "severity": "P2",
                "type": "ATTENDANCE_LOCATION_DIFFERS_FROM_OFFICIAL_PROJECT_DAY",
                "period": period,
                "employee_day_count": official_location_conflicts,
                "action": "核定生产考勤项目日优先；钉钉地理位置差异仅保留为观察，不改写核定项目",
            }
        )
    return assignments, sources, {
        "period": period,
        "selected_day_count": len(selected),
        "employee_days_with_location": employee_days_with_location,
        "mapped_employee_days": mapped_employee_days,
        "ambiguous_employee_days": ambiguous_employee_days,
        "official_project_attendance": official_meta,
        "official_location_conflicts": official_location_conflicts,
    }, reviews


def parse_payroll_and_attendance(
    payroll_workbooks: Sequence[Path],
    attendance_roots: Sequence[Path],
    employer_burden_workbooks: Sequence[Path],
    projects: Sequence[Mapping[str, Any]],
    status_map: Mapping[str, Mapping[str, Any]],
    ledger_events: Sequence[Mapping[str, Any]],
    roots: Sequence[Path],
    *,
    year: int,
    as_of: str,
    password_env: Optional[str] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Allocate wage and employer burden using the same approved project days."""

    (
        burden_records_by_period,
        burden_sources,
        burden_reviews,
        burden_meta,
    ) = parse_employer_burden_workbooks(
        employer_burden_workbooks,
        roots,
        year=year,
        as_of=as_of,
        password_env=password_env,
    )
    if not payroll_workbooks:
        reviews = list(burden_reviews)
        if burden_meta["employer_burden_control_cents"]:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "EMPLOYER_BURDEN_WITHOUT_PAYROLL",
                    "period_count": burden_meta["selected_period_count"],
                    "action": "单位负担来源缺少同月工资控制表；不得单独按姓名或比例分配",
                }
            )
        return [], _dedupe_review_rows(reviews), burden_sources, {
            "provided": False,
            "periods": [],
            "wage_component_control_cents": 0,
            "employer_burden_control_cents": 0,
            "fully_loaded_labor_control_cents": 0,
            "wage_allocated_accrual_cents": 0,
            "employer_burden_allocated_accrual_cents": 0,
            "allocated_accrual_cents": 0,
            "already_posted_cents": 0,
            "wage_unallocated_cents": 0,
            "employer_burden_unallocated_cents": 0,
            "unallocated_cents": 0,
            "conservation_delta_cents": 0,
            "employer_burden": burden_meta,
        }
    cutoff = date.fromisoformat(as_of)
    period_books: Dict[str, Tuple[Path, Any]] = {}
    for raw_path in payroll_workbooks:
        path = Path(raw_path)
        if path.is_symlink() or not path.is_file():
            raise ProjectCostError("PAYROLL_SOURCE_INVALID", "payroll source is unavailable or is a symlink")
        workbook = _open_payroll_workbook(path, password_env)
        period = _payroll_period(path, workbook)
        if int(period[:4]) != year or _period_end(period) > cutoff:
            continue
        if period in period_books:
            raise ProjectCostError(
                "PAYROLL_PERIOD_DUPLICATE",
                "provide exactly one finalized payroll workbook per payroll period",
            )
        period_books[period] = (path, workbook)
    posted_wage_cents: Dict[Tuple[str, str, str], int] = defaultdict(int)
    posted_combined_labor_cents: Dict[
        Tuple[str, str, str],
        int,
    ] = defaultdict(int)
    for event in ledger_events:
        if event.get("plane") != "JOB_POSTED_ACTUAL":
            continue
        scope = (
            str(event.get("project")),
            str(event.get("posting_date") or "")[:7],
            _employment_entity_key(event.get("entity")),
        )
        account_code = str(event.get("account_code") or "")
        if account_code.startswith("5001003"):
            posted_wage_cents[scope] += int(
                event.get("amount_cents") or 0
            )
        elif (
            account_code.startswith("5001006")
            and any(
                marker in normalize_text(event.get("summary"))
                for marker in ("工资", "社保", "人工成本")
            )
        ):
            posted_combined_labor_cents[scope] += int(
                event.get("amount_cents") or 0
            )
    events: List[Dict[str, Any]] = []
    reviews: List[Dict[str, Any]] = list(burden_reviews)
    sources: List[Dict[str, Any]] = list(burden_sources)
    period_meta: List[Dict[str, Any]] = []
    grand_wage_control = 0
    grand_burden_control = 0
    grand_wage_accrual = 0
    grand_burden_accrual = 0
    grand_posted = 0
    grand_wage_unallocated = 0
    grand_burden_unallocated = 0
    burden_source_by_period = {
        str((source.get("logical_metadata") or {}).get("payroll_period")): source
        for source in burden_sources
    }
    for period, (path, workbook) in sorted(period_books.items()):
        sheets = [sheet for sheet in workbook.worksheets if "分部门" in sheet.title]
        if len(sheets) != 1:
            raise ProjectCostError(
                "PAYROLL_DETAIL_SHEET",
                "payroll workbook must contain exactly one 分部门 sheet",
            )
        sheet = sheets[0]
        header_row, headers = _payroll_header(sheet)
        name_index = headers["姓名"]
        department_index = headers.get("实际部门", headers["部门"])
        company_index = headers.get("公司")
        gross_index = headers["应计工资小计"]
        outside_index = headers.get("实出勤天-厂外")
        inside_index = headers.get("实出勤天-厂内")
        payroll_rows: List[Dict[str, Any]] = []
        name_counts: Dict[str, int] = defaultdict(int)
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            employee = normalize_text(row[name_index] if name_index < len(row) else None)
            amount = cents(row[gross_index] if gross_index < len(row) else None)
            if not employee or amount is None or amount <= 0:
                continue
            department = normalize_text(
                row[department_index] if department_index < len(row) else None
            )
            approved_days = Decimal(0)
            if outside_index is not None and outside_index < len(row):
                approved_days += _decimal_units(row[outside_index])
            if inside_index is not None and inside_index < len(row):
                approved_days += _decimal_units(row[inside_index])
            payroll_rows.append(
                {
                    "employee": employee,
                    "entity": _employment_entity_key(
                        row[company_index]
                        if company_index is not None and company_index < len(row)
                        else None
                    ),
                    "department": department,
                    "gross_cents": amount,
                    "approved_days": approved_days,
                }
            )
            name_counts[employee] += 1
        assignments, attendance_sources, attendance_meta, attendance_reviews = _attendance_assignments(
            attendance_roots,
            period,
            projects,
            status_map,
            roots,
        )
        payroll_source = dict(
            source_record(
                path,
                tuple(roots) + (path.parent,),
                selected=True,
                reason="caller-supplied finalized payroll workbook; wage component only",
            ),
            source_slot="payroll_and_time.payroll",
            logical_metadata={
                "payroll_period": period,
                "employee_row_count": len(payroll_rows),
                "contains_personal_data": True,
                "personal_data_not_copied_to_output": True,
            },
        )
        sources.append(payroll_source)
        sources.extend(attendance_sources)
        reviews.extend(attendance_reviews)
        wage_allocated_by_scope: Dict[Tuple[str, str], int] = defaultdict(
            int
        )
        burden_allocated_by_scope: Dict[
            Tuple[str, str],
            int,
        ] = defaultdict(int)
        allocated_days_by_project: Dict[str, Decimal] = defaultdict(Decimal)
        allocated_people_by_project: Dict[str, Set[str]] = defaultdict(set)
        wage_control = 0
        burden_control = 0
        wage_accrued = 0
        burden_accrued = 0
        already_posted = 0
        wage_unallocated = 0
        burden_unallocated = 0
        duplicate_name_rows = 0
        invalid_time_rows = 0
        non_direct_excluded = 0
        non_direct_burden_excluded = 0
        missing_burden_rows = 0
        unmatched_direct_burden_rows = 0
        burden_records = burden_records_by_period.get(period, {})
        burden_keys_by_employee: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
        for burden_key in burden_records:
            burden_keys_by_employee[burden_key[1]].append(burden_key)
        consumed_burden_keys: Set[Tuple[str, str]] = set()
        for payroll_row in payroll_rows:
            employee = str(payroll_row["employee"])
            amount = int(payroll_row["gross_cents"])
            entity = str(payroll_row["entity"])
            department = str(payroll_row["department"])
            approved_days = payroll_row["approved_days"]
            project_days = {
                project: Decimal(len(days))
                for project, days in assignments.get(employee, {}).items()
                if days
            }
            direct_scope = any(
                token in department for token in DIRECT_LABOR_DEPARTMENT_TOKENS
            ) or bool(project_days)
            burden_key: Optional[Tuple[str, str]] = None
            employee_burden_keys = burden_keys_by_employee.get(
                employee,
                (),
            )
            if entity:
                matched_entity = _compatible_entity_key(
                    entity,
                    (key[0] for key in employee_burden_keys),
                )
                if (
                    matched_entity is not None
                    and (matched_entity, employee) in burden_records
                ):
                    burden_key = (matched_entity, employee)
            elif len(employee_burden_keys) == 1:
                burden_key = employee_burden_keys[0]
            burden_record = (
                burden_records.get(burden_key)
                if burden_key is not None and burden_key not in consumed_burden_keys
                else None
            )
            burden_amount = (
                int(burden_record["amount_cents"])
                if burden_record is not None
                else None
            )
            if burden_key is not None and burden_record is not None:
                consumed_burden_keys.add(burden_key)
            if not direct_scope:
                non_direct_excluded += amount
                if burden_amount is not None:
                    non_direct_burden_excluded += burden_amount
                continue
            wage_control += amount
            if burden_amount is not None:
                burden_control += burden_amount
            if name_counts[employee] != 1:
                duplicate_name_rows += 1
                wage_unallocated += amount
                if burden_amount is not None:
                    burden_unallocated += burden_amount
                continue
            if burden_amount is None:
                missing_burden_rows += 1
            mapped_days = sum(project_days.values(), Decimal(0))
            if approved_days <= 0 or mapped_days > approved_days:
                invalid_time_rows += 1
                wage_unallocated += amount
                if burden_amount is not None:
                    burden_unallocated += burden_amount
                continue
            weights = dict(project_days)
            remainder_days = approved_days - mapped_days
            if remainder_days > 0:
                weights["__UNALLOCATED__"] = remainder_days
            if not weights:
                wage_unallocated += amount
                if burden_amount is not None:
                    burden_unallocated += burden_amount
                continue
            wage_allocation = largest_remainder_allocate(amount, weights)
            wage_unallocated += wage_allocation.pop("__UNALLOCATED__", 0)
            burden_allocation = (
                largest_remainder_allocate(burden_amount, weights)
                if burden_amount is not None
                else {}
            )
            burden_unallocated += burden_allocation.pop("__UNALLOCATED__", 0)
            allocation_entity = (
                burden_key[0] if burden_key is not None else entity
            )
            employee_token = "emp_" + sha256_bytes(stable_json([period, employee]))[:16]
            for project, project_amount in wage_allocation.items():
                if project_amount <= 0:
                    continue
                wage_allocated_by_scope[
                    (project, allocation_entity)
                ] += project_amount
                allocated_days_by_project[project] += project_days[project]
                allocated_people_by_project[project].add(employee_token)
            for project, project_amount in burden_allocation.items():
                if project_amount > 0:
                    burden_allocated_by_scope[
                        (project, allocation_entity)
                    ] += project_amount
        for burden_key, burden_record in burden_records.items():
            if burden_key in consumed_burden_keys:
                continue
            employee = burden_key[1]
            department = str(burden_record.get("department") or "")
            amount = int(burden_record["amount_cents"])
            direct_scope = any(
                token in department for token in DIRECT_LABOR_DEPARTMENT_TOKENS
            ) or bool(assignments.get(employee))
            if direct_scope:
                burden_control += amount
                burden_unallocated += amount
                unmatched_direct_burden_rows += 1
            else:
                non_direct_burden_excluded += amount
        wage_residual_by_project: Dict[str, int] = defaultdict(int)
        burden_residual_by_project: Dict[str, int] = defaultdict(int)
        allocation_scopes = sorted(
            set(wage_allocated_by_scope)
            | set(burden_allocated_by_scope)
        )
        for project, allocation_entity in allocation_scopes:
            wage_amount = wage_allocated_by_scope.get(
                (project, allocation_entity),
                0,
            )
            burden_amount = burden_allocated_by_scope.get(
                (project, allocation_entity),
                0,
            )

            def resolved_posted_amount(
                values: Mapping[Tuple[str, str, str], int],
            ) -> Tuple[int, Optional[str], int]:
                candidate_entities = {
                    key[2]
                    for key, amount in values.items()
                    if key[0] == project
                    and key[1] == period
                    and amount
                    and key[2]
                }
                matched_entity = _compatible_entity_key(
                    allocation_entity,
                    candidate_entities,
                )
                amount = (
                    int(
                        values.get(
                            (project, period, matched_entity),
                            0,
                        )
                    )
                    if matched_entity is not None
                    else 0
                )
                return amount, matched_entity, len(candidate_entities)

            direct_posted, direct_entity, direct_candidate_count = (
                resolved_posted_amount(posted_wage_cents)
            )
            combined_posted, combined_entity, combined_candidate_count = (
                resolved_posted_amount(posted_combined_labor_cents)
            )
            if direct_posted < 0 or combined_posted < 0:
                reviews.append(
                    {
                        "severity": "P1",
                        "type": "LABOR_POSTED_CONTROL_NEGATIVE",
                        "project": project,
                        "period": period,
                        "direct_wage_posted_cents": direct_posted,
                        "combined_wage_burden_posted_cents": (
                            combined_posted
                        ),
                        "action": "人工过账控制额为负；未用于压减工资社保应计，需核对冲销关系",
                    }
                )
                direct_posted = max(0, direct_posted)
                combined_posted = max(0, combined_posted)
            reconciliation = labor_posted_component_reconciliation(
                wage_amount,
                burden_amount,
                direct_posted,
                combined_posted,
            )
            matched = reconciliation["matched_cents"]
            wage_residual = reconciliation["wage_accrual_cents"]
            burden_residual = reconciliation[
                "employer_burden_accrual_cents"
            ]
            already_posted += matched
            wage_accrued += wage_residual
            burden_accrued += burden_residual
            wage_residual_by_project[project] += wage_residual
            burden_residual_by_project[project] += burden_residual
            if direct_posted or combined_posted:
                reviews.append(
                    {
                        "severity": "P2",
                        "type": "LABOR_POSTED_COMPONENT_MATCHED",
                        "project": project,
                        "period": period,
                        "allocated_wage_component_cents": wage_amount,
                        "allocated_employer_burden_cents": burden_amount,
                        "posted_5001003_wage_cents": direct_posted,
                        "posted_5001006_explicit_labor_cents": (
                            combined_posted
                        ),
                        "matched_cents": matched,
                        "wage_accrual_residual_cents": wage_residual,
                        "employer_burden_accrual_residual_cents": (
                            burden_residual
                        ),
                        "entity_match_proven": bool(
                            direct_entity or combined_entity
                        ),
                        "action": (
                            "同项目、同期间、唯一兼容雇佣主体下，5001003仅抵工资；"
                            "摘要明示工资/社保的5001006抵剩余工资与单位负担，"
                            "按最大余数拆分且总分守恒"
                        ),
                    }
                )
            posted_excess = (
                reconciliation["direct_wage_posted_excess_cents"]
                + reconciliation["combined_posted_excess_cents"]
            )
            if posted_excess:
                reviews.append(
                    {
                        "severity": "P1",
                        "type": "LABOR_POSTED_EXCEEDS_ALLOCATED_COMPONENTS",
                        "project": project,
                        "period": period,
                        "excess_cents": posted_excess,
                        "action": "项目人工过账超过同主体工资与单位负担分配额；保持毛利阻断并核对范围",
                    }
                )
            if (
                allocation_entity
                and not direct_entity
                and direct_candidate_count
            ) or (
                allocation_entity
                and not combined_entity
                and combined_candidate_count
            ):
                reviews.append(
                    {
                        "severity": "P1",
                        "type": "LABOR_POSTED_ENTITY_UNRESOLVED",
                        "project": project,
                        "period": period,
                        "direct_wage_candidate_entity_count": (
                            direct_candidate_count
                        ),
                        "combined_candidate_entity_count": (
                            combined_candidate_count
                        ),
                        "action": "同项目期间存在人工过账但雇佣主体不唯一兼容；不得跨主体抵减工资应计",
                    }
                )
        burden_source = burden_source_by_period.get(period)
        for project, amount in sorted(wage_residual_by_project.items()):
            if amount <= 0:
                continue
            key = [period, project, amount, payroll_source["source_id"]]
            events.append(
                {
                    "event_id": "labor_" + sha256_bytes(stable_json(key))[:24],
                    "project": project,
                    "plane": "COST_ACCRUED",
                    "category": "自有人工-工资应计",
                    "amount_cents": amount,
                    "posting_date": _period_end(period).isoformat(),
                    "payroll_period": period,
                    "summary": "工资组件×工资表批准出勤控制额内的核定项目日（钉钉唯一定位仅补空白）；最大余数法",
                    "source_id": payroll_source["source_id"],
                    "source_member": path.name,
                    "identity_reason": "PAYROLL_APPROVED_TIME_PLUS_CONTROLLED_PROJECT_DAY",
                    "approved_project_days": str(allocated_days_by_project[project]),
                    "allocated_employee_count": len(allocated_people_by_project[project]),
                    "evidence_source_ids": [
                        source["source_id"] for source in attendance_sources
                    ],
                }
            )
        for project, amount in sorted(
            burden_residual_by_project.items()
        ):
            if amount <= 0:
                continue
            source_id = (
                burden_source.get("source_id")
                if burden_source is not None
                else payroll_source["source_id"]
            )
            key = [period, project, amount, source_id]
            events.append(
                {
                    "event_id": "labor_burden_"
                    + sha256_bytes(stable_json(key))[:24],
                    "project": project,
                    "plane": "COST_ACCRUED",
                    "category": "自有人工-雇主社保医保应计",
                    "amount_cents": amount,
                    "posting_date": _period_end(period).isoformat(),
                    "payroll_period": period,
                    "summary": "单位承担社保医保×同员工工资表批准出勤控制额内的核定项目日；最大余数法",
                    "source_id": source_id,
                    "source_member": (
                        burden_source.get("relative_path")
                        if burden_source is not None
                        else None
                    ),
                    "identity_reason": "EMPLOYER_ENTITY_EMPLOYEE_PLUS_CONTROLLED_PROJECT_DAY",
                    "approved_project_days": str(
                        allocated_days_by_project[project]
                    ),
                    "allocated_employee_count": len(
                        allocated_people_by_project[project]
                    ),
                    "evidence_source_ids": [
                        source["source_id"] for source in attendance_sources
                    ],
                }
            )
        conservation_delta = (
            wage_control
            + burden_control
            - wage_accrued
            - burden_accrued
            - already_posted
            - wage_unallocated
            - burden_unallocated
        )
        if conservation_delta:
            raise ProjectCostError(
                "LABOR_CONTROL_DRIFT",
                "wage + employer burden must equal accrual + matched labor postings + unallocated",
            )
        if wage_unallocated:
            reviews.append(
                {
                    "severity": "P2",
                    "type": "LABOR_WAGE_COMPONENT_UNALLOCATED",
                    "source": payroll_source["relative_path"],
                    "amount_cents": wage_unallocated,
                    "action": (
                        "工资组件已按项目日证据分配并守恒；剩余部分保留为"
                        "未分配控制池，不向任何项目塞数"
                    ),
                }
            )
        if burden_unallocated:
            reviews.append(
                {
                    "severity": "P2",
                    "type": "LABOR_EMPLOYER_BURDEN_UNALLOCATED",
                    "period": period,
                    "amount_cents": burden_unallocated,
                    "action": "单位负担成本未唯一归属部分保留控制池，不按比例向项目塞数",
                }
            )
        if not attendance_sources:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "LABOR_TIME_PERIOD_NOT_MAPPED",
                    "source": payroll_source["relative_path"],
                    "action": "该工资期间无可绑定的钉钉逐日项目定位；工资控制额全部保留未分配",
                }
            )
        if duplicate_name_rows or invalid_time_rows:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "LABOR_EMPLOYEE_TIME_REJECTED",
                    "source": payroll_source["relative_path"],
                    "duplicate_name_rows": duplicate_name_rows,
                    "invalid_time_rows": invalid_time_rows,
                    "action": "歧义姓名或项目日超过批准出勤的行不分配，且不输出人员身份",
                }
            )
        if missing_burden_rows:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "LABOR_EMPLOYER_BURDEN_MISSING",
                    "period": period,
                    "employee_row_count": missing_burden_rows,
                    "action": "项目人工工资行缺少同月同雇佣主体单位负担来源；该期间毛利率保持阻断",
                }
            )
        if unmatched_direct_burden_rows:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "LABOR_EMPLOYER_BURDEN_WITHOUT_PAYROLL_ROW",
                    "period": period,
                    "employee_row_count": unmatched_direct_burden_rows,
                    "action": "生产或项目人员单位负担缺少同月工资行；金额已守恒保留未分配",
                }
            )
        reviews.append(
            {
                "severity": "P2",
                "type": "LABOR_COMPONENT_SCOPE_CONTROLLED",
                "source": payroll_source["relative_path"],
                "action": "工资与单位承担社保医保分别分配并守恒；个人扣款不反推单位成本，公积金无单位来源时不猜测",
            }
        )
        period_meta.append(
            {
                "period": period,
                "employee_row_count": len(payroll_rows),
                "direct_wage_component_control_cents": wage_control,
                "direct_employer_burden_control_cents": burden_control,
                "fully_loaded_labor_control_cents": (
                    wage_control + burden_control
                ),
                "wage_allocated_accrual_cents": wage_accrued,
                "employer_burden_allocated_accrual_cents": burden_accrued,
                "allocated_accrual_cents": wage_accrued + burden_accrued,
                "already_posted_cents": already_posted,
                "wage_unallocated_cents": wage_unallocated,
                "employer_burden_unallocated_cents": burden_unallocated,
                "unallocated_cents": wage_unallocated + burden_unallocated,
                "non_direct_payroll_cents_excluded": non_direct_excluded,
                "non_direct_employer_burden_cents_excluded": (
                    non_direct_burden_excluded
                ),
                "missing_employer_burden_employee_row_count": (
                    missing_burden_rows
                ),
                "unmatched_direct_employer_burden_row_count": (
                    unmatched_direct_burden_rows
                ),
                "conservation_delta_cents": conservation_delta,
                "attendance": attendance_meta,
            }
        )
        grand_wage_control += wage_control
        grand_burden_control += burden_control
        grand_wage_accrual += wage_accrued
        grand_burden_accrual += burden_accrued
        grand_posted += already_posted
        grand_wage_unallocated += wage_unallocated
        grand_burden_unallocated += burden_unallocated
    burden_without_payroll = sorted(
        set(burden_records_by_period) - set(period_books)
    )
    if burden_without_payroll:
        reviews.append(
            {
                "severity": "P1",
                "type": "EMPLOYER_BURDEN_PERIOD_WITHOUT_PAYROLL",
                "period_count": len(burden_without_payroll),
                "action": "单位负担来源存在无同月工资表期间；不得跨月或按姓名强配",
            }
        )
    return events, _dedupe_review_rows(reviews), sources, {
        "provided": True,
        "periods": period_meta,
        "wage_component_control_cents": grand_wage_control,
        "employer_burden_control_cents": grand_burden_control,
        "fully_loaded_labor_control_cents": (
            grand_wage_control + grand_burden_control
        ),
        "wage_allocated_accrual_cents": grand_wage_accrual,
        "employer_burden_allocated_accrual_cents": grand_burden_accrual,
        "allocated_accrual_cents": (
            grand_wage_accrual + grand_burden_accrual
        ),
        "already_posted_cents": grand_posted,
        "wage_unallocated_cents": grand_wage_unallocated,
        "employer_burden_unallocated_cents": grand_burden_unallocated,
        "unallocated_cents": (
            grand_wage_unallocated + grand_burden_unallocated
        ),
        "conservation_delta_cents": (
            grand_wage_control
            + grand_burden_control
            - grand_wage_accrual
            - grand_burden_accrual
            - grand_posted
            - grand_wage_unallocated
            - grand_burden_unallocated
        ),
        "fixed_daily_rate_used": False,
        "allocation_method": "largest_remainder",
        "personal_data_in_output": False,
        "personal_deductions_used": False,
        "employer_burden": burden_meta,
    }


def inventory_sources(roots: Sequence[Path]) -> Dict[str, Any]:
    files = list(iter_source_files(roots))
    candidates = discover_candidates(files)
    counts = {key: len(value) for key, value in candidates.items()}
    extension_counts: Dict[str, int] = defaultdict(int)
    total_bytes = 0
    for path in files:
        extension_counts[path.suffix.lower() or "<none>"] += 1
        total_bytes += path.stat().st_size
    return {
        "schema_version": "kmfa.project_cost.inventory.v1",
        "generated_at": utc_now(),
        "roots": [str(Path(root).resolve()) for root in roots],
        "file_count": len(files),
        "total_bytes": total_bytes,
        "candidate_counts": counts,
        "extension_counts": dict(sorted(extension_counts.items())),
    }


def build_snapshot(
    roots: Sequence[Path],
    *,
    year: int,
    as_of: str,
    ocr_jsonl: Optional[Path] = None,
    payroll_workbooks: Sequence[Path] = (),
    employer_burden_workbooks: Sequence[Path] = (),
    attendance_roots: Sequence[Path] = (),
    payroll_password_env: Optional[str] = None,
    private_input_manifest_sha256: Optional[str] = None,
) -> Dict[str, Any]:
    try:
        date.fromisoformat(as_of)
    except ValueError as exc:
        raise ProjectCostError("AS_OF_INVALID", "as-of must use YYYY-MM-DD") from exc
    if private_input_manifest_sha256 is not None and re.fullmatch(
        r"[0-9a-f]{64}",
        str(private_input_manifest_sha256),
    ) is None:
        raise ProjectCostError(
            "PRIVATE_MANIFEST_BINDING",
            "private input manifest digest must be lowercase SHA-256",
        )
    subject_binding = subject_source_binding()
    files = list(iter_source_files(roots))
    candidates = discover_candidates(files)
    _, all_projects, master_sources, master_errors = select_master(candidates["master"], roots)
    target_prefix = "%04d-" % year
    status_candidates = [
        project
        for project in all_projects
        if project.get("year") in (year - 1, year)
        or str(project.get("completion_date_master") or "").startswith(
            target_prefix
        )
    ]
    candidate_bases = [
        str(project["contract_base"]) for project in status_candidates
    ]
    if len(candidate_bases) != len(set(candidate_bases)):
        raise ProjectCostError(
            "COHORT_IDENTITY_CONFLICT",
            "target-year and carryover candidate contracts are not unique",
        )
    (
        status_path,
        candidate_status_map,
        status_sources,
        status_reviews,
        status_meta,
    ) = select_latest_tabular(
        candidates["status"],
        roots,
        parse_status,
        source_slot="production_status",
        parser_args=(status_candidates, year),
    )
    if candidate_status_map is None:
        candidate_status_map = {}
    projects = reporting_project_cohort(
        status_candidates,
        candidate_status_map,
        year=year,
    )
    if not projects:
        raise ProjectCostError(
            "YEAR_PROJECTS_EMPTY",
            "red-circle master/status has no target-year or carryover projects for %d"
            % year,
        )
    year_bases = [str(project["contract_base"]) for project in projects]
    if len(year_bases) != len(set(year_bases)):
        raise ProjectCostError(
            "YEAR_IDENTITY_CONFLICT",
            "reporting-cohort red-circle contracts have duplicate normalized identities",
        )
    status_map = {
        base: candidate_status_map[base]
        for base in year_bases
        if base in candidate_status_map
    }
    status_meta = dict(status_meta)
    status_meta["reporting_cohort_project_count"] = len(projects)
    status_meta["target_contract_year_project_count"] = sum(
        project.get("year") == year for project in projects
    )
    status_meta["carryover_project_count"] = sum(
        project.get("year") != year for project in projects
    )
    books, ledger_sources, ledger_errors = collect_ledger_books(candidates["ledger_zip"], roots)
    ledger_events, ledger_reviews, ledger_meta = parse_ledger_books(
        books,
        projects,
        year,
        as_of,
        status_map,
    )
    labor_events, labor_reviews, labor_sources, labor_meta = parse_payroll_and_attendance(
        payroll_workbooks,
        attendance_roots,
        employer_burden_workbooks,
        projects,
        status_map,
        ledger_events,
        roots,
        year=year,
        as_of=as_of,
        password_env=payroll_password_env,
    )
    payment_path, payment_events, payment_sources, payment_reviews, payment_meta = select_latest_tabular(
        candidates["payment"],
        roots,
        parse_payment,
        source_slot="payment_approval",
        parser_args=(projects, year),
    )
    if payment_events is None:
        payment_events = []
    dingtalk_sources, dingtalk_errors = inspect_dingtalk_archives(
        candidates["dingtalk_zip"],
        roots,
    )
    dws_events, dws_reviews, dws_sources, dws_meta = parse_dws_approvals(
        candidates["dws_zip"],
        roots,
        projects,
        status_map,
        year,
        as_of,
    )
    (
        approved_cost_detail_path,
        approved_cost_detail_events,
        approved_cost_detail_sources,
        approved_cost_detail_reviews,
        approved_cost_detail_meta,
    ) = select_latest_tabular(
        candidates["approved_cost_detail"],
        roots,
        parse_approved_cost_detail,
        source_slot="approved_project_cost_detail",
        parser_args=(projects, status_map, year, as_of),
    )
    if approved_cost_detail_events is None:
        approved_cost_detail_events = []
    (
        project_invoice_path,
        project_invoice_tax_events,
        project_invoice_sources,
        project_invoice_reviews,
        project_invoice_meta,
    ) = select_latest_tabular(
        candidates["project_invoice"],
        roots,
        parse_project_invoice_tax,
        source_slot="project_invoice_output_vat",
        parser_args=(projects, year, as_of),
    )
    if project_invoice_tax_events is None:
        project_invoice_tax_events = []
    (
        approved_cost_events,
        approved_cost_merge_reviews,
        approved_cost_merge_meta,
    ) = merge_approved_cost_sources(
        dws_events,
        approved_cost_detail_events,
    )
    ocr_events, ocr_reviews, ocr_sources, ocr_meta = parse_ocr_paid_project_costs(
        ocr_jsonl,
        roots,
        projects,
        status_map,
        year,
        as_of,
    )
    accrual_events, accrual_reviews, accrual_meta = qualify_cost_accruals(
        ledger_events,
        approved_cost_events,
        ocr_events,
    )
    sources = (
        master_sources
        + status_sources
        + ledger_sources
        + payment_sources
        + dingtalk_sources
        + dws_sources
        + approved_cost_detail_sources
        + project_invoice_sources
        + ocr_sources
        + labor_sources
    )
    reviews: List[Dict[str, Any]] = []
    reviews.extend(status_reviews)
    reviews.extend(ledger_reviews)
    reviews.extend(payment_reviews)
    reviews.extend(dws_reviews)
    reviews.extend(approved_cost_detail_reviews)
    reviews.extend(project_invoice_reviews)
    reviews.extend(approved_cost_merge_reviews)
    reviews.extend(ocr_reviews)
    reviews.extend(labor_reviews)
    reviews.extend(accrual_reviews)
    for error in master_errors + ledger_errors + dingtalk_errors:
        reviews.append(
            {
                "severity": "P1",
                "type": "SOURCE_CANDIDATE_REJECTED",
                "source": error.get("source"),
                "detail": error.get("error"),
                "action": "已跳过失败候选；如其应为权威来源需修复后以新快照重跑",
            }
        )
    ledger_available = bool(books)
    ledger_periods = ledger_meta.get("period_ends_by_entity") or {}
    project_ledger_entities: Dict[str, Set[str]] = defaultdict(set)
    for event in ledger_events:
        if event.get("entity"):
            project_ledger_entities[str(event.get("project"))].add(
                normalize_text(event.get("entity"))
            )
    contractor_entity_evidence: Dict[str, Set[str]] = defaultdict(set)
    for project in projects:
        base = str(project["contract_base"])
        observed_entities = project_ledger_entities.get(base, set())
        contractor = normalize_text(project.get("contractor"))
        if contractor and len(observed_entities) == 1:
            contractor_entity_evidence[contractor].update(observed_entities)
    for entity, period_end in sorted(ledger_periods.items()):
        if period_end and period_end < as_of[:7]:
            reviews.append(
                {
                    "severity": "P1",
                    "type": "LEDGER_PERIOD_BEFORE_AS_OF",
                    "entity": entity,
                    "ledger_period_end": period_end,
                    "as_of": as_of,
                    "action": "结果仅覆盖该账簿截止月及其后的合格应计；禁止标成截至日全期账簿",
                }
            )
    payment_available = payment_path is not None
    aggregate: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for event in (
        list(ledger_events)
        + list(payment_events)
        + list(accrual_events)
        + list(project_invoice_tax_events)
        + list(labor_events)
    ):
        aggregate[str(event["project"])][str(event["plane"])] += int(event["amount_cents"])
    detail_events: List[Dict[str, Any]] = (
        list(ledger_events)
        + list(payment_events)
        + list(dws_events)
        + list(approved_cost_detail_events)
        + list(project_invoice_tax_events)
        + list(ocr_events)
        + list(accrual_events)
        + list(labor_events)
    )
    project_rows: List[Dict[str, Any]] = []
    for project in projects:
        base = str(project["contract_base"])
        status = status_map.get(base)
        revenue = governed_contract_revenue(project, status)
        observed_entities = project_ledger_entities.get(base, set())
        contractor_entities = contractor_entity_evidence.get(
            normalize_text(project.get("contractor")),
            set(),
        )
        if len(observed_entities) == 1:
            project_ledger_period_end = _entity_ledger_period_end(
                next(iter(observed_entities)),
                ledger_periods,
            )
            ledger_entity_resolution = "EXACT_PROJECT_LEDGER_EVENT"
        elif len(contractor_entities) == 1:
            project_ledger_period_end = _entity_ledger_period_end(
                next(iter(contractor_entities)),
                ledger_periods,
            )
            ledger_entity_resolution = "SAME_CONTRACTOR_EXACT_EVENT_EVIDENCE"
        else:
            project_ledger_period_end = _project_ledger_period_end(
                project,
                ledger_periods,
            )
            ledger_entity_resolution = (
                "UNIQUE_CONTRACTOR_LEDGER_ENTITY"
                if project_ledger_period_end
                else "UNRESOLVED"
            )
        project_ledger_available = bool(
            ledger_available and project_ledger_period_end
        )
        job_posted = (
            aggregate[base]["JOB_POSTED_ACTUAL"]
            if project_ledger_available
            else None
        )
        accrued = (
            aggregate[base]["COST_ACCRUED"]
            if project_ledger_available
            else None
        )
        incurred = (
            int(job_posted) + int(accrued)
            if job_posted is not None and accrued is not None
            else None
        )
        recognized = (
            aggregate[base]["GL_RECOGNIZED_COGS"]
            if project_ledger_available
            else None
        )
        paid_observed = (
            aggregate[base]["PAYMENT_SYSTEM_PAID_OBSERVED"] if payment_available else None
        )
        business_reported = (
            status.get("business_reported_direct_cost_cents")
            if status is not None
            else None
        )
        row = dict(project)
        row.update(
            {
                "construction_status": (
                    status.get("construction_status")
                    if status is not None
                    else project.get("construction_status_master")
                ),
                "status_source_contract": status.get("source_contract") if status else None,
                "project_type": status.get("project_type") if status else None,
                "owner": status.get("owner") if status else None,
                "start_date": status.get("start_date") if status else None,
                "completion_date": (
                    status.get("completion_date")
                    if status and status.get("completion_date")
                    else project.get("completion_date_master")
                ),
                "settlement_date": status.get("settlement_date") if status else None,
                "invoice_date": status.get("invoice_date") if status else None,
                "cash_in_date": status.get("cash_in_date") if status else None,
                "own_work_units": status.get("own_work_units") if status else None,
                "external_work_units": status.get("external_work_units") if status else None,
                "status_business_components_cents": (
                    status.get("business_components_cents") if status else {}
                ),
                "status_settlement_amount_cents": (
                    status.get("settlement_amount_cents") if status else None
                ),
                "status_invoice_amount_cents": (
                    status.get("invoice_amount_cents") if status else None
                ),
                "status_contract_amount_cents": (
                    status.get("status_contract_amount_cents") if status else None
                ),
                "project_cost_report_deadline": (
                    status.get("project_cost_report_deadline") if status else None
                ),
                "project_cost_report_provided": (
                    status.get("project_cost_report_provided") if status else False
                ),
                "commission_calculated": (
                    status.get("commission_calculated") if status else False
                ),
                "job_posted_actual_cents": job_posted,
                "cost_accrued_cents": accrued,
                "job_cost_incurred_cents": incurred,
                # Actual-to-date is a lower bound.  Final project cost/FAC and
                # its governed revenue basis remain empty until a controlled
                # close process proves them.
                "gross_margin_cost_basis_cents": None,
                "effective_revenue_cents": revenue["effective_revenue_cents"],
                "revenue_basis_status": revenue["status"],
                "revenue_basis": revenue["basis"],
                "gross_margin_status": (
                    "BLOCKED_COST_COMPLETENESS"
                    if revenue["effective_revenue_cents"] is not None
                    else "BLOCKED_REVENUE_AND_COST_COMPLETENESS"
                ),
                "gl_recognized_cogs_cents": recognized,
                "business_reported_direct_cost_cents": business_reported,
                "payment_system_paid_observed_cents": paid_observed,
                "ledger_period_end": project_ledger_period_end,
                "ledger_entity_resolution": ledger_entity_resolution,
                "job_cost_coverage": (
                    "GL_SELECTED_THROUGH_%s;POSTING_PRESENT"
                    % project_ledger_period_end
                    if project_ledger_available and job_posted
                    else (
                        "GL_SELECTED_THROUGH_%s;NO_QUALIFIED_EVENT"
                        % project_ledger_period_end
                        if project_ledger_available
                        else "SOURCE_UNAVAILABLE"
                    )
                ),
                "accrual_coverage": (
                    "QUALIFIED_NOT_YET_POSTED_EVENT"
                    if project_ledger_available and accrued
                    else (
                        "NO_QUALIFIED_ACCRUAL_EVENT"
                        if project_ledger_available
                        else "SOURCE_UNAVAILABLE"
                    )
                ),
                "status_coverage": (
                    "STATUS_ROW_PRESENT" if status is not None else "NO_MATCHED_STATUS_ROW"
                ),
                "payment_coverage": (
                    "EXACT_SINGLE_PROJECT_MAPPING_ONLY"
                    if payment_available
                    else "SOURCE_UNAVAILABLE"
                ),
            }
        )
        project_rows.append(row)
        if status:
            for category, amount in status.get("business_components_cents", {}).items():
                if amount is None:
                    continue
                detail_events.append(
                    {
                        "event_id": "status_"
                        + sha256_bytes(
                            stable_json([base, category, amount, status.get("source_row")])
                        )[:24],
                        "project": base,
                        "plane": "BUSINESS_REPORTED_DIRECT_COST",
                        "category": category,
                        "amount_cents": amount,
                        "posting_date": status.get("completion_date"),
                        "account_code": None,
                        "voucher": None,
                        "summary": "生产项目状态表金额观察",
                        "entity": None,
                        "source_id": (
                            next(
                                (
                                    source["source_id"]
                                    for source in status_sources
                                    if source.get("selected")
                                ),
                                None,
                            )
                        ),
                        "source_member": None,
                        "sheet": None,
                        "row": status.get("source_row"),
                        "identity_reason": status.get("identity_reason"),
                    }
                )
    for project in project_rows:
        if project["job_cost_incurred_cents"] is None:
            reviews.append(
                {
                    "severity": "P0",
                    "type": "PROJECT_COST_SOURCE_UNAVAILABLE",
                    "project": project["contract_base"],
                    "action": "修复金蝶来源读取；不得把空值显示为 0",
                }
            )
        if (
            project["business_reported_direct_cost_cents"] is not None
            and project["job_cost_incurred_cents"] is not None
            and project["business_reported_direct_cost_cents"]
            != project["job_cost_incurred_cents"]
        ):
            reviews.append(
                {
                    "severity": "P2",
                    "type": "OBSERVATION_PLANE_DIFFERENCE",
                    "project": project["contract_base"],
                    "job_cost_incurred_cents": project["job_cost_incurred_cents"],
                    "business_reported_direct_cost_cents": project[
                        "business_reported_direct_cost_cents"
                    ],
                    "action": "并列保留；禁止相加、覆盖或取最大值",
                }
            )
    reviews = _dedupe_review_rows(reviews)
    reviews_control = review_summary(reviews)
    selected_source_binding = [
        {
            "source_id": source["source_id"],
            "sha256": source["sha256"],
            "logical_metadata": source.get("logical_metadata"),
        }
        for source in sources
        if source.get("selected")
    ]
    selected_source_binding_digest = sha256_bytes(
        stable_json(selected_source_binding)
    )
    input_manifest_binding = {
        "kind": (
            "PRIVATE_MANIFEST_SHA256"
            if private_input_manifest_sha256 is not None
            else "SELECTED_SOURCE_DERIVED_SHA256"
        ),
        "digest": (
            private_input_manifest_sha256
            if private_input_manifest_sha256 is not None
            else selected_source_binding_digest
        ),
    }
    binding = {
        "skill_version": SKILL_VERSION,
        "subject_digest": subject_binding["digest"],
        "input_manifest_binding": input_manifest_binding,
        "selected_source_binding_digest": selected_source_binding_digest,
        "year": year,
        "as_of": as_of,
        "projects": [
            {
                "project": row["contract_base"],
                "incurred": row["job_cost_incurred_cents"],
                "cogs": row["gl_recognized_cogs_cents"],
            }
            for row in project_rows
        ],
        "sources": selected_source_binding,
    }
    snapshot_id = "kmfa-pc-%d-%s" % (year, sha256_bytes(stable_json(binding))[:12])
    return {
        "schema_version": "kmfa.project_cost.snapshot.v2",
        "snapshot_id": snapshot_id,
        "generated_at": utc_now(),
        "skill_version": SKILL_VERSION,
        "core_version": CORE_VERSION,
        "year": year,
        "as_of": as_of,
        "currency": "CNY",
        "money_unit": "integer_cents",
        "subject_binding": subject_binding,
        "private_input_manifest_sha256": private_input_manifest_sha256,
        "input_manifest_binding": input_manifest_binding,
        "selected_source_binding_digest": selected_source_binding_digest,
        "project_count": len(project_rows),
        "projects": project_rows,
        "events": sorted(
            detail_events,
            key=lambda event: (
                str(event.get("project") or ""),
                str(event.get("posting_date") or ""),
                str(event.get("plane") or ""),
                str(event.get("event_id") or ""),
            ),
        ),
        "sources": sources,
        "reviews": reviews,
        "review_summary": reviews_control,
        "coverage": {
            "master_selected": True,
            "status_selected": status_path is not None,
            "ledger_selected_book_count": len(books),
            "payment_selected": payment_path is not None,
            "approved_cost_detail_selected": (
                approved_cost_detail_path is not None
            ),
            "project_invoice_selected": (
                project_invoice_path is not None
            ),
            "ledger_logical_period_end": ledger_meta.get("selected_period_end"),
            "ledger_minimum_period_end": ledger_meta.get("minimum_period_end"),
            "ledger_entity_count": len(ledger_meta.get("period_ends_by_entity") or {}),
            "ledger_stale_entity_count": sum(
                bool(period_end and period_end < as_of[:7])
                for period_end in (ledger_meta.get("period_ends_by_entity") or {}).values()
            ),
            "status_max_observed_date": status_meta.get("max_observed_date"),
            "payment_max_observed_date": payment_meta.get("max_observed_date"),
            "ocr_formal_amount_use": bool(ocr_events),
            "dws_reaction_observed_event_count": len(dws_events),
            "dws_reaction_formal_amount_use": False,
            "approved_cost_detail_event_count": len(
                approved_cost_detail_events
            ),
            "project_invoice_tax_event_count": len(
                project_invoice_tax_events
            ),
            "approved_cost_combined_event_count": len(
                approved_cost_events
            ),
            "ocr_paid_project_cost_event_count": len(ocr_events),
            "qualified_accrual_event_count": len(accrual_events),
            "labor_wage_component_event_count": sum(
                event.get("category") == "自有人工-工资应计"
                for event in labor_events
            ),
            "labor_employer_burden_event_count": sum(
                event.get("category") == "自有人工-雇主社保医保应计"
                for event in labor_events
            ),
            "labor_wage_component_control_cents": labor_meta.get(
                "wage_component_control_cents"
            ),
            "labor_employer_burden_control_cents": labor_meta.get(
                "employer_burden_control_cents"
            ),
            "labor_fully_loaded_control_cents": labor_meta.get(
                "fully_loaded_labor_control_cents"
            ),
            "labor_allocated_accrual_cents": labor_meta.get(
                "allocated_accrual_cents"
            ),
            "labor_wage_allocated_accrual_cents": labor_meta.get(
                "wage_allocated_accrual_cents"
            ),
            "labor_employer_burden_allocated_accrual_cents": labor_meta.get(
                "employer_burden_allocated_accrual_cents"
            ),
            "labor_unallocated_cents": labor_meta.get("unallocated_cents"),
            "labor_wage_unallocated_cents": labor_meta.get(
                "wage_unallocated_cents"
            ),
            "labor_employer_burden_unallocated_cents": labor_meta.get(
                "employer_burden_unallocated_cents"
            ),
        },
        "diagnostics": {
            "ledger": ledger_meta,
            "status": status_meta,
            "payment": payment_meta,
            "dws": dws_meta,
            "approved_cost_detail": approved_cost_detail_meta,
            "project_invoice": project_invoice_meta,
            "approved_cost_source_merge": approved_cost_merge_meta,
            "ocr": ocr_meta,
            "accrual": accrual_meta,
            "labor": labor_meta,
            "candidate_counts": {key: len(value) for key, value in candidates.items()},
        },
        "formula_contract": {
            "JOB_COST_INCURRED": "JOB_POSTED_ACTUAL + COST_ACCRUED",
            "planes_not_combined": [
                "GL_RECOGNIZED_COGS",
                "BUSINESS_REPORTED_DIRECT_COST",
                "PAYMENT_SYSTEM_PAID_OBSERVED",
            ],
            "automatic_fixed_labor_rate": False,
            "labor_allocation": (
                "auditable wage + employer-paid social/medical burden × "
                "approved payroll days; entity+employee match; largest remainder"
            ),
            "labor_unallocated_preserved": True,
            "personal_deductions_used": False,
            "employer_burden_inferred_from_personal_deduction": False,
            "automatic_management_fee_percent": False,
            "project_output_vat": (
                "per approved issued invoice/project line: "
                "ROUND_HALF_UP(gross_invoice_cents * rate / (1 + rate))"
            ),
            "company_tax_default_allocation": False,
            "historical_reference_in_calculate": False,
        },
    }


def _style_workbook(workbook: Any, snapshot_id: str) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    navy = "17365D"
    blue = "D9EAF7"
    gray = "E7E6E6"
    white = "FFFFFF"
    thin = Side(style="thin", color="B7B7B7")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A4"
        sheet["A1"] = "snapshot_id"
        sheet["B1"] = snapshot_id
        sheet["A1"].font = Font(bold=True, color=white)
        sheet["B1"].font = Font(bold=True, color=white)
        sheet["A1"].fill = PatternFill("solid", fgColor=navy)
        sheet["B1"].fill = PatternFill("solid", fgColor=navy)
        for cell in sheet[3]:
            if cell.value is None:
                continue
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        for row in sheet.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin)
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            sample = [str(cell.value or "") for cell in list(column_cells)[:120]]
            width = min(42, max(10, max((len(text) for text in sample), default=10) + 2))
            sheet.column_dimensions[letter].width = width
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[3].height = 32
        if sheet.max_row >= 4:
            for cell in sheet[4]:
                if cell.fill.fill_type is None:
                    cell.fill = PatternFill("solid", fgColor=blue)


def _append_table(
    sheet: Any,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    money_columns: Sequence[int] = (),
) -> None:
    for column, header in enumerate(headers, 1):
        sheet.cell(row=3, column=column, value=spreadsheet_text(header))
    row_number = 4
    for row in rows:
        for column, value in enumerate(row, 1):
            sheet.cell(row=row_number, column=column, value=spreadsheet_text(value))
        row_number += 1
    for column in money_columns:
        for row_number in range(4, sheet.max_row + 1):
            sheet.cell(row=row_number, column=column).number_format = MONEY_FORMAT


def write_workbook(path: Path, snapshot: Mapping[str, Any]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ProjectCostError("DEPENDENCY_MISSING", "openpyxl is required") from exc
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in FORMULA_SHEETS}
    projects = snapshot["projects"]
    _append_table(
        sheets["01_项目成本表"],
        (
            "合同编号",
            "项目名称",
            "客户/甲方",
            "施工状态",
            "含税合同额（元）",
            "项目过账实际（元）",
            "项目应计（元）",
            "项目已发生成本（元）",
            "主营成本已结转（元）",
            "状态表已报直接成本（元）",
            "支付系统已付观察（元）",
            "自有工数",
            "劳务工数",
            "项目成本覆盖",
            "状态覆盖",
        ),
        (
            (
                project["canonical_contract_id"],
                project["project_name"],
                project["customer"],
                project.get("construction_status"),
                yuan_from_cents(project.get("contract_amount_cents")),
                yuan_from_cents(project.get("job_posted_actual_cents")),
                yuan_from_cents(project.get("cost_accrued_cents")),
                yuan_from_cents(project.get("job_cost_incurred_cents")),
                yuan_from_cents(project.get("gl_recognized_cogs_cents")),
                yuan_from_cents(project.get("business_reported_direct_cost_cents")),
                yuan_from_cents(project.get("payment_system_paid_observed_cents")),
                project.get("own_work_units"),
                project.get("external_work_units"),
                project.get("job_cost_coverage"),
                project.get("status_coverage"),
            )
            for project in projects
        ),
        money_columns=(5, 6, 7, 8, 9, 10, 11),
    )
    _append_table(
        sheets["02_成本明细"],
        (
            "事件ID",
            "合同基准号",
            "观察面",
            "成本分类",
            "金额（元）",
            "日期",
            "科目",
            "凭证/付款编号",
            "摘要",
            "核算主体",
            "来源ID",
            "来源成员",
            "工作表",
            "行号",
            "身份映射依据",
        ),
        (
            (
                event.get("event_id"),
                event.get("project"),
                event.get("plane"),
                event.get("category"),
                yuan_from_cents(event.get("amount_cents")),
                event.get("posting_date"),
                event.get("account_code"),
                event.get("voucher") or event.get("payment_id"),
                event.get("summary"),
                event.get("entity"),
                event.get("source_id"),
                event.get("source_member"),
                event.get("sheet"),
                event.get("row") or event.get("source_row"),
                event.get("identity_reason"),
            )
            for event in snapshot["events"]
        ),
        money_columns=(5,),
    )
    _append_table(
        sheets["03_生命周期对照"],
        (
            "合同编号",
            "创建日期",
            "开工日期",
            "完工日期",
            "验收日期",
            "结算日期",
            "开票日期",
            "回款日期",
            "施工状态",
            "项目类型",
            "负责人",
        ),
        (
            (
                project["canonical_contract_id"],
                project.get("created_date"),
                project.get("start_date"),
                project.get("completion_date"),
                project.get("acceptance_date"),
                project.get("settlement_date"),
                project.get("invoice_date"),
                project.get("cash_in_date"),
                project.get("construction_status"),
                project.get("project_type"),
                project.get("owner"),
            )
            for project in projects
        ),
    )
    _append_table(
        sheets["04_收入与现金"],
        (
            "合同编号",
            "含税合同额（元）",
            "红圈结算额（元）",
            "状态表结算额（元）",
            "累计开票（元）",
            "状态表开票额（元）",
            "项目收款（元）",
            "保证金支付（元）",
            "保证金退还（元）",
            "支付系统项目已付观察（元）",
            "说明",
        ),
        (
            (
                project["canonical_contract_id"],
                yuan_from_cents(project.get("contract_amount_cents")),
                yuan_from_cents(project.get("settlement_amount_cents")),
                yuan_from_cents(project.get("status_settlement_amount_cents")),
                yuan_from_cents(project.get("invoiced_cents")),
                yuan_from_cents(project.get("status_invoice_amount_cents")),
                yuan_from_cents(project.get("cash_in_cents")),
                yuan_from_cents(project.get("deposit_paid_cents")),
                yuan_from_cents(project.get("deposit_returned_cents")),
                yuan_from_cents(project.get("payment_system_paid_observed_cents")),
                "收入/现金与成本观察面独立；不得直接混算毛利",
            )
            for project in projects
        ),
        money_columns=(2, 3, 4, 5, 6, 7, 8, 9, 10),
    )
    _append_table(
        sheets["05_来源与核销"],
        (
            "来源ID",
            "槽位",
            "相对路径/成员",
            "SHA-256",
            "大小（字节）",
            "是否选中",
            "选择理由",
            "逻辑元数据",
        ),
        (
            (
                source.get("source_id"),
                source.get("source_slot"),
                source.get("relative_path"),
                source.get("sha256"),
                source.get("size_bytes"),
                "是" if source.get("selected") else "否",
                source.get("selection_reason"),
                json.dumps(source.get("logical_metadata"), ensure_ascii=False, sort_keys=True)
                if source.get("logical_metadata") is not None
                else None,
            )
            for source in snapshot["sources"]
        ),
    )
    _append_table(
        sheets["06_差异与待确认"],
        (
            "严重级别",
            "问题类型",
            "合同基准号",
            "来源",
            "来源行",
            "金额A（元）",
            "金额B（元）",
            "说明/动作",
            "证据JSON",
        ),
        (
            (
                review.get("severity"),
                review.get("type"),
                review.get("project"),
                review.get("source"),
                review.get("row") or review.get("source_row"),
                yuan_from_cents(
                    review.get("job_cost_incurred_cents")
                    if review.get("job_cost_incurred_cents") is not None
                    else review.get("amount_cents")
                ),
                yuan_from_cents(review.get("business_reported_direct_cost_cents")),
                review.get("action") or review.get("detail"),
                json.dumps(review, ensure_ascii=False, sort_keys=True),
            )
            for review in snapshot["reviews"]
        ),
        money_columns=(6, 7),
    )
    _append_table(
        sheets["07_项目身份"],
        (
            "Canonical 合同编号",
            "合同基准号",
            "项目名称",
            "客户/甲方",
            "合同乙方",
            "状态表合同号",
            "状态身份依据",
            "来源行",
        ),
        (
            (
                project["canonical_contract_id"],
                project["contract_base"],
                project["project_name"],
                project["customer"],
                project["contractor"],
                project.get("status_source_contract"),
                (
                    "EXACT/CONTROLLED_OR_REVIEWED"
                    if project.get("status_source_contract")
                    else "NO_MATCHED_STATUS_ROW"
                ),
                project.get("source_row"),
            )
            for project in projects
        ),
    )
    notes = (
        ("Skill 版本", snapshot["skill_version"]),
        ("年份", snapshot["year"]),
        ("as_of", snapshot["as_of"]),
        ("金额真值", "整数分；工作簿仅显示人民币元"),
        ("正式公式", "JOB_COST_INCURRED = JOB_POSTED_ACTUAL + COST_ACCRUED"),
        ("独立观察面", "GL_RECOGNIZED_COGS / 状态表成本 / 支付系统已付"),
        ("禁止", "观察面相加、取最大值、参考报表回填、固定工资率、自动2%管理费"),
        ("参考边界", "销售绩效考核/PDF 仅用于 validate-reference"),
        ("金蝶逻辑期间截止", snapshot["coverage"].get("ledger_logical_period_end")),
        ("状态数据日期", snapshot["coverage"].get("status_max_observed_date")),
        ("支付数据日期", snapshot["coverage"].get("payment_max_observed_date")),
        (
            "截图 OCR",
            "已限定用于项目成本实付观察并经未入账去重"
            if snapshot["coverage"].get("ocr_formal_amount_use")
            else "本次未提供或无合格项目成本行",
        ),
        (
            "人工工资组件控制额（元）",
            yuan_from_cents(
                snapshot["coverage"].get("labor_wage_component_control_cents")
            ),
        ),
        (
            "人工单位社保医保控制额（元）",
            yuan_from_cents(
                snapshot["coverage"].get(
                    "labor_employer_burden_control_cents"
                )
            ),
        ),
        (
            "人工全成本已分配（元）",
            yuan_from_cents(
                snapshot["coverage"].get("labor_allocated_accrual_cents")
            ),
        ),
        (
            "人工全成本未分配（元）",
            yuan_from_cents(
                snapshot["coverage"].get("labor_unallocated_cents")
            ),
        ),
        (
            "人工分摊",
            "工资及单位承担社保医保×工资表批准出勤控制额内的核定项目日；"
            "同主体同人员匹配；最大余数法；固定日薪和个人扣款反推禁用",
        ),
        ("项目数", snapshot["project_count"]),
        ("事件数", len(snapshot["events"])),
        ("待确认数", len(snapshot["reviews"])),
    )
    _append_table(
        sheets["08_运行说明"],
        ("项目", "值"),
        notes,
    )
    _style_workbook(workbook, str(snapshot["snapshot_id"]))
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.save(path)


def _safe_filename(value: str, limit: int = 90) -> str:
    cleaned = re.sub(r"[/\\:\x00-\x1f]+", "_", value).strip(" .")
    return cleaned[:limit] or "project"


def _sum_present_cents(*values: Optional[int]) -> Optional[int]:
    present = [int(value) for value in values if value is not None]
    return sum(present) if present else None


def _statement_rows(
    project: Mapping[str, Any],
    buckets: Mapping[str, int],
) -> List[Tuple[str, Optional[int], str]]:
    """Map formal cents into the supplied A-family statement without invention."""

    material = buckets.get("material")
    fuel_power = buckets.get("fuel_power")
    rental = buckets.get("rental")
    logistics = buckets.get("logistics")
    ticket = buckets.get("travel")
    lodging = buckets.get("lodging")
    living = buckets.get("living")
    vehicle_fuel = buckets.get("vehicle")
    road_parking = buckets.get("road_parking")
    other = buckets.get("other")
    own_labor = buckets.get("own_labor")
    subcontract_labor = buckets.get("subcontract_labor")
    total = project.get("job_cost_incurred_cents")

    material_total = _sum_present_cents(material, fuel_power)
    rental_total = _sum_present_cents(rental, logistics)
    travel = _sum_present_cents(ticket, lodging)
    vehicle = _sum_present_cents(vehicle_fuel, road_parking)
    site = _sum_present_cents(own_labor, travel, living, vehicle, other)
    classified = _sum_present_cents(
        material_total,
        rental_total,
        site,
        subcontract_labor,
    )
    if total == 0 and classified is None:
        classified = 0
    if total is not None and classified != total:
        raise ProjectCostError(
            "PDF_CATEGORY_CONSERVATION",
            "formal statement categories do not conserve project cents",
        )

    own_units = project.get("own_work_units")
    external_units = project.get("external_work_units")
    values: Dict[str, Tuple[Optional[int], str]] = {
        "contract": (
            project.get("contract_amount_cents"),
            "原始合同额；不等于有效合同额或收入确认额",
        ),
        "sec2": (total, "正式项目成本；政策行未自动计提"),
        "l2_material": (material_total, ""),
        "d_material": (material, ""),
        "d_fuel_power": (fuel_power, "燃料及动力"),
        "l2_rental": (
            rental_total,
            "设备租赁未细分吊车/脚手架" if rental else "",
        ),
        "d_logistics": (logistics, ""),
        "l2_site": (
            site,
            (
                "含未能安全细分到模板行的正式成本 %s 元；已计入本小计"
                % _pdf_money(other)
            )
            if other
            else "",
        ),
        "d_own_labor": (
            own_labor,
            (
                "自有 %s 个工" % own_units
            )
            if own_units is not None
            else ("正式自有人工成本" if own_labor is not None else ""),
        ),
        "d_travel": (travel, ""),
        "d_ticket": (ticket, ""),
        "d_lodging": (lodging, ""),
        "d_living": (living, ""),
        "d_vehicle": (vehicle, ""),
        "d_vehicle_fuel": (vehicle_fuel, ""),
        "d_road_parking": (road_parking, ""),
        "l2_subcontract_labor": (
            subcontract_labor,
            "外协 %s 个工" % external_units
            if external_units is not None
            else "",
        ),
        "allocation": (
            None,
            "模板原行；无合格管理费政策，禁止按合同额2%自动生成",
        ),
        "interest": (None, "无合格资金占用政策，留空"),
        "total": (total, ""),
        "profit": (
            None,
            "有效合同变更链与收入确认口径未闭合，禁止生成毛利",
        ),
    }
    rows: List[Tuple[str, Optional[int], str]] = []
    for label, kind in STATEMENT_TEMPLATE_A:
        amount, note = values.get(kind, (None, ""))
        if (
            amount is not None
            and total not in (None, 0)
            and kind not in ("contract", "total", "sec2")
        ):
            share = Decimal(amount) / Decimal(total)
            suffix = "%.2f%%" % (share * 100)
            note = "%s %s" % (note, suffix) if note else suffix
        rows.append((label, amount, note))
    return rows


def _pdf_money(value: Optional[int]) -> str:
    if value is None:
        return ""
    amount = Decimal(value) / 100
    if amount < 0:
        return "(%s)" % format(abs(amount), ",.2f")
    return format(amount, ",.2f")


def _register_pdf_cjk_font(pdfmetrics: Any, UnicodeCIDFont: Any, TTFont: Any) -> str:
    """Prefer an embedded local CJK font; retain a portable CID fallback."""

    configured = os.environ.get("KMFA_CJK_FONT")
    candidates = [
        configured,
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for raw in candidates:
        if not raw:
            continue
        path = Path(raw)
        if not path.is_file():
            continue
        try:
            pdfmetrics.registerFont(TTFont("KMFA-CJK", str(path), subfontIndex=0))
            return "KMFA-CJK"
        except Exception:
            try:
                pdfmetrics.registerFont(TTFont("KMFA-CJK", str(path)))
                return "KMFA-CJK"
            except Exception:
                continue
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    except Exception:
        pass
    return "STSong-Light"


def write_project_pdfs(directory: Path, snapshot: Mapping[str, Any]) -> List[Path]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from xml.sax.saxutils import escape
    except ImportError as exc:
        raise ProjectCostError("DEPENDENCY_MISSING", "reportlab is required for PDF output") from exc
    directory.mkdir(parents=True, exist_ok=False)
    font_name = _register_pdf_cjk_font(pdfmetrics, UnicodeCIDFont, TTFont)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=14,
        leading=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17365D"),
    )
    body_style = ParagraphStyle(
        "ChineseBody",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=6.2,
        leading=7.4,
    )
    header_style = ParagraphStyle(
        "ChineseHeader",
        parent=body_style,
        fontSize=6.8,
        leading=8,
    )

    def paragraph(value: Any, style: Any = body_style) -> Any:
        return Paragraph(escape("" if value is None else str(value)), style)

    categories_by_project = _formal_cost_categories(snapshot)
    paths: List[Path] = []
    for project in snapshot["projects"]:
        buckets = _statement_buckets(
            categories_by_project.get(str(project["contract_base"]), {})
        )
        name = "%s_%s.pdf" % (
            _safe_filename(str(project["canonical_contract_id"])),
            _safe_filename(str(project["project_name"]), 45),
        )
        path = directory / name
        document = SimpleDocTemplate(
            str(path),
            pagesize=A4,
            rightMargin=7 * mm,
            leftMargin=7 * mm,
            topMargin=7 * mm,
            bottomMargin=7 * mm,
            title="项目财务分析表",
            author="KMFA Project Cost Skill",
        )
        title = Paragraph("项目财务分析表", title_style)
        header = Table(
            [
                [paragraph("项目名称：", header_style), paragraph(project["project_name"], header_style)],
                [paragraph("合同编号", header_style), paragraph(project["canonical_contract_id"], header_style)],
                [paragraph("开工时间", header_style), paragraph(str(project.get("start_date") or "").replace("-", "/"), header_style)],
                [paragraph("完工时间", header_style), paragraph(str(project.get("completion_date") or "").replace("-", "/"), header_style)],
            ],
            colWidths=[24 * mm, 157 * mm],
            rowHeights=[4.4 * mm] * 4,
        )
        header.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 6.8),
                    ("FONTNAME", (0, 0), (0, -1), font_name),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B7B7B7")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DDEBF7")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 1),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ]
            )
        )
        data = [
            [
                paragraph("项目", header_style),
                paragraph("金额（元）", header_style),
                paragraph("备注", header_style),
            ]
        ]
        statement = _statement_rows(project, buckets)
        for label, amount, note in statement:
            data.append(
                [
                    paragraph(label),
                    paragraph(_pdf_money(amount)),
                    paragraph(note),
                ]
            )
        table = Table(
            data,
            colWidths=[63 * mm, 31 * mm, 87 * mm],
            repeatRows=1,
        )
        section_labels = {"二、资金运用及各项支出"}
        subtotal_labels = {
            "（一）原材料",
            "（二）租赁费",
            "（三）保险费",
            "（四）现场管理费",
            "（五）工资（承包费）支出",
            "（六）信息费",
        }
        total_labels = {"一、合同额", "合计支出", "（七）毛利"}
        policy_labels = {
            "三 1.1分摊的管理费用（合同的2%）",
            "1.2占用的资金利息",
        }
        table_commands = [
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 6.2),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#7F8C8D")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
            ("TOPPADDING", (0, 0), (-1, -1), 1.1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.1),
        ]
        for index, (label, _amount, _note) in enumerate(statement, start=1):
            fill = None
            if label in total_labels:
                fill = colors.HexColor("#F4B183")
            elif label in section_labels:
                fill = colors.HexColor("#9DC3E6")
            elif label in subtotal_labels:
                fill = colors.HexColor("#DDEBF7")
            elif label in policy_labels:
                fill = colors.HexColor("#FFF2CC")
                table_commands.append(
                    ("TEXTCOLOR", (0, index), (-1, index), colors.HexColor("#9C0006"))
                )
            if fill is not None:
                table_commands.append(("BACKGROUND", (0, index), (-1, index), fill))
        table.setStyle(TableStyle(table_commands))
        footer = Table(
            [[
                paragraph("项目经理：", header_style),
                paragraph(
                    "正式成本＝过账实际＋合格应计；管理费、利息、毛利无合格输入时留空。",
                    header_style,
                ),
                paragraph("日期：", header_style),
            ]],
            colWidths=[35 * mm, 111 * mm, 35 * mm],
        )
        footer.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 6.8),
                    ("LINEABOVE", (0, 0), (-1, 0), 0.4, colors.HexColor("#7F8C8D")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ]
            )
        )
        document.build(
            [
                title,
                Spacer(1, 1.5 * mm),
                header,
                Spacer(1, 1.5 * mm),
                table,
                Spacer(1, 1.5 * mm),
                footer,
                Spacer(1, 0.6 * mm),
                paragraph(
                    "快照ID：" + str(snapshot["snapshot_id"]),
                    header_style,
                ),
            ]
        )
        paths.append(path)
    return paths


def _money_label(value: Optional[int]) -> str:
    if value is None:
        return "不可计算"
    if value == 0:
        return "-"
    amount = Decimal(value) / 100
    if amount < 0:
        return "(¥%s)" % format(abs(amount), ",.2f")
    return "¥%s" % format(amount, ",.2f")


def write_summary_csv(path: Path, snapshot: Mapping[str, Any]) -> None:
    headers = (
        "canonical_contract_id",
        "contract_base",
        "project_name",
        "customer",
        "construction_status",
        "contract_amount_cents",
        "job_posted_actual_cents",
        "cost_accrued_cents",
        "job_cost_incurred_cents",
        "gl_recognized_cogs_cents",
        "business_reported_direct_cost_cents",
        "payment_system_paid_observed_cents",
        "job_cost_coverage",
    )
    with path.open("x", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for project in snapshot["projects"]:
            writer.writerow({key: project.get(key) for key in headers})


def _yuan_text(value: Optional[int]) -> Optional[str]:
    if value is None:
        return None
    return format(Decimal(value) / 100, ".2f")


def governed_gross_margin(
    *,
    revenue_cents: Optional[int],
    cost_cents: Optional[int],
    basis_status: str,
) -> Dict[str, Any]:
    """Calculate a publishable project margin only from a closed basis.

    ``JOB_COST_INCURRED`` is an actual-to-date lower bound while payroll,
    approved expenses, remaining commitments, or final settlement are still
    open.  Treating that lower bound as final cost is what produced the
    implausible 89%--100% margins in the withdrawn run.  A caller must
    explicitly close the revenue and cost basis before this function returns a
    number.  The 70% owner control is a release invariant, never a clamp or a
    target used to back-solve cost.
    """

    status = str(basis_status or "BLOCKED_COST_COMPLETENESS")
    if status != "READY":
        return {
            "gross_profit_cents": None,
            "gross_margin_bps": None,
            "status": status,
        }
    if (
        isinstance(revenue_cents, bool)
        or not isinstance(revenue_cents, int)
        or revenue_cents <= 0
        or isinstance(cost_cents, bool)
        or not isinstance(cost_cents, int)
    ):
        raise ProjectCostError(
            "GROSS_MARGIN_BASIS_INVALID",
            "READY gross margin requires positive integer-cents revenue and integer-cents cost",
        )
    gross_profit = revenue_cents - cost_cents
    gross_margin_bps = int(
        (
            Decimal(gross_profit)
            * Decimal(10_000)
            / Decimal(revenue_cents)
        ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    )
    if gross_margin_bps > MAX_GROSS_MARGIN_BPS:
        raise ProjectCostError(
            "GROSS_MARGIN_SANITY_GATE",
            "a project gross margin above 70% blocks the complete runtime; "
            "the engine never clamps or back-solves cost",
        )
    return {
        "gross_profit_cents": gross_profit,
        "gross_margin_bps": gross_margin_bps,
        "status": "READY",
    }


def _formal_cost_categories(
    snapshot: Mapping[str, Any],
) -> Dict[str, Dict[str, int]]:
    """Aggregate only the two planes that form ``JOB_COST_INCURRED``.

    Observation planes (6401, status-table values and payment-system values)
    are deliberately absent. The runtime projection and statement renderer use
    this helper so presentation code cannot silently invent a second formula.
    """

    totals: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for event in snapshot.get("events", ()):
        if event.get("plane") not in ("JOB_POSTED_ACTUAL", "COST_ACCRUED"):
            continue
        project = str(event.get("project") or "")
        category = str(event.get("category") or "未分类")
        amount = event.get("amount_cents")
        if not project or isinstance(amount, bool) or not isinstance(amount, int):
            raise ProjectCostError(
                "RUNTIME_CATEGORY_EVENT_INVALID",
                "formal cost event is missing a project, category, or integer-cents amount",
            )
        totals[project][category] += amount
    return {project: dict(categories) for project, categories in totals.items()}


def _statement_buckets(categories: Mapping[str, int]) -> Dict[str, int]:
    """Map formal event categories to the owner statement without losing cents."""

    mapping = {
        "材料": "material",
        "燃料及动力": "fuel_power",
        "设备租赁": "rental",
        "物流运杂": "logistics",
        "住宿": "lodging",
        "生活补助": "living",
        "交通/差旅": "travel",
        "过路停车": "road_parking",
        "车辆油费": "vehicle",
        "自有人工过账": "own_labor",
        "自有人工-工资应计": "own_labor",
        "自有人工-雇主社保医保应计": "own_labor",
        "外协": "subcontract_labor",
        "劳务/分包": "subcontract_labor",
        "劳务/人工": "subcontract_labor",
        "已过账制造费用分配": "other",
        "其他直接成本": "other",
        "其他5001": "other",
        "未分类": "other",
    }
    result: Dict[str, int] = defaultdict(int)
    for category, amount in categories.items():
        result[mapping.get(category, "other")] += int(amount)
    return dict(result)


def runtime_projection(
    snapshot: Mapping[str, Any],
    sealed_workbook: Optional[Mapping[str, Any]] = None,
) -> Dict[str, Any]:
    """Create the private payload consumed by the KMFA web app.

    It contains business values and therefore belongs only in the private
    runtime volume or Private-Database. It must never be committed.
    """

    reviews_control = review_summary(snapshot.get("reviews") or ())
    if reviews_control["by_severity"]["P0"]:
        raise ProjectCostError(
            "P0_REVIEW_OPEN",
            "P0 review rows block runtime publication",
        )
    if reviews_control["by_severity"]["P1"]:
        raise ProjectCostError(
            "P1_REVIEW_OPEN",
            "P1 review rows block formal runtime publication",
        )
    categories_by_project = _formal_cost_categories(snapshot)
    rows: List[Dict[str, Any]] = []
    for project in snapshot.get("projects", ()):
        base = str(project["contract_base"])
        categories = categories_by_project.get(base, {})
        buckets = _statement_buckets(categories)
        formal_total = project.get("job_cost_incurred_cents")
        if formal_total is not None and sum(buckets.values()) != formal_total:
            raise ProjectCostError(
                "RUNTIME_CATEGORY_CONSERVATION",
                "formal statement categories do not conserve project cents",
            )
        observed = project.get("status_business_components_cents") or {}
        margin_cost_basis = project.get("gross_margin_cost_basis_cents")
        if (
            str(project.get("gross_margin_status") or "") == "READY"
            and formal_total is not None
            and (
                isinstance(margin_cost_basis, bool)
                or not isinstance(margin_cost_basis, int)
                or margin_cost_basis < formal_total
            )
        ):
            raise ProjectCostError(
                "GROSS_MARGIN_COST_BELOW_INCURRED",
                "closed project cost/FAC must be integer cents and cannot be below incurred cost",
            )
        margin = governed_gross_margin(
            revenue_cents=project.get("effective_revenue_cents"),
            cost_cents=margin_cost_basis,
            basis_status=str(
                project.get("gross_margin_status")
                or "BLOCKED_COST_COMPLETENESS"
            ),
        )
        margin_bps = margin["gross_margin_bps"]
        rows.append(
            {
                "合同编号": project.get("canonical_contract_id"),
                "合同基准号": base,
                "项目名称": project.get("project_name"),
                "甲方名称": project.get("customer"),
                "施工状态": project.get("construction_status"),
                "项目类型": project.get("project_type"),
                "负责人": project.get("owner"),
                "开工时间": project.get("start_date"),
                "完工日期": project.get("completion_date"),
                "结算时间": project.get("settlement_date"),
                "开票时间": project.get("invoice_date"),
                "回款时间": project.get("cash_in_date"),
                "含税合同金额": _yuan_text(project.get("contract_amount_cents")),
                "合同额口径": "红圈原始合同；不含未经完整批准链验证的变更",
                "有效合同额": _yuan_text(project.get("effective_revenue_cents")),
                "项目成本": (
                    _yuan_text(margin_cost_basis)
                    if margin["status"] == "READY"
                    else None
                ),
                "收入确认额": None,
                "毛利": _yuan_text(margin["gross_profit_cents"]),
                "毛利率": (
                    "%s%%"
                    % format(Decimal(int(margin_bps)) / Decimal(100), ".2f")
                    if margin_bps is not None
                    else None
                ),
                "毛利率基点": margin_bps,
                "收入与毛利状态": margin["status"],
                "结算金额": _yuan_text(project.get("status_settlement_amount_cents")),
                "开票金额": _yuan_text(project.get("status_invoice_amount_cents")),
                "自有人工工时": project.get("own_work_units"),
                "劳务人工工时": project.get("external_work_units"),
                "生活住宿费": _yuan_text(observed.get("生活住宿费")),
                "交通费": _yuan_text(observed.get("交通费")),
                "材料费": _yuan_text(observed.get("材料费")),
                "其他费用": _yuan_text(observed.get("其他费用")),
                "项目过账实际": _yuan_text(project.get("job_posted_actual_cents")),
                "项目应计": _yuan_text(project.get("cost_accrued_cents")),
                "项目已发生成本": _yuan_text(formal_total),
                "主营成本已结转": _yuan_text(project.get("gl_recognized_cogs_cents")),
                "状态表已报直接成本": _yuan_text(
                    project.get("business_reported_direct_cost_cents")
                ),
                "支付系统已付观察": _yuan_text(
                    project.get("payment_system_paid_observed_cents")
                ),
                "正式成本分类": {
                    category: _yuan_text(amount)
                    for category, amount in sorted(categories.items())
                },
                "报表归类": {
                    bucket: _yuan_text(amount)
                    for bucket, amount in sorted(buckets.items())
                },
                "项目成本覆盖": project.get("job_cost_coverage"),
                "账簿截至月份": project.get("ledger_period_end"),
                "应计覆盖": project.get("accrual_coverage"),
                "状态覆盖": project.get("status_coverage"),
                "支付覆盖": project.get("payment_coverage"),
            }
        )
    coverage = snapshot.get("coverage") or {}
    labor = snapshot.get("diagnostics", {}).get("labor", {})
    subject_binding = snapshot.get("subject_binding")
    input_manifest_binding = snapshot.get("input_manifest_binding")
    private_manifest_digest = snapshot.get("private_input_manifest_sha256")
    selected_source_digest = snapshot.get("selected_source_binding_digest")
    if (
        not isinstance(subject_binding, dict)
        or subject_binding.get("recipe") != SUBJECT_DIGEST_RECIPE
        or re.fullmatch(
            r"[0-9a-f]{64}",
            str(subject_binding.get("digest") or ""),
        )
        is None
        or not isinstance(subject_binding.get("file_count"), int)
        or subject_binding.get("file_count", 0) <= 0
        or re.fullmatch(
            r"[0-9a-f]{64}",
            str(selected_source_digest or ""),
        )
        is None
        or not isinstance(input_manifest_binding, dict)
        or re.fullmatch(
            r"[0-9a-f]{64}",
            str(input_manifest_binding.get("digest") or ""),
        )
        is None
    ):
        raise ProjectCostError(
            "RUNTIME_SOURCE_BINDING",
            "runtime projection requires valid source and selected-input bindings",
        )
    input_kind = input_manifest_binding.get("kind")
    input_digest = input_manifest_binding.get("digest")
    if (
        input_kind == "PRIVATE_MANIFEST_SHA256"
        and (
            re.fullmatch(r"[0-9a-f]{64}", str(private_manifest_digest or ""))
            is None
            or input_digest != private_manifest_digest
        )
    ) or (
        input_kind == "SELECTED_SOURCE_DERIVED_SHA256"
        and (
            private_manifest_digest is not None
            or input_digest != selected_source_digest
        )
    ) or input_kind not in (
        "PRIVATE_MANIFEST_SHA256",
        "SELECTED_SOURCE_DERIVED_SHA256",
    ):
        raise ProjectCostError(
            "RUNTIME_INPUT_MANIFEST_BINDING",
            "runtime input-manifest binding kind and digest are inconsistent",
        )
    result = {
        "schema_version": "kmfa.project_cost.current.v4",
        "生成时间": snapshot.get("generated_at"),
        "快照ID": snapshot.get("snapshot_id"),
        "年份": snapshot.get("year"),
        "截至日期": snapshot.get("as_of"),
        "币种": snapshot.get("currency"),
        "金额单位": "元；正式计算内部使用整数分",
        "计算状态": reviews_control["status"],
        "项目数": len(rows),
        "正式成本口径": "项目已发生成本 = 项目过账实际 + 合格应计",
        "观察面": ["主营成本已结转", "状态表已报直接成本", "支付系统已付观察"],
        "禁止": ["观察面相加", "取最大值", "固定人工单价", "自动合同额2%管理费", "参考报表回填"],
        "封印来源": {
            "源码摘要算法": subject_binding["recipe"],
            "源码SHA256": subject_binding["digest"],
            "源码文件数": subject_binding["file_count"],
            "输入清单类型": input_kind,
            "输入清单SHA256": input_digest,
            "私有输入清单SHA256": private_manifest_digest,
            "选中来源绑定SHA256": selected_source_digest,
        },
        "账簿覆盖": {
            "最早截至月份": coverage.get("ledger_minimum_period_end"),
            "最晚截至月份": coverage.get("ledger_logical_period_end"),
            "账簿主体数": coverage.get("ledger_entity_count"),
            "早于报表截至月主体数": coverage.get("ledger_stale_entity_count"),
            "说明": "过账实际只覆盖各主体所列截至月；随后仅纳入满足资格且未见过账冲突的应计",
        },
        "待确认": {
            "状态": reviews_control["status"],
            "P0阻断数": reviews_control["by_severity"]["P0"],
            "P1开放复核数": reviews_control["by_severity"]["P1"],
            "P2已排除或提示数": reviews_control["by_severity"]["P2"],
            "说明": "P1 观察未唯一归属或未满足成本资格，均已排除在正式公式之外；P0 会阻断发布",
        },
        "项目": rows,
        "校验": {
            "项目数一致": len(rows) == snapshot.get("project_count"),
            "P0阻断数": reviews_control["by_severity"]["P0"],
            "人工分配分差": coverage.get(
                "labor_fully_loaded_control_cents",
                coverage.get("labor_wage_component_control_cents", 0),
            )
            - coverage.get("labor_allocated_accrual_cents", 0)
            - coverage.get("labor_unallocated_cents", 0)
            - labor.get("already_posted_cents", 0),
            "金蝶明细账本数": coverage.get("ledger_selected_book_count"),
            "合格应计事件数": coverage.get("qualified_accrual_event_count"),
            "工资应计事件数": coverage.get("labor_wage_component_event_count"),
            "单位社保医保应计事件数": coverage.get(
                "labor_employer_burden_event_count"
            ),
        },
    }
    if sealed_workbook is not None:
        filename = str(sealed_workbook.get("filename") or "")
        digest = str(sealed_workbook.get("sha256") or "").lower()
        size = sealed_workbook.get("size_bytes")
        if (
            not filename
            or PurePosixPath(filename).name != filename
            or "/" in filename
            or "\\" in filename
            or not re.fullmatch(r"[0-9a-f]{64}", digest)
            or not isinstance(size, int)
            or isinstance(size, bool)
            or size <= 0
        ):
            raise ProjectCostError(
                "RUNTIME_WORKBOOK_BINDING",
                "sealed workbook binding is invalid",
            )
        result["封印工作簿"] = {
            "文件名": filename,
            "SHA256": digest,
            "字节数": size,
            "快照ID": snapshot.get("snapshot_id"),
        }
    if (
        not result["校验"]["项目数一致"]
        or result["校验"]["人工分配分差"] != 0
        or result["校验"]["P0阻断数"] != 0
    ):
        raise ProjectCostError(
            "RUNTIME_PROJECTION_CONTROL",
            "runtime projection failed project-count or labor-conservation control",
        )
    return result


def write_runtime_projection(
    path: Path,
    snapshot: Mapping[str, Any],
    sealed_workbook: Optional[Mapping[str, Any]] = None,
) -> Dict[str, Any]:
    """Atomically publish one private runtime JSON payload."""

    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.is_symlink() or destination.parent.is_symlink():
        raise ProjectCostError(
            "RUNTIME_OUTPUT_UNSAFE",
            "runtime JSON destination and parent must not be symbolic links",
        )
    payload = runtime_projection(snapshot, sealed_workbook)
    temporary = destination.parent / (".%s.tmp-%s" % (destination.name, uuid.uuid4().hex))
    try:
        with temporary.open("xb") as handle:
            handle.write(pretty_json(payload))
            handle.flush()
            os.fsync(handle.fileno())
        temporary.chmod(0o600)
        os.replace(str(temporary), str(destination))
    finally:
        if temporary.exists():
            temporary.unlink()
    return payload


def _file_manifest(root: Path, exclude: Sequence[str] = ()) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    excluded = set(exclude)
    for path in sorted((item for item in root.rglob("*") if item.is_file()), key=lambda item: item.as_posix()):
        relative = path.relative_to(root).as_posix()
        if relative in excluded:
            continue
        rows.append(
            {
                "path": relative,
                "sha256": sha256_file(path),
                "size_bytes": path.stat().st_size,
            }
        )
    return rows


def generate_outputs(output_dir: Path, snapshot: Mapping[str, Any]) -> Dict[str, Any]:
    output = Path(output_dir)
    reviews_control = review_summary(snapshot.get("reviews") or ())
    if reviews_control["by_severity"]["P0"]:
        raise ProjectCostError(
            "P0_REVIEW_OPEN",
            "P0 review rows block output generation",
        )
    if not output.is_absolute():
        raise ProjectCostError("OUTPUT_NOT_ABSOLUTE", "output directory must be absolute")
    if output.exists():
        raise ProjectCostError("OUTPUT_EXISTS", "output directory must not already exist")
    if not output.parent.is_dir():
        raise ProjectCostError("OUTPUT_PARENT_MISSING", "output parent directory does not exist")
    temporary = output.parent / (".%s.tmp-%s" % (output.name, uuid.uuid4().hex))
    temporary.mkdir(mode=0o700)
    try:
        snapshot_path = temporary / "project_cost_snapshot.json"
        snapshot_path.write_bytes(pretty_json(snapshot))
        csv_path = temporary / "project_cost_summary.csv"
        write_summary_csv(csv_path, snapshot)
        workbook_path = temporary / (
            "KMFA_项目成本报表_%d_%s.xlsx"
            % (snapshot["year"], str(snapshot["snapshot_id"]).rsplit("-", 1)[-1])
        )
        write_workbook(workbook_path, snapshot)
        pdf_paths = write_project_pdfs(temporary / "项目单页PDF", snapshot)
        business_files = _file_manifest(temporary)
        manifest = {
            "schema_version": "kmfa.project_cost.run_manifest.v2",
            "snapshot_id": snapshot["snapshot_id"],
            "generated_at": snapshot["generated_at"],
            "skill_version": snapshot["skill_version"],
            "year": snapshot["year"],
            "as_of": snapshot["as_of"],
            "project_count": snapshot["project_count"],
            "event_count": len(snapshot["events"]),
            "review_count": len(snapshot["reviews"]),
            "review_summary": reviews_control,
            "calculation_status": reviews_control["status"],
            "subject_binding": snapshot["subject_binding"],
            "input_manifest_binding": snapshot["input_manifest_binding"],
            "private_input_manifest_sha256": snapshot.get(
                "private_input_manifest_sha256"
            ),
            "selected_source_binding_digest": snapshot[
                "selected_source_binding_digest"
            ],
            "workbook": workbook_path.name,
            "pdf_count": len(pdf_paths),
            "coverage": snapshot["coverage"],
            "formula_contract": snapshot["formula_contract"],
            "files": business_files,
        }
        (temporary / "run_manifest.json").write_bytes(pretty_json(manifest))
        seal_rows = _file_manifest(temporary, exclude=("run_seal.sha256",))
        seal_text = "".join("%s  %s\n" % (row["sha256"], row["path"]) for row in seal_rows)
        (temporary / "run_seal.sha256").write_text(seal_text, encoding="utf-8")
        os.replace(str(temporary), str(output))
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    result = verify_output(output)
    result.update(
        {
            "output_dir": str(output),
            "workbook": str(output / workbook_path.name),
            "snapshot": str(output / "project_cost_snapshot.json"),
            "summary_csv": str(output / "project_cost_summary.csv"),
            "pdf_directory": str(output / "项目单页PDF"),
        }
    )
    return result


def _seal_map(path: Path) -> Dict[str, str]:
    observed: Dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if "  " not in line:
            raise ProjectCostError("SEAL_FORMAT", "run seal contains a malformed line")
        digest, relative = line.split("  ", 1)
        if not re.fullmatch(r"[0-9a-f]{64}", digest):
            raise ProjectCostError("SEAL_FORMAT", "run seal contains a malformed digest")
        _safe_archive_name(relative)
        if relative in observed:
            raise ProjectCostError("SEAL_DUPLICATE", "run seal repeats a path")
        observed[relative] = digest
    return observed


def _verify_xlsx_security(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        audit_archive(archive)
        names = set(archive.namelist())
        forbidden_exact = {
            "xl/vbaProject.bin",
            "xl/connections.xml",
            "xl/externalLinks/externalLink1.xml",
        }
        if names.intersection(forbidden_exact):
            raise ProjectCostError("XLSX_ACTIVE_CONTENT", "workbook contains active or external content")
        if any(
            name.startswith(("xl/externalLinks/", "xl/embeddings/", "xl/activeX/"))
            for name in names
        ):
            raise ProjectCostError("XLSX_ACTIVE_CONTENT", "workbook contains external or embedded content")
        for name in names:
            if not name.endswith((".xml", ".rels")):
                continue
            data = archive.read(name)
            lowered = data.lower()
            if b"<dde" in lowered or b"TargetMode=\"External\"" in data:
                raise ProjectCostError("XLSX_EXTERNAL_LINK", "workbook contains an external-link marker")


def verify_output(
    output_dir: Path,
    *,
    expected_private_input_manifest_sha256: Optional[str] = None,
) -> Dict[str, Any]:
    root = Path(output_dir)
    if root.is_symlink() or not root.is_dir():
        raise ProjectCostError("OUTPUT_INVALID", "output directory is unavailable")
    seal_path = root / "run_seal.sha256"
    manifest_path = root / "run_manifest.json"
    snapshot_path = root / "project_cost_snapshot.json"
    csv_path = root / "project_cost_summary.csv"
    for required in (seal_path, manifest_path, snapshot_path, csv_path):
        if not required.is_file():
            raise ProjectCostError("OUTPUT_REQUIRED_MISSING", "required output is missing: %s" % required.name)
    expected = _seal_map(seal_path)
    actual_files = {
        path.relative_to(root).as_posix()
        for path in root.rglob("*")
        if path.is_file() and path != seal_path
    }
    if set(expected) != actual_files:
        raise ProjectCostError("SEAL_FILE_SET", "run seal file set differs from output directory")
    for relative, digest in expected.items():
        if sha256_file(root / Path(relative)) != digest:
            raise ProjectCostError("SEAL_HASH", "run seal hash mismatch: %s" % relative)
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if snapshot.get("schema_version") != "kmfa.project_cost.snapshot.v2":
        raise ProjectCostError(
            "SNAPSHOT_SCHEMA",
            "sealed snapshot does not use the current source-bound schema",
        )
    if manifest.get("schema_version") != "kmfa.project_cost.run_manifest.v2":
        raise ProjectCostError(
            "MANIFEST_SCHEMA",
            "run manifest does not use the current source-bound schema",
        )
    current_subject = subject_source_binding()
    if snapshot.get("subject_binding") != current_subject:
        raise ProjectCostError(
            "SUBJECT_BINDING_MISMATCH",
            "sealed output was not generated by the current complete Skill source tree",
        )
    if manifest.get("subject_binding") != current_subject:
        raise ProjectCostError(
            "MANIFEST_SUBJECT_BINDING",
            "run manifest subject binding differs from current Skill source",
        )
    input_manifest_binding = snapshot.get("input_manifest_binding")
    if (
        not isinstance(input_manifest_binding, dict)
        or input_manifest_binding.get("kind")
        not in (
            "PRIVATE_MANIFEST_SHA256",
            "SELECTED_SOURCE_DERIVED_SHA256",
        )
        or re.fullmatch(
            r"[0-9a-f]{64}",
            str(input_manifest_binding.get("digest") or ""),
        )
        is None
    ):
        raise ProjectCostError(
            "INPUT_MANIFEST_BINDING",
            "sealed snapshot lacks a valid non-empty input-manifest binding",
        )
    if manifest.get("input_manifest_binding") != input_manifest_binding:
        raise ProjectCostError(
            "MANIFEST_INPUT_BINDING",
            "run manifest and snapshot input-manifest bindings differ",
        )
    private_manifest_digest = snapshot.get("private_input_manifest_sha256")
    if private_manifest_digest is not None and re.fullmatch(
        r"[0-9a-f]{64}",
        str(private_manifest_digest),
    ) is None:
        raise ProjectCostError(
            "PRIVATE_MANIFEST_BINDING",
            "sealed private input manifest digest is invalid",
        )
    if manifest.get("private_input_manifest_sha256") != private_manifest_digest:
        raise ProjectCostError(
            "MANIFEST_PRIVATE_INPUT_BINDING",
            "run manifest and snapshot private-input bindings differ",
        )
    if (
        expected_private_input_manifest_sha256 is not None
        and private_manifest_digest != expected_private_input_manifest_sha256
    ):
        raise ProjectCostError(
            "PRIVATE_MANIFEST_BINDING_MISMATCH",
            "sealed output does not bind the expected private input manifest",
        )
    selected_source_binding = [
        {
            "source_id": source["source_id"],
            "sha256": source["sha256"],
            "logical_metadata": source.get("logical_metadata"),
        }
        for source in snapshot.get("sources", ())
        if source.get("selected")
    ]
    selected_source_binding_digest = sha256_bytes(
        stable_json(selected_source_binding)
    )
    if (
        snapshot.get("selected_source_binding_digest")
        != selected_source_binding_digest
        or manifest.get("selected_source_binding_digest")
        != selected_source_binding_digest
    ):
        raise ProjectCostError(
            "SELECTED_SOURCE_BINDING_MISMATCH",
            "selected source binding differs from the sealed source inventory",
        )
    if (
        input_manifest_binding["kind"] == "PRIVATE_MANIFEST_SHA256"
        and (
            private_manifest_digest is None
            or input_manifest_binding["digest"] != private_manifest_digest
        )
    ) or (
        input_manifest_binding["kind"] == "SELECTED_SOURCE_DERIVED_SHA256"
        and (
            private_manifest_digest is not None
            or input_manifest_binding["digest"]
            != selected_source_binding_digest
        )
    ):
        raise ProjectCostError(
            "INPUT_MANIFEST_BINDING_MISMATCH",
            "input-manifest binding does not match its declared source",
        )
    reviews_control = review_summary(snapshot.get("reviews") or ())
    if reviews_control["by_severity"]["P0"]:
        raise ProjectCostError(
            "P0_REVIEW_OPEN",
            "P0 review rows block output verification",
        )
    if manifest.get("review_summary") != reviews_control:
        raise ProjectCostError(
            "MANIFEST_REVIEW_SUMMARY",
            "manifest review summary differs from the sealed snapshot",
        )
    if manifest.get("calculation_status") != reviews_control["status"]:
        raise ProjectCostError(
            "MANIFEST_CALCULATION_STATUS",
            "manifest calculation status differs from the sealed snapshot",
        )
    workbook_path = root / manifest["workbook"]
    if not workbook_path.is_file():
        raise ProjectCostError("WORKBOOK_MISSING", "manifest workbook is missing")
    _verify_xlsx_security(workbook_path)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ProjectCostError("DEPENDENCY_MISSING", "openpyxl is required") from exc
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if tuple(workbook.sheetnames) != FORMULA_SHEETS:
        workbook.close()
        raise ProjectCostError("WORKBOOK_SHEETS", "workbook sheet order differs from the output contract")
    workbook_rows: Dict[str, int] = {}
    for sheet in workbook.worksheets:
        try:
            sheet.reset_dimensions()
        except Exception:
            pass
        if sheet["B1"].value != snapshot["snapshot_id"]:
            workbook.close()
            raise ProjectCostError("WORKBOOK_SNAPSHOT", "worksheet snapshot ID mismatch")
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    workbook.close()
                    raise ProjectCostError("WORKBOOK_FORMULA", "workbook contains a formula")
    summary_sheet = workbook["01_项目成本表"]
    headers = {
        str(cell.value): index
        for index, cell in enumerate(next(summary_sheet.iter_rows(min_row=3, max_row=3)), 1)
        if cell.value is not None
    }
    required_headers = ("合同编号", "项目已发生成本（元）")
    if any(header not in headers for header in required_headers):
        workbook.close()
        raise ProjectCostError("WORKBOOK_SCHEMA", "summary sheet headers are incomplete")
    for row in summary_sheet.iter_rows(min_row=4, values_only=True):
        contract = row[headers["合同编号"] - 1]
        if not contract:
            continue
        amount = row[headers["项目已发生成本（元）"] - 1]
        workbook_rows[str(contract).lstrip("'")] = cents(amount) if amount is not None else None
    workbook.close()
    snapshot_rows = {
        str(project["canonical_contract_id"]): project.get("job_cost_incurred_cents")
        for project in snapshot["projects"]
    }
    if workbook_rows != snapshot_rows:
        raise ProjectCostError("WORKBOOK_JSON_PARITY", "workbook and JSON project costs differ")
    csv_rows: Dict[str, Optional[int]] = {}
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            value = row.get("job_cost_incurred_cents")
            csv_rows[str(row["canonical_contract_id"])] = int(value) if value not in (None, "") else None
    if csv_rows != snapshot_rows:
        raise ProjectCostError("CSV_JSON_PARITY", "CSV and JSON project costs differ")
    pdf_paths = sorted((root / "项目单页PDF").glob("*.pdf"))
    if len(pdf_paths) != snapshot["project_count"]:
        raise ProjectCostError("PDF_COUNT", "single-project PDF count differs from project count")
    expected_pdf_projects = {
        "%s_%s.pdf"
        % (
            _safe_filename(str(project["canonical_contract_id"])),
            _safe_filename(str(project["project_name"]), 45),
        ): project
        for project in snapshot["projects"]
    }
    if set(expected_pdf_projects) != {path.name for path in pdf_paths}:
        raise ProjectCostError(
            "PDF_PROJECT_SET",
            "single-project PDF filenames differ from the snapshot project set",
        )
    try:
        import pdfplumber
    except ImportError as exc:
        raise ProjectCostError("DEPENDENCY_MISSING", "pdfplumber is required for PDF verification") from exc
    for path in pdf_paths:
        expected_project = expected_pdf_projects[path.name]
        with pdfplumber.open(path) as pdf:
            if len(pdf.pages) != 1:
                raise ProjectCostError("PDF_PAGE_COUNT", "single-project PDF must have exactly one page")
            page = pdf.pages[0]
            if abs(float(page.width) - 595.28) > 3 or abs(float(page.height) - 841.89) > 3:
                raise ProjectCostError("PDF_PAGE_SIZE", "single-project PDF is not A4")
            text = page.extract_text() or ""
            if str(expected_project["canonical_contract_id"]) not in text:
                raise ProjectCostError(
                    "PDF_CONTRACT_IDENTITY",
                    "single-project PDF contract identity differs from the snapshot",
                )
            if str(snapshot["snapshot_id"]) not in text:
                raise ProjectCostError(
                    "PDF_SNAPSHOT_ID",
                    "single-project PDF snapshot ID is missing",
                )
            expected_cost = expected_project.get("job_cost_incurred_cents")
            if (
                isinstance(expected_cost, int)
                and not isinstance(expected_cost, bool)
                and not _reference_amount_present(text, expected_cost)
            ):
                raise ProjectCostError(
                    "PDF_PROJECT_COST",
                    "single-project PDF formal cost differs from the snapshot",
                )
    if manifest.get("project_count") != snapshot.get("project_count"):
        raise ProjectCostError("MANIFEST_PROJECT_COUNT", "manifest project count mismatch")
    return {
        "status": reviews_control["status"],
        "snapshot_id": snapshot["snapshot_id"],
        "project_count": snapshot["project_count"],
        "event_count": len(snapshot["events"]),
        "review_count": len(snapshot["reviews"]),
        "p0_review_count": reviews_control["by_severity"]["P0"],
        "p1_review_count": reviews_control["by_severity"]["P1"],
        "p2_review_count": reviews_control["by_severity"]["P2"],
        "pdf_count": len(pdf_paths),
        "sealed_file_count": len(expected),
        "workbook_json_csv_parity": True,
        "workbook_formula_free": True,
        "workbook_external_link_free": True,
        "pdf_snapshot_and_amount_parity": True,
        "subject_digest": current_subject["digest"],
        "input_manifest_binding": input_manifest_binding,
        "private_input_manifest_sha256": private_manifest_digest,
        "selected_source_binding_digest": selected_source_binding_digest,
    }


def _digits(value: str) -> str:
    return re.sub(r"\D", "", value)


def _reference_amount_present(text: str, amount_minor: int) -> bool:
    amount = Decimal(amount_minor) / 100
    decimal_forms = {
        format(amount, ".2f"),
        format(amount, ",.2f"),
        format(amount.normalize(), "f"),
    }
    compact = re.sub(r"[\s,，￥¥]", "", text)
    if any(form.replace(",", "") in compact for form in decimal_forms):
        return True
    minor_digits = str(abs(amount_minor))
    return len(minor_digits) >= 4 and minor_digits in _digits(text)


def _baseline_records(path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    workbook = open_xlsx_payload(read_path_bytes(path))
    if "项目总览" not in workbook.sheetnames or "毛利复核" not in workbook.sheetnames:
        workbook.close()
        raise ProjectCostError("REFERENCE_BASELINE_SCHEMA", "reference baseline workbook is missing required sheets")
    overview = workbook["项目总览"]
    header_row, headers = locate_header(
        overview,
        ("合同编号", "参考总成本（元）", "参考回放"),
        max_rows=10,
    )
    records: List[Dict[str, Any]] = []
    for row in overview.iter_rows(min_row=header_row + 1, values_only=True):
        values = _row_dict(row, headers)
        contract = str(values.get("合同编号") or "").strip()
        if not contract.startswith("KMX"):
            continue
        records.append(
            {
                "contract": contract,
                "project_name": str(values.get("项目名称") or ""),
                "reference_total_cents": cents(values.get("参考总成本（元）")),
                "reference_status": str(values.get("参考回放") or ""),
            }
        )
    profit_sheet = workbook["毛利复核"]
    profit_header, profit_headers = locate_header(
        profit_sheet,
        ("合同编号", "收入/结算额（元）", "参考总成本（元）", "来源毛利（元）"),
        max_rows=10,
    )
    profit_rows: List[Dict[str, Any]] = []
    for row in profit_sheet.iter_rows(min_row=profit_header + 1, values_only=True):
        values = _row_dict(row, profit_headers)
        contract = str(values.get("合同编号") or "").strip()
        if not contract.startswith("KMX"):
            continue
        revenue = cents(values.get("收入/结算额（元）"))
        total = cents(values.get("参考总成本（元）"))
        source_profit = cents(values.get("来源毛利（元）"))
        recomputed = revenue - total if revenue is not None and total is not None else None
        delta = source_profit - recomputed if source_profit is not None and recomputed is not None else None
        profit_rows.append(
            {
                "contract": contract,
                "revenue_cents": revenue,
                "reference_total_cents": total,
                "source_profit_cents": source_profit,
                "recomputed_profit_cents": recomputed,
                "delta_cents": delta,
                "status": "SOURCE_ARITHMETIC_CONFLICT" if delta else "PASS",
            }
        )
    workbook.close()
    return records, profit_rows


def validate_reference(
    reference_root: Path,
    baseline_workbook: Path,
) -> Dict[str, Any]:
    root = Path(reference_root)
    if root.is_symlink() or not root.is_dir():
        raise ProjectCostError("REFERENCE_ROOT_INVALID", "reference root is unavailable")
    records, profit_rows = _baseline_records(Path(baseline_workbook))
    try:
        import pdfplumber
    except ImportError as exc:
        raise ProjectCostError("DEPENDENCY_MISSING", "pdfplumber is required") from exc
    pdfs: List[Dict[str, Any]] = []
    for path in sorted(root.rglob("*.pdf")):
        with pdfplumber.open(path) as pdf:
            text = "\n".join((page.extract_text() or "") for page in pdf.pages)
            a4 = all(
                abs(float(page.width) - 595.28) <= 4
                and abs(float(page.height) - 841.89) <= 4
                for page in pdf.pages
            )
            layout = "B" if "项目产值" in text and "采购材料" in text else "A"
            pdfs.append(
                {
                    "path": str(path),
                    "name": path.name,
                    "text": text,
                    "normalized": normalize_text(text),
                    "pages": len(pdf.pages),
                    "a4": a4,
                    "layout_family": layout,
                }
            )
    results: List[Dict[str, Any]] = []
    used: Set[str] = set()
    for record in records:
        contract = record["contract"]
        sequence = contract.rsplit("-", 1)[-1]
        matches = [
            pdf
            for pdf in pdfs
            if normalize_text(contract) in pdf["normalized"]
            or re.search(r"(^|\D)%s(\D|$)" % re.escape(sequence), pdf["name"])
        ]
        matches = [pdf for pdf in matches if pdf["path"] not in used]
        if len(matches) != 1:
            results.append(
                dict(
                    record,
                    pdf=None,
                    pdf_match_count=len(matches),
                    amount_match=False,
                    a4=False,
                    layout_family=None,
                    status="FAIL_PDF_IDENTITY",
                )
            )
            continue
        pdf = matches[0]
        used.add(pdf["path"])
        total_minor = record["reference_total_cents"]
        amount_match = (
            total_minor is not None and _reference_amount_present(pdf["text"], total_minor)
        )
        status = (
            "PASS"
            if amount_match and pdf["a4"] and record["reference_status"] == "PASS"
            else "FAIL"
        )
        results.append(
            dict(
                record,
                pdf=pdf["name"],
                pdf_match_count=1,
                amount_match=amount_match,
                a4=pdf["a4"],
                pages=pdf["pages"],
                layout_family=pdf["layout_family"],
                status=status,
            )
        )
    conflict_rows = [row for row in profit_rows if row["status"] == "SOURCE_ARITHMETIC_CONFLICT"]
    passed = (
        len(records) == 8
        and len(results) == 8
        and all(result["status"] == "PASS" for result in results)
        and len(conflict_rows) == 3
        and sum(result.get("layout_family") == "A" for result in results) == 5
        and sum(result.get("layout_family") == "B" for result in results) == 3
    )
    return {
        "schema_version": "kmfa.project_cost.reference_validation.v1",
        "generated_at": utc_now(),
        "status": "PASS" if passed else "FAIL",
        "reference_project_count": len(records),
        "amount_and_layout_pass_count": sum(result["status"] == "PASS" for result in results),
        "source_arithmetic_conflict_count": len(conflict_rows),
        "layout_family_counts": {
            "A": sum(result.get("layout_family") == "A" for result in results),
            "B": sum(result.get("layout_family") == "B" for result in results),
        },
        "results": results,
        "profit_reconciliation": profit_rows,
        "boundary": "REFERENCE_VALIDATION_ONLY;NOT_USED_BY_CALCULATE",
    }


def calculate_and_generate(
    roots: Sequence[Path],
    *,
    year: int,
    as_of: str,
    output_dir: Path,
    ocr_jsonl: Optional[Path] = None,
    payroll_workbooks: Sequence[Path] = (),
    employer_burden_workbooks: Sequence[Path] = (),
    attendance_roots: Sequence[Path] = (),
    payroll_password_env: Optional[str] = None,
    private_input_manifest_sha256: Optional[str] = None,
) -> Dict[str, Any]:
    output = Path(output_dir).resolve()
    input_paths = (
        tuple(Path(root).resolve() for root in roots)
        + tuple(Path(path).resolve() for path in payroll_workbooks)
        + tuple(Path(path).resolve() for path in employer_burden_workbooks)
        + tuple(Path(root).resolve() for root in attendance_roots)
    )
    for raw in input_paths:
        if output == raw or raw in output.parents or output in raw.parents:
            raise ProjectCostError(
                "OUTPUT_OVERLAPS_INPUT",
                "output directory must be disjoint from every raw data root",
            )
    snapshot = build_snapshot(
        roots,
        year=year,
        as_of=as_of,
        ocr_jsonl=ocr_jsonl,
        payroll_workbooks=payroll_workbooks,
        employer_burden_workbooks=employer_burden_workbooks,
        attendance_roots=attendance_roots,
        payroll_password_env=payroll_password_env,
        private_input_manifest_sha256=private_input_manifest_sha256,
    )
    result = generate_outputs(output, snapshot)
    result["coverage"] = snapshot["coverage"]
    result["job_cost_total_cents"] = sum(
        project["job_cost_incurred_cents"] or 0 for project in snapshot["projects"]
    )
    result["gl_recognized_cogs_total_cents"] = sum(
        project["gl_recognized_cogs_cents"] or 0 for project in snapshot["projects"]
    )
    return result


def verify_skill(skill_root: Path) -> Dict[str, Any]:
    source = Path(skill_root)
    cleanup: Optional[Path] = None
    if source.is_file() and source.suffix.lower() == ".zip":
        cleanup = Path(tempfile.mkdtemp(prefix="kmfa-skill-verify-"))
        with zipfile.ZipFile(source) as archive:
            audit_archive(archive)
            archive.extractall(cleanup)
        directories = [item for item in cleanup.iterdir() if item.is_dir()]
        files = [item for item in cleanup.iterdir() if item.is_file()]
        if len(directories) != 1 or files:
            shutil.rmtree(cleanup, ignore_errors=True)
            raise ProjectCostError("SKILL_ZIP_ROOT", "Skill ZIP must contain one top-level directory")
        root = directories[0]
    else:
        root = source
    try:
        required = (
            "SKILL.md",
            "VERSION",
            "OPERATIONAL_VERSION",
            "requirements.txt",
            "agents/openai.yaml",
            "formula_registry.yaml",
            "config/metric_catalog.yml",
            "config/money_profile.yml",
            "scripts/run_operational_report.py",
            "scripts/run_private_refresh.py",
            "scripts/validate_skill_package.py",
            "src/project_cost_table/operational.py",
            "schemas/operational_private_input_manifest.schema.json",
            "references/MONEY_AND_FILE_SAFETY.md",
            "references/REFERENCE_REPLAY_ISOLATION.md",
        )
        missing = [relative for relative in required if not (root / relative).is_file()]
        if missing:
            raise ProjectCostError("SKILL_REQUIRED_MISSING", "missing files: %s" % ", ".join(missing))
        skill_text = (root / "SKILL.md").read_text(encoding="utf-8")
        if not skill_text.startswith("---\n") or "name: project-cost-table-skill" not in skill_text:
            raise ProjectCostError("SKILL_FRONTMATTER", "SKILL.md frontmatter is invalid")
        core_version = (root / "VERSION").read_text(encoding="ascii").strip()
        version = (root / "OPERATIONAL_VERSION").read_text(encoding="ascii").strip()
        if core_version != CORE_VERSION or version != SKILL_VERSION:
            raise ProjectCostError(
                "SKILL_VERSION",
                "core or operational version differs from engine constants",
            )
        operational_text = (
            root / "src/project_cost_table/operational.py"
        ).read_text(encoding="utf-8")
        for token in (
            "job_posted_actual_cents",
            "cost_accrued_cents",
            "job_cost_incurred_cents",
            "automatic_management_fee_percent",
            "historical_reference_in_calculate",
        ):
            if token not in operational_text:
                raise ProjectCostError(
                    "POLICY_FORMULA",
                    "formal formula or prohibited-default control is missing",
                )
        disallowed_suffixes = {".xlsx", ".xls", ".pdf", ".sqlite", ".db", ".rar"}
        text_extensions = {".md", ".py", ".json", ".yaml", ".yml", ".m", ".txt"}
        compiled = 0
        text_scanned = 0
        privacy_pattern = re.compile(r"KMX20(?:1\d|2\d|3[0-5])\d*-\d{3}")
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if path.name in (".DS_Store",) or "__pycache__" in path.parts:
                raise ProjectCostError("SKILL_JUNK_FILE", "Skill contains cache or metadata junk")
            if path.suffix.lower() in disallowed_suffixes:
                raise ProjectCostError("SKILL_PRIVATE_ARTIFACT", "Skill contains a raw/report artifact")
            if path.suffix.lower() == ".py":
                content = path.read_text(encoding="utf-8")
                compile(content, str(path), "exec")
                compiled += 1
            if path.suffix.lower() in text_extensions:
                content = path.read_text(encoding="utf-8")
                text_scanned += 1
                if privacy_pattern.search(content):
                    raise ProjectCostError(
                        "SKILL_REAL_CONTRACT_LEAK",
                        "Skill contains a real-looking contract ID",
                    )
        return {
            "status": "PASS",
            "skill_root": str(root),
            "version": version,
            "core_version": core_version,
            "required_file_count": len(required),
            "python_files_compiled": compiled,
            "text_files_privacy_scanned": text_scanned,
            "raw_private_artifacts": 0,
        }
    finally:
        if cleanup is not None:
            shutil.rmtree(cleanup, ignore_errors=True)


def _synthetic_master(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "合同名称",
            "合同编号",
            "创建时间",
            "施工状态",
            "甲方",
            "完工日期（产值上报）",
            "乙方",
            "含税合同额(元)",
            "税率(%)",
            "结算金额(元)",
            "结算应收款(元)",
            "累计开票",
            "开票应收款(元)",
            "项目收款金额",
            "保证金支付金额",
            "保证金退还金额",
            "验收日期",
        ]
    )
    sheet.append(
        [
            "合成项目甲",
            "KMX20990101-001",
            "2099-01-01",
            "已完工",
            "合成客户甲",
            "2099-02-01",
            "合成企业甲",
            1000,
            "9%",
            1000,
            0,
            1000,
            0,
            900,
            0,
            0,
            "2099-02-01",
        ]
    )
    sheet.append(
        [
            "合成项目乙",
            "KMX20990102-002-XF",
            "2099-01-02",
            "施工中",
            "合成客户乙",
            None,
            "合成企业乙",
            2000,
            "13%",
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            None,
        ]
    )
    workbook.save(path)


def _synthetic_status(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "甲方名称",
            "合同号",
            "施工状态",
            "项目类型",
            "负责人",
            "开工时间",
            "完工时间",
            "结算时间",
            "开票时间",
            "回款时间",
            "结算金额",
            "开票金额",
            "自有人工工时",
            "劳务人工工时",
            "生活住宿费",
            "交通费",
            "材料费",
            "其他费用",
            "含税合同金额",
            "税率",
            "项目成本表截止提供时间",
            "是否提供项目成本表",
            "是否已计算提成",
        ]
    )
    sheet.append(
        [
            "合成客户甲",
            "KMX20990101-001-Z",
            "已完工",
            "自有人员",
            "测试",
            "2099-01-10",
            "2099-02-01",
            "2099-02-02",
            "2099-02-03",
            "2099-02-04",
            1000,
            1000,
            2,
            None,
            10,
            20,
            30,
            40,
            1000,
            "9%",
            "2099-03-01",
            "已提供",
            "是",
        ]
    )
    workbook.save(path)


def _synthetic_ledger_book(
    *,
    entity: str,
    contract: str,
    customer: str,
    amount: Decimal,
    account: str,
    include_research_column: bool,
    cogs: Optional[Decimal] = None,
    duplicate_auxiliary_view: bool = False,
    same_sheet_occurrences: int = 1,
    posting_date: str = "2099-02-01",
) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    headers = ["科目", "客户", "职员", "供应商", "部门", "销售合同号"]
    if include_research_column:
        headers.append("研发项目")
    headers.extend(["往来", "日期", "凭证字号", "摘要", "对方科目", "借方", "贷方", "方向", "余额"])

    def add_sheet(title: str, row_account: str, debit: Decimal, summary: str) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append(["明细账"] * len(headers))
        period_row = [entity] * len(headers)
        period_row[-1] = "2099年1期 至 2099年2期"
        sheet.append(period_row)
        sheet.append(headers)
        opening = [None] * len(headers)
        opening[0] = row_account
        opening[1] = customer
        opening[5] = contract
        opening[headers.index("日期")] = "2099-01-01"
        opening[headers.index("摘要")] = "期初余额"
        opening[headers.index("借方")] = 0
        opening[headers.index("贷方")] = 0
        sheet.append(opening)
        event = list(opening)
        event[headers.index("日期")] = posting_date
        event[headers.index("凭证字号")] = "记-1"
        event[headers.index("摘要")] = summary
        event[headers.index("借方")] = debit
        for _ in range(same_sheet_occurrences):
            sheet.append(event)

    add_sheet(account.replace("-", "_")[:28], account, amount, "合成项目成本")
    if duplicate_auxiliary_view:
        subject, _, label = account.partition("-")
        auxiliary_account = subject + "_06.037.001-" + label
        add_sheet(
            (subject + "_06.037_辅助核算")[:28],
            auxiliary_account,
            amount,
            "合成项目成本",
        )
    if cogs is not None:
        add_sheet("6401_合成项目", "6401-主营业务成本", cogs, "结转生产成本")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _synthetic_payment(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "付款编号",
            "支付状态(系统)",
            "申请日期",
            "支付日期",
            "申请支付金额",
            "实际支付金额",
            "审批状态",
            "收款账户",
            "付款内容",
            "备注",
        ]
    )
    sheet.append(
        [
            "SYN-1",
            "全部支付",
            "2099-02-01",
            "2099-02-02",
            12.34,
            12.34,
            "已通过",
            "合成账户",
            "合成客户乙 项目费用",
            "",
        ]
    )
    workbook.save(path)


def self_test() -> Dict[str, Any]:
    root = Path(tempfile.mkdtemp(prefix="kmfa-project-cost-selftest-"))
    try:
        payroll_stub = type("PayrollWorkbookStub", (), {"sheetnames": []})()
        period_cases = {
            "26-05月份工资表-个税已扣-20260629.xlsx": "2026-05",
            "26-06月份工资表-个税已扣-26.7.30.xlsx": "2026-06",
        }
        for name, expected_period in period_cases.items():
            actual_period = _payroll_period(Path(name), payroll_stub)
            if actual_period != expected_period:
                raise ProjectCostError(
                    "SELF_TEST_PAYROLL_PERIOD",
                    "%s resolved to %s instead of %s"
                    % (name, actual_period, expected_period),
                )
        allocation = largest_remainder_allocate(
            100,
            {
                "project-a": Decimal(1),
                "project-b": Decimal(1),
                "unallocated": Decimal(1),
            },
        )
        if allocation != {"project-a": 34, "project-b": 33, "unallocated": 33}:
            raise ProjectCostError(
                "SELF_TEST_LABOR_ALLOCATION",
                "largest-remainder tie-breaking or cent conservation differs",
            )
        data_root = root / "data"
        data_root.mkdir()
        _synthetic_master(data_root / "红圈主合同 2099.xlsx")
        status_dir = data_root / "2099-02-05"
        status_dir.mkdir()
        _synthetic_status(status_dir / "生产项目状态表.xlsx")
        _synthetic_payment(data_root / "付款审批（日常费用）_合成.xlsx")
        ledger_zip = data_root / "金蝶账套导出明细账2099.zip"
        with zipfile.ZipFile(ledger_zip, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            duplicate_voucher_view = _synthetic_ledger_book(
                entity="合成企业甲",
                contract="KMX20990101-001-Z",
                customer="合成客户甲",
                amount=Decimal("123.45"),
                account="5001001-生产成本_原材料",
                include_research_column=True,
                cogs=Decimal("80.00"),
                duplicate_auxiliary_view=True,
            )
            archive.writestr(
                "企业甲-明细账.xlsx",
                duplicate_voucher_view,
            )
            archive.writestr(
                "金蝶账套导出明细账2099/企业甲_凭证列表.xlsx",
                duplicate_voucher_view,
            )
            archive.writestr(
                "企业乙-明细账.xlsx",
                _synthetic_ledger_book(
                    entity="合成企业乙",
                    contract="KMX20990102-002-XF-Z",
                    customer="合成客户乙",
                    amount=Decimal("5.00"),
                    account="5001003-生产成本_工资",
                    include_research_column=True,
                ),
            )
        snapshot = build_snapshot((data_root,), year=2099, as_of="2099-02-05")
        if snapshot["coverage"]["ledger_selected_book_count"] != 2:
            raise ProjectCostError(
                "SELF_TEST_VOUCHER_PROJECTION",
                "voucher-list projection was admitted as a ledger workbook",
            )
        amounts = {
            project["contract_base"]: (
                project["job_cost_incurred_cents"],
                project["gl_recognized_cogs_cents"],
                project["payment_system_paid_observed_cents"],
            )
            for project in snapshot["projects"]
        }
        expected = {
            "KMX20990101-001": (12345, 8000, 0),
            "KMX20990102-002": (500, 0, 1234),
        }
        if amounts != expected:
            raise ProjectCostError(
                "SELF_TEST_AMOUNTS",
                "synthetic exact-cents result differs: %r" % (amounts,),
            )
        output = root / "output"
        generated = generate_outputs(output, snapshot)
        if generated.get("status") != "PASS":
            raise ProjectCostError("SELF_TEST_OUTPUT", "synthetic output verification failed")
        return {
            "status": "PASS",
            "python_compatibility": "3.9+",
            "synthetic_project_count": snapshot["project_count"],
            "schema_variants_tested": 2,
            "auxiliary_account_duplicate_deduped": True,
            "voucher_projection_excluded": True,
            "original_payroll_filename_periods": True,
            "integer_cents": True,
            "labor_largest_remainder_conserved": True,
            "dual_basis_separated": True,
            "output_verified": True,
        }
    finally:
        shutil.rmtree(root, ignore_errors=True)
